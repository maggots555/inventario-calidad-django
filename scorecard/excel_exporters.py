"""
Funciones auxiliares para exportación de Excel avanzada
Sistema de Score Card - Reportes Completos con Múltiples Hojas
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.db.models import Count, Avg, Q
from collections import defaultdict


# ==========================================
# FUNCIONES DE ESTILOS
# ==========================================

def get_header_style():
    """
    Retorna el estilo para encabezados de tablas
    Fondo azul con texto blanco en negrita
    """
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_title_style():
    """
    Retorna el estilo para títulos de secciones
    Texto grande en negrita con fondo gris claro
    """
    return {
        'font': Font(bold=True, size=14, color="1f4e78"),
        'fill': PatternFill(start_color="e7e6e6", end_color="e7e6e6", fill_type="solid"),
        'alignment': Alignment(horizontal="left", vertical="center")
    }


def get_kpi_title_style():
    """
    Retorna el estilo para títulos de KPIs
    """
    return {
        'font': Font(bold=True, size=11, color="495057"),
        'alignment': Alignment(horizontal="left", vertical="center")
    }


def get_kpi_value_style():
    """
    Retorna el estilo para valores de KPIs
    Números grandes y destacados
    """
    return {
        'font': Font(bold=True, size=16, color="0d6efd"),
        'alignment': Alignment(horizontal="center", vertical="center")
    }


def get_severidad_color(severidad):
    """
    Retorna el color de fondo según la severidad
    - Crítico: Rojo intenso
    - Alto: Naranja
    - Medio: Amarillo
    - Bajo: Verde
    """
    colores = {
        'critico': 'dc3545',
        'alto': 'fd7e14',
        'medio': 'ffc107',
        'bajo': '28a745'
    }
    return colores.get(severidad.lower(), 'ffffff')


def apply_cell_style(cell, style_dict):
    """
    Aplica un diccionario de estilos a una celda
    
    Args:
        cell: Celda de openpyxl
        style_dict: Diccionario con claves 'font', 'fill', 'alignment', 'border'
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def auto_adjust_column_width(ws, min_width=10, max_width=50):
    """
    Ajusta automáticamente el ancho de las columnas según el contenido
    
    Args:
        ws: Worksheet de openpyxl
        min_width: Ancho mínimo
        max_width: Ancho máximo
    """
    for column_cells in ws.columns:
        # Obtener la letra de la columna de forma segura (evitar MergedCell)
        column_letter = None
        for cell in column_cells:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                break
        
        if not column_letter:
            continue
            
        # Calcular el ancho basado en el contenido
        length = max(len(str(cell.value or '')) for cell in column_cells if hasattr(cell, 'value'))
        adjusted_width = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


# ==========================================
# FUNCIONES DE CÁLCULO Y ANÁLISIS
# ==========================================

def calcular_metricas_generales(incidencias):
    """
    Calcula métricas generales del sistema
    
    Returns:
        dict con KPIs principales
    """
    total = incidencias.count()
    abiertas = incidencias.filter(estado='abierta').count()
    criticas = incidencias.filter(grado_severidad='critico').count()
    cerradas = incidencias.filter(estado='cerrada').count()
    reincidencias = incidencias.filter(es_reincidencia=True).count()
    
    # Calcular porcentajes
    porcentaje_reincidencias = round((reincidencias / total * 100), 2) if total > 0 else 0
    porcentaje_cerradas = round((cerradas / total * 100), 2) if total > 0 else 0
    porcentaje_abiertas = round((abiertas / total * 100), 2) if total > 0 else 0
    
    # Promedio de días para cerrar
    incidencias_cerradas = incidencias.filter(estado='cerrada')
    if incidencias_cerradas.exists():
        dias_totales = sum([inc.dias_abierta for inc in incidencias_cerradas])
        promedio_dias_cierre = round(dias_totales / incidencias_cerradas.count(), 1)
    else:
        promedio_dias_cierre = 0
    
    return {
        'total_incidencias': total,
        'incidencias_abiertas': abiertas,
        'incidencias_criticas': criticas,
        'incidencias_cerradas': cerradas,
        'reincidencias': reincidencias,
        'porcentaje_reincidencias': porcentaje_reincidencias,
        'porcentaje_cerradas': porcentaje_cerradas,
        'porcentaje_abiertas': porcentaje_abiertas,
        'promedio_dias_cierre': promedio_dias_cierre
    }


def calcular_tendencia_mensual(incidencias, meses=6):
    """
    Calcula la tendencia mensual de incidencias
    
    Args:
        incidencias: QuerySet de incidencias
        meses: Número de meses hacia atrás
        
    Returns:
        Lista de tuplas (mes, cantidad)
    """
    hoy = datetime.now()
    fecha_inicio = hoy - timedelta(days=meses * 30)
    
    incidencias_periodo = incidencias.filter(fecha_deteccion__gte=fecha_inicio)
    
    # Agrupar por mes
    meses_data = defaultdict(int)
    for inc in incidencias_periodo:
        mes_key = inc.fecha_deteccion.strftime('%Y-%m')
        meses_data[mes_key] += 1
    
    # Ordenar cronológicamente
    return sorted(meses_data.items())


def calcular_distribucion_severidad(incidencias):
    """
    Calcula la distribución de incidencias por severidad
    
    Returns:
        Lista de tuplas (severidad, cantidad)
    """
    return incidencias.values('grado_severidad').annotate(
        count=Count('id')
    ).order_by('grado_severidad')


def calcular_estadisticas_por_sucursal(incidencias):
    """
    Calcula estadísticas detalladas por sucursal
    
    Returns:
        Lista de diccionarios con estadísticas por sucursal
    """
    sucursales_stats = []
    
    sucursales = incidencias.values_list('sucursal__nombre', flat=True).distinct()
    
    for sucursal in sucursales:
        if not sucursal:
            continue
            
        inc_sucursal = incidencias.filter(sucursal__nombre=sucursal)
        total = inc_sucursal.count()
        
        sucursales_stats.append({
            'sucursal': sucursal,
            'total': total,
            'abiertas': inc_sucursal.filter(estado='abierta').count(),
            'cerradas': inc_sucursal.filter(estado='cerrada').count(),
            'criticas': inc_sucursal.filter(grado_severidad='critico').count(),
            'reincidencias': inc_sucursal.filter(es_reincidencia=True).count(),
            'porcentaje_total': round((total / incidencias.count() * 100), 2) if incidencias.count() > 0 else 0
        })
    
    return sorted(sucursales_stats, key=lambda x: x['total'], reverse=True)


def calcular_analisis_temporal(incidencias):
    """
    Calcula análisis temporal por Quarter, mes y semana
    
    Returns:
        dict con análisis por período
    """
    hoy = datetime.now()
    
    # Calcular Quarter actual
    quarter_actual = (hoy.month - 1) // 3 + 1
    
    # Inicio y fin del quarter actual
    mes_inicio_q = (quarter_actual - 1) * 3 + 1
    fecha_inicio_q = datetime(hoy.year, mes_inicio_q, 1)
    
    # Último mes
    primer_dia_mes = datetime(hoy.year, hoy.month, 1)
    
    # Última semana
    hace_7_dias = hoy - timedelta(days=7)
    
    return {
        'quarter_actual': {
            'numero': f'Q{quarter_actual}',
            'total': incidencias.filter(fecha_deteccion__gte=fecha_inicio_q).count(),
            'criticas': incidencias.filter(
                fecha_deteccion__gte=fecha_inicio_q,
                grado_severidad='critico'
            ).count()
        },
        'ultimo_mes': {
            'total': incidencias.filter(fecha_deteccion__gte=primer_dia_mes).count(),
            'criticas': incidencias.filter(
                fecha_deteccion__gte=primer_dia_mes,
                grado_severidad='critico'
            ).count()
        },
        'ultima_semana': {
            'total': incidencias.filter(fecha_deteccion__gte=hace_7_dias).count(),
            'criticas': incidencias.filter(
                fecha_deteccion__gte=hace_7_dias,
                grado_severidad='critico'
            ).count()
        }
    }


def calcular_estadisticas_por_empleado(incidencias):
    """
    Calcula estadísticas consolidadas por empleado
    
    Returns:
        Lista de diccionarios con estadísticas por técnico
    """
    empleados_stats = []
    
    tecnicos = incidencias.values(
        'tecnico_responsable__id',
        'tecnico_responsable__nombre_completo'
    ).distinct()
    
    for tecnico in tecnicos:
        if not tecnico['tecnico_responsable__id']:
            continue
            
        inc_tecnico = incidencias.filter(
            tecnico_responsable__id=tecnico['tecnico_responsable__id']
        )
        
        total = inc_tecnico.count()
        criticas = inc_tecnico.filter(grado_severidad='critico').count()
        reincidencias = inc_tecnico.filter(es_reincidencia=True).count()
        
        empleados_stats.append({
            'id': tecnico['tecnico_responsable__id'],
            'nombre': tecnico['tecnico_responsable__nombre_completo'],
            'total': total,
            'criticas': criticas,
            'reincidencias': reincidencias,
            'porcentaje_criticas': round((criticas / total * 100), 2) if total > 0 else 0,
            'porcentaje_reincidencias': round((reincidencias / total * 100), 2) if total > 0 else 0
        })
    
    return sorted(empleados_stats, key=lambda x: x['total'], reverse=True)
