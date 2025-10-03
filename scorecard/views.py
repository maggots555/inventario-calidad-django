"""
Vistas para el sistema de Score Card
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from .models import Incidencia, CategoriaIncidencia, ComponenteEquipo, EvidenciaIncidencia
from .forms import IncidenciaForm, EvidenciaIncidenciaForm
from .emails import enviar_notificacion_incidencia, obtener_destinatarios_disponibles
from inventario.models import Empleado, Sucursal
from datetime import datetime, timedelta
from collections import defaultdict
import json


def dashboard(request):
    """
    Dashboard principal con KPIs y gráficos
    """
    # Obtener estadísticas
    total_incidencias = Incidencia.objects.count()
    incidencias_abiertas = Incidencia.objects.filter(estado='abierta').count()
    incidencias_criticas = Incidencia.objects.filter(grado_severidad='critico').count()
    
    # Contexto para el template
    context = {
        'total_incidencias': total_incidencias,
        'incidencias_abiertas': incidencias_abiertas,
        'incidencias_criticas': incidencias_criticas,
    }
    
    return render(request, 'scorecard/dashboard.html', context)


def lista_incidencias(request):
    """
    Lista todas las incidencias con filtros
    """
    incidencias = Incidencia.objects.all().select_related(
        'sucursal',
        'tecnico_responsable',
        'inspector_calidad',
        'tipo_incidencia'
    )
    
    context = {
        'incidencias': incidencias,
    }
    
    return render(request, 'scorecard/lista_incidencias.html', context)


def detalle_incidencia(request, incidencia_id):
    """
    Detalle completo de una incidencia
    """
    incidencia = get_object_or_404(
        Incidencia.objects.select_related(
            'sucursal',
            'tecnico_responsable',
            'inspector_calidad',
            'tipo_incidencia',
            'componente_afectado'
        ),
        id=incidencia_id
    )
    
    # Obtener evidencias
    evidencias = incidencia.evidencias.all()
    
    # Obtener historial de notificaciones enviadas (ordenadas por más reciente)
    from .models import NotificacionIncidencia
    notificaciones_enviadas = NotificacionIncidencia.objects.filter(
        incidencia=incidencia
    ).order_by('-fecha_envio')
    
    context = {
        'incidencia': incidencia,
        'evidencias': evidencias,
        'notificaciones_enviadas': notificaciones_enviadas,
    }
    
    return render(request, 'scorecard/detalle_incidencia.html', context)


def crear_incidencia(request):
    """
    Crear una nueva incidencia con formulario completo
    NOTA: Todas las incidencias nuevas se crean con estado='abierta' automáticamente
    """
    if request.method == 'POST':
        form = IncidenciaForm(request.POST)
        
        if form.is_valid():
            incidencia = form.save(commit=False)
            
            # ESTADO AUTOMÁTICO: Toda nueva incidencia inicia como "Abierta"
            incidencia.estado = 'abierta'
            
            # Auto-completar el área del técnico desde el empleado seleccionado
            tecnico = form.cleaned_data['tecnico_responsable']
            if tecnico and tecnico.area:
                incidencia.area_tecnico = tecnico.area
            
            incidencia.save()
            
            # Manejar las imágenes subidas
            imagenes = request.FILES.getlist('evidencias')
            inspector = form.cleaned_data['inspector_calidad']
            
            for imagen in imagenes:
                EvidenciaIncidencia.objects.create(
                    incidencia=incidencia,
                    imagen=imagen,
                    descripcion='',
                    subido_por=inspector
                )
            
            messages.success(
                request,
                f'Incidencia {incidencia.folio} registrada exitosamente.'
            )
            return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = IncidenciaForm()
    
    # Obtener datos para autocompletado
    marcas_comunes = ['HP', 'Dell', 'Lenovo', 'Acer', 'ASUS', 'Toshiba', 'Apple', 'MSI']
    
    context = {
        'form': form,
        'marcas_comunes': marcas_comunes,
    }
    
    return render(request, 'scorecard/form_incidencia.html', context)


def editar_incidencia(request, incidencia_id):
    """
    Editar una incidencia existente
    NOTA: El estado NO se modifica aquí, se gestiona mediante formularios específicos
    """
    incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
    if request.method == 'POST':
        form = IncidenciaForm(request.POST, instance=incidencia)
        
        if form.is_valid():
            # Guardar el estado actual antes de hacer save
            estado_actual = incidencia.estado
            
            incidencia = form.save(commit=False)
            
            # PRESERVAR EL ESTADO: No permitir que se modifique al editar
            incidencia.estado = estado_actual
            
            # Auto-completar el área del técnico desde el empleado seleccionado
            tecnico = form.cleaned_data['tecnico_responsable']
            if tecnico and tecnico.area:
                incidencia.area_tecnico = tecnico.area
            
            incidencia.save()
            
            # Manejar nuevas imágenes si se suben
            imagenes = request.FILES.getlist('evidencias')
            inspector = form.cleaned_data['inspector_calidad']
            
            for imagen in imagenes:
                EvidenciaIncidencia.objects.create(
                    incidencia=incidencia,
                    imagen=imagen,
                    descripcion='',
                    subido_por=inspector
                )
            
            messages.success(
                request,
                f'Incidencia {incidencia.folio} actualizada exitosamente.'
            )
            return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = IncidenciaForm(instance=incidencia)
    
    # Obtener evidencias existentes
    evidencias = incidencia.evidencias.all()
    
    # Obtener datos para autocompletado
    marcas_comunes = ['HP', 'Dell', 'Lenovo', 'Acer', 'ASUS', 'Toshiba', 'Apple', 'MSI']
    
    context = {
        'form': form,
        'incidencia': incidencia,
        'evidencias': evidencias,
        'marcas_comunes': marcas_comunes,
        'editando': True,
    }
    
    return render(request, 'scorecard/form_incidencia.html', context)


def eliminar_incidencia(request, incidencia_id):
    """
    Eliminar una incidencia
    """
    incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
    if request.method == 'POST':
        folio = incidencia.folio
        incidencia.delete()
        messages.success(request, f'Incidencia {folio} eliminada exitosamente.')
        return redirect('scorecard:lista_incidencias')
    
    return render(request, 'scorecard/confirmar_eliminacion.html', {'incidencia': incidencia})


def reportes(request):
    """
    Página de reportes y análisis
    """
    return render(request, 'scorecard/reportes.html')


def lista_categorias(request):
    """
    Lista de categorías de incidencias
    """
    categorias = CategoriaIncidencia.objects.all()
    return render(request, 'scorecard/lista_categorias.html', {'categorias': categorias})


def lista_componentes(request):
    """
    Lista de componentes de equipos
    """
    componentes = ComponenteEquipo.objects.all()
    return render(request, 'scorecard/lista_componentes.html', {'componentes': componentes})


# APIs para JavaScript
def api_empleado_data(request, empleado_id):
    """
    API para obtener datos de un empleado (para autocompletar)
    """
    try:
        empleado = Empleado.objects.get(id=empleado_id)
        data = {
            'success': True,
            'empleado': {
                'id': empleado.id,
                'nombre': empleado.nombre_completo,
                'area': empleado.area,
                'cargo': empleado.cargo,
                'email': empleado.email or '',
                'sucursal_id': empleado.sucursal.id if empleado.sucursal else None,
                'sucursal_nombre': empleado.sucursal.nombre if empleado.sucursal else '',
            }
        }
    except Empleado.DoesNotExist:
        data = {'success': False, 'error': 'Empleado no encontrado'}
    
    return JsonResponse(data)


def api_buscar_reincidencias(request):
    """
    API para buscar incidencias previas por número de serie
    """
    numero_serie = request.GET.get('numero_serie', '').strip().upper()
    
    if not numero_serie:
        return JsonResponse({'success': False, 'error': 'Número de serie requerido'})
    
    # Buscar incidencias previas con el mismo número de serie
    incidencias_previas = Incidencia.objects.filter(
        numero_serie__iexact=numero_serie
    ).order_by('-fecha_deteccion')[:5]  # Últimas 5
    
    data = {
        'success': True,
        'count': incidencias_previas.count(),
        'incidencias': []
    }
    
    for inc in incidencias_previas:
        data['incidencias'].append({
            'id': inc.id,
            'folio': inc.folio,
            'fecha': inc.fecha_deteccion.strftime('%d/%m/%Y'),
            'tipo_equipo': inc.get_tipo_equipo_display(),
            'marca': inc.marca,
            'tecnico': inc.tecnico_responsable.nombre_completo,
            'categoria': inc.tipo_incidencia.nombre,
            'estado': inc.get_estado_display(),
            'severidad': inc.get_grado_severidad_display(),
        })
    
    return JsonResponse(data)


def api_componentes_por_tipo(request):
    """
    API para obtener componentes filtrados por tipo de equipo
    """
    tipo_equipo = request.GET.get('tipo', 'todos')
    
    componentes = ComponenteEquipo.objects.filter(
        Q(tipo_equipo=tipo_equipo) | Q(tipo_equipo='todos'),
        activo=True
    ).order_by('nombre')
    
    data = {
        'success': True,
        'componentes': [
            {
                'id': comp.id,
                'nombre': comp.nombre,
                'tipo': comp.get_tipo_equipo_display()
            }
            for comp in componentes
        ]
    }
    
    return JsonResponse(data)


# === APIs para Gráficos y Reportes (Fase 3) ===

def api_datos_dashboard(request):
    """
    API que proporciona todos los datos para gráficos del dashboard
    Retorna: tendencias, distribuciones, rankings, etc.
    """
    # 1. Tendencia mensual (últimos 6 meses)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    incidencias_recientes = Incidencia.objects.filter(
        fecha_deteccion__gte=hace_6_meses
    )
    
    # Agrupar por mes
    meses_data = defaultdict(int)
    for inc in incidencias_recientes:
        mes_key = inc.fecha_deteccion.strftime('%Y-%m')
        meses_data[mes_key] += 1
    
    # Ordenar por fecha
    meses_ordenados = sorted(meses_data.items())
    
    tendencia_mensual = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes, _ in meses_ordenados],
        'data': [count for _, count in meses_ordenados]
    }
    
    # 2. Distribución por severidad
    severidad_counts = Incidencia.objects.values('grado_severidad').annotate(
        count=Count('id')
    ).order_by('grado_severidad')
    
    severidad_labels = {
        'critico': 'Crítico',
        'alto': 'Alto',
        'medio': 'Medio',
        'bajo': 'Bajo'
    }
    
    severidad_colors = {
        'critico': '#dc3545',
        'alto': '#fd7e14',
        'medio': '#ffc107',
        'bajo': '#28a745'
    }
    
    distribucion_severidad = {
        'labels': [severidad_labels.get(item['grado_severidad'], item['grado_severidad']) 
                   for item in severidad_counts],
        'data': [item['count'] for item in severidad_counts],
        'colors': [severidad_colors.get(item['grado_severidad'], '#6c757d') 
                   for item in severidad_counts]
    }
    
    # 3. Top 10 técnicos con más incidencias
    top_tecnicos = Incidencia.objects.values(
        'tecnico_responsable__nombre_completo'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    ranking_tecnicos = {
        'labels': [item['tecnico_responsable__nombre_completo'] or 'Sin técnico' 
                   for item in top_tecnicos],
        'data': [item['total'] for item in top_tecnicos]
    }
    
    # 4. Distribución por categoría
    categorias_counts = Incidencia.objects.values(
        'tipo_incidencia__nombre',
        'tipo_incidencia__color'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    distribucion_categorias = {
        'labels': [item['tipo_incidencia__nombre'] or 'Sin categoría' 
                   for item in categorias_counts],
        'data': [item['count'] for item in categorias_counts],
        'colors': [item['tipo_incidencia__color'] or '#6c757d' 
                   for item in categorias_counts]
    }
    
    # 5. Análisis por sucursal
    sucursales_counts = Incidencia.objects.values(
        'sucursal__nombre'
    ).annotate(
        total=Count('id')
    ).order_by('-total')
    
    analisis_sucursales = {
        'labels': [item['sucursal__nombre'] or 'Sin sucursal' 
                   for item in sucursales_counts],
        'data': [item['total'] for item in sucursales_counts]
    }
    
    # 6. Componentes más afectados
    componentes_counts = Incidencia.objects.values(
        'componente_afectado__nombre'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    componentes_afectados = {
        'labels': [item['componente_afectado__nombre'] or 'Sin componente' 
                   for item in componentes_counts],
        'data': [item['count'] for item in componentes_counts]
    }
    
    # 7. KPIs adicionales
    total_incidencias = Incidencia.objects.count()
    incidencias_abiertas = Incidencia.objects.filter(estado='abierta').count()
    incidencias_criticas = Incidencia.objects.filter(grado_severidad='critico').count()
    incidencias_cerradas = Incidencia.objects.filter(estado='cerrada').count()
    reincidencias = Incidencia.objects.filter(es_reincidencia=True).count()
    
    # Calcular porcentaje de reincidencias
    porcentaje_reincidencias = round((reincidencias / total_incidencias * 100), 2) if total_incidencias > 0 else 0
    
    # Promedio de días para cerrar incidencias
    incidencias_cerradas_obj = Incidencia.objects.filter(estado='cerrada')
    if incidencias_cerradas_obj.exists():
        dias_totales = sum([inc.dias_abierta for inc in incidencias_cerradas_obj])
        promedio_dias_cierre = round(dias_totales / incidencias_cerradas_obj.count(), 1)
    else:
        promedio_dias_cierre = 0
    
    kpis = {
        'total_incidencias': total_incidencias,
        'incidencias_abiertas': incidencias_abiertas,
        'incidencias_criticas': incidencias_criticas,
        'incidencias_cerradas': incidencias_cerradas,
        'reincidencias': reincidencias,
        'porcentaje_reincidencias': porcentaje_reincidencias,
        'promedio_dias_cierre': promedio_dias_cierre
    }
    
    # Retornar todos los datos
    data = {
        'success': True,
        'kpis': kpis,
        'tendencia_mensual': tendencia_mensual,
        'distribucion_severidad': distribucion_severidad,
        'ranking_tecnicos': ranking_tecnicos,
        'distribucion_categorias': distribucion_categorias,
        'analisis_sucursales': analisis_sucursales,
        'componentes_afectados': componentes_afectados
    }
    
    return JsonResponse(data)


def api_exportar_excel(request):
    """
    API para exportar incidencias a Excel
    Requiere: openpyxl instalado (pip install openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({
            'success': False,
            'error': 'Librería openpyxl no instalada. Ejecuta: pip install openpyxl'
        })
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    
    # Encabezados
    headers = [
        'Folio', 'Fecha Detección', 'Tipo Equipo', 'Marca', 'Modelo',
        'No. Serie', 'Sucursal', 'Técnico', 'Inspector', 'Categoría',
        'Severidad', 'Estado', 'Componente', 'Descripción', 'Días Abierta'
    ]
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Obtener incidencias
    incidencias = Incidencia.objects.all().select_related(
        'sucursal', 'tecnico_responsable', 'inspector_calidad',
        'tipo_incidencia', 'componente_afectado'
    ).order_by('-fecha_deteccion')
    
    # Escribir datos
    for row_num, inc in enumerate(incidencias, 2):
        ws.cell(row=row_num, column=1).value = inc.folio
        ws.cell(row=row_num, column=2).value = inc.fecha_deteccion.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=3).value = inc.get_tipo_equipo_display()
        ws.cell(row=row_num, column=4).value = inc.marca
        ws.cell(row=row_num, column=5).value = inc.modelo
        ws.cell(row=row_num, column=6).value = inc.numero_serie
        ws.cell(row=row_num, column=7).value = inc.sucursal.nombre if inc.sucursal else ''
        ws.cell(row=row_num, column=8).value = inc.tecnico_responsable.nombre_completo if inc.tecnico_responsable else ''
        ws.cell(row=row_num, column=9).value = inc.inspector_calidad.nombre_completo if inc.inspector_calidad else ''
        ws.cell(row=row_num, column=10).value = inc.tipo_incidencia.nombre if inc.tipo_incidencia else ''
        ws.cell(row=row_num, column=11).value = inc.get_grado_severidad_display()
        ws.cell(row=row_num, column=12).value = inc.get_estado_display()
        ws.cell(row=row_num, column=13).value = inc.componente_afectado.nombre if inc.componente_afectado else ''
        ws.cell(row=row_num, column=14).value = inc.descripcion_incidencia[:100]  # Truncar para Excel
        ws.cell(row=row_num, column=15).value = inc.dias_abierta
    
    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=incidencias_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


# ==========================================
# SISTEMA DE NOTIFICACIONES POR EMAIL
# ==========================================

@require_http_methods(["GET"])
def api_obtener_destinatarios(request, incidencia_id):
    """
    API: Obtiene la lista de destinatarios disponibles para una incidencia
    GET /scorecard/api/incidencias/<id>/destinatarios/
    
    Retorna JSON con lista de destinatarios disponibles
    """
    try:
        incidencia = get_object_or_404(Incidencia, id=incidencia_id)
        destinatarios = obtener_destinatarios_disponibles(incidencia)
        
        return JsonResponse({
            'success': True,
            'destinatarios': destinatarios
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener destinatarios: {str(e)}'
        }, status=400)


@require_http_methods(["POST"])
def api_enviar_notificacion(request, incidencia_id):
    """
    API: Envía notificación por email sobre una incidencia
    POST /scorecard/api/incidencias/<id>/enviar-notificacion/
    
    Body (JSON):
    {
        "destinatarios": [
            {"nombre": "Juan Pérez", "email": "juan@email.com", "rol": "Técnico"},
            ...
        ],
        "mensaje_adicional": "Texto opcional"
    }
    
    Retorna JSON con resultado del envío
    """
    try:
        incidencia = get_object_or_404(Incidencia, id=incidencia_id)
        
        # Validar que la incidencia NO esté cerrada
        if incidencia.estado == 'cerrada':
            return JsonResponse({
                'success': False,
                'message': 'No se puede enviar notificación para una incidencia cerrada'
            }, status=400)
        
        # Parsear el body JSON
        data = json.loads(request.body)
        destinatarios_seleccionados = data.get('destinatarios', [])
        mensaje_adicional = data.get('mensaje_adicional', '')
        
        # Validar que hay destinatarios
        if not destinatarios_seleccionados:
            return JsonResponse({
                'success': False,
                'message': 'Debe seleccionar al menos un destinatario'
            }, status=400)
        
        # Enviar la notificación
        resultado = enviar_notificacion_incidencia(
            incidencia=incidencia,
            destinatarios_seleccionados=destinatarios_seleccionados,
            mensaje_adicional=mensaje_adicional,
            enviado_por=request.user.username if request.user.is_authenticated else 'Sistema'
        )
        
        # CAMBIO AUTOMÁTICO DE ESTADO: Si la notificación se envió correctamente
        # y la incidencia está en estado "abierta", cambiarla a "en_revision"
        if resultado['success'] and incidencia.estado == 'abierta':
            incidencia.estado = 'en_revision'
            incidencia.save()
            resultado['estado_cambiado'] = True
            resultado['nuevo_estado'] = 'En Revisión'
            resultado['message'] += ' | Estado cambiado automáticamente a "En Revisión".'
        
        if resultado['success']:
            return JsonResponse(resultado)
        else:
            return JsonResponse(resultado, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error al parsear JSON del request'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error inesperado: {str(e)}'
        }, status=500)


def cambiar_estado_incidencia(request, incidencia_id):
    """
    Vista para cambiar el estado de una incidencia manualmente
    
    ESTADOS PERMITIDOS:
    - 'en_revision': Solo para casos excepcionales (sin enviar notificación)
    - 'reincidente': Para marcar/desmarcar reincidencias
    
    NOTAS:
    - Esta vista NO permite cerrar incidencias (usar formulario de cierre)
    - Esta vista NO permite volver a "abierta" (crear nueva si es necesario)
    - El cambio a "en_revision" se hace automáticamente al enviar notificación manual
    - NO se puede usar en incidencias cerradas
    """
    incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
    # Validar que la incidencia NO esté cerrada
    if incidencia.estado == 'cerrada':
        messages.warning(request, 'No se puede cambiar el estado de una incidencia cerrada.')
        return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    
    if request.method == 'POST':
        from .forms import CambiarEstadoIncidenciaForm
        form = CambiarEstadoIncidenciaForm(request.POST)
        
        if form.is_valid():
            nuevo_estado = form.cleaned_data['estado']
            notas = form.cleaned_data.get('notas', '')
            
            # Guardar estado anterior
            estado_anterior = incidencia.get_estado_display()
            
            # Actualizar estado
            incidencia.estado = nuevo_estado
            
            # Si hay notas, agregarlas a acciones_tomadas
            if notas:
                if incidencia.acciones_tomadas:
                    incidencia.acciones_tomadas += f"\n\n--- Cambio de estado ({timezone.now().strftime('%d/%m/%Y %H:%M')}) ---\n{notas}"
                else:
                    incidencia.acciones_tomadas = notas
            
            incidencia.save()
            
            messages.success(request, f'Estado cambiado de "{estado_anterior}" a "{incidencia.get_estado_display()}"')
            
            return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    else:
        from .forms import CambiarEstadoIncidenciaForm
        form = CambiarEstadoIncidenciaForm(initial={'estado': incidencia.estado})
    
    context = {
        'incidencia': incidencia,
        'form': form,
    }
    
    return render(request, 'scorecard/cambiar_estado.html', context)


def marcar_no_atribuible(request, incidencia_id):
    """
    Vista para marcar una incidencia como NO atribuible al técnico
    Requiere justificación y envía notificación automática
    """
    incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
    # Validar que no esté ya marcada como no atribuible
    if not incidencia.es_atribuible:
        messages.warning(request, 'Esta incidencia ya está marcada como NO atribuible.')
        return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    
    if request.method == 'POST':
        from .forms import MarcarNoAtribuibleForm
        form = MarcarNoAtribuibleForm(request.POST)
        
        if form.is_valid():
            justificacion = form.cleaned_data['justificacion']
            
            # Actualizar la incidencia
            incidencia.es_atribuible = False
            incidencia.justificacion_no_atribuible = justificacion
            incidencia.fecha_marcado_no_atribuible = timezone.now()
            
            # Intentar obtener el empleado que marca (si está logueado)
            if request.user.is_authenticated:
                try:
                    from inventario.models import Empleado
                    empleado = Empleado.objects.filter(email=request.user.email).first()
                    if empleado:
                        incidencia.marcado_no_atribuible_por = empleado
                except:
                    pass
            
            incidencia.save()
            
            # Enviar notificación automática al técnico
            from .emails import enviar_notificacion_no_atribuible
            resultado = enviar_notificacion_no_atribuible(
                incidencia=incidencia,
                justificacion=justificacion,
                marcado_por=request.user.username if request.user.is_authenticated else 'Sistema'
            )
            
            if resultado['success']:
                messages.success(request, f'Incidencia marcada como NO atribuible y técnico notificado: {resultado["message"]}')
            else:
                messages.warning(request, f'Incidencia marcada como NO atribuible, pero hubo un error en la notificación: {resultado["message"]}')
            
            return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    else:
        from .forms import MarcarNoAtribuibleForm
        form = MarcarNoAtribuibleForm()
    
    # Obtener destinatarios históricos para mostrar en el formulario
    from .emails import obtener_destinatarios_historicos
    destinatarios_historicos = obtener_destinatarios_historicos(incidencia)
    
    context = {
        'incidencia': incidencia,
        'form': form,
        'destinatarios_historicos': destinatarios_historicos,
    }
    
    return render(request, 'scorecard/marcar_no_atribuible.html', context)


def cerrar_incidencia(request, incidencia_id):
    """
    Vista para cerrar una incidencia con formulario de cierre
    Envía notificación automática (diferente según sea atribuible o no)
    """
    incidencia = get_object_or_404(Incidencia, id=incidencia_id)
    
    # Validar que no esté ya cerrada
    if incidencia.estado == 'cerrada':
        messages.warning(request, 'Esta incidencia ya está cerrada.')
        return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    
    if request.method == 'POST':
        from .forms import CerrarIncidenciaForm
        form = CerrarIncidenciaForm(request.POST)
        
        if form.is_valid():
            acciones_tomadas = form.cleaned_data['acciones_tomadas']
            causa_raiz = form.cleaned_data.get('causa_raiz', '')
            
            # Actualizar la incidencia
            incidencia.estado = 'cerrada'
            incidencia.acciones_tomadas = acciones_tomadas
            if causa_raiz:
                incidencia.causa_raiz = causa_raiz
            incidencia.fecha_cierre = timezone.now()
            incidencia.save()
            
            # Enviar notificación de cierre
            from .emails import enviar_notificacion_cierre_incidencia
            mensaje_adicional = acciones_tomadas
            if causa_raiz:
                mensaje_adicional += f"\n\nCausa Raíz:\n{causa_raiz}"
            
            resultado = enviar_notificacion_cierre_incidencia(
                incidencia=incidencia,
                mensaje_adicional=mensaje_adicional,
                enviado_por=request.user.username if request.user.is_authenticated else 'Sistema'
            )
            
            if resultado['success']:
                tipo_cierre = "NO atribuible" if not incidencia.es_atribuible else "normal"
                messages.success(request, f'Incidencia cerrada ({tipo_cierre}) y técnico notificado: {resultado["message"]}')
            else:
                messages.warning(request, f'Incidencia cerrada, pero hubo un error en la notificación: {resultado["message"]}')
            
            return redirect('scorecard:detalle_incidencia', incidencia_id=incidencia.id)
    else:
        from .forms import CerrarIncidenciaForm
        # Pre-cargar con los datos existentes si los hay
        initial_data = {}
        if incidencia.acciones_tomadas:
            initial_data['acciones_tomadas'] = incidencia.acciones_tomadas
        if incidencia.causa_raiz:
            initial_data['causa_raiz'] = incidencia.causa_raiz
        
        form = CerrarIncidenciaForm(initial=initial_data)
    
    # Obtener destinatarios históricos para mostrar en el formulario
    from .emails import obtener_destinatarios_historicos
    destinatarios_historicos = obtener_destinatarios_historicos(incidencia)
    
    context = {
        'incidencia': incidencia,
        'form': form,
        'destinatarios_historicos': destinatarios_historicos,
    }
    
    return render(request, 'scorecard/cerrar_incidencia.html', context)


# ==========================================
# APIS PARA REPORTES AVANZADOS - FASE 2
# ==========================================

def aplicar_filtros_reporte(queryset, request):
    """
    Función auxiliar para aplicar filtros a las consultas de reportes
    Recibe un queryset de Incidencia y los parámetros GET del request
    Retorna el queryset filtrado
    """
    # Filtro por rango de fechas
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        queryset = queryset.filter(fecha_deteccion__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_deteccion__lte=fecha_fin)
    
    # Filtro por sucursal
    sucursal = request.GET.get('sucursal')
    if sucursal:
        queryset = queryset.filter(equipo__sucursal__nombre=sucursal)
    
    # Filtro por técnico
    tecnico = request.GET.get('tecnico')
    if tecnico:
        queryset = queryset.filter(tecnico_responsable__nombre_completo=tecnico)
    
    # Filtro por área detectora
    area = request.GET.get('area')
    if area:
        queryset = queryset.filter(area_detectora=area)
    
    # Filtro por severidad
    severidad = request.GET.get('severidad')
    if severidad:
        queryset = queryset.filter(severidad=severidad)
    
    # Filtro por estado
    estado = request.GET.get('estado')
    if estado:
        queryset = queryset.filter(estado=estado)
    
    return queryset


def api_analisis_atribuibilidad(request):
    """
    API: Análisis completo de atribuibilidad de incidencias
    Retorna datos sobre incidencias atribuibles vs no atribuibles
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    # Obtener queryset base y aplicar filtros
    incidencias = Incidencia.objects.all()
    incidencias = aplicar_filtros_reporte(incidencias, request)
    
    # Total de incidencias (filtradas)
    total_incidencias = incidencias.count()
    
    # Contadores de atribuibilidad
    atribuibles = incidencias.filter(es_atribuible=True).count()
    no_atribuibles = incidencias.filter(es_atribuible=False).count()
    
    # Porcentajes
    porcentaje_atribuibles = round((atribuibles / total_incidencias * 100), 2) if total_incidencias > 0 else 0
    porcentaje_no_atribuibles = round((no_atribuibles / total_incidencias * 100), 2) if total_incidencias > 0 else 0
    
    # Gráfico de dona: Atribuibles vs No Atribuibles
    distribucion_atribuibilidad = {
        'labels': ['Atribuibles', 'No Atribuibles'],
        'data': [atribuibles, no_atribuibles],
        'colors': ['#4CAF50', '#FF9800']
    }
    
    # Tendencia mensual de atribuibilidad (últimos 6 meses)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    # Aplicar filtros también a las incidencias recientes
    incidencias_recientes = incidencias.filter(
        fecha_deteccion__gte=hace_6_meses
    ).order_by('fecha_deteccion')
    
    # Agrupar por mes
    meses_atribuibles = defaultdict(int)
    meses_no_atribuibles = defaultdict(int)
    
    for inc in incidencias_recientes:
        mes_key = inc.fecha_deteccion.strftime('%Y-%m')
        if inc.es_atribuible:
            meses_atribuibles[mes_key] += 1
        else:
            meses_no_atribuibles[mes_key] += 1
    
    # Obtener todos los meses en el rango
    meses_unicos = sorted(set(list(meses_atribuibles.keys()) + list(meses_no_atribuibles.keys())))
    
    tendencia_atribuibilidad = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes in meses_unicos],
        'atribuibles': [meses_atribuibles[mes] for mes in meses_unicos],
        'no_atribuibles': [meses_no_atribuibles[mes] for mes in meses_unicos]
    }
    
    # Top técnicos con más incidencias NO atribuibles (con filtros aplicados)
    top_tecnicos_no_atribuibles = incidencias.filter(
        es_atribuible=False
    ).values(
        'tecnico_responsable__nombre_completo'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    ranking_no_atribuibles = {
        'labels': [item['tecnico_responsable__nombre_completo'] or 'Sin técnico' 
                   for item in top_tecnicos_no_atribuibles],
        'data': [item['total'] for item in top_tecnicos_no_atribuibles]
    }
    
    # Obtener justificaciones de NO atribuibilidad (últimas 10, con filtros aplicados)
    justificaciones_recientes = incidencias.filter(
        es_atribuible=False,
        justificacion_no_atribuible__isnull=False
    ).exclude(
        justificacion_no_atribuible=''
    ).order_by('-fecha_marcado_no_atribuible')[:10]
    
    lista_justificaciones = []
    for inc in justificaciones_recientes:
        lista_justificaciones.append({
            'folio': inc.folio,
            'fecha': inc.fecha_marcado_no_atribuible.strftime('%d/%m/%Y %H:%M') if inc.fecha_marcado_no_atribuible else 'N/A',
            'tecnico': inc.tecnico_responsable.nombre_completo if inc.tecnico_responsable else 'N/A',
            'justificacion': inc.justificacion_no_atribuible[:150] + '...' if len(inc.justificacion_no_atribuible) > 150 else inc.justificacion_no_atribuible,
            'marcado_por': inc.marcado_no_atribuible_por.nombre_completo if inc.marcado_no_atribuible_por else 'Sistema'
        })
    
    # Análisis de razones comunes (palabras clave en justificaciones, con filtros aplicados)
    todas_justificaciones = incidencias.filter(
        es_atribuible=False,
        justificacion_no_atribuible__isnull=False
    ).exclude(justificacion_no_atribuible='').values_list('justificacion_no_atribuible', flat=True)
    
    # Categorizar justificaciones por palabras clave
    razones = {
        'Fallo de fábrica': 0,
        'Componente defectuoso': 0,
        'Error del cliente': 0,
        'Falta de refacción': 0,
        'Condiciones externas': 0,
        'Otros': 0
    }
    
    palabras_clave = {
        'Fallo de fábrica': ['fábrica', 'fabrica', 'fabricación', 'manufactura', 'defecto de origen'],
        'Componente defectuoso': ['defectuoso', 'defecto', 'dañado', 'componente malo', 'pieza mala'],
        'Error del cliente': ['cliente', 'usuario', 'mal uso', 'maltrato', 'golpe'],
        'Falta de refacción': ['refacción', 'repuesto', 'sin pieza', 'no disponible', 'agotado'],
        'Condiciones externas': ['humedad', 'polvo', 'temperatura', 'voltaje', 'ambiente']
    }
    
    for justificacion in todas_justificaciones:
        justificacion_lower = justificacion.lower()
        categorizada = False
        
        for razon, keywords in palabras_clave.items():
            if any(keyword in justificacion_lower for keyword in keywords):
                razones[razon] += 1
                categorizada = True
                break
        
        if not categorizada:
            razones['Otros'] += 1
    
    distribucion_razones = {
        'labels': list(razones.keys()),
        'data': list(razones.values()),
        'colors': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#C9CBCF']
    }
    
    # Retornar todos los datos
    data = {
        'success': True,
        'atribuibilidad': {
            'total': total_incidencias,
            'atribuibles': atribuibles,
            'no_atribuibles': no_atribuibles,
            'porcentaje_atribuibles': porcentaje_atribuibles,
            'porcentaje_no_atribuibles': porcentaje_no_atribuibles
        },
        'distribucion_atribuibilidad': distribucion_atribuibilidad,
        'tendencia_atribuibilidad': tendencia_atribuibilidad,
        'ranking_no_atribuibles': ranking_no_atribuibles,
        'justificaciones_recientes': lista_justificaciones,
        'distribucion_razones': distribucion_razones
    }
    
    return JsonResponse(data)


def api_analisis_tecnicos(request):
    """
    API: Análisis detallado por técnico con múltiples métricas
    Retorna scorecard completo de cada técnico
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    # Obtener queryset base y aplicar filtros
    incidencias_base = Incidencia.objects.all()
    incidencias_base = aplicar_filtros_reporte(incidencias_base, request)
    
    # Obtener todos los técnicos que tienen incidencias (en el conjunto filtrado)
    tecnicos_con_incidencias = Empleado.objects.filter(
        id__in=incidencias_base.values_list('tecnico_responsable', flat=True).distinct()
    ).distinct()
    
    scorecard_tecnicos = []
    
    for tecnico in tecnicos_con_incidencias:
        # Incidencias del técnico (ya filtradas)
        incidencias_tecnico = incidencias_base.filter(tecnico_responsable=tecnico)
        
        total_incidencias = incidencias_tecnico.count()
        
        if total_incidencias == 0:
            continue
        
        # Métricas del técnico
        criticas = incidencias_tecnico.filter(grado_severidad='critico').count()
        reincidencias = incidencias_tecnico.filter(es_reincidencia=True).count()
        atribuibles = incidencias_tecnico.filter(es_atribuible=True).count()
        no_atribuibles = incidencias_tecnico.filter(es_atribuible=False).count()
        cerradas = incidencias_tecnico.filter(estado='cerrada').count()
        
        # Porcentajes
        porcentaje_criticas = round((criticas / total_incidencias * 100), 2)
        porcentaje_reincidencias = round((reincidencias / total_incidencias * 100), 2)
        porcentaje_atribuibilidad = round((atribuibles / total_incidencias * 100), 2)
        tasa_cierre = round((cerradas / total_incidencias * 100), 2)
        
        # Promedio de días de resolución (solo cerradas)
        incidencias_cerradas = incidencias_tecnico.filter(estado='cerrada')
        if incidencias_cerradas.exists():
            dias_totales = sum([inc.dias_abierta for inc in incidencias_cerradas])
            promedio_dias = round(dias_totales / incidencias_cerradas.count(), 1)
        else:
            promedio_dias = 0
        
        # Calcular score de calidad (0-100)
        # Fórmula: 100 - (peso_criticas + peso_reincidencias + peso_no_atribuibles + peso_tiempo)
        peso_criticas = (porcentaje_criticas * 0.3)  # 30% del peso
        peso_reincidencias = (porcentaje_reincidencias * 0.3)  # 30% del peso
        peso_no_atribuibles = ((100 - porcentaje_atribuibilidad) * 0.25)  # 25% del peso
        peso_tiempo = min((promedio_dias / 10) * 15, 15)  # 15% del peso, max 10 días óptimo
        
        score_calidad = max(0, round(100 - (peso_criticas + peso_reincidencias + peso_no_atribuibles + peso_tiempo), 2))
        
        scorecard_tecnicos.append({
            'tecnico_id': tecnico.id,
            'tecnico': tecnico.nombre_completo,
            'area': tecnico.area,
            'sucursal': tecnico.sucursal.nombre if tecnico.sucursal else 'Sin sucursal',
            'total_incidencias': total_incidencias,
            'criticas': criticas,
            'porcentaje_criticas': porcentaje_criticas,
            'reincidencias': reincidencias,
            'porcentaje_reincidencias': porcentaje_reincidencias,
            'atribuibles': atribuibles,
            'no_atribuibles': no_atribuibles,
            'porcentaje_atribuibilidad': porcentaje_atribuibilidad,
            'cerradas': cerradas,
            'tasa_cierre': tasa_cierre,
            'promedio_dias': promedio_dias,
            'score_calidad': score_calidad
        })
    
    # Ordenar por score de calidad (de mayor a mejor)
    scorecard_tecnicos.sort(key=lambda x: x['score_calidad'], reverse=True)
    
    # Top 10 técnicos con mejor score
    top_10_mejores = scorecard_tecnicos[:10]
    
    # Top 10 técnicos con peor score (que requieren atención)
    top_10_atencion = sorted(scorecard_tecnicos, key=lambda x: x['score_calidad'])[:10]
    
    # Ranking por total de incidencias
    ranking_total = sorted(scorecard_tecnicos, key=lambda x: x['total_incidencias'], reverse=True)[:10]
    
    ranking_datos = {
        'labels': [t['tecnico'] for t in ranking_total],
        'total': [t['total_incidencias'] for t in ranking_total],
        'criticas': [t['criticas'] for t in ranking_total],
        'reincidencias': [t['reincidencias'] for t in ranking_total],
        'no_atribuibles': [t['no_atribuibles'] for t in ranking_total]
    }
    
    # Estadísticas generales
    if scorecard_tecnicos:
        promedio_score = round(sum(t['score_calidad'] for t in scorecard_tecnicos) / len(scorecard_tecnicos), 2)
        promedio_dias_todos = round(sum(t['promedio_dias'] for t in scorecard_tecnicos) / len(scorecard_tecnicos), 2)
    else:
        promedio_score = 0
        promedio_dias_todos = 0
    
    data = {
        'success': True,
        'scorecard_completo': scorecard_tecnicos,
        'top_10_mejores': top_10_mejores,
        'top_10_atencion': top_10_atencion,
        'ranking_datos': ranking_datos,
        'estadisticas': {
            'total_tecnicos': len(scorecard_tecnicos),
            'promedio_score': promedio_score,
            'promedio_dias_resolucion': promedio_dias_todos
        }
    }
    
    return JsonResponse(data)


def api_analisis_reincidencias(request):
    """
    API: Análisis profundo de reincidencias con cadenas de relaciones
    Retorna datos sobre incidencias reincidentes y sus patrones
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    # Obtener queryset base y aplicar filtros
    incidencias = Incidencia.objects.all()
    incidencias = aplicar_filtros_reporte(incidencias, request)
    
    # Total de incidencias y reincidencias (con filtros aplicados)
    total_incidencias = incidencias.count()
    total_reincidencias = incidencias.filter(es_reincidencia=True).count()
    porcentaje_reincidencias = round((total_reincidencias / total_incidencias * 100), 2) if total_incidencias > 0 else 0
    
    # 1. Incidencias con 2+ reincidencias (cadenas largas) - del conjunto filtrado
    # Encontrar incidencias originales que tienen múltiples reincidencias
    incidencias_originales = incidencias.filter(
        reincidencias__isnull=False
    ).annotate(
        num_reincidencias=Count('reincidencias')
    ).filter(num_reincidencias__gte=2).order_by('-num_reincidencias')[:10]
    
    cadenas_reincidencias = []
    for original in incidencias_originales:
        # Obtener todas las reincidencias de esta incidencia
        reincidencias_list = original.reincidencias.all().order_by('fecha_deteccion')
        
        cadenas_reincidencias.append({
            'folio_original': original.folio,
            'fecha_original': original.fecha_deteccion.strftime('%d/%m/%Y'),
            'numero_serie': original.numero_serie,
            'tipo_equipo': original.get_tipo_equipo_display(),
            'marca': original.marca,
            'tecnico': original.tecnico_responsable.nombre_completo if original.tecnico_responsable else 'N/A',
            'total_reincidencias': reincidencias_list.count(),
            'reincidencias': [
                {
                    'folio': r.folio,
                    'fecha': r.fecha_deteccion.strftime('%d/%m/%Y'),
                    'dias_desde_original': (r.fecha_deteccion - original.fecha_deteccion).days,
                    'estado': r.get_estado_display(),
                    'severidad': r.get_grado_severidad_display()
                }
                for r in reincidencias_list
            ]
        })
    
    # 2. Top 10 equipos (números de serie) con más reincidencias
    top_equipos = Incidencia.objects.filter(
        es_reincidencia=True
    ).values('numero_serie', 'marca', 'tipo_equipo').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    top_equipos_reincidentes = {
        'labels': [f"{item['marca']} - {item['numero_serie'][:10]}" for item in top_equipos],
        'data': [item['total'] for item in top_equipos],
        'detalles': [
            {
                'numero_serie': item['numero_serie'],
                'marca': item['marca'],
                'tipo_equipo': item['tipo_equipo'],
                'total_reincidencias': item['total']
            }
            for item in top_equipos
        ]
    }
    
    # 3. % Reincidencias por técnico (del conjunto filtrado)
    tecnicos_reincidencias = incidencias.values(
        'tecnico_responsable__nombre_completo'
    ).annotate(
        total_incidencias=Count('id'),
        total_reincidencias=Count('id', filter=Q(es_reincidencia=True))
    ).filter(total_incidencias__gt=0).order_by('-total_reincidencias')[:10]
    
    ranking_reincidencias_tecnico = {
        'labels': [item['tecnico_responsable__nombre_completo'] or 'Sin técnico' for item in tecnicos_reincidencias],
        'reincidencias': [item['total_reincidencias'] for item in tecnicos_reincidencias],
        'porcentajes': [
            round((item['total_reincidencias'] / item['total_incidencias'] * 100), 2) if item['total_incidencias'] > 0 else 0
            for item in tecnicos_reincidencias
        ]
    }
    
    # 4. Tendencia mensual de reincidencias (últimos 6 meses, con filtros aplicados)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    reincidencias_recientes = incidencias.filter(
        fecha_deteccion__gte=hace_6_meses,
        es_reincidencia=True
    ).order_by('fecha_deteccion')
    
    meses_reincidencias = defaultdict(int)
    for inc in reincidencias_recientes:
        mes_key = inc.fecha_deteccion.strftime('%Y-%m')
        meses_reincidencias[mes_key] += 1
    
    meses_ordenados = sorted(meses_reincidencias.items())
    
    tendencia_reincidencias = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes, _ in meses_ordenados],
        'data': [count for _, count in meses_ordenados]
    }
    
    # 5. Tiempo promedio entre reincidencias (del conjunto filtrado)
    tiempos_entre_reincidencias = []
    for inc in incidencias.filter(es_reincidencia=True, incidencia_relacionada__isnull=False):
        if inc.incidencia_relacionada:
            dias_entre = (inc.fecha_deteccion - inc.incidencia_relacionada.fecha_deteccion).days
            tiempos_entre_reincidencias.append(dias_entre)
    
    if tiempos_entre_reincidencias:
        tiempo_promedio_entre_reincidencias = round(sum(tiempos_entre_reincidencias) / len(tiempos_entre_reincidencias), 1)
        tiempo_minimo = min(tiempos_entre_reincidencias)
        tiempo_maximo = max(tiempos_entre_reincidencias)
    else:
        tiempo_promedio_entre_reincidencias = 0
        tiempo_minimo = 0
        tiempo_maximo = 0
    
    # 6. Distribución por categoría de fallo en reincidencias (del conjunto filtrado)
    categorias_reincidencias = incidencias.filter(
        es_reincidencia=True
    ).values('categoria_fallo').annotate(
        count=Count('id')
    ).order_by('-count')
    
    distribucion_categorias_reincidencias = {
        'labels': [dict(Incidencia.CATEGORIA_FALLO_CHOICES).get(item['categoria_fallo'], item['categoria_fallo']) for item in categorias_reincidencias],
        'data': [item['count'] for item in categorias_reincidencias],
        'colors': ['#dc3545', '#ffc107', '#17a2b8', '#28a745', '#6c757d', '#fd7e14']
    }
    
    # Retornar todos los datos
    data = {
        'success': True,
        'kpis': {
            'total_reincidencias': total_reincidencias,
            'porcentaje_reincidencias': porcentaje_reincidencias,
            'tiempo_promedio_entre_reincidencias': tiempo_promedio_entre_reincidencias,
            'tiempo_minimo': tiempo_minimo,
            'tiempo_maximo': tiempo_maximo,
            'total_cadenas_largas': len(cadenas_reincidencias)
        },
        'cadenas_reincidencias': cadenas_reincidencias,
        'top_equipos_reincidentes': top_equipos_reincidentes,
        'ranking_reincidencias_tecnico': ranking_reincidencias_tecnico,
        'tendencia_reincidencias': tendencia_reincidencias,
        'distribucion_categorias_reincidencias': distribucion_categorias_reincidencias
    }
    
    return JsonResponse(data)


def api_analisis_tiempos(request):
    """
    API: Análisis de tiempos de resolución por estados
    Retorna métricas de tiempo de todas las incidencias
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    # Obtener queryset base y aplicar filtros
    incidencias = Incidencia.objects.all()
    incidencias = aplicar_filtros_reporte(incidencias, request)
    
    # 1. Análisis de tiempos de cierre (del conjunto filtrado)
    incidencias_cerradas = incidencias.filter(estado='cerrada', fecha_cierre__isnull=False)
    
    tiempos_cierre = []
    for inc in incidencias_cerradas:
        dias = inc.dias_abierta
        tiempos_cierre.append(dias)
    
    if tiempos_cierre:
        tiempo_promedio_cierre = round(sum(tiempos_cierre) / len(tiempos_cierre), 1)
        tiempo_minimo_cierre = min(tiempos_cierre)
        tiempo_maximo_cierre = max(tiempos_cierre)
        
        # Calcular mediana
        tiempos_ordenados = sorted(tiempos_cierre)
        n = len(tiempos_ordenados)
        mediana_cierre = tiempos_ordenados[n // 2] if n % 2 != 0 else (tiempos_ordenados[n // 2 - 1] + tiempos_ordenados[n // 2]) / 2
    else:
        tiempo_promedio_cierre = 0
        tiempo_minimo_cierre = 0
        tiempo_maximo_cierre = 0
        mediana_cierre = 0
    
    # 2. Distribución de tiempos (histograma)
    # Rangos: 0-3 días, 4-7 días, 8-15 días, 16-30 días, 31+ días
    rangos_tiempos = {
        '0-3 días': 0,
        '4-7 días': 0,
        '8-15 días': 0,
        '16-30 días': 0,
        '31+ días': 0
    }
    
    for dias in tiempos_cierre:
        if dias <= 3:
            rangos_tiempos['0-3 días'] += 1
        elif dias <= 7:
            rangos_tiempos['4-7 días'] += 1
        elif dias <= 15:
            rangos_tiempos['8-15 días'] += 1
        elif dias <= 30:
            rangos_tiempos['16-30 días'] += 1
        else:
            rangos_tiempos['31+ días'] += 1
    
    distribucion_tiempos = {
        'labels': list(rangos_tiempos.keys()),
        'data': list(rangos_tiempos.values()),
        'colors': ['#28a745', '#4CAF50', '#ffc107', '#ff9800', '#dc3545']
    }
    
    # 3. Técnicos más rápidos vs más lentos
    tecnicos_tiempos = []
    
    for tecnico in Empleado.objects.filter(
        id__in=incidencias_cerradas.values_list('tecnico_responsable', flat=True).distinct()
    ).distinct():
        incidencias_tecnico_cerradas = incidencias_cerradas.filter(
            tecnico_responsable=tecnico
        )
        
        if incidencias_tecnico_cerradas.exists():
            dias_totales = sum([inc.dias_abierta for inc in incidencias_tecnico_cerradas])
            promedio_dias = round(dias_totales / incidencias_tecnico_cerradas.count(), 1)
            
            tecnicos_tiempos.append({
                'tecnico': tecnico.nombre_completo,
                'promedio_dias': promedio_dias,
                'total_cerradas': incidencias_tecnico_cerradas.count()
            })
    
    # Ordenar por promedio de días
    tecnicos_tiempos.sort(key=lambda x: x['promedio_dias'])
    
    # Top 10 más rápidos
    top_rapidos = tecnicos_tiempos[:10]
    ranking_rapidos = {
        'labels': [t['tecnico'] for t in top_rapidos],
        'data': [t['promedio_dias'] for t in top_rapidos],
        'totales': [t['total_cerradas'] for t in top_rapidos]
    }
    
    # Top 10 más lentos
    top_lentos = tecnicos_tiempos[-10:][::-1]  # Invertir para mostrar del más lento al menos lento
    ranking_lentos = {
        'labels': [t['tecnico'] for t in top_lentos],
        'data': [t['promedio_dias'] for t in top_lentos],
        'totales': [t['total_cerradas'] for t in top_lentos]
    }
    
    # 4. Tendencia mensual de tiempo promedio de cierre (últimos 6 meses)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    incidencias_recientes_cerradas = Incidencia.objects.filter(
        fecha_cierre__gte=hace_6_meses,
        estado='cerrada',
        fecha_cierre__isnull=False
    ).order_by('fecha_cierre')
    
    meses_tiempos = defaultdict(list)
    for inc in incidencias_recientes_cerradas:
        mes_key = inc.fecha_cierre.strftime('%Y-%m')
        meses_tiempos[mes_key].append(inc.dias_abierta)
    
    # Calcular promedio por mes
    meses_promedios = []
    for mes, dias_list in sorted(meses_tiempos.items()):
        promedio_mes = round(sum(dias_list) / len(dias_list), 1)
        meses_promedios.append((mes, promedio_mes))
    
    tendencia_tiempos = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes, _ in meses_promedios],
        'data': [promedio for _, promedio in meses_promedios]
    }
    
    # 5. Análisis por severidad (del conjunto filtrado)
    tiempos_por_severidad = {}
    for severidad_key, severidad_label in Incidencia.GRADO_SEVERIDAD_CHOICES:
        incidencias_sev = incidencias_cerradas.filter(
            grado_severidad=severidad_key
        )
        
        if incidencias_sev.exists():
            dias_totales = sum([inc.dias_abierta for inc in incidencias_sev])
            promedio = round(dias_totales / incidencias_sev.count(), 1)
        else:
            promedio = 0
        
        tiempos_por_severidad[severidad_label] = promedio
    
    analisis_por_severidad = {
        'labels': list(tiempos_por_severidad.keys()),
        'data': list(tiempos_por_severidad.values()),
        'colors': ['#dc3545', '#fd7e14', '#ffc107', '#28a745']
    }
    
    # 6. Incidencias abiertas hace mucho tiempo (alertas)
    # 6. Alertas de tiempo - incidencias abiertas con más de 15 días (del conjunto filtrado)
    incidencias_abiertas = incidencias.filter(
        Q(estado='abierta') | Q(estado='en_revision')
    ).order_by('fecha_deteccion')
    
    alertas_tiempos = []
    for inc in incidencias_abiertas:
        dias_abierta = inc.dias_abierta
        
        # Alertar si lleva más de 15 días abierta
        if dias_abierta > 15:
            alertas_tiempos.append({
                'folio': inc.folio,
                'dias_abierta': dias_abierta,
                'fecha_deteccion': inc.fecha_deteccion.strftime('%d/%m/%Y'),
                'tecnico': inc.tecnico_responsable.nombre_completo if inc.tecnico_responsable else 'N/A',
                'severidad': inc.get_grado_severidad_display(),
                'estado': inc.get_estado_display(),
                'tipo_equipo': inc.get_tipo_equipo_display(),
                'marca': inc.marca
            })
    
    # Ordenar por días abierta (descendente)
    alertas_tiempos.sort(key=lambda x: x['dias_abierta'], reverse=True)
    
    # Retornar todos los datos
    data = {
        'success': True,
        'kpis': {
            'tiempo_promedio_cierre': tiempo_promedio_cierre,
            'tiempo_minimo_cierre': tiempo_minimo_cierre,
            'tiempo_maximo_cierre': tiempo_maximo_cierre,
            'mediana_cierre': mediana_cierre,
            'total_cerradas': len(tiempos_cierre),
            'total_alertas': len(alertas_tiempos)
        },
        'distribucion_tiempos': distribucion_tiempos,
        'ranking_rapidos': ranking_rapidos,
        'ranking_lentos': ranking_lentos,
        'tendencia_tiempos': tendencia_tiempos,
        'analisis_por_severidad': analisis_por_severidad,
        'alertas_tiempos': alertas_tiempos[:20]  # Top 20 alertas
    }
    
    return JsonResponse(data)


# ==========================================
# APIS PARA REPORTES AVANZADOS - FASE 3
# ==========================================

def api_analisis_componentes(request):
    """
    API: Análisis detallado de componentes afectados
    Retorna datos sobre componentes con más fallos y su impacto
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    # Obtener queryset base y aplicar filtros
    incidencias = Incidencia.objects.all()
    incidencias = aplicar_filtros_reporte(incidencias, request)
    
    # Total de incidencias con componente especificado
    incidencias_con_componente = incidencias.exclude(componente_afectado__isnull=True).exclude(componente_afectado='')
    total_con_componente = incidencias_con_componente.count()
    total_incidencias = incidencias.count()
    
    porcentaje_con_componente = round((total_con_componente / total_incidencias * 100), 2) if total_incidencias > 0 else 0
    
    # 1. Top 10 componentes con más fallos
    top_componentes = incidencias_con_componente.values('componente_afectado').annotate(
        total=Count('id'),
        criticas=Count('id', filter=Q(grado_severidad='critico')),
        atribuibles=Count('id', filter=Q(es_atribuible=True))
    ).order_by('-total')[:10]
    
    top_componentes_data = {
        'labels': [item['componente_afectado'] for item in top_componentes],
        'data': [item['total'] for item in top_componentes],
        'criticas': [item['criticas'] for item in top_componentes],
        'colors': ['#e74c3c' if item['criticas'] > item['total'] * 0.3 else '#3498db' for item in top_componentes]
    }
    
    # 2. Componentes por tipo de equipo (matriz)
    tipos_equipo = incidencias_con_componente.values_list('tipo_equipo', flat=True).distinct()
    componentes_unicos = list(incidencias_con_componente.values_list('componente_afectado', flat=True).distinct())[:15]  # Top 15 componentes
    
    matriz_componentes_equipo = {}
    for tipo in tipos_equipo:
        matriz_componentes_equipo[tipo] = {}
        for componente in componentes_unicos:
            count = incidencias_con_componente.filter(
                tipo_equipo=tipo,
                componente_afectado=componente
            ).count()
            matriz_componentes_equipo[tipo][componente] = count
    
    # Preparar datos para heatmap
    heatmap_componentes_equipo = {
        'tipos_equipo': [dict(Incidencia.TIPO_EQUIPO_CHOICES).get(tipo, tipo) for tipo in tipos_equipo],
        'componentes': componentes_unicos,
        'data': [
            [matriz_componentes_equipo[tipo].get(comp, 0) for comp in componentes_unicos]
            for tipo in tipos_equipo
        ]
    }
    
    # 3. Severidad por componente (Top 10)
    severidad_por_componente = {}
    for item in top_componentes[:10]:
        componente = item['componente_afectado']
        incidencias_comp = incidencias_con_componente.filter(componente_afectado=componente)
        
        severidad_por_componente[componente] = {
            'baja': incidencias_comp.filter(grado_severidad='bajo').count(),
            'media': incidencias_comp.filter(grado_severidad='medio').count(),
            'alta': incidencias_comp.filter(grado_severidad='alto').count(),
            'critica': incidencias_comp.filter(grado_severidad='critico').count()
        }
    
    # Preparar datos para gráfico apilado
    severidad_componentes_chart = {
        'labels': list(severidad_por_componente.keys()),
        'datasets': [
            {
                'label': 'Baja',
                'data': [severidad_por_componente[comp]['baja'] for comp in severidad_por_componente],
                'backgroundColor': '#27ae60',
                'stack': 'Stack 0'
            },
            {
                'label': 'Media',
                'data': [severidad_por_componente[comp]['media'] for comp in severidad_por_componente],
                'backgroundColor': '#f39c12',
                'stack': 'Stack 0'
            },
            {
                'label': 'Alta',
                'data': [severidad_por_componente[comp]['alta'] for comp in severidad_por_componente],
                'backgroundColor': '#e67e22',
                'stack': 'Stack 0'
            },
            {
                'label': 'Crítica',
                'data': [severidad_por_componente[comp]['critica'] for comp in severidad_por_componente],
                'backgroundColor': '#e74c3c',
                'stack': 'Stack 0'
            }
        ]
    }
    
    # 4. Tendencia mensual de componentes principales (Top 5 últimos 6 meses)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    top_5_componentes = [item['componente_afectado'] for item in top_componentes[:5]]
    
    incidencias_recientes = incidencias_con_componente.filter(
        fecha_deteccion__gte=hace_6_meses,
        componente_afectado__in=top_5_componentes
    ).order_by('fecha_deteccion')
    
    # Agrupar por mes y componente
    tendencias_componentes = {comp: defaultdict(int) for comp in top_5_componentes}
    
    for inc in incidencias_recientes:
        mes_key = inc.fecha_deteccion.strftime('%Y-%m')
        tendencias_componentes[inc.componente_afectado][mes_key] += 1
    
    # Obtener todos los meses únicos
    todos_meses = set()
    for comp_data in tendencias_componentes.values():
        todos_meses.update(comp_data.keys())
    meses_ordenados = sorted(todos_meses)
    
    # Preparar datos para gráfico de líneas múltiples
    tendencia_componentes_chart = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes in meses_ordenados] if meses_ordenados else [],
        'datasets': []
    }
    
    colores_lineas = ['#e74c3c', '#3498db', '#2ecc71', '#f39c12', '#9b59b6']
    for idx, componente in enumerate(top_5_componentes):
        tendencia_componentes_chart['datasets'].append({
            'label': componente,
            'data': [tendencias_componentes[componente].get(mes, 0) for mes in meses_ordenados],
            'borderColor': colores_lineas[idx],
            'backgroundColor': colores_lineas[idx] + '33',
            'tension': 0.4,
            'fill': False
        })
    
    # 5. Componentes más críticos (mayor % de incidencias críticas)
    componentes_criticos = []
    for item in top_componentes:
        componente = item['componente_afectado']
        total_comp = item['total']
        criticas_comp = item['criticas']
        porcentaje_criticas = round((criticas_comp / total_comp * 100), 2) if total_comp > 0 else 0
        
        if porcentaje_criticas > 0:
            componentes_criticos.append({
                'componente': componente,
                'total': total_comp,
                'criticas': criticas_comp,
                'porcentaje_criticas': porcentaje_criticas
            })
    
    # Ordenar por porcentaje de críticas
    componentes_criticos.sort(key=lambda x: x['porcentaje_criticas'], reverse=True)
    
    # Retornar todos los datos
    data = {
        'success': True,
        'kpis': {
            'total_con_componente': total_con_componente,
            'porcentaje_con_componente': porcentaje_con_componente,
            'total_componentes_unicos': len(componentes_unicos),
            'componente_mas_frecuente': top_componentes[0]['componente_afectado'] if top_componentes else 'N/A'
        },
        'top_componentes': top_componentes_data,
        'heatmap_componentes_equipo': heatmap_componentes_equipo,
        'severidad_componentes': severidad_componentes_chart,
        'tendencia_componentes': tendencia_componentes_chart,
        'componentes_criticos': componentes_criticos[:10]
    }
    
    return JsonResponse(data)


def api_analisis_notificaciones(request):
    """
    API: Análisis del historial de notificaciones enviadas
    Retorna métricas sobre notificaciones del sistema
    Soporta filtros: fecha_inicio, fecha_fin, sucursal, tecnico, area, severidad, estado
    """
    from .models import HistorialNotificacion
    
    # Obtener queryset base de notificaciones
    notificaciones = HistorialNotificacion.objects.all()
    
    # Aplicar filtros de fecha si están presentes
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        notificaciones = notificaciones.filter(fecha_envio__gte=fecha_inicio)
    if fecha_fin:
        notificaciones = notificaciones.filter(fecha_envio__lte=fecha_fin)
    
    # Filtros relacionados con la incidencia
    sucursal = request.GET.get('sucursal')
    tecnico = request.GET.get('tecnico')
    area = request.GET.get('area')
    severidad = request.GET.get('severidad')
    estado = request.GET.get('estado')
    
    if sucursal:
        notificaciones = notificaciones.filter(incidencia__equipo__sucursal__nombre=sucursal)
    if tecnico:
        notificaciones = notificaciones.filter(incidencia__tecnico_responsable__nombre_completo=tecnico)
    if area:
        notificaciones = notificaciones.filter(incidencia__area_detectora=area)
    if severidad:
        notificaciones = notificaciones.filter(incidencia__grado_severidad=severidad)
    if estado:
        notificaciones = notificaciones.filter(incidencia__estado=estado)
    
    # Total de notificaciones
    total_notificaciones = notificaciones.count()
    
    # KPIs básicos
    exitosas = notificaciones.filter(exitosa=True).count()
    fallidas = notificaciones.filter(exitosa=False).count()
    tasa_exito = round((exitosas / total_notificaciones * 100), 2) if total_notificaciones > 0 else 0
    
    # 1. Distribución por tipo de notificación
    tipos_notificacion = notificaciones.values('tipo_notificacion').annotate(
        total=Count('id')
    ).order_by('-total')
    
    distribucion_tipos = {
        'labels': [dict(HistorialNotificacion.TIPO_NOTIFICACION_CHOICES).get(item['tipo_notificacion'], item['tipo_notificacion']) 
                   for item in tipos_notificacion],
        'data': [item['total'] for item in tipos_notificacion],
        'colors': ['#3498db', '#e74c3c', '#2ecc71', '#f39c12']
    }
    
    # 2. Tendencia mensual de notificaciones (últimos 6 meses)
    hoy = timezone.now()
    hace_6_meses = hoy - timedelta(days=180)
    
    notificaciones_recientes = notificaciones.filter(
        fecha_envio__gte=hace_6_meses
    ).order_by('fecha_envio')
    
    meses_notificaciones = defaultdict(int)
    meses_exitosas = defaultdict(int)
    
    for notif in notificaciones_recientes:
        mes_key = notif.fecha_envio.strftime('%Y-%m')
        meses_notificaciones[mes_key] += 1
        if notif.exitosa:
            meses_exitosas[mes_key] += 1
    
    meses_ordenados = sorted(meses_notificaciones.keys())
    
    tendencia_notificaciones = {
        'labels': [datetime.strptime(mes, '%Y-%m').strftime('%b %Y') for mes in meses_ordenados] if meses_ordenados else [],
        'datasets': [
            {
                'label': 'Total Enviadas',
                'data': [meses_notificaciones[mes] for mes in meses_ordenados],
                'borderColor': '#3498db',
                'backgroundColor': 'rgba(52, 152, 219, 0.1)',
                'fill': True,
                'tension': 0.4
            },
            {
                'label': 'Exitosas',
                'data': [meses_exitosas[mes] for mes in meses_ordenados],
                'borderColor': '#2ecc71',
                'backgroundColor': 'rgba(46, 204, 113, 0.1)',
                'fill': True,
                'tension': 0.4
            }
        ]
    }
    
    # 3. Top destinatarios más frecuentes
    # Obtener todos los destinatarios únicos de las notificaciones
    destinatarios_count = defaultdict(int)
    
    for notif in notificaciones:
        if notif.destinatarios:
            # Los destinatarios están separados por comas
            destinatarios_list = [d.strip() for d in notif.destinatarios.split(',')]
            for dest in destinatarios_list:
                if dest:
                    destinatarios_count[dest] += 1
    
    # Ordenar y tomar top 10
    top_destinatarios = sorted(destinatarios_count.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_destinatarios_chart = {
        'labels': [dest for dest, _ in top_destinatarios],
        'data': [count for _, count in top_destinatarios]
    }
    
    # 4. Tiempo promedio entre detección de incidencia y primera notificación
    tiempos_notificacion = []
    
    for notif in notificaciones.select_related('incidencia'):
        if notif.incidencia and notif.incidencia.fecha_deteccion:
            diferencia = notif.fecha_envio - notif.incidencia.fecha_deteccion
            minutos = diferencia.total_seconds() / 60
            tiempos_notificacion.append(minutos)
    
    if tiempos_notificacion:
        tiempo_promedio_minutos = round(sum(tiempos_notificacion) / len(tiempos_notificacion), 1)
        tiempo_minimo_minutos = round(min(tiempos_notificacion), 1)
        tiempo_maximo_minutos = round(max(tiempos_notificacion), 1)
    else:
        tiempo_promedio_minutos = 0
        tiempo_minimo_minutos = 0
        tiempo_maximo_minutos = 0
    
    # 5. Distribución de notificaciones por día de la semana
    dias_semana = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}
    notificaciones_por_dia = defaultdict(int)
    
    for notif in notificaciones:
        dia_num = notif.fecha_envio.weekday()
        notificaciones_por_dia[dia_num] += 1
    
    distribucion_dias_semana = {
        'labels': [dias_semana[i] for i in range(7)],
        'data': [notificaciones_por_dia.get(i, 0) for i in range(7)],
        'colors': ['#3498db'] * 7
    }
    
    # 6. Notificaciones por severidad de incidencia
    notificaciones_por_severidad = notificaciones.values(
        'incidencia__grado_severidad'
    ).annotate(
        total=Count('id')
    ).order_by('-total')
    
    distribucion_severidad = {
        'labels': [dict(Incidencia.GRADO_SEVERIDAD_CHOICES).get(item['incidencia__grado_severidad'], 'N/A') 
                   for item in notificaciones_por_severidad],
        'data': [item['total'] for item in notificaciones_por_severidad],
        'colors': ['#27ae60', '#f39c12', '#e67e22', '#e74c3c']
    }
    
    # Retornar todos los datos
    data = {
        'success': True,
        'kpis': {
            'total_notificaciones': total_notificaciones,
            'exitosas': exitosas,
            'fallidas': fallidas,
            'tasa_exito': tasa_exito,
            'tiempo_promedio_minutos': tiempo_promedio_minutos,
            'tiempo_minimo_minutos': tiempo_minimo_minutos,
            'tiempo_maximo_minutos': tiempo_maximo_minutos
        },
        'distribucion_tipos': distribucion_tipos,
        'tendencia_notificaciones': tendencia_notificaciones,
        'top_destinatarios': top_destinatarios_chart,
        'distribucion_dias_semana': distribucion_dias_semana,
        'distribucion_severidad': distribucion_severidad
    }
    
    return JsonResponse(data)
