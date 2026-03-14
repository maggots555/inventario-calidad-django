from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db.models import Q, Sum, Count, F
from django.db import models
from django.utils import timezone
from datetime import datetime, timedelta, date
from functools import wraps
from .models import Producto, Movimiento, Sucursal, Empleado
from .forms import ProductoForm, MovimientoForm, SucursalForm, MovimientoRapidoForm, EmpleadoForm, MovimientoFraccionarioForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ===== DECORADORES DE PERMISOS =====
def permission_required_with_message(perm, message=None):
    """
    Decorador personalizado que verifica permisos de Django y redirige a página de acceso denegado
    
    Args:
        perm (str): Permiso requerido en formato 'app.codename' (ej: 'inventario.add_producto')
        message (str): Mensaje personalizado de error (opcional)
    
    Uso:
        @login_required
        @permission_required_with_message('inventario.add_producto')
        def crear_producto(request):
            # Solo usuarios con permiso add_producto pueden ejecutar esto
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.has_perm(perm):
                error_msg = message or f'No tienes permisos para realizar esta acción.'
                # Redirigir a página de acceso denegado con el mensaje
                return redirect(f"{reverse('acceso_denegado')}?mensaje={error_msg}&permiso={perm}")
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def staff_required(view_func):
    """
    Decorador que requiere que el usuario sea staff o superusuario
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Este decorador se usa en vistas que solo pueden acceder administradores.
    Verifica que el usuario tenga is_staff=True o is_superuser=True.
    Si no tiene permisos, muestra un mensaje de error y redirige.
    
    Uso:
        @login_required
        @staff_required
        def mi_vista(request):
            # Solo usuarios staff pueden ejecutar esto
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_staff and not request.user.is_superuser:
            messages.error(
                request, 
                '⛔ No tienes permisos para realizar esta acción. '
                'Solo los administradores pueden gestionar empleados.'
            )
            return redirect('lista_empleados')
        return view_func(request, *args, **kwargs)
    return wrapper

# Importar modelos de scorecard para el dashboard principal
try:
    from scorecard.models import Incidencia
    SCORECARD_AVAILABLE = True
except ImportError:
    SCORECARD_AVAILABLE = False

try:
    from almacen.models import ProductoAlmacen, CompraProducto, SolicitudBaja
    ALMACEN_AVAILABLE = True
except ImportError:
    ALMACEN_AVAILABLE = False

# Función auxiliar para convertir fechas a zona horaria local
def fecha_local(fecha_utc):
    """
    Convierte una fecha UTC a la hora local del país activo.

    EXPLICACIÓN PARA PRINCIPIANTES:
    Antes esta función siempre convertía a hora de México (America/Mexico_City).
    Ahora usa get_pais_actual() para detectar el país del request actual
    y convierte a la zona horaria correcta (ej: Buenos Aires para Argentina).
    Si no hay request activo (manage.py shell, cron), usa México como fallback.
    """
    from config.paises_config import get_pais_actual, fecha_local_pais
    pais = get_pais_actual()
    return fecha_local_pais(fecha_utc, pais)

# ===== DASHBOARD PRINCIPAL UNIFICADO =====
@login_required
def dashboard_principal(request):
    """
    Dashboard principal del sistema - Punto de entrada unificado
    Muestra resumen de los tres módulos: Inventario, Control de Calidad y Servicio Técnico
    """
    # ========== ESTADÍSTICAS DE INVENTARIO ==========
    total_productos = Producto.objects.count()
    productos_stock_bajo = Producto.objects.filter(cantidad__lte=F('stock_minimo')).count()
    movimientos_hoy = Movimiento.objects.filter(fecha_movimiento__date=date.today()).count()
    valor_inventario = Producto.objects.aggregate(
        total=Sum(F('cantidad') * F('costo_unitario'))
    )['total'] or 0
    
    # ========== ESTADÍSTICAS DE CONTROL DE CALIDAD ==========
    incidencias_totales = 0
    incidencias_abiertas = 0
    incidencias_criticas = 0
    incidencias_recientes = []
    
    if SCORECARD_AVAILABLE:
        incidencias_totales = Incidencia.objects.count()
        incidencias_abiertas = Incidencia.objects.filter(estado='abierta').count()
        incidencias_criticas = Incidencia.objects.filter(
            grado_severidad='critico',
            estado__in=['abierta', 'en_revision']
        ).count()
        incidencias_recientes = Incidencia.objects.select_related(
            'tecnico_responsable', 'tipo_incidencia'
        ).order_by('-fecha_registro')[:5]
    
    # ========== ESTADÍSTICAS DE SERVICIO TÉCNICO ==========
    ordenes_activas = 0
    ordenes_pendientes = 0
    ordenes_hoy = 0
    ordenes_completadas = 0
    
    # Verificar si el módulo de servicio técnico está disponible
    try:
        from servicio_tecnico.models import OrdenServicio
        SERVICIO_TECNICO_AVAILABLE = True
    except ImportError:
        SERVICIO_TECNICO_AVAILABLE = False
    
    if SERVICIO_TECNICO_AVAILABLE:
        # Estados considerados como "activas" (en proceso)
        estados_activos = [
            'espera', 'recepcion', 'diagnostico', 'equipo_diagnosticado',
            'diagnostico_enviado_cliente', 'cotizacion', 'cliente_acepta_cotizacion',
            'partes_solicitadas_proveedor', 'esperando_piezas', 'piezas_recibidas',
            'reparacion', 'control_calidad'
        ]
        
        # Estados considerados como "pendientes" (esperando acción)
        estados_pendientes = [
            'espera', 'diagnostico_enviado_cliente', 'cotizacion',
            'cotizacion_enviada_proveedor', 'esperando_piezas'
        ]
        
        # Órdenes activas (todas las que no están finalizadas, entregadas o canceladas)
        ordenes_activas = OrdenServicio.objects.filter(estado__in=estados_activos).count()
        
        # Órdenes pendientes (esperando alguna acción o respuesta)
        ordenes_pendientes = OrdenServicio.objects.filter(estado__in=estados_pendientes).count()
        
        # Órdenes creadas hoy
        ordenes_hoy = OrdenServicio.objects.filter(fecha_ingreso__date=date.today()).count()
        
        # Órdenes completadas (finalizadas + entregadas)
        ordenes_completadas = OrdenServicio.objects.filter(
            estado__in=['finalizado', 'entregado']
        ).count()

    # ========== ESTADÍSTICAS DE ALMACÉN CENTRAL ==========
    total_almacen = 0
    compras_pendientes = 0
    bajas_pendientes = 0
    valor_almacen = 0

    if ALMACEN_AVAILABLE:
        # Total de productos en catálogo de almacén
        total_almacen = ProductoAlmacen.objects.filter(activo=True).count()
        
        # Compras o cotizaciones que requieren atención
        compras_pendientes = CompraProducto.objects.filter(
            estado__in=['cotizacion', 'pendiente_llegada']
        ).count()
        
        # Solicitudes de baja esperando aprobación
        bajas_pendientes = SolicitudBaja.objects.filter(
            estado='pendiente_aprobacion'
        ).count()
        
        # Valor total estimado del stock en almacén central
        valor_almacen = ProductoAlmacen.objects.filter(activo=True).aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0
    
    # ========== ACTIVIDADES RECIENTES ==========
    # Últimos 5 movimientos
    movimientos_recientes = Movimiento.objects.select_related(
        'producto', 'sucursal_destino', 'usuario_registro_empleado', 'empleado_destinatario'
    ).order_by('-fecha_movimiento')[:5]
    
    # ========== ESTADÍSTICAS RÁPIDAS ==========
    total_sucursales = Sucursal.objects.count()
    total_empleados = Empleado.objects.count()
    
    # Calcular porcentaje de éxito (incidencias cerradas vs totales)
    if incidencias_totales > 0 and SCORECARD_AVAILABLE:
        incidencias_cerradas = Incidencia.objects.filter(estado='cerrada').count()
        tasa_exito = round((incidencias_cerradas / incidencias_totales) * 100, 1)
    else:
        tasa_exito = 100
    
    context = {
        # Inventario
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'movimientos_hoy': movimientos_hoy,
        'valor_inventario': valor_inventario,
        
        # Control de Calidad
        'incidencias_totales': incidencias_totales,
        'incidencias_abiertas': incidencias_abiertas,
        'incidencias_criticas': incidencias_criticas,
        'incidencias_recientes': incidencias_recientes,
        
        # Servicio Técnico
        'ordenes_activas': ordenes_activas,
        'ordenes_pendientes': ordenes_pendientes,
        'ordenes_hoy': ordenes_hoy,
        'ordenes_completadas': ordenes_completadas,
        'servicio_tecnico_disponible': SERVICIO_TECNICO_AVAILABLE,

        # Almacén Central
        'total_almacen': total_almacen,
        'compras_pendientes': compras_pendientes,
        'bajas_pendientes': bajas_pendientes,
        'valor_almacen': valor_almacen,
        'almacen_disponible': ALMACEN_AVAILABLE,
        
        # General
        'total_sucursales': total_sucursales,
        'total_empleados': total_empleados,
        'tasa_exito': tasa_exito,
        'movimientos_recientes': movimientos_recientes,
        'scorecard_disponible': SCORECARD_AVAILABLE,
    }
    
    return render(request, 'dashboard_principal.html', context)


# ===== DASHBOARD DE INVENTARIO =====
@login_required
@permission_required_with_message('inventario.view_producto', message='No tienes permisos para acceder al dashboard de inventario.')
def dashboard_inventario(request):
    """
    Vista del dashboard de inventario con métricas e información resumida
    """
    # Estadísticas generales
    total_productos = Producto.objects.count()
    productos_stock_bajo = Producto.objects.filter(cantidad__lte=F('stock_minimo')).count()
    valor_total_inventario = Producto.objects.aggregate(
        total=Sum(F('cantidad') * F('costo_unitario'))
    )['total'] or 0
    
    # Productos con stock bajo (alertas tradicionales)
    alertas_stock = Producto.objects.filter(cantidad__lte=F('stock_minimo'))[:10]
    
    # Alertas específicas para productos fraccionables (lógica mejorada)
    alertas_fraccionables = []
    productos_fraccionables = Producto.objects.filter(es_fraccionable=True)
    for producto in productos_fraccionables:
        # Verificar múltiples condiciones de alerta
        porcentaje = producto.porcentaje_disponible()
        
        # Condiciones de alerta:
        # 1. Stock fraccionario bajo (método original)
        # 2. Porcentaje disponible <= 70% (medio/bajo)
        # 3. Cantidad mínima de alerta específica
        if (producto.stock_fraccionario_bajo() or 
            porcentaje <= 70 or
            (producto.cantidad_minima_alerta > 0 and producto.cantidad_actual <= producto.cantidad_minima_alerta)):
            alertas_fraccionables.append(producto)
    
    alertas_fraccionables = alertas_fraccionables[:10]
    
    # Movimientos recientes
    movimientos_recientes = Movimiento.objects.select_related('producto', 'sucursal_destino')[:10]
    
    # Productos más utilizados (últimos 30 días)
    fecha_limite = timezone.now() - timedelta(days=30)
    productos_mas_usados = (
        Movimiento.objects
        .filter(fecha_movimiento__gte=fecha_limite, tipo='salida')
        .values('producto__nombre')
        .annotate(total_usado=Sum('cantidad'))
        .order_by('-total_usado')[:5]
    )
    
    # Movimientos por categoría (últimos 30 días)
    movimientos_por_categoria = (
        Movimiento.objects
        .filter(fecha_movimiento__gte=fecha_limite)
        .values('producto__categoria')
        .annotate(total_movimientos=Count('id'))
        .order_by('-total_movimientos')
    )
    
    context = {
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'valor_total_inventario': valor_total_inventario,
        'alertas_stock': alertas_stock,
        'alertas_fraccionables': alertas_fraccionables,  # Nuevas alertas fraccionables
        'movimientos_recientes': movimientos_recientes,
        'productos_mas_usados': productos_mas_usados,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'inventario/dashboard.html', context)

# ===== GESTIÓN DE PRODUCTOS =====
@login_required
@permission_required_with_message('inventario.view_producto', message='No tienes permisos para ver la lista de productos del inventario.')
def lista_productos(request):
    """
    Lista de productos con filtros y búsqueda
    """
    productos = Producto.objects.all()
    
    # Filtros de búsqueda
    busqueda = request.GET.get('busqueda', '')
    categoria = request.GET.get('categoria', '')
    stock_bajo = request.GET.get('stock_bajo', '')
    
    # Aplicar filtros
    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo_qr__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if categoria:
        productos = productos.filter(categoria=categoria)
    
    # Corregir filtro de stock bajo - verificar que sea exactamente 'true'
    if stock_bajo == 'true':
        productos = productos.filter(cantidad__lte=F('stock_minimo'))
    
    # Obtener opciones para filtros
    categorias = Producto.CATEGORIA_CHOICES
    
    # Calcular estadísticas de resumen correctamente en la vista
    total_productos = productos.count()
    productos_stock_bajo = productos.filter(cantidad__lte=F('stock_minimo')).count()
    
    # Calcular unidades totales
    unidades_totales = productos.aggregate(
        total=Sum('cantidad')
    )['total'] or 0
    
    # Calcular valor total del inventario
    valor_total = productos.aggregate(
        total=Sum(F('cantidad') * F('costo_unitario'))
    )['total'] or 0
    
    context = {
        'productos': productos,
        'categorias': categorias,
        'busqueda': busqueda,
        'categoria_seleccionada': categoria,
        'stock_bajo_seleccionado': stock_bajo == 'true',  # Pasar como booleano
        # Estadísticas calculadas en la vista
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'unidades_totales': unidades_totales,
        'valor_total': valor_total,
    }
    
    return render(request, 'inventario/lista_productos.html', context)

@login_required
@permission_required_with_message('inventario.view_producto', message='No tienes permisos para ver los detalles de productos.')
def detalle_producto(request, producto_id):
    """
    Vista detallada de un producto con su historial de movimientos
    """
    producto = get_object_or_404(Producto, id=producto_id)
    movimientos = Movimiento.objects.filter(producto=producto).select_related(
        'empleado_destinatario', 'usuario_registro_empleado', 'sucursal_destino'
    ).order_by('-fecha_movimiento')[:20]
    
    context = {
        'producto': producto,
        'movimientos': movimientos,
        'qr_image': producto.generar_qr_image(),
    }
    
    return render(request, 'inventario/detalle_producto.html', context)

@login_required
@permission_required_with_message('inventario.add_producto', message='No tienes permisos para crear nuevos productos en el inventario.')
def crear_producto(request):
    """
    Crear un nuevo producto
    """
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente con código QR: {producto.codigo_qr}')
            return redirect('detalle_producto', producto_id=producto.id)
    else:
        form = ProductoForm()
    
    return render(request, 'inventario/form_producto.html', {
        'form': form,
        'titulo': 'Crear Producto',
        'boton_texto': 'Crear Producto'
    })

@login_required
@permission_required_with_message('inventario.change_producto', message='No tienes permisos para modificar productos existentes.')
def editar_producto(request, producto_id):
    """
    Editar un producto existente
    """
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
            return redirect('detalle_producto', producto_id=producto.id)
    else:
        form = ProductoForm(instance=producto)
    
    return render(request, 'inventario/form_producto.html', {
        'form': form,
        'producto': producto,
        'titulo': f'Editar: {producto.nombre}',
        'boton_texto': 'Guardar Cambios'
    })

@login_required
@permission_required_with_message('inventario.delete_producto', message='No tienes permisos para eliminar productos del inventario.')
def eliminar_producto(request, producto_id):
    """
    Eliminar un producto (con confirmación)
    """
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        nombre = producto.nombre
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado exitosamente.')
        return redirect('lista_productos')
    
    return render(request, 'inventario/confirmar_eliminacion.html', {
        'producto': producto,
        'movimientos_count': Movimiento.objects.filter(producto=producto).count()
    })

# ===== GESTIÓN DE MOVIMIENTOS =====
@login_required
@permission_required_with_message('inventario.view_movimiento', message='No tienes permisos para ver el historial de movimientos.')
def lista_movimientos(request):
    """
    Lista de todos los movimientos con filtros
    """
    movimientos = Movimiento.objects.select_related('producto', 'sucursal_destino').order_by('-fecha_movimiento')
    
    # Filtros
    tipo = request.GET.get('tipo')
    producto_id = request.GET.get('producto')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    fraccionario = request.GET.get('fraccionario')
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_hasta)
    
    if fraccionario == 'true':
        movimientos = movimientos.filter(es_movimiento_fraccionario=True)
    elif fraccionario == 'false':
        movimientos = movimientos.filter(es_movimiento_fraccionario=False)
    
    # Paginación (opcional: mostrar solo los primeros 50)
    movimientos_limitados = movimientos[:50]
    
    # Calcular estadísticas de resumen correctamente en la vista
    total_movimientos = movimientos.count()
    entradas_count = movimientos.filter(tipo='entrada').count()
    salidas_count = movimientos.filter(tipo='salida').count()
    ajustes_count = movimientos.filter(tipo='ajuste').count()
    devoluciones_count = movimientos.filter(tipo='devolucion').count()
    
    context = {
        'movimientos': movimientos_limitados,
        'tipos_movimiento': Movimiento.TIPO_CHOICES,
        'productos': Producto.objects.all()[:20],  # Primeros 20 para el selector
        # Estadísticas calculadas en la vista
        'total_movimientos': total_movimientos,
        'entradas_count': entradas_count,
        'salidas_count': salidas_count,
        'ajustes_count': ajustes_count,
        'devoluciones_count': devoluciones_count,
    }
    
    return render(request, 'inventario/lista_movimientos.html', context)

@login_required
@permission_required_with_message('inventario.add_movimiento', message='No tienes permisos para registrar movimientos de inventario.')
def crear_movimiento(request):
    """
    Crear un nuevo movimiento de inventario
    """
    if request.method == 'POST':
        form = MovimientoForm(request.POST)
        if form.is_valid():
            movimiento = form.save()
            messages.success(request, f'Movimiento registrado exitosamente. Nuevo stock: {movimiento.stock_posterior}')
            return redirect('lista_movimientos')
    else:
        form = MovimientoForm()
    
    return render(request, 'inventario/form_movimiento.html', {
        'form': form,
        'titulo': 'Registrar Movimiento',
        'boton_texto': 'Registrar Movimiento'
    })

@login_required
@permission_required_with_message('inventario.add_movimiento', message='No tienes permisos para registrar movimientos rápidos (Scanner QR).')
def movimiento_rapido(request):
    """
    Formulario rápido para movimientos usando código QR
    """
    if request.method == 'POST':
        form = MovimientoRapidoForm(request.POST)
        if form.is_valid():
            movimiento = form.save()
            producto = movimiento.producto
            messages.success(request, f'Movimiento registrado: {producto.nombre}. Nuevo stock: {movimiento.stock_posterior}')
            return redirect('movimiento_rapido')
    else:
        form = MovimientoRapidoForm()
    
    return render(request, 'inventario/movimiento_rapido.html', {
        'form': form,
        'titulo': 'Movimiento Rápido (Scanner QR)'
    })

@login_required
@permission_required_with_message('inventario.add_movimiento', message='No tienes permisos para registrar movimientos fraccionarios.')
def movimiento_fraccionario(request):
    """
    Vista para movimientos fraccionarios (consumo parcial de productos)
    """
    if request.method == 'POST':
        form = MovimientoFraccionarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save()
            producto = movimiento.producto
            
            # Mensaje personalizado para movimientos fraccionarios
            if movimiento.es_movimiento_fraccionario:
                cantidad_restante = producto.cantidad_actual
                porcentaje = producto.porcentaje_disponible()
                messages.success(
                    request, 
                    f'Movimiento fraccionario registrado: {producto.nombre}. '
                    f'Consumido: {movimiento.cantidad_fraccionaria} {movimiento.unidad_utilizada}. '
                    f'Restante: {cantidad_restante:.2f} {producto.unidad_base} ({porcentaje:.1f}%)'
                )
            else:
                messages.success(request, f'Movimiento registrado: {producto.nombre}')
                
            return redirect('movimiento_fraccionario')
    else:
        form = MovimientoFraccionarioForm()
    
    return render(request, 'inventario/movimiento_fraccionario.html', {
        'form': form,
        'titulo': 'Movimiento Fraccionario (Consumo Parcial)'
    })

# ===== GESTIÓN DE SUCURSALES =====
@login_required
@permission_required_with_message('inventario.view_sucursal', message='No tienes permisos para ver la lista de sucursales.')
def lista_sucursales(request):
    """
    Lista de todas las sucursales
    """
    sucursales = Sucursal.objects.all()
    return render(request, 'inventario/lista_sucursales.html', {'sucursales': sucursales})

@login_required
@permission_required_with_message('inventario.add_sucursal', message='No tienes permisos para crear nuevas sucursales.')
def crear_sucursal(request):
    """
    Crear una nueva sucursal
    """
    if request.method == 'POST':
        form = SucursalForm(request.POST)
        if form.is_valid():
            sucursal = form.save()
            messages.success(request, f'Sucursal "{sucursal.nombre}" creada exitosamente.')
            return redirect('lista_sucursales')
    else:
        form = SucursalForm()
    
    return render(request, 'inventario/form_sucursal.html', {
        'form': form,
        'titulo': 'Crear Sucursal',
        'boton_texto': 'Crear Sucursal'
    })

@login_required
@permission_required_with_message('inventario.change_sucursal', message='No tienes permisos para modificar sucursales existentes.')
def editar_sucursal(request, sucursal_id):
    """
    Editar una sucursal existente
    """
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    if request.method == 'POST':
        form = SucursalForm(request.POST, instance=sucursal)
        if form.is_valid():
            form.save()
            messages.success(request, f'Sucursal "{sucursal.nombre}" actualizada exitosamente.')
            return redirect('lista_sucursales')
    else:
        form = SucursalForm(instance=sucursal)
    
    return render(request, 'inventario/form_sucursal.html', {
        'form': form,
        'sucursal': sucursal,
        'titulo': f'Editar: {sucursal.nombre}',
        'boton_texto': 'Guardar Cambios'
    })

@login_required
@permission_required_with_message('inventario.delete_sucursal', message='No tienes permisos para eliminar sucursales.')
def eliminar_sucursal(request, sucursal_id):
    """
    Eliminar una sucursal (con validaciones de seguridad)
    """
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    
    if request.method == 'POST':
        # Verificar si la sucursal tiene movimientos asociados
        movimientos_asociados = Movimiento.objects.filter(sucursal_destino=sucursal).count()
        
        if movimientos_asociados > 0:
            messages.error(
                request, 
                f'No se puede eliminar la sucursal "{sucursal.nombre}" porque tiene {movimientos_asociados} movimiento(s) asociado(s). '
                'Primero debe eliminar o reasignar estos movimientos.'
            )
        else:
            nombre_sucursal = sucursal.nombre
            try:
                sucursal.delete()
                messages.success(request, f'Sucursal "{nombre_sucursal}" eliminada exitosamente.')
            except Exception as e:
                messages.error(request, f'Error al eliminar la sucursal: {str(e)}')
        
        return redirect('lista_sucursales')
    
    # Si es GET, mostrar confirmación con información básica de la sucursal
    context = {
        'sucursal': sucursal,
        'movimientos_asociados': Movimiento.objects.filter(sucursal_destino=sucursal).count()
    }
    return render(request, 'inventario/confirmar_eliminar_sucursal.html', context)

# ===== FUNCIONES AJAX/API =====
def buscar_producto_qr(request):
    """
    API endpoint para buscar producto por código QR (AJAX)
    Con limpieza de caracteres invisibles del scanner
    """
    codigo_qr_raw = request.GET.get('codigo_qr')
    
    if not codigo_qr_raw:
        return JsonResponse({'error': 'Código QR requerido'}, status=400)
    
    # Limpiar código de caracteres invisibles y problemáticos
    import re
    codigo_qr = re.sub(r'[\r\n\t\x00-\x1f\x7f-\x9f]', '', codigo_qr_raw)  # Remover caracteres de control
    codigo_qr = codigo_qr.strip()  # Remover espacios al inicio y final
    codigo_qr = re.sub(r'\s+', '', codigo_qr)  # Remover espacios internos
    
    if not codigo_qr:
        return JsonResponse({
            'error': 'Código QR vacío después de limpieza',
            'codigo_recibido': repr(codigo_qr_raw),
            'debug': f'Código original tenía {len(codigo_qr_raw)} caracteres'
        }, status=400)
    
    try:
        # Búsqueda insensible a mayúsculas/minúsculas
        producto = Producto.objects.get(codigo_qr__iexact=codigo_qr)
    except Producto.DoesNotExist:
        return JsonResponse({
            'error': 'Producto no encontrado',
            'codigo_buscado': codigo_qr,
            'codigo_original': repr(codigo_qr_raw),
            'debug': f'Buscando: "{codigo_qr}" (longitud: {len(codigo_qr)})'
        }, status=404)
    
    # VALIDACIÓN: Verificar si es un producto fraccionario
    if producto.es_fraccionable:
        return JsonResponse({
            'error': 'Producto fraccionable detectado',
            'es_fraccionable': True,
            'nombre': producto.nombre,
            'codigo_qr': producto.codigo_qr,
            'unidad_base': producto.unidad_base,
            'cantidad_unitaria': producto.cantidad_unitaria,
            'mensaje': f'Este producto ({producto.nombre}) requiere manejo fraccionario en {producto.unidad_base}. Use el formulario de movimientos fraccionarios.',
            'url_fraccionario': reverse('movimiento_fraccionario')
        }, status=400)
    
    data = {
        'id': producto.id,
        'nombre': producto.nombre,
        'codigo_qr': producto.codigo_qr,
        'cantidad_actual': producto.cantidad,
        'categoria': producto.get_categoria_display(),
        'ubicacion': producto.ubicacion,
    }
    return JsonResponse(data)


@login_required
def buscar_producto_fraccionable_qr(request):
    """
    API endpoint para buscar producto fraccionable por código QR
    Proporciona información detallada sobre el estado fraccionario
    """
    codigo_qr_raw = request.GET.get('codigo_qr')
    
    if not codigo_qr_raw:
        return JsonResponse({'error': 'Código QR requerido'}, status=400)
    
    # Limpiar código de caracteres invisibles y problemáticos
    import re
    codigo_qr = re.sub(r'[\r\n\t\x00-\x1f\x7f-\x9f]', '', codigo_qr_raw)
    codigo_qr = codigo_qr.strip()
    codigo_qr = re.sub(r'\s+', '', codigo_qr)
    
    if not codigo_qr:
        return JsonResponse({
            'error': 'Código QR vacío después de limpieza',
            'codigo_recibido': repr(codigo_qr_raw)
        }, status=400)
    
    try:
        producto = Producto.objects.get(codigo_qr__iexact=codigo_qr)
    except Producto.DoesNotExist:
        return JsonResponse({
            'error': 'Producto no encontrado',
            'codigo_buscado': codigo_qr
        }, status=404)
    
    # Verificar si el producto es fraccionable
    if not producto.es_fraccionable:
        return JsonResponse({
            'error': 'Este producto no es fraccionable',
            'producto_nombre': producto.nombre,
            'es_fraccionable': False
        }, status=400)
    
    # Preparar datos específicos para productos fraccionables
    data = {
        'id': producto.id,
        'nombre': producto.nombre,
        'codigo_qr': producto.codigo_qr,
        'es_fraccionable': producto.es_fraccionable,
        'unidad_base': producto.unidad_base,
        'cantidad_unitaria': producto.cantidad_unitaria,
        'cantidad_actual': producto.cantidad_actual,
        'cantidad_minima_alerta': producto.cantidad_minima_alerta,
        'cantidad_total_disponible': producto.cantidad_total_disponible(),
        'porcentaje_disponible': producto.porcentaje_disponible(),
        'stock_fraccionario_bajo': producto.stock_fraccionario_bajo(),
        'unidades_completas': producto.cantidad,
        'categoria': producto.get_categoria_display(),
        'ubicacion': producto.ubicacion,
    }
    
    return JsonResponse(data)


# def limpiar_codigo_scanner(codigo):
#     """
#     Limpia códigos QR que pueden venir corruptos de scanners físicos
#     """
#     import re
#     
#     # Remover caracteres problemáticos comunes
#     codigo_limpio = codigo.replace('"', '').replace('=', '').replace('%', '')
#     codigo_limpio = codigo_limpio.replace(')', '').replace('#', '').replace('!', '')
#     codigo_limpio = codigo_limpio.replace('$', '').replace('&', '')
#     
#     # Solo mantener caracteres alfanuméricos válidos
#     codigo_limpio = re.sub(r'[^A-Z0-9]', '', codigo_limpio)
#     
#     return codigo_limpio


# def buscar_producto_fuzzy(codigo_corrupto):
#     """
#     Búsqueda aproximada para códigos muy corruptos
#     """
#     codigo_limpio = limpiar_codigo_scanner(codigo_corrupto)
#     
#     # Buscar productos que contengan la parte "INV" + números
#     if 'INV' in codigo_limpio:
#         parte_numerica = codigo_limpio.replace('INV', '')
#         if len(parte_numerica) >= 8:  # Al menos 8 dígitos (fecha parcial)
#             productos = Producto.objects.filter(
#                 codigo_qr__contains=parte_numerica[:8]
#             )
#             if productos.exists():
#                 return productos.first()
#     
#     raise Producto.DoesNotExist("No se pudo encontrar el producto")

@login_required
def generar_qr_producto(request, producto_id):
    """
    Generar y mostrar código QR de un producto
    """
    producto = get_object_or_404(Producto, id=producto_id)
    
    # Generar QR como imagen
    import qrcode
    from io import BytesIO
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(producto.codigo_qr)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Crear respuesta HTTP con la imagen
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    
    return response


# ===== GESTIÓN DE EMPLEADOS =====
@login_required
@permission_required_with_message('inventario.view_empleado', message='No tienes permisos para ver la lista de empleados.')
def lista_empleados(request):
    """
    Lista de empleados con filtros
    """
    empleados = Empleado.objects.all()
    
    # Filtros de búsqueda
    busqueda = request.GET.get('busqueda', '')
    area = request.GET.get('area', '')
    activo = request.GET.get('activo', '')
    acceso_sistema = request.GET.get('acceso_sistema', '')
    
    if busqueda:
        empleados = empleados.filter(
            Q(nombre_completo__icontains=busqueda) |
            Q(cargo__icontains=busqueda) |
            Q(area__icontains=busqueda)
        )
    
    if area:
        empleados = empleados.filter(area__icontains=area)
    
    if activo == 'true':
        empleados = empleados.filter(activo=True)
    elif activo == 'false':
        empleados = empleados.filter(activo=False)
    
    # Filtro de acceso al sistema
    if acceso_sistema == 'activo':
        # Empleados con usuario activo y contraseña configurada
        empleados = empleados.filter(
            user__isnull=False, 
            user__is_active=True,
            contraseña_configurada=True
        )
    elif acceso_sistema == 'pendiente':
        # Empleados con usuario activo pero sin configurar contraseña
        empleados = empleados.filter(
            user__isnull=False,
            user__is_active=True,
            contraseña_configurada=False
        )
    elif acceso_sistema == 'revocado':
        # Empleados con usuario pero desactivado (acceso revocado)
        empleados = empleados.filter(
            user__isnull=False,
            user__is_active=False
        )
    elif acceso_sistema == 'sin_acceso':
        # Empleados sin usuario de sistema
        empleados = empleados.filter(user__isnull=True)
    
    # Ordenar por área y luego por nombre
    empleados = empleados.order_by('area', 'nombre_completo')
    
    # Obtener áreas únicas para filtro
    areas_disponibles = Empleado.objects.values_list('area', flat=True).distinct().order_by('area')
    
    # Verificar si el usuario tiene permisos de administrador
    es_admin = request.user.is_staff or request.user.is_superuser
    
    context = {
        'empleados': empleados,
        'busqueda': busqueda,
        'area_seleccionada': area,
        'activo_seleccionado': activo,
        'acceso_seleccionado': acceso_sistema,
        'areas_disponibles': areas_disponibles,
        'es_admin': es_admin,  # Nuevo: Indica si el usuario puede modificar
    }
    
    return render(request, 'inventario/lista_empleados.html', context)

@login_required
@staff_required
def crear_empleado(request):
    """
    Crear un nuevo empleado
    Solo accesible para usuarios staff/superusuario
    """
    if request.method == 'POST':
        # request.FILES permite manejar archivos subidos (como imágenes)
        form = EmpleadoForm(request.POST, request.FILES)
        if form.is_valid():
            empleado = form.save()
            messages.success(request, f'Empleado {empleado.nombre_completo} creado exitosamente.')
            return redirect('lista_empleados')
    else:
        form = EmpleadoForm()
    
    return render(request, 'inventario/form_empleado.html', {
        'form': form,
        'titulo': 'Agregar Empleado',
        'boton_texto': 'Crear Empleado'
    })

@login_required
@staff_required
def editar_empleado(request, empleado_id):
    """
    Editar un empleado existente
    Solo accesible para usuarios staff/superusuario
    """
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == 'POST':
        # request.FILES permite manejar archivos subidos (como imágenes)
        form = EmpleadoForm(request.POST, request.FILES, instance=empleado)
        if form.is_valid():
            empleado = form.save()
            messages.success(request, f'Empleado {empleado.nombre_completo} actualizado exitosamente.')
            return redirect('lista_empleados')
    else:
        form = EmpleadoForm(instance=empleado)
    
    return render(request, 'inventario/form_empleado.html', {
        'form': form,
        'titulo': f'Editar Empleado: {empleado.nombre_completo}',
        'boton_texto': 'Actualizar Empleado',
        'empleado': empleado
    })

@login_required
@staff_required
def eliminar_empleado(request, empleado_id):
    """
    Marcar empleado como inactivo (soft delete)
    Solo accesible para usuarios staff/superusuario
    """
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == 'POST':
        empleado.activo = False
        empleado.save()
        messages.success(request, f'Empleado {empleado.nombre_completo} marcado como inactivo.')
        return redirect('lista_empleados')
    
    return render(request, 'inventario/confirmar_eliminacion.html', {
        'objeto': empleado,
        'tipo': 'empleado',
        'url_cancelar': 'lista_empleados'
    })


# ===== GESTIÓN DE ACCESO AL SISTEMA =====

@login_required
@staff_required
def dar_acceso_empleado(request, empleado_id):
    """
    Otorga acceso al sistema a un empleado
    Crea su usuario de Django y envía credenciales por email
    Solo accesible para usuarios staff/superusuario
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista hace varias cosas importantes:
    1. Verifica que el empleado tenga email
    2. Crea un usuario de Django para el empleado
    3. Genera una contraseña temporal aleatoria
    4. Envía un email con las credenciales
    5. Redirige de vuelta a la lista de empleados
    """
    from .utils import crear_usuario_para_empleado, enviar_credenciales_empleado
    
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # Validaciones
    if empleado.user:
        messages.warning(request, f'{empleado.nombre_completo} ya tiene acceso al sistema.')
        return redirect('lista_empleados')
    
    if not empleado.email:
        messages.error(request, f'El empleado {empleado.nombre_completo} no tiene un email registrado. Por favor, actualiza su información primero.')
        return redirect('editar_empleado', empleado_id=empleado.id)
    
    try:
        # Crear usuario y generar contraseña
        user, contraseña_temporal = crear_usuario_para_empleado(empleado)
        
        # Enviar email con credenciales (retorna tupla: exito, mensaje_error)
        email_enviado, error_mensaje = enviar_credenciales_empleado(empleado, contraseña_temporal, es_reenvio=False)
        
        if email_enviado:
            messages.success(
                request, 
                f'✅ Acceso otorgado a {empleado.nombre_completo}. '
                f'Las credenciales han sido enviadas a {empleado.email}'
            )
        else:
            messages.warning(
                request,
                f'⚠️ Usuario creado para {empleado.nombre_completo}, pero hubo un problema al enviar el email. '
                f'Error: {error_mensaje}. '
                f'Contraseña temporal: {contraseña_temporal} (guárdala y compártela de forma segura)'
            )
            
    except ValueError as e:
        # Error de validación (email duplicado, etc.)
        messages.error(request, f'Error: {str(e)}')
    except Exception as e:
        # Cualquier otro error
        messages.error(request, f'Error al crear el usuario: {str(e)}')
    
    return redirect('lista_empleados')


@login_required
@staff_required
def reenviar_credenciales(request, empleado_id):
    """
    Reenvía las credenciales al empleado
    Genera una nueva contraseña temporal y envía email
    Solo accesible para usuarios staff/superusuario
    """
    from .utils import generar_contraseña_temporal, enviar_credenciales_empleado
    
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # Validaciones
    if not empleado.user:
        messages.warning(request, f'{empleado.nombre_completo} no tiene acceso al sistema todavía.')
        return redirect('lista_empleados')
    
    if not empleado.email:
        messages.error(request, 'El empleado no tiene un email registrado.')
        return redirect('lista_empleados')
    
    try:
        # Generar nueva contraseña temporal
        nueva_contraseña = generar_contraseña_temporal()
        
        # Actualizar contraseña del usuario
        empleado.user.set_password(nueva_contraseña)
        empleado.user.save()
        
        # Marcar que necesita configurar contraseña nuevamente
        empleado.contraseña_configurada = False
        empleado.save()
        
        # Enviar email (retorna tupla: exito, mensaje_error)
        email_enviado, error_mensaje = enviar_credenciales_empleado(empleado, nueva_contraseña, es_reenvio=True)
        
        if email_enviado:
            messages.success(
                request,
                f'✅ Credenciales reenviadas a {empleado.nombre_completo} ({empleado.email})'
            )
        else:
            messages.warning(
                request,
                f'⚠️ Contraseña reseteada pero hubo un problema al enviar el email. '
                f'Error: {error_mensaje}. '
                f'Nueva contraseña: {nueva_contraseña} (compártela de forma segura)'
            )
            
    except Exception as e:
        messages.error(request, f'Error al reenviar credenciales: {str(e)}')
    
    return redirect('lista_empleados')


@login_required
@staff_required
def resetear_contraseña_empleado(request, empleado_id):
    """
    Resetea la contraseña del empleado a una nueva temporal
    Similar a reenviar_credenciales pero con mensaje diferente
    Solo accesible para usuarios staff/superusuario
    """
    from .utils import generar_contraseña_temporal, enviar_credenciales_empleado
    
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # Validaciones
    if not empleado.user:
        messages.warning(request, f'{empleado.nombre_completo} no tiene acceso al sistema.')
        return redirect('lista_empleados')
    
    if not empleado.email:
        messages.error(request, 'El empleado no tiene un email registrado.')
        return redirect('lista_empleados')
    
    try:
        # Generar nueva contraseña temporal
        nueva_contraseña = generar_contraseña_temporal()
        
        # Actualizar contraseña del usuario
        empleado.user.set_password(nueva_contraseña)
        empleado.user.save()
        
        # Marcar que necesita configurar contraseña
        empleado.contraseña_configurada = False
        empleado.fecha_activacion_acceso = None  # Resetear fecha de activación
        empleado.save()
        
        # Enviar email (retorna tupla: exito, mensaje_error)
        email_enviado, error_mensaje = enviar_credenciales_empleado(empleado, nueva_contraseña, es_reenvio=True)
        
        if email_enviado:
            messages.success(
                request,
                f'🔄 Contraseña reseteada para {empleado.nombre_completo}. '
                f'Nueva contraseña enviada a {empleado.email}'
            )
        else:
            messages.warning(
                request,
                f'⚠️ Contraseña reseteada pero no se pudo enviar el email. '
                f'Error: {error_mensaje}. '
                f'Nueva contraseña: {nueva_contraseña}'
            )
            
    except Exception as e:
        messages.error(request, f'Error al resetear contraseña: {str(e)}')
    
    return redirect('lista_empleados')


@login_required
@staff_required
def revocar_acceso_empleado(request, empleado_id):
    """
    Revoca el acceso al sistema de un empleado
    Desactiva su usuario pero mantiene el registro del empleado
    Solo accesible para usuarios staff/superusuario
    
    EXPLICACIÓN:
    No elimina el usuario ni el empleado, solo desactiva el acceso.
    El empleado sigue en la base de datos pero no puede iniciar sesión.
    """
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # Validar que tenga usuario
    if not empleado.user:
        messages.warning(request, f'{empleado.nombre_completo} no tiene acceso al sistema.')
        return redirect('lista_empleados')
    
    if request.method == 'POST':
        try:
            # Desactivar el usuario de Django
            empleado.user.is_active = False
            empleado.user.save()
            
            # Actualizar campos del empleado
            empleado.tiene_acceso_sistema = False
            empleado.contraseña_configurada = False
            empleado.save()
            
            messages.success(
                request,
                f'🔒 Acceso revocado para {empleado.nombre_completo}. '
                f'El empleado ya no puede iniciar sesión en el sistema.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al revocar acceso: {str(e)}')
        
        return redirect('lista_empleados')
    
    # Mostrar confirmación
    return render(request, 'inventario/confirmar_revocar_acceso.html', {
        'empleado': empleado,
    })


@login_required
@staff_required
def reactivar_acceso_empleado(request, empleado_id):
    """
    Reactiva el acceso al sistema de un empleado previamente revocado
    Vuelve a activar su usuario y opcionalmente genera nueva contraseña
    Solo accesible para usuarios staff/superusuario
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista se usa cuando un empleado tuvo acceso, se le revocó, y ahora
    necesitamos dárselo de vuelta. A diferencia de "dar acceso" (que crea
    un nuevo usuario), esta función solo reactiva el usuario existente.
    
    Flujo:
    1. Verificar que el empleado tenga un usuario desactivado
    2. Reactivar el usuario (is_active = True)
    3. Generar nueva contraseña temporal (por seguridad)
    4. Enviar email con las nuevas credenciales
    5. Actualizar campos de control
    """
    from .utils import generar_contraseña_temporal, enviar_credenciales_empleado
    
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    # ===== VALIDACIONES =====
    
    # Validación 1: Debe tener un usuario existente
    if not empleado.user:
        messages.warning(
            request, 
            f'{empleado.nombre_completo} no tiene usuario creado. '
            f'Usa "Dar Acceso" en lugar de "Reactivar".'
        )
        return redirect('lista_empleados')
    
    # Validación 2: El usuario debe estar inactivo (revocado)
    if empleado.user.is_active:
        messages.info(
            request,
            f'{empleado.nombre_completo} ya tiene acceso activo al sistema.'
        )
        return redirect('lista_empleados')
    
    # Validación 3: Debe tener email para recibir credenciales
    if not empleado.email:
        messages.error(
            request,
            f'El empleado {empleado.nombre_completo} no tiene un email registrado. '
            f'Por favor, actualiza su información primero.'
        )
        return redirect('editar_empleado', empleado_id=empleado.id)
    
    # ===== PROCESO DE REACTIVACIÓN =====
    
    try:
        # Paso 1: Generar nueva contraseña temporal (por seguridad)
        # No usamos la contraseña anterior por seguridad
        nueva_contraseña = generar_contraseña_temporal()
        
        # Paso 2: Actualizar la contraseña del usuario
        empleado.user.set_password(nueva_contraseña)
        
        # Paso 3: Reactivar el usuario
        empleado.user.is_active = True
        empleado.user.save()
        
        # Paso 4: Actualizar campos del empleado
        empleado.tiene_acceso_sistema = True
        empleado.contraseña_configurada = False  # Debe cambiarla en primer login
        empleado.fecha_activacion_acceso = None  # Se establecerá cuando cambie la contraseña
        empleado.save()
        
        # Paso 5: Enviar email con las nuevas credenciales (retorna tupla: exito, mensaje_error)
        email_enviado, error_mensaje = enviar_credenciales_empleado(
            empleado, 
            nueva_contraseña, 
            es_reenvio=True  # Usar template de reenvío
        )
        
        # Paso 6: Mostrar mensaje de éxito
        if email_enviado:
            messages.success(
                request,
                f'✅ Acceso reactivado para {empleado.nombre_completo}. '
                f'Las nuevas credenciales han sido enviadas a {empleado.email}'
            )
        else:
            messages.warning(
                request,
                f'⚠️ Acceso reactivado pero hubo un problema al enviar el email. '
                f'Error: {error_mensaje}. '
                f'Contraseña temporal: {nueva_contraseña} '
                f'(guárdala y compártela de forma segura con el empleado)'
            )
    
    except Exception as e:
        # Si hay cualquier error, mostrarlo
        messages.error(
            request,
            f'❌ Error al reactivar acceso: {str(e)}'
        )
    
    return redirect('lista_empleados')


# ===== REPORTES =====
@login_required
def descargar_reporte_excel(request):
    """
    Genera y descarga un reporte completo en Excel con información del inventario
    """
    # Crear el workbook (archivo Excel)
    wb = openpyxl.Workbook()
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Título del reporte
    ws_resumen['A1'] = 'REPORTE DE INVENTARIO'
    ws_resumen['A1'].font = Font(bold=True, size=16)
    ws_resumen['A2'] = f'Generado el: {fecha_local(timezone.now()).strftime("%d/%m/%Y %H:%M")}'
    
    # Estadísticas generales
    ws_resumen['A4'] = 'ESTADÍSTICAS GENERALES'
    ws_resumen['A4'].font = header_font
    ws_resumen['A4'].fill = header_fill
    
    # Calcular estadísticas generales
    total_productos = Producto.objects.count()
    productos_normales = Producto.objects.filter(es_fraccionable=False, es_objeto_unico=False)
    productos_fraccionarios = Producto.objects.filter(es_fraccionable=True)
    productos_objetos_unicos = Producto.objects.filter(es_objeto_unico=True)
    
    # Stock bajo por tipo de producto
    productos_stock_bajo_normal = productos_normales.filter(cantidad__lte=F('stock_minimo')).count()
    productos_stock_bajo_fraccionario = sum(1 for p in productos_fraccionarios if p.stock_fraccionario_bajo())
    productos_objetos_unicos_no_disponibles = productos_objetos_unicos.filter(cantidad=0).count()
    productos_stock_bajo = productos_stock_bajo_normal + productos_stock_bajo_fraccionario + productos_objetos_unicos_no_disponibles
    
    # Cálculos de inventario
    valor_total_inventario = Producto.objects.aggregate(
        total=Sum(F('cantidad') * F('costo_unitario'))
    )['total'] or 0
    total_unidades = Producto.objects.aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Escribir estadísticas mejoradas
    estadisticas = [
        ['Total de Productos:', total_productos],
        ['  - Productos Normales:', productos_normales.count()],
        ['  - Productos Fraccionarios:', productos_fraccionarios.count()],
        ['  - Objetos Únicos:', productos_objetos_unicos.count()],
        ['Productos con Alertas:', productos_stock_bajo],
        ['  - Stock Bajo Normal:', productos_stock_bajo_normal],
        ['  - Stock Bajo Fraccionario:', productos_stock_bajo_fraccionario],
        ['  - Objetos Únicos No Disponibles:', productos_objetos_unicos_no_disponibles],
        ['Total de Unidades en Stock:', total_unidades],
        ['Valor Total del Inventario:', f'${valor_total_inventario:,.2f}']
    ]
    
    for i, (descripcion, valor) in enumerate(estadisticas, start=5):
        ws_resumen[f'A{i}'] = descripcion
        ws_resumen[f'B{i}'] = valor
        ws_resumen[f'A{i}'].font = Font(bold=True)
    
    # === HOJA 2: PRODUCTOS CON STOCK BAJO ===
    ws_stock_bajo = wb.create_sheet("Stock Bajo")
    
    # Encabezados mejorados
    headers_stock = ['Código QR', 'Nombre', 'Categoría', 'Tipo', 'Stock Actual', 'Stock Mínimo', 'Diferencia', 'Estado', 'Costo Unitario', 'Valor en Stock']
    for col, header in enumerate(headers_stock, start=1):
        cell = ws_stock_bajo.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener productos con stock bajo (normales, fraccionarios y objetos únicos)
    productos_stock_bajo_list = []
    
    # Productos normales con stock bajo (excluyendo objetos únicos)
    for producto in Producto.objects.filter(es_fraccionable=False, es_objeto_unico=False, cantidad__lte=F('stock_minimo')).order_by('cantidad'):
        productos_stock_bajo_list.append(producto)
    
    # Productos fraccionarios con stock bajo
    for producto in Producto.objects.filter(es_fraccionable=True):
        if producto.stock_fraccionario_bajo():
            productos_stock_bajo_list.append(producto)
    
    # Objetos únicos no disponibles (cantidad = 0)
    for producto in Producto.objects.filter(es_objeto_unico=True, cantidad=0).order_by('nombre'):
        productos_stock_bajo_list.append(producto)
    
    for row, producto in enumerate(productos_stock_bajo_list, start=2):
        valor_stock = producto.cantidad * producto.costo_unitario
        
        if producto.es_fraccionable:
            # Información para productos fraccionarios
            tipo_producto = f"Fraccionario ({producto.unidad_base})"
            stock_actual = f"{producto.cantidad_actual:.1f} {producto.unidad_base} ({producto.porcentaje_disponible():.0f}%)"
            stock_minimo = f"{producto.cantidad_minima_alerta:.1f} {producto.unidad_base}"
            diferencia = f"{max(0, producto.cantidad_minima_alerta - producto.cantidad_actual):.1f} {producto.unidad_base}"
            estado = "Crítico Fraccionario" if producto.cantidad_actual <= producto.cantidad_minima_alerta else "Bajo Fraccionario"
        elif producto.es_objeto_unico:
            # Información para objetos únicos
            tipo_producto = "Objeto Único"
            stock_actual = f"{producto.cantidad} unidades"
            stock_minimo = f"{producto.stock_minimo} unidades (permite 0)"
            diferencia = f"N/A (solo alerta cuando = 0)"
            estado = "No Disponible" if producto.cantidad == 0 else "Disponible"
        else:
            # Información para productos normales
            tipo_producto = "Normal (unidades)"
            stock_actual = f"{producto.cantidad} unidades"
            stock_minimo = f"{producto.stock_minimo} unidades"
            diferencia = f"{max(0, producto.stock_minimo - producto.cantidad)} unidades"
            estado = "Crítico" if producto.cantidad == 0 else "Bajo"
        
        datos = [
            producto.codigo_qr,
            producto.nombre,
            producto.get_categoria_display(),
            tipo_producto,
            stock_actual,
            stock_minimo,
            diferencia,
            estado,
            f'${producto.costo_unitario:.2f}',
            f'${valor_stock:.2f}'
        ]
        
        for col, dato in enumerate(datos, start=1):
            cell = ws_stock_bajo.cell(row=row, column=col, value=dato)
            # Resaltar productos fraccionarios críticos
            if producto.es_fraccionable and producto.cantidad_actual <= producto.cantidad_minima_alerta:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # === HOJA 3: PRODUCTOS MÁS UTILIZADOS ===
    ws_mas_usados = wb.create_sheet("Productos Más Utilizados")
    
    # Encabezados
    headers_usados = ['Producto', 'Categoría', 'Total Utilizado (30 días)', 'Stock Actual', 'Promedio Diario']
    for col, header in enumerate(headers_usados, start=1):
        cell = ws_mas_usados.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener productos más utilizados (últimos 30 días)
    fecha_limite = timezone.now() - timedelta(days=30)
    productos_mas_usados = (
        Movimiento.objects
        .filter(fecha_movimiento__gte=fecha_limite, tipo='salida')
        .values('producto__nombre', 'producto__categoria', 'producto__cantidad')
        .annotate(total_usado=Sum('cantidad'))
        .order_by('-total_usado')[:20]  # Top 20
    )
    
    for row, item in enumerate(productos_mas_usados, start=2):
        promedio_diario = item['total_usado'] / 30
        
        datos = [
            item['producto__nombre'],
            dict(Producto.CATEGORIA_CHOICES).get(item['producto__categoria'], item['producto__categoria']),
            item['total_usado'],
            item['producto__cantidad'],
            f'{promedio_diario:.1f}'
        ]
        
        for col, dato in enumerate(datos, start=1):
            ws_mas_usados.cell(row=row, column=col, value=dato)
    
    # === HOJA 4: TODOS LOS PRODUCTOS ===
    ws_todos = wb.create_sheet("Inventario Completo")
    
    # Encabezados ampliados para incluir información fraccionaria
    headers_todos = ['Código QR', 'Nombre', 'Descripción', 'Categoría', 'Tipo', 'Stock Actual', 'Stock Disponible', 'Stock Mínimo', 'Estado Stock', 'Estado Calidad', 'Costo Unitario', 'Valor Total', 'Fecha Ingreso']
    for col, header in enumerate(headers_todos, start=1):
        cell = ws_todos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener todos los productos
    todos_productos = Producto.objects.all().order_by('nombre')
    
    for row, producto in enumerate(todos_productos, start=2):
        valor_total = producto.cantidad * producto.costo_unitario
        
        if producto.es_fraccionable:
            # Información detallada para productos fraccionarios
            tipo_producto = f"Fraccionario ({producto.unidad_base})"
            stock_actual = f"{producto.cantidad} unidades × {producto.cantidad_unitaria} {producto.unidad_base}"
            stock_disponible = f"{producto.cantidad_total_disponible():.1f} {producto.unidad_base} (Actual: {producto.cantidad_actual:.1f} {producto.unidad_base} - {producto.porcentaje_disponible():.0f}%)"
            stock_minimo = f"{producto.cantidad_minima_alerta:.1f} {producto.unidad_base}"
            estado_stock = "Crítico Fraccionario" if producto.stock_fraccionario_bajo() else "Normal"
        elif producto.es_objeto_unico:
            # Información para objetos únicos
            tipo_producto = "Objeto Único"
            stock_actual = f"{producto.cantidad} unidades"
            stock_disponible = f"{producto.cantidad} unidades ({'Disponible' if producto.cantidad > 0 else 'No Disponible'})"
            stock_minimo = f"{producto.stock_minimo} unidades (alerta solo en 0)"
            estado_stock = "No Disponible" if producto.cantidad == 0 else "Disponible"
        else:
            # Información para productos normales
            tipo_producto = "Normal"
            stock_actual = f"{producto.cantidad} unidades"
            stock_disponible = f"{producto.cantidad} unidades"
            stock_minimo = f"{producto.stock_minimo} unidades"
            estado_stock = "Bajo" if producto.stock_bajo() else "Normal"
        
        datos = [
            producto.codigo_qr,
            producto.nombre,
            producto.descripcion[:50] + '...' if len(producto.descripcion) > 50 else producto.descripcion,
            producto.get_categoria_display(),
            tipo_producto,
            stock_actual,
            stock_disponible,
            stock_minimo,
            estado_stock,
            producto.get_estado_calidad_display(),
            f'${producto.costo_unitario:.2f}',
            f'${valor_total:.2f}',
            fecha_local(producto.fecha_ingreso).strftime('%d/%m/%Y')
        ]
        
        for col, dato in enumerate(datos, start=1):
            cell = ws_todos.cell(row=row, column=col, value=dato)
            # Resaltar diferentes tipos de productos
            if producto.es_fraccionable:
                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Azul claro
            elif producto.es_objeto_unico:
                cell.fill = PatternFill(start_color="F3E7FF", end_color="F3E7FF", fill_type="solid")  # Morado claro
                # Resaltar especialmente si no está disponible
                if producto.cantidad == 0:
                    cell.fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")  # Rojo claro
    
    # === HOJA 5: MOVIMIENTOS RECIENTES ===
    ws_movimientos = wb.create_sheet("Movimientos Recientes")
    
    # Encabezados mejorados para incluir información fraccionaria
    headers_mov = ['Fecha', 'Tipo Movimiento', 'Tipo Producto', 'Producto', 'Cantidad/Detalle', 'Destinatario/Área', 'Usuario Registro', 'Observaciones']
    for col, header in enumerate(headers_mov, start=1):
        cell = ws_movimientos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener movimientos recientes (últimos 100)
    movimientos_recientes = Movimiento.objects.select_related('producto', 'empleado_destinatario').order_by('-fecha_movimiento')[:100]
    
    for row, mov in enumerate(movimientos_recientes, start=2):
        destinatario = ""
        if mov.empleado_destinatario:
            destinatario = f"{mov.empleado_destinatario.nombre_completo}"
            if mov.empleado_destinatario.area:
                destinatario += f" ({mov.empleado_destinatario.area})"
        elif mov.destinatario:
            destinatario = mov.destinatario
        
        usuario_registro = ""
        if mov.usuario_registro_empleado:
            usuario_registro = mov.usuario_registro_empleado.nombre_completo
        elif mov.usuario_registro:
            usuario_registro = mov.usuario_registro
        
        # Información de cantidad diferenciada
        if mov.es_movimiento_fraccionario:
            tipo_producto = f"Fraccionario ({mov.unidad_utilizada})"
            detalle_cantidad = f"{mov.cantidad_fraccionaria:.1f} {mov.unidad_utilizada}"
            if mov.tipo == 'entrada':
                detalle_cantidad = f"+{detalle_cantidad}"
            elif mov.tipo == 'salida':
                detalle_cantidad = f"-{detalle_cantidad}"
        else:
            tipo_producto = "Normal"
            detalle_cantidad = f"{mov.cantidad} unidades"
            if mov.tipo == 'entrada':
                detalle_cantidad = f"+{detalle_cantidad}"
            elif mov.tipo == 'salida':
                detalle_cantidad = f"-{detalle_cantidad}"
        
        datos = [
            mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M'),
            mov.get_tipo_display(),
            tipo_producto,
            mov.producto.nombre,
            detalle_cantidad,
            destinatario,
            usuario_registro,
            mov.observaciones[:100] + '...' if len(mov.observaciones) > 100 else mov.observaciones
        ]
        
        for col, dato in enumerate(datos, start=1):
            cell = ws_movimientos.cell(row=row, column=col, value=dato)
            # Resaltar movimientos fraccionarios
            if mov.es_movimiento_fraccionario:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # === HOJA 6: PRODUCTOS FRACCIONARIOS (DETALLE) ===
    ws_fraccionarios = wb.create_sheet("Productos Fraccionarios")
    
    # Encabezados específicos para productos fraccionarios
    headers_frac = ['Código QR', 'Nombre', 'Categoría', 'Unidad Base', 'Cantidad por Unidad', 'Unidades en Stock', 'Cantidad Actual', 'Total Disponible', 'Porcentaje Actual', 'Mínimo Alerta', 'Estado', 'Costo Unitario', 'Valor Total']
    for col, header in enumerate(headers_frac, start=1):
        cell = ws_fraccionarios.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul para fraccionarios
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener solo productos fraccionarios
    productos_fraccionarios_list = Producto.objects.filter(es_fraccionable=True).order_by('nombre')
    
    for row, producto in enumerate(productos_fraccionarios_list, start=2):
        valor_total = producto.cantidad * producto.costo_unitario
        estado_fraccionario = ""
        
        if producto.cantidad_actual <= 0:
            estado_fraccionario = "Sin Stock"
        elif producto.stock_fraccionario_bajo():
            estado_fraccionario = "Crítico"
        elif producto.porcentaje_disponible() <= 30:
            estado_fraccionario = "Bajo"
        elif producto.porcentaje_disponible() <= 70:
            estado_fraccionario = "Medio"
        else:
            estado_fraccionario = "Óptimo"
        
        datos = [
            producto.codigo_qr,
            producto.nombre,
            producto.get_categoria_display(),
            producto.unidad_base,
            f"{producto.cantidad_unitaria:.1f} {producto.unidad_base}",
            producto.cantidad,
            f"{producto.cantidad_actual:.1f} {producto.unidad_base}",
            f"{producto.cantidad_total_disponible():.1f} {producto.unidad_base}",
            f"{producto.porcentaje_disponible():.1f}%",
            f"{producto.cantidad_minima_alerta:.1f} {producto.unidad_base}",
            estado_fraccionario,
            f'${producto.costo_unitario:.2f}',
            f'${valor_total:.2f}'
        ]
        
        for col, dato in enumerate(datos, start=1):
            cell = ws_fraccionarios.cell(row=row, column=col, value=dato)
            
            # Colorear según el estado
            if estado_fraccionario == "Sin Stock":
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Rojo
            elif estado_fraccionario == "Crítico":
                cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Naranja
            elif estado_fraccionario == "Bajo":
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Amarillo
            elif estado_fraccionario == "Óptimo":
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Verde claro
    
    # === HOJA 7: OBJETOS ÚNICOS (DETALLE) ===
    ws_objetos_unicos = wb.create_sheet("Objetos Únicos")
    
    # Encabezados específicos para objetos únicos
    headers_unicos = ['Código QR', 'Nombre', 'Descripción', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Estado Disponibilidad', 'Estado Calidad', 'Costo Unitario', 'Valor Total', 'Fecha Ingreso', 'Ubicación']
    for col, header in enumerate(headers_unicos, start=1):
        cell = ws_objetos_unicos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")  # Morado para objetos únicos
        cell.alignment = Alignment(horizontal='center')
    
    # Obtener solo objetos únicos
    objetos_unicos_list = Producto.objects.filter(es_objeto_unico=True).order_by('nombre')
    
    for row, producto in enumerate(objetos_unicos_list, start=2):
        valor_total = producto.cantidad * producto.costo_unitario
        estado_disponibilidad = producto.estado_disponibilidad()
        
        datos = [
            producto.codigo_qr,
            producto.nombre,
            producto.descripcion[:50] + '...' if len(producto.descripcion) > 50 else producto.descripcion,
            producto.get_categoria_display(),
            f"{producto.cantidad} unidades",
            f"{producto.stock_minimo} unidades (permite 0)",
            estado_disponibilidad,
            producto.get_estado_calidad_display(),
            f'${producto.costo_unitario:.2f}',
            f'${valor_total:.2f}',
            fecha_local(producto.fecha_ingreso).strftime('%d/%m/%Y'),
            producto.ubicacion or "No especificada"
        ]
        
        for col, dato in enumerate(datos, start=1):
            cell = ws_objetos_unicos.cell(row=row, column=col, value=dato)
            
            # Colorear según disponibilidad
            if estado_disponibilidad == "No Disponible":
                cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")  # Rojo claro
            else:
                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Verde claro
    
    # Ajustar ancho de columnas para todas las hojas
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_{fecha_local(timezone.now()).strftime("%Y%m%d_%H%M")}.xlsx"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response


# ===== CAMBIO DE CONTRASEÑA INICIAL (FASE 5) =====
@login_required
def cambiar_contraseña_inicial(request):
    """
    Vista para cambio obligatorio de contraseña en el primer login
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista se muestra cuando un empleado inicia sesión por primera vez
    con su contraseña temporal. El middleware redirige aquí automáticamente
    hasta que el empleado cambie su contraseña.
    
    Flujo:
    1. Empleado recibe email con contraseña temporal
    2. Inicia sesión por primera vez
    3. Middleware detecta que contraseña_configurada=False
    4. Lo redirige a esta vista
    5. Empleado cambia su contraseña
    6. Se actualiza contraseña_configurada=True
    7. Puede acceder normalmente al sistema
    
    IMPORTANTE:
    - Solo usuarios que tienen un perfil de Empleado llegan aquí
    - Staff/superusers no pasan por este flujo
    - Si el empleado ya cambió su contraseña, se redirige al dashboard
    """
    
    # Verificar que el usuario tenga un perfil de empleado asociado
    try:
        empleado = request.user.empleado
    except Empleado.DoesNotExist:
        # Si el usuario no tiene perfil de empleado (es staff/admin), redirigir
        messages.warning(request, 'Esta página es solo para empleados con acceso inicial.')
        return redirect('home')
    
    # Si ya configuró su contraseña, redirigir al dashboard
    if empleado.contraseña_configurada:
        messages.info(request, 'Ya has configurado tu contraseña previamente.')
        return redirect('home')
    
    if request.method == 'POST':
        # Procesar el formulario enviado
        from .forms import CambioContraseñaInicialForm
        form = CambioContraseñaInicialForm(request.user, request.POST)
        
        if form.is_valid():
            # Guardar la nueva contraseña
            form.save()
            
            # Actualizar campos del empleado
            empleado.contraseña_configurada = True
            empleado.fecha_activacion_acceso = timezone.now()
            empleado.save()
            
            # Mantener la sesión activa (actualizar session auth hash)
            # Sin esto, el usuario sería deslogueado al cambiar su contraseña
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            
            messages.success(
                request, 
                '✅ ¡Contraseña actualizada correctamente! '
                'Ahora puedes acceder a todas las funciones del sistema.'
            )
            return redirect('home')  # Redirigir al dashboard principal
        else:
            # Si hay errores, se mostrarán automáticamente en el template
            messages.error(
                request,
                '❌ Por favor corrige los errores en el formulario.'
            )
    else:
        # Primera carga de la página (GET request)
        from .forms import CambioContraseñaInicialForm
        form = CambioContraseñaInicialForm(request.user)
    
    # Renderizar el template con el formulario
    context = {
        'form': form,
        'empleado': empleado,
    }
    
    return render(request, 'inventario/cambiar_contraseña_inicial.html', context)


# ============================================================================
# VISTAS DE ADMINISTRACIÓN - MONITOR DE ALMACENAMIENTO
# ============================================================================

@login_required
@staff_required
def admin_storage_monitor(request):
    """
    Vista para monitorear el espacio de almacenamiento en los discos.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista muestra información sobre el espacio disponible en el disco principal
    y el disco alterno. Solo los administradores (staff) pueden acceder.
    
    Muestra:
    - Espacio total, usado y libre en cada disco
    - Qué disco está activo actualmente
    - Alertas si el espacio es bajo
    - Configuración del umbral mínimo
    
    Returns:
        HttpResponse con el template renderizado
    """
    from config.storage_utils import get_storage_info
    
    # Obtener información de almacenamiento de ambos discos
    storage_info = get_storage_info()
    
    context = {
        'storage_info': storage_info,
    }
    
    return render(request, 'admin_storage_monitor.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser, login_url='/acceso-denegado/')
def admin_clear_redis_cache(request):
    """
    Vista para limpiar la caché de Redis (base de datos 2).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Redis es una base de datos en memoria que usa Django para guardar resultados
    de dashboards y gráficas (cache). Esto acelera la carga de páginas pesadas.
    
    Esta función limpia toda la caché almacenada en Redis BD #2, forzando que
    los dashboards se regeneren con datos frescos la próxima vez.
    
    Casos de uso:
    - Después de actualizar datos masivos y quieres ver cambios inmediatos
    - Debugging de problemas de cache
    - Después de cambios en la lógica de visualizaciones
    
    Solo superusuarios pueden ejecutar esta acción (is_superuser=True).
    
    Returns:
        Redirect al storage monitor con mensaje de confirmación
    """
    from django.core.cache import cache
    
    try:
        # Limpiar toda la caché de Redis (BD 2)
        # EXPLICACIÓN: cache.clear() borra todos los datos almacenados en la cache de Django.
        # En este proyecto, eso incluye:
        # - Gráficas Plotly del dashboard de cotizaciones
        # - Resultados de consultas pesadas con @cache_page()
        # - Notificaciones cacheadas (TTL 10 segundos)
        cache.clear()
        
        messages.success(
            request,
            '✅ Caché de Redis limpiada exitosamente. '
            'Los dashboards se regenerarán con datos actualizados en la próxima carga.'
        )
    except Exception as e:
        messages.error(
            request,
            f'❌ Error al limpiar la caché de Redis: {str(e)}'
        )
    
    # Redirigir al monitor de almacenamiento o a la página anterior
    return redirect('admin_storage_monitor')


@login_required
def acceso_denegado(request):
    """
    Página de acceso denegado cuando el usuario no tiene permisos
    
    Parámetros GET:
        - mensaje: Mensaje descriptivo del error
        - permiso: Permiso que se requería (formato: app.codename)
    """
    mensaje = request.GET.get('mensaje', 'No tienes permisos para acceder a esta sección.')
    permiso = request.GET.get('permiso', '')
    
    context = {
        'mensaje': mensaje,
        'permiso': permiso,
        'user_groups': request.user.groups.all(),
    }
    
    return render(request, 'inventario/acceso_denegado.html', context)
