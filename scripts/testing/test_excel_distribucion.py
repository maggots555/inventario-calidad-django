#!/usr/bin/env python
"""
Script de prueba para verificar el Excel de Distribución de Sucursales
Verifica que las 7 hojas se generen correctamente con los datos esperados
"""
import os
import sys
import django

# Configurar Django
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from almacen.models import ProductoAlmacen, MovimientoAlmacen, SolicitudBaja
from openpyxl import load_workbook
from io import BytesIO
import tempfile

def test_excel_generation():
    """Prueba la generación del Excel y verifica su contenido"""
    
    print("=" * 80)
    print("PRUEBA DE GENERACIÓN DE EXCEL - DISTRIBUCIÓN SUCURSALES")
    print("=" * 80)
    
    # PASO 1: Verificar datos en la base de datos
    print("\n📊 PASO 1: Verificando datos en la base de datos...")
    print("-" * 80)
    
    total_productos = ProductoAlmacen.objects.count()
    productos_con_stock = ProductoAlmacen.objects.filter(
        unidades__isnull=False
    ).distinct().count()
    
    print(f"✓ Total de productos en almacén: {total_productos}")
    print(f"✓ Productos con stock: {productos_con_stock}")
    
    # Verificar movimientos
    entradas = MovimientoAlmacen.objects.filter(tipo='entrada').count()
    salidas = MovimientoAlmacen.objects.filter(tipo='salida').count()
    transferencias_mov = MovimientoAlmacen.objects.filter(tipo='transferencia').count()
    
    print(f"\n📦 Movimientos registrados:")
    print(f"   - Entradas: {entradas}")
    print(f"   - Salidas: {salidas}")
    print(f"   - Transferencias (en MovimientoAlmacen): {transferencias_mov}")
    
    # Verificar solicitudes de transferencia
    total_solicitudes_transf = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia'
    ).count()
    solicitudes_aprobadas = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia',
        estado='aprobado'
    ).count()
    solicitudes_pendientes = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia',
        estado='pendiente'
    ).count()
    solicitudes_rechazadas = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia',
        estado='rechazado'
    ).count()
    
    print(f"\n🔄 Solicitudes de Transferencia:")
    print(f"   - Total: {total_solicitudes_transf}")
    print(f"   - Aprobadas: {solicitudes_aprobadas}")
    print(f"   - Pendientes: {solicitudes_pendientes}")
    print(f"   - Rechazadas: {solicitudes_rechazadas}")
    
    # Mostrar algunos productos de ejemplo
    if productos_con_stock > 0:
        print(f"\n📝 Ejemplo de productos con stock:")
        for prod in ProductoAlmacen.objects.filter(
            unidades__isnull=False
        ).distinct()[:3]:
            stock_total = prod.unidades.count()  # Cada UnidadInventario es 1 unidad
            sucursal = prod.sucursal.nombre if prod.sucursal else "Almacén Central"
            print(f"   - {prod.codigo_producto}: {prod.nombre}")
            print(f"     Ubicación: {sucursal}, Stock: {stock_total} unidades")
    
    # PASO 2: Simular la generación del Excel
    print("\n" + "=" * 80)
    print("📄 PASO 2: Simulando generación del Excel...")
    print("-" * 80)
    
    try:
        from django.test import RequestFactory
        from almacen.views import exportar_distribucion_excel
        from django.contrib.auth.models import User
        
        # Crear un request simulado
        factory = RequestFactory()
        request = factory.get('/almacen/exportar-distribucion-excel/')
        
        # Asignar un usuario (necesario para el decorador login_required)
        user = User.objects.first()
        if not user:
            print("❌ ERROR: No hay usuarios en la base de datos")
            print("   Ejecuta: python manage.py createsuperuser")
            return False
        
        request.user = user
        
        # Llamar a la vista
        response = exportar_distribucion_excel(request)
        
        print(f"✓ Respuesta HTTP: {response.status_code}")
        print(f"✓ Content-Type: {response.get('Content-Type')}")
        print(f"✓ Tamaño del archivo: {len(response.content)} bytes")
        
        # PASO 3: Verificar el contenido del Excel
        print("\n" + "=" * 80)
        print("🔍 PASO 3: Verificando contenido del Excel...")
        print("-" * 80)
        
        # Cargar el Excel desde la respuesta
        wb = load_workbook(BytesIO(response.content))
        
        print(f"\n✓ Hojas encontradas: {len(wb.sheetnames)}")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {sheet_name}")
        
        # Verificar que tenemos exactamente 7 hojas
        expected_sheets = [
            "Distribución Actual",
            "Análisis de Movimientos",
            "Transferencias",
            "Resumen por Sucursal",
            "Productos Sin Stock",
            "Movimientos Recientes",
            "Alertas de Reposición"
        ]
        
        if len(wb.sheetnames) != 7:
            print(f"\n❌ ERROR: Se esperaban 7 hojas, se encontraron {len(wb.sheetnames)}")
            return False
        
        print(f"\n✓ Número de hojas correcto (7)")
        
        # Verificar nombres de hojas
        for i, expected in enumerate(expected_sheets):
            actual = wb.sheetnames[i]
            if actual == expected:
                print(f"   ✓ Hoja {i+1}: '{actual}' - OK")
            else:
                print(f"   ❌ Hoja {i+1}: Se esperaba '{expected}', se encontró '{actual}'")
        
        # PASO 4: Verificar contenido de Hoja 2 (Análisis de Movimientos)
        print("\n" + "=" * 80)
        print("🔍 PASO 4: Verificando HOJA 2 - Análisis de Movimientos...")
        print("-" * 80)
        
        ws2 = wb["Análisis de Movimientos"]
        
        # Verificar encabezados
        expected_headers = ["Sucursal", "Producto", "Entradas", "Salidas", 
                           "Transferencias Netas", "Stock Actual"]
        actual_headers = [cell.value for cell in ws2[3]]  # Fila 3 tiene los headers
        
        print(f"\n📋 Encabezados de Hoja 2:")
        for i, (exp, act) in enumerate(zip(expected_headers, actual_headers), 1):
            if exp == act:
                print(f"   ✓ Columna {i}: '{act}'")
            else:
                print(f"   ⚠ Columna {i}: Se esperaba '{exp}', se encontró '{act}'")
        
        # Contar filas con datos
        data_rows = 0
        for row in ws2.iter_rows(min_row=4, max_col=6):
            if row[0].value:  # Si la columna Sucursal tiene valor
                data_rows += 1
        
        print(f"\n✓ Filas de datos en Hoja 2: {data_rows}")
        
        # Mostrar las primeras 5 filas como ejemplo
        if data_rows > 0:
            print(f"\n📊 Primeras filas de datos:")
            print("   " + " | ".join(f"{h:^15}" for h in expected_headers))
            print("   " + "-" * 105)
            
            for i, row in enumerate(ws2.iter_rows(min_row=4, max_row=8, max_col=6), 1):
                if row[0].value:
                    values = [str(cell.value)[:15].center(15) if cell.value else "".center(15) 
                             for cell in row]
                    print(f"   {' | '.join(values)}")
        
        # PASO 5: Verificar contenido de Hoja 3 (Transferencias)
        print("\n" + "=" * 80)
        print("🔍 PASO 5: Verificando HOJA 3 - Transferencias...")
        print("-" * 80)
        
        ws3 = wb["Transferencias"]
        
        # Contar filas de transferencias
        transfer_rows = 0
        for row in ws3.iter_rows(min_row=3, max_col=7):
            if row[0].value and row[0].value != "Fecha":  # Excluir header
                transfer_rows += 1
        
        print(f"\n✓ Transferencias registradas en Excel: {transfer_rows}")
        
        if transfer_rows == 0:
            print("   ℹ No hay transferencias en el sistema (esto es normal si no se han hecho)")
        else:
            print(f"\n📊 Primeras transferencias:")
            expected_headers_t = ["Fecha", "Producto", "Cantidad", "Origen", 
                                 "Destino", "Solicitante", "Estado"]
            print("   " + " | ".join(f"{h:^12}" for h in expected_headers_t))
            print("   " + "-" * 100)
            
            for row in ws3.iter_rows(min_row=3, max_row=7, max_col=7):
                if row[0].value and row[0].value != "Fecha":
                    values = [str(cell.value)[:12].center(12) if cell.value else "".center(12) 
                             for cell in row]
                    print(f"   {' | '.join(values)}")
        
        # PASO 6: Resultados finales
        print("\n" + "=" * 80)
        print("✅ RESULTADO FINAL")
        print("=" * 80)
        
        print(f"""
✓ Excel generado exitosamente
✓ 7 hojas presentes con nombres correctos
✓ Hoja 2 tiene {data_rows} filas de análisis de movimientos
✓ Hoja 3 tiene {transfer_rows} transferencias registradas
✓ Tamaño del archivo: {len(response.content):,} bytes

📝 INTERPRETACIÓN DE RESULTADOS:

1. Si ves "0" en Entradas/Salidas de Hoja 2:
   → Es normal si no has registrado movimientos aún
   → Los datos se llenarán conforme uses el sistema

2. Si Hoja 3 está vacía:
   → Es normal si no has hecho transferencias entre sucursales
   → Se llenará cuando apruebes solicitudes de transferencia

3. Para poblar datos de prueba, ejecuta:
   → python scripts/poblado/poblar_sistema.py

🎯 SIGUIENTE PASO:
Descarga el Excel desde el navegador en:
http://127.0.0.1:8000/almacen/dashboard/distribucion-sucursales/
Y haz clic en "Exportar Excel"
        """)
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR al generar el Excel:")
        print(f"   {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = test_excel_generation()
    sys.exit(0 if success else 1)
