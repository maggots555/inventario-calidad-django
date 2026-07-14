"""
Vistas para el módulo Almacén - Sistema de Inventario de Almacén Central

EXPLICACIÓN PARA PRINCIPIANTES:
-------------------------------
Este archivo contiene las vistas (views) que procesan las solicitudes HTTP.
Cada vista:
1. Recibe una solicitud (request) del navegador
2. Procesa los datos (consulta BD, valida formularios)
3. Retorna una respuesta (HTML renderizado)

Patrones utilizados:
- Vistas basadas en funciones (function-based views)
- Decoradores para control de acceso (@login_required)
- Mensajes flash (messages) para feedback al usuario
- Paginación para listas largas

Organización:
- Dashboard y vistas principales
- CRUD de Proveedores
- CRUD de Categorías  
- CRUD de Productos
- Gestión de Solicitudes
- Auditorías
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django.views.decorators.clickjacking import xframe_options_sameorigin
from functools import wraps

from .models import (
    Proveedor,
    CategoriaAlmacen,
    ProductoAlmacen,
    CompraProducto,
    UnidadCompra,
    MovimientoAlmacen,
    SolicitudBaja,
    Auditoria,
    DiferenciaAuditoria,
    UnidadInventario,
    SolicitudCotizacion,
    LineaCotizacion,
    ImagenLineaCotizacion,
    ImagenSolicitudCotizacion,
    LineaServicioAdicional,
)
from .forms import (
    ProveedorForm,
    CategoriaAlmacenForm,
    ProductoAlmacenForm,
    CompraProductoForm,
    UnidadCompraForm,
    UnidadCompraFormSet,
    RecepcionCompraForm,
    ProblemaCompraForm,
    RechazoCotizacionForm,
    DevolucionCompraForm,
    MovimientoAlmacenForm,
    SolicitudBajaForm,
    ProcesarSolicitudForm,
    AuditoriaForm,
    DiferenciaAuditoriaForm,
    BusquedaProductoForm,
    UnidadInventarioForm,
    UnidadInventarioFiltroForm,
    SolicitudCotizacionForm,
    LineaCotizacionForm,
    LineaCotizacionFormSet,
    LineaCotizacionFormSetCreacion,
    SolicitudCotizacionFiltroForm,
    RespuestaLineaCotizacionForm,
    ImagenLineaCotizacionForm,
    LineaServicioAdicionalForm,
    LineaServicioAdicionalFormSet,
)
from inventario.models import Empleado

import logging

logger = logging.getLogger('almacen')


# ============================================================================
# DECORADOR PERSONALIZADO PARA PERMISOS
# ============================================================================
def permission_required_with_message(perm, message=None):
    """
    Decorador personalizado para verificar permisos con redirección a página de acceso denegado.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Este decorador verifica que el usuario tenga el permiso requerido.
    Si NO lo tiene, redirige a una página amigable explicando el problema.
    
    Args:
        perm (str): Permiso requerido en formato 'app.permiso_modelo'
                   Ejemplo: 'almacen.view_productoalmacen'
        message (str, opcional): Mensaje personalizado de error
    
    Uso:
        @login_required
        @permission_required_with_message('almacen.view_productoalmacen')
        def lista_productos(request):
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            # Verificar si el usuario tiene el permiso
            if not request.user.has_perm(perm):
                # Mensaje personalizado o genérico
                error_msg = message or 'No tienes permisos para realizar esta acción.'
                
                # Redirigir a la página de acceso denegado con parámetros
                return redirect(
                    f"{reverse('almacen:acceso_denegado_almacen')}?mensaje={error_msg}&permiso={perm}"
                )
            
            # Si tiene permiso, ejecutar la vista normalmente
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ============================================================================
# DASHBOARD PRINCIPAL
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def dashboard_almacen(request):
    """
    Dashboard principal del módulo Almacén.
    
    Muestra KPIs y resúmenes:
    - Solicitudes pendientes
    - Productos con stock bajo
    - Valor total del inventario
    - Últimos movimientos
    - Auditorías en proceso
    """
    
    # KPIs principales
    solicitudes_pendientes = SolicitudBaja.objects.filter(
        estado='pendiente'
    ).count()
    
    productos_stock_bajo = ProductoAlmacen.objects.filter(
        activo=True,
        tipo_producto='resurtible',
        stock_actual__lte=F('stock_minimo')
    ).count()
    
    productos_agotados = ProductoAlmacen.objects.filter(
        activo=True,
        stock_actual=0
    ).count()
    
    # Valor total del inventario
    valor_inventario = ProductoAlmacen.objects.filter(
        activo=True
    ).aggregate(
        total=Coalesce(
            Sum(F('stock_actual') * F('costo_unitario'), output_field=DecimalField()),
            0,
            output_field=DecimalField()
        )
    )['total']
    
    # Total de productos activos
    total_productos = ProductoAlmacen.objects.filter(activo=True).count()
    
    # Últimas solicitudes pendientes (para cola de trabajo)
    ultimas_solicitudes = SolicitudBaja.objects.filter(
        estado='pendiente'
    ).select_related(
        'producto', 'solicitante', 'orden_servicio'
    ).order_by('-fecha_solicitud')[:5]
    
    # Productos que requieren reposición
    productos_reposicion = ProductoAlmacen.objects.filter(
        activo=True,
        tipo_producto='resurtible',
        stock_actual__lte=F('stock_minimo')
    ).select_related('categoria', 'proveedor_principal')[:5]
    
    # Últimos movimientos
    ultimos_movimientos = MovimientoAlmacen.objects.select_related(
        'producto', 'empleado'
    ).order_by('-fecha')[:10]
    
    # Auditorías en proceso
    auditorias_en_proceso = Auditoria.objects.filter(
        estado='en_proceso'
    ).count()
    
    context = {
        # KPIs
        'solicitudes_pendientes': solicitudes_pendientes,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_agotados': productos_agotados,
        'valor_inventario': valor_inventario,
        'total_productos': total_productos,
        'auditorias_en_proceso': auditorias_en_proceso,
        # Listas
        'ultimas_solicitudes': ultimas_solicitudes,
        'productos_reposicion': productos_reposicion,
        'ultimos_movimientos': ultimos_movimientos,
    }
    
    return render(request, 'almacen/dashboard_almacen.html', context)


# ============================================================================
# CRUD: PROVEEDORES
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_proveedor')
def lista_proveedores(request):
    """
    Lista todos los proveedores con búsqueda y paginación.
    """
    proveedores = Proveedor.objects.all()
    
    # Búsqueda
    q = request.GET.get('q', '').strip()
    if q:
        proveedores = proveedores.filter(
            Q(nombre__icontains=q) |
            Q(contacto__icontains=q) |
            Q(email__icontains=q)
        )
    
    # Filtro por estado activo
    activo = request.GET.get('activo', '')
    if activo == '1':
        proveedores = proveedores.filter(activo=True)
    elif activo == '0':
        proveedores = proveedores.filter(activo=False)
    
    # Paginación
    paginator = Paginator(proveedores.order_by('nombre'), 20)
    page = request.GET.get('page', 1)
    proveedores_page = paginator.get_page(page)
    
    context = {
        'proveedores': proveedores_page,
        'q': q,
        'activo_filtro': activo,
    }
    
    return render(request, 'almacen/proveedores/lista_proveedores.html', context)


@login_required
@permission_required_with_message('almacen.add_proveedor')
def crear_proveedor(request):
    """
    Crea un nuevo proveedor.
    """
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" creado correctamente.')
            return redirect('almacen:lista_proveedores')
    else:
        form = ProveedorForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Proveedor',
        'boton': 'Crear Proveedor',
    }
    
    return render(request, 'almacen/proveedores/form_proveedor.html', context)


@login_required
@permission_required_with_message('almacen.change_proveedor')
def editar_proveedor(request, pk):
    """
    Edita un proveedor existente.
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" actualizado.')
            return redirect('almacen:lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'titulo': f'Editar: {proveedor.nombre}',
        'boton': 'Guardar Cambios',
    }
    
    return render(request, 'almacen/proveedores/form_proveedor.html', context)


@login_required
@permission_required_with_message('almacen.delete_proveedor')
def eliminar_proveedor(request, pk):
    """
    Elimina un proveedor (o lo desactiva si tiene productos asociados).
    """
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        # Verificar si tiene productos asociados
        if proveedor.productos_principales.exists():
            # Desactivar en lugar de eliminar
            proveedor.activo = False
            proveedor.save()
            messages.warning(
                request, 
                f'El proveedor "{proveedor.nombre}" tiene productos asociados. '
                f'Se ha desactivado en lugar de eliminar.'
            )
        else:
            nombre = proveedor.nombre
            proveedor.delete()
            messages.success(request, f'Proveedor "{nombre}" eliminado.')
        
        return redirect('almacen:lista_proveedores')
    
    context = {
        'proveedor': proveedor,
        'tiene_productos': proveedor.productos_principales.exists(),
    }
    
    return render(request, 'almacen/proveedores/eliminar_proveedor.html', context)


# ============================================================================
# CRUD: CATEGORÍAS
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_categoriaalmacen')
def lista_categorias(request):
    """
    Lista todas las categorías con conteo de productos.
    """
    categorias = CategoriaAlmacen.objects.annotate(
        num_productos=Count('productos', filter=Q(productos__activo=True))
    ).order_by('nombre')
    
    context = {
        'categorias': categorias,
    }
    
    return render(request, 'almacen/categorias/lista_categorias.html', context)


@login_required
@permission_required_with_message('almacen.add_categoriaalmacen')
def crear_categoria(request):
    """
    Crea una nueva categoría.
    """
    if request.method == 'POST':
        form = CategoriaAlmacenForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            messages.success(request, f'Categoría "{categoria.nombre}" creada.')
            return redirect('almacen:lista_categorias')
    else:
        form = CategoriaAlmacenForm()
    
    context = {
        'form': form,
        'titulo': 'Nueva Categoría',
        'boton': 'Crear Categoría',
    }
    
    return render(request, 'almacen/categorias/form_categoria.html', context)


@login_required
@permission_required_with_message('almacen.change_categoriaalmacen')
def editar_categoria(request, pk):
    """
    Edita una categoría existente.
    """
    categoria = get_object_or_404(CategoriaAlmacen, pk=pk)
    
    if request.method == 'POST':
        form = CategoriaAlmacenForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, f'Categoría "{categoria.nombre}" actualizada.')
            return redirect('almacen:lista_categorias')
    else:
        form = CategoriaAlmacenForm(instance=categoria)
    
    context = {
        'form': form,
        'categoria': categoria,
        'titulo': f'Editar: {categoria.nombre}',
        'boton': 'Guardar Cambios',
    }
    
    return render(request, 'almacen/categorias/form_categoria.html', context)


# ============================================================================
# CRUD: PRODUCTOS DE ALMACÉN
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def lista_productos(request):
    """
    Lista productos con búsqueda, filtros y paginación.
    """
    productos = ProductoAlmacen.objects.filter(activo=True).select_related(
        'categoria', 'proveedor_principal', 'sucursal'
    )
    
    # Procesar formulario de búsqueda
    form = BusquedaProductoForm(request.GET)
    
    if form.is_valid():
        # Búsqueda por texto
        q = form.cleaned_data.get('q')
        if q:
            productos = productos.filter(
                Q(codigo_producto__icontains=q) |
                Q(nombre__icontains=q) |
                Q(descripcion__icontains=q)
            )
        
        # Filtro por tipo
        tipo = form.cleaned_data.get('tipo')
        if tipo:
            productos = productos.filter(tipo_producto=tipo)
        
        # Filtro por categoría
        categoria = form.cleaned_data.get('categoria')
        if categoria:
            productos = productos.filter(categoria=categoria)
        
        # Filtro por stock
        stock = form.cleaned_data.get('stock')
        if stock == 'bajo':
            productos = productos.filter(
                tipo_producto='resurtible',
                stock_actual__lte=F('stock_minimo')
            )
        elif stock == 'agotado':
            productos = productos.filter(stock_actual=0)
        elif stock == 'disponible':
            productos = productos.filter(stock_actual__gt=0)
    
    # Contar por tipo
    total_resurtibles = productos.filter(tipo_producto='resurtible').count()
    total_unicos = productos.filter(tipo_producto='unico').count()
    
    # Paginación
    paginator = Paginator(productos.order_by('nombre'), 20)
    page = request.GET.get('page', 1)
    productos_page = paginator.get_page(page)
    
    context = {
        'productos': productos_page,
        'form': form,
        'total_resurtibles': total_resurtibles,
        'total_unicos': total_unicos,
        'total': productos.count(),
    }
    
    return render(request, 'almacen/productos/lista_productos.html', context)


@login_required
@permission_required_with_message('almacen.add_productoalmacen')
def crear_producto(request):
    """
    Crea un nuevo producto de almacén.
    """
    if request.method == 'POST':
        form = ProductoAlmacenForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.creado_por = request.user
            producto.save()
            messages.success(
                request, 
                f'Producto "{producto.codigo_producto} - {producto.nombre}" creado.'
            )
            return redirect('almacen:detalle_producto', pk=producto.pk)
    else:
        form = ProductoAlmacenForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Producto',
        'boton': 'Crear Producto',
    }
    
    return render(request, 'almacen/productos/form_producto.html', context)


@login_required
@permission_required_with_message('almacen.change_productoalmacen')
def editar_producto(request, pk):
    """
    Edita un producto existente.
    """
    producto = get_object_or_404(ProductoAlmacen, pk=pk)
    
    if request.method == 'POST':
        form = ProductoAlmacenForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado.')
            return redirect('almacen:detalle_producto', pk=producto.pk)
    else:
        form = ProductoAlmacenForm(instance=producto)
    
    context = {
        'form': form,
        'producto': producto,
        'titulo': f'Editar: {producto.codigo_producto}',
        'boton': 'Guardar Cambios',
    }
    
    return render(request, 'almacen/productos/form_producto.html', context)


@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def detalle_producto(request, pk):
    """
    Muestra el detalle completo de un producto.
    
    Incluye:
    - Información general
    - Estado del stock (total y distribución por sucursal)
    - Historial de movimientos
    - Historial de compras
    - Órdenes de servicio vinculadas
    
    ACTUALIZADO (Enero 2026): Agregada distribución por sucursal
    """
    from inventario.models import Sucursal
    from django.db.models import Count
    
    producto = get_object_or_404(
        ProductoAlmacen.objects.select_related(
            'categoria', 'proveedor_principal', 'sucursal', 'creado_por'
        ),
        pk=pk
    )
    
    # ========== DISTRIBUCIÓN POR SUCURSAL (NUEVO - Enero 2026) ==========
    # EXPLICACIÓN: Solo contamos unidades DISPONIBLES (no asignadas ni vendidas)
    # Cuando se aprueba una solicitud de servicio, la disponibilidad cambia a 'asignada'
    # y automáticamente se descuenta del conteo de la sucursal
    distribucion_sucursales = []
    
    # Almacén Central (unidades sin sucursal asignada)
    central_count = producto.unidades.filter(
        sucursal_actual__isnull=True,
        disponibilidad='disponible'  # Solo disponibles
    ).count()
    
    if central_count > 0:
        distribucion_sucursales.append({
            'nombre': 'Almacén Central',
            'codigo': 'central',
            'cantidad': central_count,
            'es_central': True
        })
    
    # Sucursales activas con unidades disponibles
    for sucursal in Sucursal.objects.filter(activa=True).annotate(
        cantidad_unidades=Count('unidades_almacenadas', 
                                filter=Q(unidades_almacenadas__producto=producto,
                                       unidades_almacenadas__disponibilidad='disponible'))  # Solo disponibles
    ):
        if sucursal.cantidad_unidades > 0:
            distribucion_sucursales.append({
                'nombre': sucursal.nombre,
                'codigo': sucursal.codigo,
                'cantidad': sucursal.cantidad_unidades,
                'es_central': False
            })
    
    # Últimos movimientos
    movimientos = producto.movimientos.select_related(
        'empleado', 'orden_servicio'
    ).order_by('-fecha')[:20]
    
    # Historial de compras
    compras = producto.historial_compras.select_related(
        'proveedor', 'orden_servicio'
    ).order_by('-fecha_recepcion', '-fecha_pedido')[:10]
    
    # Estadísticas de compras
    if compras.exists():
        from django.db.models import Avg, Min, Max
        stats_compras = producto.historial_compras.aggregate(
            costo_promedio=Avg('costo_unitario'),
            costo_minimo=Min('costo_unitario'),
            costo_maximo=Max('costo_unitario'),
            total_comprado=Sum('cantidad'),
        )
    else:
        stats_compras = None
    
    # Solicitudes pendientes para este producto
    solicitudes_pendientes = producto.solicitudes_baja.filter(
        estado__in=['pendiente', 'en_espera']
    ).select_related('solicitante')
    
    context = {
        'producto': producto,
        'distribucion_sucursales': distribucion_sucursales,  # NUEVO
        'movimientos': movimientos,
        'compras': compras,
        'stats_compras': stats_compras,
        'solicitudes_pendientes': solicitudes_pendientes,
    }
    
    return render(request, 'almacen/productos/detalle_producto.html', context)


# ============================================================================
# SOLICITUDES DE BAJA
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_solicitudbaja')
def lista_solicitudes(request):
    """
    Lista de solicitudes de baja con filtros por estado.
    """
    solicitudes = SolicitudBaja.objects.select_related(
        'producto', 'unidad_inventario', 'solicitante', 'agente_almacen', 'orden_servicio'
    ).prefetch_related('unidades_seleccionadas')
    
    # Filtro por estado
    estado = request.GET.get('estado', '')
    if estado:
        solicitudes = solicitudes.filter(estado=estado)
    
    # Por defecto, mostrar pendientes primero
    solicitudes = solicitudes.order_by(
        # Pendientes primero, luego por fecha
        '-fecha_solicitud'
    )
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page = request.GET.get('page', 1)
    solicitudes_page = paginator.get_page(page)
    
    # Contadores
    contadores = {
        'pendientes': SolicitudBaja.objects.filter(estado='pendiente').count(),
        'aprobadas': SolicitudBaja.objects.filter(estado='aprobada').count(),
        'rechazadas': SolicitudBaja.objects.filter(estado='rechazada').count(),
    }
    
    context = {
        'solicitudes': solicitudes_page,
        'estado_filtro': estado,
        'contadores': contadores,
    }
    
    return render(request, 'almacen/solicitudes/lista_solicitudes.html', context)


@login_required
@permission_required_with_message('almacen.add_solicitudbaja')
def crear_solicitud(request):
    """
    Crea una nueva solicitud de salida de productos del almacén.

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Renderiza el formulario SolicitudBajaForm con autocompletado de producto
    y orden (TypeScript en solicitud_baja_form.ts). Al enviar POST valida
    stock, técnico, prefijo OOW/FL y unidades seleccionadas.

    ACTUALIZADO (Enero 2026):
    - Procesa unidades seleccionadas del formulario
    - Valida que el empleado tenga sucursal asignada
    - Filtra unidades por sucursal (empleados) o todas (agentes de almacén)
    """
    # Obtener empleado del usuario actual
    try:
        empleado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        messages.error(
            request, 
            'No tienes un perfil de empleado asociado. Contacta al administrador.'
        )
        return redirect('almacen:dashboard_almacen')
    
    # ========== VALIDACIÓN: Empleado debe tener sucursal asignada (NUEVO) ==========
    # Solo se exige para empleados normales, agentes de almacén pueden no tenerla
    es_agente_almacen = request.user.is_staff
    
    if not es_agente_almacen and not empleado.sucursal:
        messages.error(
            request,
            'Tu perfil no tiene una sucursal asignada. '
            'No puedes crear solicitudes hasta que un administrador te asigne una sucursal. '
            'Por favor, contacta al departamento de sistemas.'
        )
        return redirect('almacen:dashboard_almacen')
    
    if request.method == 'POST':
        # Pasar parámetros extras al formulario
        form = SolicitudBajaForm(
            request.POST,
            empleado_actual=empleado,
            es_agente_almacen=es_agente_almacen
        )
        
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.solicitante = empleado
            solicitud.save()
            
            # ========== GUARDAR UNIDADES SELECCIONADAS ==========
            # Obtener IDs de unidades seleccionadas del formulario validado
            unidades_ids = form.cleaned_data.get('unidades_ids', [])
            
            if unidades_ids:
                # Obtener las unidades y agregarlas al ManyToMany
                unidades = UnidadInventario.objects.filter(id__in=unidades_ids)
                solicitud.unidades_seleccionadas.set(unidades)
            
            # Guardar formulario completo (incluyendo ManyToMany)
            form.save_m2m()
            
            messages.success(request, 'Solicitud creada correctamente.')
            return redirect('almacen:lista_solicitudes')
    else:
        # Pasar parámetros extras al formulario vacío
        form = SolicitudBajaForm(
            empleado_actual=empleado,
            es_agente_almacen=es_agente_almacen
        )
    
    context = {
        'form': form,
        'titulo': 'Nueva solicitud',
        'empleado': empleado,
        'es_agente_almacen': es_agente_almacen,
    }
    
    return render(request, 'almacen/solicitudes/form_solicitud.html', context)


@login_required
@permission_required_with_message('almacen.change_solicitudbaja')
def procesar_solicitud(request, pk):
    """
    Procesa (aprueba o rechaza) una solicitud de baja.
    Solo para agentes de almacén.
    """
    solicitud = get_object_or_404(
        SolicitudBaja.objects.select_related(
            'producto', 
            'unidad_inventario', 
            'solicitante', 
            'orden_servicio',
            'orden_servicio__detalle_equipo',  # Cargar detalle_equipo para acceder a orden_cliente
            'tecnico_asignado'  # Cargar técnico asignado
        ).prefetch_related('unidades_seleccionadas'),
        pk=pk, 
        estado='pendiente'
    )
    
    if request.method == 'POST':
        form = ProcesarSolicitudForm(request.POST)
        if form.is_valid():
            accion = form.cleaned_data['accion']
            observaciones = form.cleaned_data['observaciones']
            
            # Obtener empleado del usuario actual (agente)
            try:
                agente = Empleado.objects.get(user=request.user)
            except Empleado.DoesNotExist:
                messages.error(request, 'No tienes perfil de empleado asociado.')
                return redirect('almacen:lista_solicitudes')
            
            if accion == 'aprobar':
                solicitud.aprobar(agente, observaciones)
                messages.success(
                    request, 
                    f'Solicitud #{solicitud.pk} aprobada. Stock actualizado.'
                )
            else:
                solicitud.rechazar(agente, observaciones)
                messages.warning(request, f'Solicitud #{solicitud.pk} rechazada.')
            
            return redirect('almacen:lista_solicitudes')
    else:
        form = ProcesarSolicitudForm()
    
    context = {
        'form': form,
        'solicitud': solicitud,
    }
    
    return render(request, 'almacen/solicitudes/procesar_solicitud.html', context)


# ============================================================================
# MOVIMIENTOS (ENTRADAS/SALIDAS)
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_movimientoalmacen')
def lista_movimientos(request):
    """
    Lista de movimientos de almacén (entradas y salidas).
    """
    movimientos = MovimientoAlmacen.objects.select_related(
        'producto', 'empleado', 'orden_servicio', 'solicitud_baja'
    )
    
    # Filtro por tipo
    tipo = request.GET.get('tipo', '')
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    # Filtro por producto
    producto_id = request.GET.get('producto', '')
    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    
    movimientos = movimientos.order_by('-fecha')
    
    # Paginación
    paginator = Paginator(movimientos, 30)
    page = request.GET.get('page', 1)
    movimientos_page = paginator.get_page(page)
    
    context = {
        'movimientos': movimientos_page,
        'tipo_filtro': tipo,
    }
    
    return render(request, 'almacen/movimientos/lista_movimientos.html', context)


# ============================================================================
# API / AJAX ENDPOINTS
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def api_buscar_productos(request):
    """
    API para búsqueda de productos (autocompletado).
    Retorna JSON con productos que coinciden con la búsqueda.
    """
    q = request.GET.get('q', '').strip()
    
    if len(q) < 2:
        return JsonResponse({'productos': []})
    
    productos = ProductoAlmacen.objects.filter(
        activo=True
    ).filter(
        Q(codigo_producto__icontains=q) |
        Q(nombre__icontains=q)
    )[:10]
    
    data = {
        'productos': [
            {
                'id': p.pk,
                'codigo': p.codigo_producto,
                'nombre': p.nombre,
                'stock': p.stock_actual,
                'costo': float(p.costo_unitario),
                'tipo': p.tipo_producto,
            }
            for p in productos
        ]
    }
    
    return JsonResponse(data)


@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def api_info_producto(request, pk):
    """
    API para obtener información de un producto específico.
    """
    try:
        producto = ProductoAlmacen.objects.get(pk=pk, activo=True)
        data = {
            'success': True,
            'producto': {
                'id': producto.pk,
                'codigo': producto.codigo_producto,
                'nombre': producto.nombre,
                'stock': producto.stock_actual,
                'costo': float(producto.costo_unitario),
                'tipo': producto.tipo_producto,
                'stock_minimo': producto.stock_minimo,
                'stock_maximo': producto.stock_maximo,
            }
        }
    except ProductoAlmacen.DoesNotExist:
        data = {'success': False, 'error': 'Producto no encontrado'}
    
    return JsonResponse(data)


# ============================================================================
# VISTAS: UNIDADES DE INVENTARIO
# ============================================================================

@login_required
@permission_required_with_message('almacen.view_unidadinventario')
def lista_unidades(request):
    """
    Lista de todas las unidades individuales de inventario.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra todas las unidades físicas individuales registradas.
    
    Funcionalidades:
    - Filtrado por producto, marca, estado, disponibilidad, origen
    - Búsqueda por código interno, número de serie, modelo
    - Paginación para manejar grandes cantidades de unidades
    - Contadores de resumen (total, disponibles, por revisar)
    
    La diferencia con "lista_productos" es que aquí vemos cada unidad física
    individual, no el producto consolidado. Por ejemplo:
    - lista_productos: "SSD 1TB - Stock: 20"
    - lista_unidades: 20 registros individuales, cada uno con su marca/modelo/serie
    """
    
    # Inicializar queryset base
    unidades = UnidadInventario.objects.select_related(
        'producto',
        'producto__categoria',
        'compra',
        'orden_servicio_origen',
        'orden_servicio_destino',
        'sucursal_actual',  # NUEVO: Para mostrar ubicación
    ).order_by('-fecha_registro')
    
    # Procesar filtros
    form_filtro = UnidadInventarioFiltroForm(request.GET or None)
    
    if form_filtro.is_valid():
        # Filtrar por producto
        producto = form_filtro.cleaned_data.get('producto')
        if producto:
            unidades = unidades.filter(producto=producto)
        
        # Filtrar por marca
        marca = form_filtro.cleaned_data.get('marca')
        if marca:
            unidades = unidades.filter(marca=marca)
        
        # Filtrar por estado
        estado = form_filtro.cleaned_data.get('estado')
        if estado:
            unidades = unidades.filter(estado=estado)
        
        # Filtrar por disponibilidad
        disponibilidad = form_filtro.cleaned_data.get('disponibilidad')
        if disponibilidad:
            unidades = unidades.filter(disponibilidad=disponibilidad)
        
        # Filtrar por origen
        origen = form_filtro.cleaned_data.get('origen')
        if origen:
            unidades = unidades.filter(origen=origen)
        
        # Búsqueda de texto
        buscar = form_filtro.cleaned_data.get('buscar')
        if buscar:
            unidades = unidades.filter(
                Q(codigo_interno__icontains=buscar) |
                Q(numero_serie__icontains=buscar) |
                Q(modelo__icontains=buscar) |
                Q(producto__nombre__icontains=buscar) |
                Q(notas__icontains=buscar)
            )
    
    # ========== FILTRO POR SUCURSAL (NUEVO) ==========
    from inventario.models import Sucursal
    sucursal_filtro = request.GET.get('sucursal', '')
    
    if sucursal_filtro == 'central':
        unidades = unidades.filter(sucursal_actual__isnull=True)
    elif sucursal_filtro:
        try:
            unidades = unidades.filter(sucursal_actual_id=int(sucursal_filtro))
        except (ValueError, TypeError):
            pass
    
    # Contadores por sucursal (para pestañas)
    # IMPORTANTE: Solo contamos unidades con disponibilidad='disponible'
    # Las asignadas, vendidas o descartadas NO deben aparecer en los contadores
    from django.db.models import Count
    resumen_sucursales = []
    
    # Almacén Central (solo disponibles)
    central_count = UnidadInventario.objects.filter(
        sucursal_actual__isnull=True,
        disponibilidad='disponible'  # Solo unidades disponibles
    ).count()
    resumen_sucursales.append({
        'codigo': 'central',
        'nombre': 'Almacén Central',
        'count': central_count
    })
    
    # Sucursales (solo disponibles)
    for sucursal in Sucursal.objects.filter(activa=True):
        # Contar solo unidades disponibles en esta sucursal
        sucursal_count = UnidadInventario.objects.filter(
            sucursal_actual=sucursal,
            disponibilidad='disponible'
        ).count()
        
        resumen_sucursales.append({
            'codigo': str(sucursal.id),
            'nombre': sucursal.nombre,
            'count': sucursal_count
        })
    
    # Contadores para resumen
    total_unidades = unidades.count()
    unidades_disponibles = unidades.filter(disponibilidad='disponible').count()
    unidades_para_revision = unidades.filter(estado='para_revision').count()
    unidades_defectuosas = unidades.filter(estado='defectuoso').count()
    
    # Paginación
    paginator = Paginator(unidades, 25)  # 25 unidades por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form_filtro': form_filtro,
        'total_unidades': total_unidades,
        'unidades_disponibles': unidades_disponibles,
        'unidades_para_revision': unidades_para_revision,
        'unidades_defectuosas': unidades_defectuosas,
        'resumen_sucursales': resumen_sucursales,  # NUEVO
        'sucursal_filtro': sucursal_filtro,  # NUEVO
        'titulo': 'Unidades de Inventario',
    }
    
    return render(request, 'almacen/lista_unidades.html', context)


@login_required
@permission_required_with_message('almacen.add_unidadinventario')
def crear_unidad(request, producto_id=None):
    """
    Crear una nueva unidad de inventario.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista permite registrar una nueva unidad física individual.
    
    Si se proporciona producto_id, el formulario viene pre-seleccionado
    con ese producto (útil cuando se accede desde el detalle de un producto).
    
    Flujo:
    1. GET: Muestra formulario vacío (o con producto preseleccionado)
    2. POST: Valida datos y crea la unidad
    3. Redirige a la lista o al detalle del producto
    
    Parámetros:
    - producto_id (opcional): ID del producto para preseleccionar
    """
    
    # Si viene de un producto específico, preseleccionarlo
    producto_inicial = None
    if producto_id:
        producto_inicial = get_object_or_404(ProductoAlmacen, pk=producto_id, activo=True)
    
    if request.method == 'POST':
        form = UnidadInventarioForm(request.POST)
        
        if form.is_valid():
            unidad = form.save()
            
            messages.success(
                request,
                f'Unidad "{unidad.codigo_interno}" creada exitosamente.'
            )
            
            # Si vino de un producto, regresar al detalle del producto
            if producto_id:
                return redirect('almacen:detalle_producto', pk=producto_id)
            
            # Si no, ir a la lista de unidades
            return redirect('almacen:lista_unidades')
    else:
        # Preparar datos iniciales
        initial_data = {}
        if producto_inicial:
            initial_data['producto'] = producto_inicial
            initial_data['costo_unitario'] = producto_inicial.costo_unitario
        
        form = UnidadInventarioForm(initial=initial_data)
        
        # Si hay producto inicial, deshabilitar el campo para evitar cambios
        if producto_inicial:
            form.fields['producto'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'titulo': 'Registrar Nueva Unidad',
        'producto_inicial': producto_inicial,
    }
    
    return render(request, 'almacen/crear_unidad.html', context)


@login_required
@permission_required_with_message('almacen.view_unidadinventario')
def detalle_unidad(request, pk):
    """
    Detalle de una unidad específica de inventario.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Muestra toda la información de una unidad individual:
    - Datos del producto padre
    - Marca, modelo, número de serie
    - Estado y disponibilidad actual
    - Origen y trazabilidad (de dónde vino, a dónde fue)
    - Imágenes de referencia de la cotización (si aplica)
    - Historial de cambios (si implementado)
    
    También permite acciones rápidas como:
    - Cambiar estado/disponibilidad
    - Asignar a una orden de servicio
    - Marcar como defectuosa
    
    Trazabilidad de imágenes:
    - Si la unidad proviene de una cotización (compra.linea_cotizacion_origen)
    - Se muestran las imágenes de referencia que se subieron en la cotización
    - Esto permite verificar visualmente que la pieza recibida es correcta
    """
    
    unidad = get_object_or_404(
        UnidadInventario.objects.select_related(
            'producto',
            'producto__categoria',
            'producto__proveedor_principal',
            'compra',
            'compra__linea_cotizacion_origen',
            'compra__linea_cotizacion_origen__solicitud',
            'orden_servicio_origen',
            'orden_servicio_destino',
        ),
        pk=pk
    )
    
    # Obtener imágenes de cotización si existen
    # La trazabilidad es: UnidadInventario → compra → linea_cotizacion_origen → imagenes
    imagenes_cotizacion = None
    linea_cotizacion = None
    solicitud_cotizacion = None
    
    if unidad.compra:
        try:
            linea_cotizacion = unidad.compra.linea_cotizacion_origen
            if linea_cotizacion:
                solicitud_cotizacion = linea_cotizacion.solicitud
                imagenes_cotizacion = linea_cotizacion.imagenes.all()
        except:
            pass
    
    context = {
        'unidad': unidad,
        'titulo': f'Unidad: {unidad.codigo_interno}',
        'imagenes_cotizacion': imagenes_cotizacion,
        'linea_cotizacion': linea_cotizacion,
        'solicitud_cotizacion': solicitud_cotizacion,
    }
    
    return render(request, 'almacen/detalle_unidad.html', context)


@login_required
@permission_required_with_message('almacen.change_unidadinventario')
def editar_unidad(request, pk):
    """
    Editar una unidad de inventario existente.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Permite modificar los datos de una unidad existente.
    
    Restricciones:
    - El código interno no se puede cambiar (es autogenerado)
    - El producto padre no se puede cambiar (por integridad)
    - Los campos de sistema (fechas) son solo lectura
    
    Flujo:
    1. GET: Muestra formulario con datos actuales
    2. POST: Valida y guarda cambios
    3. Redirige al detalle de la unidad
    """
    
    unidad = get_object_or_404(UnidadInventario, pk=pk)
    
    if request.method == 'POST':
        form = UnidadInventarioForm(request.POST, instance=unidad)
        
        if form.is_valid():
            form.save()
            
            messages.success(
                request,
                f'Unidad "{unidad.codigo_interno}" actualizada exitosamente.'
            )
            
            return redirect('almacen:detalle_unidad', pk=pk)
    else:
        form = UnidadInventarioForm(instance=unidad)
        # El producto no se puede cambiar
        form.fields['producto'].widget.attrs['disabled'] = True
    
    context = {
        'form': form,
        'unidad': unidad,
        'titulo': f'Editar Unidad: {unidad.codigo_interno}',
    }
    
    return render(request, 'almacen/editar_unidad.html', context)


@login_required
@permission_required_with_message('almacen.delete_unidadinventario')
def eliminar_unidad(request, pk):
    """
    Eliminar una unidad de inventario.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista maneja la eliminación de una unidad individual.
    
    ⚠️ IMPORTANTE: La eliminación es permanente. En un sistema de producción,
    considera implementar "soft delete" (marcar como eliminado en lugar de borrar).
    
    Restricciones:
    - Solo se pueden eliminar unidades que no estén asignadas a órdenes
    - Requiere confirmación (método POST)
    
    Flujo:
    1. GET: Muestra página de confirmación
    2. POST: Elimina la unidad y redirige
    """
    
    unidad = get_object_or_404(UnidadInventario, pk=pk)
    
    # Verificar que no esté asignada a una orden activa
    if unidad.orden_servicio_destino:
        messages.error(
            request,
            f'No se puede eliminar la unidad "{unidad.codigo_interno}" '
            f'porque está asignada a la orden {unidad.orden_servicio_destino}.'
        )
        return redirect('almacen:detalle_unidad', pk=pk)
    
    if request.method == 'POST':
        codigo = unidad.codigo_interno
        producto_id = unidad.producto.pk
        
        unidad.delete()
        
        messages.success(
            request,
            f'Unidad "{codigo}" eliminada exitosamente.'
        )
        
        # Regresar al detalle del producto
        return redirect('almacen:detalle_producto', pk=producto_id)
    
    context = {
        'unidad': unidad,
        'titulo': f'Eliminar Unidad: {unidad.codigo_interno}',
    }
    
    return render(request, 'almacen/eliminar_unidad.html', context)


@login_required
@permission_required_with_message('almacen.change_unidadinventario')
def cambiar_estado_unidad(request, pk):
    """
    Cambiar rápidamente el estado de una unidad (AJAX).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta es una vista API que permite cambiar el estado de una unidad
    sin recargar toda la página (usando AJAX desde JavaScript).
    
    Acepta POST con:
    - estado: nuevo estado (nuevo, usado_bueno, defectuoso, etc.)
    - disponibilidad: nueva disponibilidad (disponible, reservada, etc.)
    
    Retorna JSON con el resultado.
    """
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    unidad = get_object_or_404(UnidadInventario, pk=pk)
    
    # Obtener nuevos valores
    nuevo_estado = request.POST.get('estado')
    nueva_disponibilidad = request.POST.get('disponibilidad')
    
    cambios = []
    
    # Validar y aplicar estado
    if nuevo_estado:
        estados_validos = [e[0] for e in UnidadInventario._meta.get_field('estado').choices]
        if nuevo_estado in estados_validos:
            unidad.estado = nuevo_estado
            cambios.append(f'estado a {unidad.get_estado_display()}')
    
    # Validar y aplicar disponibilidad
    if nueva_disponibilidad:
        disponibilidades_validas = [d[0] for d in UnidadInventario._meta.get_field('disponibilidad').choices]
        if nueva_disponibilidad in disponibilidades_validas:
            unidad.disponibilidad = nueva_disponibilidad
            cambios.append(f'disponibilidad a {unidad.get_disponibilidad_display()}')
    
    if cambios:
        unidad.save()
        return JsonResponse({
            'success': True,
            'message': f'Unidad actualizada: {", ".join(cambios)}',
            'estado': unidad.estado,
            'estado_display': unidad.get_estado_display(),
            'disponibilidad': unidad.disponibilidad,
            'disponibilidad_display': unidad.get_disponibilidad_display(),
        })
    
    return JsonResponse({'success': False, 'error': 'No se proporcionaron cambios válidos'})


@login_required
@permission_required_with_message('almacen.view_unidadinventario')
def unidades_por_producto(request, producto_id):
    """
    Lista de unidades para un producto específico.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra solo las unidades de un producto específico.
    
    Es útil cuando se accede desde el detalle de un producto y se quiere
    ver todas sus unidades individuales con más detalle.
    
    Incluye:
    - Resumen por marca (cuántas de cada marca)
    - Resumen por estado (cuántas nuevas, usadas, defectuosas)
    - Lista AGRUPADA de unidades (expandibles por grupo)
    
    AGRUPACIÓN:
    Las unidades se agrupan por: Producto + Marca + Modelo + Estado + Origen
    Esto permite mostrar "10 unidades" en lugar de 10 filas separadas.
    """
    
    producto = get_object_or_404(ProductoAlmacen, pk=producto_id)
    
    # Obtener unidades del producto
    unidades = UnidadInventario.objects.filter(
        producto=producto
    ).select_related(
        'compra',
        'orden_servicio_origen',
        'orden_servicio_destino',
    ).order_by('-fecha_registro')
    
    # Resumen por marca
    resumen_marcas = unidades.values('marca').annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')
    
    # Resumen por estado
    resumen_estados = unidades.values('estado').annotate(
        cantidad=Count('id')
    ).order_by('estado')
    
    # Resumen por disponibilidad
    resumen_disponibilidad = unidades.values('disponibilidad').annotate(
        cantidad=Count('id')
    ).order_by('disponibilidad')
    
    # AGRUPACIÓN DE UNIDADES
    # Agrupar por marca, modelo, estado y origen para mostrar grupos expandibles
    from itertools import groupby
    from operator import attrgetter
    
    # Ordenar para que groupby funcione correctamente
    unidades_ordenadas = unidades.order_by('marca', 'modelo', 'estado', 'origen', '-fecha_registro')
    
    # Agrupar unidades
    grupos = []
    for key, group in groupby(unidades_ordenadas, key=lambda u: (u.marca or 'Sin marca', u.modelo or 'Sin modelo', u.estado, u.origen)):
        unidades_grupo = list(group)
        marca, modelo, estado, origen = key
        
        # Calcular estadísticas del grupo
        disponibles_grupo = sum(1 for u in unidades_grupo if u.disponibilidad == 'disponible')
        costo_promedio = sum(u.costo_unitario or 0 for u in unidades_grupo) / len(unidades_grupo) if unidades_grupo else 0
        
        grupos.append({
            'marca': marca,
            'modelo': modelo,
            'estado': estado,
            'origen': origen,
            'cantidad': len(unidades_grupo),
            'unidades': unidades_grupo,
            'disponibles': disponibles_grupo,
            'costo_promedio': costo_promedio,
        })
    
    # Paginación de grupos (no de unidades individuales)
    paginator = Paginator(grupos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'producto': producto,
        'page_obj': page_obj,
        'total_unidades': unidades.count(),
        'unidades_disponibles': unidades.filter(disponibilidad='disponible').count(),
        'resumen_marcas': resumen_marcas,
        'resumen_estados': resumen_estados,
        'resumen_disponibilidad': resumen_disponibilidad,
        'titulo': f'Unidades de: {producto.nombre}',
        'vista_agrupada': True,  # Indicador para el template
    }
    
    return render(request, 'almacen/unidades_por_producto.html', context)


@login_required
@permission_required_with_message('almacen.view_unidadinventario')
def api_unidad_info(request, pk):
    """
    API para obtener información de una unidad específica (JSON).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista devuelve información de una unidad en formato JSON.
    
    Es útil para:
    - Mostrar tooltips o modales sin recargar la página
    - Integración con otros sistemas
    - Consultas AJAX desde JavaScript
    """
    
    try:
        unidad = UnidadInventario.objects.select_related('producto').get(pk=pk)
        
        data = {
            'success': True,
            'unidad': {
                'id': unidad.pk,
                'codigo_interno': unidad.codigo_interno,
                'producto_nombre': unidad.producto.nombre,
                'producto_id': unidad.producto.pk,
                'numero_serie': unidad.numero_serie or '',
                'marca': unidad.marca or '',
                'modelo': unidad.modelo or '',
                'estado': unidad.estado,
                'estado_display': unidad.get_estado_display(),
                'disponibilidad': unidad.disponibilidad,
                'disponibilidad_display': unidad.get_disponibilidad_display(),
                'origen': unidad.origen,
                'origen_display': unidad.get_origen_display(),
                'costo_unitario': float(unidad.costo_unitario) if unidad.costo_unitario else 0,
                'ubicacion_especifica': unidad.ubicacion_especifica or '',
                'fecha_registro': unidad.fecha_registro.strftime('%d/%m/%Y %H:%M') if unidad.fecha_registro else '',
                'notas': unidad.notas or '',
            }
        }
    except UnidadInventario.DoesNotExist:
        data = {'success': False, 'error': 'Unidad no encontrada'}
    
    return JsonResponse(data)


@login_required
@permission_required_with_message('almacen.view_unidadinventario')
def api_unidades_producto(request):
    """
    API para obtener las unidades disponibles de un producto (JSON).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista se usa en el formulario de solicitud de baja.
    Cuando el usuario selecciona un producto, JavaScript llama a esta API
    para obtener las unidades específicas (con marca/modelo/serie) disponibles.
    
    ACTUALIZACIÓN (Enero 2026):
    Ahora retorna las unidades AGRUPADAS por marca/modelo/estado para
    una mejor visualización en el formulario (similar a unidades_por_producto.html)
    
    FILTRADO POR SUCURSAL (Enero 2026):
    - Empleados normales: Solo ven unidades de su sucursal
    - Agentes de almacén (is_staff): Ven todas las unidades
    
    Parámetros GET:
    - producto_id: ID del ProductoAlmacen
    
    Retorna:
    - grupos: Lista de grupos de unidades (marca/modelo/estado)
    - unidades: Lista plana de todas las unidades (para compatibilidad)
    - stock_info: Información del stock del producto
    """
    producto_id = request.GET.get('producto_id')
    
    if not producto_id:
        return JsonResponse({
            'success': False,
            'error': 'Se requiere producto_id',
            'unidades': [],
            'grupos': [],
            'stock_info': ''
        })
    
    try:
        # ========== OBTENER EMPLEADO Y VERIFICAR PERMISOS (NUEVO - Enero 2026) ==========
        try:
            empleado = Empleado.objects.get(user=request.user)
        except Empleado.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'No tienes un perfil de empleado asociado',
                'unidades': [],
                'grupos': [],
                'stock_info': ''
            })
        
        es_agente_almacen = request.user.is_staff
        
        producto = ProductoAlmacen.objects.get(pk=producto_id)
        
        # Obtener unidades disponibles (con sucursal_actual para filtrado)
        unidades = UnidadInventario.objects.filter(
            producto_id=producto_id,
            disponibilidad='disponible'
        ).select_related('sucursal_actual')
        
        # ========== FILTRADO POR SUCURSAL DEL EMPLEADO (NUEVO - Enero 2026) ==========
        # Si NO es agente de almacén Y tiene sucursal asignada, filtrar por su sucursal
        if not es_agente_almacen and empleado.sucursal:
            unidades = unidades.filter(sucursal_actual=empleado.sucursal)
        
        # Ordenar después del filtrado
        unidades = unidades.order_by('marca', 'modelo', 'estado', 'fecha_registro')
        
        # Construir lista plana de unidades (para compatibilidad)
        unidades_data = []
        for u in unidades:
            # ========== DETECTAR SOLICITUDES PENDIENTES (NUEVO - Enero 2026) ==========
            # Verificar si esta unidad tiene una solicitud pendiente
            solicitud_pendiente = None
            tiene_solicitud_pendiente = False
            
            # Buscar en solicitudes pendientes que tienen esta unidad seleccionada
            solicitudes_pendientes = SolicitudBaja.objects.filter(
                unidades_seleccionadas=u,
                estado='pendiente'
            ).select_related('solicitante', 'orden_servicio').first()
            
            if solicitudes_pendientes:
                tiene_solicitud_pendiente = True
                solicitud_pendiente = {
                    'id': solicitudes_pendientes.pk,
                    'solicitante': solicitudes_pendientes.solicitante.nombre_completo if solicitudes_pendientes.solicitante else 'Desconocido',
                    'fecha': solicitudes_pendientes.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
                    'tipo': solicitudes_pendientes.get_tipo_solicitud_display(),
                    'cantidad': solicitudes_pendientes.cantidad,
                }
            
            unidades_data.append({
                'id': u.pk,
                'codigo_interno': u.codigo_interno or '',
                'numero_serie': u.numero_serie or '',
                'marca': u.marca or '',
                'modelo': u.modelo or '',
                'estado': u.estado,
                'estado_display': u.get_estado_display(),
                'disponibilidad': u.disponibilidad,
                'origen': u.origen,
                'origen_display': u.get_origen_display(),
                'costo_unitario': float(u.costo_unitario or 0),
                'tiene_solicitud_pendiente': tiene_solicitud_pendiente,  # NUEVO
                'solicitud_pendiente': solicitud_pendiente,  # NUEVO
                # Información de sucursal (Enero 2026)
                'sucursal_actual': {
                    'codigo': u.sucursal_actual.codigo,
                    'nombre': u.sucursal_actual.nombre,
                } if u.sucursal_actual else None,
            })
        
        # ========== AGRUPACIÓN DE UNIDADES ==========
        # Similar a unidades_por_producto view
        from itertools import groupby
        
        grupos_data = []
        for key, group in groupby(unidades, key=lambda u: (u.marca or 'Sin marca', u.modelo or 'Sin modelo', u.estado)):
            unidades_grupo = list(group)
            marca, modelo, estado = key
            
            # Construir lista de unidades del grupo
            unidades_grupo_data = []
            for u in unidades_grupo:
                # Detectar solicitudes pendientes
                solicitudes_pendientes = SolicitudBaja.objects.filter(
                    unidades_seleccionadas=u,
                    estado='pendiente'
                ).select_related('solicitante').first()
                
                tiene_solicitud_pendiente = False
                solicitud_pendiente = None
                
                if solicitudes_pendientes:
                    tiene_solicitud_pendiente = True
                    solicitud_pendiente = {
                        'id': solicitudes_pendientes.pk,
                        'solicitante': solicitudes_pendientes.solicitante.nombre_completo if solicitudes_pendientes.solicitante else 'Desconocido',
                        'fecha': solicitudes_pendientes.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
                        'tipo': solicitudes_pendientes.get_tipo_solicitud_display(),
                    }
                
                unidades_grupo_data.append({
                    'id': u.pk,
                    'codigo_interno': u.codigo_interno or '',
                    'numero_serie': u.numero_serie or '',
                    'costo_unitario': float(u.costo_unitario or 0),
                    'fecha_registro': u.fecha_registro.strftime('%d/%m/%Y'),
                    'origen': u.origen,
                    'origen_display': u.get_origen_display(),
                    'tiene_solicitud_pendiente': tiene_solicitud_pendiente,  # NUEVO
                    'solicitud_pendiente': solicitud_pendiente,  # NUEVO
                    # Información de sucursal (Enero 2026)
                    'sucursal_actual': {
                        'codigo': u.sucursal_actual.codigo,
                        'nombre': u.sucursal_actual.nombre,
                    } if u.sucursal_actual else None,
                })
            
            grupos_data.append({
                'marca': marca,
                'modelo': modelo,
                'estado': estado,
                'estado_display': dict(unidades_grupo[0]._meta.get_field('estado').choices).get(estado, estado),
                'cantidad': len(unidades_grupo),
                'unidades': unidades_grupo_data,
            })
        
        # Info de stock
        stock_info = f"Stock disponible: {producto.stock_actual} unidades"
        if unidades_data:
            stock_info += f" ({len(unidades_data)} con seguimiento individual)"
        
        return JsonResponse({
            'success': True,
            'producto_id': producto.pk,
            'producto_nombre': producto.nombre,
            'stock_actual': producto.stock_actual,
            'stock_info': stock_info,
            'unidades': unidades_data,  # Lista plana (compatibilidad)
            'grupos': grupos_data,  # Grupos (nuevo)
            'total_unidades': len(unidades_data),
            'total_grupos': len(grupos_data),
        })
        
    except ProductoAlmacen.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Producto no encontrado',
            'unidades': [],
            'grupos': [],
            'stock_info': ''
        })


@login_required
@permission_required_with_message('almacen.view_solicitudbaja')
def api_tecnicos_disponibles(request):
    """
    API para obtener la lista de técnicos de laboratorio disponibles (JSON).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista se usa en el formulario de solicitud de baja.
    Cuando el usuario selecciona "Servicio Técnico" como tipo de solicitud,
    JavaScript llama a esta API para obtener la lista de técnicos disponibles.
    
    Los técnicos se filtran por:
    - activo=True: Solo empleados activos
    - cargo__icontains='TECNICO DE LABORATORIO': Solo técnicos de laboratorio
    
    Retorna:
    - success: bool indicando si la operación fue exitosa
    - tecnicos: Lista de técnicos con id y nombre
    - total: Cantidad de técnicos disponibles
    
    Ejemplo de respuesta:
    {
        "success": true,
        "tecnicos": [
            {"id": 1, "nombre": "Juan Pérez", "sucursal": "Matriz"},
            {"id": 2, "nombre": "María García", "sucursal": "Sucursal Norte"}
        ],
        "total": 2
    }
    """
    # Importar Empleado desde inventario
    from inventario.models import Empleado
    
    try:
        # Filtrar técnicos de laboratorio activos
        tecnicos = Empleado.objects.filter(
            activo=True,
            cargo__icontains='TECNICO DE LABORATORIO'
        ).select_related('sucursal').order_by('nombre_completo')
        
        # Construir lista de técnicos
        tecnicos_data = []
        for tecnico in tecnicos:
            tecnicos_data.append({
                'id': tecnico.pk,
                'nombre': tecnico.nombre_completo,
                'cargo': tecnico.cargo,
                'sucursal': tecnico.sucursal.nombre if tecnico.sucursal else 'Sin asignar',
            })
        
        return JsonResponse({
            'success': True,
            'tecnicos': tecnicos_data,
            'total': len(tecnicos_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'tecnicos': [],
            'total': 0
        })


# ============================================================================
# API: BUSCAR O CREAR ORDEN DE SERVICIO POR ORDEN_CLIENTE
# ============================================================================
@login_required
@permission_required_with_message('almacen.add_solicitudbaja')
def api_buscar_crear_orden_cliente(request):
    """
    API para buscar una orden de servicio por orden_cliente o crearla si no existe.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Usada en el formulario de Nueva solicitud del almacén. Permite:
    1. Buscar órdenes existentes por número de orden del cliente (coincidencia exacta)
    2. Crear automáticamente una orden si no existe al enviar el formulario
    
    El campo orden_cliente vive en DetalleEquipo; la búsqueda cruza esa relación.
    
    VALIDACIÓN DE FORMATO Y PREFIJO:
    - Formato base: debe empezar con 'OOW-' o 'FL-'
    - Con tipo_solicitud en GET/POST:
      - servicio_tecnico  → solo acepta OOW-
      - venta_mostrador   → solo acepta FL-
    
    MÉTODOS HTTP:
    - GET: Buscar orden existente
    - POST: Crear nueva orden si no existe
    
    PARÁMETROS GET:
    - orden_cliente (str): Número a buscar (ej. OOW-12345)
    - tipo_solicitud (str, opcional): 'servicio_tecnico' o 'venta_mostrador' para validar prefijo
    
    PARÁMETROS POST (JSON):
    - orden_cliente (str): Número de orden del cliente
    - tipo_solicitud (str, opcional): Valida prefijo OOW/FL según tipo
    - sucursal_id (int): Sucursal donde se registra la orden nueva
    - tecnico_id (int): Técnico asignado (obligatorio para servicio técnico)
    
    RETORNA:
    {
        "success": true/false,
        "found": true/false,
        "created": true/false,
        "orden_id": int,
        "orden_cliente": str,
        "numero_orden_interno": str,
        "estado": str,
        "sucursal": str,
        "error": str
    }
    """
    import re
    import json
    from servicio_tecnico.models import OrdenServicio, DetalleEquipo
    from inventario.models import Sucursal, Empleado
    
    # Validar formato de orden_cliente (debe empezar con OOW- o FL-)
    def validar_formato_orden(orden_cliente: str) -> tuple[bool, str]:
        """
        Valida que el número de orden tenga el formato correcto.
        
        Retorna:
        - (True, '') si el formato es válido
        - (False, 'mensaje de error') si el formato es inválido
        """
        if not orden_cliente:
            return False, 'El número de orden es requerido'
        
        orden_cliente = orden_cliente.strip().upper()
        
        # Verificar que empiece con OOW- o FL-
        if not (orden_cliente.startswith('OOW-') or orden_cliente.startswith('FL-')):
            return False, 'El número de orden debe empezar con "OOW-" o "FL-"'
        
        return True, ''
    
    def validar_prefijo_por_tipo(orden_cliente: str, tipo_solicitud: str) -> tuple[bool, str]:
        """
        Valida que el prefijo de la orden coincida con el tipo de solicitud.
        
        - servicio_tecnico → solo OOW- (diagnóstico)
        - venta_mostrador → solo FL- (venta mostrador)
        """
        if not tipo_solicitud or not orden_cliente:
            return True, ''
        
        orden_cliente = orden_cliente.strip().upper()
        
        if tipo_solicitud == 'servicio_tecnico' and not orden_cliente.startswith('OOW-'):
            return False, 'Para Servicio Técnico el número de orden debe empezar con "OOW-"'
        if tipo_solicitud == 'venta_mostrador' and not orden_cliente.startswith('FL-'):
            return False, 'Para Venta Mostrador el número de orden debe empezar con "FL-"'
        
        return True, ''
    
    if request.method == 'GET':
        # ========== MODO BÚSQUEDA ==========
        orden_cliente = request.GET.get('orden_cliente', '').strip().upper()
        tipo_solicitud = request.GET.get('tipo_solicitud', '').strip()
        
        if not orden_cliente:
            return JsonResponse({
                'success': False,
                'error': 'Se requiere el parámetro orden_cliente',
                'found': False
            })
        
        # Validar formato
        formato_valido, error_formato = validar_formato_orden(orden_cliente)
        if not formato_valido:
            return JsonResponse({
                'success': False,
                'error': error_formato,
                'found': False,
                'formato_invalido': True
            })
        
        # Validar prefijo según tipo de solicitud (si se envía)
        prefijo_valido, error_prefijo = validar_prefijo_por_tipo(orden_cliente, tipo_solicitud)
        if not prefijo_valido:
            return JsonResponse({
                'success': False,
                'error': error_prefijo,
                'found': False,
                'formato_invalido': True
            })
        
        # Buscar en DetalleEquipo por orden_cliente
        try:
            detalle = DetalleEquipo.objects.select_related(
                'orden', 'orden__sucursal'
            ).get(orden_cliente__iexact=orden_cliente)
            
            orden = detalle.orden
            
            return JsonResponse({
                'success': True,
                'found': True,
                'created': False,
                'orden_id': orden.pk,
                'orden_cliente': detalle.orden_cliente,
                'numero_orden_interno': orden.numero_orden_interno,
                'estado': orden.estado,
                'estado_display': orden.get_estado_display(),
                'sucursal': orden.sucursal.nombre if orden.sucursal else 'Sin asignar',
            })
            
        except DetalleEquipo.DoesNotExist:
            # No se encontró, indicar que se puede crear
            return JsonResponse({
                'success': True,
                'found': False,
                'created': False,
                'orden_cliente': orden_cliente,
                'mensaje': f'No se encontró orden con número "{orden_cliente}". Se puede crear automáticamente.'
            })
            
        except DetalleEquipo.MultipleObjectsReturned:
            # Caso raro: múltiples órdenes con mismo orden_cliente
            return JsonResponse({
                'success': False,
                'error': f'Se encontraron múltiples órdenes con el número "{orden_cliente}". Contacte al administrador.',
                'found': False
            })
    
    elif request.method == 'POST':
        # ========== MODO CREACIÓN ==========
        try:
            # Parsear datos JSON del body
            data = json.loads(request.body)
            orden_cliente = data.get('orden_cliente', '').strip().upper()
            sucursal_id = data.get('sucursal_id')
            tecnico_id = data.get('tecnico_id')
            
            # Determinar tipo de servicio según el tipo de solicitud
            # 'servicio_tecnico' → 'diagnostico', 'venta_mostrador' → 'venta_mostrador'
            tipo_solicitud = data.get('tipo_solicitud', 'servicio_tecnico')
            tipo_servicio = 'venta_mostrador' if tipo_solicitud == 'venta_mostrador' else 'diagnostico'
            
            # Validar formato
            formato_valido, error_formato = validar_formato_orden(orden_cliente)
            if not formato_valido:
                return JsonResponse({
                    'success': False,
                    'error': error_formato,
                    'created': False,
                    'formato_invalido': True
                })
            
            # Validar prefijo según tipo de solicitud
            prefijo_valido, error_prefijo = validar_prefijo_por_tipo(orden_cliente, tipo_solicitud)
            if not prefijo_valido:
                return JsonResponse({
                    'success': False,
                    'error': error_prefijo,
                    'created': False,
                    'formato_invalido': True
                })
            
            # Validar que no exista ya
            if DetalleEquipo.objects.filter(orden_cliente__iexact=orden_cliente).exists():
                return JsonResponse({
                    'success': False,
                    'error': f'Ya existe una orden con el número "{orden_cliente}"',
                    'created': False
                })
            
            # Validar sucursal
            if not sucursal_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Se requiere seleccionar una sucursal',
                    'created': False
                })
            
            try:
                sucursal = Sucursal.objects.get(pk=sucursal_id)
            except Sucursal.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Sucursal no válida',
                    'created': False
                })
            
            # Validar técnico
            if not tecnico_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Se requiere seleccionar un técnico',
                    'created': False
                })
            
            try:
                tecnico = Empleado.objects.get(pk=tecnico_id)
            except Empleado.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Técnico no válido',
                    'created': False
                })
            
            # Obtener empleado del usuario actual (responsable de seguimiento)
            try:
                responsable = Empleado.objects.get(user=request.user)
            except Empleado.DoesNotExist:
                # Si el usuario no tiene empleado asociado, usar el técnico como responsable
                responsable = tecnico
            
            # ========== CREAR ORDEN DE SERVICIO ==========
            # tipo_servicio: 'diagnostico' para Servicio Técnico, 'venta_mostrador' para Venta Mostrador
            orden = OrdenServicio.objects.create(
                sucursal=sucursal,
                responsable_seguimiento=responsable,
                tecnico_asignado_actual=tecnico,
                estado='almacen',  # Estado especial: Proveniente de Almacén
                tipo_servicio=tipo_servicio,  # Dinámico según tipo de solicitud
            )
            
            # ========== CREAR DETALLE DE EQUIPO ==========
            # Crear DetalleEquipo con datos mínimos requeridos
            DetalleEquipo.objects.create(
                orden=orden,
                orden_cliente=orden_cliente,
                tipo_equipo='Laptop',  # Valor por defecto
                marca='Otra',  # Marca genérica - se actualizará después
                modelo='Por definir',  # Se actualizará después
                numero_serie='ALMACEN-' + orden_cliente,  # Placeholder
                gama='media',  # Valor por defecto
                falla_principal='Orden creada desde Almacén - Pendiente de registrar falla',
                email_cliente='pendiente@actualizar.com',  # Placeholder
            )
            
            # Determinar descripción del tipo de orden para el mensaje
            tipo_orden_desc = 'Venta Mostrador' if tipo_servicio == 'venta_mostrador' else 'Diagnóstico'
            
            return JsonResponse({
                'success': True,
                'found': False,
                'created': True,
                'orden_id': orden.pk,
                'orden_cliente': orden_cliente,
                'numero_orden_interno': orden.numero_orden_interno,
                'estado': orden.estado,
                'estado_display': orden.get_estado_display(),
                'sucursal': sucursal.nombre,
                'tipo_servicio': tipo_servicio,
                'mensaje': f'Orden "{orden_cliente}" ({tipo_orden_desc}) creada exitosamente con estado "Proveniente de Almacén"'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Datos JSON inválidos',
                'created': False
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear la orden: {str(e)}',
                'created': False
            })
    
    else:
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido. Use GET para buscar o POST para crear.',
            'found': False
        }, status=405)


# ============================================================================
# COMPRAS Y COTIZACIONES
# ============================================================================

@login_required
@permission_required_with_message('almacen.view_compraproducto')
def lista_compras(request):
    """
    Lista de todas las compras y cotizaciones.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra una tabla con todas las compras/cotizaciones,
    permitiendo filtrar por tipo, estado, producto y proveedor.
    
    Filtros disponibles:
    - tipo: 'cotizacion' o 'compra'
    - estado: cualquier estado del workflow
    - producto: ID del producto
    - proveedor: ID del proveedor
    - orden_cliente: búsqueda por número de orden visible
    """
    compras = CompraProducto.objects.select_related(
        'producto', 'proveedor', 'orden_servicio', 'registrado_por'
    ).prefetch_related('unidades_compra')
    
    # Filtro por tipo
    tipo = request.GET.get('tipo', '')
    if tipo:
        compras = compras.filter(tipo=tipo)
    
    # Filtro por estado
    estado = request.GET.get('estado', '')
    if estado:
        compras = compras.filter(estado=estado)
    
    # Filtro por producto
    producto_id = request.GET.get('producto', '')
    if producto_id:
        compras = compras.filter(producto_id=producto_id)
    
    # Filtro por proveedor
    proveedor_id = request.GET.get('proveedor', '')
    if proveedor_id:
        compras = compras.filter(proveedor_id=proveedor_id)
    
    # Búsqueda por orden_cliente
    orden_cliente = request.GET.get('orden_cliente', '').strip()
    if orden_cliente:
        compras = compras.filter(orden_cliente__icontains=orden_cliente)
    
    compras = compras.order_by('-fecha_registro')
    
    # Paginación
    paginator = Paginator(compras, 25)
    page = request.GET.get('page', 1)
    compras_page = paginator.get_page(page)
    
    # Datos para filtros
    productos = ProductoAlmacen.objects.filter(activo=True).order_by('nombre')
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'compras': compras_page,
        'productos': productos,
        'proveedores': proveedores,
        'tipo_filtro': tipo,
        'estado_filtro': estado,
        'producto_filtro': producto_id,
        'proveedor_filtro': proveedor_id,
        'orden_cliente_filtro': orden_cliente,
    }
    
    return render(request, 'almacen/compras/lista_compras.html', context)


@login_required
@permission_required_with_message('almacen.view_solicitudcotizacion')
def panel_cotizaciones(request):
    """
    Panel de cotizaciones pendientes de aprobación.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra un dashboard específico para las cotizaciones
    que están esperando respuesta del cliente.
    
    IMPORTANTE - NUEVO SISTEMA DE COTIZACIONES:
    Ahora las cotizaciones se manejan con el modelo SolicitudCotizacion,
    que permite múltiples proveedores por cotización.
    
    Estados del nuevo sistema:
    - borrador: En preparación
    - enviada_front: Enviada a recepción para revisión
    - enviada_cliente: Enviada al cliente, esperando respuesta
    - parcialmente_aprobada: Algunas líneas aprobadas
    - totalmente_aprobada: Todas las líneas aprobadas
    - totalmente_rechazada: Todas las líneas rechazadas
    - completada: Proceso finalizado
    
    Incluye:
    - Cotizaciones en diferentes estados
    - Alertas de cotizaciones con muchos días sin respuesta
    - Estadísticas de aprobación/rechazo
    """
    from datetime import timedelta
    from django.db.models import Count, Q
    
    # Cotizaciones pendientes de respuesta del cliente
    # Incluye enviada_front (en revisión por recepción) y enviada_cliente (con el cliente)
    cotizaciones_pendientes = SolicitudCotizacion.objects.filter(
        estado__in=['enviada_front', 'enviada_cliente']
    ).select_related(
        'orden_servicio', 'creado_por'
    ).prefetch_related('lineas').order_by('-fecha_creacion')
    
    # Cotizaciones en borrador (aún no enviadas)
    cotizaciones_borrador = SolicitudCotizacion.objects.filter(
        estado='borrador'
    ).count()
    
    # Alertas: cotizaciones con más de 3 días sin respuesta
    fecha_limite = timezone.now() - timedelta(days=3)
    cotizaciones_urgentes = cotizaciones_pendientes.filter(
        fecha_creacion__lt=fecha_limite
    ).count()
    
    # Estadísticas del mes
    inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    cotizaciones_mes = SolicitudCotizacion.objects.filter(
        fecha_creacion__gte=inicio_mes
    )
    
    total_mes = cotizaciones_mes.count()
    aprobadas_mes = cotizaciones_mes.filter(
        estado__in=['parcialmente_aprobada', 'totalmente_aprobada', 'completada']
    ).count()
    rechazadas_mes = cotizaciones_mes.filter(estado='totalmente_rechazada').count()
    
    tasa_aprobacion = (aprobadas_mes / total_mes * 100) if total_mes > 0 else 0
    
    context = {
        'cotizaciones': cotizaciones_pendientes,
        'cotizaciones_urgentes': cotizaciones_urgentes,
        'cotizaciones_borrador': cotizaciones_borrador,
        'total_pendientes': cotizaciones_pendientes.count(),
        'estadisticas': {
            'total_mes': total_mes,
            'aprobadas_mes': aprobadas_mes,
            'rechazadas_mes': rechazadas_mes,
            'tasa_aprobacion': round(tasa_aprobacion, 1),
        }
    }
    
    return render(request, 'almacen/cotizaciones/panel_cotizaciones.html', context)


@login_required
@permission_required_with_message('almacen.add_compraproducto')
def crear_compra(request):
    """
    Crear nueva COMPRA DIRECTA con unidades individuales.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista maneja un formulario con "formset", que es una técnica
    de Django para manejar múltiples formularios relacionados.
    
    IMPORTANTE - SISTEMA DE COTIZACIONES:
    Las cotizaciones ahora se manejan en un sistema separado (SolicitudCotizacion)
    que permite múltiples proveedores por cotización. Esta vista es 
    EXCLUSIVAMENTE para compras directas.
    
    Estructura:
    - Formulario principal: CompraProductoForm (producto, cantidad, etc.)
    - Formset: UnidadCompraFormSet (detalles de cada pieza individual)
    
    Cuando el usuario guarda:
    1. Se valida el formulario principal
    2. Se validan todas las unidades del formset
    3. Se guarda la compra (tipo='compra' automáticamente)
    4. Se guardan las unidades vinculadas a la compra
    """
    if request.method == 'POST':
        form = CompraProductoForm(request.POST)
        formset = UnidadCompraFormSet(request.POST, prefix='unidades')
        
        if form.is_valid():
            # Guardar compra sin commit para agregar campos adicionales
            compra = form.save(commit=False)
            compra.registrado_por = request.user
            
            # SIEMPRE es compra directa (cotizaciones usan SolicitudCotizacion)
            compra.tipo = 'compra'
            compra.estado = 'pendiente_llegada'
            
            # Costo unitario inicial en 0 (se calculará después)
            compra.costo_unitario = 0
            
            compra.save()
            
            # Ahora procesar el formset de unidades
            formset = UnidadCompraFormSet(request.POST, prefix='unidades', instance=compra)
            
            if formset.is_valid():
                # VALIDACIÓN 1: Verificar que haya al menos una unidad
                unidades_validas = [
                    f for f in formset 
                    if f.cleaned_data and not f.cleaned_data.get('DELETE', False)
                ]
                
                if not unidades_validas:
                    compra.delete()
                    messages.error(
                        request,
                        'Error: Debes especificar al menos una línea de detalle con marca y costo.'
                    )
                    form = CompraProductoForm(request.POST)
                    formset = UnidadCompraFormSet(request.POST, prefix='unidades')
                    context = {
                        'form': form,
                        'formset': formset,
                        'titulo': 'Nueva Compra Directa',
                        'es_creacion': True,
                    }
                    return render(request, 'almacen/compras/form_compra.html', context)
                
                # VALIDACIÓN 2: Verificar que la suma de cantidades = cantidad total
                suma_cantidades = sum(
                    f.cleaned_data.get('cantidad', 0) for f in unidades_validas
                )
                
                if suma_cantidades != compra.cantidad:
                    compra.delete()
                    messages.error(
                        request,
                        f'Error: La suma de cantidades ({suma_cantidades}) '
                        f'no coincide con la cantidad total ({compra.cantidad}). '
                        f'Ajusta las cantidades para que sumen exactamente {compra.cantidad}.'
                    )
                    form = CompraProductoForm(request.POST)
                    formset = UnidadCompraFormSet(request.POST, prefix='unidades')
                    context = {
                        'form': form,
                        'formset': formset,
                        'titulo': 'Nueva Compra Directa',
                        'es_creacion': True,
                    }
                    return render(request, 'almacen/compras/form_compra.html', context)
                
                # VALIDACIÓN 3: Verificar que todas las unidades tengan marca y costo
                for i, unidad_form in enumerate(unidades_validas, start=1):
                    marca = unidad_form.cleaned_data.get('marca')
                    costo = unidad_form.cleaned_data.get('costo_unitario')
                    
                    if not marca:
                        compra.delete()
                        messages.error(
                            request,
                            f'Error en línea {i}: La marca es obligatoria.'
                        )
                        form = CompraProductoForm(request.POST)
                        formset = UnidadCompraFormSet(request.POST, prefix='unidades')
                        context = {
                            'form': form,
                            'formset': formset,
                            'titulo': 'Nueva Compra Directa',
                            'es_creacion': True,
                        }
                        return render(request, 'almacen/compras/form_compra.html', context)
                    
                    if not costo or costo <= 0:
                        compra.delete()
                        messages.error(
                            request,
                            f'Error en línea {i}: El costo unitario es obligatorio y debe ser mayor a 0.'
                        )
                        form = CompraProductoForm(request.POST)
                        formset = UnidadCompraFormSet(request.POST, prefix='unidades')
                        context = {
                            'form': form,
                            'formset': formset,
                            'titulo': 'Nueva Compra Directa',
                            'es_creacion': True,
                        }
                        return render(request, 'almacen/compras/form_compra.html', context)
                
                # Guardar las unidades
                unidades = formset.save(commit=False)
                
                # Asignar número de línea secuencial
                for i, unidad in enumerate(unidades, start=1):
                    unidad.numero_linea = i
                    unidad.save()
                
                # Eliminar las marcadas para borrar
                for obj in formset.deleted_objects:
                    obj.delete()
                
                # CALCULAR Y ACTUALIZAR COSTO PROMEDIO
                compra.actualizar_costo_desde_unidades()
                
                messages.success(
                    request,
                    f'Compra #{compra.pk} creada exitosamente para {compra.producto.nombre} '
                    f'({compra.cantidad} unidades @ ${compra.costo_unitario:.2f} promedio)'
                )
                return redirect('almacen:detalle_compra', pk=compra.pk)
            else:
                # Si el formset tiene errores, eliminar la compra creada
                compra.delete()
                messages.error(request, 'Error en los detalles de unidades. Verifica que todas las líneas tengan marca y costo.')
        else:
            messages.error(request, 'Error en el formulario. Verifica los datos.')
            formset = UnidadCompraFormSet(request.POST, prefix='unidades')
    else:
        form = CompraProductoForm(initial={
            'fecha_pedido': timezone.now().date(),
            'costo_unitario': 0,  # Se calculará automáticamente
        })
        formset = UnidadCompraFormSet(prefix='unidades')
    
    context = {
        'form': form,
        'formset': formset,
        'titulo': 'Nueva Compra Directa',
        'es_creacion': True,
    }
    
    return render(request, 'almacen/compras/form_compra.html', context)


@login_required
@permission_required_with_message('almacen.view_compraproducto')
def detalle_compra(request, pk):
    """
    Detalle de una compra o cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Muestra toda la información de una compra/cotización:
    - Datos generales (producto, cantidad, costos)
    - Estado actual y botones de acción disponibles
    - Lista de unidades individuales con sus estados
    - Historial de cambios (si aplica)
    """
    # linea_cotizacion_origen: necesario para ETA (tiempo_entrega_estimado)
    # sin disparar una query extra al leer las propiedades de llegada.
    compra = get_object_or_404(
        CompraProducto.objects.select_related(
            'producto',
            'proveedor',
            'orden_servicio',
            'registrado_por',
            'linea_cotizacion_origen',
        ).prefetch_related('unidades_compra'),
        pk=pk
    )
    
    # Obtener unidades ordenadas por número de línea
    unidades = compra.unidades_compra.all().order_by('numero_linea')
    
    # Calcular estadísticas de unidades
    total_unidades = unidades.count()
    unidades_recibidas = unidades.filter(estado='recibida').count()
    unidades_problema = unidades.filter(estado__in=['wpb', 'doa']).count()
    
    context = {
        'compra': compra,
        'unidades': unidades,
        'estadisticas_unidades': {
            'total': total_unidades,
            'recibidas': unidades_recibidas,
            'problema': unidades_problema,
            'pendientes': total_unidades - unidades_recibidas - unidades_problema,
        },
    }
    
    return render(request, 'almacen/compras/detalle_compra.html', context)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def editar_compra(request, pk):
    """
    Editar una compra o cotización existente.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Permite modificar una compra antes de que sea recibida.
    
    VALIDACIONES:
    1. La compra no debe estar en estado final (recibida, devuelta, cancelada)
    2. Debe haber al menos una línea de detalle con marca y costo
    3. La suma de cantidades debe coincidir con la cantidad total
    4. El costo promedio se recalcula automáticamente
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    # Validar que se puede editar
    if compra.estado in ['recibida', 'devuelta', 'cancelada']:
        messages.error(request, 'No se puede editar una compra en estado final.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        form = CompraProductoForm(request.POST, instance=compra)
        formset = UnidadCompraFormSet(request.POST, prefix='unidades', instance=compra)
        
        if form.is_valid() and formset.is_valid():
            compra_actualizada = form.save(commit=False)
            
            # VALIDACIÓN 1: Verificar que haya al menos una unidad
            unidades_validas = [
                f for f in formset 
                if f.cleaned_data and not f.cleaned_data.get('DELETE', False)
            ]
            
            if not unidades_validas:
                messages.error(
                    request,
                    'Error: Debes especificar al menos una línea de detalle con marca y costo.'
                )
                context = {
                    'form': form,
                    'formset': formset,
                    'compra': compra,
                    'titulo': f'Editar Compra #{compra.pk}',
                    'es_creacion': False,
                }
                return render(request, 'almacen/compras/form_compra.html', context)
            
            # VALIDACIÓN 2: Verificar que la suma de cantidades = cantidad total
            suma_cantidades = sum(
                f.cleaned_data.get('cantidad', 0) for f in unidades_validas
            )
            
            if suma_cantidades != compra_actualizada.cantidad:
                messages.error(
                    request,
                    f'Error: La suma de cantidades ({suma_cantidades}) '
                    f'no coincide con la cantidad total ({compra_actualizada.cantidad}). '
                    f'Ajusta las cantidades para que sumen exactamente {compra_actualizada.cantidad}.'
                )
                context = {
                    'form': form,
                    'formset': formset,
                    'compra': compra,
                    'titulo': f'Editar Compra #{compra.pk}',
                    'es_creacion': False,
                }
                return render(request, 'almacen/compras/form_compra.html', context)
            
            # VALIDACIÓN 3: Verificar que todas las unidades tengan marca y costo
            for i, unidad_form in enumerate(unidades_validas, start=1):
                marca = unidad_form.cleaned_data.get('marca')
                costo = unidad_form.cleaned_data.get('costo_unitario')
                
                if not marca:
                    messages.error(request, f'Error en línea {i}: La marca es obligatoria.')
                    context = {
                        'form': form,
                        'formset': formset,
                        'compra': compra,
                        'titulo': f'Editar Compra #{compra.pk}',
                        'es_creacion': False,
                    }
                    return render(request, 'almacen/compras/form_compra.html', context)
                
                if not costo or costo <= 0:
                    messages.error(request, f'Error en línea {i}: El costo unitario es obligatorio y debe ser mayor a 0.')
                    context = {
                        'form': form,
                        'formset': formset,
                        'compra': compra,
                        'titulo': f'Editar Compra #{compra.pk}',
                        'es_creacion': False,
                    }
                    return render(request, 'almacen/compras/form_compra.html', context)
            
            # Guardar la compra
            compra_actualizada.save()
            
            # Guardar unidades
            unidades = formset.save(commit=False)
            
            # Reasignar números de línea
            for i, unidad in enumerate(unidades, start=1):
                unidad.numero_linea = i
                unidad.save()
            
            for obj in formset.deleted_objects:
                obj.delete()
            
            # RECALCULAR COSTO PROMEDIO
            compra_actualizada.actualizar_costo_desde_unidades()
            
            messages.success(
                request, 
                f'Compra actualizada exitosamente. Costo promedio: ${compra_actualizada.costo_unitario:.2f}'
            )
            return redirect('almacen:detalle_compra', pk=pk)
    else:
        form = CompraProductoForm(instance=compra)
        formset = UnidadCompraFormSet(prefix='unidades', instance=compra)
    
    context = {
        'form': form,
        'formset': formset,
        'compra': compra,
        'titulo': f'Editar Compra #{compra.pk}',
        'es_creacion': False,
    }
    
    return render(request, 'almacen/compras/form_compra.html', context)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def aprobar_cotizacion(request, pk):
    """
    Aprobar una cotización y convertirla en compra.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando el cliente acepta la cotización:
    1. El tipo cambia de 'cotizacion' a 'compra'
    2. El estado cambia a 'pendiente_llegada'
    3. Se registra la fecha de aprobación
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if not compra.puede_aprobar():
        messages.error(request, 'Esta cotización no puede ser aprobada.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        if compra.aprobar(usuario=request.user):
            messages.success(
                request,
                f'Cotización #{compra.pk} aprobada. Estado: Pendiente de Llegada.'
            )
        else:
            messages.error(request, 'Error al aprobar la cotización.')
    
    return redirect('almacen:detalle_compra', pk=pk)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def rechazar_cotizacion(request, pk):
    """
    Rechazar una cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando el cliente no acepta la cotización:
    1. El estado cambia a 'rechazada'
    2. Se registra el motivo del rechazo
    3. La cotización queda cerrada (no se puede reactivar)
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if not compra.puede_rechazar():
        messages.error(request, 'Esta cotización no puede ser rechazada.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        form = RechazoCotizacionForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data.get('motivo', '')
            if compra.rechazar(motivo=motivo, usuario=request.user):
                messages.success(request, f'Cotización #{compra.pk} rechazada.')
            else:
                messages.error(request, 'Error al rechazar la cotización.')
            return redirect('almacen:detalle_compra', pk=pk)
    else:
        form = RechazoCotizacionForm()
    
    context = {
        'compra': compra,
        'form': form,
    }
    
    return render(request, 'almacen/compras/rechazar_cotizacion.html', context)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def recibir_compra(request, pk):
    """
    Confirmar la recepción de una compra.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando llega la compra al almacén:
    1. Se registra la fecha de recepción
    2. Se crean MovimientoAlmacen de entrada (actualiza stock)
    3. Se crean UnidadInventario automáticamente desde las UnidadCompra
    4. El estado cambia a 'recibida'
    
    FLUJO SIMPLIFICADO:
    -------------------
    Cada UnidadCompra tiene un campo 'cantidad' que indica cuántas piezas
    son de esa marca/modelo. El método UnidadCompra.recibir() crea
    N UnidadInventario según esa cantidad.
    
    Ejemplo:
    - UnidadCompra #1: cantidad=5, marca=Kingston → crea 5 UnidadInventario
    - UnidadCompra #2: cantidad=5, marca=Samsung → crea 5 UnidadInventario
    - Total: 10 UnidadInventario
    """
    compra = get_object_or_404(
        CompraProducto.objects.select_related('producto').prefetch_related('unidades_compra'),
        pk=pk
    )
    
    if not compra.puede_recibir():
        messages.error(request, 'Esta compra no puede ser recibida en su estado actual.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        form = RecepcionCompraForm(request.POST)
        if form.is_valid():
            fecha_recepcion = form.cleaned_data['fecha_recepcion']
            crear_unidades = form.cleaned_data['crear_unidades']
            observaciones = form.cleaned_data.get('observaciones', '')
            
            # Obtener empleado para el movimiento
            try:
                empleado = Empleado.objects.get(user=request.user)
            except Empleado.DoesNotExist:
                messages.error(request, 'No tienes perfil de empleado asociado.')
                return redirect('almacen:detalle_compra', pk=pk)
            
            # Obtener orden de servicio SOLO si viene de cotización
            # Esto hace que las unidades se creen como 'asignadas' (comprometidas)
            # Las compras directas quedan con orden_servicio=None → unidades 'disponibles'
            orden_servicio = None
            if compra.tipo == 'cotizacion':
                # Intentar obtener de la relación con SolicitudCotizacion
                if hasattr(compra, 'linea_cotizacion_origen') and compra.linea_cotizacion_origen:
                    if compra.linea_cotizacion_origen.solicitud:
                        orden_servicio = compra.linea_cotizacion_origen.solicitud.orden_servicio
                # Si no se encontró, intentar del campo directo
                if not orden_servicio:
                    orden_servicio = compra.orden_servicio
            
            # Recibir la compra
            if compra.recibir(fecha_recepcion=fecha_recepcion, crear_unidades=False):
                # Crear movimiento de entrada
                MovimientoAlmacen.objects.create(
                    tipo='entrada',
                    producto=compra.producto,
                    cantidad=compra.cantidad,
                    costo_unitario=compra.costo_unitario,
                    empleado=empleado,
                    compra=compra,
                    observaciones=f'Recepción de compra #{compra.pk}. {observaciones}'.strip(),
                )
                
                # Crear UnidadInventario si se solicitó
                total_unidades_creadas = 0
                
                if crear_unidades:
                    unidades_compra = compra.unidades_compra.filter(estado='pendiente')
                    
                    # NUEVO FLUJO SIMPLIFICADO:
                    # Cada UnidadCompra.recibir() crea N UnidadInventario según su cantidad
                    for unidad_compra in unidades_compra:
                        unidades_creadas = unidad_compra.recibir(
                            crear_unidad_inventario=True,
                            orden_servicio_destino=orden_servicio,
                            registrado_por=request.user
                        )
                        total_unidades_creadas += len(unidades_creadas)
                
                # Si es COTIZACIÓN, crear movimiento de SALIDA automático
                # porque la pieza va directo al servicio (no se queda en almacén)
                if compra.tipo == 'cotizacion' and orden_servicio:
                    MovimientoAlmacen.objects.create(
                        tipo='salida',
                        producto=compra.producto,
                        cantidad=compra.cantidad,
                        costo_unitario=compra.costo_unitario,
                        empleado=empleado,
                        compra=compra,
                        orden_servicio=orden_servicio,
                        observaciones=f'Asignación automática a servicio (Cotización #{compra.pk}). Orden: {orden_servicio.detalle_equipo.orden_cliente if orden_servicio.detalle_equipo else orden_servicio.pk}',
                    )
                
                mensaje_resultado = f'Compra #{compra.pk} recibida exitosamente. {total_unidades_creadas} unidades agregadas al inventario.'
                if compra.tipo == 'cotizacion' and orden_servicio:
                    mensaje_resultado += f' (Asignadas automáticamente a servicio)'
                
                messages.success(request, mensaje_resultado)
            else:
                messages.error(request, 'Error al recibir la compra.')
            
            return redirect('almacen:detalle_compra', pk=pk)
    else:
        form = RecepcionCompraForm(initial={
            'fecha_recepcion': timezone.now().date(),
            'crear_unidades': True,
        })
    
    context = {
        'compra': compra,
        'form': form,
    }
    
    return render(request, 'almacen/compras/recibir_compra.html', context)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def reportar_problema_compra(request, pk):
    """
    Reportar problema con una compra recibida (WPB o DOA).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Si la pieza recibida tiene problemas:
    - WPB (Wrong Part): Enviaron una pieza incorrecta
    - DOA (Dead On Arrival): La pieza está dañada/no funciona
    
    Al reportar:
    1. Se registra el tipo de problema y descripción
    2. El estado cambia a 'wpb' o 'doa'
    3. Se puede iniciar proceso de devolución
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if not compra.puede_marcar_problema():
        messages.error(request, 'No se puede reportar problema en esta compra.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        form = ProblemaCompraForm(request.POST)
        if form.is_valid():
            tipo_problema = form.cleaned_data['tipo_problema']
            motivo = form.cleaned_data['motivo']
            
            if tipo_problema == 'wpb':
                compra.marcar_wpb(motivo=motivo)
                messages.warning(request, f'Compra #{compra.pk} marcada como WPB (Pieza Incorrecta).')
            else:
                compra.marcar_doa(motivo=motivo)
                messages.warning(request, f'Compra #{compra.pk} marcada como DOA (Dañada al Llegar).')
            
            return redirect('almacen:detalle_compra', pk=pk)
    else:
        form = ProblemaCompraForm()
    
    context = {
        'compra': compra,
        'form': form,
    }
    
    return render(request, 'almacen/compras/problema_compra.html', context)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def iniciar_devolucion(request, pk):
    """
    Iniciar proceso de devolución al proveedor.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Después de reportar un problema (WPB/DOA), se puede iniciar
    la devolución al proveedor:
    1. El estado cambia a 'devolucion_garantia'
    2. Se prepara la pieza para envío de vuelta
    3. Cuando llegue al proveedor, se confirma la devolución
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if not compra.puede_devolver():
        messages.error(request, 'No se puede iniciar devolución para esta compra.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        if compra.iniciar_devolucion():
            messages.info(
                request,
                f'Devolución iniciada para compra #{compra.pk}. Confirma cuando sea recibida por el proveedor.'
            )
        else:
            messages.error(request, 'Error al iniciar devolución.')
    
    return redirect('almacen:detalle_compra', pk=pk)


@login_required
@permission_required_with_message('almacen.change_compraproducto')
def confirmar_devolucion(request, pk):
    """
    Confirmar que la devolución fue completada.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando el proveedor recibe la pieza devuelta:
    1. El estado cambia a 'devuelta'
    2. Se crea un MovimientoAlmacen de salida (descuenta del stock)
    3. La compra queda cerrada
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if not compra.puede_confirmar_devolucion():
        messages.error(request, 'No se puede confirmar devolución para esta compra.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        form = DevolucionCompraForm(request.POST)
        if form.is_valid():
            observaciones = form.cleaned_data.get('observaciones', '')
            numero_guia = form.cleaned_data.get('numero_guia', '')
            
            if numero_guia:
                observaciones = f'Guía: {numero_guia}. {observaciones}'.strip()
            
            # Obtener empleado
            try:
                empleado = Empleado.objects.get(user=request.user)
            except Empleado.DoesNotExist:
                empleado = None
            
            if compra.confirmar_devolucion(empleado=empleado, observaciones=observaciones):
                messages.success(
                    request,
                    f'Devolución confirmada para compra #{compra.pk}. Stock actualizado.'
                )
            else:
                messages.error(request, 'Error al confirmar devolución.')
            
            return redirect('almacen:detalle_compra', pk=pk)
    else:
        form = DevolucionCompraForm()
    
    context = {
        'compra': compra,
        'form': form,
        # Stock final después de la devolución (stock actual - cantidad a devolver)
        'stock_final': compra.producto.stock_actual - compra.cantidad if compra.producto else 0,
    }
    
    return render(request, 'almacen/compras/confirmar_devolucion.html', context)


@login_required
@permission_required_with_message('almacen.delete_compraproducto')
def cancelar_compra(request, pk):
    """
    Cancelar una compra o cotización.
    
    NOTA: No se puede cancelar si ya fue recibida sin problemas.
    """
    compra = get_object_or_404(CompraProducto, pk=pk)
    
    if compra.estado == 'recibida':
        messages.error(request, 'No se puede cancelar una compra ya recibida.')
        return redirect('almacen:detalle_compra', pk=pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '')
        if compra.cancelar(motivo=motivo):
            messages.success(request, f'Compra #{compra.pk} cancelada.')
        else:
            messages.error(request, 'Error al cancelar la compra.')
    
    return redirect('almacen:detalle_compra', pk=pk)


@login_required
@permission_required_with_message('almacen.change_unidadcompra')
def recibir_unidad_compra(request, compra_pk, pk):
    """
    Recibir una unidad individual de una compra.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Si la compra tiene UnidadCompra definidas, se puede recibir
    cada pieza individualmente en lugar de todas a la vez.
    
    Esto es útil cuando:
    - Las piezas llegan en diferentes momentos
    - Se quiere verificar cada pieza antes de darla por recibida
    """
    compra = get_object_or_404(CompraProducto, pk=compra_pk)
    unidad = get_object_or_404(UnidadCompra, pk=pk, compra=compra)
    
    if not unidad.puede_recibir():
        messages.error(request, 'Esta unidad no puede ser recibida.')
        return redirect('almacen:detalle_compra', pk=compra_pk)
    
    if request.method == 'POST':
        unidad_inv = unidad.recibir(crear_unidad_inventario=True)
        if unidad_inv:
            messages.success(
                request,
                f'Unidad #{unidad.numero_linea} recibida. Inventario: {unidad_inv.codigo_interno}'
            )
        else:
            messages.error(request, 'Error al recibir la unidad.')
    
    return redirect('almacen:detalle_compra', pk=compra_pk)


@login_required
@permission_required_with_message('almacen.change_unidadcompra')
def problema_unidad_compra(request, compra_pk, pk):
    """
    Reportar problema con una unidad específica de una compra.
    """
    compra = get_object_or_404(CompraProducto, pk=compra_pk)
    unidad = get_object_or_404(UnidadCompra, pk=pk, compra=compra)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo_problema', 'wpb')
        motivo = request.POST.get('motivo', '')
        
        if tipo == 'doa':
            unidad.marcar_doa(motivo=motivo)
            messages.warning(request, f'Unidad #{unidad.numero_linea} marcada como DOA.')
        else:
            unidad.marcar_wpb(motivo=motivo)
            messages.warning(request, f'Unidad #{unidad.numero_linea} marcada como WPB.')
    
    return redirect('almacen:detalle_compra', pk=compra_pk)


# ============================================================================
# SOLICITUDES DE COTIZACIÓN (MULTI-PROVEEDOR)
# ============================================================================

@login_required
@permission_required_with_message('almacen.view_solicitudcotizacion')
def lista_solicitudes_cotizacion(request):
    """
    Lista todas las solicitudes de cotización con filtros.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra todas las solicitudes de cotización en una tabla.
    Permite filtrar por:
    - Estado (borrador, enviada, aprobada, etc.)
    - Fecha de creación
    - Búsqueda por número de solicitud u orden
    
    También muestra un resumen con contadores por estado para una
    visión rápida del flujo de trabajo.
    """
    # Obtener todas las solicitudes
    solicitudes = SolicitudCotizacion.objects.select_related(
        'orden_servicio',
        'creado_por'
    ).prefetch_related('lineas').order_by('-fecha_creacion')
    
    # Aplicar filtros
    filtro_form = SolicitudCotizacionFiltroForm(request.GET)
    
    if filtro_form.is_valid():
        estado = filtro_form.cleaned_data.get('estado')
        fecha_desde = filtro_form.cleaned_data.get('fecha_desde')
        fecha_hasta = filtro_form.cleaned_data.get('fecha_hasta')
        buscar = filtro_form.cleaned_data.get('buscar')
        
        if estado:
            solicitudes = solicitudes.filter(estado=estado)
        
        if fecha_desde:
            solicitudes = solicitudes.filter(fecha_creacion__date__gte=fecha_desde)
        
        if fecha_hasta:
            solicitudes = solicitudes.filter(fecha_creacion__date__lte=fecha_hasta)
        
        if buscar:
            solicitudes = solicitudes.filter(
                Q(numero_solicitud__icontains=buscar) |
                Q(numero_orden_cliente__icontains=buscar)
            )
    
    # Contadores por estado para el resumen
    contadores = SolicitudCotizacion.objects.values('estado').annotate(
        total=Count('id')
    )
    contadores_dict = {c['estado']: c['total'] for c in contadores}
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page = request.GET.get('page', 1)
    solicitudes_page = paginator.get_page(page)
    
    context = {
        'solicitudes': solicitudes_page,
        'filtro_form': filtro_form,
        'contadores': contadores_dict,
        'titulo': 'Solicitudes de Cotización',
        # Para resaltar el KPI activo y mostrar el total en el encabezado
        'estado_filtro': request.GET.get('estado', ''),
        'total_general': sum(contadores_dict.values()),
    }
    
    return render(request, 'almacen/cotizaciones/lista_solicitudes.html', context)


@login_required
@permission_required_with_message('almacen.add_solicitudcotizacion')
def crear_solicitud_cotizacion(request):
    """
    Crear una nueva solicitud de cotización con múltiples líneas.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista maneja la creación de una solicitud de cotización.
    
    El proceso tiene tres partes:
    1. El formulario principal (SolicitudCotizacionForm) captura:
       - Número de orden del cliente (para vincular con servicio técnico)
       - O modo sin orden: datos del cliente + service tag + marca/modelo
       - Observaciones internas
    
    2. El formset (LineaCotizacionFormSet) captura las líneas:
       - Cada línea tiene: producto, descripción, proveedor, cantidad, costo
       - Se pueden agregar múltiples líneas dinámicamente con JavaScript
    
    3. Las imágenes de referencia (request.FILES):
       - Hasta 6 imágenes del equipo/piezas que el cliente quiere cotizar
       - Se procesan después de guardar la solicitud
    
    Flujo:
    1. GET: Muestra formularios vacíos
    2. POST: Valida ambos formularios
       - Si válidos: Guarda solicitud, líneas e imágenes, redirige a detalle
       - Si inválidos: Muestra errores
    """
    if request.method == 'POST':
        form = SolicitudCotizacionForm(request.POST)
        formset = LineaCotizacionFormSetCreacion(request.POST)
        
        if form.is_valid() and formset.is_valid():
            # Guardar la solicitud (cabecera)
            solicitud = form.save(commit=False)
            solicitud.creado_por = request.user
            solicitud.save()
            
            # Guardar las líneas (detalle)
            formset.instance = solicitud
            formset.save()
            
            # Procesar imágenes de referencia (hasta 6)
            imagenes = request.FILES.getlist('imagenes_referencia')
            descripciones = request.POST.getlist('descripcion_imagen')
            imagenes_guardadas = 0
            
            for i, imagen in enumerate(imagenes):
                if imagenes_guardadas >= ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD:
                    break
                try:
                    descripcion = descripciones[i] if i < len(descripciones) else ''
                    ImagenSolicitudCotizacion.objects.create(
                        solicitud=solicitud,
                        imagen=imagen,
                        descripcion=descripcion,
                        subido_por=request.user,
                    )
                    imagenes_guardadas += 1
                except (ValueError, Exception) as e:
                    messages.warning(request, f'Error al subir imagen: {str(e)}')
            
            if imagenes_guardadas > 0:
                messages.info(request, f'{imagenes_guardadas} imagen(es) de referencia subidas.')

            # =================================================================
            # NOTIFICAR A COMPRAS cuando la solicitud es "Sin Orden Activa"
            # =================================================================
            # EXPLICACIÓN PARA PRINCIPIANTES:
            # Cuando una solicitud se crea sin una orden de servicio vinculada
            # (modo "sin orden activa"), el área de Compras necesita saberlo
            # de inmediato para procesar la cotización. Les enviamos:
            #   1. Push al dispositivo (notificación en tiempo real)
            #   2. Campanita interna (notificación del sistema)
            #   3. Email en segundo plano vía Celery (no bloquea al usuario)
            if solicitud.sin_orden_activa:
                try:
                    from notificaciones.push_service import enviar_push_a_usuario
                    from notificaciones.utils import notificar_info
                    from .tasks import notificar_compras_nueva_cotizacion_task

                    # Buscar todos los empleados con rol "Compras" que tengan
                    # usuario activo en el sistema (pueden recibir notificaciones)
                    compradores = Empleado.objects.filter(
                        rol='compras',
                        user__is_active=True,
                    ).select_related('user')

                    if compradores.exists():
                        # Construir la URL al detalle de la solicitud para que
                        # al hacer clic en la notificación los lleve directamente
                        url_solicitud = reverse(
                            'almacen:detalle_solicitud_cotizacion',
                            kwargs={'pk': solicitud.pk}
                        )

                        # Texto de la notificación — incluye el nombre del
                        # cliente y el service tag para identificación rápida
                        titulo_push = f'📋 Nueva cotización sin orden: {solicitud.numero_solicitud}'
                        mensaje_push = (
                            f'Cliente: {solicitud.nombre_cliente or "Sin nombre"} — '
                            f'S/T: {solicitud.service_tag or "N/A"}. '
                            f'Requiere atención para procesar la cotización.'
                        )

                        # Enviar push + campanita a cada empleado de Compras
                        for comprador in compradores:
                            try:
                                enviar_push_a_usuario(
                                    usuario=comprador.user,
                                    titulo=titulo_push,
                                    mensaje=mensaje_push,
                                    url=url_solicitud,
                                )
                            except Exception as push_err:
                                logger.warning(
                                    f"[COTIZACION] Error enviando push a {comprador.nombre_completo}: {push_err}"
                                )

                            try:
                                notificar_info(
                                    titulo=titulo_push,
                                    mensaje=mensaje_push,
                                    usuario=comprador.user,
                                    url=url_solicitud,
                                    app_origen='almacen',
                                )
                            except Exception as notif_err:
                                logger.warning(
                                    f"[COTIZACION] Error creando notificación para {comprador.nombre_completo}: {notif_err}"
                                )

                        # Enviar email en segundo plano vía Celery (no bloquea)
                        from config.paises_config import get_pais_actual
                        notificar_compras_nueva_cotizacion_task.delay(
                            solicitud.pk,
                            request.user.pk,
                            get_pais_actual()['db_alias'],
                        )
                        logger.info(
                            f"[COTIZACION] Notificaciones enviadas a {compradores.count()} "
                            f"empleado(s) de Compras para solicitud {solicitud.numero_solicitud}"
                        )
                except Exception as e:
                    # Si falla la notificación, NO debe impedir que la solicitud
                    # se haya creado correctamente. Solo registramos el error.
                    logger.error(f"[COTIZACION] Error al notificar a Compras: {e}")

            messages.success(
                request,
                f'Solicitud de cotización {solicitud.numero_solicitud} creada exitosamente.'
            )
            return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = SolicitudCotizacionForm()
        formset = LineaCotizacionFormSetCreacion()

    # Obtener la sucursal del usuario logueado para mostrarla en el formulario
    # cuando se usa el modo "sin orden activa". Se muestra como dato informativo
    # (solo lectura) para que el técnico sepa desde qué sucursal está cotizando.
    sucursal_usuario = None
    try:
        sucursal_usuario = request.user.empleado.sucursal
    except Exception:
        # El usuario puede no tener perfil de empleado o sucursal asignada — es válido
        pass

    context = {
        'form': form,
        'formset': formset,
        'titulo': 'Nueva Solicitud de Cotización',
        'es_creacion': True,
        'max_imagenes_referencia': ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD,
        # Sucursal del usuario actual, para pre-mostrar en modo sin orden activa
        'sucursal_usuario': sucursal_usuario,
    }
    
    return render(request, 'almacen/cotizaciones/form_solicitud.html', context)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def editar_solicitud_cotizacion(request, pk):
    """
    Editar una solicitud de cotización existente.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Similar a crear, pero carga los datos existentes en los formularios.
    
    Solo se puede editar si la solicitud está en estado 'borrador'.
    Una vez enviada al cliente, no se puede modificar.
    
    También permite agregar más imágenes de referencia.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    # Solo se puede editar en estado borrador
    if solicitud.estado != 'borrador':
        messages.error(
            request,
            'Solo se pueden editar solicitudes en estado borrador.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
    
    if request.method == 'POST':
        form = SolicitudCotizacionForm(request.POST, instance=solicitud)
        formset = LineaCotizacionFormSet(request.POST, instance=solicitud)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            
            # Procesar imágenes de referencia nuevas (hasta el límite)
            imagenes = request.FILES.getlist('imagenes_referencia')
            descripciones = request.POST.getlist('descripcion_imagen')
            imagenes_guardadas = 0
            
            for i, imagen in enumerate(imagenes):
                if not ImagenSolicitudCotizacion.puede_agregar_imagen(solicitud):
                    messages.warning(
                        request,
                        f'Se alcanzó el límite de {ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD} imágenes.'
                    )
                    break
                try:
                    descripcion = descripciones[i] if i < len(descripciones) else ''
                    ImagenSolicitudCotizacion.objects.create(
                        solicitud=solicitud,
                        imagen=imagen,
                        descripcion=descripcion,
                        subido_por=request.user,
                    )
                    imagenes_guardadas += 1
                except (ValueError, Exception) as e:
                    messages.warning(request, f'Error al subir imagen: {str(e)}')
            
            if imagenes_guardadas > 0:
                messages.info(request, f'{imagenes_guardadas} imagen(es) de referencia agregadas.')
            
            messages.success(request, 'Solicitud actualizada exitosamente.')
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = SolicitudCotizacionForm(instance=solicitud)
        formset = LineaCotizacionFormSet(instance=solicitud)
    
    # Calcular imágenes restantes
    imagenes_actuales = ImagenSolicitudCotizacion.objects.filter(solicitud=solicitud).count()
    imagenes_restantes = max(0, ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD - imagenes_actuales)
    
    context = {
        'form': form,
        'formset': formset,
        'solicitud': solicitud,
        'titulo': f'Editar Solicitud {solicitud.numero_solicitud}',
        'es_creacion': False,
        'max_imagenes_referencia': ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD,
        'imagenes_referencia_actuales': imagenes_actuales,
        'imagenes_referencia_restantes': imagenes_restantes,
    }
    
    return render(request, 'almacen/cotizaciones/form_solicitud.html', context)


@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def editar_lineas_cotizacion(request, pk):
    """
    Editar líneas de cotización cuando la solicitud está en estado 'enviada_front'.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista permite modificar ciertos campos de las líneas de cotización
    DESPUÉS de que la solicitud fue enviada a Front, pero ANTES de que el
    cliente apruebe o se genere la compra.
    
    ¿Qué se puede editar?
    - Proveedor (si el original no tiene stock)
    - Costo unitario (si cambió el precio)
    - Cantidad (si el cliente pide más/menos)
    - Tiempo de entrega estimado
    - Notas adicionales
    
    ¿Qué NO se puede editar?
    - Producto (ya fue definido)
    - Descripción de la pieza (es la identidad)
    
    ¿Qué líneas se pueden editar?
    - Solo las que tienen estado 'pendiente' o 'rechazada'
    - Las líneas 'aprobada' o 'compra_generada' se muestran como solo lectura
    
    IMPORTANTE: El formset solo incluye las líneas editables.
    Las líneas bloqueadas se muestran como texto plano fuera del formset.
    """
    from .forms import EditarLineaCotizacionFormSet
    from .models import LineaCotizacion
    
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    # Se puede editar líneas en enviada_front o enviada_cliente (por si se necesita recotizar)
    if solicitud.estado not in ['enviada_front', 'enviada_cliente']:
        messages.error(
            request,
            'Solo se pueden editar líneas cuando la solicitud está en estado "Enviada a Front" o "Enviada a Cliente".'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
    
    # Filtrar solo líneas editables (pendiente o rechazada)
    lineas_editables_qs = solicitud.lineas.filter(
        estado_cliente__in=['pendiente', 'rechazada']
    ).order_by('numero_linea')
    
    # Líneas bloqueadas (aprobada o compra_generada) - solo para mostrar
    lineas_bloqueadas = solicitud.lineas.filter(
        estado_cliente__in=['aprobada', 'compra_generada']
    ).order_by('numero_linea')
    
    if request.method == 'POST':
        # El formset solo procesa líneas editables
        formset = EditarLineaCotizacionFormSet(
            request.POST,
            instance=solicitud,
            queryset=lineas_editables_qs
        )
        
        if formset.is_valid():
            # Guardar cambios y contar cuántas líneas se modificaron
            lineas_modificadas = 0
            
            for form in formset.forms:
                if form.has_changed():
                    form.save()
                    lineas_modificadas += 1
            
            if lineas_modificadas > 0:
                messages.success(
                    request,
                    f'Se actualizaron {lineas_modificadas} línea(s) de la cotización.'
                )
            else:
                messages.info(request, 'No se realizaron cambios en las líneas.')
            
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        formset = EditarLineaCotizacionFormSet(
            instance=solicitud,
            queryset=lineas_editables_qs
        )
    
    # Preparar información para el template
    lineas_editables_info = []
    for form, linea in zip(formset.forms, lineas_editables_qs):
        lineas_editables_info.append({
            'form': form,
            'linea': linea,
        })
    
    context = {
        'formset': formset,
        'solicitud': solicitud,
        'lineas_editables_info': lineas_editables_info,
        'lineas_bloqueadas': lineas_bloqueadas,
        'titulo': f'Editar Líneas - {solicitud.numero_solicitud}',
    }
    
    return render(request, 'almacen/cotizaciones/editar_lineas.html', context)


import json as _json  # alias para no colisionar con variables de vistas


def _serializar_profit_config() -> str:
    """
    Lee PROFIT_CONFIG desde el módulo de PDF (que a su vez lo lee del .env)
    y lo convierte a una cadena JSON lista para inyectar en el template.

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    El template necesita pasar estos valores al JavaScript del navegador.
    Al usar |safe en el template, Django inserta el JSON sin escapar las
    comillas, de modo que el navegador lo interpreta como objeto JS válido.

    Importamos dentro de la función (importación diferida) para evitar
    importaciones circulares y para mantener el módulo ligero.

    Returns:
        str: Cadena JSON con la configuración de profit por perfil.
    """
    # Importación diferida para evitar ciclos de importación entre módulos
    from .utils.pdf_cotizacion_cliente import PROFIT_CONFIG

    # Construir un diccionario serializable (las listas de costos_fijos ya lo son)
    datos = {
        perfil: {
            'profit_target':  cfg['profit_target'],
            'costos_fijos':   cfg['costos_fijos'],
            'diagnostico':    cfg['diagnostico'],
        }
        for perfil, cfg in PROFIT_CONFIG.items()
    }
    # Convertir a JSON compacto — se incrustará dentro de un <script>
    return _json.dumps(datos, separators=(',', ':'))


def _serializar_costeo_reacondicionado_config() -> str:
    """
    Serializa la configuración del costeo de reacondicionados para el modal TypeScript.

    Returns:
        str: JSON compacto con porcentajes y montos del .env.
    """
    from .utils.costeo_reacondicionado import serializar_config_costeo
    return _json.dumps(serializar_config_costeo(), separators=(',', ':'))


def _actualizar_estado_st_esperando_aprobacion_cliente(solicitud, usuario=None):
    """
    Al enviar la cotización al cliente desde Almacén, pone la orden de ST
    en estado ``cotizacion`` («Esperando Aprobación Cliente»).

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Antes, ese cambio de estado ocurría al crear la cotización en ST.
    Ahora la MO y la cotización se separaron, y el momento correcto de
    avisar «estamos esperando al cliente» es cuando realmente se le
    envía la cotización por correo desde este módulo.

    Args:
        solicitud: SolicitudCotizacion (debe tener orden_servicio vinculada).
        usuario: User opcional; si tiene empleado, se asocia al historial.

    Returns:
        bool: True si se cambió el estado; False si no había orden,
        ya estaba en ``cotizacion``, o no aplica.
    """
    # Sin orden vinculada (modo sin_orden_activa) no hay nada que actualizar en ST
    orden = getattr(solicitud, 'orden_servicio', None)
    if not orden:
        return False

    # Ya está esperando al cliente: no duplicar historial ni push
    if orden.estado == 'cotizacion':
        return False

    estado_anterior = orden.estado
    # Código de estado en constants: 'cotizacion' → «Esperando Aprobación Cliente»
    orden.estado = 'cotizacion'
    # OrdenServicio.save() registra solo el cambio de estado en el historial
    orden.save(update_fields=['estado'])

    # Enriquecer el último historial con el usuario de Almacén y un comentario claro
    empleado = None
    if usuario is not None and hasattr(usuario, 'empleado'):
        empleado = getattr(usuario, 'empleado', None)

    ultimo = (
        orden.historial.filter(tipo_evento='cambio_estado', estado_nuevo='cotizacion')
        .order_by('-fecha_evento')
        .first()
    )
    if ultimo:
        from config.constants import ESTADO_ORDEN_CHOICES
        ultimo.comentario = (
            f'Cambio de estado al enviar cotización al cliente desde Almacén: '
            f'{dict(ESTADO_ORDEN_CHOICES).get(estado_anterior, estado_anterior)} → '
            f'Esperando Aprobación Cliente '
            f'(solicitud {solicitud.numero_solicitud})'
        )
        if empleado is not None:
            ultimo.usuario = empleado
        ultimo.es_sistema = True
        ultimo.save(update_fields=['comentario', 'usuario', 'es_sistema'])

    logger.info(
        f"[API_COTIZACION_CLIENTE] Orden ST {orden.numero_orden_interno}: "
        f"{estado_anterior} → cotizacion (envío al cliente, solicitud {solicitud.numero_solicitud})"
    )
    return True


def _extraer_datos_reacondicionado_post(post) -> dict:
    """
    Lee del POST los campos del equipo reacondicionado capturados en el modal.

    Args:
        post: request.POST de Django.

    Returns:
        dict: Datos del equipo y parámetros de costeo.
    """
    return {
        'costo_proveedor': post.get('reac_costo_proveedor', '').strip(),
        'dias_front_desk': post.get('reac_dias_front_desk', '1').strip(),
        'marca': post.get('reac_marca', '').strip(),
        'modelo': post.get('reac_modelo', '').strip(),
        'procesador': post.get('reac_procesador', '').strip(),
        'ram': post.get('reac_ram', '').strip(),
        'sistema_operativo': post.get('reac_sistema_operativo', '').strip(),
        'incluye_cargador': post.get('reac_incluye_cargador') == '1',
        'especificaciones': post.get('reac_especificaciones', '').strip(),
    }


def _validar_y_calcular_reacondicionado(datos: dict):
    """
    Valida campos obligatorios y ejecuta calcular_costeo().

    Returns:
        tuple: (ok: bool, resultado_o_error: dict|str)
    """
    from .utils.costeo_reacondicionado import calcular_costeo

    if not datos.get('marca'):
        return False, 'La marca del equipo es obligatoria.'
    if not datos.get('modelo'):
        return False, 'El modelo del equipo es obligatorio.'

    try:
        costo = float(datos.get('costo_proveedor') or 0)
        if costo <= 0:
            return False, 'El costo de proveedor debe ser mayor a cero.'
    except ValueError:
        return False, 'El costo de proveedor no es un número válido.'

    try:
        dias = int(datos.get('dias_front_desk') or 1)
        if dias < 1:
            dias = 1
    except ValueError:
        return False, 'Los días de front desk deben ser un número entero válido.'

    costeo = calcular_costeo(costo_proveedor=costo, dias_front_desk=dias)
    datos_equipo = {
        'marca': datos['marca'],
        'modelo': datos['modelo'],
        'procesador': datos.get('procesador', ''),
        'ram': datos.get('ram', ''),
        'sistema_operativo': datos.get('sistema_operativo', ''),
        'incluye_cargador': datos.get('incluye_cargador', False),
        'especificaciones': datos.get('especificaciones', ''),
    }
    return True, {'costeo': costeo, 'datos_equipo': datos_equipo, 'dias_front_desk': dias, 'costo_proveedor': costo}


def _guardar_snapshot_reacondicionado(solicitud, datos_equipo, costeo, dias, costo):
    """Persiste en la solicitud el snapshot de la propuesta reacondicionada."""
    from decimal import Decimal
    solicitud.modo_cotizacion_cliente = 'reacondicionado'
    solicitud.costo_proveedor_reac = Decimal(str(costo))
    solicitud.dias_front_desk_reac = dias
    solicitud.reac_marca = datos_equipo.get('marca', '')
    solicitud.reac_modelo = datos_equipo.get('modelo', '')
    solicitud.reac_procesador = datos_equipo.get('procesador', '')
    solicitud.reac_ram = datos_equipo.get('ram', '')
    solicitud.reac_sistema_operativo = datos_equipo.get('sistema_operativo', '')
    solicitud.reac_incluye_cargador = bool(datos_equipo.get('incluye_cargador'))
    solicitud.reac_especificaciones = datos_equipo.get('especificaciones', '')
    solicitud.resultado_costeo_reac = costeo
    solicitud.save(update_fields=[
        'modo_cotizacion_cliente',
        'costo_proveedor_reac',
        'dias_front_desk_reac',
        'reac_marca',
        'reac_modelo',
        'reac_procesador',
        'reac_ram',
        'reac_sistema_operativo',
        'reac_incluye_cargador',
        'reac_especificaciones',
        'resultado_costeo_reac',
    ])


# SKU del catálogo para equipos reacondicionados ofertados al cliente
CODIGO_PRODUCTO_REACONDICIONADO = 'P0125'


def _construir_descripcion_linea_reac(datos_equipo: dict) -> str:
    """
    Arma una descripción compacta del equipo para LineaCotizacion.descripcion_pieza.

    Args:
        datos_equipo: dict con marca, modelo, procesador, ram, sistema_operativo, incluye_cargador.

    Returns:
        str: Texto truncado a 255 caracteres (límite del campo).
    """
    partes = []
    marca = (datos_equipo.get('marca') or '').strip()
    modelo = (datos_equipo.get('modelo') or '').strip()
    if marca or modelo:
        partes.append(f'{marca} {modelo}'.strip())
    if datos_equipo.get('procesador'):
        partes.append(str(datos_equipo['procesador']).strip())
    if datos_equipo.get('ram'):
        partes.append(str(datos_equipo['ram']).strip())
    if datos_equipo.get('sistema_operativo'):
        partes.append(str(datos_equipo['sistema_operativo']).strip())
    if datos_equipo.get('incluye_cargador'):
        partes.append('Con cargador')
    return ' | '.join(partes)[:255]


def _crear_o_actualizar_linea_reacondicionado(solicitud, datos_equipo, costeo, costo_proveedor):
    """
    Crea o actualiza la LineaCotizacion P0125 al enviar propuesta de equipo reacondicionado.

    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta línea permite a Front aprobar/rechazar la oferta de equipo igual que las piezas
    de reparación. Al aprobar y generar compras, el equipo va a PiezaVentaMostrador en ST.

    Args:
        solicitud: SolicitudCotizacion vinculada.
        datos_equipo: Especificaciones capturadas en el modal.
        costeo: Resultado de calcular_costeo().
        costo_proveedor: float, costo de adquisición sin IVA.

    Returns:
        tuple: (ok: bool, error: str|None)
    """
    from decimal import Decimal
    from .models import LineaCotizacion, ProductoAlmacen

    try:
        producto = ProductoAlmacen.objects.get(codigo_producto=CODIGO_PRODUCTO_REACONDICIONADO)
    except ProductoAlmacen.DoesNotExist:
        logger.error(
            f'[REAC] Producto {CODIGO_PRODUCTO_REACONDICIONADO} no existe en ProductoAlmacen. '
            f'Solicitud {solicitud.numero_solicitud}'
        )
        return False, (
            f'El producto {CODIGO_PRODUCTO_REACONDICIONADO} (equipo reacondicionado) '
            'no está en el catálogo de almacén. Contacte al administrador.'
        )

    subtotal_sin_iva = Decimal(str(costeo.get('subtotal_sin_iva', 0)))
    descripcion = _construir_descripcion_linea_reac(datos_equipo)

    notas_partes = []
    especificaciones = (datos_equipo.get('especificaciones') or '').strip()
    if especificaciones:
        notas_partes.append(especificaciones)
    total_contado = costeo.get('total_precio_contado_mxn')
    if total_contado is not None:
        notas_partes.append(f'Precio contado (IVA incl.): ${total_contado}')

    defaults = {
        'descripcion_pieza': descripcion,
        'cantidad': 1,
        'costo_unitario': Decimal(str(costo_proveedor)),
        'precio_unitario_cliente': subtotal_sin_iva,
        'subtotal_cliente_sin_iva': subtotal_sin_iva,
        'es_linea_reacondicionado': True,
        'es_necesaria': False,
        'estado_cliente': 'pendiente',
        'opcion_pago_reac': '',
        'notas': '\n'.join(notas_partes),
    }

    linea = solicitud.lineas.filter(
        producto=producto,
        es_linea_reacondicionado=True,
    ).first()

    if linea:
        for campo, valor in defaults.items():
            setattr(linea, campo, valor)
        linea.save()
        logger.info(
            f'[REAC] Línea reacondicionado actualizada en solicitud {solicitud.numero_solicitud}'
        )
    else:
        LineaCotizacion.objects.create(
            solicitud=solicitud,
            producto=producto,
            **defaults,
        )
        logger.info(
            f'[REAC] Línea reacondicionado creada en solicitud {solicitud.numero_solicitud}'
        )

    return True, None


def _opciones_servicios_adicionales():
    """
    Construye la lista de servicios adicionales para el dropdown del modal.

    EXPLICACIÓN PARA PRINCIPIANTES:
    Los nombres vienen de TIPO_SERVICIO_ADICIONAL_CHOICES y los precios de
    PRECIOS_SERVICIOS_ADICIONALES (constants.py). Así el template no repite
    valores hardcodeados que pueden quedar desactualizados.

    Returns:
        list[dict]: Opciones con codigo, nombre y precio (IVA incluido).
    """
    from config.constants import (
        TIPO_SERVICIO_ADICIONAL_CHOICES,
        PRECIOS_SERVICIOS_ADICIONALES,
    )

    return [
        {
            'codigo': codigo,
            'nombre': nombre,
            'precio': PRECIOS_SERVICIOS_ADICIONALES.get(codigo, 0),
        }
        for codigo, nombre in TIPO_SERVICIO_ADICIONAL_CHOICES
    ]


@login_required
@permission_required_with_message('almacen.view_solicitudcotizacion')
def detalle_solicitud_cotizacion(request, pk):
    """
    Ver detalle completo de una solicitud de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra toda la información de una solicitud:
    - Datos de la cabecera (número, orden vinculada, estado)
    - Tabla con todas las líneas y sus estados
    - Imágenes de referencia de cada línea
    - Totales y resúmenes
    - Acciones disponibles según el estado
    
    Las acciones cambian según el estado:
    - Borrador: Editar, Enviar a cliente, Cancelar, Subir imágenes
    - Enviada: Registrar respuestas del cliente
    - Aprobada: Generar compras
    - Completada: Solo visualización
    
    También permite subir imágenes a las líneas cuando está en borrador.
    """
    solicitud = get_object_or_404(
        SolicitudCotizacion.objects.select_related(
            'orden_servicio',
            'orden_servicio__sucursal',        # Sucursal de la orden vinculada
            'creado_por',
            'creado_por__empleado',            # Perfil de empleado del creador
            'creado_por__empleado__sucursal',  # Sucursal del creador (para sin_orden_activa)
        ).prefetch_related(
            'lineas__producto',
            'lineas__proveedor',
            'lineas__compra_generada',
            'lineas__imagenes',  # Incluir imágenes de cada línea
            'imagenes_referencia',  # Incluir imágenes de referencia de la solicitud
        ),
        pk=pk
    )
    
    # Información del equipo desde DetalleEquipo si está vinculada la orden
    info_orden = None
    if solicitud.orden_servicio:
        try:
            info_orden = solicitud.orden_servicio.detalle_equipo
        except Exception:
            pass

    # --- Datos extra para el modal de envío de cotización al cliente ---
    # Obtenemos gama, mano de obra y email del cliente para pre-llenar el modal

    # Gama del equipo (alta/media/baja) — solo disponible si hay orden vinculada
    gama_equipo = ''
    if info_orden:
        gama_equipo = getattr(info_orden, 'gama', '') or ''

    # Costo de mano de obra desde la Cotizacion de Servicio Técnico
    # (el usuario puede sobreescribirlo en el modal si lo desea)
    costo_mano_obra = None
    if solicitud.orden_servicio:
        try:
            # La Cotizacion está vinculada 1:1 a la OrdenServicio
            cotizacion_st = solicitud.orden_servicio.cotizacion
            costo_mano_obra = float(cotizacion_st.costo_mano_obra)
        except Exception:
            # La orden puede no tener cotización aún — es normal
            pass

    # Email del cliente para el campo "Destinatario" del modal
    email_cliente_modal = ''
    if info_orden:
        email_raw = getattr(info_orden, 'email_cliente', '') or ''
        # Excluir el email placeholder que usa ST cuando no hay email real
        if email_raw and email_raw != 'cliente@ejemplo.com':
            email_cliente_modal = email_raw
    elif solicitud.email_cliente:
        email_cliente_modal = solicitud.email_cliente

    # Asunto sugerido para el modal: prefijo + orden cliente o service tag
    from .utils.cotizacion_email_context import construir_asunto_correo_default
    asunto_correo_modal = construir_asunto_correo_default(solicitud, info_orden=info_orden)
    
    # Procesar subida de imagen (solo en estado borrador)
    mensaje_imagen = None
    if request.method == 'POST' and solicitud.estado == 'borrador':
        linea_pk = request.POST.get('linea_pk')
        if linea_pk and 'imagen' in request.FILES:
            try:
                linea = solicitud.lineas.get(pk=linea_pk)
                form = ImagenLineaCotizacionForm(
                    request.POST,
                    request.FILES,
                    linea=linea
                )
                if form.is_valid():
                    imagen = form.save(commit=False)
                    imagen.linea = linea
                    imagen.subido_por = request.user
                    imagen.save()
                    messages.success(
                        request,
                        f'Imagen subida exitosamente a la línea #{linea.numero_linea}.'
                    )
                else:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f'{error}')
            except LineaCotizacion.DoesNotExist:
                messages.error(request, 'Línea no encontrada.')
            except ValueError as e:
                messages.error(request, str(e))
            
            # Redirigir para evitar reenvío del formulario
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
    
    # Verificar si se puede subir imágenes
    puede_subir_imagenes = solicitud.estado == 'borrador'
    
    # Imágenes de referencia de la solicitud
    imagenes_referencia = solicitud.imagenes_referencia.all()
    max_imagenes_referencia = ImagenSolicitudCotizacion.MAX_IMAGENES_POR_SOLICITUD
    
    # Empleados disponibles para copia en notificación a front
    empleados_copia = Empleado.objects.filter(
        Q(area='CALIDAD') | Q(area='FRONTDESK') | Q(area='CARRY IN'),
        activo=True,
        email__isnull=False
    ).exclude(
        email=''
    ).order_by('area', 'nombre_completo')
    
    # Verificar si el usuario actual está en la lista de CC
    usuario_en_lista_cc = False
    if hasattr(request.user, 'empleado') and request.user.empleado:
        usuario_en_lista_cc = empleados_copia.filter(id=request.user.empleado.id).exists()

    # --- Sucursal para mostrar en el encabezado de la solicitud ---
    # Con orden: se obtiene directamente de la FK orden_servicio → sucursal
    sucursal_orden = None
    if solicitud.orden_servicio:
        sucursal_orden = solicitud.orden_servicio.sucursal

    # Sin orden activa: se obtiene del perfil de empleado de quien creó la solicitud.
    # El select_related ya cargó el camino creado_por → empleado → sucursal,
    # por lo que esto no genera queries adicionales.
    sucursal_creador = None
    try:
        sucursal_creador = solicitud.creado_por.empleado.sucursal
    except Exception:
        # El usuario puede no tener perfil de empleado o sucursal asignada — es válido
        pass

    context = {
        'solicitud': solicitud,
        'info_orden': info_orden,
        'titulo': f'Solicitud {solicitud.numero_solicitud}',
        'puede_subir_imagenes': puede_subir_imagenes,
        'max_imagenes_por_linea': ImagenLineaCotizacion.MAX_IMAGENES_POR_LINEA,
        'imagenes_referencia': imagenes_referencia,
        'max_imagenes_referencia': max_imagenes_referencia,
        'empleados_copia': empleados_copia,
        'usuario_en_lista_cc': usuario_en_lista_cc,
        # Sucursal del encabezado: con orden → de la orden; sin orden → del creador
        'sucursal_orden': sucursal_orden,
        'sucursal_creador': sucursal_creador,
        # Datos para el modal "Enviar Cotización al Cliente"
        'gama_equipo': gama_equipo,
        'costo_mano_obra': costo_mano_obra,
        'email_cliente_modal': email_cliente_modal,
        'asunto_correo_modal': asunto_correo_modal,
        # Configuración de profit serializada como JSON para inyectarla en el
        # template y leerla desde TypeScript. Los valores vienen del .env
        # (nunca del código fuente), así que no aparecen en el repositorio.
        'profit_config_json': _serializar_profit_config(),
        'costeo_reac_config_json': _serializar_costeo_reacondicionado_config(),
        # Opciones del dropdown "Agregar Servicio Adicional" (precios desde constants.py)
        'servicios_adicionales_opciones': _opciones_servicios_adicionales(),
    }

    from .utils.cotizacion_items_cliente import (
        solicitud_puede_descargar_pdf_final,
        solicitud_tiene_items_cotizables,
    )
    context['puede_descargar_pdf_final'] = solicitud_puede_descargar_pdf_final(solicitud)
    context['tiene_items_cotizables'] = solicitud_tiene_items_cotizables(solicitud)

    return render(request, 'almacen/cotizaciones/detalle_solicitud.html', context)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def enviar_solicitud_cliente(request, pk):
    """
    Cambiar estado de la solicitud a 'enviada_front'.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta acción marca la solicitud como lista para que Recepción la revise.
    Recepción podrá entonces enviarla al cliente y registrar las respuestas.
    
    Requisitos:
    - Estado debe ser 'borrador'
    - Debe tener al menos una línea
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        if solicitud.enviar_a_front():
            messages.success(
                request,
                f'Solicitud {solicitud.numero_solicitud} enviada a Front. '
                'Recepción puede ahora revisarla y compartirla con el cliente.'
            )
        else:
            messages.error(
                request,
                'No se puede enviar la solicitud. Verifica que esté en estado '
                'borrador y tenga al menos una línea.'
            )
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def enviar_solicitud_a_cliente(request, pk):
    """
    Cambiar estado de la solicitud de 'enviada_front' a 'enviada_cliente'.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta acción la realiza Recepción (Front) cuando ya compartió la cotización
    con el cliente final. A partir de este momento:
    - El cliente puede aprobar o rechazar cada línea
    - Ya no se pueden editar líneas ni reenviar notificaciones
    - Aparecen los botones de aprobar/rechazar en el detalle
    
    Requisitos:
    - Estado debe ser 'enviada_front'
    - Debe tener al menos una línea
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        if solicitud.enviar_a_cliente():
            messages.success(
                request,
                f'Solicitud {solicitud.numero_solicitud} enviada al cliente. '
                'Ahora se pueden registrar las respuestas de aprobación/rechazo.'
            )
        else:
            messages.error(
                request,
                'No se puede enviar al cliente. Verifica que la solicitud esté '
                'en estado "Enviada a Front" y tenga al menos una línea.'
            )
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


# =============================================================================
# NUEVAS VISTAS: ENVÍO DE COTIZACIÓN DIRECTAMENTE AL CLIENTE FINAL
# =============================================================================

@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
@require_http_methods(["POST"])
def api_enviar_cotizacion_cliente(request, pk):
    """
    Envía la cotización directamente al cliente final por correo con PDF adjunto.

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista es el corazón del nuevo modal "Enviar Cotización al Cliente".
    Recibe los parámetros del modal (tipo de servicio, modo de agrupación,
    email del cliente, etc.), genera los PDF necesarios y los envía al cliente.

    Flujo:
    1. Valida los datos del POST (tipo_servicio, email_cliente, modo_agrupacion)
    2. Cambia el estado de la solicitud a 'enviada_cliente'
    3. Agrupa los ítems según el modo elegido (todo junto / piezas vs servicios / etc.)
    4. Para cada grupo, dispara una tarea Celery que genera el PDF y lo envía por email
    5. Si hay orden de ST vinculada, cambia su estado a 'cotizacion'
       («Esperando Aprobación Cliente») — este es el momento real de espera al cliente
    6. Retorna JsonResponse inmediato (el email se procesa en background)

    Args:
        request: HttpRequest POST con los datos del modal.
        pk     : ID de la SolicitudCotizacion.

    Returns:
        JsonResponse con {'success': bool, 'mensaje': str}
    """
    import json as _json
    from .tasks import enviar_cotizacion_cliente_task
    from config.paises_config import get_pais_actual

    try:
        solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)

        # --- 1. VALIDACIÓN DE ESTADO ---
        # La solicitud debe estar en un estado que permita el envío al cliente
        estados_validos = ['enviada_front', 'enviada_cliente', 'parcialmente_aprobada']
        if solicitud.estado not in estados_validos:
            return JsonResponse({
                'success': False,
                'error': f'Estado "{solicitud.get_estado_display()}" no permite envío al cliente. '
                         f'La solicitud debe estar en estado "Enviada a Front", "Enviada al Cliente" '
                         f'o "Parcialmente Aprobada".'
            })

        # --- 2. EXTRAER PARÁMETROS DEL POST ---
        modo_cotizacion = request.POST.get('modo_cotizacion', 'reparacion')
        tipo_servicio  = request.POST.get('tipo_servicio', 'estandar')
        email_cliente  = request.POST.get('email_cliente', '').strip()
        modo_agrupacion = request.POST.get('modo_agrupacion', 'todo_junto')
        mensaje_personalizado = request.POST.get('mensaje_personalizado', '').strip()
        incluir_descuento = request.POST.get('incluir_descuento_diagnostico') == '1'

        # Asunto personalizado del correo.
        # Si el usuario lo dejó vacío o con el prefijo por defecto sin texto adicional,
        # la tarea lo generará automáticamente con el perfil y folio.
        asunto_correo = request.POST.get('asunto_correo', '').strip()
        # Normalizar: si solo enviaron el prefijo vacío, tratar como sin asunto personalizado
        from .utils.cotizacion_email_context import es_asunto_correo_vacio
        if es_asunto_correo_vacio(asunto_correo):
            asunto_correo = ''

        # Mano de obra override — el campo es informativo, se envía como 0 desde el frontend.
        # Se mantiene la lectura por compatibilidad con llamadas directas a la API.
        mano_obra_raw = request.POST.get('mano_de_obra_override', '')
        mano_de_obra_override = None
        if mano_obra_raw:
            try:
                val = float(mano_obra_raw)
                mano_de_obra_override = val if val > 0 else None
            except ValueError:
                pass

        # Emails con copia (CC) — pueden venir múltiples campos con el mismo nombre
        copia_empleados = request.POST.getlist('copia_empleados')

        # --- 3. VALIDAR EMAIL DEL CLIENTE ---
        if not email_cliente:
            return JsonResponse({'success': False, 'error': 'El email del cliente es requerido.'})

        # Validación simple de formato email
        import re as _re
        if not _re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_cliente):
            return JsonResponse({'success': False, 'error': f'El email "{email_cliente}" no tiene formato válido.'})

        # --- 4a. MODO REACONDICIONADO (propuesta de equipo certificado) ---
        if modo_cotizacion == 'reacondicionado':
            datos_post = _extraer_datos_reacondicionado_post(request.POST)
            ok, resultado = _validar_y_calcular_reacondicionado(datos_post)
            if not ok:
                return JsonResponse({'success': False, 'error': resultado})

            if solicitud.estado == 'enviada_front':
                solicitud.enviar_a_cliente()

            _guardar_snapshot_reacondicionado(
                solicitud,
                resultado['datos_equipo'],
                resultado['costeo'],
                resultado['dias_front_desk'],
                resultado['costo_proveedor'],
            )

            ok_linea, error_linea = _crear_o_actualizar_linea_reacondicionado(
                solicitud,
                resultado['datos_equipo'],
                resultado['costeo'],
                resultado['costo_proveedor'],
            )
            if not ok_linea:
                return JsonResponse({'success': False, 'error': error_linea})

            _db = get_pais_actual()['db_alias']
            enviar_cotizacion_cliente_task.delay(
                solicitud_id=solicitud.pk,
                email_cliente=email_cliente,
                copia_empleados=copia_empleados,
                tipo_servicio='reacondicionado',
                items=[],
                titulo_propuesta='Propuesta de Equipo Reacondicionado — Certificado SIC',
                incluir_descuento_diagnostico=False,
                mano_de_obra_override=None,
                mensaje_personalizado=mensaje_personalizado,
                asunto_correo=asunto_correo,
                usuario_id=request.user.pk,
                db_alias=_db,
                modo_cotizacion='reacondicionado',
                datos_equipo_reac=resultado['datos_equipo'],
                costeo_reac=resultado['costeo'],
            )
            # Al enviar al cliente: ST pasa a «Esperando Aprobación Cliente»
            _actualizar_estado_st_esperando_aprobacion_cliente(solicitud, usuario=request.user)
            return JsonResponse({
                'success': True,
                'mensaje': (
                    f'Propuesta de equipo reacondicionado enviada a {email_cliente}. '
                    'El correo se procesa en segundo plano.'
                ),
                'grupos_enviados': 1,
            })

        # --- 4. VALIDAR TIPO DE SERVICIO (modo reparación) ---
        from .utils.pdf_cotizacion_cliente import PROFIT_CONFIG
        tipos_validos = list(PROFIT_CONFIG.keys())
        if tipo_servicio not in tipos_validos:
            return JsonResponse({'success': False, 'error': f'Tipo de servicio "{tipo_servicio}" no válido.'})

        # --- 5. CAMBIAR ESTADO A 'enviada_cliente' (si aún no lo está) ---
        if solicitud.estado == 'enviada_front':
            # Avanzar el estado usando el método del modelo
            solicitud.enviar_a_cliente()

        # Snapshot del perfil de profit (los precios se calculan al aprobar líneas)
        if not solicitud.fecha_precios_cliente:
            solicitud.tipo_servicio_cliente = tipo_servicio
            solicitud.incluir_descuento_diagnostico_cliente = incluir_descuento
            solicitud.save(update_fields=[
                'tipo_servicio_cliente',
                'incluir_descuento_diagnostico_cliente',
            ])

        # --- 6. CONSTRUIR GRUPOS DE ÍTEMS (solo pendiente + aprobada; excluye rechazadas) ---
        from .utils.cotizacion_items_cliente import (
            construir_grupos_cotizacion,
            obtener_lineas_cotizables,
            obtener_servicios_cotizables,
            serializar_linea_cotizacion,
            serializar_servicio_cotizacion,
            solicitud_tiene_items_cotizables,
        )

        if not solicitud_tiene_items_cotizables(solicitud):
            return JsonResponse({
                'success': False,
                'error': (
                    'No hay piezas ni servicios pendientes o aprobados para cotizar. '
                    'Las líneas rechazadas no se incluyen.'
                ),
            })

        items_piezas_todos = [
            serializar_linea_cotizacion(l) for l in obtener_lineas_cotizables(solicitud)
        ]
        items_servicios = [
            serializar_servicio_cotizacion(s) for s in obtener_servicios_cotizables(solicitud)
        ]
        grupos = construir_grupos_cotizacion(
            items_piezas_todos, items_servicios, modo_agrupacion
        )

        if not grupos:
            return JsonResponse({
                'success': False,
                'error': (
                    'No hay piezas ni servicios pendientes o aprobados para cotizar. '
                    'Las líneas rechazadas no se incluyen.'
                ),
            })

        # --- 8. DISPARAR TAREA CELERY PARA CADA GRUPO ---
        _db = get_pais_actual()['db_alias']
        usuario_id = request.user.pk

        for grupo in grupos:
            enviar_cotizacion_cliente_task.delay(
                solicitud_id=solicitud.pk,
                email_cliente=email_cliente,
                copia_empleados=copia_empleados,
                tipo_servicio=tipo_servicio,
                items=grupo['items'],
                titulo_propuesta=grupo['titulo'],
                incluir_descuento_diagnostico=incluir_descuento,
                mano_de_obra_override=mano_de_obra_override,
                mensaje_personalizado=mensaje_personalizado,
                asunto_correo=asunto_correo,
                usuario_id=usuario_id,
                db_alias=_db,
            )

        # Al enviar al cliente: orden ST → «Esperando Aprobación Cliente»
        # (si hay orden vinculada y aún no estaba en ese estado)
        _actualizar_estado_st_esperando_aprobacion_cliente(solicitud, usuario=request.user)

        # Mensaje de éxito según cuántos grupos se enviaron
        n_grupos = len(grupos)
        if n_grupos == 1:
            mensaje = f'Cotización enviada al cliente {email_cliente}. El correo se procesa en segundo plano.'
        else:
            mensaje = (
                f'{n_grupos} cotizaciones separadas enviadas a {email_cliente}. '
                f'Los correos se procesan en segundo plano.'
            )

        return JsonResponse({'success': True, 'mensaje': mensaje, 'grupos_enviados': n_grupos})

    except Exception as e:
        import traceback as _tb
        logger.error(f"[API_COTIZACION_CLIENTE] Error: {e}\n{_tb.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})


@login_required
@permission_required_with_message('almacen.view_solicitudcotizacion')
@require_http_methods(["GET"])
@xframe_options_sameorigin
def preview_pdf_cotizacion(request, pk):
    """
    Genera y devuelve un PDF de previsualización para el modal de envío al cliente.

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista es la que llama el botón "Ver Previsualización" del modal.
    Genera el PDF con los parámetros actuales del modal (tipo de servicio,
    modo de agrupación) y lo devuelve directamente como respuesta PDF
    para mostrarlo en el iframe del modal.

    Args:
        request: HttpRequest GET con parámetros:
                 - tipo_servicio: str
                 - modo_agrupacion: str
                 - incluir_descuento: '1'/'0'
                 - grupo_idx: int (0=primero, 1=segundo — para modo separado)
        pk     : ID de la SolicitudCotizacion.

    Returns:
        HttpResponse con content_type='application/pdf'
    """
    from django.http import HttpResponse
    from .utils.pdf_cotizacion_cliente import PDFCotizacionCliente
    from config.paises_config import get_pais_actual

    try:
        solicitud = get_object_or_404(
            SolicitudCotizacion.objects.select_related(
                'orden_servicio',
                'orden_servicio__detalle_equipo',
                'orden_servicio__sucursal',
                'creado_por',
                'creado_por__empleado__sucursal',
            ).prefetch_related('lineas__producto', 'servicios_adicionales'),
            pk=pk
        )

        tipo_servicio     = request.GET.get('tipo_servicio', 'estandar')
        modo_cotizacion   = request.GET.get('modo_cotizacion', 'reparacion')
        modo_agrupacion   = request.GET.get('modo_agrupacion', 'todo_junto')
        incluir_descuento = request.GET.get('incluir_descuento_diagnostico', '0') == '1'
        grupo_idx         = int(request.GET.get('grupo_idx', 0))

        _pais = get_pais_actual()

        # Preview de propuesta reacondicionada (motor distinto al de reparación)
        if modo_cotizacion == 'reacondicionado':
            from .utils.pdf_cotizacion_reacondicionado import PDFCotizacionReacondicionado

            datos_post = {
                'costo_proveedor': request.GET.get('reac_costo_proveedor', ''),
                'dias_front_desk': request.GET.get('reac_dias_front_desk', '1'),
                'marca': request.GET.get('reac_marca', ''),
                'modelo': request.GET.get('reac_modelo', ''),
                'procesador': request.GET.get('reac_procesador', ''),
                'ram': request.GET.get('reac_ram', ''),
                'sistema_operativo': request.GET.get('reac_sistema_operativo', ''),
                'incluye_cargador': request.GET.get('reac_incluye_cargador', '0'),
                'especificaciones': request.GET.get('reac_especificaciones', ''),
            }
            datos_post['incluye_cargador'] = datos_post['incluye_cargador'] == '1'
            ok, resultado = _validar_y_calcular_reacondicionado(datos_post)
            if not ok:
                return HttpResponse(resultado.encode(), content_type='text/plain', status=400)

            generador_reac = PDFCotizacionReacondicionado(
                solicitud=solicitud,
                datos_equipo=resultado['datos_equipo'],
                costeo=resultado['costeo'],
                pais_config=_pais,
            )
            resultado_pdf = generador_reac.generar_pdf()
            if not resultado_pdf['success']:
                return HttpResponse(
                    f'Error al generar PDF: {resultado_pdf.get("error", "desconocido")}'.encode(),
                    content_type='text/plain',
                    status=500,
                )
            pdf_bytes = resultado_pdf['buffer'].getvalue()
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = (
                f'inline; filename="{resultado_pdf["nombre_archivo"]}"'
            )
            return response

        from .utils.cotizacion_items_cliente import (
            construir_grupos_cotizacion,
            obtener_lineas_cotizables,
            obtener_servicios_cotizables,
            serializar_linea_cotizacion,
            serializar_servicio_cotizacion,
        )

        items_todos_piezas = [
            serializar_linea_cotizacion(l) for l in obtener_lineas_cotizables(solicitud)
        ]
        items_servicios = [
            serializar_servicio_cotizacion(s) for s in obtener_servicios_cotizables(solicitud)
        ]
        grupos = construir_grupos_cotizacion(
            items_todos_piezas, items_servicios, modo_agrupacion
        )

        if not grupos:
            return HttpResponse(
                b'%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
                b'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n'
                b'3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj\n'
                b'xref\n0 4\n0000000000 65535 f \n0000000009 00000 n \n'
                b'0000000058 00000 n \n0000000115 00000 n \n'
                b'trailer<</Size 4/Root 1 0 R>>\nstartxref\n190\n%%EOF',
                content_type='application/pdf'
            )

        # Elegir el grupo según el índice (para modo separado)
        grupo_idx = min(grupo_idx, len(grupos) - 1)
        grupo = grupos[grupo_idx]

        generador = PDFCotizacionCliente(
            solicitud=solicitud,
            tipo_servicio=tipo_servicio,
            items=grupo['items'],
            titulo_propuesta=grupo['titulo'],
            incluir_descuento_diagnostico=incluir_descuento,
            pais_config=_pais,
        )

        resultado = generador.generar_pdf()
        if not resultado['success']:
            return HttpResponse(
                f'Error al generar PDF: {resultado.get("error", "desconocido")}'.encode(),
                content_type='text/plain',
                status=500
            )

        pdf_bytes = resultado['buffer'].getvalue()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="{resultado["nombre_archivo"]}"'
        )
        return response

    except Exception as e:
        import traceback as _tb
        logger.error(f"[PREVIEW_PDF_COTIZACION] Error: {e}\n{_tb.format_exc()}")
        return HttpResponse(f'Error: {str(e)}'.encode(), content_type='text/plain', status=500)


@login_required
@permission_required_with_message('almacen.view_solicitudcotizacion')
@require_http_methods(["GET"])
def descargar_pdf_cotizacion_final(request, pk):
    """
    Genera y descarga el PDF final con piezas/servicios aceptados y precios persistidos.

    Si solo se aceptó el equipo reacondicionado (línea P0125), genera el PDF de
    propuesta reacondicionada en lugar del PDF de piezas de reparación.

    Args:
        request: HttpRequest GET.
        pk     : ID de la SolicitudCotizacion.

    Returns:
        HttpResponse PDF inline o error en texto plano.
    """
    from django.http import HttpResponse
    from .utils.pdf_cotizacion_cliente import PDFCotizacionCliente
    from .utils.cotizacion_items_cliente import (
        construir_items_cotizacion_final,
        solicitud_puede_descargar_pdf_final,
        solicitud_pdf_final_es_solo_reacondicionado,
        extraer_datos_equipo_desde_solicitud,
    )
    from .utils.cotizacion_precios_cliente import obtener_tipo_servicio_solicitud
    from config.paises_config import get_pais_actual

    try:
        solicitud = get_object_or_404(
            SolicitudCotizacion.objects.select_related(
                'orden_servicio',
                'orden_servicio__detalle_equipo',
            ).prefetch_related('lineas__producto', 'servicios_adicionales'),
            pk=pk,
        )

        if not solicitud_puede_descargar_pdf_final(solicitud):
            return HttpResponse(
                'No hay ítems aceptados con precios guardados para generar el PDF final.'.encode(),
                content_type='text/plain',
                status=400,
            )

        _pais = get_pais_actual()

        # PDF final de equipo reacondicionado (solo línea P0125 aceptada)
        if solicitud_pdf_final_es_solo_reacondicionado(solicitud):
            from .utils.pdf_cotizacion_reacondicionado import PDFCotizacionReacondicionado
            from .utils.cotizacion_items_cliente import obtener_lineas_aceptadas_final

            linea_reac = obtener_lineas_aceptadas_final(solicitud).filter(
                es_linea_reacondicionado=True,
            ).first()

            generador_reac = PDFCotizacionReacondicionado(
                solicitud=solicitud,
                datos_equipo=extraer_datos_equipo_desde_solicitud(solicitud),
                costeo=solicitud.resultado_costeo_reac or {},
                pais_config=_pais,
                opcion_pago_aceptada=linea_reac.opcion_pago_reac if linea_reac else 'contado',
                modo_final=True,
            )
            resultado = generador_reac.generar_pdf()
            if not resultado['success']:
                return HttpResponse(
                    f'Error al generar PDF: {resultado.get("error", "desconocido")}'.encode(),
                    content_type='text/plain',
                    status=500,
                )
            pdf_bytes = resultado['buffer'].getvalue()
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = (
                f'inline; filename="{resultado["nombre_archivo"]}"'
            )
            return response

        items = construir_items_cotizacion_final(solicitud)
        if not items:
            return HttpResponse(
                'No se encontraron líneas aceptadas con precio al cliente.'.encode(),
                content_type='text/plain',
                status=400,
            )

        tipo_servicio = obtener_tipo_servicio_solicitud(solicitud)
        incluir_descuento = bool(
            getattr(solicitud, 'incluir_descuento_diagnostico_cliente', True)
        )

        generador = PDFCotizacionCliente(
            solicitud=solicitud,
            tipo_servicio=tipo_servicio,
            items=items,
            incluir_descuento_diagnostico=incluir_descuento,
            pais_config=_pais,
            modo_final=True,
        )

        resultado = generador.generar_pdf()
        if not resultado['success']:
            return HttpResponse(
                f'Error al generar PDF: {resultado.get("error", "desconocido")}'.encode(),
                content_type='text/plain',
                status=500,
            )

        pdf_bytes = resultado['buffer'].getvalue()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="{resultado["nombre_archivo"]}"'
        )
        return response

    except Exception as e:
        import traceback as _tb
        logger.error(f"[PDF_COTIZACION_FINAL] Error: {e}\n{_tb.format_exc()}")
        return HttpResponse(f'Error: {str(e)}'.encode(), content_type='text/plain', status=500)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
@require_http_methods(["POST"])
def notificar_front(request, pk):
    """
    Enviar notificación de cotización a recepción (FRONTDESK) por correo.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista reemplaza el antiguo "Enviar a Cliente". En lugar de cambiar
    solo el estado, ahora envía un correo HTML a los empleados de FRONTDESK
    con el detalle de la cotización (piezas, costos, imágenes) para que
    recepción la comparta con el cliente.
    
    Flujo:
    1. Valida que la solicitud esté en estado 'borrador' y tenga líneas
    2. Cambia el estado a 'enviada_front'
    3. Dispara la tarea Celery para enviar el correo en segundo plano
    4. Devuelve JsonResponse inmediato
    
    Args:
        request: HttpRequest con datos POST del formulario
        pk: ID de la SolicitudCotizacion
    
    Returns:
        JsonResponse — el correo se procesa en background via Celery
    """
    from .tasks import notificar_front_cotizacion_task
    
    try:
        solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
        
        # Validar estado: permitir envío inicial (borrador) o reenvío (enviada_front)
        if solicitud.estado not in ['borrador', 'enviada_front']:
            return JsonResponse({
                'success': False,
                'error': 'Solo se puede notificar cuando la solicitud está en borrador o ya enviada a front.'
            }, status=400)
        
        # Validar que tenga al menos una línea
        if not solicitud.lineas.exists():
            return JsonResponse({
                'success': False,
                'error': 'La solicitud debe tener al menos una línea para notificar.'
            }, status=400)
        
        # Obtener destinatarios del formulario (los seleccionados en el modal)
        copia_empleados = request.POST.getlist('copia_empleados', [])
        copia_tecnico = request.POST.getlist('copia_tecnico', [])
        
        # Los destinatarios principales son los que el usuario seleccionó
        destinatarios = list(set(copia_empleados + copia_tecnico))
        
        if not destinatarios:
            return JsonResponse({
                'success': False,
                'error': 'Debes seleccionar al menos un destinatario.'
            }, status=400)
        
        mensaje_personalizado = request.POST.get('mensaje_personalizado', '').strip()
        
        # Cambiar estado de la solicitud a 'enviada_front' solo si está en borrador
        if solicitud.estado == 'borrador':
            solicitud.enviar_a_front(usuario=request.user)
        
        # Disparar tarea Celery
        usuario_id = request.user.pk if request.user.is_authenticated else None
        from config.paises_config import get_pais_actual

        tarea = notificar_front_cotizacion_task.delay(
            solicitud_id=pk,
            destinatarios=destinatarios,
            mensaje_personalizado=mensaje_personalizado,
            usuario_id=usuario_id,
            db_alias=get_pais_actual()['db_alias'],
        )
        
        return JsonResponse({
            'success': True,
            'message': (
                f'Notificación en proceso de envío a {len(destinatarios)} '
                f'destinatario(s).'
            ),
            'data': {
                'task_id': tarea.id,
                'destinatario': ', '.join(destinatarios),
                'solicitud': solicitud.numero_solicitud,
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar la solicitud: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def responder_linea_cotizacion(request, solicitud_pk, linea_pk):
    """
    Registrar la respuesta del cliente para una línea específica.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista permite a Recepción registrar si el cliente aprobó o
    rechazó una línea específica de la cotización.
    
    Si el cliente aprueba: La línea queda marcada para generar compra.
    Si el cliente rechaza: Se debe indicar el motivo.
    
    Después de cada respuesta, se actualiza automáticamente el estado
    general de la solicitud.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=solicitud_pk)
    linea = get_object_or_404(LineaCotizacion, pk=linea_pk, solicitud=solicitud)
    
    # Solo se puede responder si la solicitud está enviada
    if solicitud.estado not in ['enviada_cliente', 'parcialmente_aprobada']:
        messages.error(
            request,
            'Solo se pueden registrar respuestas en solicitudes enviadas al cliente.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)
    
    if request.method == 'POST':
        form = RespuestaLineaCotizacionForm(request.POST)
        
        if form.is_valid():
            decision = form.cleaned_data['decision']
            motivo = form.cleaned_data.get('motivo_rechazo', '')
            
            if decision == 'aprobar':
                if linea.es_linea_reacondicionado:
                    opcion = form.cleaned_data.get('opcion_pago_reac', '')
                    if not opcion:
                        messages.error(
                            request,
                            'Debes seleccionar la forma de pago del equipo reacondicionado.',
                        )
                        return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)
                    aprobado = linea.aprobar(opcion_pago_reac=opcion)
                else:
                    aprobado = linea.aprobar()
                if aprobado:
                    messages.success(
                        request,
                        f'Línea #{linea.numero_linea} aprobada por el cliente.'
                    )
                else:
                    if linea.es_linea_reacondicionado:
                        messages.error(
                            request,
                            'No se pudo aprobar el equipo reacondicionado. '
                            'Verifica el costeo guardado y la forma de pago.',
                        )
                    else:
                        messages.error(request, 'No se pudo aprobar la línea.')
            else:  # rechazar
                if linea.rechazar(motivo=motivo):
                    messages.warning(
                        request,
                        f'Línea #{linea.numero_linea} rechazada por el cliente.'
                    )
                else:
                    messages.error(request, 'No se pudo rechazar la línea.')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)


@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def aprobar_todas_lineas(request, pk):
    """
    Aprobar todas las líneas pendientes de una solicitud.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Atajo para cuando el cliente aprueba toda la cotización.
    En lugar de aprobar línea por línea, se aprueban todas a la vez.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        lineas_pendientes = solicitud.lineas.filter(
            estado_cliente='pendiente',
            es_linea_reacondicionado=False,
        )
        aprobadas = 0
        
        for linea in lineas_pendientes:
            if linea.aprobar():
                aprobadas += 1
        
        if aprobadas > 0:
            messages.success(
                request,
                f'Se aprobaron {aprobadas} línea(s) de la cotización.'
            )
        reac_pendientes = solicitud.lineas.filter(
            estado_cliente='pendiente',
            es_linea_reacondicionado=True,
        ).count()
        if reac_pendientes > 0:
            messages.info(
                request,
                f'Quedan {reac_pendientes} equipo(s) reacondicionado(s) pendiente(s). '
                'Apruébalos uno por uno para elegir la forma de pago.',
            )
        elif aprobadas == 0:
            messages.info(request, 'No había líneas pendientes por aprobar.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def rechazar_todas_lineas(request, pk):
    """
    Rechazar todas las líneas pendientes de una solicitud.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Atajo para cuando el cliente rechaza toda la cotización.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', 'Rechazado por el cliente')
        lineas_pendientes = solicitud.lineas.filter(estado_cliente='pendiente')
        rechazadas = 0
        
        for linea in lineas_pendientes:
            if linea.rechazar(motivo=motivo):
                rechazadas += 1
        
        if rechazadas > 0:
            messages.warning(
                request,
                f'Se rechazaron {rechazadas} línea(s) de la cotización.'
            )
        else:
            messages.info(request, 'No había líneas pendientes por rechazar.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


# ============================================================================
# VISTAS PARA SERVICIOS ADICIONALES (Venta Mostrador en Cotizaciones)
# ============================================================================

@login_required
@permission_required_with_message('almacen.add_lineaservicioadicional')
def agregar_servicio_adicional(request, solicitud_pk):
    """
    Agregar un servicio adicional a una solicitud de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista permite agregar servicios de Venta Mostrador (limpieza,
    reinstalación de SO, paquetes, etc.) a una cotización existente.
    
    Los servicios adicionales aparecen debajo de las líneas de cotización
    y el cliente puede aprobarlos/rechazarlos por separado.
    
    Cuando el cliente aprueba y se generan las compras, estos servicios
    se crean automáticamente en el VentaMostrador de la orden.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=solicitud_pk)
    
    # Solo se pueden agregar servicios en estados iniciales
    if solicitud.estado not in ['borrador', 'enviada_front', 'enviada_cliente']:
        messages.error(
            request,
            'No se pueden agregar servicios adicionales en este estado de la solicitud.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)
    
    if request.method == 'POST':
        form = LineaServicioAdicionalForm(request.POST)
        
        if form.is_valid():
            servicio = form.save(commit=False)
            servicio.solicitud = solicitud
            servicio.save()
            
            messages.success(
                request,
                f'Servicio "{servicio.get_tipo_servicio_display()}" agregado a la cotización.'
            )
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        # GET: redirigir al detalle
        pass
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)


@login_required
@permission_required_with_message('almacen.delete_lineaservicioadicional')
def eliminar_servicio_adicional(request, solicitud_pk, servicio_pk):
    """
    Eliminar un servicio adicional de una solicitud de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Elimina un servicio adicional que fue agregado por error o que el
    cliente ya no quiere. Solo se puede eliminar si aún no ha sido
    aprobado/rechazado por el cliente.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=solicitud_pk)
    servicio = get_object_or_404(
        LineaServicioAdicional,
        pk=servicio_pk,
        solicitud=solicitud
    )
    
    # Solo se pueden eliminar servicios pendientes
    if servicio.estado_cliente != 'pendiente':
        messages.error(
            request,
            'No se puede eliminar un servicio que ya fue respondido por el cliente.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)
    
    if request.method == 'POST':
        nombre_servicio = servicio.get_tipo_servicio_display()
        servicio.delete()
        messages.success(
            request,
            f'Servicio "{nombre_servicio}" eliminado de la cotización.'
        )
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)


@login_required
@permission_required_with_message('almacen.change_lineaservicioadicional')
def responder_servicio_adicional(request, solicitud_pk, servicio_pk):
    """
    Registrar la respuesta del cliente para un servicio adicional.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Similar a responder_linea_cotizacion, pero para servicios adicionales.
    Permite registrar si el cliente aprobó o rechazó el servicio.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=solicitud_pk)
    servicio = get_object_or_404(
        LineaServicioAdicional,
        pk=servicio_pk,
        solicitud=solicitud
    )
    
    # Solo se puede responder si la solicitud está enviada
    if solicitud.estado not in ['enviada_cliente', 'parcialmente_aprobada']:
        messages.error(
            request,
            'Solo se pueden registrar respuestas en solicitudes enviadas al cliente.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)
    
    if request.method == 'POST':
        decision = request.POST.get('decision', '')
        motivo = request.POST.get('motivo_rechazo', '')
        
        if decision == 'aprobar':
            if servicio.aprobar():
                messages.success(
                    request,
                    f'Servicio "{servicio.get_tipo_servicio_display()}" aprobado por el cliente.'
                )
            else:
                messages.error(request, 'No se pudo aprobar el servicio.')
        elif decision == 'rechazar':
            if servicio.rechazar(motivo=motivo):
                messages.warning(
                    request,
                    f'Servicio "{servicio.get_tipo_servicio_display()}" rechazado por el cliente.'
                )
            else:
                messages.error(request, 'No se pudo rechazar el servicio.')
        else:
            messages.error(request, 'Decisión no válida.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=solicitud_pk)


@login_required
@permission_required_with_message('almacen.change_lineaservicioadicional')
def aprobar_todos_servicios(request, pk):
    """
    Aprobar todos los servicios adicionales pendientes de una solicitud.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Atajo para cuando el cliente aprueba todos los servicios adicionales.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        servicios_pendientes = solicitud.servicios_adicionales.filter(estado_cliente='pendiente')
        aprobados = 0
        
        for servicio in servicios_pendientes:
            if servicio.aprobar():
                aprobados += 1
        
        if aprobados > 0:
            messages.success(
                request,
                f'Se aprobaron {aprobados} servicio(s) adicional(es).'
            )
        else:
            messages.info(request, 'No había servicios pendientes por aprobar.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.change_lineaservicioadicional')
def rechazar_todos_servicios(request, pk):
    """
    Rechazar todos los servicios adicionales pendientes de una solicitud.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', 'Rechazado por el cliente')
        servicios_pendientes = solicitud.servicios_adicionales.filter(estado_cliente='pendiente')
        rechazados = 0
        
        for servicio in servicios_pendientes:
            if servicio.rechazar(motivo=motivo):
                rechazados += 1
        
        if rechazados > 0:
            messages.warning(
                request,
                f'Se rechazaron {rechazados} servicio(s) adicional(es).'
            )
        else:
            messages.info(request, 'No había servicios pendientes por rechazar.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.add_compraproducto')
def generar_compras_solicitud(request, pk):
    """
    Genera CompraProducto para las líneas aprobadas y VentaMostrador para servicios adicionales.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Una vez que el cliente ha aprobado las líneas y/o servicios adicionales,
    esta acción crea:
    
    1. CompraProducto para cada línea de pieza aprobada
       - Queda vinculado a la línea de cotización
       - Tiene estado 'pendiente_llegada'
       - Hereda el producto, proveedor, cantidad y costo de la línea
       - Se vincula a la misma orden de servicio
    
    2. VentaMostrador (o actualiza si ya existe) para servicios adicionales aprobados
       - Mapea cada servicio a su campo correspondiente en VentaMostrador
       - Ejemplo: 'limpieza' → incluye_limpieza=True, costo_limpieza=$450
    
    3. En órdenes OOW de reparación (no FL-): crea SeguimientoPieza en ST
       agrupados por proveedor y pasa la orden a «Esperando Llegada de Piezas».
    
    Esto integra el flujo de cotizaciones con el flujo existente de compras
    y ventas mostrador.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        # Bloquear compras en modo sin orden activa hasta vincular orden de servicio
        if solicitud.compras_pendientes_sin_orden():
            messages.error(
                request,
                'Debes crear o vincular una orden de servicio antes de generar las compras.'
            )
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

        puede_generar_compras = solicitud.puede_generar_compras()
        puede_generar_venta = solicitud.puede_generar_venta_mostrador()
        
        # Validar que haya algo que generar
        if not puede_generar_compras and not puede_generar_venta:
            messages.error(
                request,
                'No hay líneas ni servicios aprobados pendientes de procesar.'
            )
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
        
        mensajes_exito = []

        # ====== Piezas en Venta Mostrador (FL- o equipos reacondicionados en OOW) ======
        # Se llama ANTES de generar_compras() porque generar_compras() pasa las líneas
        # a estado 'compra_generada' y este método filtra por estado='aprobada'.
        necesita_piezas_vm = (
            puede_generar_compras
            and solicitud.orden_servicio
            and (
                solicitud.orden_servicio.tipo_servicio == 'venta_mostrador'
                or solicitud.lineas.filter(
                    es_linea_reacondicionado=True,
                    estado_cliente='aprobada',
                ).exists()
            )
        )
        if necesita_piezas_vm:
            n_piezas = solicitud.generar_piezas_venta_mostrador()
            if n_piezas:
                mensajes_exito.append(
                    f'{n_piezas} pieza(s) registrada(s) en sección Venta Mostrador'
                )

        # Generar compras para piezas (CompraProducto para control de inventario/almacén)
        if puede_generar_compras:
            compras = solicitud.generar_compras(usuario=request.user)
            if compras:
                mensajes_exito.append(
                    f'{len(compras)} compra(s) de piezas generada(s)'
                )
            # Mensaje del sync ST (seguimiento de piezas + estado esperando_piezas)
            sync_st = getattr(solicitud, '_resultado_sync_seguimiento_st', None) or {}
            n_seg = sync_st.get('seguimientos_creados', 0)
            if n_seg:
                mensajes_exito.append(
                    f'{n_seg} seguimiento(s) de piezas registrado(s) en Servicio Técnico'
                )
            if sync_st.get('estado_actualizado'):
                mensajes_exito.append(
                    'orden ST actualizada a «Esperando Llegada de Piezas»'
                )
        
        # Generar VentaMostrador para servicios adicionales (paquetes, limpieza, etc.)
        if puede_generar_venta:
            venta = solicitud.generar_venta_mostrador()
            if venta:
                mensajes_exito.append(
                    f'Venta Mostrador creada/actualizada ({venta.folio_venta})'
                )
        
        # Mostrar mensajes al usuario
        if mensajes_exito:
            messages.success(
                request,
                'Se generaron exitosamente: ' + '. '.join(mensajes_exito) + '.'
            )
        else:
            messages.warning(request, 'No se pudieron generar las compras/servicios.')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def vincular_orden_solicitud(request, pk):
    """
    Vincular una solicitud de cotización (sin orden activa) a una orden de servicio.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando una cotización se crea sin orden activa (el equipo aún no ingresa),
    esta vista permite buscar y vincular la orden de servicio correspondiente
    cuando el equipo ya ingresó formalmente.
    
    BÚSQUEDA:
    - Por número de orden interno (ORD-2025-0001)
    - Por número de orden cliente (OOW-12345, FL-67890)
    - Por service tag / número de serie
    - Por nombre del cliente
    """
    from servicio_tecnico.models import OrdenServicio, DetalleEquipo
    
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if not solicitud.puede_vincular_orden():
        messages.error(
            request,
            'No se puede vincular una orden. La solicitud ya tiene orden activa '
            'o está completada/cancelada.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
    
    resultados = []
    termino_busqueda = ''
    
    if request.method == 'POST':
        termino_busqueda = request.POST.get('busqueda', '').strip()
        orden_pk = request.POST.get('orden_pk', '')
        
        # Si se seleccionó una orden específica, vincularla
        if orden_pk:
            orden = get_object_or_404(OrdenServicio, pk=orden_pk)
            
            try:
                solicitud.vincular_orden(orden)
                messages.success(
                    request,
                    f'Solicitud {solicitud.numero_solicitud} vinculada exitosamente '
                    f'a la orden {orden.numero_orden_interno}.'
                )
            except ValueError as e:
                messages.error(request, str(e))
            
            return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
        
        # Buscar órdenes que coincidan
        if termino_busqueda:
            from django.db.models import Q
            
            # Buscar por número de orden interno, orden cliente, o service tag
            resultados = OrdenServicio.objects.filter(
                Q(numero_orden_interno__icontains=termino_busqueda) |
                Q(detalle_equipo__orden_cliente__icontains=termino_busqueda) |
                Q(detalle_equipo__numero_serie__icontains=termino_busqueda) |
                Q(detalle_equipo__nombre_cliente__icontains=termino_busqueda)
            ).select_related('detalle_equipo').order_by('-fecha_ingreso')[:20]
            
            # Si no hay resultados y tenemos service_tag de la solicitud, buscar por eso
            if not resultados and solicitud.service_tag:
                resultados = OrdenServicio.objects.filter(
                    detalle_equipo__numero_serie__icontains=solicitud.service_tag
                ).select_related('detalle_equipo').order_by('-fecha_ingreso')[:20]
    
    # Si es GET, buscar automáticamente por service_tag si existe
    elif solicitud.service_tag:
        from servicio_tecnico.models import OrdenServicio
        resultados = OrdenServicio.objects.filter(
            detalle_equipo__numero_serie__icontains=solicitud.service_tag
        ).select_related('detalle_equipo').order_by('-fecha_ingreso')[:20]
        termino_busqueda = solicitud.service_tag
    
    return render(request, 'almacen/cotizaciones/vincular_orden.html', {
        'solicitud': solicitud,
        'resultados': resultados,
        'termino_busqueda': termino_busqueda,
    })


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def crear_orden_fl_desde_cotizacion(request, pk):
    """
    Crea una OrdenServicio tipo FL- (Venta Mostrador / Servicio Directo) directamente
    desde el detalle de una SolicitudCotizacion que no tiene orden vinculada.

    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cuando una cotización se creó en modo "sin orden activa" (el cliente solicitó
    precio antes de ingresar físicamente el equipo) y el cliente acepta la cotización,
    en lugar de obligar al usuario a ir a Servicio Técnico a crear la orden manualmente,
    esta vista crea la orden FL- (Servicio Directo sin Diagnóstico) aquí mismo,
    usando todos los datos del cliente y equipo que ya fueron capturados en la solicitud.

    Flujo:
    1. GET  → muestra modal con el formulario mínimo (técnico + número FL- sugerido)
    2. POST → crea OrdenServicio + DetalleEquipo con datos reales de la solicitud,
              llama a solicitud.vincular_orden() para sincronizar con ST,
              re-guarda las LineaCotizacion para que se sincronicen como PiezaCotizada,
              redirige al detalle con mensaje de éxito.

    Restricciones:
    - Solo aplica si solicitud.puede_vincular_orden() es True (sin_orden_activa + no completada/cancelada)
    - El tipo de servicio es siempre 'venta_mostrador' (Servicio Directo sin Diagnóstico)
    - El número FL- se auto-sugiere pero puede ser editado por el usuario

    Args:
        request: HttpRequest de Django
        pk     : PK de la SolicitudCotizacion

    Efectos secundarios:
        - Crea OrdenServicio y DetalleEquipo en servicio_tecnico
        - Modifica SolicitudCotizacion (vincula orden, desactiva sin_orden_activa)
        - Crea Cotizacion en servicio_tecnico (vía solicitud.vincular_orden → save)
        - Re-sincroniza LineaCotizacion → PiezaCotizada en servicio_tecnico
    """
    from servicio_tecnico.models import OrdenServicio, DetalleEquipo

    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)

    # Validar que la solicitud esté en modo sin_orden_activa y pueda vincularse
    if not solicitud.puede_vincular_orden():
        messages.error(
            request,
            'No se puede crear una orden. La solicitud ya tiene orden activa '
            'o está completada/cancelada.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    if request.method == 'GET':
        # ====== PREPARAR DATOS PARA EL MODAL ======

        # Obtener lista de técnicos de laboratorio activos para el selector
        tecnicos = Empleado.objects.filter(
            activo=True,
            cargo__icontains='TECNICO DE LABORATORIO'
        ).select_related('sucursal').order_by('nombre_completo')

        # Auto-generar sugerencia de número FL- con formato FL-YYYY-NNNN
        # Busca el último FL- de este año en DetalleEquipo para sugerir el siguiente
        año_actual = timezone.now().year
        ultimo_fl = DetalleEquipo.objects.filter(
            orden_cliente__startswith=f'FL-{año_actual}'
        ).order_by('-orden_cliente').first()

        if ultimo_fl:
            # Extraer el número secuencial del último FL- y sumar 1
            try:
                ultimo_num = int(ultimo_fl.orden_cliente.split('-')[-1])
                siguiente_num = ultimo_num + 1
            except (ValueError, IndexError):
                siguiente_num = 1
        else:
            # Si no hay ningún FL- este año, empezar desde 1
            siguiente_num = 1

        numero_fl_sugerido = f"FL-{año_actual}-{siguiente_num:04d}"

        return JsonResponse({
            'success': True,
            'tecnicos': [
                {
                    'id': t.pk,
                    'nombre': t.nombre_completo,
                    'sucursal': t.sucursal.nombre if t.sucursal else 'Sin asignar',
                }
                for t in tecnicos
            ],
            'numero_fl_sugerido': numero_fl_sugerido,
            # Datos del cliente/equipo para mostrar resumen en el modal
            'resumen': {
                'nombre_cliente': solicitud.nombre_cliente or '(sin nombre)',
                'email_cliente': solicitud.email_cliente or '(sin email)',
                'telefono_cliente': solicitud.telefono_cliente or '(sin teléfono)',
                'tipo_equipo': solicitud.tipo_equipo or 'Por definir',
                'marca': solicitud.marca or 'Por definir',
                'modelo': solicitud.modelo or 'Por definir',
                'numero_serie': solicitud.service_tag or '(sin service tag)',
            }
        })

    # ====== POST: Crear la orden FL- ======
    tecnico_id = request.POST.get('tecnico_id', '').strip()
    numero_fl = request.POST.get('numero_fl', '').strip().upper()

    # --- Validación de campos requeridos ---
    if not tecnico_id:
        messages.error(request, 'Debes seleccionar un técnico para crear la orden.')
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    if not numero_fl:
        messages.error(request, 'El número de folio FL- es obligatorio.')
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    # Validar que el número FL- tenga el formato correcto
    if not numero_fl.startswith('FL-'):
        messages.error(request, 'El folio debe comenzar con "FL-" (ej: FL-2026-0001).')
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    # Verificar que el número FL- no esté ya en uso en DetalleEquipo
    if DetalleEquipo.objects.filter(orden_cliente=numero_fl).exists():
        messages.error(
            request,
            f'El folio {numero_fl} ya está registrado en otra orden. '
            'Por favor usa un número diferente.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    # --- Obtener el técnico ---
    try:
        tecnico = Empleado.objects.get(pk=tecnico_id)
    except Empleado.DoesNotExist:
        messages.error(request, 'El técnico seleccionado no es válido.')
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    # --- Determinar la sucursal desde el usuario que creó la solicitud ---
    # Se intenta obtener la sucursal del empleado creador; si no tiene, se usa la del técnico
    try:
        empleado_creador = Empleado.objects.get(user=solicitud.creado_por)
        sucursal = empleado_creador.sucursal
    except (Empleado.DoesNotExist, AttributeError):
        # Fallback: usar la sucursal del técnico asignado
        sucursal = tecnico.sucursal

    if not sucursal:
        messages.error(
            request,
            'No se pudo determinar la sucursal. '
            'El creador de la solicitud o el técnico no tienen sucursal asignada.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)

    # --- Obtener el responsable de seguimiento (el usuario actual) ---
    try:
        responsable = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        # Si el usuario actual no tiene empleado asociado, usar el técnico como responsable
        responsable = tecnico

    try:
        # ====== PASO 1: Crear la OrdenServicio ======
        # tipo_servicio='venta_mostrador' porque estas órdenes son Servicio Directo sin Diagnóstico
        # estado='almacen' porque se origina desde el módulo Almacén
        orden = OrdenServicio.objects.create(
            sucursal=sucursal,
            responsable_seguimiento=responsable,
            tecnico_asignado_actual=tecnico,
            estado='almacen',
            tipo_servicio='venta_mostrador',
        )

        # ====== PASO 2: Crear el DetalleEquipo con los datos REALES del cliente ======
        # A diferencia del flujo genérico de solicitudes de baja (que usa placeholders),
        # aquí ya tenemos datos reales del cliente capturados en la SolicitudCotizacion.

        # Determinar tipo_equipo: usar el de la solicitud o fallback a 'Laptop'
        tipo_equipo_real = solicitud.tipo_equipo if solicitud.tipo_equipo else 'Laptop'

        # Determinar marca: usar la de la solicitud o fallback a 'Otra'
        marca_real = solicitud.marca if solicitud.marca else 'Otra'

        # Determinar modelo: usar el de la solicitud o placeholder
        modelo_real = solicitud.modelo if solicitud.modelo else 'Por definir'

        # Determinar número de serie: usar service_tag de la solicitud o placeholder
        numero_serie_real = solicitud.service_tag if solicitud.service_tag else f'ALMACEN-{numero_fl}'

        # Determinar falla_principal: usar observaciones de la solicitud o texto genérico
        falla_real = (
            solicitud.observaciones
            if solicitud.observaciones
            else 'Servicio directo sin diagnóstico - Cotización aprobada por cliente'
        )

        # Email del cliente: usar el de la solicitud o placeholder
        email_real = solicitud.email_cliente if solicitud.email_cliente else 'pendiente@actualizar.com'

        DetalleEquipo.objects.create(
            orden=orden,
            orden_cliente=numero_fl,        # Número FL- define es_fuera_garantia=True automáticamente
            tipo_equipo=tipo_equipo_real,
            marca=marca_real,
            modelo=modelo_real,
            numero_serie=numero_serie_real,
            gama='media',                   # Valor por defecto; el técnico puede ajustar después
            falla_principal=falla_real,
            email_cliente=email_real,
            # Datos adicionales del cliente (campos opcionales en DetalleEquipo)
            nombre_cliente=solicitud.nombre_cliente or '',
            telefono_cliente=solicitud.telefono_cliente or '',
            rfc_cliente=solicitud.rfc_cliente or '',
        )

        # ====== PASO 3: Vincular la solicitud con la nueva orden ======
        # vincular_orden() hace:
        #   - self.orden_servicio = orden
        #   - self.sin_orden_activa = False
        #   - Sincroniza numero_orden_cliente desde DetalleEquipo
        #   - self.save() → crea Cotizacion en ST vía _sincronizar_cotizacion_st()
        solicitud.vincular_orden(orden)

        # ====== PASO 4: Crear VentaMostrador vacío para la nueva orden ======
        # Para órdenes FL- (Venta Mostrador / Servicio Directo), NO se sincronizan
        # las líneas a PiezaCotizada (flujo de diagnóstico). En su lugar se crea el
        # VentaMostrador vacío ahora para que exista el objeto receptor; las piezas
        # individuales se crearán como PiezaVentaMostrador cuando el usuario pulse
        # "Generar Compras" en el detalle de la cotización.
        from servicio_tecnico.models import VentaMostrador as VentaMostradorModel
        VentaMostradorModel.objects.get_or_create(
            orden=orden,
            defaults={'fecha_venta': timezone.now()}
        )

        # Construir mensaje de éxito con información de la orden creada
        mensaje = (
            f'Orden {numero_fl} creada exitosamente y vinculada a la solicitud. '
            f'Orden interna: {orden.numero_orden_interno}. '
            f'Las piezas aprobadas se registrarán en Venta Mostrador al generar compras.'
        )

        messages.success(request, mensaje)

        logger.info(
            f"OrdenServicio {orden.numero_orden_interno} (FL: {numero_fl}) creada desde "
            f"SolicitudCotizacion {solicitud.numero_solicitud} por usuario {request.user}"
        )

    except ValueError as e:
        # Error de validación del método vincular_orden (ej: ya tiene otra solicitud activa)
        messages.error(request, f'Error al vincular la orden: {str(e)}')
        # Si la orden fue creada pero falló el vínculo, eliminarla para evitar órfanos
        try:
            orden.delete()
        except Exception:
            pass
    except Exception as e:
        messages.error(request, f'Error inesperado al crear la orden: {str(e)}')
        logger.error(
            f"Error al crear OrdenServicio desde SolicitudCotizacion {pk}: {e}",
            exc_info=True
        )
        # Intentar eliminar la orden parcialmente creada si existe
        try:
            orden.delete()
        except Exception:
            pass

    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.change_solicitudcotizacion')
def cancelar_solicitud_cotizacion(request, pk):
    """
    Cancelar una solicitud de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Cancela la solicitud completa. Solo se puede cancelar si no está
    completada (es decir, si aún no se generaron todas las compras).
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '')
        
        if solicitud.cancelar(motivo=motivo):
            messages.success(
                request,
                f'Solicitud {solicitud.numero_solicitud} cancelada.'
            )
        else:
            messages.error(
                request,
                'No se puede cancelar esta solicitud (ya está completada o cancelada).'
            )
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


@login_required
@permission_required_with_message('almacen.delete_solicitudcotizacion')
def eliminar_solicitud_cotizacion(request, pk):
    """
    Eliminar una solicitud de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Elimina permanentemente la solicitud y todas sus líneas.
    Solo se puede eliminar si está en estado 'borrador' o 'cancelada'.
    """
    solicitud = get_object_or_404(SolicitudCotizacion, pk=pk)
    
    if solicitud.estado not in ['borrador', 'cancelada']:
        messages.error(
            request,
            'Solo se pueden eliminar solicitudes en estado borrador o canceladas.'
        )
        return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)
    
    if request.method == 'POST':
        numero = solicitud.numero_solicitud
        solicitud.delete()
        messages.success(request, f'Solicitud {numero} eliminada.')
        return redirect('almacen:lista_solicitudes_cotizacion')
    
    return redirect('almacen:detalle_solicitud_cotizacion', pk=pk)


# ============================================================================
# GESTIÓN DE IMÁGENES DE LÍNEAS DE COTIZACIÓN
# ============================================================================

@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def gestionar_imagenes_linea(request, solicitud_pk, linea_pk):
    """
    Gestionar imágenes de referencia de una línea de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista permite ver, subir y eliminar imágenes de referencia
    para una línea específica de cotización.
    
    ¿Por qué es útil?
    - El proveedor ve exactamente qué pieza se necesita
    - El cliente puede verificar las especificaciones
    - Queda evidencia visual para trazabilidad
    
    Restricciones:
    - Solo se pueden gestionar imágenes si la solicitud está en estado 'borrador'
    - Máximo 5 imágenes por línea
    - Las imágenes mayores a 2MB se comprimen automáticamente
    
    Args:
        request: Solicitud HTTP
        solicitud_pk: ID de la SolicitudCotizacion
        linea_pk: ID de la LineaCotizacion
    
    Returns:
        Renderizado del template con el formulario y las imágenes actuales
    """
    # Obtener la solicitud y validar acceso
    solicitud = get_object_or_404(
        SolicitudCotizacion.objects.select_related('orden_servicio'),
        pk=solicitud_pk
    )
    
    # Obtener la línea y validar que pertenece a la solicitud
    linea = get_object_or_404(
        LineaCotizacion.objects.select_related('producto', 'proveedor'),
        pk=linea_pk,
        solicitud=solicitud
    )
    
    # Solo se pueden gestionar imágenes en estado borrador
    puede_editar = solicitud.estado == 'borrador'
    
    # Obtener imágenes existentes
    imagenes = linea.imagenes.all().order_by('fecha_subida')
    
    # Calcular información de límites
    imagenes_restantes = ImagenLineaCotizacion.imagenes_restantes(linea)
    puede_agregar = imagenes_restantes > 0 and puede_editar
    
    # Procesar formulario de subida
    if request.method == 'POST' and puede_agregar:
        form = ImagenLineaCotizacionForm(
            request.POST,
            request.FILES,
            linea=linea
        )
        
        if form.is_valid():
            try:
                # Guardar la imagen asociándola a la línea y al usuario
                imagen = form.save(commit=False)
                imagen.linea = linea
                imagen.subido_por = request.user
                imagen.save()
                
                messages.success(
                    request,
                    f'Imagen subida exitosamente. '
                    f'Quedan {ImagenLineaCotizacion.imagenes_restantes(linea)} espacios disponibles.'
                )
                
                # Redirigir para evitar reenvío del formulario
                return redirect(
                    'almacen:gestionar_imagenes_linea',
                    solicitud_pk=solicitud_pk,
                    linea_pk=linea_pk
                )
                
            except ValueError as e:
                messages.error(request, str(e))
        else:
            # Mostrar errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ImagenLineaCotizacionForm(linea=linea) if puede_agregar else None
    
    context = {
        'solicitud': solicitud,
        'linea': linea,
        'imagenes': imagenes,
        'form': form,
        'puede_editar': puede_editar,
        'puede_agregar': puede_agregar,
        'imagenes_restantes': imagenes_restantes,
        'max_imagenes': ImagenLineaCotizacion.MAX_IMAGENES_POR_LINEA,
        'titulo': f'Imágenes - Línea #{linea.numero_linea}',
    }
    
    return render(request, 'almacen/cotizaciones/gestionar_imagenes_linea.html', context)


@login_required
@permission_required_with_message('almacen.change_lineacotizacion')
def eliminar_imagen_linea(request, solicitud_pk, linea_pk, imagen_pk):
    """
    Eliminar una imagen de una línea de cotización.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista elimina una imagen específica de una línea de cotización.
    
    Validaciones:
    - La imagen debe pertenecer a la línea indicada
    - La solicitud debe estar en estado 'borrador'
    - Solo se procesa si es una solicitud POST (para evitar eliminaciones accidentales)
    
    Args:
        request: Solicitud HTTP
        solicitud_pk: ID de la SolicitudCotizacion
        linea_pk: ID de la LineaCotizacion
        imagen_pk: ID de la ImagenLineaCotizacion a eliminar
    
    Returns:
        Redirección a la vista de gestión de imágenes
    """
    # Validar la cadena completa: solicitud → línea → imagen
    solicitud = get_object_or_404(SolicitudCotizacion, pk=solicitud_pk)
    linea = get_object_or_404(LineaCotizacion, pk=linea_pk, solicitud=solicitud)
    imagen = get_object_or_404(ImagenLineaCotizacion, pk=imagen_pk, linea=linea)
    
    # Solo se pueden eliminar imágenes en estado borrador
    if solicitud.estado != 'borrador':
        messages.error(
            request,
            'Solo se pueden eliminar imágenes cuando la solicitud está en borrador.'
        )
        return redirect(
            'almacen:detalle_solicitud_cotizacion',
            pk=solicitud_pk
        )
    
    # Solo procesar eliminación con método POST
    if request.method == 'POST':
        nombre_archivo = imagen.nombre_archivo
        imagen.delete()
        messages.success(
            request,
            f'Imagen "{nombre_archivo}" eliminada correctamente.'
        )
    
    return redirect(
        'almacen:detalle_solicitud_cotizacion',
        pk=solicitud_pk
    )


@login_required
@permission_required_with_message('almacen.view_lineacotizacion')
def api_imagenes_linea(request, linea_pk):
    """
    API para obtener las imágenes de una línea en formato JSON.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista retorna las imágenes de una línea en formato JSON,
    útil para actualizar la interfaz con JavaScript sin recargar la página.
    
    La respuesta incluye:
    - Lista de imágenes con su URL, descripción y metadata
    - Información sobre cuántas imágenes más se pueden subir
    - Si se puede agregar más imágenes
    
    Args:
        request: Solicitud HTTP
        linea_pk: ID de la LineaCotizacion
    
    Returns:
        JsonResponse con la información de las imágenes
    """
    linea = get_object_or_404(
        LineaCotizacion.objects.select_related('solicitud'),
        pk=linea_pk
    )
    
    imagenes = linea.imagenes.all().order_by('fecha_subida')
    
    imagenes_data = []
    for img in imagenes:
        imagenes_data.append({
            'id': img.pk,
            'url': img.imagen.url,
            'nombre': img.nombre_archivo,
            'descripcion': img.descripcion or '',
            'fecha_subida': img.fecha_subida.strftime('%d/%m/%Y %H:%M'),
            'fue_comprimida': img.fue_comprimida,
            'tamano_kb': img.tamano_final_kb,
        })
    
    return JsonResponse({
        'success': True,
        'imagenes': imagenes_data,
        'total_imagenes': len(imagenes_data),
        'imagenes_restantes': ImagenLineaCotizacion.imagenes_restantes(linea),
        'puede_agregar': ImagenLineaCotizacion.puede_agregar_imagen(linea),
        'max_imagenes': ImagenLineaCotizacion.MAX_IMAGENES_POR_LINEA,
        'puede_editar': linea.solicitud.estado == 'borrador',
    })


# ============================================================================
# VISTA DE ACCESO DENEGADO
# ============================================================================
@login_required
def acceso_denegado(request):
    """
    Vista para mostrar página de acceso denegado con información del error.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista se ejecuta cuando un usuario intenta acceder a una funcionalidad
    del módulo Almacén pero NO tiene el permiso necesario.
    
    Muestra información útil:
    - Mensaje de error personalizado
    - Permiso específico que se requería
    - Grupos/roles a los que pertenece el usuario
    - Sugerencia de contactar al administrador
    
    Parámetros GET:
        - mensaje: Descripción del error
        - permiso: Permiso que se requería (formato: 'app.permiso_modelo')
    
    Returns:
        Renderiza template almacen/acceso_denegado.html
    """
    # Obtener parámetros de la URL
    mensaje = request.GET.get(
        'mensaje', 
        'No tienes permisos para acceder a esta sección del módulo Almacén.'
    )
    permiso = request.GET.get('permiso', 'N/A')
    
    # Obtener grupos del usuario para mostrar información útil
    grupos = request.user.groups.all()
    
    context = {
        'mensaje': mensaje,
        'permiso_requerido': permiso,
        'grupos_usuario': grupos,
    }
    
    return render(request, 'almacen/acceso_denegado.html', context)


# ============================================================================
# DASHBOARD: DISTRIBUCIÓN MULTI-SUCURSAL
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def dashboard_distribucion_sucursales(request):
    """
    Dashboard de distribución de inventario multi-sucursal.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra cómo están distribuidos los productos de almacén
    entre las diferentes sucursales de la empresa.
    
    Para cada producto muestra:
    - Entradas: Unidades que llegaron a cada sucursal
    - Salidas: Unidades que salieron (asignadas/vendidas)
    - TOTAL: Stock actual disponible en cada sucursal
    
    LÓGICA DE CÁLCULO:
    ------------------
    - Se basa en UnidadInventario.sucursal_actual (ubicación física real)
    - Solo cuenta unidades con disponibilidad='disponible'
    - Las unidades asignadas/vendidas NO aparecen en el conteo
    - Las unidades en almacén central tienen sucursal_actual=NULL
    
    EXPORTACIÓN:
    ------------
    Incluye botón para exportar a Excel con múltiples hojas de análisis.
    Si se detecta el parámetro ?export=excel, se genera el archivo Excel.
    
    Returns:
        Renderiza template con tabla interactiva y filtros, o archivo Excel
    """
    from inventario.models import Sucursal
    from django.db.models import Count, Q, F
    from django.db.models.functions import Coalesce
    from datetime import timedelta
    
    # ========== DETECTAR SOLICITUD DE EXPORTACIÓN ==========
    if request.GET.get('export') == 'excel':
        return exportar_distribucion_excel(request)
    
    # ========== OBTENER SUCURSALES ACTIVAS ==========
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')
    
    # ========== OBTENER PRODUCTOS ACTIVOS ==========
    productos = ProductoAlmacen.objects.filter(
        activo=True
    ).select_related(
        'categoria', 
        'proveedor_principal'
    ).prefetch_related(
        'unidades'  # Prefetch para optimizar consultas
    ).order_by('nombre')
    
    # ========== FILTROS ==========
    # Filtro por búsqueda de texto
    q = request.GET.get('q', '').strip()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    # Filtro por categoría
    categoria_id = request.GET.get('categoria', '')
    if categoria_id:
        try:
            productos = productos.filter(categoria_id=int(categoria_id))
        except (ValueError, TypeError):
            pass
    
    # Filtro por sucursal (mostrar solo productos con stock en esa sucursal)
    sucursal_filtro = request.GET.get('sucursal', '')
    if sucursal_filtro:
        if sucursal_filtro == 'central':
            productos = productos.filter(
                unidades__sucursal_actual__isnull=True,
                unidades__disponibilidad='disponible'
            ).distinct()
        else:
            try:
                productos = productos.filter(
                    unidades__sucursal_actual_id=int(sucursal_filtro),
                    unidades__disponibilidad='disponible'
                ).distinct()
            except (ValueError, TypeError):
                pass
    
    # ========== CONSTRUIR DATOS DE DISTRIBUCIÓN ==========
    productos_data = []
    
    for producto in productos:
        # Diccionario para almacenar inventario por sucursal
        inventario_sucursales = {}
        
        # ========== ALMACÉN CENTRAL ==========
        # Contar unidades en almacén central (sin sucursal asignada)
        central_disponibles = producto.unidades.filter(
            sucursal_actual__isnull=True,
            disponibilidad='disponible'
        ).count()
        
        # Para simplificar, en esta primera versión solo mostramos el total actual
        # En versiones futuras se puede agregar tracking de entradas/salidas
        inventario_sucursales['central'] = {
            'entradas': central_disponibles,  # Por ahora igual al total
            'salidas': 0,  # Calculable con historial de movimientos
            'total': central_disponibles
        }
        
        # ========== SUCURSALES ==========
        for sucursal in sucursales:
            # Contar unidades disponibles en esta sucursal
            sucursal_disponibles = producto.unidades.filter(
                sucursal_actual=sucursal,
                disponibilidad='disponible'
            ).count()
            
            inventario_sucursales[sucursal.codigo] = {
                'entradas': sucursal_disponibles,  # Por ahora igual al total
                'salidas': 0,  # Calculable con historial
                'total': sucursal_disponibles
            }
        
        # ========== CALCULAR TOTAL GENERAL ==========
        total_general = sum(
            inv['total'] for inv in inventario_sucursales.values()
        )
        
        # ========== CALCULAR DÍAS SIN MOVIMIENTO ==========
        # Obtener la unidad más reciente del producto
        ultima_unidad = producto.unidades.order_by('-fecha_registro').first()
        dias_sin_movimiento = None
        
        if ultima_unidad:
            delta = timezone.now() - ultima_unidad.fecha_registro
            dias_sin_movimiento = delta.days
        
        # ========== AGREGAR A LA LISTA ==========
        # Solo agregar productos que tengan al menos 1 unidad O que coincidan con filtros
        if total_general > 0 or q or categoria_id:
            productos_data.append({
                'id': producto.pk,
                'codigo': producto.codigo_producto,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'categoria_id': producto.categoria.pk if producto.categoria else None,
                'inventario': inventario_sucursales,
                'total_general': total_general,
                'dias_sin_movimiento': dias_sin_movimiento,
                'tipo': producto.tipo_producto,
            })
    
    # ========== PAGINACIÓN ==========
    paginator = Paginator(productos_data, 50)  # 50 productos por página
    page = request.GET.get('page', 1)
    productos_page = paginator.get_page(page)
    
    # ========== KPIs ==========
    total_productos_con_stock = len([p for p in productos_data if p['total_general'] > 0])
    total_unidades_sistema = sum(p['total_general'] for p in productos_data)
    
    # Productos sin stock
    productos_sin_stock = ProductoAlmacen.objects.filter(
        activo=True
    ).exclude(
        id__in=[p['id'] for p in productos_data if p['total_general'] > 0]
    ).count()
    
    # Sucursales con más stock
    stock_por_sucursal = {}
    stock_por_sucursal['Almacén Central'] = sum(
        p['inventario'].get('central', {}).get('total', 0) 
        for p in productos_data
    )
    for sucursal in sucursales:
        stock_por_sucursal[sucursal.nombre] = sum(
            p['inventario'].get(sucursal.codigo, {}).get('total', 0)
            for p in productos_data
        )
    
    # ========== CATEGORÍAS DISPONIBLES (para filtro) ==========
    categorias_disponibles = CategoriaAlmacen.objects.filter(
        activo=True,
        productos__activo=True
    ).distinct().order_by('nombre')
    
    # ========== CONTEXTO ==========
    context = {
        'productos': productos_page,
        'sucursales': sucursales,
        'categorias': categorias_disponibles,
        'total_productos_con_stock': total_productos_con_stock,
        'total_unidades_sistema': total_unidades_sistema,
        'productos_sin_stock': productos_sin_stock,
        'stock_por_sucursal': stock_por_sucursal,
        # Filtros aplicados
        'q': q,
        'categoria_filtro': categoria_id,
        'sucursal_filtro': sucursal_filtro,
    }
    
    return render(request, 'almacen/dashboard_distribucion_sucursales.html', context)


# ============================================================================
# EXPORTACIÓN EXCEL: DISTRIBUCIÓN MULTI-SUCURSAL
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def exportar_distribucion_excel(request):
    """
    Exporta el dashboard de distribución multi-sucursal a Excel con análisis completo.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta función genera un archivo Excel profesional con 7 hojas especializadas:
    
    1. DISTRIBUCIÓN ACTUAL: Stock actual por sucursal (vista simple y clara)
    2. ANÁLISIS DE MOVIMIENTOS: Entradas/Salidas históricas por sucursal
    3. HISTORIAL DE TRANSFERENCIAS: Todas las transferencias entre sucursales
    4. RESUMEN POR SUCURSAL: Estadísticas y porcentajes por ubicación
    5. PRODUCTOS SIN STOCK: Lista de productos agotados que necesitan reposición
    6. MOVIMIENTOS RECIENTES: Últimos 30 días de actividad en el almacén
    7. ALERTAS DE REPOSICIÓN: Productos con stock crítico (≤10 unidades)
    
    La Vista Web muestra solo el stock actual (simple), pero el Excel proporciona
    análisis completo con histórico de entradas/salidas para toma de decisiones.
    
    Utiliza openpyxl para crear el archivo Excel con formato profesional y colores.
    
    Returns:
        HttpResponse con archivo Excel (.xlsx) para descarga inmediata
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from inventario.models import Sucursal
    from django.db.models import Count, Q, Sum
    from datetime import datetime, timedelta
    
    # ========== OBTENER DATOS (MISMA LÓGICA QUE LA VISTA) ==========
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')
    
    productos = ProductoAlmacen.objects.filter(
        activo=True
    ).select_related(
        'categoria', 
        'proveedor_principal'
    ).prefetch_related(
        'unidades'
    ).order_by('nombre')
    
    # Aplicar filtros si existen
    q = request.GET.get('q', '').strip()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    categoria_id = request.GET.get('categoria', '')
    if categoria_id:
        try:
            productos = productos.filter(categoria_id=int(categoria_id))
        except (ValueError, TypeError):
            pass
    
    sucursal_filtro = request.GET.get('sucursal', '')
    if sucursal_filtro:
        if sucursal_filtro == 'central':
            productos = productos.filter(
                unidades__sucursal_actual__isnull=True,
                unidades__disponibilidad='disponible'
            ).distinct()
        else:
            try:
                productos = productos.filter(
                    unidades__sucursal_actual_id=int(sucursal_filtro),
                    unidades__disponibilidad='disponible'
                ).distinct()
            except (ValueError, TypeError):
                pass
    
    # Construir datos de distribución
    productos_data = []
    
    for producto in productos:
        inventario_sucursales = {}
        
        # Almacén Central
        central_disponibles = producto.unidades.filter(
            sucursal_actual__isnull=True,
            disponibilidad='disponible'
        ).count()
        
        inventario_sucursales['central'] = {
            'entradas': central_disponibles,
            'salidas': 0,
            'total': central_disponibles
        }
        
        # Sucursales
        for sucursal in sucursales:
            sucursal_disponibles = producto.unidades.filter(
                sucursal_actual=sucursal,
                disponibilidad='disponible'
            ).count()
            
            inventario_sucursales[sucursal.codigo] = {
                'entradas': sucursal_disponibles,
                'salidas': 0,
                'total': sucursal_disponibles
            }
        
        total_general = sum(inv['total'] for inv in inventario_sucursales.values())
        
        ultima_unidad = producto.unidades.order_by('-fecha_registro').first()
        dias_sin_movimiento = None
        
        if ultima_unidad:
            delta = timezone.now() - ultima_unidad.fecha_registro
            dias_sin_movimiento = delta.days
        
        if total_general > 0 or q or categoria_id:
            productos_data.append({
                'id': producto.pk,
                'codigo': producto.codigo_producto,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'inventario': inventario_sucursales,
                'total_general': total_general,
                'dias_sin_movimiento': dias_sin_movimiento,
                'tipo': producto.tipo_producto,
                'proveedor': producto.proveedor_principal.nombre if producto.proveedor_principal else 'Sin proveedor',
                'costo_unitario': float(producto.costo_unitario),
            })
    
    # ========== CREAR WORKBOOK ==========
    wb = Workbook()
    
    # ========== HOJA 1: DISTRIBUCIÓN GENERAL ==========
    ws1 = wb.active
    ws1.title = "Distribución General"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws1.merge_cells('A1:' + get_column_letter(4 + len(sucursales) + 1) + '1')
    ws1['A1'] = 'DISTRIBUCIÓN MULTI-SUCURSAL DE INVENTARIO'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información del reporte
    ws1['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(italic=True, size=9)
    
    if q:
        ws1['A3'] = f'Filtro aplicado: "{q}"'
        ws1['A3'].font = Font(italic=True, size=9)
    
    # Encabezados principales
    row = 5
    col = 1
    
    # Columnas fijas
    headers_fijos = ['Código', 'Producto', 'Categoría', 'Proveedor']
    for header in headers_fijos:
        cell = ws1.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col += 1
    
    # Almacén Central (una sola columna)
    cell = ws1.cell(row=row, column=col)
    cell.value = 'ALMACÉN CENTRAL'
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    col += 1
    
    # Sucursales (una columna por sucursal)
    for sucursal in sucursales:
        cell = ws1.cell(row=row, column=col)
        cell.value = sucursal.nombre.upper()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col += 1
    
    # Columna TOTAL GENERAL
    cell = ws1.cell(row=row, column=col)
    cell.value = 'TOTAL GENERAL'
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # ========== DATOS ==========
    row = 6  # Ahora empezamos en la fila 6 (sin subencabezados)
    for producto in productos_data:
        col = 1
        
        # Datos fijos
        ws1.cell(row=row, column=col, value=producto['codigo']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['nombre']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['categoria']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['proveedor']).border = border
        col += 1
        
        # Almacén Central (solo total con color)
        central = producto['inventario']['central']
        cell_total = ws1.cell(row=row, column=col, value=central['total'])
        cell_total.border = border
        cell_total.alignment = Alignment(horizontal='center')
        
        if central['total'] == 0:
            cell_total.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True)
        elif central['total'] <= 10:
            cell_total.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_total.font = Font(bold=True)
        else:
            cell_total.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True)
        col += 1
        
        # Sucursales (solo total con color)
        for sucursal in sucursales:
            inv = producto['inventario'].get(sucursal.codigo, {'total': 0})
            
            cell_total = ws1.cell(row=row, column=col, value=inv['total'])
            cell_total.border = border
            cell_total.alignment = Alignment(horizontal='center')
            
            if inv['total'] == 0:
                cell_total.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell_total.font = Font(color="FFFFFF", bold=True)
            elif inv['total'] <= 10:
                cell_total.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell_total.font = Font(bold=True)
            else:
                cell_total.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                cell_total.font = Font(color="FFFFFF", bold=True)
            col += 1
        
        # Total General
        cell_total = ws1.cell(row=row, column=col, value=producto['total_general'])
        cell_total.border = border
        cell_total.alignment = Alignment(horizontal='center')
        cell_total.font = Font(bold=True, size=11)
        
        if producto['total_general'] == 0:
            cell_total.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True, size=11)
        elif producto['total_general'] <= 10:
            cell_total.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell_total.font = Font(bold=True, size=11)
        else:
            cell_total.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True, size=11)
        
        row += 1
    
    # Ajustar ancho de columnas
    for col in range(1, 4 + len(sucursales) + 2):
        column_letter = get_column_letter(col)
        if col <= 2:
            ws1.column_dimensions[column_letter].width = 15
        elif col <= 4:
            ws1.column_dimensions[column_letter].width = 20
        else:
            ws1.column_dimensions[column_letter].width = 12
    
    # ========== HOJA 2: ANÁLISIS DE MOVIMIENTOS (ENTRADAS/SALIDAS HISTÓRICAS) ==========
    ws2 = wb.create_sheet("Análisis de Movimientos")
    
    ws2['A1'] = 'ANÁLISIS HISTÓRICO DE ENTRADAS Y SALIDAS POR SUCURSAL'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:F1')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws2['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Datos históricos completos desde el inicio del sistema'
    ws2['A2'].font = Font(italic=True, size=9, color="666666")
    ws2.merge_cells('A2:F2')
    
    # Encabezados
    row = 4
    headers_movimientos = ['Sucursal', 'Producto', 'Entradas', 'Salidas', 'Transferencias Netas', 'Stock Actual']
    for col, header in enumerate(headers_movimientos, start=1):
        cell = ws2.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos - Análisis por producto y sucursal
    row = 5
    
    # EXPLICACIÓN: Calculamos entradas/salidas basándonos en MovimientoAlmacen
    # - Entradas: movimientos tipo='entrada' + transferencias entrantes
    # - Salidas: movimientos tipo='salida' + transferencias salientes
    # - Stock Actual: lo que realmente hay ahora (de UnidadInventario)
    
    for producto in productos_data:
        producto_obj = ProductoAlmacen.objects.get(pk=producto['id'])
        
        # Almacén Central
        # Entradas: todos los movimientos de entrada sin considerar sucursal específica
        entradas_central = MovimientoAlmacen.objects.filter(
            producto=producto_obj,
            tipo='entrada'
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Salidas: movimientos de salida desde central
        salidas_central = MovimientoAlmacen.objects.filter(
            producto=producto_obj,
            tipo='salida'
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Transferencias salientes de central (SolicitudBaja aprobadas desde central)
        transferencias_salientes_central = SolicitudBaja.objects.filter(
            producto=producto_obj,
            tipo_solicitud='transferencia',
            estado='aprobada',
            producto__sucursal__isnull=True  # Origen: Central
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Transferencias entrantes a central (poco común, pero posible)
        transferencias_entrantes_central = SolicitudBaja.objects.filter(
            producto=producto_obj,
            tipo_solicitud='transferencia',
            estado='aprobada',
            sucursal_destino__isnull=True  # Destino: Central
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        transferencias_netas_central = transferencias_entrantes_central - transferencias_salientes_central
        stock_actual_central = producto['inventario']['central']['total']
        
        ws2.cell(row=row, column=1, value='Almacén Central').border = border
        ws2.cell(row=row, column=2, value=producto['nombre']).border = border
        ws2.cell(row=row, column=3, value=entradas_central).border = border
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=4, value=salidas_central).border = border
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=5, value=transferencias_netas_central).border = border
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        
        # Color en stock actual según nivel
        cell_stock = ws2.cell(row=row, column=6, value=stock_actual_central)
        cell_stock.border = border
        cell_stock.alignment = Alignment(horizontal='center')
        if stock_actual_central == 0:
            cell_stock.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_stock.font = Font(color="FFFFFF", bold=True)
        elif stock_actual_central <= 10:
            cell_stock.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_stock.font = Font(bold=True)
        else:
            cell_stock.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_stock.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Sucursales
        for sucursal in sucursales:
            # Transferencias entrantes a esta sucursal
            transferencias_entrantes = SolicitudBaja.objects.filter(
                producto=producto_obj,
                tipo_solicitud='transferencia',
                estado='aprobada',
                sucursal_destino=sucursal
            ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
            
            # Transferencias salientes desde esta sucursal
            transferencias_salientes = SolicitudBaja.objects.filter(
                producto=producto_obj,
                tipo_solicitud='transferencia',
                estado='aprobada',
                producto__sucursal=sucursal
            ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
            
            transferencias_netas = transferencias_entrantes - transferencias_salientes
            stock_actual = producto['inventario'].get(sucursal.codigo, {'total': 0})['total']
            
            # Las entradas/salidas directas en sucursales son principalmente transferencias
            # No hay movimientos de "entrada" directos a sucursales (vienen del central)
            entradas_suc = transferencias_entrantes
            salidas_suc = transferencias_salientes
            
            ws2.cell(row=row, column=1, value=sucursal.nombre).border = border
            ws2.cell(row=row, column=2, value=producto['nombre']).border = border
            ws2.cell(row=row, column=3, value=entradas_suc).border = border
            ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=4, value=salidas_suc).border = border
            ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=5, value=transferencias_netas).border = border
            ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            
            # Color en stock actual
            cell_stock = ws2.cell(row=row, column=6, value=stock_actual)
            cell_stock.border = border
            cell_stock.alignment = Alignment(horizontal='center')
            if stock_actual == 0:
                cell_stock.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell_stock.font = Font(color="FFFFFF", bold=True)
            elif stock_actual <= 10:
                cell_stock.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell_stock.font = Font(bold=True)
            else:
                cell_stock.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                cell_stock.font = Font(color="FFFFFF", bold=True)
            
            row += 1
    
    # Nota explicativa
    ws2.cell(row=row + 2, column=1, value='NOTA:').font = Font(bold=True, size=10)
    ws2.cell(row=row + 3, column=1, value='• Entradas: Movimientos de tipo "entrada" registrados en el sistema')
    ws2.cell(row=row + 4, column=1, value='• Salidas: Movimientos de tipo "salida" (consumos, servicios técnicos, etc.)')
    ws2.cell(row=row + 5, column=1, value='• Transferencias Netas: Diferencia entre transferencias entrantes y salientes')
    ws2.cell(row=row + 6, column=1, value='• Stock Actual: Unidades físicamente disponibles en cada ubicación')
    
    for r in range(row + 2, row + 7):
        ws2.merge_cells(f'A{r}:F{r}')
        ws2.cell(row=r, column=1).font = Font(italic=True, size=9, color="666666")
    
    # Ajustar anchos
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 15
    
    # ========== HOJA 3: HISTORIAL DE TRANSFERENCIAS ==========
    ws3 = wb.create_sheet("Transferencias")
    
    ws3['A1'] = 'HISTORIAL COMPLETO DE TRANSFERENCIAS ENTRE SUCURSALES'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:G1')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws3['A2'] = 'Registro de todos los movimientos de inventario entre ubicaciones'
    ws3['A2'].font = Font(italic=True, size=9, color="666666")
    ws3.merge_cells('A2:G2')
    
    # Encabezados
    row = 4
    headers_transferencias = ['Fecha', 'Producto', 'Cantidad', 'Origen', 'Destino', 'Solicitante', 'Estado']
    for col, header in enumerate(headers_transferencias, start=1):
        cell = ws3.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Obtener todas las transferencias
    transferencias = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia'
    ).select_related(
        'producto', 
        'producto__sucursal', 
        'sucursal_destino', 
        'solicitante'
    ).order_by('-fecha_solicitud')
    
    row = 5
    for trans in transferencias:
        origen = trans.producto.sucursal.nombre if trans.producto.sucursal else 'Almacén Central'
        destino = trans.sucursal_destino.nombre if trans.sucursal_destino else 'Almacén Central'
        
        ws3.cell(row=row, column=1, value=trans.fecha_solicitud.strftime('%d/%m/%Y %H:%M')).border = border
        ws3.cell(row=row, column=2, value=trans.producto.nombre).border = border
        ws3.cell(row=row, column=3, value=trans.cantidad).border = border
        ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws3.cell(row=row, column=4, value=origen).border = border
        ws3.cell(row=row, column=5, value=destino).border = border
        ws3.cell(row=row, column=6, value=trans.solicitante.nombre_completo if trans.solicitante else 'N/A').border = border
        
        # Estado con color
        cell_estado = ws3.cell(row=row, column=7, value=trans.get_estado_display())
        cell_estado.border = border
        cell_estado.alignment = Alignment(horizontal='center')
        
        if trans.estado == 'aprobado':
            cell_estado.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_estado.font = Font(color="FFFFFF", bold=True)
        elif trans.estado == 'rechazado':
            cell_estado.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_estado.font = Font(color="FFFFFF", bold=True)
        else:  # pendiente
            cell_estado.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_estado.font = Font(bold=True)
        
        row += 1
    
    if not transferencias.exists():
        ws3.cell(row=5, column=1, value='No hay transferencias registradas en el sistema.')
        ws3.merge_cells('A5:G5')
        ws3['A5'].alignment = Alignment(horizontal='center')
        ws3['A5'].font = Font(italic=True, size=11, color="999999")
    
    # Ajustar anchos
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 20
    ws3.column_dimensions['F'].width = 25
    ws3.column_dimensions['G'].width = 15
    
    # ========== HOJA 4: RESUMEN POR SUCURSAL ==========
    ws4 = wb.create_sheet("Resumen por Sucursal")
    
    ws4['A1'] = 'RESUMEN DE INVENTARIO POR UBICACIÓN'
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.merge_cells('A1:D1')
    ws4['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Sucursal', 'Total Unidades', 'Productos Diferentes', '% del Total']
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    total_sistema = sum(p['total_general'] for p in productos_data)
    
    row = 4
    
    # Almacén Central
    total_central = sum(p['inventario']['central']['total'] for p in productos_data)
    productos_central = sum(1 for p in productos_data if p['inventario']['central']['total'] > 0)
    porcentaje_central = (total_central / total_sistema * 100) if total_sistema > 0 else 0
    
    ws4.cell(row=row, column=1, value='Almacén Central').border = border
    ws4.cell(row=row, column=2, value=total_central).border = border
    ws4.cell(row=row, column=3, value=productos_central).border = border
    ws4.cell(row=row, column=4, value=f'{porcentaje_central:.1f}%').border = border
    row += 1
    
    # Sucursales
    for sucursal in sucursales:
        total_suc = sum(p['inventario'].get(sucursal.codigo, {'total': 0})['total'] for p in productos_data)
        productos_suc = sum(1 for p in productos_data if p['inventario'].get(sucursal.codigo, {'total': 0})['total'] > 0)
        porcentaje_suc = (total_suc / total_sistema * 100) if total_sistema > 0 else 0
        
        ws4.cell(row=row, column=1, value=sucursal.nombre).border = border
        ws4.cell(row=row, column=2, value=total_suc).border = border
        ws4.cell(row=row, column=3, value=productos_suc).border = border
        ws4.cell(row=row, column=4, value=f'{porcentaje_suc:.1f}%').border = border
        row += 1
    
    # Total
    ws4.cell(row=row, column=1, value='TOTAL SISTEMA').font = Font(bold=True)
    ws4.cell(row=row, column=1).border = border
    ws4.cell(row=row, column=2, value=total_sistema).font = Font(bold=True)
    ws4.cell(row=row, column=2).border = border
    ws4.cell(row=row, column=3, value=len(productos_data)).font = Font(bold=True)
    ws4.cell(row=row, column=3).border = border
    ws4.cell(row=row, column=4, value='100%').font = Font(bold=True)
    ws4.cell(row=row, column=4).border = border
    
    # Ajustar anchos
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 15
    
    # ========== HOJA 5: PRODUCTOS SIN STOCK ==========
    ws5 = wb.create_sheet("Productos Sin Stock")
    
    ws5['A1'] = 'PRODUCTOS SIN STOCK EN TODO EL SISTEMA'
    ws5['A1'].font = Font(bold=True, size=14)
    ws5.merge_cells('A1:E1')
    ws5['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Código', 'Producto', 'Categoría', 'Proveedor', 'Días sin Movimiento']
    for col, header in enumerate(headers, start=1):
        cell = ws5.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    productos_sin_stock = [p for p in productos_data if p['total_general'] == 0]
    
    row = 4
    for producto in productos_sin_stock:
        ws5.cell(row=row, column=1, value=producto['codigo']).border = border
        ws5.cell(row=row, column=2, value=producto['nombre']).border = border
        ws5.cell(row=row, column=3, value=producto['categoria']).border = border
        ws5.cell(row=row, column=4, value=producto['proveedor']).border = border
        ws5.cell(row=row, column=5, value=producto['dias_sin_movimiento'] or 'N/A').border = border
        row += 1
    
    if not productos_sin_stock:
        ws5.cell(row=4, column=1, value='¡Excelente! Todos los productos tienen stock disponible.')
        ws5.merge_cells('A4:E4')
        ws5['A4'].alignment = Alignment(horizontal='center')
        ws5['A4'].font = Font(italic=True, color="70AD47")
    
    # Ajustar anchos
    for col in range(1, 6):
        ws5.column_dimensions[get_column_letter(col)].width = 20
    
    # ========== HOJA 6: MOVIMIENTOS RECIENTES ==========
    ws6 = wb.create_sheet("Movimientos Recientes")
    
    ws6['A1'] = 'MOVIMIENTOS DE INVENTARIO - ÚLTIMOS 30 DÍAS'
    ws6['A1'].font = Font(bold=True, size=14)
    ws6.merge_cells('A1:F1')
    ws6['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Fecha', 'Producto', 'Tipo', 'Cantidad', 'Empleado', 'Observaciones']
    for col, header in enumerate(headers, start=1):
        cell = ws6.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener movimientos recientes
    fecha_limite = timezone.now() - timedelta(days=30)
    movimientos_recientes = MovimientoAlmacen.objects.filter(
        fecha__gte=fecha_limite
    ).select_related('producto', 'empleado').order_by('-fecha')[:100]
    
    row = 4
    for movimiento in movimientos_recientes:
        ws6.cell(row=row, column=1, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M')).border = border
        ws6.cell(row=row, column=2, value=movimiento.producto.nombre).border = border
        ws6.cell(row=row, column=3, value=movimiento.get_tipo_display()).border = border
        ws6.cell(row=row, column=4, value=movimiento.cantidad).border = border
        ws6.cell(row=row, column=5, value=movimiento.empleado.nombre_completo if movimiento.empleado else 'N/A').border = border
        ws6.cell(row=row, column=6, value=movimiento.observaciones or '').border = border
        row += 1
    
    if not movimientos_recientes.exists():
        ws6.cell(row=4, column=1, value='No hay movimientos registrados en los últimos 30 días.')
        ws6.merge_cells('A4:F4')
        ws6['A4'].alignment = Alignment(horizontal='center')
        ws6['A4'].font = Font(italic=True)
    
    # Ajustar anchos
    ws6.column_dimensions['A'].width = 18
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 15
    ws6.column_dimensions['D'].width = 12
    ws6.column_dimensions['E'].width = 25
    ws6.column_dimensions['F'].width = 40
    
    # ========== HOJA 7: ALERTAS DE REPOSICIÓN ==========
    ws7 = wb.create_sheet("Alertas de Reposición")
    
    ws7['A1'] = 'PRODUCTOS QUE REQUIEREN REABASTECIMIENTO'
    ws7['A1'].font = Font(bold=True, size=14)
    ws7.merge_cells('A1:F1')
    ws7['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Código', 'Producto', 'Stock Actual', 'Stock Mínimo', 'Proveedor', 'Costo Unit.']
    for col, header in enumerate(headers, start=1):
        cell = ws7.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener productos con stock bajo (total <= 10)
    productos_alerta = [p for p in productos_data if 0 < p['total_general'] <= 10]
    productos_alerta.sort(key=lambda x: x['total_general'])
    
    row = 4
    for producto in productos_alerta:
        ws7.cell(row=row, column=1, value=producto['codigo']).border = border
        ws7.cell(row=row, column=2, value=producto['nombre']).border = border
        
        cell = ws7.cell(row=row, column=3, value=producto['total_general'])
        cell.border = border
        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        ws7.cell(row=row, column=4, value=10).border = border
        ws7.cell(row=row, column=5, value=producto['proveedor']).border = border
        ws7.cell(row=row, column=6, value=f"${producto['costo_unitario']:.2f}").border = border
        row += 1
    
    if not productos_alerta:
        ws7.cell(row=4, column=1, value='No hay productos que requieran reposición urgente.')
        ws7.merge_cells('A4:F4')
        ws7['A4'].alignment = Alignment(horizontal='center')
        ws7['A4'].font = Font(italic=True, color="70AD47")
    
    # Ajustar anchos
    for col in range(1, 7):
        ws7.column_dimensions[get_column_letter(col)].width = 20
    
    # ========== GUARDAR Y RETORNAR ==========
    from io import BytesIO
    
    # Crear buffer en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Distribucion_Multi_Sucursal_{fecha_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

