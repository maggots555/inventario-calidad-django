"""
Dashboard de distribución multi-sucursal + exportación Excel.

EXPLICACIÓN PARA PRINCIPIANTES:
-------------------------------
Extraído de views.py (Fase 1 de modularización Almacén).
urls.py sigue usando views.dashboard_distribucion_sucursales porque
views.py reexporta estos nombres.

Contiene:
- dashboard_distribucion_sucursales: tabla web con filtros/paginación
- exportar_distribucion_excel: archivo .xlsx (también vía ?export=excel)

Efectos secundarios:
- Consulta ProductoAlmacen / UnidadInventario / MovimientoAlmacen / SolicitudBaja
- La exportación genera un HttpResponse binario (no guarda archivo en disco)
"""

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.shortcuts import render
from django.utils import timezone

from .decorators import permission_required_with_message
from .models import (
    CategoriaAlmacen,
    MovimientoAlmacen,
    ProductoAlmacen,
    SolicitudBaja,
)


# ============================================================================
# DASHBOARD: DISTRIBUCIÓN MULTI-SUCURSAL
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def dashboard_distribucion_sucursales(request):
    """
    Dashboard de distribución de inventario multi-sucursal.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta vista muestra cómo están distribuidos los productos de almacén
    entre las diferentes sucursales de la empresa.
    
    Para cada producto muestra:
    - Entradas: Unidades que llegaron a cada sucursal
    - Salidas: Unidades que salieron (asignadas/vendidas)
    - TOTAL: Stock actual disponible en cada sucursal
    
    LÓGICA DE CÁLCULO:
    ------------------
    - Se basa en UnidadInventario.sucursal_actual (ubicación física real)
    - Solo cuenta unidades con disponibilidad='disponible'
    - Las unidades asignadas/vendidas NO aparecen en el conteo
    - Las unidades en almacén central tienen sucursal_actual=NULL
    
    EXPORTACIÓN:
    ------------
    Incluye botón para exportar a Excel con múltiples hojas de análisis.
    Si se detecta el parámetro ?export=excel, se genera el archivo Excel.
    
    Returns:
        Renderiza template con tabla interactiva y filtros, o archivo Excel
    """
    from inventario.models import Sucursal
    from django.db.models import Count, Q, F
    from django.db.models.functions import Coalesce
    from datetime import timedelta
    
    # ========== DETECTAR SOLICITUD DE EXPORTACIÓN ==========
    if request.GET.get('export') == 'excel':
        return exportar_distribucion_excel(request)
    
    # ========== OBTENER SUCURSALES ACTIVAS ==========
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')
    
    # ========== OBTENER PRODUCTOS ACTIVOS ==========
    productos = ProductoAlmacen.objects.filter(
        activo=True
    ).select_related(
        'categoria', 
        'proveedor_principal'
    ).prefetch_related(
        'unidades'  # Prefetch para optimizar consultas
    ).order_by('nombre')
    
    # ========== FILTROS ==========
    # Filtro por búsqueda de texto
    q = request.GET.get('q', '').strip()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    # Filtro por categoría
    categoria_id = request.GET.get('categoria', '')
    if categoria_id:
        try:
            productos = productos.filter(categoria_id=int(categoria_id))
        except (ValueError, TypeError):
            pass
    
    # Filtro por sucursal (mostrar solo productos con stock en esa sucursal)
    sucursal_filtro = request.GET.get('sucursal', '')
    if sucursal_filtro:
        if sucursal_filtro == 'central':
            productos = productos.filter(
                unidades__sucursal_actual__isnull=True,
                unidades__disponibilidad='disponible'
            ).distinct()
        else:
            try:
                productos = productos.filter(
                    unidades__sucursal_actual_id=int(sucursal_filtro),
                    unidades__disponibilidad='disponible'
                ).distinct()
            except (ValueError, TypeError):
                pass
    
    # ========== CONSTRUIR DATOS DE DISTRIBUCIÓN ==========
    productos_data = []
    
    for producto in productos:
        # Diccionario para almacenar inventario por sucursal
        inventario_sucursales = {}
        
        # ========== ALMACÉN CENTRAL ==========
        # Contar unidades en almacén central (sin sucursal asignada)
        central_disponibles = producto.unidades.filter(
            sucursal_actual__isnull=True,
            disponibilidad='disponible'
        ).count()
        
        # Para simplificar, en esta primera versión solo mostramos el total actual
        # En versiones futuras se puede agregar tracking de entradas/salidas
        inventario_sucursales['central'] = {
            'entradas': central_disponibles,  # Por ahora igual al total
            'salidas': 0,  # Calculable con historial de movimientos
            'total': central_disponibles
        }
        
        # ========== SUCURSALES ==========
        for sucursal in sucursales:
            # Contar unidades disponibles en esta sucursal
            sucursal_disponibles = producto.unidades.filter(
                sucursal_actual=sucursal,
                disponibilidad='disponible'
            ).count()
            
            inventario_sucursales[sucursal.codigo] = {
                'entradas': sucursal_disponibles,  # Por ahora igual al total
                'salidas': 0,  # Calculable con historial
                'total': sucursal_disponibles
            }
        
        # ========== CALCULAR TOTAL GENERAL ==========
        total_general = sum(
            inv['total'] for inv in inventario_sucursales.values()
        )
        
        # ========== CALCULAR DÍAS SIN MOVIMIENTO ==========
        # Obtener la unidad más reciente del producto
        ultima_unidad = producto.unidades.order_by('-fecha_registro').first()
        dias_sin_movimiento = None
        
        if ultima_unidad:
            delta = timezone.now() - ultima_unidad.fecha_registro
            dias_sin_movimiento = delta.days
        
        # ========== AGREGAR A LA LISTA ==========
        # Solo agregar productos que tengan al menos 1 unidad O que coincidan con filtros
        if total_general > 0 or q or categoria_id:
            productos_data.append({
                'id': producto.pk,
                'codigo': producto.codigo_producto,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'categoria_id': producto.categoria.pk if producto.categoria else None,
                'inventario': inventario_sucursales,
                'total_general': total_general,
                'dias_sin_movimiento': dias_sin_movimiento,
                'tipo': producto.tipo_producto,
            })
    
    # ========== PAGINACIÓN ==========
    paginator = Paginator(productos_data, 50)  # 50 productos por página
    page = request.GET.get('page', 1)
    productos_page = paginator.get_page(page)
    
    # ========== KPIs ==========
    total_productos_con_stock = len([p for p in productos_data if p['total_general'] > 0])
    total_unidades_sistema = sum(p['total_general'] for p in productos_data)
    
    # Productos sin stock
    productos_sin_stock = ProductoAlmacen.objects.filter(
        activo=True
    ).exclude(
        id__in=[p['id'] for p in productos_data if p['total_general'] > 0]
    ).count()
    
    # Sucursales con más stock
    stock_por_sucursal = {}
    stock_por_sucursal['Almacén Central'] = sum(
        p['inventario'].get('central', {}).get('total', 0) 
        for p in productos_data
    )
    for sucursal in sucursales:
        stock_por_sucursal[sucursal.nombre] = sum(
            p['inventario'].get(sucursal.codigo, {}).get('total', 0)
            for p in productos_data
        )
    
    # ========== CATEGORÍAS DISPONIBLES (para filtro) ==========
    categorias_disponibles = CategoriaAlmacen.objects.filter(
        activo=True,
        productos__activo=True
    ).distinct().order_by('nombre')
    
    # ========== CONTEXTO ==========
    context = {
        'productos': productos_page,
        'sucursales': sucursales,
        'categorias': categorias_disponibles,
        'total_productos_con_stock': total_productos_con_stock,
        'total_unidades_sistema': total_unidades_sistema,
        'productos_sin_stock': productos_sin_stock,
        'stock_por_sucursal': stock_por_sucursal,
        # Filtros aplicados
        'q': q,
        'categoria_filtro': categoria_id,
        'sucursal_filtro': sucursal_filtro,
    }
    
    return render(request, 'almacen/dashboard_distribucion_sucursales.html', context)


# ============================================================================
# EXPORTACIÓN EXCEL: DISTRIBUCIÓN MULTI-SUCURSAL
# ============================================================================
@login_required
@permission_required_with_message('almacen.view_productoalmacen')
def exportar_distribucion_excel(request):
    """
    Exporta el dashboard de distribución multi-sucursal a Excel con análisis completo.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    --------------------------------
    Esta función genera un archivo Excel profesional con 7 hojas especializadas:
    
    1. DISTRIBUCIÓN ACTUAL: Stock actual por sucursal (vista simple y clara)
    2. ANÁLISIS DE MOVIMIENTOS: Entradas/Salidas históricas por sucursal
    3. HISTORIAL DE TRANSFERENCIAS: Todas las transferencias entre sucursales
    4. RESUMEN POR SUCURSAL: Estadísticas y porcentajes por ubicación
    5. PRODUCTOS SIN STOCK: Lista de productos agotados que necesitan reposición
    6. MOVIMIENTOS RECIENTES: Últimos 30 días de actividad en el almacén
    7. ALERTAS DE REPOSICIÓN: Productos con stock crítico (≤10 unidades)
    
    La Vista Web muestra solo el stock actual (simple), pero el Excel proporciona
    análisis completo con histórico de entradas/salidas para toma de decisiones.
    
    Utiliza openpyxl para crear el archivo Excel con formato profesional y colores.
    
    Returns:
        HttpResponse con archivo Excel (.xlsx) para descarga inmediata
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from inventario.models import Sucursal
    from django.db.models import Count, Q, Sum
    from datetime import datetime, timedelta
    
    # ========== OBTENER DATOS (MISMA LÓGICA QUE LA VISTA) ==========
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')
    
    productos = ProductoAlmacen.objects.filter(
        activo=True
    ).select_related(
        'categoria', 
        'proveedor_principal'
    ).prefetch_related(
        'unidades'
    ).order_by('nombre')
    
    # Aplicar filtros si existen
    q = request.GET.get('q', '').strip()
    if q:
        productos = productos.filter(
            Q(codigo_producto__icontains=q) |
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )
    
    categoria_id = request.GET.get('categoria', '')
    if categoria_id:
        try:
            productos = productos.filter(categoria_id=int(categoria_id))
        except (ValueError, TypeError):
            pass
    
    sucursal_filtro = request.GET.get('sucursal', '')
    if sucursal_filtro:
        if sucursal_filtro == 'central':
            productos = productos.filter(
                unidades__sucursal_actual__isnull=True,
                unidades__disponibilidad='disponible'
            ).distinct()
        else:
            try:
                productos = productos.filter(
                    unidades__sucursal_actual_id=int(sucursal_filtro),
                    unidades__disponibilidad='disponible'
                ).distinct()
            except (ValueError, TypeError):
                pass
    
    # Construir datos de distribución
    productos_data = []
    
    for producto in productos:
        inventario_sucursales = {}
        
        # Almacén Central
        central_disponibles = producto.unidades.filter(
            sucursal_actual__isnull=True,
            disponibilidad='disponible'
        ).count()
        
        inventario_sucursales['central'] = {
            'entradas': central_disponibles,
            'salidas': 0,
            'total': central_disponibles
        }
        
        # Sucursales
        for sucursal in sucursales:
            sucursal_disponibles = producto.unidades.filter(
                sucursal_actual=sucursal,
                disponibilidad='disponible'
            ).count()
            
            inventario_sucursales[sucursal.codigo] = {
                'entradas': sucursal_disponibles,
                'salidas': 0,
                'total': sucursal_disponibles
            }
        
        total_general = sum(inv['total'] for inv in inventario_sucursales.values())
        
        ultima_unidad = producto.unidades.order_by('-fecha_registro').first()
        dias_sin_movimiento = None
        
        if ultima_unidad:
            delta = timezone.now() - ultima_unidad.fecha_registro
            dias_sin_movimiento = delta.days
        
        if total_general > 0 or q or categoria_id:
            productos_data.append({
                'id': producto.pk,
                'codigo': producto.codigo_producto,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'inventario': inventario_sucursales,
                'total_general': total_general,
                'dias_sin_movimiento': dias_sin_movimiento,
                'tipo': producto.tipo_producto,
                'proveedor': producto.proveedor_principal.nombre if producto.proveedor_principal else 'Sin proveedor',
                'costo_unitario': float(producto.costo_unitario),
            })
    
    # ========== CREAR WORKBOOK ==========
    wb = Workbook()
    
    # ========== HOJA 1: DISTRIBUCIÓN GENERAL ==========
    ws1 = wb.active
    ws1.title = "Distribución General"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws1.merge_cells('A1:' + get_column_letter(4 + len(sucursales) + 1) + '1')
    ws1['A1'] = 'DISTRIBUCIÓN MULTI-SUCURSAL DE INVENTARIO'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información del reporte
    ws1['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(italic=True, size=9)
    
    if q:
        ws1['A3'] = f'Filtro aplicado: "{q}"'
        ws1['A3'].font = Font(italic=True, size=9)
    
    # Encabezados principales
    row = 5
    col = 1
    
    # Columnas fijas
    headers_fijos = ['Código', 'Producto', 'Categoría', 'Proveedor']
    for header in headers_fijos:
        cell = ws1.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col += 1
    
    # Almacén Central (una sola columna)
    cell = ws1.cell(row=row, column=col)
    cell.value = 'ALMACÉN CENTRAL'
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    col += 1
    
    # Sucursales (una columna por sucursal)
    for sucursal in sucursales:
        cell = ws1.cell(row=row, column=col)
        cell.value = sucursal.nombre.upper()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col += 1
    
    # Columna TOTAL GENERAL
    cell = ws1.cell(row=row, column=col)
    cell.value = 'TOTAL GENERAL'
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    
    # ========== DATOS ==========
    row = 6  # Ahora empezamos en la fila 6 (sin subencabezados)
    for producto in productos_data:
        col = 1
        
        # Datos fijos
        ws1.cell(row=row, column=col, value=producto['codigo']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['nombre']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['categoria']).border = border
        col += 1
        ws1.cell(row=row, column=col, value=producto['proveedor']).border = border
        col += 1
        
        # Almacén Central (solo total con color)
        central = producto['inventario']['central']
        cell_total = ws1.cell(row=row, column=col, value=central['total'])
        cell_total.border = border
        cell_total.alignment = Alignment(horizontal='center')
        
        if central['total'] == 0:
            cell_total.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True)
        elif central['total'] <= 10:
            cell_total.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_total.font = Font(bold=True)
        else:
            cell_total.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True)
        col += 1
        
        # Sucursales (solo total con color)
        for sucursal in sucursales:
            inv = producto['inventario'].get(sucursal.codigo, {'total': 0})
            
            cell_total = ws1.cell(row=row, column=col, value=inv['total'])
            cell_total.border = border
            cell_total.alignment = Alignment(horizontal='center')
            
            if inv['total'] == 0:
                cell_total.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell_total.font = Font(color="FFFFFF", bold=True)
            elif inv['total'] <= 10:
                cell_total.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell_total.font = Font(bold=True)
            else:
                cell_total.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                cell_total.font = Font(color="FFFFFF", bold=True)
            col += 1
        
        # Total General
        cell_total = ws1.cell(row=row, column=col, value=producto['total_general'])
        cell_total.border = border
        cell_total.alignment = Alignment(horizontal='center')
        cell_total.font = Font(bold=True, size=11)
        
        if producto['total_general'] == 0:
            cell_total.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True, size=11)
        elif producto['total_general'] <= 10:
            cell_total.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell_total.font = Font(bold=True, size=11)
        else:
            cell_total.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell_total.font = Font(color="FFFFFF", bold=True, size=11)
        
        row += 1
    
    # Ajustar ancho de columnas
    for col in range(1, 4 + len(sucursales) + 2):
        column_letter = get_column_letter(col)
        if col <= 2:
            ws1.column_dimensions[column_letter].width = 15
        elif col <= 4:
            ws1.column_dimensions[column_letter].width = 20
        else:
            ws1.column_dimensions[column_letter].width = 12
    
    # ========== HOJA 2: ANÁLISIS DE MOVIMIENTOS (ENTRADAS/SALIDAS HISTÓRICAS) ==========
    ws2 = wb.create_sheet("Análisis de Movimientos")
    
    ws2['A1'] = 'ANÁLISIS HISTÓRICO DE ENTRADAS Y SALIDAS POR SUCURSAL'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:F1')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws2['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Datos históricos completos desde el inicio del sistema'
    ws2['A2'].font = Font(italic=True, size=9, color="666666")
    ws2.merge_cells('A2:F2')
    
    # Encabezados
    row = 4
    headers_movimientos = ['Sucursal', 'Producto', 'Entradas', 'Salidas', 'Transferencias Netas', 'Stock Actual']
    for col, header in enumerate(headers_movimientos, start=1):
        cell = ws2.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos - Análisis por producto y sucursal
    row = 5
    
    # EXPLICACIÓN: Calculamos entradas/salidas basándonos en MovimientoAlmacen
    # - Entradas: movimientos tipo='entrada' + transferencias entrantes
    # - Salidas: movimientos tipo='salida' + transferencias salientes
    # - Stock Actual: lo que realmente hay ahora (de UnidadInventario)
    
    for producto in productos_data:
        producto_obj = ProductoAlmacen.objects.get(pk=producto['id'])
        
        # Almacén Central
        # Entradas: todos los movimientos de entrada sin considerar sucursal específica
        entradas_central = MovimientoAlmacen.objects.filter(
            producto=producto_obj,
            tipo='entrada'
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Salidas: movimientos de salida desde central
        salidas_central = MovimientoAlmacen.objects.filter(
            producto=producto_obj,
            tipo='salida'
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Transferencias salientes de central (SolicitudBaja aprobadas desde central)
        transferencias_salientes_central = SolicitudBaja.objects.filter(
            producto=producto_obj,
            tipo_solicitud='transferencia',
            estado='aprobada',
            producto__sucursal__isnull=True  # Origen: Central
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        # Transferencias entrantes a central (poco común, pero posible)
        transferencias_entrantes_central = SolicitudBaja.objects.filter(
            producto=producto_obj,
            tipo_solicitud='transferencia',
            estado='aprobada',
            sucursal_destino__isnull=True  # Destino: Central
        ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
        
        transferencias_netas_central = transferencias_entrantes_central - transferencias_salientes_central
        stock_actual_central = producto['inventario']['central']['total']
        
        ws2.cell(row=row, column=1, value='Almacén Central').border = border
        ws2.cell(row=row, column=2, value=producto['nombre']).border = border
        ws2.cell(row=row, column=3, value=entradas_central).border = border
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=4, value=salidas_central).border = border
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=5, value=transferencias_netas_central).border = border
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        
        # Color en stock actual según nivel
        cell_stock = ws2.cell(row=row, column=6, value=stock_actual_central)
        cell_stock.border = border
        cell_stock.alignment = Alignment(horizontal='center')
        if stock_actual_central == 0:
            cell_stock.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_stock.font = Font(color="FFFFFF", bold=True)
        elif stock_actual_central <= 10:
            cell_stock.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_stock.font = Font(bold=True)
        else:
            cell_stock.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_stock.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Sucursales
        for sucursal in sucursales:
            # Transferencias entrantes a esta sucursal
            transferencias_entrantes = SolicitudBaja.objects.filter(
                producto=producto_obj,
                tipo_solicitud='transferencia',
                estado='aprobada',
                sucursal_destino=sucursal
            ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
            
            # Transferencias salientes desde esta sucursal
            transferencias_salientes = SolicitudBaja.objects.filter(
                producto=producto_obj,
                tipo_solicitud='transferencia',
                estado='aprobada',
                producto__sucursal=sucursal
            ).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
            
            transferencias_netas = transferencias_entrantes - transferencias_salientes
            stock_actual = producto['inventario'].get(sucursal.codigo, {'total': 0})['total']
            
            # Las entradas/salidas directas en sucursales son principalmente transferencias
            # No hay movimientos de "entrada" directos a sucursales (vienen del central)
            entradas_suc = transferencias_entrantes
            salidas_suc = transferencias_salientes
            
            ws2.cell(row=row, column=1, value=sucursal.nombre).border = border
            ws2.cell(row=row, column=2, value=producto['nombre']).border = border
            ws2.cell(row=row, column=3, value=entradas_suc).border = border
            ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=4, value=salidas_suc).border = border
            ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=5, value=transferencias_netas).border = border
            ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            
            # Color en stock actual
            cell_stock = ws2.cell(row=row, column=6, value=stock_actual)
            cell_stock.border = border
            cell_stock.alignment = Alignment(horizontal='center')
            if stock_actual == 0:
                cell_stock.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell_stock.font = Font(color="FFFFFF", bold=True)
            elif stock_actual <= 10:
                cell_stock.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell_stock.font = Font(bold=True)
            else:
                cell_stock.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                cell_stock.font = Font(color="FFFFFF", bold=True)
            
            row += 1
    
    # Nota explicativa
    ws2.cell(row=row + 2, column=1, value='NOTA:').font = Font(bold=True, size=10)
    ws2.cell(row=row + 3, column=1, value='• Entradas: Movimientos de tipo "entrada" registrados en el sistema')
    ws2.cell(row=row + 4, column=1, value='• Salidas: Movimientos de tipo "salida" (consumos, servicios técnicos, etc.)')
    ws2.cell(row=row + 5, column=1, value='• Transferencias Netas: Diferencia entre transferencias entrantes y salientes')
    ws2.cell(row=row + 6, column=1, value='• Stock Actual: Unidades físicamente disponibles en cada ubicación')
    
    for r in range(row + 2, row + 7):
        ws2.merge_cells(f'A{r}:F{r}')
        ws2.cell(row=r, column=1).font = Font(italic=True, size=9, color="666666")
    
    # Ajustar anchos
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 15
    
    # ========== HOJA 3: HISTORIAL DE TRANSFERENCIAS ==========
    ws3 = wb.create_sheet("Transferencias")
    
    ws3['A1'] = 'HISTORIAL COMPLETO DE TRANSFERENCIAS ENTRE SUCURSALES'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:G1')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws3['A2'] = 'Registro de todos los movimientos de inventario entre ubicaciones'
    ws3['A2'].font = Font(italic=True, size=9, color="666666")
    ws3.merge_cells('A2:G2')
    
    # Encabezados
    row = 4
    headers_transferencias = ['Fecha', 'Producto', 'Cantidad', 'Origen', 'Destino', 'Solicitante', 'Estado']
    for col, header in enumerate(headers_transferencias, start=1):
        cell = ws3.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Obtener todas las transferencias
    transferencias = SolicitudBaja.objects.filter(
        tipo_solicitud='transferencia'
    ).select_related(
        'producto', 
        'producto__sucursal', 
        'sucursal_destino', 
        'solicitante'
    ).order_by('-fecha_solicitud')
    
    row = 5
    for trans in transferencias:
        origen = trans.producto.sucursal.nombre if trans.producto.sucursal else 'Almacén Central'
        destino = trans.sucursal_destino.nombre if trans.sucursal_destino else 'Almacén Central'
        
        ws3.cell(row=row, column=1, value=trans.fecha_solicitud.strftime('%d/%m/%Y %H:%M')).border = border
        ws3.cell(row=row, column=2, value=trans.producto.nombre).border = border
        ws3.cell(row=row, column=3, value=trans.cantidad).border = border
        ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws3.cell(row=row, column=4, value=origen).border = border
        ws3.cell(row=row, column=5, value=destino).border = border
        ws3.cell(row=row, column=6, value=trans.solicitante.nombre_completo if trans.solicitante else 'N/A').border = border
        
        # Estado con color
        cell_estado = ws3.cell(row=row, column=7, value=trans.get_estado_display())
        cell_estado.border = border
        cell_estado.alignment = Alignment(horizontal='center')
        
        if trans.estado == 'aprobado':
            cell_estado.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            cell_estado.font = Font(color="FFFFFF", bold=True)
        elif trans.estado == 'rechazado':
            cell_estado.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell_estado.font = Font(color="FFFFFF", bold=True)
        else:  # pendiente
            cell_estado.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            cell_estado.font = Font(bold=True)
        
        row += 1
    
    if not transferencias.exists():
        ws3.cell(row=5, column=1, value='No hay transferencias registradas en el sistema.')
        ws3.merge_cells('A5:G5')
        ws3['A5'].alignment = Alignment(horizontal='center')
        ws3['A5'].font = Font(italic=True, size=11, color="999999")
    
    # Ajustar anchos
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 20
    ws3.column_dimensions['F'].width = 25
    ws3.column_dimensions['G'].width = 15
    
    # ========== HOJA 4: RESUMEN POR SUCURSAL ==========
    ws4 = wb.create_sheet("Resumen por Sucursal")
    
    ws4['A1'] = 'RESUMEN DE INVENTARIO POR UBICACIÓN'
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.merge_cells('A1:D1')
    ws4['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Sucursal', 'Total Unidades', 'Productos Diferentes', '% del Total']
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    total_sistema = sum(p['total_general'] for p in productos_data)
    
    row = 4
    
    # Almacén Central
    total_central = sum(p['inventario']['central']['total'] for p in productos_data)
    productos_central = sum(1 for p in productos_data if p['inventario']['central']['total'] > 0)
    porcentaje_central = (total_central / total_sistema * 100) if total_sistema > 0 else 0
    
    ws4.cell(row=row, column=1, value='Almacén Central').border = border
    ws4.cell(row=row, column=2, value=total_central).border = border
    ws4.cell(row=row, column=3, value=productos_central).border = border
    ws4.cell(row=row, column=4, value=f'{porcentaje_central:.1f}%').border = border
    row += 1
    
    # Sucursales
    for sucursal in sucursales:
        total_suc = sum(p['inventario'].get(sucursal.codigo, {'total': 0})['total'] for p in productos_data)
        productos_suc = sum(1 for p in productos_data if p['inventario'].get(sucursal.codigo, {'total': 0})['total'] > 0)
        porcentaje_suc = (total_suc / total_sistema * 100) if total_sistema > 0 else 0
        
        ws4.cell(row=row, column=1, value=sucursal.nombre).border = border
        ws4.cell(row=row, column=2, value=total_suc).border = border
        ws4.cell(row=row, column=3, value=productos_suc).border = border
        ws4.cell(row=row, column=4, value=f'{porcentaje_suc:.1f}%').border = border
        row += 1
    
    # Total
    ws4.cell(row=row, column=1, value='TOTAL SISTEMA').font = Font(bold=True)
    ws4.cell(row=row, column=1).border = border
    ws4.cell(row=row, column=2, value=total_sistema).font = Font(bold=True)
    ws4.cell(row=row, column=2).border = border
    ws4.cell(row=row, column=3, value=len(productos_data)).font = Font(bold=True)
    ws4.cell(row=row, column=3).border = border
    ws4.cell(row=row, column=4, value='100%').font = Font(bold=True)
    ws4.cell(row=row, column=4).border = border
    
    # Ajustar anchos
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 15
    
    # ========== HOJA 5: PRODUCTOS SIN STOCK ==========
    ws5 = wb.create_sheet("Productos Sin Stock")
    
    ws5['A1'] = 'PRODUCTOS SIN STOCK EN TODO EL SISTEMA'
    ws5['A1'].font = Font(bold=True, size=14)
    ws5.merge_cells('A1:E1')
    ws5['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Código', 'Producto', 'Categoría', 'Proveedor', 'Días sin Movimiento']
    for col, header in enumerate(headers, start=1):
        cell = ws5.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    productos_sin_stock = [p for p in productos_data if p['total_general'] == 0]
    
    row = 4
    for producto in productos_sin_stock:
        ws5.cell(row=row, column=1, value=producto['codigo']).border = border
        ws5.cell(row=row, column=2, value=producto['nombre']).border = border
        ws5.cell(row=row, column=3, value=producto['categoria']).border = border
        ws5.cell(row=row, column=4, value=producto['proveedor']).border = border
        ws5.cell(row=row, column=5, value=producto['dias_sin_movimiento'] or 'N/A').border = border
        row += 1
    
    if not productos_sin_stock:
        ws5.cell(row=4, column=1, value='¡Excelente! Todos los productos tienen stock disponible.')
        ws5.merge_cells('A4:E4')
        ws5['A4'].alignment = Alignment(horizontal='center')
        ws5['A4'].font = Font(italic=True, color="70AD47")
    
    # Ajustar anchos
    for col in range(1, 6):
        ws5.column_dimensions[get_column_letter(col)].width = 20
    
    # ========== HOJA 6: MOVIMIENTOS RECIENTES ==========
    ws6 = wb.create_sheet("Movimientos Recientes")
    
    ws6['A1'] = 'MOVIMIENTOS DE INVENTARIO - ÚLTIMOS 30 DÍAS'
    ws6['A1'].font = Font(bold=True, size=14)
    ws6.merge_cells('A1:F1')
    ws6['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Fecha', 'Producto', 'Tipo', 'Cantidad', 'Empleado', 'Observaciones']
    for col, header in enumerate(headers, start=1):
        cell = ws6.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener movimientos recientes
    fecha_limite = timezone.now() - timedelta(days=30)
    movimientos_recientes = MovimientoAlmacen.objects.filter(
        fecha__gte=fecha_limite
    ).select_related('producto', 'empleado').order_by('-fecha')[:100]
    
    row = 4
    for movimiento in movimientos_recientes:
        ws6.cell(row=row, column=1, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M')).border = border
        ws6.cell(row=row, column=2, value=movimiento.producto.nombre).border = border
        ws6.cell(row=row, column=3, value=movimiento.get_tipo_display()).border = border
        ws6.cell(row=row, column=4, value=movimiento.cantidad).border = border
        ws6.cell(row=row, column=5, value=movimiento.empleado.nombre_completo if movimiento.empleado else 'N/A').border = border
        ws6.cell(row=row, column=6, value=movimiento.observaciones or '').border = border
        row += 1
    
    if not movimientos_recientes.exists():
        ws6.cell(row=4, column=1, value='No hay movimientos registrados en los últimos 30 días.')
        ws6.merge_cells('A4:F4')
        ws6['A4'].alignment = Alignment(horizontal='center')
        ws6['A4'].font = Font(italic=True)
    
    # Ajustar anchos
    ws6.column_dimensions['A'].width = 18
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 15
    ws6.column_dimensions['D'].width = 12
    ws6.column_dimensions['E'].width = 25
    ws6.column_dimensions['F'].width = 40
    
    # ========== HOJA 7: ALERTAS DE REPOSICIÓN ==========
    ws7 = wb.create_sheet("Alertas de Reposición")
    
    ws7['A1'] = 'PRODUCTOS QUE REQUIEREN REABASTECIMIENTO'
    ws7['A1'].font = Font(bold=True, size=14)
    ws7.merge_cells('A1:F1')
    ws7['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Código', 'Producto', 'Stock Actual', 'Stock Mínimo', 'Proveedor', 'Costo Unit.']
    for col, header in enumerate(headers, start=1):
        cell = ws7.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener productos con stock bajo (total <= 10)
    productos_alerta = [p for p in productos_data if 0 < p['total_general'] <= 10]
    productos_alerta.sort(key=lambda x: x['total_general'])
    
    row = 4
    for producto in productos_alerta:
        ws7.cell(row=row, column=1, value=producto['codigo']).border = border
        ws7.cell(row=row, column=2, value=producto['nombre']).border = border
        
        cell = ws7.cell(row=row, column=3, value=producto['total_general'])
        cell.border = border
        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        ws7.cell(row=row, column=4, value=10).border = border
        ws7.cell(row=row, column=5, value=producto['proveedor']).border = border
        ws7.cell(row=row, column=6, value=f"${producto['costo_unitario']:.2f}").border = border
        row += 1
    
    if not productos_alerta:
        ws7.cell(row=4, column=1, value='No hay productos que requieran reposición urgente.')
        ws7.merge_cells('A4:F4')
        ws7['A4'].alignment = Alignment(horizontal='center')
        ws7['A4'].font = Font(italic=True, color="70AD47")
    
    # Ajustar anchos
    for col in range(1, 7):
        ws7.column_dimensions[get_column_letter(col)].width = 20
    
    # ========== GUARDAR Y RETORNAR ==========
    from io import BytesIO
    
    # Crear buffer en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Distribucion_Multi_Sucursal_{fecha_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

