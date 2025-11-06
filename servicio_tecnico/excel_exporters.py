"""
Funciones auxiliares para exportación de Excel avanzada
Dashboard de Seguimiento OOW-/FL- - Reportes Completos con Múltiples Hojas

EXPLICACIÓN PARA PRINCIPIANTES:
================================
Este archivo contiene funciones que ayudan a crear archivos Excel profesionales
con múltiples hojas, estilos, colores y formato. Es como una "caja de herramientas"
que usaremos desde la vista principal para generar el Excel.

Dependencias necesarias:
- openpyxl: Librería para crear y manipular archivos Excel (.xlsx)
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from django.db.models import Count, Sum, Avg, Q
from collections import defaultdict
from decimal import Decimal


# ==========================================
# FUNCIONES DE ESTILOS
# ==========================================

def get_header_style():
    """
    Retorna el estilo para encabezados de tablas
    
    EXPLICACIÓN: Crea un formato visual para los encabezados (primera fila de tablas)
    con fondo azul, texto blanco en negrita y bordes.
    
    Returns:
        dict: Diccionario con configuración de Font, Fill, Alignment y Border
    """
    return {
        'font': Font(bold=True, color="FFFFFF", size=11),
        'fill': PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_title_style():
    """
    Retorna el estilo para títulos de secciones
    
    EXPLICACIÓN: Formato para títulos grandes de cada sección del Excel
    (por ejemplo: "RESUMEN GENERAL", "ANÁLISIS POR RESPONSABLE")
    
    Returns:
        dict: Diccionario con estilo de título
    """
    return {
        'font': Font(bold=True, size=14, color="1f4e78"),
        'fill': PatternFill(start_color="e7e6e6", end_color="e7e6e6", fill_type="solid"),
        'alignment': Alignment(horizontal="left", vertical="center")
    }


def get_kpi_title_style():
    """
    Retorna el estilo para títulos de KPIs (indicadores clave)
    
    EXPLICACIÓN: Formato para los nombres de las métricas
    (por ejemplo: "Total de Órdenes", "Ventas Mostrador")
    
    Returns:
        dict: Diccionario con estilo de KPI
    """
    return {
        'font': Font(bold=True, size=11, color="495057"),
        'alignment': Alignment(horizontal="left", vertical="center")
    }


def get_kpi_value_style():
    """
    Retorna el estilo para valores de KPIs
    
    EXPLICACIÓN: Formato para los números grandes que representan métricas importantes
    Los muestra grandes, en azul y centrados para que destaquen
    
    Returns:
        dict: Diccionario con estilo para valores numéricos
    """
    return {
        'font': Font(bold=True, size=16, color="0d6efd"),
        'alignment': Alignment(horizontal="center", vertical="center")
    }


def get_estado_color(estado):
    """
    Retorna el color de fondo según el estado de la orden
    
    EXPLICACIÓN: Cada estado de servicio tiene un color asociado para identificarlo
    rápidamente en el Excel (verde = entregado, amarillo = en proceso, etc.)
    
    ACTUALIZACIÓN (Nov 2025): Se agregaron colores para todos los estados nuevos
    para evitar que aparezcan en blanco en el Excel.
    
    Args:
        estado (str): Estado de la orden (entregado, finalizado, reparacion, etc.)
        
    Returns:
        str: Código hexadecimal del color (sin #)
    """
    colores = {
        # === ESTADOS BÁSICOS (Ya existentes) ===
        'entregado': '28a745',              # Verde - completado exitosamente
        'finalizado': '17a2b8',             # Azul claro - terminado pero no entregado
        'reparacion': 'ffc107',             # Amarillo - en proceso
        'cancelado': 'dc3545',              # Rojo - cancelado
        'diagnostico': 'fd7e14',            # Naranja - en diagnóstico
        'espera': '6c757d',                 # Gris - en espera
        'cotizacion': 'e83e8c',             # Rosa - esperando cotización
        'control_calidad': '20c997',        # Verde agua - en control de calidad
        'recepcion': 'a0c4ff',              # Azul claro suave - recepción
        
        # === ESTADOS NUEVOS AGREGADOS (Oct 2025) ===
        'equipo_diagnosticado': '6f42c1',   # Púrpura - diagnóstico completado
        'diagnostico_enviado_cliente': '9b59b6',  # Morado claro - diagnóstico enviado
        'cotizacion_enviada_proveedor': 'e74c3c', # Rojo ladrillo - cotización enviada a proveedor
        'cotizacion_recibida_proveedor': 'f39c12', # Naranja - cotización recibida de proveedor
        'cliente_acepta_cotizacion': '2ecc71',    # Verde esmeralda - cliente aceptó
        'rechazada': 'c0392b',              # Rojo oscuro - cotización rechazada
        'partes_solicitadas_proveedor': '3498db', # Azul - partes solicitadas
        'esperando_piezas': 'f1c40f',       # Amarillo dorado - esperando llegada
        'piezas_recibidas': '27ae60',       # Verde medio - piezas recibidas
        'wpb_pieza_incorrecta': 'e67e22',   # Naranja fuerte - WPB pieza incorrecta
        'doa_pieza_danada': 'c0392b',       # Rojo granate - DOA pieza dañada
        'pnc_parte_no_disponible': '95a5a6', # Gris medio - PNC parte no disponible
    }
    return colores.get(estado.lower(), 'cccccc')  # Gris claro por defecto (en vez de blanco)


def apply_cell_style(cell, style_dict):
    """
    Aplica un diccionario de estilos a una celda de Excel
    
    EXPLICACIÓN: Esta función toma una celda de Excel y le aplica los estilos
    definidos en un diccionario (fuente, color de fondo, alineación, bordes)
    
    Args:
        cell: Objeto celda de openpyxl
        style_dict: Diccionario con claves 'font', 'fill', 'alignment', 'border'
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']


def auto_adjust_column_width(ws, min_width=10, max_width=50):
    """
    Ajusta automáticamente el ancho de las columnas según el contenido
    
    EXPLICACIÓN: Revisa el contenido de cada columna y ajusta el ancho para que
    todo sea legible sin cortar texto. Tiene límites mínimo y máximo para evitar
    columnas demasiado estrechas o excesivamente anchas.
    
    Args:
        ws: Worksheet (hoja) de openpyxl
        min_width: Ancho mínimo en caracteres (default: 10)
        max_width: Ancho máximo en caracteres (default: 50)
    """
    for column_cells in ws.columns:
        # Obtener la letra de la columna de forma segura (evitar MergedCell)
        column_letter = None
        for cell in column_cells:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                break
        
        if not column_letter:
            continue
            
        # Calcular el ancho basado en el contenido más largo
        length = max(len(str(cell.value or '')) for cell in column_cells if hasattr(cell, 'value'))
        adjusted_width = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


# ==========================================
# FUNCIONES DE CÁLCULO Y ANÁLISIS
# ==========================================

def calcular_metricas_generales(ordenes):
    """
    Calcula métricas generales del dashboard OOW-/FL-
    
    EXPLICACIÓN: Esta función analiza todas las órdenes y calcula los indicadores
    principales que se muestran en las tarjetas del dashboard (totales, promedios,
    porcentajes, montos, etc.)
    
    Args:
        ordenes: QuerySet de OrdenServicio filtradas
        
    Returns:
        dict: Diccionario con todas las métricas calculadas
    """
    total_ordenes = ordenes.count()
    
    # Contar por estado
    ordenes_activas = ordenes.exclude(estado__in=['entregado', 'cancelado']).count()
    ordenes_finalizadas = ordenes.filter(estado='finalizado').count()
    ordenes_entregadas = ordenes.filter(estado='entregado').count()
    
    # Contar ventas mostrador
    total_ventas_mostrador = ordenes.filter(venta_mostrador__isnull=False).count()
    
    # Contar con cotización
    total_con_cotizacion = ordenes.filter(cotizacion__isnull=False).count()
    cotizaciones_aceptadas = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto=True
    ).count()
    cotizaciones_pendientes = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto__isnull=True
    ).count()
    cotizaciones_rechazadas = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto=False
    ).count()
    
    # Calcular montos totales
    monto_total_ventas_mostrador = Decimal('0.00')
    monto_total_cotizaciones = Decimal('0.00')
    
    for orden in ordenes:
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            monto_total_ventas_mostrador += orden.venta_mostrador.total_venta
        
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto:
                monto_total_cotizaciones += orden.cotizacion.costo_total_final
    
    monto_total_general = monto_total_ventas_mostrador + monto_total_cotizaciones
    
    # Calcular tiempo promedio (días hábiles)
    total_dias_habiles = 0
    ordenes_con_tiempo = 0
    
    for orden in ordenes:
        dias = orden.dias_habiles_en_servicio
        if dias >= 0:
            total_dias_habiles += dias
            ordenes_con_tiempo += 1
    
    tiempo_promedio = round(total_dias_habiles / ordenes_con_tiempo, 1) if ordenes_con_tiempo > 0 else 0
    
    # Calcular % en tiempo (<= 15 días hábiles)
    ordenes_en_tiempo = sum(1 for orden in ordenes if orden.dias_habiles_en_servicio <= 15)
    porcentaje_en_tiempo = round((ordenes_en_tiempo / total_ordenes) * 100, 1) if total_ordenes > 0 else 0
    
    return {
        'total_ordenes': total_ordenes,
        'ordenes_activas': ordenes_activas,
        'ordenes_finalizadas': ordenes_finalizadas,
        'ordenes_entregadas': ordenes_entregadas,
        'total_ventas_mostrador': total_ventas_mostrador,
        'monto_ventas_mostrador': float(monto_total_ventas_mostrador),
        'total_con_cotizacion': total_con_cotizacion,
        'cotizaciones_aceptadas': cotizaciones_aceptadas,
        'cotizaciones_pendientes': cotizaciones_pendientes,
        'cotizaciones_rechazadas': cotizaciones_rechazadas,
        'monto_cotizaciones': float(monto_total_cotizaciones),
        'monto_total': float(monto_total_general),
        'tiempo_promedio': tiempo_promedio,
        'porcentaje_en_tiempo': porcentaje_en_tiempo,
    }


def calcular_distribucion_estados(ordenes):
    """
    Calcula la distribución de órdenes por estado
    
    EXPLICACIÓN: Cuenta cuántas órdenes hay en cada estado (entregado, reparación, etc.)
    para hacer gráficos o tablas de distribución.
    
    Args:
        ordenes: QuerySet de OrdenServicio
        
    Returns:
        dict: Diccionario con {estado: cantidad}
    """
    distribucion = {}
    for orden in ordenes:
        estado = orden.get_estado_display()
        distribucion[estado] = distribucion.get(estado, 0) + 1
    
    return distribucion


def calcular_estadisticas_por_responsable(ordenes):
    """
    Calcula estadísticas consolidadas por responsable de seguimiento
    
    EXPLICACIÓN: Agrupa todas las órdenes por responsable y calcula sus métricas
    individuales (total órdenes, ventas, cotizaciones, tiempos, etc.)
    
    Args:
        ordenes: QuerySet de OrdenServicio
        
    Returns:
        list: Lista de diccionarios con estadísticas por responsable
    """
    responsables_data = {}
    
    for orden in ordenes:
        resp_id = orden.responsable_seguimiento.id
        resp_nombre = orden.responsable_seguimiento.nombre_completo
        
        if resp_id not in responsables_data:
            responsables_data[resp_id] = {
                'id': resp_id,
                'nombre': resp_nombre,
                'total_ordenes': 0,
                'ordenes_activas': 0,
                'ordenes_finalizadas': 0,
                'ordenes_entregadas': 0,
                'ventas_mostrador': 0,
                'con_cotizacion': 0,
                'cotizaciones_aceptadas': 0,
                'cotizaciones_pendientes': 0,
                'cotizaciones_rechazadas': 0,
                'monto_ventas_mostrador': Decimal('0.00'),
                'monto_cotizaciones': Decimal('0.00'),
                'dias_acumulados': 0,
            }
        
        # Acumular estadísticas
        responsables_data[resp_id]['total_ordenes'] += 1
        
        if orden.estado not in ['entregado', 'cancelado']:
            responsables_data[resp_id]['ordenes_activas'] += 1
        
        if orden.estado == 'finalizado':
            responsables_data[resp_id]['ordenes_finalizadas'] += 1
        
        if orden.estado == 'entregado':
            responsables_data[resp_id]['ordenes_entregadas'] += 1
        
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            responsables_data[resp_id]['ventas_mostrador'] += 1
            responsables_data[resp_id]['monto_ventas_mostrador'] += orden.venta_mostrador.total_venta
        
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            responsables_data[resp_id]['con_cotizacion'] += 1
            if orden.cotizacion.usuario_acepto is True:
                responsables_data[resp_id]['cotizaciones_aceptadas'] += 1
                responsables_data[resp_id]['monto_cotizaciones'] += orden.cotizacion.costo_total_final
            elif orden.cotizacion.usuario_acepto is False:
                responsables_data[resp_id]['cotizaciones_rechazadas'] += 1
            else:
                responsables_data[resp_id]['cotizaciones_pendientes'] += 1
        
        responsables_data[resp_id]['dias_acumulados'] += orden.dias_habiles_en_servicio
    
    # Calcular promedios y montos totales por responsable
    for resp_id, data in responsables_data.items():
        if data['total_ordenes'] > 0:
            data['tiempo_promedio'] = round(data['dias_acumulados'] / data['total_ordenes'], 1)
            data['tasa_entrega'] = round(
                (data['ordenes_entregadas'] / data['total_ordenes']) * 100,
                1
            )
        else:
            data['tiempo_promedio'] = 0
            data['tasa_entrega'] = 0
        
        data['monto_total'] = float(data['monto_ventas_mostrador'] + data['monto_cotizaciones'])
        data['monto_ventas_mostrador'] = float(data['monto_ventas_mostrador'])
        data['monto_cotizaciones'] = float(data['monto_cotizaciones'])
    
    # Convertir a lista y ordenar por total de órdenes (descendente)
    responsables_lista = sorted(
        responsables_data.values(),
        key=lambda x: x['total_ordenes'],
        reverse=True
    )
    
    return responsables_lista


def calcular_top_productos(ordenes, limite=10):
    """
    Calcula el top de productos más vendidos en ventas mostrador
    
    EXPLICACIÓN: Analiza todas las ventas mostrador y cuenta qué productos/servicios
    se vendieron más veces y generaron más ingresos.
    
    Args:
        ordenes: QuerySet de OrdenServicio
        limite: Número máximo de productos a retornar (default: 10)
        
    Returns:
        list: Lista de diccionarios con {descripcion, cantidad, monto}
    """
    productos = {}
    
    for orden in ordenes:
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            venta = orden.venta_mostrador
            
            # Incluir paquetes
            if venta.paquete != 'ninguno':
                desc_paquete = f"Paquete {venta.paquete.upper()}"
                costo_paq = float(venta.costo_paquete)
                
                if desc_paquete not in productos:
                    productos[desc_paquete] = {'cantidad': 0, 'monto': 0}
                
                productos[desc_paquete]['cantidad'] += 1
                productos[desc_paquete]['monto'] += costo_paq
            
            # Incluir servicios
            if venta.incluye_cambio_pieza and venta.costo_cambio_pieza > 0:
                desc = 'Cambio de Pieza'
                if desc not in productos:
                    productos[desc] = {'cantidad': 0, 'monto': 0}
                productos[desc]['cantidad'] += 1
                productos[desc]['monto'] += float(venta.costo_cambio_pieza)
            
            if venta.incluye_limpieza and venta.costo_limpieza > 0:
                desc = 'Limpieza y Mantenimiento'
                if desc not in productos:
                    productos[desc] = {'cantidad': 0, 'monto': 0}
                productos[desc]['cantidad'] += 1
                productos[desc]['monto'] += float(venta.costo_limpieza)
            
            if venta.incluye_kit_limpieza and venta.costo_kit > 0:
                desc = 'Kit de Limpieza'
                if desc not in productos:
                    productos[desc] = {'cantidad': 0, 'monto': 0}
                productos[desc]['cantidad'] += 1
                productos[desc]['monto'] += float(venta.costo_kit)
            
            if venta.incluye_reinstalacion_so and venta.costo_reinstalacion > 0:
                desc = 'Reinstalación SO'
                if desc not in productos:
                    productos[desc] = {'cantidad': 0, 'monto': 0}
                productos[desc]['cantidad'] += 1
                productos[desc]['monto'] += float(venta.costo_reinstalacion)
            
            # Incluir piezas individuales
            piezas = venta.piezas_vendidas.all()
            for pieza in piezas:
                desc = pieza.descripcion_pieza[:50]
                
                if desc not in productos:
                    productos[desc] = {'cantidad': 0, 'monto': 0}
                
                productos[desc]['cantidad'] += pieza.cantidad
                productos[desc]['monto'] += float(pieza.subtotal)
    
    # Ordenar por cantidad vendida (descendente)
    productos_ordenados = sorted(
        [{'descripcion': k, 'cantidad': v['cantidad'], 'monto': v['monto']} 
         for k, v in productos.items()],
        key=lambda x: x['cantidad'],
        reverse=True
    )
    
    return productos_ordenados[:limite]


def calcular_estadisticas_por_sucursal(ordenes):
    """
    Calcula estadísticas por sucursal
    
    EXPLICACIÓN: Agrupa órdenes por sucursal y calcula totales, montos, etc.
    
    Args:
        ordenes: QuerySet de OrdenServicio
        
    Returns:
        list: Lista de diccionarios con estadísticas por sucursal
    """
    sucursales_data = {}
    
    for orden in ordenes:
        suc_id = orden.sucursal.id
        suc_nombre = orden.sucursal.nombre
        
        if suc_id not in sucursales_data:
            sucursales_data[suc_id] = {
                'nombre': suc_nombre,
                'total_ordenes': 0,
                'ventas_mostrador': 0,
                'cotizaciones': 0,
                'monto_total': Decimal('0.00'),
            }
        
        sucursales_data[suc_id]['total_ordenes'] += 1
        
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            sucursales_data[suc_id]['ventas_mostrador'] += 1
            sucursales_data[suc_id]['monto_total'] += orden.venta_mostrador.total_venta
        
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            sucursales_data[suc_id]['cotizaciones'] += 1
            if orden.cotizacion.usuario_acepto:
                sucursales_data[suc_id]['monto_total'] += orden.cotizacion.costo_total_final
    
    # Convertir montos a float
    for data in sucursales_data.values():
        data['monto_total'] = float(data['monto_total'])
    
    # Convertir a lista y ordenar
    sucursales_lista = sorted(
        sucursales_data.values(),
        key=lambda x: x['total_ordenes'],
        reverse=True
    )
    
    return sucursales_lista
