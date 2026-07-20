"""
Dashboard de seguimiento OOW-/FL- + export Excel (Fase 8).

urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
"""

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from inventario.models import Empleado, Sucursal

from .decorators import cache_page_dashboard, permission_required_with_message
from .models import OrdenServicio
from .services.ventas_mostrador_analytics import (
    determinar_categoria_venta,
    obtener_top_productos_vendidos,
)


# ============================================================================
# DASHBOARD DE SEGUIMIENTO ESPECIALIZADO OOW-/FL-
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@cache_page_dashboard
def dashboard_seguimiento_oow_fl(request):
    """
    Dashboard especializado para seguimiento de órdenes con prefijo OOW- y FL-.

    Returns:
        HttpResponse: Renderiza el template con todo el contexto de datos
    """
    from django.db.models import Q, Count, Sum, Avg, F, When, Case, Value, CharField
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    from datetime import timedelta
    from .utils_rhitso import (
        calcular_dias_habiles,
        calcular_dias_por_estatus,
        calcular_promedio_dias_por_estatus,
        agrupar_ordenes_por_mes
    )
    
    # =========================================================================
    # PASO 1: OBTENER FILTROS DE LA URL
    # =========================================================================
    
    responsable_id = request.GET.get('responsable_id', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado_filtro = request.GET.get('estado', '')
    sucursal_id = request.GET.get('sucursal_id', '')
    prefijo_filtro = request.GET.get('prefijo', 'ambos')  # 'OOW', 'FL', o 'ambos'
    
    # =========================================================================
    # PASO 2: CONSTRUIR QUERY BASE (FILTRO PRINCIPAL POR PREFIJO)
    # =========================================================================
    
    # Query base: órdenes con prefijo OOW- o FL- en orden_cliente
    if prefijo_filtro == 'OOW':
        ordenes = OrdenServicio.objects.filter(
            detalle_equipo__orden_cliente__istartswith='OOW-'
        )
    elif prefijo_filtro == 'FL':
        ordenes = OrdenServicio.objects.filter(
            detalle_equipo__orden_cliente__istartswith='FL-'
        )
    else:  # 'ambos' (default)
        ordenes = OrdenServicio.objects.filter(
            Q(detalle_equipo__orden_cliente__istartswith='OOW-') |
            Q(detalle_equipo__orden_cliente__istartswith='FL-')
        )
    
    # Optimizar consultas con select_related y prefetch_related
    ordenes = ordenes.select_related(
        'detalle_equipo',
        'sucursal',
        'responsable_seguimiento',
        'tecnico_asignado_actual',
        'venta_mostrador',
        'cotizacion'
    ).prefetch_related(
        'historial'
    )
    
    # =========================================================================
    # PASO 3: APLICAR FILTROS ADICIONALES
    # =========================================================================
    
    if responsable_id == 'sin_asignar':
        # Filtrar solo órdenes sin responsable asignado
        ordenes = ordenes.filter(responsable_seguimiento__isnull=True)
    elif responsable_id:
        ordenes = ordenes.filter(responsable_seguimiento_id=responsable_id)
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__gte=fecha_desde_obj)
        except ValueError:
            pass  # Ignorar si el formato es inválido
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)
    
    if sucursal_id:
        ordenes = ordenes.filter(sucursal_id=sucursal_id)
    
    # =========================================================================
    # PASO 4: CALCULAR MÉTRICAS GENERALES
    # =========================================================================
    
    total_ordenes = ordenes.count()
    
    # Contar por estado
    ordenes_activas = ordenes.exclude(estado__in=['entregado', 'cancelado']).count()
    ordenes_finalizadas = ordenes.filter(estado='finalizado').count()
    ordenes_entregadas = ordenes.filter(estado='entregado').count()
    
    # Contar ventas mostrador
    total_ventas_mostrador = ordenes.filter(venta_mostrador__isnull=False).count()
    
    # Contar con cotización
    total_con_cotizacion = ordenes.filter(cotizacion__isnull=False).count()
    cotizaciones_aceptadas = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto=True
    ).count()
    cotizaciones_pendientes = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto__isnull=True
    ).count()
    cotizaciones_rechazadas = ordenes.filter(
        cotizacion__isnull=False,
        cotizacion__usuario_acepto=False
    ).count()
    
    # Calcular montos totales
    monto_total_ventas_mostrador = Decimal('0.00')
    monto_total_cotizaciones = Decimal('0.00')
    
    for orden in ordenes:
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            monto_total_ventas_mostrador += orden.venta_mostrador.total_venta
        
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto:
                monto_total_cotizaciones += orden.cotizacion.costo_total_final
    
    monto_total_general = monto_total_ventas_mostrador + monto_total_cotizaciones
    
    # Calcular tiempo promedio (días hábiles)
    total_dias_habiles = 0
    ordenes_con_tiempo = 0
    
    for orden in ordenes:
        dias = orden.dias_habiles_en_servicio
        if dias >= 0:
            total_dias_habiles += dias
            ordenes_con_tiempo += 1
    
    tiempo_promedio = round(total_dias_habiles / ordenes_con_tiempo, 1) if ordenes_con_tiempo > 0 else 0
    
    # Calcular % en tiempo (<= 15 días hábiles)
    ordenes_en_tiempo = sum(1 for orden in ordenes if orden.dias_habiles_en_servicio <= 15)
    porcentaje_en_tiempo = round((ordenes_en_tiempo / total_ordenes) * 100, 1) if total_ordenes > 0 else 0
    
    # Calcular ingreso promedio diario
    if fecha_desde and fecha_hasta:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            dias_rango = calcular_dias_habiles(fecha_desde_obj, fecha_hasta_obj)
            if dias_rango > 0:
                ingreso_promedio_dia = round(total_ordenes / dias_rango, 1)
            else:
                ingreso_promedio_dia = 0
        except:
            ingreso_promedio_dia = 0
    else:
        ingreso_promedio_dia = 0
    
    # =========================================================================
    # PASO 5: AGRUPAR POR RESPONSABLE DE SEGUIMIENTO
    # =========================================================================
    
    responsables_data = {}
    
    for orden in ordenes:
        # EXPLICACIÓN: Algunas órdenes OOW/FL pueden no tener responsable asignado.
        # En ese caso usamos id=0 y nombre "Sin asignar" para no romper el dashboard.
        if orden.responsable_seguimiento:
            resp_id = orden.responsable_seguimiento.id
            resp_nombre = orden.responsable_seguimiento.nombre_completo
        else:
            resp_id = 0
            resp_nombre = "Sin asignar"
        
        if resp_id not in responsables_data:
            responsables_data[resp_id] = {
                'id': resp_id,
                'nombre': resp_nombre,
                'total_ordenes': 0,
                'ordenes_activas': 0,
                'ordenes_finalizadas': 0,
                'ordenes_entregadas': 0,
                'ventas_mostrador': 0,
                'con_cotizacion': 0,
                'cotizaciones_aceptadas': 0,
                'cotizaciones_pendientes': 0,
                'cotizaciones_rechazadas': 0,
                'monto_ventas_mostrador': Decimal('0.00'),
                'monto_cotizaciones': Decimal('0.00'),
                'dias_acumulados': 0,
            }
        
        # Acumular estadísticas
        responsables_data[resp_id]['total_ordenes'] += 1
        
        if orden.estado not in ['entregado', 'cancelado']:
            responsables_data[resp_id]['ordenes_activas'] += 1
        
        if orden.estado == 'finalizado':
            responsables_data[resp_id]['ordenes_finalizadas'] += 1
        
        if orden.estado == 'entregado':
            responsables_data[resp_id]['ordenes_entregadas'] += 1
        
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            responsables_data[resp_id]['ventas_mostrador'] += 1
            responsables_data[resp_id]['monto_ventas_mostrador'] += orden.venta_mostrador.total_venta
        
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            responsables_data[resp_id]['con_cotizacion'] += 1
            if orden.cotizacion.usuario_acepto is True:
                responsables_data[resp_id]['cotizaciones_aceptadas'] += 1
                responsables_data[resp_id]['monto_cotizaciones'] += orden.cotizacion.costo_total_final
            elif orden.cotizacion.usuario_acepto is False:
                responsables_data[resp_id]['cotizaciones_rechazadas'] += 1
            else:
                # Si es None (null), está pendiente
                responsables_data[resp_id]['cotizaciones_pendientes'] += 1
        
        responsables_data[resp_id]['dias_acumulados'] += orden.dias_habiles_en_servicio
    
    # Calcular promedios y montos totales por responsable
    for resp_id, data in responsables_data.items():
        if data['total_ordenes'] > 0:
            data['tiempo_promedio'] = round(data['dias_acumulados'] / data['total_ordenes'], 1)
            # Tasa de finalización = porcentaje de órdenes ENTREGADAS
            data['tasa_finalizacion'] = round(
                (data['ordenes_entregadas'] / data['total_ordenes']) * 100,
                1
            )
        else:
            data['tiempo_promedio'] = 0
            data['tasa_finalizacion'] = 0
        
        data['monto_total'] = data['monto_ventas_mostrador'] + data['monto_cotizaciones']
    
    # Convertir a lista y ordenar por total de órdenes (descendente)
    responsables_lista = sorted(
        responsables_data.values(),
        key=lambda x: x['total_ordenes'],
        reverse=True
    )
    
    # =========================================================================
    # PASO 5.5: PREPARAR GRÁFICO 1 - VENTAS MOSTRADOR POR RESPONSABLE + CATEGORÍA
    # =========================================================================
    
    # Crear estructura con datos de ventas mostrador por responsable
    # Incluyendo la categoría de producto vendido
    grafico_ventas_mostrador_responsables = {
        'labels': [],  # Nombres de responsables
        'data': [],  # Montos en $
        'categorias': [],  # Categoría de producto vendido
        'iconos': [],  # Emojis para visual
        'desglose': [],  # NUEVO: Desglose completo de productos por responsable
    }
    
    # Filtrar responsables que tengan ventas mostrador > 0
    responsables_con_ventas = [
        r for r in responsables_lista 
        if r['ventas_mostrador'] > 0
    ]
    
    # Para cada responsable con ventas, obtener categoría de producto
    for responsable in responsables_con_ventas:
        resp_id = responsable['id']
        
        # Obtener todas las órdenes de este responsable con venta mostrador
        # EXPLICACIÓN: Mismo manejo seguro — si no tiene responsable, su id es 0
        ordenes_resp = [
            o for o in ordenes
            if (o.responsable_seguimiento.id if o.responsable_seguimiento else 0) == resp_id
        ]
        
        # NUEVO: Crear desglose detallado de productos vendidos por este responsable
        productos_responsable = {}  # {descripcion: {cantidad, subtotal, categoria}}
        categorias_contador = {}
        
        for orden in ordenes_resp:
            if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
                venta = orden.venta_mostrador
                
                # ===== INCLUIR PAQUETES =====
                if venta.paquete != 'ninguno':
                    desc_paquete = f"Paquete {venta.paquete.upper()}"
                    costo_paq = float(venta.costo_paquete)
                    
                    if desc_paquete not in productos_responsable:
                        productos_responsable[desc_paquete] = {
                            'cantidad': 0,
                            'subtotal': 0
                        }
                    
                    productos_responsable[desc_paquete]['cantidad'] += 1
                    productos_responsable[desc_paquete]['subtotal'] += costo_paq
                
                # ===== INCLUIR SERVICIOS =====
                servicios_venta = []
                
                if venta.incluye_cambio_pieza and venta.costo_cambio_pieza > 0:
                    servicios_venta.append({
                        'nombre': 'Cambio de Pieza',
                        'costo': float(venta.costo_cambio_pieza)
                    })
                
                if venta.incluye_limpieza and venta.costo_limpieza > 0:
                    servicios_venta.append({
                        'nombre': 'Limpieza y Mantenimiento',
                        'costo': float(venta.costo_limpieza)
                    })
                
                if venta.incluye_kit_limpieza and venta.costo_kit > 0:
                    servicios_venta.append({
                        'nombre': 'Kit de Limpieza',
                        'costo': float(venta.costo_kit)
                    })
                
                if venta.incluye_reinstalacion_so and venta.costo_reinstalacion > 0:
                    servicios_venta.append({
                        'nombre': 'Reinstalación SO',
                        'costo': float(venta.costo_reinstalacion)
                    })
                
                # Agregar servicios al desglose
                for servicio in servicios_venta:
                    desc_servicio = servicio['nombre']
                    costo_servicio = servicio['costo']
                    
                    if desc_servicio not in productos_responsable:
                        productos_responsable[desc_servicio] = {
                            'cantidad': 0,
                            'subtotal': 0
                        }
                    
                    productos_responsable[desc_servicio]['cantidad'] += 1
                    productos_responsable[desc_servicio]['subtotal'] += costo_servicio
                
                # ===== INCLUIR PIEZAS INDIVIDUALES =====
                piezas = venta.piezas_vendidas.all()
                for pieza in piezas:
                    desc = pieza.descripcion_pieza[:50]  # Truncar descripción
                    
                    if desc not in productos_responsable:
                        productos_responsable[desc] = {
                            'cantidad': 0,
                            'subtotal': 0
                        }
                    
                    productos_responsable[desc]['cantidad'] += pieza.cantidad
                    productos_responsable[desc]['subtotal'] += float(pieza.subtotal)
                
                # Contar categorías
                cat_info = determinar_categoria_venta(venta)
                categoria = cat_info['categoria']
                categorias_contador[categoria] = categorias_contador.get(categoria, 0) + 1
        
        # Obtener la categoría más vendida
        if categorias_contador:
            categoria_principal = max(categorias_contador, key=categorias_contador.get)
            # Obtener información de la primera venta para obtener icono
            primera_venta = None
            for orden in ordenes_resp:
                if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
                    primera_venta = orden.venta_mostrador
                    break
            
            cat_info = determinar_categoria_venta(primera_venta) if primera_venta else {'categoria': 'Desconocido', 'icono': '❓'}
        else:
            cat_info = {'categoria': 'Sin categoría', 'icono': '❌'}
        
        # NUEVO: Ordenar productos por subtotal (mayor primero)
        productos_ordenados = sorted(
            productos_responsable.items(),
            key=lambda x: x[1]['subtotal'],
            reverse=True
        )
        
        # Convertir a formato para tooltip
        desglose_texto = []
        for desc, info in productos_ordenados[:10]:  # Top 10 productos
            desglose_texto.append({
                'descripcion': desc,
                'cantidad': info['cantidad'],
                'subtotal': info['subtotal']
            })
        
        # Agregar datos al gráfico
        grafico_ventas_mostrador_responsables['labels'].append(responsable['nombre'])
        grafico_ventas_mostrador_responsables['data'].append(float(responsable['monto_ventas_mostrador']))
        grafico_ventas_mostrador_responsables['categorias'].append(cat_info['categoria'])
        grafico_ventas_mostrador_responsables['iconos'].append(cat_info['icono'])
        grafico_ventas_mostrador_responsables['desglose'].append(desglose_texto)
    
    # =========================================================================
    # PASO 5.6: PREPARAR GRÁFICO 2 - TOP PRODUCTOS VENDIDOS
    # =========================================================================
    
    top_productos = obtener_top_productos_vendidos(ordenes, limite=5)
    
    grafico_top_productos = {
        'labels': [p['descripcion'][:30] for p in top_productos],  # Truncar descripciones largas
        'data': [int(p['cantidad']) for p in top_productos],
        'montos': [float(p['subtotal']) for p in top_productos],
    }
    
    # =========================================================================
    # PASO 6: CALCULAR DÍAS PROMEDIO POR ESTATUS
    # =========================================================================
    
    # Obtener estadísticas separadas: estados de proceso vs estados finales
    resultado_dias_por_estatus = calcular_promedio_dias_por_estatus(ordenes)
    
    # Estados de proceso (sin entregado/cancelado) - para la tabla principal
    dias_por_estatus_proceso = resultado_dias_por_estatus['estados_proceso']
    
    # Estados finales (entregado/cancelado) - para mostrar por separado
    dias_por_estatus_finales = resultado_dias_por_estatus['estados_finales']
    
    # =========================================================================
    # PASO 7: GENERAR DATOS MENSUALES
    # =========================================================================
    
    datos_mensuales = agrupar_ordenes_por_mes(ordenes)
    
    # =========================================================================
    # PASO 8: IDENTIFICAR ALERTAS
    # =========================================================================
    
    alertas = {
        'retrasadas': [],  # >15 días hábiles
        'sin_actualizacion': [],  # >5 días sin cambio de estado
        'cotizaciones_pendientes': [],  # >7 días sin respuesta
        'en_reparacion_larga': [],  # >10 días en estado 'reparacion'
    }
    
    # Función auxiliar para construir datos de alerta con información completa
    def construir_datos_alerta(orden, dias, tipo_dias='hábiles'):
        """
        Construye un diccionario con toda la información necesaria para mostrar
        una alerta en el dashboard de forma profesional.
        
        Args:
            orden: Instancia de OrdenServicio
            dias: Número de días de la alerta
            tipo_dias: Descripción del tipo de días (hábiles, sin respuesta, etc.)
        
        Returns:
            dict: Diccionario con datos completos de la alerta
        """
        return {
            'orden': orden,
            'dias': dias,
            'tipo_dias': tipo_dias,
            # Información adicional para tablas mejoradas
            'orden_cliente': orden.detalle_equipo.orden_cliente if orden.detalle_equipo else 'N/A',
            'estado': orden.get_estado_display(),
            'estado_codigo': orden.estado,
            'responsable': orden.responsable_seguimiento.nombre_completo if orden.responsable_seguimiento else 'Sin asignar',
            'modelo': orden.detalle_equipo.modelo if orden.detalle_equipo else 'N/A',
            'gama': orden.detalle_equipo.get_gama_display() if orden.detalle_equipo and hasattr(orden.detalle_equipo, 'get_gama_display') else orden.detalle_equipo.gama if orden.detalle_equipo else 'N/A',
            'es_candidato_rhitso': orden.es_candidato_rhitso,
        }
    
    for orden in ordenes:
        # Retrasadas (>15 días hábiles sin entregar)
        # Excluir estados finales: entregado y cancelado
        if orden.estado not in ['entregado', 'cancelado'] and orden.dias_habiles_en_servicio > 15:
            alertas['retrasadas'].append(
                construir_datos_alerta(orden, orden.dias_habiles_en_servicio, 'días hábiles')
            )
        
        # Sin actualización (>5 días hábiles)
        dias_sin_act = orden.dias_sin_actualizacion_estado
        if dias_sin_act > 5 and orden.estado not in ['entregado', 'cancelado']:
            alertas['sin_actualizacion'].append(
                construir_datos_alerta(orden, dias_sin_act, 'días sin cambio')
            )
        
        # Cotizaciones pendientes (>7 días sin respuesta)
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is None and orden.cotizacion.dias_sin_respuesta > 7:
                alertas['cotizaciones_pendientes'].append(
                    construir_datos_alerta(orden, orden.cotizacion.dias_sin_respuesta, 'días sin respuesta')
                )
        
        # En reparación prolongada (>10 días en estado reparacion)
        if orden.estado == 'reparacion':
            dias_por_estado = calcular_dias_por_estatus(orden)
            dias_reparacion = dias_por_estado.get('reparacion', 0)
            if dias_reparacion > 10:
                alertas['en_reparacion_larga'].append(
                    construir_datos_alerta(orden, dias_reparacion, 'días en reparación')
                )
    
    # Ordenar alertas por días (mayor a menor) para priorizar las más críticas
    for tipo_alerta in alertas:
        alertas[tipo_alerta].sort(key=lambda x: x['dias'], reverse=True)
    
    # =========================================================================
    # PASO 9: PREPARAR DATOS PARA GRÁFICOS (Chart.js)
    # =========================================================================
    
    # Gráfico 1: Órdenes por responsable
    grafico_responsables = {
        'labels': [r['nombre'] for r in responsables_lista],
        'data': [r['total_ordenes'] for r in responsables_lista],
    }
    
    # Gráfico 2: Días promedio por estatus
    # Ordenar estados en el orden lógico del proceso (sin estados finales)
    # Los estados finales (entregado/cancelado) se muestran por separado
    orden_estados = ['espera', 'diagnostico', 'cotizacion', 'reparacion', 'finalizado', 'control_calidad']
    estados_ordenados = []
    dias_ordenados = []
    
    # Buscar cada estado en los datos de proceso
    # Nota: dias_por_estatus_proceso tiene claves formateadas (ej: 'Control Calidad')
    for estado_codigo in orden_estados:
        # Convertir código a nombre formateado para buscar en el diccionario
        nombre_formateado = estado_codigo.replace('_', ' ').title()
        if nombre_formateado in dias_por_estatus_proceso:
            estados_ordenados.append(nombre_formateado)
            dias_ordenados.append(dias_por_estatus_proceso[nombre_formateado]['promedio'])
    
    grafico_dias_estatus = {
        'labels': estados_ordenados,
        'data': dias_ordenados,
    }
    
    # Gráfico 3: Evolución mensual (últimos 6 meses)
    meses_recientes = datos_mensuales[-6:] if len(datos_mensuales) > 6 else datos_mensuales
    
    grafico_evolucion_mensual = {
        'labels': [m['mes'] for m in meses_recientes],
        'data_ordenes': [m['total_ordenes'] for m in meses_recientes],
        'data_finalizadas': [m['ordenes_finalizadas'] for m in meses_recientes],
        'data_entregadas': [m['ordenes_entregadas'] for m in meses_recientes],
    }
    
    # Gráfico 4: Distribución por estado
    estados_distribucion = {}
    for orden in ordenes:
        estado = orden.get_estado_display()
        estados_distribucion[estado] = estados_distribucion.get(estado, 0) + 1
    
    grafico_distribucion_estados = {
        'labels': list(estados_distribucion.keys()),
        'data': list(estados_distribucion.values()),
    }
    
    # =========================================================================
    # PASO 10: OBTENER LISTAS PARA FILTROS
    # =========================================================================
    
    # Lista de responsables para filtro
    lista_responsables = Empleado.objects.filter(
        ordenes_responsable__in=ordenes
    ).distinct().order_by('nombre_completo')
    
    # Lista de sucursales para filtro
    lista_sucursales = Sucursal.objects.filter(
        ordenes_servicio__in=ordenes
    ).distinct().order_by('nombre')
    
    # Lista de estados para filtro (choices del modelo)
    from config.constants import ESTADO_ORDEN_CHOICES
    lista_estados = ESTADO_ORDEN_CHOICES
    
    # =========================================================================
    # PASO 11: PREPARAR CONTEXTO COMPLETO
    # =========================================================================
    
    context = {
        # Filtros actuales
        'filtros': {
            'responsable_id': responsable_id,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado': estado_filtro,
            'sucursal_id': sucursal_id,
            'prefijo': prefijo_filtro,
        },
        
        # Listas para selectores de filtros
        'lista_responsables': lista_responsables,
        'lista_sucursales': lista_sucursales,
        'lista_estados': lista_estados,
        
        # Métricas generales
        'metricas': {
            'total_ordenes': total_ordenes,
            'ordenes_activas': ordenes_activas,
            'ordenes_finalizadas': ordenes_finalizadas,
            'ordenes_entregadas': ordenes_entregadas,
            'total_ventas_mostrador': total_ventas_mostrador,
            'total_con_cotizacion': total_con_cotizacion,
            'cotizaciones_aceptadas': cotizaciones_aceptadas,
            'cotizaciones_pendientes': cotizaciones_pendientes,
            'cotizaciones_rechazadas': cotizaciones_rechazadas,
            'monto_ventas_mostrador': monto_total_ventas_mostrador,
            'monto_cotizaciones': monto_total_cotizaciones,
            'monto_total': monto_total_general,
            'tiempo_promedio': tiempo_promedio,
            'porcentaje_en_tiempo': porcentaje_en_tiempo,
            'ingreso_promedio_dia': ingreso_promedio_dia,
        },
        
        # Datos por responsable
        'responsables': responsables_lista,
        
        # Días por estatus - SEPARADOS: proceso vs finales
        # 'dias_por_estatus' contiene SOLO estados de proceso (sin entregado/cancelado)
        # 'dias_por_estatus_finales' contiene estadísticas de cierre (entregado/cancelado)
        'dias_por_estatus': dias_por_estatus_proceso,
        'dias_por_estatus_finales': dias_por_estatus_finales,
        
        # Datos mensuales
        'datos_mensuales': datos_mensuales,
        'meses_recientes': meses_recientes,
        
        # Alertas
        'alertas': alertas,
        'total_alertas': (
            len(alertas['retrasadas']) +
            len(alertas['sin_actualizacion']) +
            len(alertas['cotizaciones_pendientes']) +
            len(alertas['en_reparacion_larga'])
        ),
        
        # Datos para gráficos
        'grafico_responsables': grafico_responsables,
        'grafico_dias_estatus': grafico_dias_estatus,
        'grafico_evolucion_mensual': grafico_evolucion_mensual,
        'grafico_distribucion_estados': grafico_distribucion_estados,
        
        # NUEVOS GRÁFICOS: Ventas Mostrador
        'grafico_ventas_mostrador_responsables': grafico_ventas_mostrador_responsables,
        'grafico_top_productos': grafico_top_productos,
        
        # Órdenes completas (para tabla detallada)
        'ordenes': ordenes[:100],  # Limitar a 100 para rendimiento inicial
        'total_ordenes_tabla': ordenes.count(),
    }
    
    return render(request, 'servicio_tecnico/dashboard_seguimiento_oow_fl.html', context)


@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def exportar_excel_dashboard_oow_fl(request):
    """
    Exporta el dashboard OOW-/FL- a Excel con múltiples hojas de análisis

    Requiere: openpyxl instalado (pip install openpyxl)
    
    Returns:
        HttpResponse: Archivo Excel para descarga
    """
    from django.http import HttpResponse
    from django.db.models import Q
    from datetime import datetime
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from .excel_exporters import (
            get_header_style, get_title_style, get_kpi_title_style, get_kpi_value_style,
            get_estado_color, apply_cell_style, auto_adjust_column_width,
            calcular_metricas_generales, calcular_distribucion_estados,
            calcular_estadisticas_por_responsable, calcular_top_productos,
            calcular_estadisticas_por_sucursal
        )
    except ImportError as e:
        from django.http import JsonResponse
        return JsonResponse({
            'success': False,
            'error': f'Error al importar librerías: {str(e)}. Asegúrate de tener openpyxl instalado.'
        })
    
    # =========================================================================
    # PASO 1: OBTENER FILTROS Y CONSTRUIR QUERY (IGUAL QUE EL DASHBOARD)
    # =========================================================================
    
    responsable_id = request.GET.get('responsable_id', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    estado_filtro = request.GET.get('estado', '')
    sucursal_id = request.GET.get('sucursal_id', '')
    prefijo_filtro = request.GET.get('prefijo', 'ambos')
    
    # Query base: órdenes con prefijo OOW- o FL-
    if prefijo_filtro == 'OOW':
        ordenes = OrdenServicio.objects.filter(
            detalle_equipo__orden_cliente__istartswith='OOW-'
        )
    elif prefijo_filtro == 'FL':
        ordenes = OrdenServicio.objects.filter(
            detalle_equipo__orden_cliente__istartswith='FL-'
        )
    else:  # 'ambos' (default)
        ordenes = OrdenServicio.objects.filter(
            Q(detalle_equipo__orden_cliente__istartswith='OOW-') |
            Q(detalle_equipo__orden_cliente__istartswith='FL-')
        )
    
    # Optimizar consultas
    ordenes = ordenes.select_related(
        'detalle_equipo',
        'sucursal',
        'responsable_seguimiento',
        'tecnico_asignado_actual',
        'venta_mostrador',
        'cotizacion'
    ).prefetch_related('historial')
    
    # Aplicar filtros adicionales
    if responsable_id == 'sin_asignar':
        ordenes = ordenes.filter(responsable_seguimiento__isnull=True)
    elif responsable_id:
        ordenes = ordenes.filter(responsable_seguimiento_id=responsable_id)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)
    
    if sucursal_id:
        ordenes = ordenes.filter(sucursal_id=sucursal_id)
    
    # Ordenar por fecha de ingreso
    ordenes = ordenes.order_by('-fecha_ingreso')
    
    # =========================================================================
    # PASO 2: CALCULAR MÉTRICAS Y ESTADÍSTICAS
    # =========================================================================
    
    metricas = calcular_metricas_generales(ordenes)
    distribucion_estados = calcular_distribucion_estados(ordenes)
    responsables_stats = calcular_estadisticas_por_responsable(ordenes)
    top_productos = calcular_top_productos(ordenes, limite=10)
    sucursales_stats = calcular_estadisticas_por_sucursal(ordenes)
    
    # =========================================================================
    # PASO 3: CREAR WORKBOOK
    # =========================================================================
    
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja predeterminada
    
    # =========================================================================
    # HOJA 1: RESUMEN GENERAL
    # =========================================================================
    
    ws1 = wb.create_sheet("Resumen General")
    
    # Título principal
    ws1.merge_cells('A1:F1')
    title_cell = ws1['A1']
    
    # Determinar texto de filtros para el título
    filtros_texto = []
    if prefijo_filtro != 'ambos':
        filtros_texto.append(f"Prefijo: {prefijo_filtro}-")
    if responsable_id == 'sin_asignar':
        filtros_texto.append("Responsable: Sin asignar")
    elif responsable_id:
        try:
            resp = Empleado.objects.get(id=responsable_id)
            filtros_texto.append(f"Responsable: {resp.nombre_completo}")
        except:
            pass
    if fecha_desde or fecha_hasta:
        rango = f"Desde: {fecha_desde or 'inicio'} Hasta: {fecha_hasta or 'hoy'}"
        filtros_texto.append(rango)
    
    filtros_str = " | ".join(filtros_texto) if filtros_texto else "Todos los registros"
    
    title_cell.value = f"DASHBOARD OOW-/FL- - {datetime.now().strftime('%d/%m/%Y')} - {filtros_str}"
    apply_cell_style(title_cell, get_title_style())
    ws1.row_dimensions[1].height = 30
    
    # KPIs Principales
    row = 3
    ws1.merge_cells(f'A{row}:F{row}')
    kpi_section = ws1[f'A{row}']
    kpi_section.value = "📊 INDICADORES CLAVE (KPIs)"
    apply_cell_style(kpi_section, get_kpi_title_style())
    
    row += 2
    kpis_data = [
        ('Total de Órdenes OOW-/FL-', metricas['total_ordenes']),
        ('Órdenes Activas', metricas['ordenes_activas']),
        ('Órdenes Entregadas', metricas['ordenes_entregadas']),
        ('Órdenes Finalizadas', metricas['ordenes_finalizadas']),
        ('', ''),  # Separador
        ('Ventas Mostrador', metricas['total_ventas_mostrador']),
        ('Monto Ventas Mostrador', f"${metricas['monto_ventas_mostrador']:,.2f}"),
        ('', ''),  # Separador
        ('Total con Cotización', metricas['total_con_cotizacion']),
        ('Cotizaciones Aceptadas ✅', metricas['cotizaciones_aceptadas']),
        ('Cotizaciones Pendientes ⏳', metricas['cotizaciones_pendientes']),
        ('Cotizaciones Rechazadas ❌', metricas['cotizaciones_rechazadas']),
        ('Monto Cotizaciones', f"${metricas['monto_cotizaciones']:,.2f}"),
        ('', ''),  # Separador
        ('Monto Total Generado', f"${metricas['monto_total']:,.2f}"),
        ('Tiempo Promedio (días hábiles)', metricas['tiempo_promedio']),
        ('% en Tiempo (≤15 días)', f"{metricas['porcentaje_en_tiempo']}%"),
    ]
    
    for kpi_name, kpi_value in kpis_data:
        if kpi_name == '':  # Fila vacía como separador
            row += 1
            continue
            
        ws1[f'A{row}'] = kpi_name
        ws1[f'B{row}'] = kpi_value
        apply_cell_style(ws1[f'A{row}'], get_kpi_title_style())
        apply_cell_style(ws1[f'B{row}'], get_kpi_value_style())
        row += 1
    
    # Distribución por Estado
    row += 2
    ws1.merge_cells(f'A{row}:C{row}')
    dist_section = ws1[f'A{row}']
    dist_section.value = "📈 DISTRIBUCIÓN POR ESTADO"
    apply_cell_style(dist_section, get_kpi_title_style())
    
    row += 1
    ws1[f'A{row}'] = 'Estado'
    ws1[f'B{row}'] = 'Cantidad'
    ws1[f'C{row}'] = '% del Total'
    apply_cell_style(ws1[f'A{row}'], get_header_style())
    apply_cell_style(ws1[f'B{row}'], get_header_style())
    apply_cell_style(ws1[f'C{row}'], get_header_style())
    
    row += 1
    for estado, cantidad in distribucion_estados.items():
        porcentaje = round((cantidad / metricas['total_ordenes'] * 100), 1) if metricas['total_ordenes'] > 0 else 0
        ws1[f'A{row}'] = estado
        ws1[f'B{row}'] = cantidad
        ws1[f'C{row}'] = f"{porcentaje}%"
        row += 1
    
    auto_adjust_column_width(ws1)
    
    # =========================================================================
    # HOJA 2: CONSOLIDADO POR RESPONSABLE
    # =========================================================================
    
    ws2 = wb.create_sheet("Consolidado Responsables")
    
    # Título
    ws2.merge_cells('A1:K1')
    title_cell = ws2['A1']
    title_cell.value = "ANÁLISIS CONSOLIDADO POR RESPONSABLE DE SEGUIMIENTO"
    apply_cell_style(title_cell, get_title_style())
    ws2.row_dimensions[1].height = 25
    
    # Encabezados
    headers_resp = [
        'Responsable', 'Total Órdenes', 'Activas', 'Entregadas',
        'Ventas Mostrador', 'Monto VM', 'Cotizaciones Aceptadas',
        'Monto Cotizaciones', 'Monto Total', 'Tiempo Promedio (días)',
        'Tasa Entrega (%)'
    ]
    for col_num, header in enumerate(headers_resp, 1):
        cell = ws2.cell(row=3, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    
    # Datos por responsable
    row = 4
    for resp in responsables_stats:
        ws2.cell(row=row, column=1).value = resp['nombre']
        ws2.cell(row=row, column=2).value = resp['total_ordenes']
        ws2.cell(row=row, column=3).value = resp['ordenes_activas']
        ws2.cell(row=row, column=4).value = resp['ordenes_entregadas']
        ws2.cell(row=row, column=5).value = resp['ventas_mostrador']
        ws2.cell(row=row, column=6).value = f"${resp['monto_ventas_mostrador']:,.2f}"
        ws2.cell(row=row, column=7).value = resp['cotizaciones_aceptadas']
        ws2.cell(row=row, column=8).value = f"${resp['monto_cotizaciones']:,.2f}"
        ws2.cell(row=row, column=9).value = f"${resp['monto_total']:,.2f}"
        ws2.cell(row=row, column=10).value = resp['tiempo_promedio']
        ws2.cell(row=row, column=11).value = f"{resp['tasa_entrega']}%"
        row += 1
    
    auto_adjust_column_width(ws2)
    
    # =========================================================================
    # HOJA 3: TOP PRODUCTOS VENDIDOS
    # =========================================================================
    
    ws3 = wb.create_sheet("Top Productos")
    
    # Título
    ws3.merge_cells('A1:D1')
    title_cell = ws3['A1']
    title_cell.value = "TOP PRODUCTOS/SERVICIOS MÁS VENDIDOS (VENTAS MOSTRADOR)"
    apply_cell_style(title_cell, get_title_style())
    ws3.row_dimensions[1].height = 25
    
    # Encabezados
    headers_prod = ['#', 'Producto/Servicio', 'Cantidad Vendida', 'Monto Total']
    for col_num, header in enumerate(headers_prod, 1):
        cell = ws3.cell(row=3, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    
    # Datos de productos
    row = 4
    for idx, prod in enumerate(top_productos, 1):
        ws3.cell(row=row, column=1).value = idx
        ws3.cell(row=row, column=2).value = prod['descripcion']
        ws3.cell(row=row, column=3).value = prod['cantidad']
        ws3.cell(row=row, column=4).value = f"${prod['monto']:,.2f}"
        
        # Resaltar top 3
        if idx <= 3:
            color = '28a745' if idx == 1 else 'ffc107' if idx == 2 else '17a2b8'
            ws3.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws3.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
        
        row += 1
    
    auto_adjust_column_width(ws3)
    
    # =========================================================================
    # HOJA 4: ANÁLISIS POR SUCURSAL
    # =========================================================================
    
    ws4 = wb.create_sheet("Por Sucursal")
    
    # Título
    ws4.merge_cells('A1:E1')
    title_cell = ws4['A1']
    title_cell.value = "ANÁLISIS POR SUCURSAL"
    apply_cell_style(title_cell, get_title_style())
    ws4.row_dimensions[1].height = 25
    
    # Encabezados
    headers_suc = ['Sucursal', 'Total Órdenes', 'Ventas Mostrador', 'Cotizaciones', 'Monto Total']
    for col_num, header in enumerate(headers_suc, 1):
        cell = ws4.cell(row=3, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    
    # Datos por sucursal
    row = 4
    for suc in sucursales_stats:
        ws4.cell(row=row, column=1).value = suc['nombre']
        ws4.cell(row=row, column=2).value = suc['total_ordenes']
        ws4.cell(row=row, column=3).value = suc['ventas_mostrador']
        ws4.cell(row=row, column=4).value = suc['cotizaciones']
        ws4.cell(row=row, column=5).value = f"${suc['monto_total']:,.2f}"
        row += 1
    
    auto_adjust_column_width(ws4)
    
    # =========================================================================
    # HOJAS INDIVIDUALES POR RESPONSABLE (CON SEPARACIÓN ACTIVAS/CERRADAS)
    # =========================================================================
    
    for resp_stat in responsables_stats:
        # Crear hoja con nombre del responsable (límite de 31 caracteres para Excel)
        nombre_hoja = resp_stat['nombre'][:28]
        ws_resp = wb.create_sheet(nombre_hoja)
        
        # Título con nombre del responsable
        ws_resp.merge_cells('A1:P1')
        title_cell = ws_resp['A1']
        title_cell.value = f"REPORTE INDIVIDUAL - {resp_stat['nombre']}"
        apply_cell_style(title_cell, get_title_style())
        ws_resp.row_dimensions[1].height = 25
        
        # Resumen de estadísticas personales
        row = 3
        ws_resp[f'A{row}'] = "📊 ESTADÍSTICAS PERSONALES"
        apply_cell_style(ws_resp[f'A{row}'], get_kpi_title_style())
        row += 2
        
        stats_personales = [
            ('Total de Órdenes:', resp_stat['total_ordenes']),
            ('Órdenes Activas:', resp_stat['ordenes_activas']),
            ('Órdenes Entregadas:', resp_stat['ordenes_entregadas']),
            ('Tiempo Promedio:', f"{resp_stat['tiempo_promedio']} días"),
            ('Tasa de Entrega:', f"{resp_stat['tasa_entrega']}%"),
            ('', ''),  # Separador
            ('Ventas Mostrador:', resp_stat['ventas_mostrador']),
            ('Monto Ventas Mostrador:', f"${resp_stat['monto_ventas_mostrador']:,.2f}"),
            ('', ''),  # Separador
            ('Cotizaciones Aceptadas:', resp_stat['cotizaciones_aceptadas']),
            ('Cotizaciones Pendientes:', resp_stat['cotizaciones_pendientes']),
            ('Cotizaciones Rechazadas:', resp_stat['cotizaciones_rechazadas']),
            ('Monto Cotizaciones:', f"${resp_stat['monto_cotizaciones']:,.2f}"),
            ('', ''),  # Separador
            ('MONTO TOTAL GENERADO:', f"${resp_stat['monto_total']:,.2f}"),
        ]
        
        for stat_name, stat_value in stats_personales:
            if stat_name == '':
                row += 1
                continue
            
            ws_resp[f'A{row}'] = stat_name
            ws_resp[f'B{row}'] = stat_value
            apply_cell_style(ws_resp[f'A{row}'], get_kpi_title_style())
            
            # Resaltar el monto total
            if 'MONTO TOTAL' in stat_name:
                ws_resp[f'B{row}'].fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                ws_resp[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
            else:
                apply_cell_style(ws_resp[f'B{row}'], get_kpi_value_style())
            
            row += 1
        
        # Obtener órdenes del responsable.
        # EXPLICACIÓN: Si el id es 0 significa "Sin asignar" — filtramos por NULL en la BD,
        # ya que no existe ningún empleado con id=0.
        if resp_stat['id'] == 0:
            ordenes_resp = ordenes.filter(responsable_seguimiento__isnull=True)
        else:
            ordenes_resp = ordenes.filter(responsable_seguimiento__id=resp_stat['id'])
        
        # ============== SECCIÓN: ÓRDENES ACTIVAS ==============
        row += 2
        ws_resp.merge_cells(f'A{row}:P{row}')
        activas_section = ws_resp[f'A{row}']
        ordenes_activas_resp = ordenes_resp.exclude(estado__in=['entregado', 'cancelado'])
        activas_section.value = f"🔄 ÓRDENES ACTIVAS ({ordenes_activas_resp.count()})"
        activas_section.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
        activas_section.font = Font(bold=True, size=12, color="000000")
        activas_section.alignment = Alignment(horizontal="left", vertical="center")
        
        row += 1
        
        # Encabezados de la tabla de órdenes activas
        headers_orden = [
            'N° Orden Cliente', 'N° de Serie', 'Tipo Equipo', 'Marca',
            'Modelo', 'Estado', 'Días Hábiles', 'Días Sin Actualizar',
            'Tipo de Orden', 'Monto', 'Sucursal', 'Fecha Ingreso',
            'Última Actualización', 'Cotización', 'Observaciones'
        ]
        for col_num, header in enumerate(headers_orden, 1):
            cell = ws_resp.cell(row=row, column=col_num)
            cell.value = header
            apply_cell_style(cell, get_header_style())
        
        row += 1
        
        # Datos de órdenes activas
        for orden in ordenes_activas_resp.order_by('-fecha_ingreso'):
            ws_resp.cell(row=row, column=1).value = orden.detalle_equipo.orden_cliente
            ws_resp.cell(row=row, column=2).value = orden.detalle_equipo.numero_serie if orden.detalle_equipo.numero_serie else 'N/A'
            ws_resp.cell(row=row, column=3).value = orden.detalle_equipo.get_tipo_equipo_display()
            ws_resp.cell(row=row, column=4).value = orden.detalle_equipo.marca
            ws_resp.cell(row=row, column=5).value = orden.detalle_equipo.modelo[:30] if orden.detalle_equipo.modelo else 'N/A'
            ws_resp.cell(row=row, column=6).value = orden.get_estado_display()
            ws_resp.cell(row=row, column=7).value = orden.dias_habiles_en_servicio
            ws_resp.cell(row=row, column=8).value = orden.dias_sin_actualizacion_estado
            
            # Tipo de orden
            tipo_orden = 'Servicio Normal'
            monto = 0
            if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
                tipo_orden = 'Venta Mostrador'
                monto = float(orden.venta_mostrador.total_venta)
            elif hasattr(orden, 'cotizacion') and orden.cotizacion:
                if orden.cotizacion.usuario_acepto:
                    tipo_orden = 'Cotización Aceptada'
                    monto = float(orden.cotizacion.costo_total_final)
                elif orden.cotizacion.usuario_acepto is False:
                    tipo_orden = 'Cotización Rechazada'
                else:
                    tipo_orden = 'Cotización Pendiente'
            
            ws_resp.cell(row=row, column=9).value = tipo_orden
            ws_resp.cell(row=row, column=10).value = f"${monto:,.2f}" if monto > 0 else 'N/A'
            ws_resp.cell(row=row, column=11).value = orden.sucursal.nombre
            ws_resp.cell(row=row, column=12).value = orden.fecha_ingreso.strftime('%d/%m/%Y')
            
            # Última actualización (del historial)
            ultima_act = orden.historial.order_by('-fecha_evento').first()
            ws_resp.cell(row=row, column=13).value = ultima_act.fecha_evento.strftime('%d/%m/%Y') if ultima_act else 'N/A'
            
            # Estado de cotización
            cotiz_estado = 'N/A'
            if hasattr(orden, 'cotizacion') and orden.cotizacion:
                if orden.cotizacion.usuario_acepto is True:
                    cotiz_estado = '✅ Aceptada'
                elif orden.cotizacion.usuario_acepto is False:
                    cotiz_estado = '❌ Rechazada'
                else:
                    cotiz_estado = '⏳ Pendiente'
            ws_resp.cell(row=row, column=14).value = cotiz_estado
            
            # Observaciones/Alertas
            alertas = []
            if orden.dias_habiles_en_servicio > 15:
                alertas.append('⚠️ RETRASADA')
            if orden.dias_sin_actualizacion_estado > 5:
                alertas.append(f'🔴 Sin actualizar {orden.dias_sin_actualizacion_estado}d')
            ws_resp.cell(row=row, column=15).value = ' | '.join(alertas) if alertas else 'OK'
            
            # Colorear estado
            color_estado = get_estado_color(orden.estado)
            ws_resp.cell(row=row, column=6).fill = PatternFill(start_color=color_estado, end_color=color_estado, fill_type="solid")
            ws_resp.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF")
            
            # Colorear días si está retrasada
            if orden.dias_habiles_en_servicio > 15:
                ws_resp.cell(row=row, column=7).fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
                ws_resp.cell(row=row, column=7).font = Font(bold=True, color="FFFFFF")
            
            # Resaltar fila completa si es candidato RHITSO (morado claro)
            if orden.es_candidato_rhitso:
                rhitso_color = "ede9fe"  # Morado claro (igual que en dashboard)
                for col in range(1, 16):  # Columnas 1-15
                    cell = ws_resp.cell(row=row, column=col)
                    # Solo aplicar si la celda no tiene ya un color especial (estado, retrasada)
                    if col not in [6, 7] or (col == 7 and orden.dias_habiles_en_servicio <= 15):
                        cell.fill = PatternFill(start_color=rhitso_color, end_color=rhitso_color, fill_type="solid")
            
            row += 1
        
        # ============== SECCIÓN: ÓRDENES CERRADAS (ENTREGADAS) ==============
        row += 2
        ws_resp.merge_cells(f'A{row}:P{row}')
        cerradas_section = ws_resp[f'A{row}']
        ordenes_cerradas_resp = ordenes_resp.filter(estado__in=['entregado', 'cancelado'])
        cerradas_section.value = f"✅ ÓRDENES CERRADAS/ENTREGADAS ({ordenes_cerradas_resp.count()})"
        cerradas_section.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cerradas_section.font = Font(bold=True, size=12, color="FFFFFF")
        cerradas_section.alignment = Alignment(horizontal="left", vertical="center")
        
        row += 1
        
        # Encabezados (mismos que activas)
        for col_num, header in enumerate(headers_orden, 1):
            cell = ws_resp.cell(row=row, column=col_num)
            cell.value = header
            apply_cell_style(cell, get_header_style())
        
        row += 1
        
        # Datos de órdenes cerradas
        for orden in ordenes_cerradas_resp.order_by('-fecha_ingreso'):
            ws_resp.cell(row=row, column=1).value = orden.detalle_equipo.orden_cliente
            ws_resp.cell(row=row, column=2).value = orden.detalle_equipo.numero_serie if orden.detalle_equipo.numero_serie else 'N/A'
            ws_resp.cell(row=row, column=3).value = orden.detalle_equipo.get_tipo_equipo_display()
            ws_resp.cell(row=row, column=4).value = orden.detalle_equipo.marca
            ws_resp.cell(row=row, column=5).value = orden.detalle_equipo.modelo[:30] if orden.detalle_equipo.modelo else 'N/A'
            ws_resp.cell(row=row, column=6).value = orden.get_estado_display()
            ws_resp.cell(row=row, column=7).value = orden.dias_habiles_en_servicio
            ws_resp.cell(row=row, column=8).value = orden.dias_sin_actualizacion_estado
            
            # Tipo de orden y monto
            tipo_orden = 'Servicio Normal'
            monto = 0
            if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
                tipo_orden = 'Venta Mostrador'
                monto = float(orden.venta_mostrador.total_venta)
            elif hasattr(orden, 'cotizacion') and orden.cotizacion:
                if orden.cotizacion.usuario_acepto:
                    tipo_orden = 'Cotización Aceptada'
                    monto = float(orden.cotizacion.costo_total_final)
                elif orden.cotizacion.usuario_acepto is False:
                    tipo_orden = 'Cotización Rechazada'
                else:
                    tipo_orden = 'Cotización Pendiente'
            
            ws_resp.cell(row=row, column=9).value = tipo_orden
            ws_resp.cell(row=row, column=10).value = f"${monto:,.2f}" if monto > 0 else 'N/A'
            ws_resp.cell(row=row, column=11).value = orden.sucursal.nombre
            ws_resp.cell(row=row, column=12).value = orden.fecha_ingreso.strftime('%d/%m/%Y')
            
            ultima_act = orden.historial.order_by('-fecha_evento').first()
            ws_resp.cell(row=row, column=13).value = ultima_act.fecha_evento.strftime('%d/%m/%Y') if ultima_act else 'N/A'
            
            cotiz_estado = 'N/A'
            if hasattr(orden, 'cotizacion') and orden.cotizacion:
                if orden.cotizacion.usuario_acepto is True:
                    cotiz_estado = '✅ Aceptada'
                elif orden.cotizacion.usuario_acepto is False:
                    cotiz_estado = '❌ Rechazada'
                else:
                    cotiz_estado = '⏳ Pendiente'
            ws_resp.cell(row=row, column=14).value = cotiz_estado
            
            # Para órdenes cerradas, solo mostrar si fue cancelada
            ws_resp.cell(row=row, column=15).value = '❌ CANCELADA' if orden.estado == 'cancelado' else 'Completada'
            
            # Colorear estado
            color_estado = get_estado_color(orden.estado)
            ws_resp.cell(row=row, column=6).fill = PatternFill(start_color=color_estado, end_color=color_estado, fill_type="solid")
            ws_resp.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF")
            
            # Resaltar fila completa si es candidato RHITSO (morado claro)
            if orden.es_candidato_rhitso:
                rhitso_color = "ede9fe"  # Morado claro (igual que en dashboard)
                for col in range(1, 16):  # Columnas 1-15
                    cell = ws_resp.cell(row=row, column=col)
                    # Solo aplicar si la celda no tiene ya un color especial (estado)
                    if col != 6:
                        cell.fill = PatternFill(start_color=rhitso_color, end_color=rhitso_color, fill_type="solid")
            
            row += 1
        
        auto_adjust_column_width(ws_resp)
    
    # =========================================================================
    # HOJA FINAL: TODAS LAS ÓRDENES (LISTA MAESTRA COMPLETA)
    # =========================================================================
    
    ws_all = wb.create_sheet("Todas las Órdenes")
    
    # Título
    ws_all.merge_cells('A1:Q1')
    title_cell = ws_all['A1']
    title_cell.value = f"LISTA MAESTRA - TODAS LAS ÓRDENES OOW-/FL- ({ordenes.count()} registros)"
    apply_cell_style(title_cell, get_title_style())
    ws_all.row_dimensions[1].height = 25
    
    # Encabezados completos
    headers_all = [
        'N° Orden Cliente', 'N° de Serie', 'Tipo Equipo', 'Marca',
        'Modelo', 'Estado', 'Responsable Seguimiento', 'Técnico Asignado',
        'Días Hábiles', 'Días Sin Actualizar', 'Tipo de Orden', 'Monto',
        'Sucursal', 'Fecha Ingreso', 'Última Actualización', 'Cotización',
        'Observaciones/Alertas'
    ]
    
    for col_num, header in enumerate(headers_all, 1):
        cell = ws_all.cell(row=3, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    
    # Datos de todas las órdenes
    row = 4
    for orden in ordenes:
        ws_all.cell(row=row, column=1).value = orden.detalle_equipo.orden_cliente
        ws_all.cell(row=row, column=2).value = orden.detalle_equipo.numero_serie if orden.detalle_equipo.numero_serie else 'N/A'
        ws_all.cell(row=row, column=3).value = orden.detalle_equipo.get_tipo_equipo_display()
        ws_all.cell(row=row, column=4).value = orden.detalle_equipo.marca
        ws_all.cell(row=row, column=5).value = orden.detalle_equipo.modelo[:30] if orden.detalle_equipo.modelo else 'N/A'
        ws_all.cell(row=row, column=6).value = orden.get_estado_display()
        ws_all.cell(row=row, column=7).value = orden.responsable_seguimiento.nombre_completo if orden.responsable_seguimiento else 'Sin asignar'
        ws_all.cell(row=row, column=8).value = orden.tecnico_asignado_actual.nombre_completo if orden.tecnico_asignado_actual else 'No asignado'
        ws_all.cell(row=row, column=9).value = orden.dias_habiles_en_servicio
        ws_all.cell(row=row, column=10).value = orden.dias_sin_actualizacion_estado
        
        # Tipo de orden y monto
        tipo_orden = 'Servicio Normal'
        monto = 0
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            tipo_orden = 'Venta Mostrador'
            monto = float(orden.venta_mostrador.total_venta)
        elif hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto:
                tipo_orden = 'Cotización Aceptada'
                monto = float(orden.cotizacion.costo_total_final)
            elif orden.cotizacion.usuario_acepto is False:
                tipo_orden = 'Cotización Rechazada'
            else:
                tipo_orden = 'Cotización Pendiente'
        
        ws_all.cell(row=row, column=11).value = tipo_orden
        ws_all.cell(row=row, column=12).value = f"${monto:,.2f}" if monto > 0 else 'N/A'
        ws_all.cell(row=row, column=13).value = orden.sucursal.nombre
        ws_all.cell(row=row, column=14).value = orden.fecha_ingreso.strftime('%d/%m/%Y %H:%M')
        
        # Última actualización
        ultima_act = orden.historial.order_by('-fecha_evento').first()
        ws_all.cell(row=row, column=15).value = ultima_act.fecha_evento.strftime('%d/%m/%Y %H:%M') if ultima_act else 'N/A'
        
        # Estado de cotización
        cotiz_estado = 'N/A'
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                cotiz_estado = '✅ Aceptada'
            elif orden.cotizacion.usuario_acepto is False:
                cotiz_estado = '❌ Rechazada'
            else:
                cotiz_estado = '⏳ Pendiente'
        ws_all.cell(row=row, column=16).value = cotiz_estado
        
        # Observaciones/Alertas
        alertas = []
        if orden.estado not in ['entregado', 'cancelado']:
            if orden.dias_habiles_en_servicio > 15:
                alertas.append('⚠️ RETRASADA')
            if orden.dias_sin_actualizacion_estado > 5:
                alertas.append(f'🔴 Sin actualizar {orden.dias_sin_actualizacion_estado}d')
        else:
            if orden.estado == 'cancelado':
                alertas.append('❌ CANCELADA')
            else:
                alertas.append('✅ Completada')
        
        ws_all.cell(row=row, column=17).value = ' | '.join(alertas) if alertas else 'OK'
        
        # Colorear estado
        color_estado = get_estado_color(orden.estado)
        ws_all.cell(row=row, column=6).fill = PatternFill(start_color=color_estado, end_color=color_estado, fill_type="solid")
        ws_all.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF")
        
        # Colorear días si está retrasada
        if orden.estado not in ['entregado', 'cancelado'] and orden.dias_habiles_en_servicio > 15:
            ws_all.cell(row=row, column=9).fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            ws_all.cell(row=row, column=9).font = Font(bold=True, color="FFFFFF")
        
        # Resaltar fila completa si es candidato RHITSO (morado claro)
        if orden.es_candidato_rhitso:
            rhitso_color = "ede9fe"  # Morado claro (igual que en dashboard)
            for col in range(1, 18):  # Columnas 1-17
                cell = ws_all.cell(row=row, column=col)
                # Solo aplicar si la celda no tiene ya un color especial (estado en col 6, días en col 9)
                es_celda_estado = (col == 6)
                es_celda_retrasada = (col == 9 and orden.estado not in ['entregado', 'cancelado'] and orden.dias_habiles_en_servicio > 15)
                if not es_celda_estado and not es_celda_retrasada:
                    cell.fill = PatternFill(start_color=rhitso_color, end_color=rhitso_color, fill_type="solid")
        
        row += 1
    
    auto_adjust_column_width(ws_all)
    
    # =========================================================================
    # GENERAR NOMBRE DEL ARCHIVO Y RESPUESTA HTTP
    # =========================================================================
    
    # Generar nombre descriptivo del archivo
    fecha_str = datetime.now().strftime('%Y-%m-%d')
    
    # Determinar si hay filtros específicos
    nombre_archivo_partes = ['Dashboard_OOW_FL']
    
    if prefijo_filtro != 'ambos':
        nombre_archivo_partes.append(f'Prefijo_{prefijo_filtro}')
    
    if responsable_id == 'sin_asignar':
        nombre_archivo_partes.append('Resp_Sin_Asignar')
    elif responsable_id:
        try:
            resp = Empleado.objects.get(id=responsable_id)
            # Limpiar nombre para usar en archivo
            nombre_limpio = resp.nombre_completo.replace(' ', '_')[:20]
            nombre_archivo_partes.append(f'Resp_{nombre_limpio}')
        except:
            pass
    
    if estado_filtro:
        nombre_archivo_partes.append(f'Estado_{estado_filtro}')
    
    if sucursal_id:
        try:
            suc = Sucursal.objects.get(id=sucursal_id)
            nombre_limpio = suc.nombre.replace(' ', '_')[:15]
            nombre_archivo_partes.append(f'Suc_{nombre_limpio}')
        except:
            pass
    
    nombre_archivo_partes.append(fecha_str)
    nombre_archivo = '_'.join(nombre_archivo_partes) + '.xlsx'
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response


