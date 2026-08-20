"""
Tests del export de base de servicios en garantía.

EXPLICACIÓN PARA PRINCIPIANTES:
--------------------------------
El dashboard OOW/FL solo mira es_fuera_garantia=True.
Este reporte es el espejo: es_fuera_garantia=False.
Validamos tres capas:
1) El queryset no mezcla OOW/FL.
2) views.py reexporta las vistas nuevas (humo, sin BD).
3) El Excel no trae hoja Top Productos y solo lista folios de garantía.
"""

from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.urls import resolve, reverse
from openpyxl import load_workbook

from inventario.models import Empleado, Sucursal
from servicio_tecnico import views as st_views
from servicio_tecnico import views_export_garantia
from servicio_tecnico.models import DetalleEquipo, OrdenServicio
from servicio_tecnico.services.export_excel_garantia import (
    HOJAS_FIJAS,
    construir_queryset_export,
    generar_workbook_base_garantia,
    queryset_base_garantia,
)


User = get_user_model()


class QuerysetBaseGarantiaFiltroTest(TestCase):
    """
    Regla de negocio: solo órdenes en garantía.

    Objetivo:
        Confirmar que OOW (diagnóstico fuera de garantía) y FL
        (venta mostrador fuera de garantía) NO entran al queryset.
    """

    def setUp(self):
        """
        Crea 3 órdenes: OOW, FL y garantía.

        Args:
            Ninguno (Django llama setUp solo).

        Efectos secundarios:
            Inserta sucursal, usuario, empleado y 3 órdenes + detalle.
        """
        self.sucursal = Sucursal.objects.create(
            nombre='Sucursal Test Base Garantía',
            ciudad='CDMX',
        )
        self.user = User.objects.create_user(
            username='tecnico_base_garantia',
            password='testpass123',
        )
        self.empleado = Empleado.objects.create(
            nombre_completo='Técnico Base Garantía',
            cargo='Técnico',
            area='Laboratorio',
            email='tecnico.base.garantia@test.local',
            sucursal=self.sucursal,
            user=self.user,
        )
        self._contador = 0

        self.orden_oow = self._crear_orden(
            folio='OOW-GAR-TEST-01',
            tipo_servicio='diagnostico',
            es_fuera_garantia=True,
        )
        self.orden_fl = self._crear_orden(
            folio='FL-GAR-TEST-01',
            tipo_servicio='venta_mostrador',
            es_fuera_garantia=True,
        )
        self.orden_garantia = self._crear_orden(
            folio='SIC-GAR-TEST-01',
            tipo_servicio='diagnostico',
            es_fuera_garantia=False,
        )

    def _crear_orden(
        self,
        *,
        folio: str,
        tipo_servicio: str,
        es_fuera_garantia: bool,
        estado: str = 'espera',
    ) -> OrdenServicio:
        """
        Orden mínima + DetalleEquipo (el Excel lee el folio del detalle).

        Args:
            folio: orden_cliente (OOW-/FL- activa el booleano al guardar).
            tipo_servicio: diagnostico o venta_mostrador.
            es_fuera_garantia: valor inicial; DetalleEquipo.save() puede
                sobreescribirlo si el folio empieza con OOW- o FL-.
            estado: estado de la orden.

        Returns:
            OrdenServicio persistida.

        Efectos secundarios:
            Dos INSERTs (orden + detalle).
        """
        self._contador += 1
        orden = OrdenServicio.objects.create(
            sucursal=self.sucursal,
            tipo_servicio=tipo_servicio,
            estado=estado,
            es_fuera_garantia=es_fuera_garantia,
            tecnico_asignado_actual=self.empleado,
            responsable_seguimiento=self.empleado,
        )
        DetalleEquipo.objects.create(
            orden=orden,
            orden_cliente=folio,
            tipo_equipo='Laptop',
            marca='Dell',
            modelo='Latitude Test',
            numero_serie=f'SN-GAR-{self._contador:03d}',
            email_cliente='cliente.garantia@test.local',
            nombre_cliente='Cliente Garantía',
            falla_principal='No enciende',
            gama='media',
        )
        orden.refresh_from_db()
        return orden

    def test_queryset_solo_en_garantia(self):
        """
        queryset_base_garantia: entra SIC-GAR; no entran OOW ni FL.
        """
        ids = set(queryset_base_garantia().values_list('pk', flat=True))

        self.assertIn(self.orden_garantia.pk, ids)
        self.assertNotIn(self.orden_oow.pk, ids)
        self.assertNotIn(self.orden_fl.pk, ids)
        self.assertEqual(len(ids), 1)

    def test_construir_queryset_respeta_filtro_estado(self):
        """
        Un filtro de estado no debe colar órdenes fuera de garantía.
        """
        ids = set(
            construir_queryset_export(estado='espera').values_list('pk', flat=True)
        )
        self.assertEqual(ids, {self.orden_garantia.pk})


class ExportBaseGarantiaHumoTest(SimpleTestCase):
    """Humo: reexport + resolve, sin tocar BD."""

    def test_helper_y_vistas_en_modulo_propio(self):
        """
        Las vistas viven en views_export_garantia y views.py las reexporta.
        """
        self.assertTrue(callable(queryset_base_garantia))
        self.assertIs(
            st_views.exportar_base_garantia,
            views_export_garantia.exportar_base_garantia,
        )
        self.assertIs(
            st_views.exportar_excel_base_garantia,
            views_export_garantia.exportar_excel_base_garantia,
        )
        self.assertEqual(
            views_export_garantia.exportar_base_garantia.__module__,
            'servicio_tecnico.views_export_garantia',
        )

    def test_urls_resuelven_al_modulo_nuevo(self):
        """reverse/resolve apuntan a las vistas del módulo hermano."""
        pagina = resolve(reverse('servicio_tecnico:exportar_base_garantia'))
        excel = resolve(reverse('servicio_tecnico:exportar_excel_base_garantia'))
        self.assertIs(
            pagina.func,
            views_export_garantia.exportar_base_garantia,
        )
        self.assertIs(
            excel.func,
            views_export_garantia.exportar_excel_base_garantia,
        )


class ExportBaseGarantiaExcelTest(TestCase):
    """Integración ligera: el .xlsx no mezcla OOW y no trae Top Productos."""

    # PaisMiddleware / router pueden tocar alias 'mexico' al autenticar.
    databases = {'default', 'mexico'}

    def setUp(self):
        """Sucursal, usuario con permiso gerencial y órdenes de prueba."""
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.backends.db import SessionStore

        # RequestFactory evita PaisMiddleware (el Client redirige a denegado
        # porque el permiso vive en default y el middleware consulta mexico).
        self.factory = RequestFactory()
        self.SessionStore = SessionStore
        self.FallbackStorage = FallbackStorage

        self.sucursal = Sucursal.objects.create(
            nombre='Sucursal Excel Garantía',
            ciudad='CDMX',
        )
        self.user = User.objects.create_user(
            username='gerencia_base_garantia',
            password='testpass123',
        )
        self.empleado = Empleado.objects.create(
            nombre_completo='Gerente Base Garantía',
            cargo='Gerente',
            area='Administración',
            email='gerencia.base.garantia@test.local',
            sucursal=self.sucursal,
            user=self.user,
            contraseña_configurada=True,
        )
        ct = ContentType.objects.get_for_model(OrdenServicio)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type=ct,
                codename='view_dashboard_gerencial',
            ),
        )
        self.url_excel = reverse('servicio_tecnico:exportar_excel_base_garantia')

        self.orden_garantia = self._crear_orden(
            folio='SIC-EXCEL-GAR-01',
            tipo_servicio='diagnostico',
            es_fuera_garantia=False,
        )
        self.orden_oow = self._crear_orden(
            folio='OOW-EXCEL-GAR-01',
            tipo_servicio='diagnostico',
            es_fuera_garantia=True,
        )

    def _request(self, url: str, vista, query: dict | None = None):
        """
        Arma un GET autenticado y llama la vista (sin middleware de país).

        Args:
            url: path resuelto.
            vista: función vista ya decorada.
            query: query string opcional.

        Returns:
            HttpResponse de la vista.
        """
        request = self.factory.get(url, data=query or {})
        request.user = self.user
        request.session = self.SessionStore()
        request.session.create()
        request._messages = self.FallbackStorage(request)
        return vista(request)

    def _crear_orden(
        self,
        *,
        folio: str,
        tipo_servicio: str,
        es_fuera_garantia: bool,
    ) -> OrdenServicio:
        """Igual que el helper del test de queryset: orden + detalle."""
        orden = OrdenServicio.objects.create(
            sucursal=self.sucursal,
            tipo_servicio=tipo_servicio,
            estado='espera',
            es_fuera_garantia=es_fuera_garantia,
            tecnico_asignado_actual=self.empleado,
            responsable_seguimiento=self.empleado,
        )
        DetalleEquipo.objects.create(
            orden=orden,
            orden_cliente=folio,
            tipo_equipo='Laptop',
            marca='Dell',
            modelo='Latitude Excel',
            numero_serie=f'SN-{folio}',
            email_cliente='excel.garantia@test.local',
            nombre_cliente='Cliente Excel',
            falla_principal='Pantalla',
            gama='media',
        )
        orden.refresh_from_db()
        return orden

    def test_excel_sin_top_productos_y_solo_folio_garantia(self):
        """
        Descarga .xlsx: hojas fijas, sin Top Productos, folio OOW ausente.
        """
        response = self._request(
            self.url_excel,
            views_export_garantia.exportar_excel_base_garantia,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn('Base_Garantia', response['Content-Disposition'])

        wb = load_workbook(BytesIO(response.content))
        self.assertNotIn('Top Productos', wb.sheetnames)
        for hoja in HOJAS_FIJAS:
            self.assertIn(hoja, wb.sheetnames)

        ws = wb['Todas las Órdenes']
        folios = [
            fila[0]
            for fila in ws.iter_rows(min_row=4, max_col=1, values_only=True)
            if fila[0]
        ]
        self.assertIn('SIC-EXCEL-GAR-01', folios)
        self.assertNotIn('OOW-EXCEL-GAR-01', folios)

    def test_workbook_directo_omite_columnas_vm_en_resumen(self):
        """
        El resumen no escribe 'Ventas Mostrador' (KPI que sí trae el OOW).
        """
        ordenes = construir_queryset_export()
        wb = generar_workbook_base_garantia(ordenes, 'Todos los registros')
        ws = wb['Resumen General']
        textos = [
            celda.value
            for fila in ws.iter_rows(min_col=1, max_col=1)
            for celda in fila
            if celda.value
        ]
        self.assertFalse(
            any('Ventas Mostrador' in str(t) for t in textos),
            msg='El Excel de garantía no debe mostrar KPIs de Venta Mostrador',
        )
        self.assertTrue(
            any('en Garantía' in str(t) for t in textos),
        )
