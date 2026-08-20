"""
Export Excel: base de órdenes en garantía (IW).

Objetivo de negocio:
    Descargar el mismo tipo de reporte gerencial que el dashboard OOW/FL,
    pero SOLO de servicios dentro de garantía. No es un dashboard: solo
    arma el .xlsx (sin hoja Top Productos ni KPIs de Venta Mostrador).

EXPLICACIÓN PARA PRINCIPIANTES:
    La fuente de verdad es el booleano ``es_fuera_garantia``.
    - True  → fuera de garantía (OOW diagnóstico / FL venta mostrador).
    - False → en garantía (este módulo).
    No buscamos el texto del folio; el campo ya se sincroniza al guardar
    el detalle del equipo.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

from django.core.exceptions import ObjectDoesNotExist
from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from inventario.models import Empleado, Sucursal
from servicio_tecnico.excel_exporters import (
    apply_cell_style,
    auto_adjust_column_width,
    calcular_distribucion_estados,
    calcular_estadisticas_por_responsable,
    calcular_estadisticas_por_sucursal,
    calcular_metricas_generales,
    get_estado_color,
    get_header_style,
    get_kpi_title_style,
    get_kpi_value_style,
    get_title_style,
)
from servicio_tecnico.models import OrdenServicio


# Hojas fijas (sin Top Productos: eso es 100 % Venta Mostrador / FL).
HOJAS_FIJAS = (
    'Resumen General',
    'Consolidado Responsables',
    'Por Sucursal',
    'Todas las Órdenes',
)

# Encabezados de la tabla corta (hoja por responsable: 15 columnas).
HEADERS_ORDEN_CORTA = [
    'N° Orden Cliente', 'N° de Serie', 'Tipo Equipo', 'Marca',
    'Modelo', 'Estado', 'Días Hábiles', 'Días Sin Actualizar',
    'Tipo de Orden', 'Monto', 'Sucursal', 'Fecha Ingreso',
    'Última Actualización', 'Cotización', 'Observaciones',
]

# Encabezados de la lista maestra (17 columnas, igual que OOW/FL).
HEADERS_ORDEN_MAESTRA = [
    'N° Orden Cliente', 'N° de Serie', 'Tipo Equipo', 'Marca',
    'Modelo', 'Estado', 'Responsable Seguimiento', 'Técnico Asignado',
    'Días Hábiles', 'Días Sin Actualizar', 'Tipo de Orden', 'Monto',
    'Sucursal', 'Fecha Ingreso', 'Última Actualización', 'Cotización',
    'Observaciones/Alertas',
]


def queryset_base_garantia() -> QuerySet:
    """
    Queryset base del export de garantía: solo órdenes dentro de garantía.

    Objetivo de negocio:
        El reporte debe listar órdenes con ``es_fuera_garantia=False``,
        nunca las OOW/FL del dashboard hermano.

    Returns:
        QuerySet de OrdenServicio (sin select_related aún).

    Efectos secundarios:
        Ninguno (solo lectura ORM).
    """
    # EXPLICACIÓN PARA PRINCIPIANTES:
    # Invertimos el filtro del dashboard OOW/FL: allá True, aquí False.
    return OrdenServicio.objects.filter(es_fuera_garantia=False)


def optimizar_queryset_garantia(ordenes: QuerySet) -> QuerySet:
    """
    Precarga relaciones usadas al pintar filas del Excel.

    Args:
        ordenes: QuerySet ya filtrado por garantía (y filtros GET).

    Returns:
        El mismo QuerySet con select_related / prefetch_related.

    Efectos secundarios:
        Ninguno; evita N+1 al recorrer órdenes.
    """
    # detalle_equipo = folio, serie, marca; cotizacion = extras en garantía.
    return ordenes.select_related(
        'detalle_equipo',
        'sucursal',
        'responsable_seguimiento',
        'tecnico_asignado_actual',
        'venta_mostrador',
        'cotizacion',
    ).prefetch_related('historial')


def aplicar_filtros_garantia(
    ordenes: QuerySet,
    *,
    responsable_id: str = '',
    fecha_desde: str = '',
    fecha_hasta: str = '',
    estado: str = '',
    sucursal_id: str = '',
) -> QuerySet:
    """
    Aplica los mismos filtros GET que la página de descarga.

    Args:
        ordenes: QuerySet base (ya con es_fuera_garantia=False).
        responsable_id: PK del empleado, o 'sin_asignar'.
        fecha_desde / fecha_hasta: fechas YYYY-MM-DD sobre fecha_ingreso.
        estado: código de ESTADO_ORDEN_CHOICES (vacío = todos).
        sucursal_id: PK de sucursal.

    Returns:
        QuerySet filtrado y ordenado por ingreso descendente.

    Efectos secundarios:
        Ninguno.
    """
    # 1) Responsable: NULL en BD si eligieron "Sin asignar".
    if responsable_id == 'sin_asignar':
        ordenes = ordenes.filter(responsable_seguimiento__isnull=True)
    elif responsable_id:
        ordenes = ordenes.filter(responsable_seguimiento_id=responsable_id)

    # 2) Rango de ingreso: ignoramos fechas mal escritas (no tumbamos el export).
    if fecha_desde:
        try:
            desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__gte=desde)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            ordenes = ordenes.filter(fecha_ingreso__date__lte=hasta)
        except ValueError:
            pass

    # 3) Estado y sucursal son filtros simples de igualdad.
    if estado:
        ordenes = ordenes.filter(estado=estado)
    if sucursal_id:
        ordenes = ordenes.filter(sucursal_id=sucursal_id)

    return ordenes.order_by('-fecha_ingreso')


def construir_queryset_export(
    *,
    responsable_id: str = '',
    fecha_desde: str = '',
    fecha_hasta: str = '',
    estado: str = '',
    sucursal_id: str = '',
) -> QuerySet:
    """
    Encadena base + filtros + prefetch para la vista y el Excel.

    Args:
        Los mismos filtros GET que ``aplicar_filtros_garantia``.

    Returns:
        QuerySet listo para ``.count()`` o para armar el workbook.

    Efectos secundarios:
        Ninguno.
    """
    # Primero el booleano (índice), luego filtros de la pantalla, al final JOINs.
    ordenes = queryset_base_garantia()
    ordenes = aplicar_filtros_garantia(
        ordenes,
        responsable_id=responsable_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado=estado,
        sucursal_id=sucursal_id,
    )
    return optimizar_queryset_garantia(ordenes)


def texto_filtros_aplicados(
    *,
    responsable_id: str = '',
    fecha_desde: str = '',
    fecha_hasta: str = '',
    estado: str = '',
    sucursal_id: str = '',
) -> str:
    """
    Arma una leyenda humana para el título del Excel.

    Args:
        Filtros GET (strings vacíos = “sin ese filtro”).

    Returns:
        Texto tipo "Responsable: Ana | Desde: 2026-01-01 Hasta: hoy".

    Efectos secundarios:
        Lectura puntual de Empleado/Sucursal si hay PK.
    """
    partes: list[str] = []

    if responsable_id == 'sin_asignar':
        partes.append('Responsable: Sin asignar')
    elif responsable_id:
        try:
            resp = Empleado.objects.get(pk=responsable_id)
            partes.append(f'Responsable: {resp.nombre_completo}')
        except (Empleado.DoesNotExist, ValueError):
            pass

    if fecha_desde or fecha_hasta:
        partes.append(
            f"Desde: {fecha_desde or 'inicio'} Hasta: {fecha_hasta or 'hoy'}"
        )

    if estado:
        partes.append(f'Estado: {estado}')

    if sucursal_id:
        try:
            suc = Sucursal.objects.get(pk=sucursal_id)
            partes.append(f'Sucursal: {suc.nombre}')
        except (Sucursal.DoesNotExist, ValueError):
            pass

    return ' | '.join(partes) if partes else 'Todos los registros'


def nombre_archivo_base_garantia(
    *,
    responsable_id: str = '',
    estado: str = '',
    sucursal_id: str = '',
) -> str:
    """
    Nombre de descarga: Base_Garantia[_Resp_][_Estado_][_Suc_]_YYYY-MM-DD.xlsx

    Args:
        responsable_id / estado / sucursal_id: filtros GET opcionales.

    Returns:
        Nombre de archivo seguro (sin espacios raros).

    Efectos secundarios:
        Lectura puntual de Empleado/Sucursal si hay PK.
    """
    partes = ['Base_Garantia']
    fecha_str = datetime.now().strftime('%Y-%m-%d')

    if responsable_id == 'sin_asignar':
        partes.append('Resp_Sin_Asignar')
    elif responsable_id:
        try:
            resp = Empleado.objects.get(pk=responsable_id)
            limpio = resp.nombre_completo.replace(' ', '_')[:20]
            partes.append(f'Resp_{limpio}')
        except (Empleado.DoesNotExist, ValueError):
            pass

    if estado:
        partes.append(f'Estado_{estado}')

    if sucursal_id:
        try:
            suc = Sucursal.objects.get(pk=sucursal_id)
            limpio = suc.nombre.replace(' ', '_')[:15]
            partes.append(f'Suc_{limpio}')
        except (Sucursal.DoesNotExist, ValueError):
            pass

    partes.append(fecha_str)
    return '_'.join(partes) + '.xlsx'


def generar_workbook_base_garantia(
    ordenes: QuerySet,
    filtros_str: str,
) -> Workbook:
    """
    Construye el .xlsx de garantía (mismas hojas que OOW, sin Top Productos).

    Args:
        ordenes: QuerySet ya filtrado y con prefetch.
        filtros_str: leyenda para el título del resumen.

    Returns:
        Workbook de openpyxl en memoria (aún no se serializa a HTTP).

    Efectos secundarios:
        Ninguno de escritura en BD.
    """
    # Reusamos los calculadores del Excel OOW; al pintar omitimos VM.
    metricas = calcular_metricas_generales(ordenes)
    distribucion = calcular_distribucion_estados(ordenes)
    responsables = calcular_estadisticas_por_responsable(ordenes)
    sucursales = calcular_estadisticas_por_sucursal(ordenes)

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, metricas, distribucion, filtros_str)
    _hoja_consolidado_responsables(wb, responsables)
    _hoja_sucursales(wb, sucursales)
    for resp_stat in responsables:
        _hoja_un_responsable(wb, ordenes, resp_stat)
    _hoja_todas_las_ordenes(wb, ordenes)

    return wb


# ---------------------------------------------------------------------------
# Helpers de fila (evitan copiar 3 veces la misma lógica de celdas)
# ---------------------------------------------------------------------------

def _detalle_equipo(orden: OrdenServicio) -> Any | None:
    """
    Devuelve el DetalleEquipo o None si la orden no tiene ficha.

    Args:
        orden: OrdenServicio (puede no tener OneToOne aún).

    Returns:
        DetalleEquipo o None.

    Efectos secundarios:
        Ninguno.
    """
    try:
        return orden.detalle_equipo
    except ObjectDoesNotExist:
        return None


def _tipo_monto_y_cotizacion(orden: OrdenServicio) -> tuple[str, str, str]:
    """
    Clasifica la orden para las columnas Tipo / Monto / Cotización.

    Args:
        orden: Orden con cotizacion/venta_mostrador posiblemente nulos.

    Returns:
        (tipo_orden, monto_texto, estado_cotizacion).

    Efectos secundarios:
        Ninguno.
    """
    tipo = 'Servicio Normal'
    monto = 0.0
    cotiz = 'N/A'

    # En garantía casi no hay VM; si existiera una, igual la mostramos en la fila.
    if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
        tipo = 'Venta Mostrador'
        monto = float(orden.venta_mostrador.total_venta)
    elif hasattr(orden, 'cotizacion') and orden.cotizacion:
        # Cotización de extras sí es habitual en garantía.
        if orden.cotizacion.usuario_acepto is True:
            tipo = 'Cotización Aceptada'
            monto = float(orden.cotizacion.costo_total_final)
            cotiz = 'Aceptada'
        elif orden.cotizacion.usuario_acepto is False:
            tipo = 'Cotización Rechazada'
            cotiz = 'Rechazada'
        else:
            tipo = 'Cotización Pendiente'
            cotiz = 'Pendiente'
        return tipo, (f'${monto:,.2f}' if monto > 0 else 'N/A'), cotiz

    if hasattr(orden, 'cotizacion') and orden.cotizacion:
        if orden.cotizacion.usuario_acepto is True:
            cotiz = 'Aceptada'
        elif orden.cotizacion.usuario_acepto is False:
            cotiz = 'Rechazada'
        else:
            cotiz = 'Pendiente'

    return tipo, (f'${monto:,.2f}' if monto > 0 else 'N/A'), cotiz


def _texto_alertas(orden: OrdenServicio, *, es_cerrada: bool) -> str:
    """
    Arma el texto de observaciones (retraso / sin actualizar / cierre).

    Args:
        orden: OrdenServicio.
        es_cerrada: True si la fila está en la sección entregado/cancelado.

    Returns:
        Cadena lista para la celda (o 'OK' / 'Completada').

    Efectos secundarios:
        Ninguno.
    """
    if es_cerrada:
        return 'CANCELADA' if orden.estado == 'cancelado' else 'Completada'

    alertas: list[str] = []
    if orden.dias_habiles_en_servicio > 15:
        alertas.append('RETRASADA')
    if orden.dias_sin_actualizacion_estado > 5:
        alertas.append(f'Sin actualizar {orden.dias_sin_actualizacion_estado}d')
    return ' | '.join(alertas) if alertas else 'OK'


def _ultima_actualizacion(orden: OrdenServicio, con_hora: bool) -> str:
    """
    Fecha del último evento de historial, o N/A.

    Args:
        orden: Orden con historial prefetched.
        con_hora: True para la lista maestra (incluye hora).

    Returns:
        String formateado.

    Efectos secundarios:
        Ninguno.
    """
    ultima = orden.historial.order_by('-fecha_evento').first()
    if not ultima:
        return 'N/A'
    fmt = '%d/%m/%Y %H:%M' if con_hora else '%d/%m/%Y'
    return ultima.fecha_evento.strftime(fmt)


def _pintar_estado_y_retraso(
    ws: Worksheet,
    row: int,
    col_estado: int,
    col_dias: int,
    orden: OrdenServicio,
    *,
    pintar_retraso: bool,
) -> None:
    """
    Colorea la celda de estado y, si aplica, los días hábiles en rojo.

    Args:
        ws: Hoja activa.
        row / col_estado / col_dias: coordenadas.
        orden: para leer estado y días.
        pintar_retraso: False en filas ya cerradas.

    Efectos secundarios:
        Mutación de celdas del workbook en memoria.
    """
    color = get_estado_color(orden.estado)
    celda_estado = ws.cell(row=row, column=col_estado)
    celda_estado.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    celda_estado.font = Font(bold=True, color='FFFFFF')

    if pintar_retraso and orden.dias_habiles_en_servicio > 15:
        celda_dias = ws.cell(row=row, column=col_dias)
        celda_dias.fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
        celda_dias.font = Font(bold=True, color='FFFFFF')


def _resaltar_rhitso(
    ws: Worksheet,
    row: int,
    max_col: int,
    columnas_especiales: set[int],
) -> None:
    """
    Pinta de morado claro las celdas que no tienen ya un color de semáforo.

    Args:
        ws / row / max_col: rango de la fila.
        columnas_especiales: columnas que ya tienen fill (estado, retraso).

    Efectos secundarios:
        Mutación de celdas del workbook.
    """
    morado = 'ede9fe'
    for col in range(1, max_col + 1):
        if col in columnas_especiales:
            continue
        ws.cell(row=row, column=col).fill = PatternFill(
            start_color=morado, end_color=morado, fill_type='solid',
        )


def _escribir_fila_corta(
    ws: Worksheet,
    row: int,
    orden: OrdenServicio,
    *,
    es_cerrada: bool,
) -> None:
    """
    Escribe 15 columnas (hoja por responsable).

    Args:
        ws, row, orden: destino y dato.
        es_cerrada: cambia el texto de observaciones.

    Efectos secundarios:
        Escribe celdas en ``ws``.
    """
    detalle = _detalle_equipo(orden)
    tipo, monto, cotiz = _tipo_monto_y_cotizacion(orden)

    ws.cell(row=row, column=1).value = detalle.orden_cliente if detalle else 'N/A'
    ws.cell(row=row, column=2).value = (
        detalle.numero_serie if detalle and detalle.numero_serie else 'N/A'
    )
    ws.cell(row=row, column=3).value = (
        detalle.get_tipo_equipo_display() if detalle else 'N/A'
    )
    ws.cell(row=row, column=4).value = detalle.marca if detalle else 'N/A'
    ws.cell(row=row, column=5).value = (
        (detalle.modelo[:30] if detalle.modelo else 'N/A') if detalle else 'N/A'
    )
    ws.cell(row=row, column=6).value = orden.get_estado_display()
    ws.cell(row=row, column=7).value = orden.dias_habiles_en_servicio
    ws.cell(row=row, column=8).value = orden.dias_sin_actualizacion_estado
    ws.cell(row=row, column=9).value = tipo
    ws.cell(row=row, column=10).value = monto
    ws.cell(row=row, column=11).value = orden.sucursal.nombre
    ws.cell(row=row, column=12).value = orden.fecha_ingreso.strftime('%d/%m/%Y')
    ws.cell(row=row, column=13).value = _ultima_actualizacion(orden, con_hora=False)
    ws.cell(row=row, column=14).value = cotiz
    ws.cell(row=row, column=15).value = _texto_alertas(orden, es_cerrada=es_cerrada)

    _pintar_estado_y_retraso(
        ws, row, 6, 7, orden, pintar_retraso=not es_cerrada,
    )
    especiales = {6}
    if not es_cerrada and orden.dias_habiles_en_servicio > 15:
        especiales.add(7)
    if orden.es_candidato_rhitso:
        _resaltar_rhitso(ws, row, 15, especiales)


def _escribir_fila_maestra(ws: Worksheet, row: int, orden: OrdenServicio) -> None:
    """
    Escribe 17 columnas (lista maestra Todas las Órdenes).

    Args:
        ws, row, orden: destino y dato.

    Efectos secundarios:
        Escribe celdas en ``ws``.
    """
    detalle = _detalle_equipo(orden)
    tipo, monto, cotiz = _tipo_monto_y_cotizacion(orden)
    es_cerrada = orden.estado in ('entregado', 'cancelado')
    responsable = (
        orden.responsable_seguimiento.nombre_completo
        if orden.responsable_seguimiento
        else 'Sin asignar'
    )
    tecnico = (
        orden.tecnico_asignado_actual.nombre_completo
        if orden.tecnico_asignado_actual
        else 'No asignado'
    )

    ws.cell(row=row, column=1).value = detalle.orden_cliente if detalle else 'N/A'
    ws.cell(row=row, column=2).value = (
        detalle.numero_serie if detalle and detalle.numero_serie else 'N/A'
    )
    ws.cell(row=row, column=3).value = (
        detalle.get_tipo_equipo_display() if detalle else 'N/A'
    )
    ws.cell(row=row, column=4).value = detalle.marca if detalle else 'N/A'
    ws.cell(row=row, column=5).value = (
        (detalle.modelo[:30] if detalle.modelo else 'N/A') if detalle else 'N/A'
    )
    ws.cell(row=row, column=6).value = orden.get_estado_display()
    ws.cell(row=row, column=7).value = responsable
    ws.cell(row=row, column=8).value = tecnico
    ws.cell(row=row, column=9).value = orden.dias_habiles_en_servicio
    ws.cell(row=row, column=10).value = orden.dias_sin_actualizacion_estado
    ws.cell(row=row, column=11).value = tipo
    ws.cell(row=row, column=12).value = monto
    ws.cell(row=row, column=13).value = orden.sucursal.nombre
    ws.cell(row=row, column=14).value = orden.fecha_ingreso.strftime('%d/%m/%Y %H:%M')
    ws.cell(row=row, column=15).value = _ultima_actualizacion(orden, con_hora=True)
    ws.cell(row=row, column=16).value = cotiz
    ws.cell(row=row, column=17).value = _texto_alertas(orden, es_cerrada=es_cerrada)

    _pintar_estado_y_retraso(
        ws, row, 6, 9, orden, pintar_retraso=not es_cerrada,
    )
    especiales = {6}
    if not es_cerrada and orden.dias_habiles_en_servicio > 15:
        especiales.add(9)
    if orden.es_candidato_rhitso:
        _resaltar_rhitso(ws, row, 17, especiales)


def _escribir_encabezados(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Pinta una fila de encabezados con el estilo azul del export OOW."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())


def _nombre_hoja_unico(wb: Workbook, nombre: str) -> str:
    """
    Excel limita a 31 caracteres y no admite nombres duplicados.

    Args:
        wb: workbook ya con hojas.
        nombre: nombre deseado (responsable).

    Returns:
        Nombre válido y único.
    """
    # Caracteres prohibidos en nombres de hoja de Excel.
    prohibidos = set(r':\/?*[]')
    limpio = ''.join(c for c in nombre if c not in prohibidos).strip() or 'Responsable'
    base = limpio[:28]
    candidato = base
    i = 2
    while candidato in wb.sheetnames:
        sufijo = f'_{i}'
        candidato = base[: 31 - len(sufijo)] + sufijo
        i += 1
    return candidato


# ---------------------------------------------------------------------------
# Hojas
# ---------------------------------------------------------------------------

def _hoja_resumen(
    wb: Workbook,
    metricas: dict,
    distribucion: dict,
    filtros_str: str,
) -> None:
    """Hoja 1: KPIs (sin Ventas Mostrador) + distribución por estado."""
    ws = wb.create_sheet('Resumen General')
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = (
        f"BASE GARANTÍA - {datetime.now().strftime('%d/%m/%Y')} - {filtros_str}"
    )
    apply_cell_style(title, get_title_style())
    ws.row_dimensions[1].height = 30

    row = 3
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'INDICADORES CLAVE (KPIs)'
    apply_cell_style(ws[f'A{row}'], get_kpi_title_style())

    # Sin filas de VM: en garantía el dinero relevante es la cotización de extras.
    kpis = [
        ('Total de Órdenes en Garantía', metricas['total_ordenes']),
        ('Órdenes Activas', metricas['ordenes_activas']),
        ('Órdenes Entregadas', metricas['ordenes_entregadas']),
        ('Órdenes Finalizadas', metricas['ordenes_finalizadas']),
        ('', ''),
        ('Total con Cotización', metricas['total_con_cotizacion']),
        ('Cotizaciones Aceptadas', metricas['cotizaciones_aceptadas']),
        ('Cotizaciones Pendientes', metricas['cotizaciones_pendientes']),
        ('Cotizaciones Rechazadas', metricas['cotizaciones_rechazadas']),
        ('Monto Cotizaciones', f"${metricas['monto_cotizaciones']:,.2f}"),
        ('', ''),
        ('Tiempo Promedio (días hábiles)', metricas['tiempo_promedio']),
        ('% en Tiempo (≤15 días)', f"{metricas['porcentaje_en_tiempo']}%"),
    ]

    row += 2
    for nombre, valor in kpis:
        if nombre == '':
            row += 1
            continue
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = valor
        apply_cell_style(ws[f'A{row}'], get_kpi_title_style())
        apply_cell_style(ws[f'B{row}'], get_kpi_value_style())
        row += 1

    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'DISTRIBUCIÓN POR ESTADO'
    apply_cell_style(ws[f'A{row}'], get_kpi_title_style())

    row += 1
    ws[f'A{row}'] = 'Estado'
    ws[f'B{row}'] = 'Cantidad'
    ws[f'C{row}'] = '% del Total'
    apply_cell_style(ws[f'A{row}'], get_header_style())
    apply_cell_style(ws[f'B{row}'], get_header_style())
    apply_cell_style(ws[f'C{row}'], get_header_style())

    total = metricas['total_ordenes'] or 0
    row += 1
    for estado, cantidad in distribucion.items():
        pct = round((cantidad / total * 100), 1) if total else 0
        ws[f'A{row}'] = estado
        ws[f'B{row}'] = cantidad
        ws[f'C{row}'] = f'{pct}%'
        row += 1

    auto_adjust_column_width(ws)


def _hoja_consolidado_responsables(wb: Workbook, responsables: list[dict]) -> None:
    """Hoja 2: una fila por responsable, sin columnas de Venta Mostrador."""
    ws = wb.create_sheet('Consolidado Responsables')
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = 'ANÁLISIS CONSOLIDADO POR RESPONSABLE DE SEGUIMIENTO'
    apply_cell_style(title, get_title_style())
    ws.row_dimensions[1].height = 25

    headers = [
        'Responsable', 'Total Órdenes', 'Activas', 'Entregadas',
        'Cotizaciones Aceptadas', 'Monto Cotizaciones',
        'Tiempo Promedio (días)', 'Tasa Entrega (%)',
    ]
    _escribir_encabezados(ws, 3, headers)

    row = 4
    for resp in responsables:
        ws.cell(row=row, column=1).value = resp['nombre']
        ws.cell(row=row, column=2).value = resp['total_ordenes']
        ws.cell(row=row, column=3).value = resp['ordenes_activas']
        ws.cell(row=row, column=4).value = resp['ordenes_entregadas']
        ws.cell(row=row, column=5).value = resp['cotizaciones_aceptadas']
        ws.cell(row=row, column=6).value = f"${resp['monto_cotizaciones']:,.2f}"
        ws.cell(row=row, column=7).value = resp['tiempo_promedio']
        ws.cell(row=row, column=8).value = f"{resp['tasa_entrega']}%"
        row += 1

    auto_adjust_column_width(ws)


def _hoja_sucursales(wb: Workbook, sucursales: list[dict]) -> None:
    """Hoja 3: por sucursal, sin columna de Ventas Mostrador."""
    ws = wb.create_sheet('Por Sucursal')
    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = 'ANÁLISIS POR SUCURSAL'
    apply_cell_style(title, get_title_style())
    ws.row_dimensions[1].height = 25

    _escribir_encabezados(
        ws, 3, ['Sucursal', 'Total Órdenes', 'Cotizaciones', 'Monto Total'],
    )

    row = 4
    for suc in sucursales:
        ws.cell(row=row, column=1).value = suc['nombre']
        ws.cell(row=row, column=2).value = suc['total_ordenes']
        ws.cell(row=row, column=3).value = suc['cotizaciones']
        ws.cell(row=row, column=4).value = f"${suc['monto_total']:,.2f}"
        row += 1

    auto_adjust_column_width(ws)


def _hoja_un_responsable(
    wb: Workbook,
    ordenes: QuerySet,
    resp_stat: dict,
) -> None:
    """Una hoja por responsable: stats (sin VM) + activas + cerradas."""
    ws = wb.create_sheet(_nombre_hoja_unico(wb, resp_stat['nombre']))
    ws.merge_cells('A1:P1')
    title = ws['A1']
    title.value = f"REPORTE INDIVIDUAL - {resp_stat['nombre']}"
    apply_cell_style(title, get_title_style())
    ws.row_dimensions[1].height = 25

    row = 3
    ws[f'A{row}'] = 'ESTADÍSTICAS PERSONALES'
    apply_cell_style(ws[f'A{row}'], get_kpi_title_style())
    row += 2

    stats = [
        ('Total de Órdenes:', resp_stat['total_ordenes']),
        ('Órdenes Activas:', resp_stat['ordenes_activas']),
        ('Órdenes Entregadas:', resp_stat['ordenes_entregadas']),
        ('Tiempo Promedio:', f"{resp_stat['tiempo_promedio']} días"),
        ('Tasa de Entrega:', f"{resp_stat['tasa_entrega']}%"),
        ('', ''),
        ('Cotizaciones Aceptadas:', resp_stat['cotizaciones_aceptadas']),
        ('Cotizaciones Pendientes:', resp_stat['cotizaciones_pendientes']),
        ('Cotizaciones Rechazadas:', resp_stat['cotizaciones_rechazadas']),
        ('Monto Cotizaciones:', f"${resp_stat['monto_cotizaciones']:,.2f}"),
    ]
    for nombre, valor in stats:
        if nombre == '':
            row += 1
            continue
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = valor
        apply_cell_style(ws[f'A{row}'], get_kpi_title_style())
        apply_cell_style(ws[f'B{row}'], get_kpi_value_style())
        row += 1

    # id=0 es el cubo "Sin asignar" que arma excel_exporters (no existe Empleado 0).
    if resp_stat['id'] == 0:
        ordenes_resp = ordenes.filter(responsable_seguimiento__isnull=True)
    else:
        ordenes_resp = ordenes.filter(responsable_seguimiento__id=resp_stat['id'])

    row += 2
    activas = ordenes_resp.exclude(estado__in=['entregado', 'cancelado'])
    ws.merge_cells(f'A{row}:P{row}')
    seccion = ws[f'A{row}']
    seccion.value = f'ÓRDENES ACTIVAS ({activas.count()})'
    seccion.fill = PatternFill(start_color='ffc107', end_color='ffc107', fill_type='solid')
    seccion.font = Font(bold=True, size=12, color='000000')
    seccion.alignment = Alignment(horizontal='left', vertical='center')

    row += 1
    _escribir_encabezados(ws, row, HEADERS_ORDEN_CORTA)
    row += 1
    for orden in activas.order_by('-fecha_ingreso'):
        _escribir_fila_corta(ws, row, orden, es_cerrada=False)
        row += 1

    row += 2
    cerradas = ordenes_resp.filter(estado__in=['entregado', 'cancelado'])
    ws.merge_cells(f'A{row}:P{row}')
    seccion = ws[f'A{row}']
    seccion.value = f'ÓRDENES CERRADAS/ENTREGADAS ({cerradas.count()})'
    seccion.fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    seccion.font = Font(bold=True, size=12, color='FFFFFF')
    seccion.alignment = Alignment(horizontal='left', vertical='center')

    row += 1
    _escribir_encabezados(ws, row, HEADERS_ORDEN_CORTA)
    row += 1
    for orden in cerradas.order_by('-fecha_ingreso'):
        _escribir_fila_corta(ws, row, orden, es_cerrada=True)
        row += 1

    auto_adjust_column_width(ws)


def _hoja_todas_las_ordenes(wb: Workbook, ordenes: QuerySet) -> None:
    """Hoja final: lista maestra (la “base” para Excel / Power BI)."""
    ws = wb.create_sheet('Todas las Órdenes')
    ws.merge_cells('A1:Q1')
    title = ws['A1']
    title.value = (
        f'LISTA MAESTRA - ÓRDENES EN GARANTÍA ({ordenes.count()} registros)'
    )
    apply_cell_style(title, get_title_style())
    ws.row_dimensions[1].height = 25

    _escribir_encabezados(ws, 3, HEADERS_ORDEN_MAESTRA)

    row = 4
    for orden in ordenes:
        _escribir_fila_maestra(ws, row, orden)
        row += 1

    auto_adjust_column_width(ws)
