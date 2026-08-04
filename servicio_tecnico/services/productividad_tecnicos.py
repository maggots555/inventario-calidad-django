"""
Productividad de técnicos: consultas y generación de Excel.

Objetivo de negocio:
    Reportar, por técnico asignado actual, cuántas reparaciones productivas
    (órdenes finalizadas/entregadas con cotización aceptada o VM), qué ventas
    mostrador hicieron y cuántos diagnósticos registraron en un período.

EXPLICACIÓN PARA PRINCIPIANTES:
    Este módulo NO vive en views_dashboard_cotizaciones.py (ya es muy grande).
    La vista HTTP solo lee filtros GET y llama a generar_workbook_…;
    aquí está toda la lógica de QuerySets y openpyxl.
"""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from django.db.models import Exists, OuterRef, Prefetch, Q, QuerySet
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.constants import PAQUETES_CHOICES
from servicio_tecnico.models import Cotizacion, OrdenServicio, PiezaVentaMostrador, VentaMostrador

# Estados que cuentan como "trabajo terminado" (egreso / listo o ya entregado).
ESTADOS_FINALIZADOS = ('finalizado', 'entregado')

# Nombres de las 4 hojas del Excel (orden fijo para tests).
HOJAS_EXCEL = (
    'Resumen por técnico',
    'Detalle reparaciones',
    'Detalle ventas mostrador',
    'Detalle diagnósticos',
)

_PAQUETES_MAP = dict(PAQUETES_CHOICES)


def _datos_cliente_equipo(orden: OrdenServicio) -> tuple[str, str]:
    """
    Obtiene orden_cliente y Service Tag (número de serie) del detalle.

    Args:
        orden: OrdenServicio (idealmente con select_related detalle_equipo).

    Returns:
        Tupla (orden_cliente, service_tag). Si no hay detalle, strings vacíos.
    """
    # EXPLICACIÓN PARA PRINCIPIANTES: en el Excel de negocio se usa el folio
    # que ve el cliente (orden_cliente), no el ORD-interno del sistema.
    # Service Tag = numero_serie del equipo.
    # getattr NO evita DoesNotExist en OneToOne inverso de Django.
    from django.core.exceptions import ObjectDoesNotExist

    try:
        detalle = orden.detalle_equipo
    except ObjectDoesNotExist:
        return '', ''
    return (
        (detalle.orden_cliente or '').strip(),
        (detalle.numero_serie or '').strip(),
    )


def _normalizar_fecha_inicio(fecha: date | datetime | str | None) -> datetime | None:
    """
    Convierte el filtro de inicio a datetime aware (inicio del día).

    Args:
        fecha: string 'YYYY-MM-DD', date, datetime o None.

    Returns:
        datetime timezone-aware o None si no hay filtro.
    """
    if fecha is None or fecha == '':
        return None
    if isinstance(fecha, str):
        dt = datetime.strptime(fecha, '%Y-%m-%d')
        return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
    if isinstance(fecha, date) and not isinstance(fecha, datetime):
        dt = datetime.combine(fecha, time.min)
        return timezone.make_aware(dt)
    if isinstance(fecha, datetime) and timezone.is_naive(fecha):
        return timezone.make_aware(fecha)
    return fecha


def _normalizar_fecha_fin(fecha: date | datetime | str | None) -> datetime | None:
    """
    Convierte el filtro de fin a datetime aware (fin del día).

    Args:
        fecha: string 'YYYY-MM-DD', date, datetime o None.

    Returns:
        datetime timezone-aware o None si no hay filtro.
    """
    if fecha is None or fecha == '':
        return None
    if isinstance(fecha, str):
        dt = datetime.combine(datetime.strptime(fecha, '%Y-%m-%d'), time.max)
        return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
    if isinstance(fecha, date) and not isinstance(fecha, datetime):
        dt = datetime.combine(fecha, time.max)
        return timezone.make_aware(dt)
    if isinstance(fecha, datetime) and timezone.is_naive(fecha):
        return timezone.make_aware(fecha)
    return fecha


def _fecha_a_date(fecha: date | datetime | str | None) -> date | None:
    """
    Extrae solo la parte date para filtrar DateField (fecha_fin_diagnostico).

    Args:
        fecha: valor crudo del filtro GET o None.

    Returns:
        date o None.
    """
    if fecha is None or fecha == '':
        return None
    if isinstance(fecha, str):
        return datetime.strptime(fecha, '%Y-%m-%d').date()
    if isinstance(fecha, datetime):
        return fecha.date()
    if isinstance(fecha, date):
        return fecha
    return None


def _aplicar_filtros_comunes(
    qs: QuerySet[OrdenServicio],
    *,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> QuerySet[OrdenServicio]:
    """
    Aplica filtros opcionales de sucursal, técnico y gama sobre órdenes.

    Args:
        qs: QuerySet base de OrdenServicio.
        sucursal_id: PK de sucursal o None.
        tecnico_id: PK de Empleado (técnico) o None.
        gama: 'alta' | 'media' | 'baja' o None.

    Returns:
        QuerySet filtrado.
    """
    # EXPLICACIÓN PARA PRINCIPIANTES: estos filtros son los mismos del
    # Dashboard de Cotizaciones; así el Excel respeta lo que el usuario eligió.
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if tecnico_id:
        qs = qs.filter(tecnico_asignado_actual_id=tecnico_id)
    if gama:
        qs = qs.filter(detalle_equipo__gama=gama)
    return qs


def queryset_ordenes_finalizadas(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> QuerySet[OrdenServicio]:
    """
    Órdenes en estado finalizado/entregado con fecha_finalizacion en el período.

    Args:
        fecha_inicio / fecha_fin: rango sobre fecha_finalizacion.
        sucursal_id / tecnico_id / gama: filtros opcionales.

    Returns:
        QuerySet de OrdenServicio con select_related útiles.
    """
    # Paso 1: solo egresos (listos o ya entregados al cliente).
    qs = OrdenServicio.objects.filter(
        estado__in=ESTADOS_FINALIZADOS,
        fecha_finalizacion__isnull=False,
    ).select_related(
        'sucursal',
        'tecnico_asignado_actual',
        'detalle_equipo',
    )

    # Paso 2: ventana temporal por fecha de finalización (no por ingreso).
    inicio = _normalizar_fecha_inicio(fecha_inicio)
    fin = _normalizar_fecha_fin(fecha_fin)
    if inicio is not None:
        qs = qs.filter(fecha_finalizacion__gte=inicio)
    if fin is not None:
        qs = qs.filter(fecha_finalizacion__lte=fin)

    return _aplicar_filtros_comunes(
        qs,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama,
    )


def obtener_reparaciones_productivas(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> list[dict[str, Any]]:
    """
    Reparaciones productivas: finalizadas con cotización aceptada o con VM.

    Args:
        fecha_inicio / fecha_fin / sucursal_id / tecnico_id / gama: filtros.

    Returns:
        Lista de dicts listos para Excel (una fila por orden).

    Efectos secundarios:
        Solo lectura ORM; no escribe en BD.
    """
    # Subconsultas: ¿tiene cot aceptada? ¿tiene VM?
    cot_aceptada = Cotizacion.objects.filter(
        orden_id=OuterRef('pk'),
        usuario_acepto=True,
    )
    tiene_vm = VentaMostrador.objects.filter(orden_id=OuterRef('pk'))

    qs = (
        queryset_ordenes_finalizadas(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            sucursal_id=sucursal_id,
            tecnico_id=tecnico_id,
            gama=gama,
        )
        .annotate(
            _cot_aceptada=Exists(cot_aceptada),
            _tiene_vm=Exists(tiene_vm),
        )
        .filter(Q(_cot_aceptada=True) | Q(_tiene_vm=True))
        .select_related('cotizacion', 'venta_mostrador')
        .order_by('tecnico_asignado_actual__nombre_completo', 'fecha_finalizacion')
    )

    filas: list[dict[str, Any]] = []
    for orden in qs:
        # Valor aceptado solo si hay cotización aceptada (propiedad del modelo).
        cot_ok = bool(getattr(orden, '_cot_aceptada', False))
        valor_cot = Decimal('0.00')
        if cot_ok and hasattr(orden, 'cotizacion'):
            try:
                valor_cot = orden.cotizacion.costo_total_final or Decimal('0.00')
            except Cotizacion.DoesNotExist:
                valor_cot = Decimal('0.00')

        folio_vm = ''
        if getattr(orden, '_tiene_vm', False):
            try:
                folio_vm = orden.venta_mostrador.folio_venta
            except VentaMostrador.DoesNotExist:
                folio_vm = ''

        tecnico = orden.tecnico_asignado_actual
        orden_cliente, service_tag = _datos_cliente_equipo(orden)
        filas.append({
            'orden_id': orden.pk,
            # folio = orden del cliente (no el número interno ORD-…)
            'folio': orden_cliente,
            'service_tag': service_tag,
            'tecnico_id': tecnico.pk if tecnico else None,
            'tecnico': tecnico.nombre_completo if tecnico else 'Sin técnico',
            'sucursal': orden.sucursal.nombre if orden.sucursal else '',
            'estado': orden.estado,
            'fecha_finalizacion': orden.fecha_finalizacion,
            'cot_aceptada': cot_ok,
            'tiene_vm': bool(getattr(orden, '_tiene_vm', False)),
            'valor_cot_aceptada': valor_cot,
            'folio_vm': folio_vm,
        })
    return filas


def obtener_ventas_mostrador_productivas(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> list[dict[str, Any]]:
    """
    Ventas mostrador de órdenes finalizadas/entregadas en el período.

    El período se alinea con fecha_finalizacion de la orden (productividad por
    egreso), no con fecha_venta suelta.

    Args:
        fecha_inicio / fecha_fin / sucursal_id / tecnico_id / gama: filtros.

    Returns:
        Lista de dicts con desglose de paquete, servicios y piezas.
    """
    qs = (
        queryset_ordenes_finalizadas(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            sucursal_id=sucursal_id,
            tecnico_id=tecnico_id,
            gama=gama,
        )
        .filter(venta_mostrador__isnull=False)
        .select_related(
            'venta_mostrador',
            'tecnico_asignado_actual',
            'sucursal',
            'detalle_equipo',
        )
        .prefetch_related(
            Prefetch(
                'venta_mostrador__piezas_vendidas',
                queryset=PiezaVentaMostrador.objects.all(),
            ),
        )
        .order_by('tecnico_asignado_actual__nombre_completo', 'fecha_finalizacion')
    )

    filas: list[dict[str, Any]] = []
    for orden in qs:
        try:
            vm = orden.venta_mostrador
        except VentaMostrador.DoesNotExist:
            continue

        # Resumen textual de piezas para una sola celda del Excel.
        piezas = list(vm.piezas_vendidas.all())
        resumen_piezas = '; '.join(
            f'{p.descripcion_pieza} x{p.cantidad}' for p in piezas
        ) if piezas else ''

        tecnico = orden.tecnico_asignado_actual
        orden_cliente, service_tag = _datos_cliente_equipo(orden)
        filas.append({
            'orden_id': orden.pk,
            'folio': orden_cliente,
            'service_tag': service_tag,
            'folio_vm': vm.folio_venta,
            'tecnico_id': tecnico.pk if tecnico else None,
            'tecnico': tecnico.nombre_completo if tecnico else 'Sin técnico',
            'sucursal': orden.sucursal.nombre if orden.sucursal else '',
            'paquete': vm.paquete,
            'paquete_nombre': _PAQUETES_MAP.get(vm.paquete, vm.paquete),
            'costo_paquete': vm.costo_paquete or Decimal('0.00'),
            'incluye_limpieza': vm.incluye_limpieza,
            'costo_limpieza': vm.costo_limpieza or Decimal('0.00'),
            'incluye_reinstalacion': vm.incluye_reinstalacion_so,
            'costo_reinstalacion': vm.costo_reinstalacion or Decimal('0.00'),
            'incluye_respaldo': vm.incluye_respaldo,
            'costo_respaldo': vm.costo_respaldo or Decimal('0.00'),
            'incluye_cambio_pieza': vm.incluye_cambio_pieza,
            'costo_cambio_pieza': vm.costo_cambio_pieza or Decimal('0.00'),
            'incluye_kit': vm.incluye_kit_limpieza,
            'costo_kit': vm.costo_kit or Decimal('0.00'),
            'resumen_piezas': resumen_piezas,
            'total_vm': vm.total_venta,
        })
    return filas


def obtener_diagnosticos_realizados(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> list[dict[str, Any]]:
    """
    Diagnósticos SIC no vacíos atribuidos al técnico asignado actual.

    Fecha usada:
        - Preferente: detalle_equipo.fecha_fin_diagnostico en el rango.
        - Fallback: si fecha_fin_diagnostico es null, se usa fecha_ingreso
          de la orden (mismo criterio documentado en el plan).

    Args:
        fecha_inicio / fecha_fin / sucursal_id / tecnico_id / gama: filtros.

    Returns:
        Lista de dicts para la hoja de diagnósticos.
    """
    qs = OrdenServicio.objects.filter(
        detalle_equipo__isnull=False,
    ).exclude(
        detalle_equipo__diagnostico_sic__isnull=True,
    ).exclude(
        detalle_equipo__diagnostico_sic='',
    ).select_related(
        'sucursal',
        'tecnico_asignado_actual',
        'detalle_equipo',
    )

    qs = _aplicar_filtros_comunes(
        qs,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama,
    )

    # Rango: DateField vs DateTimeField requieren filtros distintos.
    d_inicio = _fecha_a_date(fecha_inicio)
    d_fin = _fecha_a_date(fecha_fin)
    inicio_dt = _normalizar_fecha_inicio(fecha_inicio)
    fin_dt = _normalizar_fecha_fin(fecha_fin)

    # EXPLICACIÓN: (tiene fecha_fin en rango) OR (sin fecha_fin Y ingreso en rango).
    cond_con_fecha = Q(detalle_equipo__fecha_fin_diagnostico__isnull=False)
    cond_sin_fecha = Q(detalle_equipo__fecha_fin_diagnostico__isnull=True)

    if d_inicio is not None:
        cond_con_fecha &= Q(detalle_equipo__fecha_fin_diagnostico__gte=d_inicio)
    if d_fin is not None:
        cond_con_fecha &= Q(detalle_equipo__fecha_fin_diagnostico__lte=d_fin)
    if inicio_dt is not None:
        cond_sin_fecha &= Q(fecha_ingreso__gte=inicio_dt)
    if fin_dt is not None:
        cond_sin_fecha &= Q(fecha_ingreso__lte=fin_dt)

    if d_inicio is not None or d_fin is not None or inicio_dt is not None or fin_dt is not None:
        qs = qs.filter(cond_con_fecha | cond_sin_fecha)

    qs = qs.order_by(
        'tecnico_asignado_actual__nombre_completo',
        'fecha_ingreso',
    )

    filas: list[dict[str, Any]] = []
    for orden in qs:
        detalle = orden.detalle_equipo
        texto = (detalle.diagnostico_sic or '').strip()
        if not texto:
            continue

        # Fecha mostrada: fin diagnóstico o, si falta, ingreso.
        fecha_diag: date | datetime | None = detalle.fecha_fin_diagnostico
        if fecha_diag is None:
            fecha_diag = orden.fecha_ingreso

        tecnico = orden.tecnico_asignado_actual
        extracto = texto if len(texto) <= 200 else texto[:197] + '...'
        orden_cliente, service_tag = _datos_cliente_equipo(orden)
        filas.append({
            'orden_id': orden.pk,
            'folio': orden_cliente,
            'service_tag': service_tag,
            'tecnico_id': tecnico.pk if tecnico else None,
            'tecnico': tecnico.nombre_completo if tecnico else 'Sin técnico',
            'sucursal': orden.sucursal.nombre if orden.sucursal else '',
            'fecha_diagnostico': fecha_diag,
            'longitud_texto': len(texto),
            'extracto': extracto,
        })
    return filas


def agregar_resumen_por_tecnico(
    reparaciones: list[dict[str, Any]],
    ventas: list[dict[str, Any]],
    diagnosticos: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Une las tres métricas en una fila por técnico.

    Incluye desglose de servicios de venta mostrador (limpieza, reinstalación,
    respaldo, cambio de pieza, kit) y paquetes (premium/oro/plata).

    Args:
        reparaciones: salida de obtener_reparaciones_productivas.
        ventas: salida de obtener_ventas_mostrador_productivas.
        diagnosticos: salida de obtener_diagnosticos_realizados.

    Returns:
        Lista ordenada por nombre de técnico con conteos y sumas.
    """
    # Clave: tecnico_id (o nombre si falta id).
    por_tecnico: dict[Any, dict[str, Any]] = {}

    def _bucket(tecnico_id: Any, tecnico: str, sucursal: str) -> dict[str, Any]:
        clave = tecnico_id if tecnico_id is not None else f'nombre:{tecnico}'
        if clave not in por_tecnico:
            # EXPLICACIÓN PARA PRINCIPIANTES: cada técnico arranca en 0;
            # luego sumamos reparaciones, VM, diagnósticos y cada servicio VM.
            por_tecnico[clave] = {
                'tecnico_id': tecnico_id,
                'tecnico': tecnico,
                'sucursales': set(),
                'reparaciones': 0,
                'ventas_mostrador': 0,
                'diagnosticos': 0,
                'valor_cot_aceptada': Decimal('0.00'),
                'ingreso_vm': Decimal('0.00'),
                # Desglose VM — cuántas veces vendió cada servicio/paquete
                'vm_limpieza': 0,
                'vm_reinstalacion': 0,
                'vm_respaldo': 0,
                'vm_cambio_pieza': 0,
                'vm_kit': 0,
                'vm_paquete_premium': 0,
                'vm_paquete_oro': 0,
                'vm_paquete_plata': 0,
            }
        if sucursal:
            por_tecnico[clave]['sucursales'].add(sucursal)
        return por_tecnico[clave]

    for fila in reparaciones:
        b = _bucket(fila['tecnico_id'], fila['tecnico'], fila.get('sucursal', ''))
        b['reparaciones'] += 1
        b['valor_cot_aceptada'] += fila.get('valor_cot_aceptada') or Decimal('0.00')

    for fila in ventas:
        b = _bucket(fila['tecnico_id'], fila['tecnico'], fila.get('sucursal', ''))
        b['ventas_mostrador'] += 1
        b['ingreso_vm'] += fila.get('total_vm') or Decimal('0.00')

        # Contar cada servicio activo en esa VM (una venta puede tener varios).
        if fila.get('incluye_limpieza'):
            b['vm_limpieza'] += 1
        if fila.get('incluye_reinstalacion'):
            b['vm_reinstalacion'] += 1
        if fila.get('incluye_respaldo'):
            b['vm_respaldo'] += 1
        if fila.get('incluye_cambio_pieza'):
            b['vm_cambio_pieza'] += 1
        if fila.get('incluye_kit'):
            b['vm_kit'] += 1

        # Paquetes (solo premium/oro/plata; "ninguno" no suma).
        paquete = fila.get('paquete') or 'ninguno'
        if paquete == 'premium':
            b['vm_paquete_premium'] += 1
        elif paquete == 'oro':
            b['vm_paquete_oro'] += 1
        elif paquete == 'plata':
            b['vm_paquete_plata'] += 1

    for fila in diagnosticos:
        b = _bucket(fila['tecnico_id'], fila['tecnico'], fila.get('sucursal', ''))
        b['diagnosticos'] += 1

    resultado: list[dict[str, Any]] = []
    for b in por_tecnico.values():
        sucursales = sorted(b['sucursales'])
        resultado.append({
            'tecnico': b['tecnico'],
            'tecnico_id': b['tecnico_id'],
            'sucursales': ', '.join(sucursales),
            'reparaciones': b['reparaciones'],
            'ventas_mostrador': b['ventas_mostrador'],
            'diagnosticos': b['diagnosticos'],
            'valor_cot_aceptada': b['valor_cot_aceptada'],
            'ingreso_vm': b['ingreso_vm'],
            'vm_limpieza': b['vm_limpieza'],
            'vm_reinstalacion': b['vm_reinstalacion'],
            'vm_respaldo': b['vm_respaldo'],
            'vm_cambio_pieza': b['vm_cambio_pieza'],
            'vm_kit': b['vm_kit'],
            'vm_paquete_premium': b['vm_paquete_premium'],
            'vm_paquete_oro': b['vm_paquete_oro'],
            'vm_paquete_plata': b['vm_paquete_plata'],
        })

    resultado.sort(key=lambda x: x['tecnico'].lower())
    return resultado


def _estilos_excel() -> dict[str, Any]:
    """Estilos reutilizables al estilo del export de aceptaciones."""
    return {
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill(start_color='198754', end_color='198754', fill_type='solid'),
        'header_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'title_font': Font(bold=True, size=14, color='FFFFFF'),
        'title_fill': PatternFill(start_color='212529', end_color='212529', fill_type='solid'),
        'title_align': Alignment(horizontal='center', vertical='center'),
        'subtitle_font': Font(italic=True, size=10, color='666666'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
        'number_fmt': '#,##0.00',
    }


def _aplicar_header(ws, fila: int, headers: list[str], estilos: dict[str, Any]) -> None:
    """Escribe y estiliza la fila de encabezados."""
    for col, texto in enumerate(headers, 1):
        cell = ws.cell(row=fila, column=col, value=texto)
        cell.font = estilos['header_font']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['header_align']
        cell.border = estilos['border']


def _auto_ajustar(ws, max_width: int = 45) -> None:
    """Ajusta anchos de columna según contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _fmt_fecha(valor: date | datetime | None) -> str:
    """Formato legible para celdas de fecha."""
    if valor is None:
        return ''
    if isinstance(valor, datetime):
        return timezone.localtime(valor).strftime('%d/%m/%Y %H:%M') if timezone.is_aware(valor) else valor.strftime('%d/%m/%Y %H:%M')
    return valor.strftime('%d/%m/%Y')


def _si_no(valor: bool) -> str:
    """Booleano a Sí/No en español."""
    return 'Sí' if valor else 'No'


def generar_workbook_productividad_tecnicos(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> Workbook | None:
    """
    Arma el Excel de productividad con 4 hojas.

    Args:
        fecha_inicio / fecha_fin / sucursal_id / tecnico_id / gama: filtros GET.

    Returns:
        Workbook openpyxl, o None si no hay ninguna fila en las tres métricas.

    Efectos secundarios:
        Solo lectura ORM; el archivo se serializa en la vista HTTP.
    """
    filtros = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'sucursal_id': sucursal_id,
        'tecnico_id': tecnico_id,
        'gama': gama,
    }

    reparaciones = obtener_reparaciones_productivas(**filtros)
    ventas = obtener_ventas_mostrador_productivas(**filtros)
    diagnosticos = obtener_diagnosticos_realizados(**filtros)

    # Sin datos en ninguna métrica → la vista mostrará warning y redirigirá.
    if not reparaciones and not ventas and not diagnosticos:
        return None

    resumen = agregar_resumen_por_tecnico(reparaciones, ventas, diagnosticos)
    estilos = _estilos_excel()

    wb = Workbook()
    wb.remove(wb.active)

    filtros_texto = (
        f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actual'}"
    )
    if sucursal_id:
        filtros_texto += f' | Sucursal ID: {sucursal_id}'
    if tecnico_id:
        filtros_texto += f' | Técnico ID: {tecnico_id}'
    if gama:
        filtros_texto += f' | Gama: {gama}'
    generado = timezone.localtime().strftime('%d/%m/%Y %H:%M:%S')

    # ------------------------------------------------------------------
    # HOJA 1: Resumen por técnico
    # ------------------------------------------------------------------
    # Columnas amplias: métricas base + desglose de servicios/paquetes VM.
    num_cols_resumen = 15
    ultima_col = get_column_letter(num_cols_resumen)

    ws1 = wb.create_sheet(HOJAS_EXCEL[0])
    ws1.merge_cells(f'A1:{ultima_col}1')
    ws1['A1'] = 'PRODUCTIVIDAD TÉCNICOS — RESUMEN'
    ws1['A1'].font = estilos['title_font']
    ws1['A1'].fill = estilos['title_fill']
    ws1['A1'].alignment = estilos['title_align']
    ws1.merge_cells(f'A2:{ultima_col}2')
    ws1['A2'] = filtros_texto
    ws1['A2'].font = estilos['subtitle_font']
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells(f'A3:{ultima_col}3')
    ws1['A3'] = (
        f'Generado: {generado} | '
        'Columnas VM = cuántas ventas incluyeron cada servicio/paquete'
    )
    ws1['A3'].font = estilos['subtitle_font']
    ws1['A3'].alignment = Alignment(horizontal='center')

    headers1 = [
        'Técnico',
        'Sucursal(es)',
        'Reparaciones',
        'Ventas mostrador',
        'Diagnósticos',
        'Valor cotización aceptada',
        'Ingreso VM',
        'Limpieza y mant.',
        'Reinstalación SO',
        'Respaldo info',
        'Cambio de pieza',
        'Kit limpieza',
        'Paquete Premium',
        'Paquete Oro',
        'Paquete Plata',
    ]
    _aplicar_header(ws1, 5, headers1, estilos)
    fila = 6
    for row in resumen:
        # Columnas 1–5: identidad y conteos principales
        valores_base = [
            row['tecnico'],
            row['sucursales'],
            row['reparaciones'],
            row['ventas_mostrador'],
            row['diagnosticos'],
        ]
        for col, val in enumerate(valores_base, 1):
            ws1.cell(row=fila, column=col, value=val).border = estilos['border']

        # Columnas 6–7: montos
        c_val = ws1.cell(row=fila, column=6, value=float(row['valor_cot_aceptada']))
        c_val.number_format = estilos['number_fmt']
        c_val.border = estilos['border']
        c_vm = ws1.cell(row=fila, column=7, value=float(row['ingreso_vm']))
        c_vm.number_format = estilos['number_fmt']
        c_vm.border = estilos['border']

        # Columnas 8–15: desglose servicios y paquetes VM
        desglose_vm = [
            row['vm_limpieza'],
            row['vm_reinstalacion'],
            row['vm_respaldo'],
            row['vm_cambio_pieza'],
            row['vm_kit'],
            row['vm_paquete_premium'],
            row['vm_paquete_oro'],
            row['vm_paquete_plata'],
        ]
        for col, val in enumerate(desglose_vm, 8):
            ws1.cell(row=fila, column=col, value=val).border = estilos['border']
        fila += 1
    _auto_ajustar(ws1)

    # ------------------------------------------------------------------
    # HOJA 2: Detalle reparaciones
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet(HOJAS_EXCEL[1])
    ws2.merge_cells('A1:J1')
    ws2['A1'] = 'DETALLE REPARACIONES PRODUCTIVAS'
    ws2['A1'].font = estilos['title_font']
    ws2['A1'].fill = estilos['title_fill']
    ws2['A1'].alignment = estilos['title_align']
    ws2.merge_cells('A2:J2')
    ws2['A2'] = filtros_texto
    ws2['A2'].font = estilos['subtitle_font']

    headers2 = [
        'Orden cliente',
        'Service Tag',
        'Técnico',
        'Sucursal',
        'Estado',
        'Fecha finalización',
        '¿Cot aceptada?',
        '¿Tiene VM?',
        'Valor aceptado cot',
        'Folio VM',
    ]
    _aplicar_header(ws2, 4, headers2, estilos)
    fila = 5
    for row in reparaciones:
        vals = [
            row['folio'],
            row['service_tag'],
            row['tecnico'],
            row['sucursal'],
            row['estado'],
            _fmt_fecha(row['fecha_finalizacion']),
            _si_no(row['cot_aceptada']),
            _si_no(row['tiene_vm']),
            float(row['valor_cot_aceptada']),
            row['folio_vm'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=fila, column=col, value=val)
            cell.border = estilos['border']
            if col == 9:
                cell.number_format = estilos['number_fmt']
        fila += 1
    _auto_ajustar(ws2)

    # ------------------------------------------------------------------
    # HOJA 3: Detalle ventas mostrador
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet(HOJAS_EXCEL[2])
    ws3.merge_cells('A1:P1')
    ws3['A1'] = 'DETALLE VENTAS MOSTRADOR'
    ws3['A1'].font = estilos['title_font']
    ws3['A1'].fill = estilos['title_fill']
    ws3['A1'].alignment = estilos['title_align']
    ws3.merge_cells('A2:P2')
    ws3['A2'] = filtros_texto
    ws3['A2'].font = estilos['subtitle_font']

    headers3 = [
        'Orden cliente',
        'Service Tag',
        'Folio VM',
        'Técnico',
        'Paquete',
        'Costo paquete',
        'Limpieza',
        'Costo limpieza',
        'Reinstalación SO',
        'Costo reinstalación',
        'Respaldo',
        'Costo respaldo',
        'Cambio pieza',
        'Kit limpieza',
        'Piezas',
        'Total VM',
    ]
    _aplicar_header(ws3, 4, headers3, estilos)
    fila = 5
    for row in ventas:
        vals = [
            row['folio'],
            row['service_tag'],
            row['folio_vm'],
            row['tecnico'],
            row['paquete_nombre'],
            float(row['costo_paquete']),
            _si_no(row['incluye_limpieza']),
            float(row['costo_limpieza']),
            _si_no(row['incluye_reinstalacion']),
            float(row['costo_reinstalacion']),
            _si_no(row['incluye_respaldo']),
            float(row['costo_respaldo']),
            _si_no(row['incluye_cambio_pieza']),
            _si_no(row['incluye_kit']),
            row['resumen_piezas'],
            float(row['total_vm']),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=fila, column=col, value=val)
            cell.border = estilos['border']
            # Costos: paquete(6), limpieza(8), reinst(10), respaldo(12), total(16)
            if col in (6, 8, 10, 12, 16):
                cell.number_format = estilos['number_fmt']
        fila += 1
    _auto_ajustar(ws3)

    # ------------------------------------------------------------------
    # HOJA 4: Detalle diagnósticos
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet(HOJAS_EXCEL[3])
    ws4.merge_cells('A1:G1')
    ws4['A1'] = 'DETALLE DIAGNÓSTICOS'
    ws4['A1'].font = estilos['title_font']
    ws4['A1'].fill = estilos['title_fill']
    ws4['A1'].alignment = estilos['title_align']
    ws4.merge_cells('A2:G2')
    ws4['A2'] = (
        f'{filtros_texto} | Fecha: fin diagnóstico o, si falta, fecha de ingreso'
    )
    ws4['A2'].font = estilos['subtitle_font']

    headers4 = [
        'Orden cliente',
        'Service Tag',
        'Técnico',
        'Sucursal',
        'Fecha diagnóstico',
        'Longitud texto',
        'Extracto diagnóstico',
    ]
    _aplicar_header(ws4, 4, headers4, estilos)
    fila = 5
    for row in diagnosticos:
        vals = [
            row['folio'],
            row['service_tag'],
            row['tecnico'],
            row['sucursal'],
            _fmt_fecha(row['fecha_diagnostico']),
            row['longitud_texto'],
            row['extracto'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws4.cell(row=fila, column=col, value=val)
            cell.border = estilos['border']
        fila += 1
    _auto_ajustar(ws4)

    return wb


def hay_datos_productividad(
    fecha_inicio: date | datetime | str | None = None,
    fecha_fin: date | datetime | str | None = None,
    sucursal_id: int | str | None = None,
    tecnico_id: int | str | None = None,
    gama: str | None = None,
) -> bool:
    """
    True si al menos una de las tres métricas tiene filas (útil para tests).

    Args:
        Mismos filtros que generar_workbook_productividad_tecnicos.

    Returns:
        bool
    """
    filtros = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'sucursal_id': sucursal_id,
        'tecnico_id': tecnico_id,
        'gama': gama,
    }
    if obtener_reparaciones_productivas(**filtros):
        return True
    if obtener_ventas_mostrador_productivas(**filtros):
        return True
    if obtener_diagnosticos_realizados(**filtros):
        return True
    return False
