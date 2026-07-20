"""
Dashboard de seguimiento de piezas en tránsito + export Excel (Fase 8).

urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
"""

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render

from inventario.models import Sucursal

from config.constants import (
    ESTADO_PIEZA_CHOICES,
    ESTADOS_PIEZA_PENDIENTES,
    ESTADOS_PIEZA_PROBLEMATICOS,
)
from .decorators import permission_required_with_message


# ============================================================================
# DASHBOARD DE SEGUIMIENTO DE PIEZAS EN TRÁNSITO
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def dashboard_seguimiento_piezas(request):
    """
    Dashboard dedicado para seguimiento de piezas en tránsito.

    """
    from datetime import datetime, timedelta, date
    import pandas as pd
    from .utils_cotizaciones import (
        obtener_dataframe_seguimientos_piezas,
        calcular_kpis_seguimientos_piezas
    )
    from .plotly_visualizations import DashboardCotizacionesVisualizer, convertir_figura_a_html
    import plotly.graph_objects as go
    import plotly.express as px
    
    # ========================================
    # 1. OBTENER Y VALIDAR FILTROS
    # ========================================
    
    # Fechas por defecto: últimos 6 meses (mayor ventana para seguimientos)
    fecha_fin_default = datetime.now().date()
    fecha_inicio_default = (datetime.now() - timedelta(days=180)).date()
    
    # Capturar parámetros GET
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    proveedor_filtro = request.GET.get('proveedor')
    estado_filtro = request.GET.get('estado')
    busqueda = request.GET.get('busqueda', '').strip()
    
    # Validar y parsear fechas
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else fecha_inicio_default
    except ValueError:
        fecha_inicio = fecha_inicio_default
    
    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else fecha_fin_default
    except ValueError:
        fecha_fin = fecha_fin_default
    
    # Convertir sucursal_id a entero
    try:
        sucursal_id = int(sucursal_id) if sucursal_id else None
    except (ValueError, TypeError):
        sucursal_id = None
    
    # ========================================
    # 2. OBTENER DATOS CON FILTROS
    # ========================================
    
    try:
        # Obtener DataFrame de seguimientos
        df_seguimientos = obtener_dataframe_seguimientos_piezas(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            sucursal_id=sucursal_id,
            proveedor=proveedor_filtro,
            estado=estado_filtro
        )
        
        # Aplicar búsqueda libre si existe
        if busqueda and not df_seguimientos.empty:
            df_seguimientos = df_seguimientos[
                df_seguimientos['orden_numero'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['proveedor'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['descripcion_piezas'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['numero_pedido'].str.contains(busqueda, case=False, na=False)
            ]
        
        # Calcular KPIs
        kpis = calcular_kpis_seguimientos_piezas(df_seguimientos)
        
    except Exception as e:
        # Si hay error, crear DataFrames vacíos
        print(f"Error al obtener datos: {e}")
        df_seguimientos = pd.DataFrame()
        kpis = calcular_kpis_seguimientos_piezas(df_seguimientos)
    
    # ========================================
    # 3. GENERAR GRÁFICOS CON PLOTLY (PYTHON)
    # ========================================
    
    graficos = {}
    
    # Gráfico 1: Distribución por Estado (Pie Chart)
    if not df_seguimientos.empty:
        try:
            estado_counts = df_seguimientos['estado_display'].value_counts()
            fig_estados = px.pie(
                values=estado_counts.values,
                names=estado_counts.index,
                title='Distribución de Seguimientos por Estado',
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            fig_estados.update_traces(textposition='inside', textinfo='percent+label')
            fig_estados.update_layout(
                height=400,
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='#1e293b',
                font=dict(color='#e2e8f0')
            )
            graficos['distribucion_estados'] = convertir_figura_a_html(fig_estados)
        except Exception as e:
            print(f"Error en gráfico de estados: {e}")
            graficos['distribucion_estados'] = None
    
    # Gráfico 2: Top Proveedores por Volumen (Bar Chart Horizontal)
    if not df_seguimientos.empty:
        try:
            top_proveedores = df_seguimientos['proveedor'].value_counts().head(10)
            fig_proveedores = go.Figure(data=[
                go.Bar(
                    y=top_proveedores.index,
                    x=top_proveedores.values,
                    orientation='h',
                    marker=dict(
                        color=top_proveedores.values,
                        colorscale='Viridis',
                        showscale=True
                    ),
                    text=top_proveedores.values,
                    textposition='auto',
                )
            ])
            fig_proveedores.update_layout(
                title='Top 10 Proveedores por Número de Pedidos',
                xaxis_title='Número de Pedidos',
                yaxis_title='Proveedor',
                height=500,
                showlegend=False,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='#1e293b',
                font=dict(color='#e2e8f0')
            )
            graficos['top_proveedores'] = convertir_figura_a_html(fig_proveedores)
        except Exception as e:
            print(f"Error en gráfico de proveedores: {e}")
            graficos['top_proveedores'] = None
    
    # Gráfico 3: Tiempos de Entrega por Proveedor (Box Plot)
    if not df_seguimientos.empty:
        try:
            # Filtrar solo piezas recibidas para análisis de tiempos
            df_recibidos = df_seguimientos[df_seguimientos['estado'] == 'recibido']
            if not df_recibidos.empty and len(df_recibidos['proveedor'].unique()) > 1:
                # Tomar top 8 proveedores por volumen
                top_prov = df_recibidos['proveedor'].value_counts().head(8).index
                df_recibidos_top = df_recibidos[df_recibidos['proveedor'].isin(top_prov)]
                
                fig_tiempos = px.box(
                    df_recibidos_top,
                    x='proveedor',
                    y='dias_desde_pedido',
                    title='Tiempos de Entrega por Proveedor (Días)',
                    color='proveedor',
                    labels={'dias_desde_pedido': 'Días desde Pedido', 'proveedor': 'Proveedor'}
                )
                fig_tiempos.update_layout(
                    height=450,
                    showlegend=False,
                    xaxis_tickangle=-45,
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='#1e293b',
                    font=dict(color='#e2e8f0')
                )
                graficos['tiempos_entrega_proveedor'] = convertir_figura_a_html(fig_tiempos)
            else:
                graficos['tiempos_entrega_proveedor'] = None
        except Exception as e:
            print(f"Error en gráfico de tiempos: {e}")
            graficos['tiempos_entrega_proveedor'] = None
    
    # Gráfico 4: Timeline de Entregas Esperadas (Gantt-like)
    if not df_seguimientos.empty:
        try:
            # Filtrar solo activos con fecha estimada futura
            # ACTUALIZADO: Usar constantes para estados pendientes
            df_activos = df_seguimientos[
                (df_seguimientos['estado'].isin(ESTADOS_PIEZA_PENDIENTES)) &
                (df_seguimientos['dias_hasta_entrega'] >= -30)  # Incluir hasta 30 días de retraso
            ].copy()
            
            if not df_activos.empty:
                # Convertir columnas de fecha a datetime si son date objects
                df_activos['fecha_pedido'] = pd.to_datetime(df_activos['fecha_pedido'])
                df_activos['fecha_entrega_estimada'] = pd.to_datetime(df_activos['fecha_entrega_estimada'])
                
                # Ordenar por fecha estimada
                df_activos = df_activos.sort_values('fecha_entrega_estimada')
                
                # Tomar máximo 20 para no saturar el gráfico
                df_activos = df_activos.head(20)
                
                # Crear etiquetas descriptivas con orden_cliente y service_tag
                df_activos['etiqueta'] = df_activos.apply(
                    lambda row: f"{row['orden_cliente']} ({row['service_tag'][:15]}) - {row['proveedor'][:20]}", axis=1
                )
                
                # Asignar colores según prioridad
                color_map = {
                    'critico': 'red',
                    'alto': 'orange',
                    'medio': 'yellow',
                    'normal': 'green'
                }
                df_activos['color'] = df_activos['prioridad'].map(color_map)
                
                fig_timeline = go.Figure()
                
                for idx, row in df_activos.iterrows():
                    fig_timeline.add_trace(go.Scatter(
                        x=[row['fecha_pedido'], row['fecha_entrega_estimada']],
                        y=[row['etiqueta'], row['etiqueta']],
                        mode='lines+markers',
                        line=dict(color=row['color'], width=8),
                        marker=dict(size=10, symbol='circle'),
                        name=row['etiqueta'],
                        showlegend=False,
                        hovertemplate=(
                            f"<b>{row['etiqueta']}</b><br>" +
                            f"Pedido: {row['fecha_pedido']}<br>" +
                            f"Estimado: {row['fecha_entrega_estimada']}<br>" +
                            f"Días restantes: {row['dias_hasta_entrega']}<br>" +
                            f"Estado: {row['estado_display']}<br>" +
                            "<extra></extra>"
                        )
                    ))
                
                # Agregar línea vertical para HOY usando add_shape
                hoy = datetime.now()
                
                fig_timeline.update_layout(
                    title='Timeline de Entregas Esperadas (Próximas 20 Piezas)',
                    xaxis_title='Fecha',
                    yaxis_title='Orden - Proveedor',
                    height=600,
                    hovermode='closest',
                    yaxis=dict(autorange="reversed"),  # Más recientes arriba
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='#1e293b',
                    font=dict(color='#e2e8f0'),
                    shapes=[
                        dict(
                            type="line",
                            x0=hoy,
                            x1=hoy,
                            y0=0,
                            y1=1,
                            yref="paper",
                            line=dict(color="#38bdf8", width=2, dash="dash"),
                        )
                    ],
                    annotations=[
                        dict(
                            x=hoy,
                            y=1,
                            yref="paper",
                            text="HOY",
                            showarrow=False,
                            xanchor="center",
                            yanchor="bottom",
                            font=dict(color="#38bdf8", size=12, family="Arial Black"),
                            bgcolor="rgba(15, 23, 42, 0.8)",
                        )
                    ]
                )
                
                graficos['timeline_entregas'] = convertir_figura_a_html(fig_timeline)
            else:
                graficos['timeline_entregas'] = None
        except Exception as e:
            import traceback
            print(f"Error en gráfico timeline: {e}")
            print(traceback.format_exc())
            graficos['timeline_entregas'] = None
    
    # ========================================
    # 4. PREPARAR DATOS PARA ALERTAS
    # ========================================
    
    # ========================================
    # 4.5 PREPARAR VISTA AGRUPADA POR ORDEN
    # ========================================
    
    # Importar función de agrupación
    from .utils_cotizaciones import agrupar_seguimientos_por_orden
    
    # Generar vista agrupada
    if not df_seguimientos.empty:
        ordenes_agrupadas = agrupar_seguimientos_por_orden(df_seguimientos)
        
        # Separar agrupadas en activas y recibidas
        ordenes_activas = [o for o in ordenes_agrupadas if o['estado_general'] != 'todos_recibidos']
        ordenes_recibidas = [o for o in ordenes_agrupadas if o['estado_general'] == 'todos_recibidos']
        
        # KPIs específicos de la vista agrupada
        kpis_agrupados = {
            'total_ordenes': len(ordenes_agrupadas),
            'ordenes_activas': len(ordenes_activas),
            'ordenes_completadas': len(ordenes_recibidas),
            'ordenes_con_retrasos': len([o for o in ordenes_agrupadas if o['tiene_retrasados']]),
            'ordenes_criticas': len([o for o in ordenes_agrupadas if o['prioridad_maxima'] == 'critico']),
        }
    else:
        ordenes_agrupadas = []
        ordenes_activas = []
        ordenes_recibidas = []
        kpis_agrupados = {
            'total_ordenes': 0,
            'ordenes_activas': 0,
            'ordenes_completadas': 0,
            'ordenes_con_retrasos': 0,
            'ordenes_criticas': 0,
        }
    
    # Filtrar piezas retrasadas para alertas (solo las que NO han llegado)
    # ACTUALIZADO: Usar constantes para estados pendientes
    if not df_seguimientos.empty:
        df_retrasados = df_seguimientos[
            (df_seguimientos['esta_retrasado'] == True) &
            (df_seguimientos['estado'].isin(ESTADOS_PIEZA_PENDIENTES))  # Solo activas, no recibidas
        ]
        piezas_retrasadas = df_retrasados.to_dict('records')
    else:
        piezas_retrasadas = []
    
    # Filtrar piezas próximas a llegar (siguientes 3 días)
    # ACTUALIZADO: Usar constantes para estados pendientes
    if not df_seguimientos.empty:
        df_proximos = df_seguimientos[
            (df_seguimientos['dias_hasta_entrega'] >= 0) &
            (df_seguimientos['dias_hasta_entrega'] <= 3) &
            (df_seguimientos['estado'].isin(ESTADOS_PIEZA_PENDIENTES))
        ]
        piezas_proximas = df_proximos.to_dict('records')
    else:
        piezas_proximas = []
    
    # NUEVO: Filtrar piezas con problemas de calidad (WPB/DOA)
    # Estas piezas llegaron físicamente pero con incidencias de calidad
    if not df_seguimientos.empty:
        df_problematicos = df_seguimientos[
            df_seguimientos['estado'].isin(ESTADOS_PIEZA_PROBLEMATICOS)
        ]
        piezas_problematicas = df_problematicos.to_dict('records')
    else:
        piezas_problematicas = []
    
    # ========================================
    # 5. PREPARAR DATOS PARA FILTROS
    # ========================================
    
    # Lista de sucursales para el filtro
    sucursales = Sucursal.objects.all().order_by('nombre')
    
    # Lista de proveedores únicos para el filtro
    if not df_seguimientos.empty:
        proveedores_lista = sorted(df_seguimientos['proveedor'].unique().tolist())
    else:
        proveedores_lista = []
    
    # Estados disponibles - Ya importados al inicio del archivo
    estados_choices = ESTADO_PIEZA_CHOICES
    
    # ========================================
    # 6. PREPARAR CONTEXTO PARA EL TEMPLATE
    # ========================================
    
    context = {
        # KPIs
        'kpis': kpis,
        'kpis_agrupados': kpis_agrupados,
        
        # Alertas
        'piezas_retrasadas': piezas_retrasadas,
        'piezas_proximas': piezas_proximas,
        'piezas_problematicas': piezas_problematicas,  # NUEVO: WPB/DOA
        
        # Datos para la tabla (vista agrupada)
        'ordenes_agrupadas': ordenes_agrupadas,
        'ordenes_activas': ordenes_activas,
        'ordenes_recibidas': ordenes_recibidas,
        
        # Gráficos
        'graficos': graficos,
        
        # Filtros para el formulario
        'sucursales': sucursales,
        'proveedores': proveedores_lista,
        'estados': estados_choices,
        
        # Filtros activos (para mantener valores en el form)
        'filtros_activos': {
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else '',
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d') if fecha_fin else '',
            'sucursal': sucursal_id,
            'proveedor': proveedor_filtro or '',
            'estado': estado_filtro or '',
            'busqueda': busqueda,
        },
        
        # Totales
        'total_ordenes': len(ordenes_agrupadas),
    }
    
    return render(request, 'servicio_tecnico/dashboard_seguimiento_piezas.html', context)


@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def exportar_dashboard_seguimiento_piezas(request):
    """
    Exporta el dashboard de seguimiento de piezas a Excel.

    Returns:
        HttpResponse: Archivo Excel descargable
    """
    from datetime import datetime, timedelta
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from django.http import HttpResponse
    from .utils_cotizaciones import (
        obtener_dataframe_seguimientos_piezas,
        calcular_kpis_seguimientos_piezas
    )
    
    # ========================================
    # 1. OBTENER MISMOS FILTROS QUE EL DASHBOARD
    # ========================================
    
    fecha_fin_default = datetime.now().date()
    fecha_inicio_default = (datetime.now() - timedelta(days=180)).date()
    
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    proveedor_filtro = request.GET.get('proveedor')
    estado_filtro = request.GET.get('estado')
    busqueda = request.GET.get('busqueda', '').strip()
    
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else fecha_inicio_default
    except ValueError:
        fecha_inicio = fecha_inicio_default
    
    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else fecha_fin_default
    except ValueError:
        fecha_fin = fecha_fin_default
    
    try:
        sucursal_id = int(sucursal_id) if sucursal_id else None
    except (ValueError, TypeError):
        sucursal_id = None
    
    # ========================================
    # 2. OBTENER DATOS
    # ========================================
    
    try:
        df_seguimientos = obtener_dataframe_seguimientos_piezas(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            sucursal_id=sucursal_id,
            proveedor=proveedor_filtro,
            estado=estado_filtro
        )
        
        if busqueda and not df_seguimientos.empty:
            df_seguimientos = df_seguimientos[
                df_seguimientos['orden_numero'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['proveedor'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['descripcion_piezas'].str.contains(busqueda, case=False, na=False) |
                df_seguimientos['numero_pedido'].str.contains(busqueda, case=False, na=False)
            ]
        
        kpis = calcular_kpis_seguimientos_piezas(df_seguimientos)
        
        # Generar vista agrupada (NUEVO)
        from .utils_cotizaciones import agrupar_seguimientos_por_orden
        ordenes_agrupadas = agrupar_seguimientos_por_orden(df_seguimientos) if not df_seguimientos.empty else []
        
    except Exception as e:
        messages.error(request, f'Error al obtener datos para exportación: {str(e)}')
        return redirect('servicio_tecnico:dashboard_seguimiento_piezas')
    
    # ========================================
    # 3. CREAR ARCHIVO EXCEL
    # ========================================
    
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # HOJA 1: Resumen de KPIs
    ws_resumen = wb.create_sheet("Resumen")
    ws_resumen.append(["DASHBOARD DE SEGUIMIENTO DE PIEZAS"])
    ws_resumen.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_resumen.append([])
    ws_resumen.append(["Métrica", "Valor"])
    
    # Escribir KPIs
    kpis_data = [
        ["Total Seguimientos", kpis['total_seguimientos']],
        ["Total Activos", kpis['total_activos']],
        ["En Tránsito", kpis['en_transito']],
        ["Pedidos", kpis['pedidos']],
        ["Recibidos", kpis['recibidos']],
        ["Retrasados", kpis['retrasados']],
        ["Próximos a Llegar (3 días)", kpis['proximos_llegar']],
        ["Promedio Días Entrega", kpis['promedio_dias_entrega']],
        ["Promedio Días Retraso", kpis['promedio_dias_retraso']],
    ]
    
    for row in kpis_data:
        ws_resumen.append(row)
    
    # Aplicar estilos al resumen
    for row in ws_resumen.iter_rows(min_row=4, max_row=4, min_col=1, max_col=2):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    
    # HOJA 2: Todos los Seguimientos
    if not df_seguimientos.empty:
        ws_seguimientos = wb.create_sheet("Seguimientos")
        
        # Seleccionar columnas para Excel
        columnas_excel = [
            'orden_numero', 'proveedor', 'estado_display', 'descripcion_piezas',
            'fecha_pedido', 'fecha_entrega_estimada', 'fecha_entrega_real',
            'dias_desde_pedido', 'dias_retraso', 'dias_hasta_entrega',
            'sucursal', 'responsable', 'numero_pedido'
        ]
        
        df_export = df_seguimientos[columnas_excel].copy()
        
        # Renombrar columnas para Excel
        df_export.columns = [
            'Orden', 'Proveedor', 'Estado', 'Descripción Piezas',
            'Fecha Pedido', 'Fecha Estimada', 'Fecha Real',
            'Días desde Pedido', 'Días Retraso', 'Días hasta Entrega',
            'Sucursal', 'Responsable', 'Nº Pedido'
        ]
        
        # Escribir DataFrame
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws_seguimientos.append(r)
        
        # Aplicar estilos a encabezados
        for cell in ws_seguimientos[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    
    # HOJA 3: Solo Retrasados
    if not df_seguimientos.empty:
        df_retrasados = df_seguimientos[df_seguimientos['esta_retrasado'] == True]
        
        if not df_retrasados.empty:
            ws_retrasados = wb.create_sheet("Retrasados")
            
            df_retrasados_export = df_retrasados[columnas_excel].copy()
            df_retrasados_export.columns = df_export.columns
            
            for r in dataframe_to_rows(df_retrasados_export, index=False, header=True):
                ws_retrasados.append(r)
            
            for cell in ws_retrasados[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.alignment = center_align
    
    # HOJA 4: Por Proveedor
    if not df_seguimientos.empty:
        ws_proveedores = wb.create_sheet("Por Proveedor")
        
        df_por_proveedor = df_seguimientos.groupby('proveedor').agg({
            'id': 'count',
            'esta_retrasado': 'sum',
            'dias_desde_pedido': 'mean',
            'dias_retraso': 'mean'
        }).reset_index()
        
        df_por_proveedor.columns = [
            'Proveedor', 'Total Pedidos', 'Retrasados',
            'Promedio Días Pedido', 'Promedio Días Retraso'
        ]
        
        df_por_proveedor = df_por_proveedor.sort_values('Total Pedidos', ascending=False)
        
        for r in dataframe_to_rows(df_por_proveedor, index=False, header=True):
            ws_proveedores.append(r)
        
        for cell in ws_proveedores[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    
    # HOJA 5: Por Sucursal
    if not df_seguimientos.empty:
        ws_sucursales = wb.create_sheet("Por Sucursal")
        
        df_por_sucursal = df_seguimientos.groupby('sucursal').agg({
            'id': 'count',
            'esta_retrasado': 'sum',
            'dias_desde_pedido': 'mean'
        }).reset_index()
        
        df_por_sucursal.columns = [
            'Sucursal', 'Total Pedidos', 'Retrasados', 'Promedio Días Pedido'
        ]
        
        df_por_sucursal = df_por_sucursal.sort_values('Total Pedidos', ascending=False)
        
        for r in dataframe_to_rows(df_por_sucursal, index=False, header=True):
            ws_sucursales.append(r)
        
        for cell in ws_sucursales[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
    
    # HOJA 6: Vista Agrupada por Orden (NUEVO)
    if ordenes_agrupadas:
        ws_agrupada = wb.create_sheet("Vista Agrupada")
        
        # Encabezados
        headers = [
            'Orden Cliente', 'Service Tag', 'Sucursal', 'Total Proveedores',
            'Seguimientos Recibidos', 'Seguimientos Pendientes', 'Estado General',
            'Tiene Retrasos', 'Días Máx Retraso', 'Prioridad Máxima',
            'Fecha Pedido Más Antigua', 'Fecha Entrega Más Próxima', 'Proveedores Detalle'
        ]
        ws_agrupada.append(headers)
        
        # Datos
        for orden in ordenes_agrupadas:
            # Construir cadena de proveedores
            proveedores_str = ' | '.join([
                f"{p['proveedor']}: {p['estado_display']} ({p['descripcion'][:30]}...)"
                for p in orden['proveedores_activos']
            ])
            
            row_data = [
                orden['orden_cliente'],
                orden['service_tag'],
                orden['sucursal'],
                orden['total_proveedores'],
                orden['seguimientos_recibidos'],
                orden['seguimientos_pendientes'],
                orden['estado_general'].upper(),
                'SÍ' if orden['tiene_retrasados'] else 'NO',
                orden['dias_maximo_retraso'] if orden['tiene_retrasados'] else 0,
                orden['prioridad_maxima'].upper(),
                orden['fecha_pedido_mas_antigua'].strftime('%Y-%m-%d') if orden['fecha_pedido_mas_antigua'] else '',
                orden['fecha_entrega_mas_proxima'].strftime('%Y-%m-%d') if orden['fecha_entrega_mas_proxima'] else '',
                proveedores_str
            ]
            ws_agrupada.append(row_data)
        
        # Aplicar estilos
        for cell in ws_agrupada[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.alignment = center_align
        
        # Ajustar ancho de columnas
        ws_agrupada.column_dimensions['A'].width = 15
        ws_agrupada.column_dimensions['B'].width = 15
        ws_agrupada.column_dimensions['C'].width = 15
        ws_agrupada.column_dimensions['D'].width = 12
        ws_agrupada.column_dimensions['E'].width = 12
        ws_agrupada.column_dimensions['F'].width = 12
        ws_agrupada.column_dimensions['G'].width = 15
        ws_agrupada.column_dimensions['H'].width = 12
        ws_agrupada.column_dimensions['I'].width = 12
        ws_agrupada.column_dimensions['J'].width = 15
        ws_agrupada.column_dimensions['K'].width = 18
        ws_agrupada.column_dimensions['L'].width = 18
        ws_agrupada.column_dimensions['M'].width = 50
        
        # Colorear filas según prioridad
        for row_idx, orden in enumerate(ordenes_agrupadas, start=2):
            if orden['prioridad_maxima'] == 'critico':
                fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for cell in ws_agrupada[row_idx]:
                    cell.fill = fill_color
            elif orden['prioridad_maxima'] == 'alto':
                fill_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                for cell in ws_agrupada[row_idx]:
                    cell.fill = fill_color
    
    # ========================================
    # 4. PREPARAR RESPUESTA HTTP
    # ========================================
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'Dashboard_Seguimiento_Piezas_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    
    return response


