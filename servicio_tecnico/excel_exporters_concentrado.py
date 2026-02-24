"""
Exportador Excel para el Concentrado Semanal CIS
================================================

EXPLICACIÓN PARA PRINCIPIANTES:
================================
Este archivo genera un archivo Excel (.xlsx) con 4 hojas a partir de los datos
del concentrado semanal. Cada hoja tiene un propósito diferente:

  Hoja 1 "Concentrado Semanal" → Las 3 tablas del reporte (Ingreso, Asignación, Egreso)
  Hoja 2 "Reporte Trimestral"  → Totales acumulados por Q1, Q2, Q3, Q4
  Hoja 3 "Gráfico Ingresos"    → Gráfico de barras con ingresos semanales del año
  Hoja 4 "Gráfico Egresos"     → Gráfico de barras con egresos semanales del año

CÓMO SE USA:
  from .excel_exporters_concentrado import generar_excel_concentrado

  wb = generar_excel_concentrado(datos_semana, datos_trimestral, datos_tendencia)
  wb.save(response)

DEPENDENCIAS:
  - openpyxl: pip install openpyxl
  - Estilos importados desde excel_exporters.py (no se duplican)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Importar utilidades de estilo del módulo principal (reutilizar, no duplicar)
from .excel_exporters import (
    apply_cell_style,
    get_header_style,
    get_title_style,
    auto_adjust_column_width,
)

# Constantes del concentrado (copiadas aquí para no importar models en Excel)
DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
TIPOS_EQUIPO = ['LENOVO', 'DELL', 'OOW', 'MIS DELL', 'MIS LENOVO']
SITIOS = ['DROP OFF', 'SATELITE']


# ===========================================================================
# ESTILOS ESPECÍFICOS DEL CONCENTRADO
# ===========================================================================

def _get_thin_border():
    """Borde fino para todas las celdas de datos."""
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _get_header_ingreso():
    """Estilo de encabezado azul para la sección de Ingreso."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _get_thin_border(),
    }


def _get_header_asignacion():
    """Estilo de encabezado índigo para la sección de Asignación."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='6610f2', end_color='6610f2', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _get_thin_border(),
    }


def _get_header_egreso():
    """Estilo de encabezado verde para la sección de Egreso."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='198754', end_color='198754', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _get_thin_border(),
    }


def _get_subheader_style(color_hex):
    """
    Estilo de sub-encabezado (fila de columnas: SITIO, EQUIPO, LUNES, etc.).

    Args:
        color_hex (str): Color de fondo en hexadecimal (sin #)
    """
    return {
        'font': Font(bold=True, color='333333', size=10),
        'fill': PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _get_thin_border(),
    }


def _get_data_cell_style():
    """Estilo para celdas de datos normales (números de conteo)."""
    return {
        'font': Font(size=10),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': _get_thin_border(),
    }


def _get_total_row_style():
    """Estilo para la fila de totales generales."""
    return {
        'font': Font(bold=True, size=10, color='212529'),
        'fill': PatternFill(start_color='e9ecef', end_color='e9ecef', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': _get_thin_border(),
    }


def _get_sitio_cell_style():
    """Estilo para la celda de SITIO (columna izquierda con rowspan visual)."""
    return {
        'font': Font(bold=True, size=10, color='495057'),
        'fill': PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': _get_thin_border(),
    }


def _get_escalados_style():
    """Estilo amarillo para la fila Escalados / Packing."""
    return {
        'font': Font(italic=True, size=10, color='664d03'),
        'fill': PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': _get_thin_border(),
    }


def _get_quarter_header_style(color_hex):
    """Estilo para encabezados de quarter en la hoja trimestral."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=12),
        'fill': PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': _get_thin_border(),
    }


# ===========================================================================
# FUNCIÓN AUXILIAR: ESCRIBIR TABLA DE INGRESO O EGRESO
# ===========================================================================

def _escribir_tabla_ingreso_egreso(
    ws,
    fila_inicio,
    titulo,
    estilo_titulo,
    estilo_subheader_hex,
    datos,
    totales,
):
    """
    Escribe una tabla de Ingreso o Egreso en la hoja de cálculo.

    EXPLICACIÓN:
    Las tablas de Ingreso y Egreso tienen exactamente la misma estructura,
    solo cambian los datos y el color. Esta función evita duplicar código.

    Args:
        ws: Worksheet de openpyxl donde escribir
        fila_inicio (int): Número de fila donde empezar
        titulo (str): Texto del título de la sección
        estilo_titulo (dict): Estilo para la fila de título
        estilo_subheader_hex (str): Color hex para los encabezados de columna
        datos (dict): {sitio: {tipo: {dia: conteo, 'total': N, 'promedio': N}}}
        totales (dict): {dia: conteo, 'total': N, 'promedio': N}

    Returns:
        int: Número de la primera fila disponible después de la tabla
    """
    # Total de columnas: SITIO(1) + EQUIPO(1) + 5 días + TOTAL(1) + PROMEDIO(1) = 10
    num_cols = 10  # columnas A..J (o las que correspondan al offset)
    col_inicio = 1

    # ---- Fila de título (merge toda la fila) ----
    ws.merge_cells(
        start_row=fila_inicio, start_column=col_inicio,
        end_row=fila_inicio, end_column=col_inicio + num_cols - 1
    )
    celda_titulo = ws.cell(row=fila_inicio, column=col_inicio, value=titulo)
    apply_cell_style(celda_titulo, estilo_titulo)
    ws.row_dimensions[fila_inicio].height = 24
    fila = fila_inicio + 1

    # ---- Fila de sub-encabezados ----
    encabezados = ['SITIO', 'EQUIPO'] + [d.upper() for d in DIAS_SEMANA] + ['TOTAL', 'PROMEDIO']
    subheader_style = _get_subheader_style(estilo_subheader_hex)
    for c_offset, texto in enumerate(encabezados):
        celda = ws.cell(row=fila, column=col_inicio + c_offset, value=texto)
        apply_cell_style(celda, subheader_style)
    ws.row_dimensions[fila].height = 18
    fila += 1

    # ---- Filas de datos por SITIO × TIPO ----
    for sitio in SITIOS:
        sitio_fila_inicio = fila  # Para calcular merge al final
        for tipo in TIPOS_EQUIPO:
            tipo_data = datos[sitio][tipo]

            # Columna SITIO (solo en la primera fila del sitio; se mergea después)
            ws.cell(row=fila, column=col_inicio, value=sitio)

            # Columna EQUIPO
            celda_tipo = ws.cell(row=fila, column=col_inicio + 1, value=tipo)
            apply_cell_style(celda_tipo, {
                'font': Font(bold=True, size=10),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': _get_thin_border(),
            })

            # Columnas de días
            for d_idx, dia in enumerate(DIAS_SEMANA):
                valor = tipo_data.get(dia, 0)
                celda = ws.cell(row=fila, column=col_inicio + 2 + d_idx, value=valor)
                apply_cell_style(celda, _get_data_cell_style())

            # Total
            celda_total = ws.cell(row=fila, column=col_inicio + 7, value=tipo_data.get('total', 0))
            apply_cell_style(celda_total, {
                'font': Font(bold=True, size=10, color='0d6efd'),
                'fill': PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': _get_thin_border(),
            })

            # Promedio
            celda_prom = ws.cell(row=fila, column=col_inicio + 8, value=tipo_data.get('promedio', 0))
            apply_cell_style(celda_prom, {
                'font': Font(italic=True, size=10, color='6c757d'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': _get_thin_border(),
            })

            ws.row_dimensions[fila].height = 16
            fila += 1

        # Merge la columna SITIO para las filas de este sitio
        if len(TIPOS_EQUIPO) > 1:
            ws.merge_cells(
                start_row=sitio_fila_inicio, start_column=col_inicio,
                end_row=fila - 1, end_column=col_inicio
            )
        celda_sitio = ws.cell(row=sitio_fila_inicio, column=col_inicio, value=sitio)
        apply_cell_style(celda_sitio, _get_sitio_cell_style())

    # ---- Fila de TOTALES ----
    celda_total_label = ws.cell(row=fila, column=col_inicio, value='TOTALES GENERALES')
    ws.merge_cells(
        start_row=fila, start_column=col_inicio,
        end_row=fila, end_column=col_inicio + 1
    )
    apply_cell_style(celda_total_label, _get_total_row_style())

    for d_idx, dia in enumerate(DIAS_SEMANA):
        celda = ws.cell(row=fila, column=col_inicio + 2 + d_idx, value=totales.get(dia, 0))
        apply_cell_style(celda, _get_total_row_style())

    celda_tot = ws.cell(row=fila, column=col_inicio + 7, value=totales.get('total', 0))
    apply_cell_style(celda_tot, _get_total_row_style())

    celda_prom = ws.cell(row=fila, column=col_inicio + 8, value=totales.get('promedio', 0))
    apply_cell_style(celda_prom, _get_total_row_style())

    ws.row_dimensions[fila].height = 18
    fila += 1

    return fila  # Siguiente fila disponible


# ===========================================================================
# HOJA 1: CONCENTRADO SEMANAL
# ===========================================================================

def _crear_hoja_concentrado(wb, datos_semana):
    """
    Crea la Hoja 1 "Concentrado Semanal" con las 3 tablas del reporte.

    Args:
        wb: Workbook de openpyxl
        datos_semana (dict): Datos devueltos por obtener_concentrado_semanal()
    """
    ws = wb.active
    ws.title = 'Concentrado Semanal'
    ws.sheet_view.showGridLines = True

    num_semana = datos_semana.get('numero_semana', '?')
    año = datos_semana.get('año', '?')
    lunes = datos_semana.get('lunes')
    viernes = datos_semana.get('viernes')

    # ---- Título principal ----
    ws.merge_cells('A1:J1')
    celda_titulo = ws.cell(
        row=1, column=1,
        value=f'CONCENTRADO SEMANAL CIS — Semana {num_semana} de {año}'
    )
    apply_cell_style(celda_titulo, {
        'font': Font(bold=True, size=14, color='1e3a5f'),
        'fill': PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    })
    ws.row_dimensions[1].height = 28

    # ---- Subtítulo con rango de fechas ----
    ws.merge_cells('A2:J2')
    fecha_str = ''
    if lunes and viernes:
        fecha_str = f'{lunes.strftime("%d/%m/%Y")} — {viernes.strftime("%d/%m/%Y")}'
    celda_fecha = ws.cell(row=2, column=1, value=fecha_str)
    apply_cell_style(celda_fecha, {
        'font': Font(size=10, color='6c757d', italic=True),
        'alignment': Alignment(horizontal='center', vertical='center'),
    })
    ws.row_dimensions[2].height = 16

    fila_actual = 4  # Espacio en blanco en fila 3

    # ---- Sección 1: Ingreso ----
    fila_actual = _escribir_tabla_ingreso_egreso(
        ws=ws,
        fila_inicio=fila_actual,
        titulo='INGRESO DE EQUIPOS EN CIS',
        estilo_titulo=_get_header_ingreso(),
        estilo_subheader_hex='cfe2ff',
        datos=datos_semana['ingreso'],
        totales=datos_semana['totales_ingreso'],
    )

    # Resumen Carry-In / MIS debajo de la tabla de ingreso
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value='Carry In – Drop Off:').font = Font(bold=True, size=10)
    ws.cell(row=fila_actual, column=2, value=datos_semana.get('carry_in_dropoff', 0)).font = Font(size=10)
    ws.cell(row=fila_actual, column=4, value='Carry In – SIC (Satélite):').font = Font(bold=True, size=10)
    ws.cell(row=fila_actual, column=5, value=datos_semana.get('carry_in_sic', 0)).font = Font(size=10)
    ws.cell(row=fila_actual, column=7, value='Mail In Service (MIS):').font = Font(bold=True, size=10)
    ws.cell(row=fila_actual, column=8, value=datos_semana.get('mail_in_service', 0)).font = Font(size=10)
    fila_actual += 2

    # ---- Sección 2: Asignación a Ingeniería ----
    # Título de sección
    ws.merge_cells(
        start_row=fila_actual, start_column=1,
        end_row=fila_actual, end_column=8
    )
    celda_asig_titulo = ws.cell(row=fila_actual, column=1, value='ASIGNACIÓN DE EQUIPOS A INGENIERÍA')
    apply_cell_style(celda_asig_titulo, _get_header_asignacion())
    ws.row_dimensions[fila_actual].height = 24
    fila_actual += 1

    # Sub-encabezados asignación
    enc_asig = ['INGENIERO'] + [d.upper() for d in DIAS_SEMANA] + ['TOTAL']
    subheader_asig_style = _get_subheader_style('e0cffc')
    for c_off, texto in enumerate(enc_asig):
        celda = ws.cell(row=fila_actual, column=1 + c_off, value=texto)
        apply_cell_style(celda, subheader_asig_style)
    ws.row_dimensions[fila_actual].height = 18
    fila_actual += 1

    # Filas de ingenieros
    for fila_ing in datos_semana.get('asignacion', []):
        es_escalados = fila_ing.get('nombre') == 'Escalados / Packing'
        estilo_fila = _get_escalados_style() if es_escalados else _get_data_cell_style()

        celda_nombre = ws.cell(row=fila_actual, column=1, value=fila_ing.get('nombre', ''))
        apply_cell_style(celda_nombre, {
            **estilo_fila,
            'alignment': Alignment(horizontal='left', vertical='center'),
        })

        for d_idx, dia in enumerate(DIAS_SEMANA):
            celda = ws.cell(row=fila_actual, column=2 + d_idx, value=fila_ing.get(dia, 0))
            apply_cell_style(celda, estilo_fila)

        celda_tot = ws.cell(row=fila_actual, column=7, value=fila_ing.get('total', 0))
        apply_cell_style(celda_tot, {
            **estilo_fila,
            'font': Font(bold=True, size=10),
        })

        ws.row_dimensions[fila_actual].height = 16
        fila_actual += 1

    # Fila total asignación
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=6)
    celda_total_asig = ws.cell(row=fila_actual, column=1, value='TOTAL EQUIPOS INGRESADOS')
    apply_cell_style(celda_total_asig, _get_total_row_style())
    celda_tot_val = ws.cell(row=fila_actual, column=7, value=datos_semana.get('total_asignados', 0))
    apply_cell_style(celda_tot_val, _get_total_row_style())
    ws.row_dimensions[fila_actual].height = 18
    fila_actual += 2

    # ---- Sección 3: Egreso ----
    fila_actual = _escribir_tabla_ingreso_egreso(
        ws=ws,
        fila_inicio=fila_actual,
        titulo='EGRESO DE EQUIPOS EN CIS',
        estilo_titulo=_get_header_egreso(),
        estilo_subheader_hex='d1e7dd',
        datos=datos_semana['egreso'],
        totales=datos_semana['totales_egreso'],
    )

    # ---- Ajuste de anchos de columna ----
    anchos = [14, 14, 10, 10, 12, 10, 10, 10, 12, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Congelar fila de encabezados
    ws.freeze_panes = 'A3'


# ===========================================================================
# HOJA 2: REPORTE TRIMESTRAL
# ===========================================================================

def _crear_hoja_trimestral(wb, datos_trimestral):
    """
    Crea la Hoja 2 "Reporte Trimestral" con Q1, Q2, Q3, Q4.

    Args:
        wb: Workbook de openpyxl
        datos_trimestral (dict): Datos devueltos por obtener_reporte_trimestral()
    """
    ws = wb.create_sheet('Reporte Trimestral')

    # Colores para cada quarter
    QUARTER_COLORES = {
        'Q1': '0d6efd',  # Azul
        'Q2': '198754',  # Verde
        'Q3': 'dc3545',  # Rojo
        'Q4': 'fd7e14',  # Naranja
    }

    fila = 1

    for q_key, q_data in datos_trimestral.items():
        color = QUARTER_COLORES.get(q_key, '6c757d')
        nombre = q_data.get('nombre', q_key)
        total_ingreso = q_data.get('total_ingreso', 0)
        total_egreso = q_data.get('total_egreso', 0)

        # Título del quarter
        num_cols_q = 1 + 1 + len(TIPOS_EQUIPO) + 1  # SITIO + sección + tipos + total
        ws.merge_cells(
            start_row=fila, start_column=1,
            end_row=fila, end_column=len(TIPOS_EQUIPO) + 3
        )
        celda_q = ws.cell(row=fila, column=1, value=f'{nombre}  —  Ingresos: {total_ingreso}  |  Egresos: {total_egreso}')
        apply_cell_style(celda_q, _get_quarter_header_style(color))
        ws.row_dimensions[fila].height = 22
        fila += 1

        # Sub-encabezados: SITIO | SECCIÓN | LENOVO | DELL | OOW | MIS DELL | MIS LENOVO | TOTAL
        encabezados_q = ['SITIO', 'SECCIÓN'] + TIPOS_EQUIPO + ['TOTAL']
        for c_off, texto in enumerate(encabezados_q):
            celda = ws.cell(row=fila, column=1 + c_off, value=texto)
            apply_cell_style(celda, _get_subheader_style('f8f9fa'))
        ws.row_dimensions[fila].height = 16
        fila += 1

        # Filas de datos por sitio y sección (ingreso/egreso)
        for sitio in SITIOS:
            for seccion_key, seccion_label in [('ingreso', 'INGRESO'), ('egreso', 'EGRESO')]:
                datos_sec = q_data.get(seccion_key, {}).get(sitio, {})

                celda_sitio = ws.cell(row=fila, column=1, value=sitio)
                apply_cell_style(celda_sitio, _get_sitio_cell_style())

                celda_sec = ws.cell(row=fila, column=2, value=seccion_label)
                apply_cell_style(celda_sec, {
                    'font': Font(bold=True, size=10),
                    'alignment': Alignment(horizontal='center', vertical='center'),
                    'border': _get_thin_border(),
                })

                total_fila = 0
                for t_idx, tipo in enumerate(TIPOS_EQUIPO):
                    valor = datos_sec.get(tipo, 0)
                    celda = ws.cell(row=fila, column=3 + t_idx, value=valor)
                    apply_cell_style(celda, _get_data_cell_style())
                    total_fila += valor

                celda_tot_fila = ws.cell(row=fila, column=3 + len(TIPOS_EQUIPO), value=total_fila)
                apply_cell_style(celda_tot_fila, {
                    'font': Font(bold=True, size=10, color='0d6efd'),
                    'fill': PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid'),
                    'alignment': Alignment(horizontal='center', vertical='center'),
                    'border': _get_thin_border(),
                })

                ws.row_dimensions[fila].height = 15
                fila += 1

        fila += 1  # Espacio entre quarters

    # Ajuste de anchos
    anchos_q = [14, 12] + [11] * len(TIPOS_EQUIPO) + [10]
    for i, ancho in enumerate(anchos_q, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


# ===========================================================================
# HOJA 3 Y 4: GRÁFICOS DE TENDENCIA
# ===========================================================================

def _crear_hoja_grafico(wb, datos_tendencia, tipo, titulo, color_barras):
    """
    Crea una hoja de gráfico de barras (ingresos o egresos por semana del año).

    EXPLICACIÓN:
    openpyxl incluye su propio módulo de gráficos. Creamos una tabla de datos
    en la hoja y luego insertamos un BarChart que referencia esa tabla.

    Args:
        wb: Workbook de openpyxl
        datos_tendencia (dict): Datos de obtener_tendencia_semanal()
        tipo (str): 'ingresos' o 'egresos' — qué columna de datos usar
        titulo (str): Título del gráfico
        color_barras (str): Color hex de las barras (sin #)
    """
    nombre_hoja = f'Gráfico {tipo.capitalize()}'
    ws = wb.create_sheet(nombre_hoja)

    etiquetas = datos_tendencia.get('etiquetas', [])
    valores = datos_tendencia.get(tipo, [])

    # ---- Tabla de datos (necesaria para que el gráfico tenga referencias) ----
    ws.cell(row=1, column=1, value='Semana').font = Font(bold=True, size=10)
    ws.cell(row=1, column=2, value=titulo).font = Font(bold=True, size=10)

    for fila_idx, (etiqueta, valor) in enumerate(zip(etiquetas, valores), start=2):
        ws.cell(row=fila_idx, column=1, value=etiqueta)
        ws.cell(row=fila_idx, column=2, value=valor)

    # ---- Crear gráfico de barras ----
    chart = BarChart()
    chart.type = 'col'          # Barras verticales
    chart.grouping = 'clustered'
    chart.title = titulo
    chart.y_axis.title = 'Equipos'
    chart.x_axis.title = 'Semana del Año'
    chart.width = 28            # Ancho en cm
    chart.height = 15           # Alto en cm
    chart.style = 10            # Estilo visual de openpyxl

    num_filas = len(etiquetas)

    # Datos del eje Y (valores)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=1 + num_filas)
    chart.add_data(data_ref, titles_from_data=True)

    # Etiquetas del eje X (semanas)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + num_filas)
    chart.set_categories(cats_ref)

    # Insertar gráfico en la hoja (posición D2)
    ws.add_chart(chart, 'D2')

    # Ancho de columnas de la tabla
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14


# ===========================================================================
# HOJA 5: REPORTE MENSUAL (tabla por mes + gráficas de ingresos y egresos)
# ===========================================================================

def _crear_hoja_mensual(wb, datos_mensual):
    """
    Crea la hoja 'Reporte Mensual' con:
      - Tabla de totales por mes agrupados en quarters (Q1-Q4)
      - Gráfica de barras de Ingresos por mes (debajo de la tabla)
      - Gráfica de barras de Egresos por mes (a la derecha de la primera gráfica)

    EXPLICACIÓN PARA PRINCIPIANTES:
    openpyxl no tiene gráficas flotantes reales; los gráficos se anclan a una
    celda. Colocamos los datos de la tabla en columnas A-C, dejamos espacio
    libre a la derecha para las gráficas y luego las pegamos con add_chart().

    Args:
        wb: Workbook de openpyxl
        datos_mensual (dict): Resultado de obtener_reporte_mensual()
    """
    ws = wb.create_sheet('Reporte Mensual')
    año = datos_mensual.get('año', '')

    # ---- Fila 1: Título general ----
    titulo_cell = ws.cell(row=1, column=1, value=f'Reporte Mensual de Equipos — {año}')
    apply_cell_style(titulo_cell, {
        'font': Font(bold=True, size=14, color='212529'),
        'alignment': Alignment(horizontal='left', vertical='center'),
    })
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.row_dimensions[1].height = 22

    # ---- Fila 2: encabezado de columnas ----
    encabezados = ['QUARTER', 'MES', 'INGRESOS', 'EGRESOS', 'DIFERENCIA']
    for col_idx, texto in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col_idx, value=texto)
        apply_cell_style(cell, _get_subheader_style('dee2e6'))

    # Anchos de columna fijos
    ws.column_dimensions['A'].width = 18   # QUARTER
    ws.column_dimensions['B'].width = 14   # MES
    ws.column_dimensions['C'].width = 14   # INGRESOS
    ws.column_dimensions['D'].width = 14   # EGRESOS
    ws.column_dimensions['E'].width = 14   # DIFERENCIA

    fila_actual = 3

    orden_quarters = ['Q1', 'Q2', 'Q3', 'Q4']

    for q_key in orden_quarters:
        q_data = datos_mensual.get(q_key)
        if not q_data:
            continue

        color_q = q_data['color']
        nombre_q = q_data['nombre']
        meses = q_data['meses']
        total_ing = q_data['total_ingreso']
        total_egr = q_data['total_egreso']
        num_meses = len(meses)

        # ---- Celda de QUARTER con merge vertical ----
        celda_q = ws.cell(row=fila_actual, column=1, value=nombre_q)
        apply_cell_style(celda_q, _get_quarter_header_style(color_q))
        if num_meses > 1:
            ws.merge_cells(
                start_row=fila_actual,
                start_column=1,
                end_row=fila_actual + num_meses - 1,
                end_column=1,
            )

        # ---- Filas de meses ----
        for mes_data in meses:
            ingreso = mes_data['ingreso']
            egreso = mes_data['egreso']
            diferencia = ingreso - egreso

            ws.cell(row=fila_actual, column=2, value=mes_data['nombre'])
            apply_cell_style(ws.cell(row=fila_actual, column=2), _get_data_cell_style())

            ws.cell(row=fila_actual, column=3, value=ingreso)
            apply_cell_style(ws.cell(row=fila_actual, column=3), _get_data_cell_style())

            ws.cell(row=fila_actual, column=4, value=egreso)
            apply_cell_style(ws.cell(row=fila_actual, column=4), _get_data_cell_style())

            ws.cell(row=fila_actual, column=5, value=diferencia)
            # La diferencia se colorea: negativo en rojo suave, positivo en verde suave
            dif_style = _get_data_cell_style()
            if diferencia < 0:
                dif_style['font'] = Font(size=10, bold=True, color='842029')
            elif diferencia > 0:
                dif_style['font'] = Font(size=10, bold=True, color='0f5132')
            apply_cell_style(ws.cell(row=fila_actual, column=5), dif_style)

            fila_actual += 1

        # ---- Fila de subtotal del quarter ----
        ws.cell(row=fila_actual, column=1, value='SUBTOTAL')
        apply_cell_style(ws.cell(row=fila_actual, column=1), _get_total_row_style())

        ws.cell(row=fila_actual, column=2, value=nombre_q)
        apply_cell_style(ws.cell(row=fila_actual, column=2), _get_total_row_style())

        ws.cell(row=fila_actual, column=3, value=total_ing)
        apply_cell_style(ws.cell(row=fila_actual, column=3), _get_total_row_style())

        ws.cell(row=fila_actual, column=4, value=total_egr)
        apply_cell_style(ws.cell(row=fila_actual, column=4), _get_total_row_style())

        ws.cell(row=fila_actual, column=5, value=total_ing - total_egr)
        apply_cell_style(ws.cell(row=fila_actual, column=5), _get_total_row_style())

        fila_actual += 1

    # ---- Fila de TOTAL ANUAL ----
    meses_lista = datos_mensual.get('meses_lista', [])
    gran_total_ing = sum(m['ingreso'] for m in meses_lista)
    gran_total_egr = sum(m['egreso'] for m in meses_lista)

    ws.cell(row=fila_actual, column=1, value='TOTAL ANUAL')
    apply_cell_style(ws.cell(row=fila_actual, column=1), _get_quarter_header_style('212529'))
    # IMPORTANTE: el merge se hace DESPUÉS de escribir el valor en la celda ancla (col 1).
    # No se debe escribir en ninguna celda subordinada del rango mergeado (col 2 en este caso).
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=2)

    ws.cell(row=fila_actual, column=3, value=gran_total_ing)
    apply_cell_style(ws.cell(row=fila_actual, column=3), _get_quarter_header_style('212529'))

    ws.cell(row=fila_actual, column=4, value=gran_total_egr)
    apply_cell_style(ws.cell(row=fila_actual, column=4), _get_quarter_header_style('212529'))

    ws.cell(row=fila_actual, column=5, value=gran_total_ing - gran_total_egr)
    apply_cell_style(ws.cell(row=fila_actual, column=5), _get_quarter_header_style('212529'))

    fila_actual += 1

    # =========================================================================
    # DATOS AUXILIARES para los gráficos (columnas G y H — fuera de la tabla)
    # =========================================================================
    # openpyxl necesita que los datos estén en celdas de la hoja para referenciarlos.
    # Escribimos los 12 meses en columnas G (etiqueta), H (ingresos), I (egresos).

    col_etiqueta = 7   # Columna G
    col_ingresos = 8   # Columna H
    col_egresos  = 9   # Columna I

    # Encabezados (fila 1 de datos auxiliares = fila 1 de la hoja)
    ws.cell(row=1, column=col_etiqueta, value='Mes')
    ws.cell(row=1, column=col_ingresos, value='Ingresos')
    ws.cell(row=1, column=col_egresos, value='Egresos')
    for c in [col_etiqueta, col_ingresos, col_egresos]:
        apply_cell_style(ws.cell(row=1, column=c), _get_subheader_style('dee2e6'))

    fila_datos_graf = 2
    for mes_data in meses_lista:
        ws.cell(row=fila_datos_graf, column=col_etiqueta, value=mes_data['nombre'])
        ws.cell(row=fila_datos_graf, column=col_ingresos, value=mes_data['ingreso'])
        ws.cell(row=fila_datos_graf, column=col_egresos,  value=mes_data['egreso'])
        fila_datos_graf += 1

    num_meses_total = len(meses_lista)  # Normalmente 12
    fila_fin_datos  = 1 + num_meses_total  # Última fila con datos (inclusive)

    # =========================================================================
    # GRÁFICA 1 — Ingresos por mes
    # =========================================================================
    chart_ing = BarChart()
    chart_ing.type = 'col'
    chart_ing.grouping = 'clustered'
    chart_ing.title = f'Ingresos por Mes — {año}'
    chart_ing.y_axis.title = 'Equipos'
    chart_ing.x_axis.title = 'Mes'
    chart_ing.width = 18
    chart_ing.height = 12
    chart_ing.style = 10

    data_ing = Reference(ws, min_col=col_ingresos, min_row=1, max_row=fila_fin_datos)
    chart_ing.add_data(data_ing, titles_from_data=True)
    cats = Reference(ws, min_col=col_etiqueta, min_row=2, max_row=fila_fin_datos)
    chart_ing.set_categories(cats)

    # Posicionar debajo de la tabla (con 2 filas de margen)
    fila_graficas = fila_actual + 2
    ancla_ing = f'A{fila_graficas}'
    ws.add_chart(chart_ing, ancla_ing)

    # =========================================================================
    # GRÁFICA 2 — Egresos por mes
    # =========================================================================
    chart_egr = BarChart()
    chart_egr.type = 'col'
    chart_egr.grouping = 'clustered'
    chart_egr.title = f'Egresos por Mes — {año}'
    chart_egr.y_axis.title = 'Equipos'
    chart_egr.x_axis.title = 'Mes'
    chart_egr.width = 18
    chart_egr.height = 12
    chart_egr.style = 10

    data_egr = Reference(ws, min_col=col_egresos, min_row=1, max_row=fila_fin_datos)
    chart_egr.add_data(data_egr, titles_from_data=True)
    chart_egr.set_categories(cats)

    # Posicionar a la derecha de la primera gráfica (~10 columnas = columna K)
    ancla_egr = f'K{fila_graficas}'
    ws.add_chart(chart_egr, ancla_egr)

    # Ancho de las columnas auxiliares de datos
    ws.column_dimensions[get_column_letter(col_etiqueta)].width = 12
    ws.column_dimensions[get_column_letter(col_ingresos)].width = 12
    ws.column_dimensions[get_column_letter(col_egresos)].width = 12


# ===========================================================================
# FUNCIÓN PRINCIPAL EXPORTADA
# ===========================================================================

def generar_excel_concentrado(datos_semana, datos_trimestral, datos_tendencia, datos_mensual=None):
    """
    Genera el archivo Excel completo del Concentrado Semanal con 4-5 hojas.

    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta función es el punto de entrada. Recibe los datos calculados por
    'concentrado_semanal.py' y los convierte en un workbook de Excel con
    hojas completamente formateadas.

    Args:
        datos_semana (dict): Resultado de obtener_concentrado_semanal()
            Contiene: ingreso, egreso, asignacion, totales, fechas, resumen.

        datos_trimestral (dict): Resultado de obtener_reporte_trimestral()
            Contiene: Q1, Q2, Q3, Q4 con ingreso/egreso por sitio×tipo.

        datos_tendencia (dict): Resultado de obtener_tendencia_semanal()
            Contiene: etiquetas, semanas, ingresos, egresos por semana del año.

        datos_mensual (dict, optional): Resultado de obtener_reporte_mensual()
            Si se pasa, se agrega una hoja extra 'Reporte Mensual' con tabla y gráficas.
            Contiene: año, Q1-Q4 por mes, meses_lista.

    Returns:
        openpyxl.Workbook: Workbook listo para guardar o enviar como respuesta HTTP.

    Uso:
        wb = generar_excel_concentrado(datos_semana, datos_trimestral, datos_tendencia, datos_mensual)
        wb.save('concentrado.xlsx')
        # O para HTTP:
        wb.save(response)
    """
    wb = openpyxl.Workbook()

    # Hoja 1: Concentrado Semanal
    _crear_hoja_concentrado(wb, datos_semana)

    # Hoja 2: Reporte Trimestral
    _crear_hoja_trimestral(wb, datos_trimestral)

    # Hoja 3: Gráfico Ingresos
    num_semana = datos_semana.get('numero_semana', '?')
    año = datos_semana.get('año', '?')
    _crear_hoja_grafico(
        wb,
        datos_tendencia,
        tipo='ingresos',
        titulo=f'Ingresos de Equipos por Semana — {año}',
        color_barras='0d6efd',
    )

    # Hoja 4: Gráfico Egresos
    _crear_hoja_grafico(
        wb,
        datos_tendencia,
        tipo='egresos',
        titulo=f'Egresos de Equipos por Semana — {año}',
        color_barras='198754',
    )

    # Hoja 5: Reporte Mensual (solo si se proporcionaron los datos)
    if datos_mensual:
        _crear_hoja_mensual(wb, datos_mensual)

    return wb
