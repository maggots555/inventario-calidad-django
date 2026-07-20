"""
Dashboard de cotizaciones (Plotly/ML) + exports Excel (Fase 8).

Incluye exports grandes de rechazos y aceptaciones.
urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
"""

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from inventario.models import Empleado, Sucursal

from .decorators import cache_page_dashboard, permission_required_with_message
from .models import OrdenServicio


# ============================================================================
# 📊 DASHBOARD DE COTIZACIONES - ANALYTICS CON PLOTLY Y MACHINE LEARNING
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@cache_page_dashboard
def dashboard_cotizaciones(request):
    """
    Dashboard analítico completo de cotizaciones tipo Power BI.

    Esta vista es el "cerebro" del dashboard. Hace lo siguiente:

    Query Parameters (filtros en URL):
        - fecha_inicio: Fecha inicio filtro (YYYY-MM-DD)
        - fecha_fin: Fecha fin filtro (YYYY-MM-DD)
        - sucursal: ID de sucursal
        - tecnico: ID de técnico
        - gama: Gama de equipo (alta/media/baja)
        - periodo: Agrupación temporal (D/W/M/Q/Y)
    
    Returns:
        HttpResponse: Página renderizada con el dashboard completo
    
    Ejemplo de URL:
        /cotizaciones/dashboard/?fecha_inicio=2025-01-01&fecha_fin=2025-12-31&sucursal=1&periodo=M
    """
    
    from datetime import datetime, timedelta
    import pandas as pd  # Necesario para pd.DataFrame() en bloques except
    from .utils_cotizaciones import (
        obtener_dataframe_cotizaciones,
        calcular_kpis_generales,
        analizar_piezas_cotizadas,
        analizar_proveedores,
        calcular_metricas_por_tecnico,
        calcular_metricas_por_sucursal,
        calcular_metricas_por_responsable,
        calcular_kpis_aceptaciones,
        analizar_servicios_vm_aceptadas,
        analizar_seguimiento_piezas_aceptadas
    )
    from .plotly_visualizations import DashboardCotizacionesVisualizer, convertir_figura_a_html
    from .ml_predictor import PredictorAceptacionCotizacion
    
    # NUEVO: Módulos ML Avanzados (Sistema Experto)
    from .ml_advanced import (
        PredictorMotivoRechazo,
        OptimizadorPrecios,
        RecomendadorAcciones
    )
    
    # ========================================
    # 1. OBTENER Y VALIDAR FILTROS DEL REQUEST
    # ========================================
    
    # Fechas por defecto: últimos 3 meses (timezone-aware)
    # EXPLICACIÓN PARA PRINCIPIANTES:
    # timezone.now() devuelve datetime con zona horaria (timezone-aware)
    # Esto previene warnings de Django cuando se compara con DateTimeFields
    from django.utils import timezone as tz
    fecha_fin_default = tz.now().date()
    fecha_inicio_default = (tz.now() - timedelta(days=90)).date()
    
    # Capturar parámetros GET
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    tecnico_id = request.GET.get('tecnico')
    gama = request.GET.get('gama')
    periodo = request.GET.get('periodo', 'M')  # Default: Mensual
    
    # Validar y parsear fechas (convertir a timezone-aware datetime)
    # EXPLICACIÓN: Los DateTimeFields en Django requieren datetimes timezone-aware
    # para evitar warnings. Convertimos date → datetime → timezone-aware
    try:
        if fecha_inicio_str:
            fecha_dt = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_inicio = tz.make_aware(fecha_dt)
        else:
            # Convertir date a datetime timezone-aware (inicio del día)
            fecha_dt = datetime.combine(fecha_inicio_default, datetime.min.time())
            fecha_inicio = tz.make_aware(fecha_dt)
    except ValueError:
        fecha_dt = datetime.combine(fecha_inicio_default, datetime.min.time())
        fecha_inicio = tz.make_aware(fecha_dt)
    
    try:
        if fecha_fin_str:
            fecha_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            # Para fecha_fin, usar fin del día (23:59:59.999999)
            fecha_dt = datetime.combine(fecha_dt.date(), datetime.max.time())
            fecha_fin = tz.make_aware(fecha_dt)
        else:
            # Convertir date a datetime timezone-aware (fin del día)
            fecha_dt = datetime.combine(fecha_fin_default, datetime.max.time())
            fecha_fin = tz.make_aware(fecha_dt)
    except ValueError:
        fecha_dt = datetime.combine(fecha_fin_default, datetime.max.time())
        fecha_fin = tz.make_aware(fecha_dt)
    
    # Validar período
    if periodo not in ['D', 'W', 'M', 'Q', 'Y']:
        periodo = 'M'
    
    # Convertir IDs a enteros si existen
    try:
        sucursal_id = int(sucursal_id) if sucursal_id else None
    except (ValueError, TypeError):
        sucursal_id = None
    
    try:
        tecnico_id = int(tecnico_id) if tecnico_id else None
    except (ValueError, TypeError):
        tecnico_id = None
    
    # ========================================
    # 2. OBTENER DATOS CON FILTROS
    # ========================================
    
    try:
        # Obtener DataFrame principal de cotizaciones
        df_cotizaciones = obtener_dataframe_cotizaciones(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            sucursal_id=sucursal_id,
            tecnico_id=tecnico_id,
            gama=gama
        )
        
        # Obtener IDs de cotizaciones para análisis relacionados
        cotizacion_ids = df_cotizaciones['cotizacion_id'].tolist() if not df_cotizaciones.empty else []
        
        # Análisis de piezas
        df_piezas = analizar_piezas_cotizadas(cotizacion_ids)
        
        # Análisis de proveedores
        df_seguimientos = analizar_proveedores(cotizacion_ids)
        
    except Exception as e:
        messages.error(request, f'Error al obtener datos: {str(e)}')
        df_cotizaciones = pd.DataFrame()
        df_piezas = pd.DataFrame()
        df_seguimientos = pd.DataFrame()
    
    # ========================================
    # 3. CALCULAR KPIs Y MÉTRICAS
    # ========================================
    
    if not df_cotizaciones.empty:
        # KPIs generales
        kpis = calcular_kpis_generales(df_cotizaciones)
        
        # Métricas por técnico
        df_metricas_tecnicos = calcular_metricas_por_tecnico(df_cotizaciones)
        
        # Métricas por sucursal
        df_metricas_sucursales = calcular_metricas_por_sucursal(df_cotizaciones)
        
        # Métricas por responsable de seguimiento
        df_metricas_responsables = calcular_metricas_por_responsable(df_cotizaciones)
    else:
        kpis = {
            'total_cotizaciones': 0,
            'aceptadas': 0,
            'rechazadas': 0,
            'pendientes': 0,
            'tasa_aceptacion': 0,
            'tasa_rechazo': 0,
            'valor_total_cotizado': 0,
            'valor_total_cotizado_fmt': '$0',
            'ticket_promedio': 0,
            'ticket_promedio_fmt': '$0'
        }
        df_metricas_tecnicos = pd.DataFrame()
        df_metricas_sucursales = pd.DataFrame()
        df_metricas_responsables = pd.DataFrame()
    
    # ========================================
    # 3.5. KPIs DE ACEPTACIONES Y ANÁLISIS VM
    # ========================================
    
    kpis_aceptaciones = {}
    analisis_vm = {}
    analisis_seguimiento = {}
    
    if not df_cotizaciones.empty:
        try:
            print("\n" + "="*50)
            print("✅ ANÁLISIS DE ACEPTACIONES Y VENTAS MOSTRADOR")
            print("="*50)
            
            kpis_aceptaciones = calcular_kpis_aceptaciones(df_cotizaciones)
            print(f"   - KPIs aceptaciones calculados: {kpis_aceptaciones.get('total_aceptadas', 0)} aceptadas")
            
            analisis_vm = analizar_servicios_vm_aceptadas(
                df_cotizaciones,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                sucursal_id=sucursal_id,
                tecnico_id=tecnico_id,
                gama=gama,
            )
            print(f"   - Análisis VM: {analisis_vm.get('total_con_vm', 0)} con venta mostrador")
            
            analisis_seguimiento = analizar_seguimiento_piezas_aceptadas(df_cotizaciones)
            print(f"   - Seguimiento piezas: {analisis_seguimiento.get('total_piezas_rastreadas', 0)} piezas rastreadas")
            
            print("✅ Análisis de aceptaciones completado")
            
        except Exception as e_acept:
            print(f"⚠️ Error en análisis de aceptaciones: {str(e_acept)}")
            kpis_aceptaciones = {}
            analisis_vm = {}
            analisis_seguimiento = {}
    
    # ========================================
    # 4. GENERAR VISUALIZACIONES
    # ========================================
    
    visualizer = DashboardCotizacionesVisualizer()
    graficos = {}
    
    if not df_cotizaciones.empty:
        try:
            # Usar función orquestadora para generar todos los gráficos
            graficos = visualizer.crear_dashboard_completo(
                df=df_cotizaciones,
                df_piezas=df_piezas if not df_piezas.empty else None,
                df_seguimientos=df_seguimientos if not df_seguimientos.empty else None,
                df_metricas_tecnicos=df_metricas_tecnicos if not df_metricas_tecnicos.empty else None,
                df_metricas_sucursales=df_metricas_sucursales if not df_metricas_sucursales.empty else None,
                df_metricas_responsables=df_metricas_responsables if not df_metricas_responsables.empty else None,
                kpis=kpis,
                ml_predictor=None,  # Lo agregamos después
                periodo=periodo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                sucursal_id=sucursal_id,
                tecnico_id=tecnico_id,
                gama=gama,
            )
        except Exception as e:
            messages.warning(request, f'Algunos gráficos no se pudieron generar: {str(e)}')
            print(f"⚠️ Error generando gráficos: {str(e)}")
    else:
        # Sin datos, mostrar mensaje
        messages.info(request, 'No hay datos de cotizaciones con los filtros aplicados.')
    
    # ========================================
    # 5. MACHINE LEARNING (Si hay datos suficientes)
    # ========================================
    
    ml_insights = {
        'modelo_disponible': False,
        'accuracy': 0,
        'sugerencias': []
    }
    
    # NUEVO: Insights avanzados (sistema experto)
    ml_insights_avanzados = {
        'disponible': False,
        'predictor_motivos_disponible': False,
        'optimizador_disponible': False,
        'recomendador_disponible': False,
        'analisis_completo': None
    }
    
    if not df_cotizaciones.empty and len(df_cotizaciones) >= 20:
        try:
            # Inicializar predictor base
            predictor = PredictorAceptacionCotizacion()
            
            # Intentar cargar modelo existente
            try:
                predictor.cargar_modelo()
                print("✅ Modelo ML base cargado exitosamente")
            except FileNotFoundError:
                # Si no existe, entrenar con datos actuales
                print("⚠️ No se encontró modelo pre-entrenado, entrenando nuevo modelo...")
                predictor.entrenar_modelo(
                    fecha_inicio=fecha_inicio.strftime('%Y-%m-%d'),
                    fecha_fin=fecha_fin.strftime('%Y-%m-%d')
                )
            
            # Obtener métricas del modelo
            metricas_ml = predictor.obtener_metricas()
            
            # Generar gráfico de factores influyentes
            feature_importance = predictor.obtener_factores_influyentes(top_n=10)
            if feature_importance:
                graficos['factores_influyentes'] = convertir_figura_a_html(
                    visualizer.grafico_factores_influyentes(feature_importance)
                )
            
            # Generar sugerencias
            sugerencias = predictor.generar_sugerencias(df_cotizaciones)
            
            ml_insights = {
                'modelo_disponible': True,
                'accuracy': metricas_ml.get('accuracy', 0) * 100,  # Convertir a porcentaje
                'precision': metricas_ml.get('precision', 0) * 100,
                'recall': metricas_ml.get('recall', 0) * 100,
                'f1_score': metricas_ml.get('f1_score', 0) * 100,
                'total_muestras': metricas_ml.get('total_muestras', 0),
                'datos_entrenamiento': metricas_ml.get('total_muestras', 0),  # Agregado para el template
                'fecha_entrenamiento': metricas_ml.get('fecha_entrenamiento', ''),
                'sugerencias': sugerencias,
                'feature_importance': feature_importance
            }
            
            # Predicción de ejemplo (última cotización pendiente)
            df_pendientes = df_cotizaciones[df_cotizaciones['aceptada'].isna()]
            if not df_pendientes.empty:
                ultima = df_pendientes.iloc[-1]
                features_ejemplo = {
                    'costo_total': ultima['costo_total'],
                    'costo_mano_obra': ultima['costo_mano_obra'],
                    'costo_total_piezas': ultima['costo_total_piezas'],
                    'total_piezas': ultima['total_piezas'],
                    'piezas_necesarias': ultima['piezas_necesarias'],
                    'porcentaje_necesarias': ultima['porcentaje_necesarias'],
                    'piezas_sugeridas_tecnico': ultima['piezas_sugeridas_tecnico'],
                    'descontar_mano_obra': ultima['descontar_mano_obra'],
                    'gama': ultima['gama'],
                    'tipo_equipo': ultima['tipo_equipo'],
                }
                
                prob_rechazo, prob_aceptacion = predictor.predecir_probabilidad(features_ejemplo)
                
                graficos['prediccion_ml_ejemplo'] = convertir_figura_a_html(
                    visualizer.grafico_prediccion_ml(prob_aceptacion, prob_rechazo)
                )
                
                # CORRECCIÓN: Cambiar 'ejemplo_prediccion' a 'prediccion_ejemplo' para que coincida con el template
                ml_insights['prediccion_ejemplo'] = {
                    'cotizacion_id': ultima['cotizacion_id'],
                    'orden': ultima['numero_orden'],
                    'orden_cliente': ultima['orden_cliente'],  # AGREGADO: Campo orden_cliente del DataFrame
                    'costo': ultima['costo_total'],
                    'prob_aceptacion': prob_aceptacion * 100,
                    'prob_rechazo': prob_rechazo * 100
                }
                
                # ========================================
                # 5.1. MÓDULOS ML AVANZADOS (Sistema Experto)
                # ========================================
                
                print("\n🔬 Iniciando análisis con módulos ML avanzados...")
                
                try:
                    # Inicializar el Recomendador (orquestador que carga todo)
                    recomendador = RecomendadorAcciones(predictor_base=predictor)
                    
                    # Análisis completo de la cotización pendiente
                    analisis_completo = recomendador.analizar_cotizacion_completa(
                        cotizacion_features=features_ejemplo,
                        incluir_optimizacion_precio=True,
                        incluir_analisis_temporal=True
                    )
                    
                    # Actualizar insights avanzados
                    ml_insights_avanzados.update({
                        'disponible': True,
                        'predictor_motivos_disponible': recomendador.predictor_motivos is not None,
                        'optimizador_disponible': recomendador.optimizador is not None,
                        'recomendador_disponible': True,
                        'analisis_completo': analisis_completo,
                        
                        # Extraer datos clave para fácil acceso en template
                        'prob_aceptacion': analisis_completo['prediccion_base']['prob_aceptacion_pct'],
                        'clasificacion': analisis_completo['prediccion_base']['clasificacion'],
                        'total_recomendaciones': len(analisis_completo['recomendaciones']),
                        'recomendaciones_criticas': len([
                            r for r in analisis_completo['recomendaciones'] 
                            if r['nivel'] <= 2
                        ]),
                        'total_alertas': len(analisis_completo['alertas_criticas']),
                        'resumen_ejecutivo': analisis_completo['resumen_ejecutivo'],
                        
                        # Datos de cotización analizada (para mostrar en UI)
                        'cotizacion_analizada': {
                            'id': ultima['cotizacion_id'],
                            'orden': ultima['numero_orden'],
                            'orden_cliente': ultima['orden_cliente'],  # AGREGADO: Campo orden_cliente
                            'costo_actual': ultima['costo_total'],
                            'total_piezas': ultima['total_piezas'],
                            'gama': ultima['gama'],
                        }
                    })
                    
                    # Si hay predicción de motivo, agregarlo
                    if analisis_completo['prediccion_motivo']:
                        # Convertir probabilidad de decimal a porcentaje (0.255 -> 25.5)
                        prob_numerica = analisis_completo['prediccion_motivo']['probabilidad'] * 100
                        
                        ml_insights_avanzados['motivo_predicho'] = {
                            'motivo': analisis_completo['prediccion_motivo']['motivo_principal'],
                            'motivo_nombre': analisis_completo['prediccion_motivo']['motivo_nombre'],
                            'probabilidad': prob_numerica,  # Valor numérico para el progress bar
                            'probabilidad_texto': analisis_completo['prediccion_motivo']['probabilidad_pct'],  # Texto formateado
                            'confianza': analisis_completo['prediccion_motivo']['confianza'],
                            'confianza_icono': analisis_completo['prediccion_motivo']['confianza_icono'],
                            'descripcion': analisis_completo['prediccion_motivo']['motivo_descripcion'],
                            'acciones': analisis_completo['prediccion_motivo']['acciones_sugeridas']
                        }
                    
                    # Si hay optimización de precio, agregarlo
                    if analisis_completo['optimizacion_precio']:
                        opt = analisis_completo['optimizacion_precio']
                        ml_insights_avanzados['optimizacion'] = {
                            'costo_actual': opt['costo_actual'],
                            'costo_optimo': opt['escenario_optimo']['costo_final'],
                            'mejora_ingreso': opt['mejora_ingreso'],
                            'mejora_probabilidad': opt['mejora_probabilidad_pct'],
                            'escenario_optimo': opt['escenario_optimo'],
                            'escenario_conservador': opt['escenario_conservador'],
                            'escenario_agresivo': opt['escenario_agresivo'],
                            'total_escenarios': opt['total_escenarios_evaluados']
                        }
                    
                    # Si hay análisis temporal, agregarlo
                    if analisis_completo['analisis_temporal']:
                        temp = analisis_completo['analisis_temporal']
                        ml_insights_avanzados['temporal'] = {
                            'dia_hoy': temp['dia_hoy'],
                            'es_dia_optimo': temp['es_dia_optimo'],
                            'mejor_dia': temp['mejor_dia'],
                            'mejora_potencial': temp['mejora_potencial'],
                            'recomendacion': temp['recomendacion'],
                            'mensaje': temp['mensaje']
                        }
                    
                    print(f"✅ Análisis ML avanzado completado:")
                    print(f"   - {ml_insights_avanzados['total_recomendaciones']} recomendaciones generadas")
                    print(f"   - {ml_insights_avanzados['total_alertas']} alertas críticas")
                    print(f"   - Estado: {ml_insights_avanzados['resumen_ejecutivo']['estado_mensaje']}")
                    
                    # Mensaje informativo para el usuario
                    if ml_insights_avanzados['total_alertas'] > 0:
                        messages.warning(
                            request,
                            f"⚠️ {ml_insights_avanzados['total_alertas']} alertas críticas detectadas en ML avanzado"
                        )
                    
                    # ========================================
                    # 5.2. GENERAR VISUALIZACIONES ML AVANZADAS
                    # ========================================
                    
                    print("📊 Generando visualizaciones ML avanzadas...")
                    
                    try:
                        # Gráfico de escenarios de precio
                        if analisis_completo['optimizacion_precio']:
                            graficos['ml_escenarios_precio'] = convertir_figura_a_html(
                                visualizer.grafico_escenarios_precio(
                                    analisis_completo['optimizacion_precio']
                                )
                            )
                            print("   ✅ Gráfico de escenarios de precio generado")
                        
                        # Matriz riesgo-beneficio
                        graficos['ml_matriz_riesgo'] = convertir_figura_a_html(
                            visualizer.grafico_matriz_riesgo_beneficio(analisis_completo)
                        )
                        print("   ✅ Matriz riesgo-beneficio generada")
                        
                        # Timeline de probabilidad por día
                        if analisis_completo['analisis_temporal']:
                            graficos['ml_probabilidad_dia'] = convertir_figura_a_html(
                                visualizer.grafico_probabilidad_por_dia(
                                    analisis_completo['analisis_temporal']
                                )
                            )
                            print("   ✅ Timeline probabilidad por día generado")
                        
                        print("✅ Todas las visualizaciones ML avanzadas generadas exitosamente")
                        
                    except Exception as e_viz:
                        print(f"⚠️ Error generando visualizaciones ML avanzadas: {str(e_viz)}")
                        # No crítico, continuar
                    
                except Exception as e_avanzado:
                    print(f"⚠️ Error en módulos ML avanzados: {str(e_avanzado)}")
                    print(f"   Stack trace: {e_avanzado.__class__.__name__}")
                    # No fallar todo el dashboard, solo deshabilitar módulos avanzados
                    ml_insights_avanzados['error'] = str(e_avanzado)
        
        except Exception as e:
            print(f"⚠️ Error en Machine Learning: {str(e)}")
            messages.warning(request, f'Machine Learning no disponible: {str(e)}')
    
    # ========================================
    # 6. PREPARAR DATOS PARA FILTROS
    # ========================================
    
    # Listas para desplegables
    sucursales = Sucursal.objects.all().order_by('nombre')
    tecnicos = Empleado.objects.filter(
        ordenes_tecnico__isnull=False
    ).distinct().order_by('nombre_completo')
    
    # Opciones de gama
    gamas = [
        ('alta', 'Alta'),
        ('media', 'Media'),
        ('baja', 'Baja')
    ]
    
    # Opciones de período
    periodos = [
        ('D', 'Diario'),
        ('W', 'Semanal'),
        ('M', 'Mensual'),
        ('Q', 'Trimestral'),
        ('Y', 'Anual')
    ]
    
    # ========================================
    # 6.5. ANÁLISIS DE TEXTO (TEXT MINING)
    # ========================================
    
    print("\n" + "="*50)
    print("📝 ANÁLISIS DE COMENTARIOS DE RECHAZO (TEXT MINING)")
    print("="*50)
    
    analisis_texto = {}
    
    try:
        from .utils_cotizaciones import analizar_comentarios_rechazo
        
        print("🔍 Analizando comentarios de rechazo...")
        
        # Llamar función de análisis de texto
        analisis_texto = analizar_comentarios_rechazo(df_cotizaciones)
        
        if analisis_texto['tiene_datos']:
            print(f"✅ Análisis de texto completado:")
            print(f"   - {analisis_texto['total_comentarios']} comentarios analizados")
            print(f"   - {analisis_texto['total_palabras_unicas']} palabras únicas encontradas")
            print(f"   - {len(analisis_texto['palabras_clave'])} palabras clave extraídas")
            print(f"   - {len(analisis_texto['frases_comunes'])} frases comunes identificadas")
            print(f"   - {len(analisis_texto['insights'])} insights generados")
            
            # Generar visualizaciones de text mining
            try:
                print("📊 Generando visualizaciones de text mining...")
                
                # Gráfico de palabras más frecuentes
                graficos['texto_palabras_frecuentes'] = convertir_figura_a_html(
                    visualizer.grafico_palabras_frecuentes(analisis_texto['palabras_clave'])
                )
                print("   ✅ Gráfico de palabras frecuentes generado")
                
                # Gráfico de frases comunes
                if analisis_texto['frases_comunes']:
                    graficos['texto_frases_comunes'] = convertir_figura_a_html(
                        visualizer.grafico_frases_comunes(analisis_texto['frases_comunes'])
                    )
                    print("   ✅ Gráfico de frases comunes generado")
                
                # Gráfico de correlación palabras → resultado
                if analisis_texto['correlaciones']:
                    graficos['texto_correlaciones'] = convertir_figura_a_html(
                        visualizer.grafico_correlacion_palabras(analisis_texto['correlaciones'])
                    )
                    print("   ✅ Gráfico de correlaciones generado")
                
                # Nube de palabras tipo burbujas
                graficos['texto_nube_palabras'] = convertir_figura_a_html(
                    visualizer.grafico_nube_palabras_simple(analisis_texto['palabras_clave'])
                )
                print("   ✅ Nube de palabras generada")
                
                print("✅ Todas las visualizaciones de text mining generadas exitosamente")
                
            except Exception as e_viz_texto:
                print(f"⚠️ Error generando visualizaciones de text mining: {str(e_viz_texto)}")
                # No crítico, continuar
        
        else:
            print("ℹ️ No hay suficientes comentarios de rechazo para análisis de texto")
            analisis_texto['mensaje'] = "No hay comentarios de rechazo suficientes para análisis"
    
    except Exception as e_texto:
        print(f"⚠️ Error en análisis de texto: {str(e_texto)}")
        analisis_texto = {
            'tiene_datos': False,
            'error': str(e_texto),
            'mensaje': 'Error al analizar comentarios'
        }
    
    # ========================================
    # 6.6. ANÁLISIS DE DIAGNÓSTICOS TÉCNICOS POR TÉCNICO
    # ========================================
    
    print("\n" + "="*50)
    print("🔬 ANÁLISIS DE DIAGNÓSTICOS TÉCNICOS POR TÉCNICO")
    print("="*50)
    
    analisis_diagnosticos = {}
    
    try:
        from .utils_cotizaciones import analizar_diagnosticos_tecnicos
        
        print("🔍 Preparando datos de órdenes de servicio para análisis...")
        
        # Obtener órdenes de servicio con diagnóstico completado
        # NOTA: Usamos tecnico_asignado_actual (siempre presente) en lugar de tecnico_diagnostico (opcional)
        ordenes_con_diagnostico = OrdenServicio.objects.filter(
            fecha_ingreso__gte=fecha_inicio,
            fecha_ingreso__lte=fecha_fin
        ).select_related('tecnico_asignado_actual', 'sucursal', 'detalle_equipo')
        
        print(f"   📋 Total órdenes en el período: {ordenes_con_diagnostico.count()}")
        
        # Aplicar filtros si existen
        if sucursal_id:
            ordenes_con_diagnostico = ordenes_con_diagnostico.filter(sucursal_id=sucursal_id)
            print(f"   🏢 Filtrado por sucursal: {ordenes_con_diagnostico.count()} órdenes")
        
        if tecnico_id:
            # Filtrar por técnico asignado actual (no por tecnico_diagnostico)
            ordenes_con_diagnostico = ordenes_con_diagnostico.filter(tecnico_asignado_actual_id=tecnico_id)
            print(f"   👨‍🔧 Filtrado por técnico: {ordenes_con_diagnostico.count()} órdenes")
        
        # Convertir a DataFrame
        if ordenes_con_diagnostico.exists():
            ordenes_data = []
            ordenes_sin_diagnostico = 0
            ordenes_sin_tecnico = 0
            
            for orden in ordenes_con_diagnostico:
                # Verificar que tenga técnico asignado (tecnico_asignado_actual es obligatorio, siempre existe)
                if not orden.tecnico_asignado_actual:
                    ordenes_sin_tecnico += 1
                    continue
                
                # Verificar que tenga detalle de equipo con diagnóstico
                if not hasattr(orden, 'detalle_equipo'):
                    ordenes_sin_diagnostico += 1
                    continue
                
                diagnostico = orden.detalle_equipo.diagnostico_sic if orden.detalle_equipo.diagnostico_sic else ''
                falla = orden.detalle_equipo.falla_principal if orden.detalle_equipo.falla_principal else ''
                
                # Solo incluir si tiene diagnóstico no vacío
                if diagnostico.strip():
                    ordenes_data.append({
                        'numero_orden': orden.numero_orden_interno,
                        'tecnico_nombre': orden.tecnico_asignado_actual.nombre_completo,
                        'diagnostico_sic': diagnostico,
                        'falla_principal': falla,
                        'fecha_diagnostico': orden.fecha_diagnostico_sic,
                    })
                else:
                    ordenes_sin_diagnostico += 1
            
            print(f"   ✅ {len(ordenes_data)} órdenes con diagnóstico válido")
            if ordenes_sin_tecnico > 0:
                print(f"   ⚠️ {ordenes_sin_tecnico} órdenes sin técnico asignado (excluidas)")
            if ordenes_sin_diagnostico > 0:
                print(f"   ⚠️ {ordenes_sin_diagnostico} órdenes sin diagnóstico escrito (excluidas)")
            
            if not ordenes_data:
                print("❌ No hay órdenes con diagnóstico válido en el período seleccionado")
                analisis_diagnosticos = {
                    'tiene_datos': False,
                    'mensaje': 'No hay órdenes con diagnóstico técnico completado en el período'
                }
            else:
                df_ordenes = pd.DataFrame(ordenes_data)
                print(f"📊 DataFrame creado con {len(df_ordenes)} registros")
                
                # Mostrar técnicos únicos encontrados
                tecnicos_unicos = df_ordenes['tecnico_nombre'].unique()
                print(f"👥 Técnicos encontrados: {', '.join(tecnicos_unicos)}")
                
                # Llamar función de análisis de diagnósticos
                analisis_diagnosticos = analizar_diagnosticos_tecnicos(df_ordenes)
                
                if analisis_diagnosticos['tiene_datos']:
                    print(f"✅ Análisis de diagnósticos completado:")
                    print(f"   - {analisis_diagnosticos['total_diagnosticos']} diagnósticos analizados")
                    print(f"   - {analisis_diagnosticos['total_tecnicos']} técnicos evaluados")
                    print(f"   - Promedio palabras: {analisis_diagnosticos['promedios_globales']['promedio_palabras']:.1f}")
                    print(f"   - Promedio tecnicidad: {analisis_diagnosticos['promedios_globales']['promedio_tecnicidad']:.1f}%")
                    print(f"   - {len(analisis_diagnosticos['insights'])} insights generados")
                    
                    # Generar visualizaciones de diagnósticos
                    try:
                        print("📊 Generando visualizaciones de análisis de diagnósticos...")
                        
                        # Gráfico: Ranking por nivel de detalle
                        graficos['diagnosticos_ranking_detalle'] = convertir_figura_a_html(
                            visualizer.grafico_ranking_tecnicos_detalle(analisis_diagnosticos['analisis_por_tecnico'])
                        )
                        print("   ✅ Ranking de detalle generado")
                        
                        # Gráfico: Ranking por tecnicidad
                        graficos['diagnosticos_ranking_tecnicidad'] = convertir_figura_a_html(
                            visualizer.grafico_ranking_tecnicos_tecnicidad(analisis_diagnosticos['analisis_por_tecnico'])
                        )
                        print("   ✅ Ranking de tecnicidad generado")
                        
                        # Gráfico: Comparativa scatter (detalle vs tecnicidad)
                        graficos['diagnosticos_comparativa_scatter'] = convertir_figura_a_html(
                            visualizer.grafico_comparativa_tecnicos_scatter(analisis_diagnosticos['analisis_por_tecnico'])
                        )
                        print("   ✅ Comparativa scatter generada")
                        
                        # Gráfico: Palabras técnicas globales
                        if analisis_diagnosticos['palabras_tecnicas_globales']:
                            graficos['diagnosticos_palabras_tecnicas'] = convertir_figura_a_html(
                                visualizer.grafico_palabras_tecnicas_globales(analisis_diagnosticos['palabras_tecnicas_globales'])
                            )
                            print("   ✅ Palabras técnicas globales generadas")
                        
                        print("✅ Todas las visualizaciones de diagnósticos generadas exitosamente")
                        
                    except Exception as e_viz_diag:
                        print(f"⚠️ Error generando visualizaciones de diagnósticos: {str(e_viz_diag)}")
                        import traceback
                        print(f"   Detalle: {traceback.format_exc()}")
                        # No crítico, continuar
                
                else:
                    print(f"ℹ️ {analisis_diagnosticos.get('mensaje', 'No hay suficientes diagnósticos')}")
        
        else:
            print("ℹ️ No se encontraron órdenes con diagnóstico en el período seleccionado")
            analisis_diagnosticos = {
                'tiene_datos': False,
                'mensaje': 'No hay órdenes con diagnóstico en el período seleccionado'
            }
    
    except Exception as e_diagnosticos:
        print(f"⚠️ Error en análisis de diagnósticos: {str(e_diagnosticos)}")
        import traceback
        print(f"   Detalle: {traceback.format_exc()}")
        analisis_diagnosticos = {
            'tiene_datos': False,
            'error': str(e_diagnosticos),
            'mensaje': 'Error al analizar diagnósticos técnicos'
        }
    
    # ========================================
    # 7. PREPARAR CONTEXTO COMPLETO
    # ========================================
    
    context = {
        # KPIs
        'kpis': kpis,
        
        # Gráficos (diccionario completo)
        'graficos': graficos,
        
        # Machine Learning (básico)
        'ml_insights': ml_insights,
        
        # Machine Learning Avanzado (Sistema Experto) - NUEVO
        'ml_insights_avanzados': ml_insights_avanzados,
        
        # Análisis de Texto (Text Mining) - NUEVO
        'analisis_texto': analisis_texto,
        
        # Análisis de Diagnósticos Técnicos por Técnico - NUEVO
        'analisis_diagnosticos': analisis_diagnosticos,
        
        # Análisis de Aceptaciones y Ventas Mostrador - NUEVO
        'kpis_aceptaciones': kpis_aceptaciones,
        'analisis_vm': analisis_vm,
        'analisis_seguimiento': analisis_seguimiento,
        
        # Filtros activos (para mantener estado en el form)
        'filtros_activos': {
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            'sucursal': sucursal_id,
            'tecnico': tecnico_id,
            'gama': gama,
            'periodo': periodo,
        },
        
        # Datos para desplegables
        'sucursales': sucursales,
        'tecnicos': tecnicos,
        'gamas': gamas,
        'periodos': periodos,
        
        # Metadatos
        'hay_datos': not df_cotizaciones.empty,
        'total_registros': len(df_cotizaciones),
        'fecha_generacion': datetime.now()
    }
    
    # ========================================
    # 8. RENDERIZAR TEMPLATE
    # ========================================
    
    return render(request, 'servicio_tecnico/dashboard_cotizaciones.html', context)


@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def exportar_dashboard_cotizaciones(request):
    """
    Exporta el dashboard de cotizaciones a Excel con múltiples hojas.

    Reutiliza los mismos filtros que el dashboard web.
    
    Returns:
        HttpResponse: Archivo Excel para descargar
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import pandas as pd  # Necesario para pd.to_datetime()
    
    from .utils_cotizaciones import (
        obtener_dataframe_cotizaciones,
        calcular_kpis_generales,
        analizar_piezas_cotizadas,
        analizar_proveedores,
        calcular_metricas_por_tecnico,
        calcular_metricas_por_sucursal
    )
    
    # Obtener filtros (mismos que dashboard)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    tecnico_id = request.GET.get('tecnico')
    gama = request.GET.get('gama')
    
    # Obtener datos
    df_cotizaciones = obtener_dataframe_cotizaciones(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama
    )
    
    if df_cotizaciones.empty:
        messages.error(request, 'No hay datos para exportar con los filtros aplicados.')
        return redirect('servicio_tecnico:dashboard_cotizaciones')
    
    # Calcular KPIs y métricas
    kpis = calcular_kpis_generales(df_cotizaciones)
    df_metricas_tecnicos = calcular_metricas_por_tecnico(df_cotizaciones)
    df_metricas_sucursales = calcular_metricas_por_sucursal(df_cotizaciones)
    
    # Obtener IDs para análisis relacionados
    cotizacion_ids = df_cotizaciones['cotizacion_id'].tolist()
    df_piezas = analizar_piezas_cotizadas(cotizacion_ids)
    df_seguimientos = analizar_proveedores(cotizacion_ids)
    
    # ========================================
    # CREAR WORKBOOK
    # ========================================
    
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # ========================================
    # HOJA 1: RESUMEN GENERAL (KPIs)
    # ========================================
    
    ws_resumen = wb.create_sheet("Resumen General")
    
    # Título
    ws_resumen.merge_cells('A1:D1')
    title_cell = ws_resumen['A1']
    title_cell.value = f"DASHBOARD DE COTIZACIONES - RESUMEN GENERAL"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    ws_resumen.row_dimensions[1].height = 30
    
    # Subtítulo con filtros
    ws_resumen.merge_cells('A2:D2')
    subtitle_cell = ws_resumen['A2']
    filtros_texto = f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Hoy'}"
    if sucursal_id:
        filtros_texto += f" | Sucursal ID: {sucursal_id}"
    if tecnico_id:
        filtros_texto += f" | Técnico ID: {tecnico_id}"
    if gama:
        filtros_texto += f" | Gama: {gama}"
    subtitle_cell.value = filtros_texto
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws_resumen.merge_cells('A3:D3')
    ws_resumen['A3'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_resumen['A3'].alignment = Alignment(horizontal='center')
    ws_resumen['A3'].font = Font(italic=True, size=10)
    
    # Espacio
    ws_resumen.row_dimensions[4].height = 5
    
    # Encabezados KPIs
    ws_resumen['A5'].value = 'Métrica'
    ws_resumen['B5'].value = 'Valor'
    ws_resumen['C5'].value = 'Porcentaje'
    ws_resumen['D5'].value = 'Observaciones'
    
    for col in ['A5', 'B5', 'C5', 'D5']:
        ws_resumen[col].font = header_font
        ws_resumen[col].fill = header_fill
        ws_resumen[col].alignment = header_alignment
    
    # Datos de KPIs
    kpis_data = [
        ['Total Cotizaciones', kpis['total_cotizaciones'], '', ''],
        ['Aceptadas', kpis['aceptadas'], f"{kpis['tasa_aceptacion']:.1f}%", 'Verde: > 60%'],
        ['Rechazadas', kpis['rechazadas'], f"{kpis['tasa_rechazo']:.1f}%", 'Rojo: > 30%'],
        ['Pendientes', kpis['pendientes'], f"{kpis['tasa_pendiente']:.1f}%", ''],
        ['Valor Total Cotizado', f"${kpis['valor_total_cotizado']:,.2f}", '', ''],
        ['Valor Aceptado', f"${kpis['valor_aceptado']:,.2f}", '', ''],
        ['Valor Rechazado', f"${kpis['valor_rechazado']:,.2f}", '', ''],
        ['Ticket Promedio', f"${kpis['ticket_promedio']:,.2f}", '', ''],
        ['Tiempo Respuesta Promedio', f"{kpis['tiempo_respuesta_promedio']:.1f} días", '', 'Ideal: < 3 días'],
        ['Piezas Promedio', f"{kpis['piezas_promedio']:.1f}", '', ''],
    ]
    
    row = 6
    for data in kpis_data:
        ws_resumen[f'A{row}'].value = data[0]
        ws_resumen[f'B{row}'].value = data[1]
        ws_resumen[f'C{row}'].value = data[2]
        ws_resumen[f'D{row}'].value = data[3]
        
        # Colorear según métrica
        if 'Aceptadas' in data[0] and kpis['tasa_aceptacion'] > 60:
            ws_resumen[f'B{row}'].fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
        elif 'Rechazadas' in data[0] and kpis['tasa_rechazo'] > 30:
            ws_resumen[f'B{row}'].fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
        
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 25
    
    # ========================================
    # HOJA 2: COTIZACIONES DETALLE
    # ========================================
    
    ws_cotiz = wb.create_sheet("Cotizaciones Detalle")
    
    # Título
    ws_cotiz.merge_cells('A1:L1')
    title_cell = ws_cotiz['A1']
    title_cell.value = f"DETALLE DE COTIZACIONES ({len(df_cotizaciones)} registros)"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    
    # Seleccionar columnas relevantes
    columnas_export = [
        'numero_orden', 'orden_cliente', 'numero_serie', 'fecha_envio', 
        'sucursal', 'tecnico', 'gama', 'tipo_equipo', 'marca', 'modelo', 
        'costo_total', 'aceptada'
    ]
    
    df_export = df_cotizaciones[columnas_export].copy()
    
    # Renombrar columnas para el Excel
    df_export.columns = [
        'Número de Orden',
        'Orden Cliente', 
        'Número de Serie',
        'Fecha Envío',
        'Sucursal',
        'Técnico',
        'Gama',
        'Tipo Equipo',
        'Marca',
        'Modelo',
        'Costo Total',
        'Estado'
    ]
    
    df_export['Fecha Envío'] = pd.to_datetime(df_export['Fecha Envío']).dt.strftime('%d/%m/%Y')
    df_export['Costo Total'] = df_export['Costo Total'].apply(lambda x: f'${x:,.2f}')
    df_export['Estado'] = df_export['Estado'].map({
        True: '✅ Aceptada',
        False: '❌ Rechazada',
        None: '⏳ Pendiente'
    })
    
    # Escribir con encabezados formateados
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws_cotiz.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 3:  # Encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # Auto-ajustar columnas
    for col_idx, column in enumerate(ws_cotiz.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_cotiz.column_dimensions[column_letter].width = adjusted_width
    
    # ========================================
    # HOJA 3: MÉTRICAS POR TÉCNICO
    # ========================================
    
    if not df_metricas_tecnicos.empty:
        ws_tecnicos = wb.create_sheet("Ranking Técnicos")
        
        ws_tecnicos.merge_cells('A1:F1')
        title_cell = ws_tecnicos['A1']
        title_cell.value = f"RANKING DE TÉCNICOS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        for r_idx, row in enumerate(dataframe_to_rows(df_metricas_tecnicos, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws_tecnicos.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
    
    # ========================================
    # HOJA 4: MÉTRICAS POR SUCURSAL
    # ========================================
    
    if not df_metricas_sucursales.empty:
        ws_sucursales = wb.create_sheet("Ranking Sucursales")
        
        ws_sucursales.merge_cells('A1:F1')
        title_cell = ws_sucursales['A1']
        title_cell.value = f"RANKING DE SUCURSALES"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        
        for r_idx, row in enumerate(dataframe_to_rows(df_metricas_sucursales, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws_sucursales.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
    
    # ========================================
    # GENERAR Y RETORNAR ARCHIVO
    # ========================================
    
    # Nombre del archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'Dashboard_Cotizaciones_{timestamp}.xlsx'
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    # Guardar workbook
    wb.save(response)
    
    return response


# ============================================================================
# EXPORTAR ANÁLISIS DETALLADO DE RECHAZOS A EXCEL (7 HOJAS)
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def exportar_analisis_rechazos(request):
    """
    Exporta un análisis exhaustivo de cotizaciones rechazadas a Excel con 7 hojas.
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Este Excel es el "detalle del detalle" de los rechazos. Mientras que el Excel
    general del dashboard muestra un resumen de todo, este se enfoca SOLO en rechazos
    y profundiza en cada ángulo: por motivo, por marca/modelo, tiempos, piezas, etc.
    
    Hojas:
        1. Resumen Rechazos - KPIs específicos de rechazos
        2. Detalle Rechazos - Cada cotización rechazada con todos los campos
        3. Rechazos por Motivo - Tabla pivote por motivo de rechazo
        4. Rechazos por Marca/Modelo - Análisis cruzado marca-modelo
        5. Tiempo de Respuesta - Análisis temporal de rechazos
        6. No Hay Partes - Apartado especial para rechazos por falta de partes
        7. Piezas Rechazadas - Detalle a nivel pieza individual
        8. Costo Alto - Detalle de rechazos por costo elevado con piezas desglosadas
        9. Servicios 3+ Piezas - Servicios con múltiples piezas cotizadas
    
    Returns:
        HttpResponse: Archivo Excel (.xlsx) para descargar
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import pandas as pd
    
    from .utils_cotizaciones import (
        obtener_dataframe_cotizaciones,
        calcular_kpis_generales,
        analizar_piezas_cotizadas,
    )
    from config.constants import MOTIVO_RECHAZO_COTIZACION
    
    # ========================================
    # OBTENER DATOS CON MISMOS FILTROS QUE EL DASHBOARD
    # ========================================
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    tecnico_id = request.GET.get('tecnico')
    gama = request.GET.get('gama')
    
    df_cotizaciones = obtener_dataframe_cotizaciones(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama
    )
    
    if df_cotizaciones.empty:
        messages.error(request, 'No hay datos para exportar con los filtros aplicados.')
        return redirect('servicio_tecnico:dashboard_cotizaciones')
    
    # Filtrar solo rechazadas
    df_rechazos = df_cotizaciones[df_cotizaciones['aceptada'] == False].copy()
    
    if df_rechazos.empty:
        messages.warning(request, 'No hay cotizaciones rechazadas en el período seleccionado.')
        return redirect('servicio_tecnico:dashboard_cotizaciones')
    
    # Diccionario de labels legibles para motivos de rechazo
    labels_motivos = dict(MOTIVO_RECHAZO_COTIZACION)
    
    # Obtener piezas de cotizaciones rechazadas
    cotizacion_ids_rechazos = df_rechazos['cotizacion_id'].tolist()
    df_piezas = analizar_piezas_cotizadas(cotizacion_ids_rechazos)
    
    # Obtener piezas con proveedor directamente del modelo
    # (analizar_piezas_cotizadas no incluye proveedor)
    from .models import PiezaCotizada
    piezas_con_proveedor = PiezaCotizada.objects.filter(
        cotizacion_id__in=cotizacion_ids_rechazos
    ).select_related(
        'componente',
        'cotizacion',
        'cotizacion__orden',
        'cotizacion__orden__detalle_equipo'
    ).values(
        'id', 'cotizacion_id', 'componente__nombre',
        'descripcion_adicional', 'sugerida_por_tecnico', 'es_necesaria',
        'cantidad', 'costo_unitario', 'proveedor',
        'aceptada_por_cliente', 'motivo_rechazo_pieza',
        'cotizacion__orden__detalle_equipo__marca',
        'cotizacion__orden__detalle_equipo__modelo',
    )
    df_piezas_proveedor = pd.DataFrame(list(piezas_con_proveedor))
    
    # KPIs generales para comparativa
    kpis = calcular_kpis_generales(df_cotizaciones)
    
    # ========================================
    # CREAR WORKBOOK
    # ========================================
    wb = Workbook()
    wb.remove(wb.active)
    
    # Estilos reutilizables
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='c0392c', end_color='c0392c', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    title_align = Alignment(horizontal='center', vertical='center')
    
    subtitle_font = Font(italic=True, size=10, color='666666')
    
    kpi_label_font = Font(bold=True, size=11)
    kpi_value_font = Font(bold=True, size=12, color='c0392c')
    
    green_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    red_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
    yellow_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
    blue_fill = PatternFill(start_color='d1ecf1', end_color='d1ecf1', fill_type='solid')
    orange_fill = PatternFill(start_color='ffeaa7', end_color='ffeaa7', fill_type='solid')
    
    section_font = Font(bold=True, size=12, color='2c3e50')
    section_fill = PatternFill(start_color='ecf0f1', end_color='ecf0f1', fill_type='solid')
    
    # Texto de filtros para subtítulos
    filtros_texto = f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Hoy'}"
    if sucursal_id:
        filtros_texto += f" | Sucursal ID: {sucursal_id}"
    if tecnico_id:
        filtros_texto += f" | Técnico ID: {tecnico_id}"
    if gama:
        filtros_texto += f" | Gama: {gama}"
    
    # Función auxiliar para escribir título y subtítulo en cada hoja
    def escribir_encabezado_hoja(ws, titulo, num_cols=8):
        ultimo_col = get_column_letter(num_cols)
        ws.merge_cells(f'A1:{ultimo_col}1')
        cell = ws['A1']
        cell.value = titulo
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_align
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells(f'A2:{ultimo_col}2')
        ws['A2'].value = f"{filtros_texto} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal='center')
    
    # Función auxiliar para escribir encabezados de tabla
    def escribir_headers(ws, headers, fila, fill=None):
        fill_usar = fill or header_fill
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=fila, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = fill_usar
            cell.alignment = header_align
    
    # Función auxiliar para auto-ajustar columnas
    def autoajustar_columnas(ws, min_width=10, max_width=45):
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = min_width
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)
    
    # ========================================
    # HOJA 1: RESUMEN RECHAZOS (KPIs)
    # ========================================
    ws1 = wb.create_sheet("Resumen Rechazos")
    escribir_encabezado_hoja(ws1, "ANÁLISIS DETALLADO DE RECHAZOS - RESUMEN", 4)
    
    # Sección 1: KPIs principales
    ws1.merge_cells('A4:D4')
    ws1['A4'].value = "INDICADORES PRINCIPALES"
    ws1['A4'].font = section_font
    ws1['A4'].fill = section_fill
    
    escribir_headers(ws1, ['Métrica', 'Valor', 'Porcentaje', 'Observaciones'], 5)
    
    total_rechazos = len(df_rechazos)
    total_cotizaciones = len(df_cotizaciones)
    pct_rechazos = (total_rechazos / total_cotizaciones * 100) if total_cotizaciones > 0 else 0
    valor_perdido = df_rechazos['costo_total'].sum()
    valor_total = df_cotizaciones['costo_total'].sum()
    pct_valor_perdido = (valor_perdido / valor_total * 100) if valor_total > 0 else 0
    ticket_prom_rechazo = df_rechazos['costo_total'].mean()
    ticket_prom_aceptado = df_cotizaciones[df_cotizaciones['aceptada'] == True]['costo_total'].mean() if len(df_cotizaciones[df_cotizaciones['aceptada'] == True]) > 0 else 0
    
    # Tiempo respuesta rechazos vs aceptados
    df_rechazos_con_resp = df_rechazos[df_rechazos['fecha_respuesta'].notna()]
    tiempo_resp_rechazos = df_rechazos_con_resp['dias_sin_respuesta'].mean() if len(df_rechazos_con_resp) > 0 else 0
    
    df_aceptados = df_cotizaciones[df_cotizaciones['aceptada'] == True]
    df_aceptados_con_resp = df_aceptados[df_aceptados['fecha_respuesta'].notna()]
    tiempo_resp_aceptados = df_aceptados_con_resp['dias_sin_respuesta'].mean() if len(df_aceptados_con_resp) > 0 else 0
    
    kpis_data = [
        ['Total Cotizaciones (todas)', total_cotizaciones, '', 'Base de referencia'],
        ['Total Rechazadas', total_rechazos, f'{pct_rechazos:.1f}%', 'Objetivo: < 30%'],
        ['Total Aceptadas', kpis.get('aceptadas', 0), f"{kpis.get('tasa_aceptacion', 0):.1f}%", ''],
        ['Valor Total Perdido', f'${valor_perdido:,.2f}', f'{pct_valor_perdido:.1f}% del total', 'Oportunidad de recuperación'],
        ['Ticket Promedio (Rechazos)', f'${ticket_prom_rechazo:,.2f}', '', ''],
        ['Ticket Promedio (Aceptados)', f'${ticket_prom_aceptado:,.2f}', '', 'Comparar con rechazos'],
        ['Tiempo Resp. Prom (Rechazos)', f'{tiempo_resp_rechazos:.1f} días', '', ''],
        ['Tiempo Resp. Prom (Aceptados)', f'{tiempo_resp_aceptados:.1f} días', '', 'Comparar con rechazos'],
        ['Piezas Prom. por Rechazo', f'{df_rechazos["total_piezas"].mean():.1f}', '', ''],
    ]
    
    for i, data in enumerate(kpis_data):
        row = 6 + i
        ws1.cell(row=row, column=1, value=data[0]).font = kpi_label_font
        ws1.cell(row=row, column=2, value=data[1]).font = kpi_value_font
        ws1.cell(row=row, column=3, value=data[2])
        ws1.cell(row=row, column=4, value=data[3])
    
    # Sección 2: Top Motivos
    fila_motivos = 6 + len(kpis_data) + 2
    ws1.merge_cells(f'A{fila_motivos}:D{fila_motivos}')
    ws1[f'A{fila_motivos}'].value = "TOP MOTIVOS DE RECHAZO"
    ws1[f'A{fila_motivos}'].font = section_font
    ws1[f'A{fila_motivos}'].fill = section_fill
    
    escribir_headers(ws1, ['Motivo', 'Cantidad', '% de Rechazos', 'Valor Perdido'], fila_motivos + 1)
    
    motivos_conteo = df_rechazos['motivo_rechazo'].value_counts()
    fila = fila_motivos + 2
    for motivo, conteo in motivos_conteo.items():
        label = labels_motivos.get(motivo, str(motivo).replace('_', ' ').title()) if motivo else 'Sin motivo especificado'
        pct = (conteo / total_rechazos * 100) if total_rechazos > 0 else 0
        valor = df_rechazos[df_rechazos['motivo_rechazo'] == motivo]['costo_total'].sum()
        
        ws1.cell(row=fila, column=1, value=label)
        ws1.cell(row=fila, column=2, value=conteo)
        ws1.cell(row=fila, column=3, value=f'{pct:.1f}%')
        ws1.cell(row=fila, column=4, value=f'${valor:,.2f}')
        
        # Colorear según criticidad
        if pct >= 25:
            ws1.cell(row=fila, column=1).fill = red_fill
        elif pct >= 15:
            ws1.cell(row=fila, column=1).fill = yellow_fill
        
        fila += 1
    
    # Sección 3: Top Marcas rechazadas
    fila_marcas = fila + 2
    ws1.merge_cells(f'A{fila_marcas}:D{fila_marcas}')
    ws1[f'A{fila_marcas}'].value = "TOP MARCAS CON MÁS RECHAZOS"
    ws1[f'A{fila_marcas}'].font = section_font
    ws1[f'A{fila_marcas}'].fill = section_fill
    
    escribir_headers(ws1, ['Marca', 'Rechazos', '% del Total', 'Tasa de Rechazo'], fila_marcas + 1)
    
    marcas_rechazos = df_rechazos['marca'].value_counts().head(10)
    fila = fila_marcas + 2
    for marca, conteo in marcas_rechazos.items():
        total_marca = len(df_cotizaciones[df_cotizaciones['marca'] == marca])
        tasa = (conteo / total_marca * 100) if total_marca > 0 else 0
        pct = (conteo / total_rechazos * 100) if total_rechazos > 0 else 0
        
        ws1.cell(row=fila, column=1, value=marca or 'Sin marca')
        ws1.cell(row=fila, column=2, value=conteo)
        ws1.cell(row=fila, column=3, value=f'{pct:.1f}%')
        ws1.cell(row=fila, column=4, value=f'{tasa:.1f}%')
        
        if tasa >= 50:
            ws1.cell(row=fila, column=4).fill = red_fill
        fila += 1
    
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 30
    
    # ========================================
    # HOJA 2: DETALLE RECHAZOS (cada cotización rechazada)
    # ========================================
    ws2 = wb.create_sheet("Detalle Rechazos")
    escribir_encabezado_hoja(ws2, f"DETALLE DE COTIZACIONES RECHAZADAS ({total_rechazos} registros)", 16)
    
    headers_detalle = [
        'Orden Cliente', 'Número Serie',
        'Marca', 'Modelo', 'Tipo Equipo', 'Gama',
        'Sucursal', 'Técnico', 'Responsable',
        'Motivo Rechazo', 'Detalle Rechazo',
        'Fecha Envío', 'Fecha Respuesta', 'Días Respuesta',
        'Costo Total', 'Costo Piezas', 'Costo Mano Obra',
        'Total Piezas', 'Piezas Necesarias'
    ]
    escribir_headers(ws2, headers_detalle, 4)
    
    fila = 5
    for _, rec in df_rechazos.iterrows():
        motivo_label = labels_motivos.get(rec.get('motivo_rechazo', ''), str(rec.get('motivo_rechazo', '')).replace('_', ' ').title()) if rec.get('motivo_rechazo') else 'Sin motivo'
        
        fecha_envio_str = ''
        if pd.notna(rec.get('fecha_envio')):
            try:
                fecha_envio_str = pd.to_datetime(rec['fecha_envio']).strftime('%d/%m/%Y')
            except:
                fecha_envio_str = str(rec['fecha_envio'])
        
        fecha_resp_str = ''
        if pd.notna(rec.get('fecha_respuesta')):
            try:
                fecha_resp_str = pd.to_datetime(rec['fecha_respuesta']).strftime('%d/%m/%Y')
            except:
                fecha_resp_str = str(rec['fecha_respuesta'])
        
        valores = [
            rec.get('orden_cliente', ''),
            rec.get('numero_serie', ''),
            rec.get('marca', ''),
            rec.get('modelo', ''),
            rec.get('tipo_equipo', ''),
            rec.get('gama', ''),
            rec.get('sucursal', ''),
            rec.get('tecnico', ''),
            rec.get('responsable', ''),
            motivo_label,
            rec.get('detalle_rechazo', ''),
            fecha_envio_str,
            fecha_resp_str,
            rec.get('dias_sin_respuesta', ''),
            f"${rec.get('costo_total', 0):,.2f}",
            f"${rec.get('costo_total_piezas', 0):,.2f}",
            f"${rec.get('costo_mano_obra', 0):,.2f}",
            rec.get('total_piezas', 0),
            rec.get('piezas_necesarias', 0),
        ]
        
        for col_idx, val in enumerate(valores, 1):
            ws2.cell(row=fila, column=col_idx, value=val)
        
        fila += 1
    
    autoajustar_columnas(ws2, max_width=40)
    # Limitar ancho de la columna de detalle de rechazo
    ws2.column_dimensions['L'].width = 50
    
    # ========================================
    # HOJA 3: RECHAZOS POR MOTIVO (tabla pivote)
    # ========================================
    ws3 = wb.create_sheet("Rechazos por Motivo")
    escribir_encabezado_hoja(ws3, "ANÁLISIS DETALLADO POR MOTIVO DE RECHAZO", 10)
    
    headers_motivo = [
        'Motivo de Rechazo', 'Cantidad', '% de Rechazos',
        'Costo Promedio', 'Costo Mediana', 'Costo Mínimo', 'Costo Máximo',
        'Tiempo Resp. Prom (días)', 'Top 3 Marcas', 'Top 3 Sucursales'
    ]
    escribir_headers(ws3, headers_motivo, 4)
    
    fila = 5
    for motivo, conteo in motivos_conteo.items():
        df_motivo = df_rechazos[df_rechazos['motivo_rechazo'] == motivo]
        label = labels_motivos.get(motivo, str(motivo).replace('_', ' ').title()) if motivo else 'Sin motivo'
        pct = (conteo / total_rechazos * 100) if total_rechazos > 0 else 0
        
        costo_promedio = df_motivo['costo_total'].mean()
        costo_mediana = df_motivo['costo_total'].median()
        costo_min = df_motivo['costo_total'].min()
        costo_max = df_motivo['costo_total'].max()
        
        # Tiempo de respuesta promedio para este motivo
        df_motivo_resp = df_motivo[df_motivo['fecha_respuesta'].notna()]
        tiempo_resp = df_motivo_resp['dias_sin_respuesta'].mean() if len(df_motivo_resp) > 0 else 0
        
        # Top 3 marcas
        top_marcas = df_motivo['marca'].value_counts().head(3)
        marcas_str = ', '.join([f"{m} ({c})" for m, c in top_marcas.items()]) if len(top_marcas) > 0 else 'N/A'
        
        # Top 3 sucursales
        top_sucursales = df_motivo['sucursal'].value_counts().head(3)
        sucursales_str = ', '.join([f"{s} ({c})" for s, c in top_sucursales.items()]) if len(top_sucursales) > 0 else 'N/A'
        
        valores = [
            label, conteo, f'{pct:.1f}%',
            f'${costo_promedio:,.2f}', f'${costo_mediana:,.2f}',
            f'${costo_min:,.2f}', f'${costo_max:,.2f}',
            f'{tiempo_resp:.1f}',
            marcas_str, sucursales_str
        ]
        
        for col_idx, val in enumerate(valores, 1):
            ws3.cell(row=fila, column=col_idx, value=val)
        
        # Colorear filas críticas
        if pct >= 25:
            for col_idx in range(1, len(valores) + 1):
                ws3.cell(row=fila, column=col_idx).fill = red_fill
        elif pct >= 15:
            for col_idx in range(1, len(valores) + 1):
                ws3.cell(row=fila, column=col_idx).fill = yellow_fill
        
        fila += 1
    
    autoajustar_columnas(ws3, max_width=50)
    
    # ========================================
    # HOJA 4: RECHAZOS POR MARCA/MODELO
    # ========================================
    ws4 = wb.create_sheet("Rechazos Marca-Modelo")
    escribir_encabezado_hoja(ws4, "ANÁLISIS DE RECHAZOS POR MARCA Y MODELO", 8)
    
    headers_marca = [
        'Marca', 'Modelo', 'Total Cotizaciones', 'Rechazadas',
        'Tasa de Rechazo', 'Motivo Más Común', 'Costo Prom. Rechazado',
        'Valor Total Perdido'
    ]
    escribir_headers(ws4, headers_marca, 4)
    
    # Agrupar por marca-modelo
    marcas_modelos = df_rechazos.groupby(['marca', 'modelo']).agg(
        rechazadas=('cotizacion_id', 'count'),
        costo_promedio=('costo_total', 'mean'),
        valor_total=('costo_total', 'sum'),
    ).reset_index()
    
    # Para cada marca-modelo, calcular total cotizaciones y motivo más común
    fila = 5
    for _, row_mm in marcas_modelos.sort_values('rechazadas', ascending=False).iterrows():
        marca = row_mm['marca'] or 'Sin marca'
        modelo = row_mm['modelo'] or 'Sin modelo'
        rechazadas = row_mm['rechazadas']
        
        # Total cotizaciones (aceptadas + rechazadas + pendientes) para este marca-modelo
        total_marca_modelo = len(df_cotizaciones[
            (df_cotizaciones['marca'] == row_mm['marca']) & 
            (df_cotizaciones['modelo'] == row_mm['modelo'])
        ])
        tasa_rechazo = (rechazadas / total_marca_modelo * 100) if total_marca_modelo > 0 else 0
        
        # Motivo más común para este marca-modelo
        df_mm = df_rechazos[
            (df_rechazos['marca'] == row_mm['marca']) & 
            (df_rechazos['modelo'] == row_mm['modelo'])
        ]
        motivo_comun = df_mm['motivo_rechazo'].mode()
        motivo_label = labels_motivos.get(motivo_comun.iloc[0], str(motivo_comun.iloc[0]).replace('_', ' ').title()) if len(motivo_comun) > 0 and motivo_comun.iloc[0] else 'N/A'
        
        valores = [
            marca, modelo, total_marca_modelo, rechazadas,
            f'{tasa_rechazo:.1f}%', motivo_label,
            f'${row_mm["costo_promedio"]:,.2f}',
            f'${row_mm["valor_total"]:,.2f}'
        ]
        
        for col_idx, val in enumerate(valores, 1):
            ws4.cell(row=fila, column=col_idx, value=val)
        
        if tasa_rechazo >= 70:
            for col_idx in range(1, len(valores) + 1):
                ws4.cell(row=fila, column=col_idx).fill = red_fill
        elif tasa_rechazo >= 50:
            for col_idx in range(1, len(valores) + 1):
                ws4.cell(row=fila, column=col_idx).fill = yellow_fill
        
        fila += 1
    
    autoajustar_columnas(ws4)
    
    # ========================================
    # HOJA 5: TIEMPO DE RESPUESTA
    # ========================================
    ws5 = wb.create_sheet("Tiempo de Respuesta")
    escribir_encabezado_hoja(ws5, "ANÁLISIS DE TIEMPOS DE RESPUESTA EN RECHAZOS", 6)
    
    # Sección 1: Rangos de tiempo vs tasa de rechazo
    ws5.merge_cells('A4:F4')
    ws5[f'A4'].value = "RANGOS DE TIEMPO DE RESPUESTA vs RESULTADO"
    ws5[f'A4'].font = section_font
    ws5[f'A4'].fill = section_fill
    
    headers_tiempo = [
        'Rango (días)', 'Total Cotizaciones', 'Aceptadas', 'Rechazadas',
        'Tasa Aceptación', 'Tasa Rechazo'
    ]
    escribir_headers(ws5, headers_tiempo, 5)
    
    # Solo cotizaciones con respuesta
    df_con_respuesta = df_cotizaciones[df_cotizaciones['fecha_respuesta'].notna()].copy()
    
    rangos = [
        ('0-2 días', 0, 2),
        ('3-5 días', 3, 5),
        ('6-10 días', 6, 10),
        ('11-15 días', 11, 15),
        ('16-30 días', 16, 30),
        ('31+ días', 31, 9999),
    ]
    
    fila = 6
    for label_rango, min_dias, max_dias in rangos:
        df_rango = df_con_respuesta[
            (df_con_respuesta['dias_sin_respuesta'] >= min_dias) & 
            (df_con_respuesta['dias_sin_respuesta'] <= max_dias)
        ]
        total_rango = len(df_rango)
        aceptadas_rango = len(df_rango[df_rango['aceptada'] == True])
        rechazadas_rango = len(df_rango[df_rango['aceptada'] == False])
        tasa_acep = (aceptadas_rango / total_rango * 100) if total_rango > 0 else 0
        tasa_rech = (rechazadas_rango / total_rango * 100) if total_rango > 0 else 0
        
        ws5.cell(row=fila, column=1, value=label_rango)
        ws5.cell(row=fila, column=2, value=total_rango)
        ws5.cell(row=fila, column=3, value=aceptadas_rango)
        ws5.cell(row=fila, column=4, value=rechazadas_rango)
        ws5.cell(row=fila, column=5, value=f'{tasa_acep:.1f}%')
        ws5.cell(row=fila, column=6, value=f'{tasa_rech:.1f}%')
        
        if tasa_rech >= 60:
            ws5.cell(row=fila, column=6).fill = red_fill
        elif tasa_rech >= 40:
            ws5.cell(row=fila, column=6).fill = yellow_fill
        else:
            ws5.cell(row=fila, column=6).fill = green_fill
        
        fila += 1
    
    # Sección 2: Tiempo promedio por motivo
    fila += 2
    ws5.merge_cells(f'A{fila}:F{fila}')
    ws5[f'A{fila}'].value = "TIEMPO PROMEDIO DE RESPUESTA POR MOTIVO DE RECHAZO"
    ws5[f'A{fila}'].font = section_font
    ws5[f'A{fila}'].fill = section_fill
    fila += 1
    
    escribir_headers(ws5, ['Motivo', 'Casos con Respuesta', 'Tiempo Promedio (días)', 'Tiempo Mediana (días)', 'Tiempo Mínimo', 'Tiempo Máximo'], fila)
    fila += 1
    
    for motivo, _ in motivos_conteo.items():
        df_motivo_resp = df_rechazos[
            (df_rechazos['motivo_rechazo'] == motivo) & 
            (df_rechazos['fecha_respuesta'].notna())
        ]
        if len(df_motivo_resp) == 0:
            continue
        
        label = labels_motivos.get(motivo, str(motivo).replace('_', ' ').title()) if motivo else 'Sin motivo'
        
        ws5.cell(row=fila, column=1, value=label)
        ws5.cell(row=fila, column=2, value=len(df_motivo_resp))
        ws5.cell(row=fila, column=3, value=f'{df_motivo_resp["dias_sin_respuesta"].mean():.1f}')
        ws5.cell(row=fila, column=4, value=f'{df_motivo_resp["dias_sin_respuesta"].median():.1f}')
        ws5.cell(row=fila, column=5, value=f'{df_motivo_resp["dias_sin_respuesta"].min():.0f}')
        ws5.cell(row=fila, column=6, value=f'{df_motivo_resp["dias_sin_respuesta"].max():.0f}')
        fila += 1
    
    # Sección 3: Tiempo por sucursal (solo rechazos)
    fila += 2
    ws5.merge_cells(f'A{fila}:F{fila}')
    ws5[f'A{fila}'].value = "TIEMPO DE RESPUESTA EN RECHAZOS POR SUCURSAL"
    ws5[f'A{fila}'].font = section_font
    ws5[f'A{fila}'].fill = section_fill
    fila += 1
    
    escribir_headers(ws5, ['Sucursal', 'Total Rechazos', 'Con Respuesta', 'Tiempo Promedio (días)', 'Sin Respuesta', '% Sin Respuesta'], fila)
    fila += 1
    
    for sucursal in df_rechazos['sucursal'].unique():
        df_suc = df_rechazos[df_rechazos['sucursal'] == sucursal]
        df_suc_resp = df_suc[df_suc['fecha_respuesta'].notna()]
        sin_resp = len(df_suc) - len(df_suc_resp)
        pct_sin = (sin_resp / len(df_suc) * 100) if len(df_suc) > 0 else 0
        tiempo_prom = df_suc_resp['dias_sin_respuesta'].mean() if len(df_suc_resp) > 0 else 0
        
        ws5.cell(row=fila, column=1, value=sucursal)
        ws5.cell(row=fila, column=2, value=len(df_suc))
        ws5.cell(row=fila, column=3, value=len(df_suc_resp))
        ws5.cell(row=fila, column=4, value=f'{tiempo_prom:.1f}')
        ws5.cell(row=fila, column=5, value=sin_resp)
        ws5.cell(row=fila, column=6, value=f'{pct_sin:.1f}%')
        fila += 1
    
    # Sección 4: Tiempo por técnico (solo rechazos)
    fila += 2
    ws5.merge_cells(f'A{fila}:F{fila}')
    ws5[f'A{fila}'].value = "TIEMPO DE RESPUESTA EN RECHAZOS POR TÉCNICO"
    ws5[f'A{fila}'].font = section_font
    ws5[f'A{fila}'].fill = section_fill
    fila += 1
    
    escribir_headers(ws5, ['Técnico', 'Total Rechazos', 'Con Respuesta', 'Tiempo Promedio (días)', 'Sin Respuesta', '% Sin Respuesta'], fila)
    fila += 1
    
    for tecnico in df_rechazos['tecnico'].unique():
        df_tec = df_rechazos[df_rechazos['tecnico'] == tecnico]
        df_tec_resp = df_tec[df_tec['fecha_respuesta'].notna()]
        sin_resp = len(df_tec) - len(df_tec_resp)
        pct_sin = (sin_resp / len(df_tec) * 100) if len(df_tec) > 0 else 0
        tiempo_prom = df_tec_resp['dias_sin_respuesta'].mean() if len(df_tec_resp) > 0 else 0
        
        ws5.cell(row=fila, column=1, value=tecnico)
        ws5.cell(row=fila, column=2, value=len(df_tec))
        ws5.cell(row=fila, column=3, value=len(df_tec_resp))
        ws5.cell(row=fila, column=4, value=f'{tiempo_prom:.1f}')
        ws5.cell(row=fila, column=5, value=sin_resp)
        ws5.cell(row=fila, column=6, value=f'{pct_sin:.1f}%')
        fila += 1
    
    autoajustar_columnas(ws5)
    
    # ========================================
    # HOJA 6: NO HAY PARTES (Apartado especial)
    # ========================================
    ws6 = wb.create_sheet("No Hay Partes")
    no_hay_partes_fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
    
    escribir_encabezado_hoja(ws6, "ANÁLISIS ESPECIAL: RECHAZOS POR FALTA DE PARTES EN EL MERCADO", 10)
    
    df_no_partes = df_rechazos[df_rechazos['motivo_rechazo'] == 'no_hay_partes'].copy()
    
    # KPIs de "No Hay Partes"
    ws6.merge_cells('A4:J4')
    ws6['A4'].value = "INDICADORES - NO HAY PARTES DISPONIBLES"
    ws6['A4'].font = Font(bold=True, size=12, color='FFFFFF')
    ws6['A4'].fill = no_hay_partes_fill
    ws6['A4'].alignment = Alignment(horizontal='center')
    
    total_no_partes = len(df_no_partes)
    pct_de_rechazos = (total_no_partes / total_rechazos * 100) if total_rechazos > 0 else 0
    pct_de_total = (total_no_partes / total_cotizaciones * 100) if total_cotizaciones > 0 else 0
    valor_perdido_partes = df_no_partes['costo_total'].sum() if total_no_partes > 0 else 0
    
    kpis_partes = [
        ['Casos "No Hay Partes"', total_no_partes],
        ['% de Todos los Rechazos', f'{pct_de_rechazos:.1f}%'],
        ['% del Total de Cotizaciones', f'{pct_de_total:.1f}%'],
        ['Valor Perdido por Falta de Partes', f'${valor_perdido_partes:,.2f}'],
    ]
    
    if total_no_partes > 0:
        kpis_partes.extend([
            ['Costo Promedio de Cotización', f'${df_no_partes["costo_total"].mean():,.2f}'],
            ['Piezas Promedio por Orden', f'{df_no_partes["total_piezas"].mean():.1f}'],
        ])
    
    for i, (kpi_label, kpi_val) in enumerate(kpis_partes):
        row_num = 6 + i
        ws6.cell(row=row_num, column=1, value=kpi_label).font = kpi_label_font
        ws6.cell(row=row_num, column=2, value=kpi_val).font = kpi_value_font
    
    if total_no_partes > 0:
        # Sección: Detalle de cada caso
        fila = 6 + len(kpis_partes) + 2
        ws6.merge_cells(f'A{fila}:J{fila}')
        ws6[f'A{fila}'].value = "DETALLE DE CASOS - NO HAY PARTES"
        ws6[f'A{fila}'].font = section_font
        ws6[f'A{fila}'].fill = section_fill
        fila += 1
        
        headers_np = [
            'Orden Cliente', 'Marca', 'Modelo', 'Tipo Equipo', 'Gama',
            'Sucursal', 'Técnico', 'Costo Total', 'Total Piezas', 'Detalle Rechazo'
        ]
        escribir_headers(ws6, headers_np, fila)
        fila += 1
        
        for _, rec in df_no_partes.iterrows():
            ws6.cell(row=fila, column=1, value=rec.get('orden_cliente', ''))
            ws6.cell(row=fila, column=2, value=rec.get('marca', ''))
            ws6.cell(row=fila, column=3, value=rec.get('modelo', ''))
            ws6.cell(row=fila, column=4, value=rec.get('tipo_equipo', ''))
            ws6.cell(row=fila, column=5, value=rec.get('gama', ''))
            ws6.cell(row=fila, column=6, value=rec.get('sucursal', ''))
            ws6.cell(row=fila, column=7, value=rec.get('tecnico', ''))
            ws6.cell(row=fila, column=8, value=f"${rec.get('costo_total', 0):,.2f}")
            ws6.cell(row=fila, column=9, value=rec.get('total_piezas', 0))
            ws6.cell(row=fila, column=10, value=rec.get('detalle_rechazo', ''))
            fila += 1
        
        # Sección: Top Marcas afectadas
        fila += 2
        ws6.merge_cells(f'A{fila}:J{fila}')
        ws6[f'A{fila}'].value = "MARCAS MÁS AFECTADAS POR FALTA DE PARTES"
        ws6[f'A{fila}'].font = section_font
        ws6[f'A{fila}'].fill = section_fill
        fila += 1
        
        escribir_headers(ws6, ['Marca', 'Casos', '% del Total "No Hay Partes"', 'Modelos Afectados', '', '', '', '', '', ''], fila)
        fila += 1
        
        marcas_np = df_no_partes['marca'].value_counts()
        for marca, conteo in marcas_np.items():
            modelos_afectados = df_no_partes[df_no_partes['marca'] == marca]['modelo'].unique()
            modelos_str = ', '.join([str(m) for m in modelos_afectados[:5]])
            pct = (conteo / total_no_partes * 100) if total_no_partes > 0 else 0
            
            ws6.cell(row=fila, column=1, value=marca or 'Sin marca')
            ws6.cell(row=fila, column=2, value=conteo)
            ws6.cell(row=fila, column=3, value=f'{pct:.1f}%')
            ws6.cell(row=fila, column=4, value=modelos_str)
            fila += 1
        
        # Sección: Piezas que se necesitaban (del modelo PiezaCotizada)
        if not df_piezas_proveedor.empty:
            # Filtrar piezas de cotizaciones "no hay partes"
            ids_no_partes = df_no_partes['cotizacion_id'].tolist()
            df_piezas_np = df_piezas_proveedor[
                df_piezas_proveedor['cotizacion_id'].isin(ids_no_partes)
            ]
            
            if not df_piezas_np.empty:
                fila += 2
                ws6.merge_cells(f'A{fila}:J{fila}')
                ws6[f'A{fila}'].value = "PIEZAS QUE SE NECESITABAN (No encontradas en el mercado)"
                ws6[f'A{fila}'].font = section_font
                ws6[f'A{fila}'].fill = section_fill
                fila += 1
                
                escribir_headers(ws6, ['Componente', 'Cant. Veces Solicitada', 'Marca Equipo', 'Modelo Equipo', 'Proveedor', 'Costo Unit. Promedio', 'Es Necesaria', '', '', ''], fila)
                fila += 1
                
                # Agrupar piezas por componente
                piezas_agrupadas = df_piezas_np.groupby('componente__nombre').agg(
                    veces=('id', 'count'),
                    costo_prom=('costo_unitario', 'mean'),
                    marcas=('cotizacion__orden__detalle_equipo__marca', lambda x: ', '.join(x.dropna().unique()[:3])),
                    modelos=('cotizacion__orden__detalle_equipo__modelo', lambda x: ', '.join(x.dropna().unique()[:3])),
                    proveedores=('proveedor', lambda x: ', '.join([str(p) for p in x.dropna().unique()[:3]]) if x.notna().any() else 'N/A'),
                    necesaria=('es_necesaria', 'mean'),
                ).reset_index().sort_values('veces', ascending=False)
                
                for _, pieza_row in piezas_agrupadas.iterrows():
                    ws6.cell(row=fila, column=1, value=pieza_row['componente__nombre'])
                    ws6.cell(row=fila, column=2, value=pieza_row['veces'])
                    ws6.cell(row=fila, column=3, value=pieza_row['marcas'])
                    ws6.cell(row=fila, column=4, value=pieza_row['modelos'])
                    ws6.cell(row=fila, column=5, value=pieza_row['proveedores'])
                    ws6.cell(row=fila, column=6, value=f"${pieza_row['costo_prom']:,.2f}")
                    es_nec_pct = pieza_row['necesaria'] * 100
                    ws6.cell(row=fila, column=7, value=f'{es_nec_pct:.0f}% necesaria')
                    fila += 1
        
        # Sección: Tendencia mensual
        fila += 2
        ws6.merge_cells(f'A{fila}:J{fila}')
        ws6[f'A{fila}'].value = "TENDENCIA MENSUAL: ¿ESTÁ MEJORANDO O EMPEORANDO?"
        ws6[f'A{fila}'].font = section_font
        ws6[f'A{fila}'].fill = section_fill
        fila += 1
        
        escribir_headers(ws6, ['Mes', 'Casos "No Hay Partes"', 'Total Rechazos del Mes', '% del Mes', 'Tendencia', '', '', '', '', ''], fila)
        fila += 1
        
        # Agrupar por mes
        if 'fecha_envio' in df_no_partes.columns and len(df_no_partes) > 0:
            # EXPLICACIÓN: tz_localize(None) quita el timezone antes de convertir a Period
            # Esto evita el warning de Pandas sobre pérdida de timezone
            df_no_partes['mes_periodo'] = pd.to_datetime(df_no_partes['fecha_envio']).dt.tz_localize(None).dt.to_period('M')
            df_rechazos_temp = df_rechazos.copy()
            df_rechazos_temp['mes_periodo'] = pd.to_datetime(df_rechazos_temp['fecha_envio']).dt.tz_localize(None).dt.to_period('M')
            
            meses_np = df_no_partes.groupby('mes_periodo').size()
            meses_total_rech = df_rechazos_temp.groupby('mes_periodo').size()
            
            prev_count = None
            for mes in sorted(meses_np.index):
                count = meses_np[mes]
                total_mes = meses_total_rech.get(mes, 0)
                pct_mes = (count / total_mes * 100) if total_mes > 0 else 0
                
                # Determinar tendencia
                if prev_count is not None:
                    if count > prev_count:
                        tendencia = 'EMPEORANDO'
                    elif count < prev_count:
                        tendencia = 'MEJORANDO'
                    else:
                        tendencia = 'ESTABLE'
                else:
                    tendencia = '-'
                
                ws6.cell(row=fila, column=1, value=str(mes))
                ws6.cell(row=fila, column=2, value=count)
                ws6.cell(row=fila, column=3, value=total_mes)
                ws6.cell(row=fila, column=4, value=f'{pct_mes:.1f}%')
                ws6.cell(row=fila, column=5, value=tendencia)
                
                if tendencia == 'EMPEORANDO':
                    ws6.cell(row=fila, column=5).fill = red_fill
                elif tendencia == 'MEJORANDO':
                    ws6.cell(row=fila, column=5).fill = green_fill
                
                prev_count = count
                fila += 1
    else:
        # No hay casos de "No hay partes"
        ws6.merge_cells('A6:J6')
        ws6['A6'].value = "No se encontraron cotizaciones rechazadas por falta de partes en el período seleccionado."
        ws6['A6'].font = Font(italic=True, size=12)
        ws6['A6'].alignment = Alignment(horizontal='center')
    
    autoajustar_columnas(ws6, max_width=50)
    
    # ========================================
    # HOJA 7: PIEZAS RECHAZADAS (detalle a nivel pieza)
    # ========================================
    ws7 = wb.create_sheet("Piezas Rechazadas")
    escribir_encabezado_hoja(ws7, "DETALLE DE PIEZAS EN COTIZACIONES RECHAZADAS", 12)
    
    if not df_piezas_proveedor.empty:
        headers_piezas = [
            'Componente', 'Descripción', 'Proveedor', 'Cantidad',
            'Costo Unitario', 'Costo Total', 'Es Necesaria', 'Sugerida por Técnico',
            'Marca Equipo', 'Modelo Equipo',
            'Motivo Rechazo Cotización', 'Motivo Rechazo Pieza'
        ]
        escribir_headers(ws7, headers_piezas, 4)
        
        # Enriquecer con motivo de rechazo de la cotización padre
        motivos_por_cotizacion = dict(zip(df_rechazos['cotizacion_id'], df_rechazos['motivo_rechazo']))
        
        fila = 5
        for _, pieza in df_piezas_proveedor.iterrows():
            cot_id = pieza.get('cotizacion_id')
            motivo_cot = motivos_por_cotizacion.get(cot_id, '')
            motivo_cot_label = labels_motivos.get(motivo_cot, str(motivo_cot).replace('_', ' ').title()) if motivo_cot else ''
            
            costo_unit = float(pieza.get('costo_unitario', 0))
            cantidad = int(pieza.get('cantidad', 1))
            costo_total_pieza = costo_unit * cantidad
            
            ws7.cell(row=fila, column=1, value=pieza.get('componente__nombre', ''))
            ws7.cell(row=fila, column=2, value=pieza.get('descripcion_adicional', ''))
            ws7.cell(row=fila, column=3, value=pieza.get('proveedor', ''))
            ws7.cell(row=fila, column=4, value=cantidad)
            ws7.cell(row=fila, column=5, value=f'${costo_unit:,.2f}')
            ws7.cell(row=fila, column=6, value=f'${costo_total_pieza:,.2f}')
            ws7.cell(row=fila, column=7, value='Si' if pieza.get('es_necesaria') else 'No')
            ws7.cell(row=fila, column=8, value='Si' if pieza.get('sugerida_por_tecnico') else 'No')
            ws7.cell(row=fila, column=9, value=pieza.get('cotizacion__orden__detalle_equipo__marca', ''))
            ws7.cell(row=fila, column=10, value=pieza.get('cotizacion__orden__detalle_equipo__modelo', ''))
            ws7.cell(row=fila, column=11, value=motivo_cot_label)
            ws7.cell(row=fila, column=12, value=pieza.get('motivo_rechazo_pieza', ''))
            fila += 1
        
        # Sección de resumen de piezas
        fila += 2
        ws7.merge_cells(f'A{fila}:L{fila}')
        ws7[f'A{fila}'].value = "RESUMEN: TOP COMPONENTES MÁS FRECUENTES EN RECHAZOS"
        ws7[f'A{fila}'].font = section_font
        ws7[f'A{fila}'].fill = section_fill
        fila += 1
        
        escribir_headers(ws7, ['Componente', 'Veces en Rechazos', 'Costo Unit. Promedio', 'Es Necesaria (%)', 'Top Proveedores', '', '', '', '', '', '', ''], fila)
        fila += 1
        
        resumen_componentes = df_piezas_proveedor.groupby('componente__nombre').agg(
            veces=('id', 'count'),
            costo_prom=('costo_unitario', 'mean'),
            necesaria=('es_necesaria', 'mean'),
            proveedores=('proveedor', lambda x: ', '.join([str(p) for p in x.dropna().unique()[:3]]) if x.notna().any() else 'N/A'),
        ).reset_index().sort_values('veces', ascending=False).head(20)
        
        for _, comp in resumen_componentes.iterrows():
            ws7.cell(row=fila, column=1, value=comp['componente__nombre'])
            ws7.cell(row=fila, column=2, value=comp['veces'])
            ws7.cell(row=fila, column=3, value=f"${comp['costo_prom']:,.2f}")
            ws7.cell(row=fila, column=4, value=f"{comp['necesaria'] * 100:.0f}%")
            ws7.cell(row=fila, column=5, value=comp['proveedores'])
            fila += 1
    else:
        ws7.merge_cells('A4:L4')
        ws7['A4'].value = "No hay datos de piezas para las cotizaciones rechazadas."
        ws7['A4'].font = Font(italic=True, size=12)
    
    autoajustar_columnas(ws7, max_width=45)
    
    # ========================================
    # HOJA 8: RECHAZOS POR COSTO ALTO (detalle con piezas)
    # ========================================
    ws8 = wb.create_sheet("Costo Alto")
    costo_alto_fill = PatternFill(start_color='e67e22', end_color='e67e22', fill_type='solid')
    
    escribir_encabezado_hoja(ws8, "ANÁLISIS DETALLADO: RECHAZOS POR COSTO ELEVADO", 10)
    
    df_costo_alto = df_rechazos[df_rechazos['motivo_rechazo'] == 'costo_alto'].copy()
    
    # KPIs de "Costo Alto"
    ws8.merge_cells('A4:J4')
    ws8['A4'].value = "INDICADORES - RECHAZOS POR COSTO ELEVADO"
    ws8['A4'].font = Font(bold=True, size=12, color='FFFFFF')
    ws8['A4'].fill = costo_alto_fill
    ws8['A4'].alignment = Alignment(horizontal='center')
    
    total_costo_alto = len(df_costo_alto)
    pct_de_rechazos_ca = (total_costo_alto / total_rechazos * 100) if total_rechazos > 0 else 0
    pct_de_total_ca = (total_costo_alto / total_cotizaciones * 100) if total_cotizaciones > 0 else 0
    valor_perdido_ca = df_costo_alto['costo_total'].sum() if total_costo_alto > 0 else 0
    
    kpis_ca = [
        ['Casos por Costo Elevado', total_costo_alto],
        ['% de Todos los Rechazos', f'{pct_de_rechazos_ca:.1f}%'],
        ['% del Total de Cotizaciones', f'{pct_de_total_ca:.1f}%'],
        ['Valor Perdido por Costo Alto', f'${valor_perdido_ca:,.2f}'],
    ]
    
    if total_costo_alto > 0:
        kpis_ca.extend([
            ['Costo Promedio de Cotización', f'${df_costo_alto["costo_total"].mean():,.2f}'],
            ['Costo Mediana de Cotización', f'${df_costo_alto["costo_total"].median():,.2f}'],
            ['Piezas Promedio por Orden', f'{df_costo_alto["total_piezas"].mean():.1f}'],
        ])
    
    for i, (kpi_label_ca, kpi_val_ca) in enumerate(kpis_ca):
        row_num = 6 + i
        ws8.cell(row=row_num, column=1, value=kpi_label_ca).font = kpi_label_font
        ws8.cell(row=row_num, column=2, value=kpi_val_ca).font = kpi_value_font
    
    if total_costo_alto > 0:
        # Mapeo de cotizacion_id → orden_cliente
        mapa_orden_cliente = dict(zip(df_rechazos['cotizacion_id'], df_rechazos['orden_cliente']))
        
        # Sección: Detalle por cada orden con sus piezas desglosadas
        fila = 6 + len(kpis_ca) + 2
        ws8.merge_cells(f'A{fila}:J{fila}')
        ws8[f'A{fila}'].value = "DETALLE POR ORDEN - COSTO ALTO (con piezas desglosadas)"
        ws8[f'A{fila}'].font = section_font
        ws8[f'A{fila}'].fill = section_fill
        fila += 1
        
        for _, rec_ca in df_costo_alto.sort_values('costo_total', ascending=False).iterrows():
            cot_id = rec_ca['cotizacion_id']
            orden_cli = rec_ca.get('orden_cliente', '') or 'Sin orden cliente'
            
            # Formatear fecha de envío
            fecha_envio_ca = ''
            if pd.notna(rec_ca.get('fecha_envio')):
                try:
                    fecha_envio_ca = pd.to_datetime(rec_ca['fecha_envio']).strftime('%d/%m/%Y')
                except:
                    fecha_envio_ca = str(rec_ca['fecha_envio'])
            
            # Encabezado de la orden
            ws8.merge_cells(f'A{fila}:J{fila}')
            ws8[f'A{fila}'].value = (
                f"Orden Cliente: {orden_cli}  |  "
                f"Fecha: {fecha_envio_ca}  |  "
                f"Marca: {rec_ca.get('marca', '')}  |  "
                f"Modelo: {rec_ca.get('modelo', '')}  |  "
                f"Sucursal: {rec_ca.get('sucursal', '')}  |  "
                f"Técnico: {rec_ca.get('tecnico', '')}  |  "
                f"Detalle: {rec_ca.get('detalle_rechazo', '')}"
            )
            ws8[f'A{fila}'].font = Font(bold=True, size=10, color='FFFFFF')
            ws8[f'A{fila}'].fill = costo_alto_fill
            fila += 1
            
            # Headers de piezas para esta orden
            headers_piezas_ca = [
                'Componente', 'Descripción', 'Proveedor',
                'Cantidad', 'Costo Unitario', 'Costo Total Pieza',
                'Es Necesaria', 'Sugerida por Técnico', '', ''
            ]
            escribir_headers(ws8, headers_piezas_ca, fila)
            fila += 1
            
            # Obtener piezas de esta cotización
            if not df_piezas_proveedor.empty:
                piezas_orden = df_piezas_proveedor[
                    df_piezas_proveedor['cotizacion_id'] == cot_id
                ]
            else:
                piezas_orden = pd.DataFrame()
            
            subtotal_piezas = 0
            if not piezas_orden.empty:
                for _, pieza_ca in piezas_orden.iterrows():
                    costo_unit = float(pieza_ca.get('costo_unitario', 0))
                    cantidad = int(pieza_ca.get('cantidad', 1))
                    costo_total_pieza = costo_unit * cantidad
                    subtotal_piezas += costo_total_pieza
                    
                    ws8.cell(row=fila, column=1, value=pieza_ca.get('componente__nombre', ''))
                    ws8.cell(row=fila, column=2, value=pieza_ca.get('descripcion_adicional', ''))
                    ws8.cell(row=fila, column=3, value=pieza_ca.get('proveedor', ''))
                    ws8.cell(row=fila, column=4, value=cantidad)
                    ws8.cell(row=fila, column=5, value=f'${costo_unit:,.2f}')
                    ws8.cell(row=fila, column=6, value=f'${costo_total_pieza:,.2f}')
                    ws8.cell(row=fila, column=7, value='Sí' if pieza_ca.get('es_necesaria') else 'No')
                    ws8.cell(row=fila, column=8, value='Sí' if pieza_ca.get('sugerida_por_tecnico') else 'No')
                    fila += 1
            else:
                ws8.cell(row=fila, column=1, value='(Sin piezas registradas)')
                ws8[f'A{fila}'].font = Font(italic=True, color='999999')
                fila += 1
            
            # Fila de totales para esta orden
            ws8.cell(row=fila, column=4, value='TOTAL ORDEN:').font = Font(bold=True)
            ws8.cell(row=fila, column=5, value=f'Piezas: ${subtotal_piezas:,.2f}').font = Font(bold=True)
            ws8.cell(row=fila, column=6, value=f'M.O.: ${rec_ca.get("costo_mano_obra", 0):,.2f}').font = Font(bold=True)
            ws8.cell(row=fila, column=7, value=f'Total: ${rec_ca.get("costo_total", 0):,.2f}')
            ws8[f'G{fila}'].font = Font(bold=True, size=11, color='c0392c')
            for c in range(1, 11):
                ws8.cell(row=fila, column=c).fill = orange_fill
            fila += 2  # Espacio entre órdenes
        
        # Sección: Resumen comparativo (tabla resumen)
        ws8.merge_cells(f'A{fila}:J{fila}')
        ws8[f'A{fila}'].value = "RESUMEN: TODAS LAS ÓRDENES RECHAZADAS POR COSTO ALTO"
        ws8[f'A{fila}'].font = section_font
        ws8[f'A{fila}'].fill = section_fill
        fila += 1
        
        escribir_headers(ws8, [
            'Orden Cliente', 'Fecha Cotización', 'Marca', 'Modelo', 'Sucursal', 'Técnico',
            'Total Piezas', 'Costo Piezas', 'Costo M.O.', 'Costo Total'
        ], fila)
        fila += 1
        
        for _, rec_ca in df_costo_alto.sort_values('costo_total', ascending=False).iterrows():
            fecha_envio_resumen = ''
            if pd.notna(rec_ca.get('fecha_envio')):
                try:
                    fecha_envio_resumen = pd.to_datetime(rec_ca['fecha_envio']).strftime('%d/%m/%Y')
                except:
                    fecha_envio_resumen = str(rec_ca['fecha_envio'])
            
            ws8.cell(row=fila, column=1, value=rec_ca.get('orden_cliente', ''))
            ws8.cell(row=fila, column=2, value=fecha_envio_resumen)
            ws8.cell(row=fila, column=3, value=rec_ca.get('marca', ''))
            ws8.cell(row=fila, column=4, value=rec_ca.get('modelo', ''))
            ws8.cell(row=fila, column=5, value=rec_ca.get('sucursal', ''))
            ws8.cell(row=fila, column=6, value=rec_ca.get('tecnico', ''))
            ws8.cell(row=fila, column=7, value=rec_ca.get('total_piezas', 0))
            ws8.cell(row=fila, column=8, value=f'${rec_ca.get("costo_total_piezas", 0):,.2f}')
            ws8.cell(row=fila, column=9, value=f'${rec_ca.get("costo_mano_obra", 0):,.2f}')
            ws8.cell(row=fila, column=10, value=f'${rec_ca.get("costo_total", 0):,.2f}')
            fila += 1
    else:
        ws8.merge_cells('A6:J6')
        ws8['A6'].value = "No se encontraron cotizaciones rechazadas por costo elevado en el período seleccionado."
        ws8['A6'].font = Font(italic=True, size=12)
        ws8['A6'].alignment = Alignment(horizontal='center')
    
    autoajustar_columnas(ws8, max_width=50)
    
    # ========================================
    # HOJA 9: SERVICIOS CON 3+ Y 4+ PIEZAS COTIZADAS
    # ========================================
    ws9 = wb.create_sheet("Servicios 3+ Piezas")
    multi_piezas_fill = PatternFill(start_color='8e44ad', end_color='8e44ad', fill_type='solid')
    
    escribir_encabezado_hoja(ws9, "SERVICIOS CON MÚLTIPLES PIEZAS COTIZADAS (RECHAZADAS)", 10)
    
    # Mapeo cotizacion_id → orden_cliente para esta hoja
    mapa_oc = dict(zip(df_rechazos['cotizacion_id'], df_rechazos['orden_cliente']))
    
    # ---- SECCIÓN A: Servicios con 3 o más piezas ----
    df_3_plus = df_rechazos[df_rechazos['total_piezas'] >= 3].copy()
    total_3_plus = len(df_3_plus)
    
    ws9.merge_cells('A4:J4')
    ws9['A4'].value = f"SECCIÓN A: SERVICIOS CON 3 O MÁS PIEZAS COTIZADAS ({total_3_plus} registros)"
    ws9['A4'].font = Font(bold=True, size=12, color='FFFFFF')
    ws9['A4'].fill = multi_piezas_fill
    ws9['A4'].alignment = Alignment(horizontal='center')
    
    if total_3_plus > 0:
        headers_multi = [
            'Orden Cliente', 'Fecha Cotización', 'Marca', 'Modelo', 'Sucursal', 'Técnico',
            'Motivo Rechazo', 'Total Piezas', 'Costo Piezas', 'Costo M.O.', 'Costo Total'
        ]
        escribir_headers(ws9, headers_multi, 5)
        
        fila = 6
        suma_total_3 = 0
        for _, rec_mp in df_3_plus.sort_values('total_piezas', ascending=False).iterrows():
            motivo_lab = labels_motivos.get(
                rec_mp.get('motivo_rechazo', ''),
                str(rec_mp.get('motivo_rechazo', '')).replace('_', ' ').title()
            ) if rec_mp.get('motivo_rechazo') else 'Sin motivo'
            
            costo_t = rec_mp.get('costo_total', 0)
            suma_total_3 += costo_t
            
            fecha_envio_3 = ''
            if pd.notna(rec_mp.get('fecha_envio')):
                try:
                    fecha_envio_3 = pd.to_datetime(rec_mp['fecha_envio']).strftime('%d/%m/%Y')
                except:
                    fecha_envio_3 = str(rec_mp['fecha_envio'])
            
            ws9.cell(row=fila, column=1, value=rec_mp.get('orden_cliente', ''))
            ws9.cell(row=fila, column=2, value=fecha_envio_3)
            ws9.cell(row=fila, column=3, value=rec_mp.get('marca', ''))
            ws9.cell(row=fila, column=4, value=rec_mp.get('modelo', ''))
            ws9.cell(row=fila, column=5, value=rec_mp.get('sucursal', ''))
            ws9.cell(row=fila, column=6, value=rec_mp.get('tecnico', ''))
            ws9.cell(row=fila, column=7, value=motivo_lab)
            ws9.cell(row=fila, column=8, value=rec_mp.get('total_piezas', 0))
            ws9.cell(row=fila, column=9, value=f'${rec_mp.get("costo_total_piezas", 0):,.2f}')
            ws9.cell(row=fila, column=10, value=f'${rec_mp.get("costo_mano_obra", 0):,.2f}')
            ws9.cell(row=fila, column=11, value=f'${costo_t:,.2f}')
            fila += 1
        
        # Fila de totales
        ws9.cell(row=fila, column=7, value='TOTAL:').font = Font(bold=True)
        ws9.cell(row=fila, column=8, value=f'{df_3_plus["total_piezas"].sum()} piezas').font = Font(bold=True)
        ws9.cell(row=fila, column=11, value=f'${suma_total_3:,.2f}')
        ws9[f'K{fila}'].font = Font(bold=True, size=11, color='8e44ad')
        for c in range(1, 12):
            ws9.cell(row=fila, column=c).fill = blue_fill
        fila += 1
    else:
        fila = 6
        ws9.cell(row=fila, column=1, value='No hay servicios rechazados con 3 o más piezas cotizadas.')
        ws9[f'A{fila}'].font = Font(italic=True)
        fila += 1
    
    # ---- SECCIÓN B: Servicios con 4 o más piezas ----
    fila += 2
    df_4_plus = df_rechazos[df_rechazos['total_piezas'] >= 4].copy()
    total_4_plus = len(df_4_plus)
    
    ws9.merge_cells(f'A{fila}:J{fila}')
    ws9[f'A{fila}'].value = f"SECCIÓN B: SERVICIOS CON 4 O MÁS PIEZAS COTIZADAS ({total_4_plus} registros)"
    ws9[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws9[f'A{fila}'].fill = multi_piezas_fill
    ws9[f'A{fila}'].alignment = Alignment(horizontal='center')
    fila += 1
    
    if total_4_plus > 0:
        escribir_headers(ws9, [
            'Orden Cliente', 'Fecha Cotización', 'Marca', 'Modelo', 'Sucursal', 'Técnico',
            'Motivo Rechazo', 'Total Piezas', 'Costo Piezas', 'Costo M.O.', 'Costo Total'
        ], fila)
        fila += 1
        
        suma_total_4 = 0
        for _, rec_mp in df_4_plus.sort_values('total_piezas', ascending=False).iterrows():
            motivo_lab = labels_motivos.get(
                rec_mp.get('motivo_rechazo', ''),
                str(rec_mp.get('motivo_rechazo', '')).replace('_', ' ').title()
            ) if rec_mp.get('motivo_rechazo') else 'Sin motivo'
            
            costo_t = rec_mp.get('costo_total', 0)
            suma_total_4 += costo_t
            
            fecha_envio_4 = ''
            if pd.notna(rec_mp.get('fecha_envio')):
                try:
                    fecha_envio_4 = pd.to_datetime(rec_mp['fecha_envio']).strftime('%d/%m/%Y')
                except:
                    fecha_envio_4 = str(rec_mp['fecha_envio'])
            
            ws9.cell(row=fila, column=1, value=rec_mp.get('orden_cliente', ''))
            ws9.cell(row=fila, column=2, value=fecha_envio_4)
            ws9.cell(row=fila, column=3, value=rec_mp.get('marca', ''))
            ws9.cell(row=fila, column=4, value=rec_mp.get('modelo', ''))
            ws9.cell(row=fila, column=5, value=rec_mp.get('sucursal', ''))
            ws9.cell(row=fila, column=6, value=rec_mp.get('tecnico', ''))
            ws9.cell(row=fila, column=7, value=motivo_lab)
            ws9.cell(row=fila, column=8, value=rec_mp.get('total_piezas', 0))
            ws9.cell(row=fila, column=9, value=f'${rec_mp.get("costo_total_piezas", 0):,.2f}')
            ws9.cell(row=fila, column=10, value=f'${rec_mp.get("costo_mano_obra", 0):,.2f}')
            ws9.cell(row=fila, column=11, value=f'${costo_t:,.2f}')
            fila += 1
        
        # Fila de totales
        ws9.cell(row=fila, column=7, value='TOTAL:').font = Font(bold=True)
        ws9.cell(row=fila, column=8, value=f'{df_4_plus["total_piezas"].sum()} piezas').font = Font(bold=True)
        ws9.cell(row=fila, column=11, value=f'${suma_total_4:,.2f}')
        ws9[f'K{fila}'].font = Font(bold=True, size=11, color='8e44ad')
        for c in range(1, 12):
            ws9.cell(row=fila, column=c).fill = blue_fill
        fila += 1
    else:
        ws9.cell(row=fila, column=1, value='No hay servicios rechazados con 4 o más piezas cotizadas.')
        ws9[f'A{fila}'].font = Font(italic=True)
        fila += 1
    
    # ---- SECCIÓN C: Desglose de piezas por orden (solo 4+) ----
    if total_4_plus > 0 and not df_piezas_proveedor.empty:
        fila += 2
        ws9.merge_cells(f'A{fila}:J{fila}')
        ws9[f'A{fila}'].value = "DESGLOSE DE PIEZAS: SERVICIOS CON 4+ PIEZAS"
        ws9[f'A{fila}'].font = section_font
        ws9[f'A{fila}'].fill = section_fill
        fila += 1
        
        for _, rec_mp in df_4_plus.sort_values('total_piezas', ascending=False).iterrows():
            cot_id = rec_mp['cotizacion_id']
            orden_cli = rec_mp.get('orden_cliente', '') or 'Sin orden cliente'
            
            # Encabezado de la orden
            ws9.merge_cells(f'A{fila}:J{fila}')
            fecha_envio_c = ''
            if pd.notna(rec_mp.get('fecha_envio')):
                try:
                    fecha_envio_c = pd.to_datetime(rec_mp['fecha_envio']).strftime('%d/%m/%Y')
                except:
                    fecha_envio_c = str(rec_mp['fecha_envio'])
            
            ws9[f'A{fila}'].value = (
                f"Orden Cliente: {orden_cli}  |  "
                f"Fecha: {fecha_envio_c}  |  "
                f"{rec_mp.get('marca', '')} {rec_mp.get('modelo', '')}  |  "
                f"Piezas: {rec_mp.get('total_piezas', 0)}  |  "
                f"Total: ${rec_mp.get('costo_total', 0):,.2f}"
            )
            ws9[f'A{fila}'].font = Font(bold=True, size=10, color='FFFFFF')
            ws9[f'A{fila}'].fill = multi_piezas_fill
            fila += 1
            
            piezas_esta_orden = df_piezas_proveedor[
                df_piezas_proveedor['cotizacion_id'] == cot_id
            ]
            
            if not piezas_esta_orden.empty:
                escribir_headers(ws9, [
                    'Componente', 'Descripción', 'Proveedor',
                    'Cantidad', 'Costo Unitario', 'Costo Total Pieza',
                    'Es Necesaria', 'Sugerida Técnico', '', ''
                ], fila)
                fila += 1
                
                sub_total = 0
                for _, pieza_mp in piezas_esta_orden.iterrows():
                    cu = float(pieza_mp.get('costo_unitario', 0))
                    cant = int(pieza_mp.get('cantidad', 1))
                    ct = cu * cant
                    sub_total += ct
                    
                    ws9.cell(row=fila, column=1, value=pieza_mp.get('componente__nombre', ''))
                    ws9.cell(row=fila, column=2, value=pieza_mp.get('descripcion_adicional', ''))
                    ws9.cell(row=fila, column=3, value=pieza_mp.get('proveedor', ''))
                    ws9.cell(row=fila, column=4, value=cant)
                    ws9.cell(row=fila, column=5, value=f'${cu:,.2f}')
                    ws9.cell(row=fila, column=6, value=f'${ct:,.2f}')
                    ws9.cell(row=fila, column=7, value='Sí' if pieza_mp.get('es_necesaria') else 'No')
                    ws9.cell(row=fila, column=8, value='Sí' if pieza_mp.get('sugerida_por_tecnico') else 'No')
                    fila += 1
                
                ws9.cell(row=fila, column=5, value='Subtotal piezas:').font = Font(bold=True)
                ws9.cell(row=fila, column=6, value=f'${sub_total:,.2f}').font = Font(bold=True)
            else:
                ws9.cell(row=fila, column=1, value='(Sin piezas registradas)')
                ws9[f'A{fila}'].font = Font(italic=True, color='999999')
            fila += 2
    
    autoajustar_columnas(ws9, max_width=50)
    
    # ========================================
    # GENERAR Y RETORNAR ARCHIVO
    # ========================================
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'Analisis_Rechazos_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    
    return response


# ============================================================================
# EXPORTACIÓN EXCEL: ANÁLISIS DE COTIZACIONES ACEPTADAS
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def exportar_analisis_aceptaciones(request):
    """
    Exporta un análisis exhaustivo de cotizaciones aceptadas a Excel con 9 hojas,
    incluyendo datos cruzados con VentaMostrador (servicios, paquetes, piezas).
    
    EXPLICACIÓN PARA PRINCIPIANTES:
    Este Excel es el espejo del 'Análisis de Rechazos', pero enfocado en lo positivo:
    qué se aceptó, cuánto se generó, qué servicios adicionales (VM) se vendieron,
    y cómo se comportó la cadena de suministro post-aceptación.
    
    Hojas:
        1. Resumen Aceptaciones - KPIs principales de aceptaciones + VM
        2. Detalle Aceptaciones - Cada cotización aceptada con datos completos
        3. Piezas Aceptadas - Detalle a nivel pieza individual
        4. Aceptación Parcial - Cotizaciones con piezas mixtas (aceptadas/rechazadas)
        5. Ventas Mostrador - VM asociadas a cotizaciones aceptadas
        6. Servicios Adicionales - Análisis de servicios VM (limpieza, reinstalación, etc.)
        7. Seguimiento de Piezas - Estado del tracking post-aceptación
        8. Rendimiento por Técnico - Métricas de aceptación + upsell por técnico
        9. Rendimiento por Sucursal - Métricas de aceptación + upsell por sucursal
    
    Returns:
        HttpResponse: Archivo Excel (.xlsx) para descargar
    """
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import pandas as pd
    
    from .utils_cotizaciones import (
        obtener_dataframe_cotizaciones,
        calcular_kpis_generales,
        calcular_kpis_aceptaciones,
        analizar_piezas_cotizadas,
        analizar_servicios_vm_aceptadas,
        analizar_seguimiento_piezas_aceptadas,
    )
    
    # ========================================
    # OBTENER DATOS CON MISMOS FILTROS QUE EL DASHBOARD
    # ========================================
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sucursal_id = request.GET.get('sucursal')
    tecnico_id = request.GET.get('tecnico')
    gama = request.GET.get('gama')
    
    df_cotizaciones = obtener_dataframe_cotizaciones(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama
    )
    
    if df_cotizaciones.empty:
        messages.error(request, 'No hay datos para exportar con los filtros aplicados.')
        return redirect('servicio_tecnico:dashboard_cotizaciones')
    
    # Filtrar solo aceptadas
    df_aceptadas = df_cotizaciones[df_cotizaciones['aceptada'] == True].copy()
    
    if df_aceptadas.empty:
        messages.warning(request, 'No hay cotizaciones aceptadas en el período seleccionado.')
        return redirect('servicio_tecnico:dashboard_cotizaciones')
    
    # KPIs
    kpis_generales = calcular_kpis_generales(df_cotizaciones)
    kpis_aceptaciones = calcular_kpis_aceptaciones(df_cotizaciones)
    
    # Análisis VM y seguimiento
    analisis_vm = analizar_servicios_vm_aceptadas(
        df_cotizaciones,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sucursal_id=sucursal_id,
        tecnico_id=tecnico_id,
        gama=gama,
    )
    analisis_seguimiento = analizar_seguimiento_piezas_aceptadas(df_cotizaciones)
    
    # Piezas aceptadas
    cotizacion_ids_aceptadas = df_aceptadas['cotizacion_id'].tolist()
    
    from .models import PiezaCotizada, SeguimientoPieza, VentaMostrador, PiezaVentaMostrador
    piezas_aceptadas_qs = PiezaCotizada.objects.filter(
        cotizacion_id__in=cotizacion_ids_aceptadas
    ).select_related(
        'componente',
        'cotizacion',
        'cotizacion__orden',
        'cotizacion__orden__detalle_equipo'
    )
    
    # ========================================
    # CREAR WORKBOOK
    # ========================================
    wb = Workbook()
    wb.remove(wb.active)
    
    # Estilos reutilizables
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')  # Verde
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    title_font = Font(bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    title_align = Alignment(horizontal='center', vertical='center')
    
    subtitle_font = Font(italic=True, size=10, color='666666')
    
    kpi_label_font = Font(bold=True, size=11)
    kpi_value_font = Font(bold=True, size=12, color='198754')  # Verde para aceptaciones
    
    green_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    blue_fill = PatternFill(start_color='d1ecf1', end_color='d1ecf1', fill_type='solid')
    yellow_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
    purple_fill = PatternFill(start_color='e2d5f1', end_color='e2d5f1', fill_type='solid')
    teal_fill = PatternFill(start_color='d1f2eb', end_color='d1f2eb', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    number_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    
    def aplicar_estilos_header(ws, fila, num_cols):
        """Aplica estilos a la fila de encabezados."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=fila, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border_thin
    
    def auto_ajustar_columnas(ws, max_width=50):
        """Autoajusta el ancho de las columnas."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)
    
    # ========================================================================
    # HOJA 1: RESUMEN ACEPTACIONES
    # ========================================================================
    ws1 = wb.create_sheet('Resumen Aceptaciones')
    
    # Título
    ws1.merge_cells('A1:F1')
    ws1['A1'] = 'ANÁLISIS DE COTIZACIONES ACEPTADAS'
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = title_align
    
    # Subtítulo con filtros
    filtros_texto = f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actual'}"
    ws1.merge_cells('A2:F2')
    ws1['A2'] = filtros_texto
    ws1['A2'].font = subtitle_font
    ws1['A2'].alignment = Alignment(horizontal='center')
    
    # Sección: KPIs principales
    fila = 4
    ws1.merge_cells(f'A{fila}:F{fila}')
    ws1[f'A{fila}'] = 'INDICADORES PRINCIPALES'
    ws1[f'A{fila}'].font = Font(bold=True, size=12)
    ws1[f'A{fila}'].fill = green_fill
    
    fila += 1
    headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws1, fila, len(headers))
    
    # Datos de KPIs
    metricas = [
        ('Total Cotizaciones', kpis_generales['total_cotizaciones'], '', 'Todas las cotizaciones en el período'),
        ('Total Aceptadas', kpis_aceptaciones['total_aceptadas'], f"{kpis_generales.get('tasa_aceptacion', 0)}%", 'Tasa de aceptación global'),
        ('Aceptación Total', kpis_aceptaciones['aceptacion_total_count'], f"{kpis_aceptaciones['aceptacion_total_pct']}%", 'Todas las piezas aceptadas'),
        ('Aceptación Parcial', kpis_aceptaciones['aceptacion_parcial_count'], f"{kpis_aceptaciones['aceptacion_parcial_pct']}%", 'Algunas piezas rechazadas'),
        ('Valor Total Aceptado', kpis_aceptaciones['valor_total_aceptado'], f"{kpis_aceptaciones['porcentaje_recuperacion']}% del cotizado", 'Monto final que paga el cliente'),
        ('Ticket Promedio Aceptado', kpis_aceptaciones['ticket_promedio_aceptado'], '', 'Monto promedio por cotización aceptada'),
        ('Piezas Promedio por Aceptada', kpis_aceptaciones['piezas_promedio_aceptadas'], '', 'Promedio de piezas aceptadas'),
        ('Con Descuento Mano de Obra', kpis_aceptaciones['descuento_count'], f"{kpis_aceptaciones['descuento_pct']}%", f"Ahorro total: {kpis_aceptaciones['descuento_monto_total_fmt']}"),
        ('Tiempo Respuesta Prom. (días)', kpis_aceptaciones['tiempo_respuesta_aceptadas'], '', 'Días promedio hasta respuesta positiva'),
        ('--- VENTA MOSTRADOR ---', '', '', ''),
        ('Aceptadas con VM', kpis_aceptaciones['con_vm_count'], f"{kpis_aceptaciones['con_vm_pct']}%", 'Tasa de upsell'),
        ('Valor VM Complementario', kpis_aceptaciones['valor_vm_complementario'], '', 'Ingreso adicional por VM'),
        ('Valor Combinado Total', kpis_aceptaciones['valor_combinado_total'], '', 'Cotización + VM'),
        ('Paquete Más Vendido', kpis_aceptaciones['paquete_mas_vendido'], '', ''),
        ('Servicio Más Vendido', kpis_aceptaciones['servicio_mas_vendido'], '', ''),
    ]
    
    for metrica, valor, pct, obs in metricas:
        fila += 1
        ws1.cell(row=fila, column=1, value=metrica).font = kpi_label_font
        cell_val = ws1.cell(row=fila, column=2, value=valor)
        if isinstance(valor, (int, float)) and valor > 100:
            cell_val.number_format = number_fmt
        cell_val.font = kpi_value_font
        ws1.cell(row=fila, column=3, value=pct)
        ws1.cell(row=fila, column=4, value=obs).font = Font(italic=True, color='666666')
        for col in range(1, 5):
            ws1.cell(row=fila, column=col).border = border_thin
    
    auto_ajustar_columnas(ws1)
    
    # ========================================================================
    # HOJA 2: DETALLE ACEPTACIONES
    # ========================================================================
    ws2 = wb.create_sheet('Detalle Aceptaciones')
    
    # Título
    ws2.merge_cells('A1:Y1')
    ws2['A1'] = f'DETALLE DE COTIZACIONES ACEPTADAS ({len(df_aceptadas)} registros)'
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = title_align
    
    # Headers
    headers_detalle = [
        'Orden Cliente', 'Número Serie', 'Marca', 'Modelo', 'Tipo Equipo',
        'Gama', 'Sucursal', 'Técnico', 'Responsable',
        'Fecha Envío', 'Fecha Respuesta', 'Días Respuesta',
        'Costo Total Cotizado', 'Costo Piezas Aceptadas', 'Costo Piezas Rechazadas',
        'Costo Mano Obra', 'Descuento MO', 'Costo Total Final',
        'Total Piezas', 'Piezas Aceptadas', 'Piezas Rechazadas',
        'Tiene VM', 'Paquete VM', 'Total VM', 'Valor Combinado',
    ]
    
    fila = 3
    for col, h in enumerate(headers_detalle, 1):
        ws2.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws2, fila, len(headers_detalle))
    
    # Datos
    for _, row in df_aceptadas.iterrows():
        fila += 1
        
        # Convertir fechas a string para evitar error de openpyxl con timezones
        # (openpyxl no soporta datetime con tzinfo != None)
        fecha_envio_str = ''
        if pd.notna(row.get('fecha_envio')):
            try:
                fecha_envio_str = pd.to_datetime(row['fecha_envio']).strftime('%d/%m/%Y')
            except Exception:
                fecha_envio_str = str(row['fecha_envio'])
        
        fecha_respuesta_str = ''
        if pd.notna(row.get('fecha_respuesta')):
            try:
                fecha_respuesta_str = pd.to_datetime(row['fecha_respuesta']).strftime('%d/%m/%Y')
            except Exception:
                fecha_respuesta_str = str(row['fecha_respuesta'])
        
        datos = [
            row.get('orden_cliente', ''),
            row.get('numero_serie', ''),
            row.get('marca', ''),
            row.get('modelo', ''),
            row.get('tipo_equipo', ''),
            row.get('gama', ''),
            row.get('sucursal', ''),
            row.get('tecnico', ''),
            row.get('responsable', ''),
            fecha_envio_str,
            fecha_respuesta_str,
            row.get('dias_sin_respuesta', ''),
            row.get('costo_total', 0),
            row.get('costo_piezas_aceptadas', 0),
            row.get('costo_piezas_rechazadas', 0),
            row.get('costo_mano_obra', 0),
            row.get('monto_descuento', 0),
            row.get('costo_total_final', 0),
            row.get('total_piezas', 0),
            row.get('piezas_aceptadas', 0),
            row.get('piezas_rechazadas', 0),
            'Sí' if row.get('tiene_venta_mostrador', False) else 'No',
            row.get('vm_paquete', 'ninguno').capitalize() if row.get('tiene_venta_mostrador', False) else '',
            row.get('vm_total_venta', 0) if row.get('tiene_venta_mostrador', False) else 0,
            row.get('valor_total_combinado', 0),
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws2.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (13, 14, 15, 16, 17, 18, 24, 25):
                cell.number_format = number_fmt
    
    auto_ajustar_columnas(ws2)
    
    # ========================================================================
    # HOJA 3: PIEZAS ACEPTADAS
    # ========================================================================
    ws3 = wb.create_sheet('Piezas Aceptadas')
    
    ws3.merge_cells('A1:L1')
    ws3['A1'] = 'DETALLE DE PIEZAS EN COTIZACIONES ACEPTADAS'
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = title_align
    
    headers_piezas = [
        'Orden', 'Componente', 'Descripción', 'Proveedor',
        'Cantidad', 'Costo Unitario', 'Costo Total',
        'Es Necesaria', 'Sugerida por Técnico',
        'Aceptada por Cliente', 'Motivo Rechazo Pieza', 'Prioridad',
    ]
    
    fila = 3
    for col, h in enumerate(headers_piezas, 1):
        ws3.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws3, fila, len(headers_piezas))
    
    for pieza in piezas_aceptadas_qs:
        fila += 1
        detalle = pieza.cotizacion.orden.detalle_equipo if hasattr(pieza.cotizacion.orden, 'detalle_equipo') else None
        orden_cliente = detalle.orden_cliente if detalle else ''
        
        # Determinar aceptación con herencia
        aceptada = pieza.aceptada_por_cliente
        if aceptada is None:
            aceptada = pieza.cotizacion.usuario_acepto
        
        datos = [
            orden_cliente,
            pieza.componente.nombre if pieza.componente else '',
            pieza.descripcion_adicional,
            pieza.proveedor,
            pieza.cantidad,
            float(pieza.costo_unitario),
            float(pieza.costo_total),
            'Sí' if pieza.es_necesaria else 'No',
            'Sí' if pieza.sugerida_por_tecnico else 'No',
            'Sí' if aceptada else ('No' if aceptada is False else 'Pendiente'),
            pieza.motivo_rechazo_pieza or '',
            pieza.orden_prioridad,
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws3.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (6, 7):
                cell.number_format = number_fmt
            # Colorear según aceptación
            if col == 10:
                if valor == 'Sí':
                    cell.fill = green_fill
                elif valor == 'No':
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
    
    # Resumen de piezas
    fila += 2
    total_piezas_todas = piezas_aceptadas_qs.count()
    piezas_si = piezas_aceptadas_qs.filter(aceptada_por_cliente=True).count()
    piezas_no = piezas_aceptadas_qs.filter(aceptada_por_cliente=False).count()
    piezas_pendientes = total_piezas_todas - piezas_si - piezas_no
    
    ws3.cell(row=fila, column=1, value='RESUMEN').font = Font(bold=True, size=12)
    fila += 1
    ws3.cell(row=fila, column=1, value='Total piezas:').font = kpi_label_font
    ws3.cell(row=fila, column=2, value=total_piezas_todas)
    fila += 1
    ws3.cell(row=fila, column=1, value='Aceptadas:').font = kpi_label_font
    ws3.cell(row=fila, column=2, value=piezas_si).fill = green_fill
    fila += 1
    ws3.cell(row=fila, column=1, value='Rechazadas:').font = kpi_label_font
    ws3.cell(row=fila, column=2, value=piezas_no)
    fila += 1
    ws3.cell(row=fila, column=1, value='Heredan aceptación:').font = kpi_label_font
    ws3.cell(row=fila, column=2, value=piezas_pendientes)
    
    # Top componentes más aceptados
    fila += 2
    ws3.cell(row=fila, column=1, value='TOP COMPONENTES MÁS ACEPTADOS').font = Font(bold=True, size=12)
    ws3.cell(row=fila, column=1).fill = green_fill
    fila += 1
    
    from django.db.models import Sum as DjSum, Count as DjCount
    top_comp = piezas_aceptadas_qs.filter(
        aceptada_por_cliente=True
    ).values('componente__nombre').annotate(
        total=DjCount('id'),
        ingreso=DjSum('costo_unitario'),
    ).order_by('-total')[:10]
    
    headers_top = ['Componente', 'Cantidad', 'Ingreso Total']
    for col, h in enumerate(headers_top, 1):
        ws3.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws3, fila, len(headers_top))
    
    for comp in top_comp:
        fila += 1
        ws3.cell(row=fila, column=1, value=comp['componente__nombre'] or 'Sin nombre').border = border_thin
        ws3.cell(row=fila, column=2, value=comp['total']).border = border_thin
        cell_ing = ws3.cell(row=fila, column=3, value=float(comp['ingreso'] or 0))
        cell_ing.number_format = number_fmt
        cell_ing.border = border_thin
    
    auto_ajustar_columnas(ws3)
    
    # ========================================================================
    # HOJA 4: ACEPTACIÓN PARCIAL
    # ========================================================================
    ws4 = wb.create_sheet('Aceptación Parcial')
    
    df_parcial = df_aceptadas[
        (df_aceptadas['piezas_rechazadas'] > 0) & (df_aceptadas['piezas_aceptadas'] > 0)
    ]
    
    ws4.merge_cells('A1:N1')
    ws4['A1'] = f'COTIZACIONES CON ACEPTACIÓN PARCIAL ({len(df_parcial)} registros)'
    ws4['A1'].font = title_font
    ws4['A1'].fill = title_fill
    ws4['A1'].alignment = title_align
    
    ws4.merge_cells('A2:N2')
    ws4['A2'] = 'Cotizaciones donde el cliente aceptó ALGUNAS piezas pero rechazó otras'
    ws4['A2'].font = subtitle_font
    ws4['A2'].alignment = Alignment(horizontal='center')
    
    headers_parcial = [
        'Orden Cliente', 'Marca', 'Modelo', 'Sucursal', 'Técnico',
        'Total Piezas', 'Piezas Aceptadas', 'Piezas Rechazadas',
        '% Aceptadas', 'Costo Cotizado', 'Costo Aceptado', 'Costo Rechazado',
        'Diferencia', 'Tiene VM',
    ]
    
    fila = 4
    for col, h in enumerate(headers_parcial, 1):
        ws4.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws4, fila, len(headers_parcial))
    
    for _, row in df_parcial.iterrows():
        fila += 1
        diferencia = row.get('costo_total', 0) - row.get('costo_total_final', 0)
        datos = [
            row.get('orden_cliente', ''),
            row.get('marca', ''),
            row.get('modelo', ''),
            row.get('sucursal', ''),
            row.get('tecnico', ''),
            row.get('total_piezas', 0),
            row.get('piezas_aceptadas', 0),
            row.get('piezas_rechazadas', 0),
            row.get('porcentaje_aceptadas', 0),
            row.get('costo_total', 0),
            row.get('costo_total_final', 0),
            row.get('costo_piezas_rechazadas', 0),
            diferencia,
            'Sí' if row.get('tiene_venta_mostrador', False) else 'No',
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws4.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (10, 11, 12, 13):
                cell.number_format = number_fmt
    
    # Resumen
    if len(df_parcial) > 0:
        fila += 2
        ws4.cell(row=fila, column=1, value='RESUMEN DE ACEPTACIÓN PARCIAL').font = Font(bold=True, size=12)
        ws4.cell(row=fila, column=1).fill = yellow_fill
        fila += 1
        valor_perdido = df_parcial['costo_piezas_rechazadas'].sum()
        ws4.cell(row=fila, column=1, value='Valor perdido por piezas rechazadas en parciales:').font = kpi_label_font
        cell_vp = ws4.cell(row=fila, column=2, value=float(valor_perdido))
        cell_vp.number_format = number_fmt
        cell_vp.font = Font(bold=True, color='c0392c', size=12)
        fila += 1
        prom_aceptacion = df_parcial['porcentaje_aceptadas'].mean()
        ws4.cell(row=fila, column=1, value='% promedio de aceptación en parciales:').font = kpi_label_font
        ws4.cell(row=fila, column=2, value=f"{prom_aceptacion:.1f}%")
    
    auto_ajustar_columnas(ws4)
    
    # ========================================================================
    # HOJA 5: VENTAS MOSTRADOR ASOCIADAS
    # ========================================================================
    ws5 = wb.create_sheet('Ventas Mostrador')
    
    df_con_vm = df_aceptadas[df_aceptadas['tiene_venta_mostrador'] == True]
    
    ws5.merge_cells('A1:R1')
    ws5['A1'] = f'VENTAS MOSTRADOR EN COTIZACIONES ACEPTADAS ({len(df_con_vm)} registros)'
    ws5['A1'].font = title_font
    ws5['A1'].fill = PatternFill(start_color='0dcaf0', end_color='0dcaf0', fill_type='solid')
    ws5['A1'].alignment = title_align
    
    headers_vm = [
        'Orden Cliente', 'Folio VM', 'Fecha Venta', 'Sucursal', 'Técnico',
        'Paquete', 'Costo Paquete',
        'Limpieza', 'Costo Limpieza',
        'Reinstalación SO', 'Costo Reinstalación',
        'Respaldo', 'Costo Respaldo',
        'Cambio Pieza', 'Costo Cambio',
        'Kit Limpieza', 'Costo Kit',
        'Total VM',
    ]
    
    fila = 3
    for col, h in enumerate(headers_vm, 1):
        ws5.cell(row=fila, column=col, value=h)
    # Usar fill azul info para headers de VM
    vm_header_fill = PatternFill(start_color='0dcaf0', end_color='0dcaf0', fill_type='solid')
    for col in range(1, len(headers_vm) + 1):
        cell = ws5.cell(row=fila, column=col)
        cell.font = Font(bold=True, color='000000', size=11)
        cell.fill = vm_header_fill
        cell.alignment = header_align
        cell.border = border_thin
    
    for _, row in df_con_vm.iterrows():
        fila += 1
        
        # Convertir fecha de venta a string para evitar error de openpyxl con timezones
        vm_fecha_venta_str = ''
        if pd.notna(row.get('vm_fecha_venta')):
            try:
                vm_fecha_venta_str = pd.to_datetime(row['vm_fecha_venta']).strftime('%d/%m/%Y')
            except Exception:
                vm_fecha_venta_str = str(row['vm_fecha_venta'])
        
        datos = [
            row.get('orden_cliente', ''),
            row.get('vm_folio', ''),
            vm_fecha_venta_str,
            row.get('sucursal', ''),
            row.get('tecnico', ''),
            row.get('vm_paquete', 'ninguno').capitalize(),
            row.get('vm_costo_paquete', 0),
            'Sí' if row.get('vm_incluye_limpieza', False) else 'No',
            row.get('vm_costo_limpieza', 0),
            'Sí' if row.get('vm_incluye_reinstalacion', False) else 'No',
            row.get('vm_costo_reinstalacion', 0),
            'Sí' if row.get('vm_incluye_respaldo', False) else 'No',
            row.get('vm_costo_respaldo', 0),
            'Sí' if row.get('vm_incluye_cambio_pieza', False) else 'No',
            row.get('vm_costo_cambio_pieza', 0),
            'Sí' if row.get('vm_incluye_kit_limpieza', False) else 'No',
            row.get('vm_costo_kit', 0),
            row.get('vm_total_venta', 0),
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws5.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (7, 9, 11, 13, 15, 17, 18):
                cell.number_format = number_fmt
            # Colorear servicios activos
            if col in (8, 10, 12, 14, 16) and valor == 'Sí':
                cell.fill = teal_fill
    
    # Totales
    if len(df_con_vm) > 0:
        fila += 1
        ws5.cell(row=fila, column=1, value='TOTALES').font = Font(bold=True, size=12)
        ws5.cell(row=fila, column=7, value=float(df_con_vm['vm_costo_paquete'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=9, value=float(df_con_vm['vm_costo_limpieza'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=11, value=float(df_con_vm['vm_costo_reinstalacion'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=13, value=float(df_con_vm['vm_costo_respaldo'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=15, value=float(df_con_vm['vm_costo_cambio_pieza'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=17, value=float(df_con_vm['vm_costo_kit'].sum())).number_format = number_fmt
        ws5.cell(row=fila, column=18, value=float(df_con_vm['vm_total_venta'].sum())).number_format = number_fmt
        for col in range(1, len(headers_vm) + 1):
            ws5.cell(row=fila, column=col).font = Font(bold=True)
            ws5.cell(row=fila, column=col).border = border_thin
    
    auto_ajustar_columnas(ws5)
    
    # ========================================================================
    # HOJA 6: SERVICIOS ADICIONALES
    # ========================================================================
    ws6 = wb.create_sheet('Servicios Adicionales')
    
    ws6.merge_cells('A1:H1')
    ws6['A1'] = 'ANÁLISIS DE SERVICIOS ADICIONALES — VM EN ACEPTADAS vs. VM PERÍODO COMPLETO'
    ws6['A1'].font = title_font
    ws6['A1'].fill = title_fill
    ws6['A1'].alignment = title_align

    # Leyenda de columnas (fila 2)
    ws6.merge_cells('A2:H2')
    ws6['A2'] = (
        'En Aceptadas = VM cuya cotización fue aceptada  |  '
        'VM Período = Todas las VM del período (incluye órdenes FL sin cotización)  |  '
        'VM Únicas FL = Diferencia (ventas directas sin cotización previa)'
    )
    ws6['A2'].font = Font(italic=True, size=9, color='444444')
    ws6['A2'].fill = PatternFill(start_color='f0f4f8', end_color='f0f4f8', fill_type='solid')

    fila = 4
    if analisis_vm.get('tiene_datos'):
        orange_fill_light = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')

        # ----------------------------------------------------------------
        # SECCIÓN 1: Distribución de paquetes
        # ----------------------------------------------------------------
        ws6.merge_cells(f'A{fila}:H{fila}')
        ws6.cell(row=fila, column=1, value='DISTRIBUCIÓN DE PAQUETES').font = Font(bold=True, size=12)
        ws6.cell(row=fila, column=1).fill = blue_fill
        fila += 1
        headers_paq = [
            'Paquete',
            'En Aceptadas', '% Aceptadas', 'Ingreso (Aceptadas)',
            'VM Período', '% VM Período', 'Ingreso (VM Período)',
            'VM Únicas FL',
        ]
        for col, h in enumerate(headers_paq, 1):
            ws6.cell(row=fila, column=col, value=h)
        aplicar_estilos_header(ws6, fila, len(headers_paq))

        for paq in analisis_vm['distribucion_paquetes']:
            fila += 1
            ws6.cell(row=fila, column=1, value=paq['nombre']).border = border_thin
            ws6.cell(row=fila, column=2, value=paq['cantidad']).border = border_thin
            ws6.cell(row=fila, column=3, value=f"{paq['porcentaje']}%").border = border_thin
            cell_ia = ws6.cell(row=fila, column=4, value=paq['ingreso_total'])
            cell_ia.number_format = number_fmt
            cell_ia.border = border_thin
            ws6.cell(row=fila, column=5, value=paq['cantidad_vm_periodo']).border = border_thin
            ws6.cell(row=fila, column=6, value=f"{paq['porcentaje_vm_periodo']}%").border = border_thin
            cell_ir = ws6.cell(row=fila, column=7, value=paq['ingreso_vm_periodo'])
            cell_ir.number_format = number_fmt
            cell_ir.border = border_thin
            dif = paq['cantidad_vm_unicas']
            cell_dif = ws6.cell(row=fila, column=8, value=dif)
            cell_dif.border = border_thin
            if dif > 0:
                cell_dif.fill = orange_fill_light
                cell_dif.font = Font(bold=True, color='e65100')

        # ----------------------------------------------------------------
        # SECCIÓN 2: Distribución de servicios individuales
        # ----------------------------------------------------------------
        fila += 2
        ws6.merge_cells(f'A{fila}:H{fila}')
        ws6.cell(row=fila, column=1, value='DISTRIBUCIÓN DE SERVICIOS INDIVIDUALES').font = Font(bold=True, size=12)
        ws6.cell(row=fila, column=1).fill = teal_fill
        fila += 1
        headers_srv = [
            'Servicio',
            'En Aceptadas', '% En Aceptadas', 'Ingreso (Aceptadas)',
            'VM Período', '% VM Período', 'Ingreso (VM Período)',
            'VM Únicas FL',
        ]
        for col, h in enumerate(headers_srv, 1):
            ws6.cell(row=fila, column=col, value=h)
        aplicar_estilos_header(ws6, fila, len(headers_srv))

        for srv in analisis_vm['distribucion_servicios']:
            fila += 1
            ws6.cell(row=fila, column=1, value=srv['servicio']).border = border_thin
            ws6.cell(row=fila, column=2, value=srv['cantidad']).border = border_thin
            ws6.cell(row=fila, column=3, value=f"{srv['porcentaje']}%").border = border_thin
            cell_ia = ws6.cell(row=fila, column=4, value=srv['ingreso_total'])
            cell_ia.number_format = number_fmt
            cell_ia.border = border_thin
            ws6.cell(row=fila, column=5, value=srv['cantidad_vm_periodo']).border = border_thin
            ws6.cell(row=fila, column=6, value=f"{srv['porcentaje_vm_periodo']}%").border = border_thin
            cell_ir = ws6.cell(row=fila, column=7, value=srv['ingreso_vm_periodo'])
            cell_ir.number_format = number_fmt
            cell_ir.border = border_thin
            dif = srv['cantidad_vm_unicas']
            cell_dif = ws6.cell(row=fila, column=8, value=dif)
            cell_dif.border = border_thin
            if dif > 0:
                cell_dif.fill = orange_fill_light
                cell_dif.font = Font(bold=True, color='e65100')

        # ----------------------------------------------------------------
        # SECCIÓN 3: Combinaciones — tabla unificada (Aceptadas + FL)
        # ----------------------------------------------------------------
        fila += 2
        ws6.merge_cells(f'A{fila}:H{fila}')
        ws6.cell(row=fila, column=1,
                 value='COMBINACIONES DE SERVICIOS MÁS FRECUENTES').font = Font(bold=True, size=12)
        ws6.cell(row=fila, column=1).fill = purple_fill
        fila += 1
        # Nota
        ws6.merge_cells(f'A{fila}:H{fila}')
        ws6.cell(row=fila, column=1,
                 value='Incluye cotizaciones aceptadas + órdenes FL (ventas directas). '
                       'Origen "Solo FL" = combinación exclusiva de órdenes sin cotización.')
        ws6.cell(row=fila, column=1).font = Font(italic=True, size=9, color='666666')
        fila += 1
        headers_combo = [
            'Combinación', 'Origen',
            'VM Período', '% VM Período',
            'En Aceptadas', '% En Aceptadas',
            'VM Únicas FL',
        ]
        for col, h in enumerate(headers_combo, 1):
            ws6.cell(row=fila, column=col, value=h)
        aplicar_estilos_header(ws6, fila, len(headers_combo))

        fl_fill = PatternFill(start_color='ede9fe', end_color='ede9fe', fill_type='solid')
        fl_font_bold = Font(bold=True, color='5b21b6')
        for combo_t in analisis_vm.get('combinaciones_frecuentes_total', []):
            fila += 1
            es_fl = combo_t.get('exclusivo_fl', False)
            cant_acept = combo_t['cantidad_aceptadas']
            cant_total = combo_t['cantidad_total']
            cant_unicas = cant_total - cant_acept

            # Columna 1: Combinación
            cell_c = ws6.cell(row=fila, column=1, value=combo_t['combinacion'])
            cell_c.border = border_thin
            if es_fl:
                cell_c.fill = fl_fill
                cell_c.font = fl_font_bold

            # Columna 2: Origen
            origen_val = 'Solo FL' if es_fl else 'Aceptadas + FL' if cant_unicas > 0 else 'En Aceptadas'
            cell_o = ws6.cell(row=fila, column=2, value=origen_val)
            cell_o.border = border_thin
            if es_fl:
                cell_o.fill = fl_fill
                cell_o.font = fl_font_bold

            # Columna 3–4: VM Período
            cell_vp = ws6.cell(row=fila, column=3, value=cant_total)
            cell_vp.border = border_thin
            if es_fl:
                cell_vp.fill = fl_fill
            ws6.cell(row=fila, column=4,
                     value=f"{combo_t['porcentaje_total']}%").border = border_thin

            # Columna 5–6: En Aceptadas
            cell_a = ws6.cell(row=fila, column=5,
                              value=cant_acept if cant_acept > 0 else '—')
            cell_a.border = border_thin
            ws6.cell(row=fila, column=6,
                     value=f"{combo_t['porcentaje_aceptadas']}%" if cant_acept > 0 else '—').border = border_thin

            # Columna 7: VM Únicas FL
            cell_u = ws6.cell(row=fila, column=7,
                              value=cant_unicas if cant_unicas > 0 else '—')
            cell_u.border = border_thin
            if cant_unicas > 0:
                cell_u.fill = orange_fill_light
                cell_u.font = Font(bold=True, color='e65100')
        
        # Top piezas VM
        if analisis_vm.get('top_piezas_vm'):
            fila += 2
            ws6.merge_cells(f'A{fila}:H{fila}')
            ws6.cell(row=fila, column=1, value='TOP PIEZAS VENDIDAS EN VENTA MOSTRADOR').font = Font(bold=True, size=12)
            ws6.cell(row=fila, column=1).fill = blue_fill
            fila += 1
            headers_tpvm = ['Pieza', 'Cantidad', 'Ingreso Total', 'Núm. Ventas']
            for col, h in enumerate(headers_tpvm, 1):
                ws6.cell(row=fila, column=col, value=h)
            aplicar_estilos_header(ws6, fila, len(headers_tpvm))
            
            for pieza in analisis_vm['top_piezas_vm']:
                fila += 1
                ws6.cell(row=fila, column=1, value=pieza['descripcion']).border = border_thin
                ws6.cell(row=fila, column=2, value=pieza['cantidad']).border = border_thin
                cell_i = ws6.cell(row=fila, column=3, value=pieza['ingreso_total'])
                cell_i.number_format = number_fmt
                cell_i.border = border_thin
                ws6.cell(row=fila, column=4, value=pieza['num_ventas']).border = border_thin
    else:
        ws6.cell(row=fila, column=1, value='No hay ventas mostrador asociadas a cotizaciones aceptadas')
        ws6.cell(row=fila, column=1).font = Font(italic=True, color='999999', size=12)
    
    auto_ajustar_columnas(ws6)
    
    # ========================================================================
    # HOJA 7: SEGUIMIENTO DE PIEZAS
    # ========================================================================
    ws7 = wb.create_sheet('Seguimiento Piezas')
    
    ws7.merge_cells('A1:H1')
    ws7['A1'] = 'SEGUIMIENTO DE PIEZAS POST-ACEPTACIÓN'
    ws7['A1'].font = title_font
    ws7['A1'].fill = title_fill
    ws7['A1'].alignment = title_align
    
    fila = 3
    if analisis_seguimiento.get('tiene_datos'):
        # Distribución de estados
        ws7.cell(row=fila, column=1, value='DISTRIBUCIÓN DE ESTADOS').font = Font(bold=True, size=12)
        ws7.cell(row=fila, column=1).fill = green_fill
        fila += 1
        headers_est = ['Estado', 'Cantidad', 'Porcentaje', 'Tipo']
        for col, h in enumerate(headers_est, 1):
            ws7.cell(row=fila, column=col, value=h)
        aplicar_estilos_header(ws7, fila, len(headers_est))
        
        for est in analisis_seguimiento['distribucion_estados']:
            fila += 1
            ws7.cell(row=fila, column=1, value=est['label']).border = border_thin
            ws7.cell(row=fila, column=2, value=est['cantidad']).border = border_thin
            ws7.cell(row=fila, column=3, value=f"{est['porcentaje']}%").border = border_thin
            tipo = 'Problemático' if est['es_problematico'] else ('Recibido' if est['es_recibido'] else 'En proceso')
            cell_tipo = ws7.cell(row=fila, column=4, value=tipo)
            cell_tipo.border = border_thin
            if est['es_problematico']:
                cell_tipo.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
            elif est['es_recibido']:
                cell_tipo.fill = green_fill
            else:
                cell_tipo.fill = yellow_fill
        
        # Tiempos de entrega
        tiempos = analisis_seguimiento.get('tiempos_entrega', {})
        if tiempos.get('total_con_fecha', 0) > 0:
            fila += 2
            ws7.cell(row=fila, column=1, value='TIEMPOS DE ENTREGA').font = Font(bold=True, size=12)
            ws7.cell(row=fila, column=1).fill = blue_fill
            fila += 1
            metricas_tiempo = [
                ('Promedio (días)', tiempos.get('promedio', 0)),
                ('Mediana (días)', tiempos.get('mediana', 0)),
                ('Mínimo (días)', tiempos.get('minimo', 0)),
                ('Máximo (días)', tiempos.get('maximo', 0)),
                ('Con fecha de entrega', tiempos.get('total_con_fecha', 0)),
                ('Sin fecha aún', tiempos.get('total_sin_fecha', 0)),
                ('Tasa de cumplimiento', f"{analisis_seguimiento.get('tasa_cumplimiento', 0)}%"),
            ]
            for label, valor in metricas_tiempo:
                fila += 1
                ws7.cell(row=fila, column=1, value=label).font = kpi_label_font
                ws7.cell(row=fila, column=1).border = border_thin
                ws7.cell(row=fila, column=2, value=valor).border = border_thin
        
        # Ranking de proveedores
        proveedores = analisis_seguimiento.get('proveedores_ranking', [])
        if proveedores:
            fila += 2
            ws7.cell(row=fila, column=1, value='RANKING DE PROVEEDORES').font = Font(bold=True, size=12)
            ws7.cell(row=fila, column=1).fill = teal_fill
            fila += 1
            headers_prov = ['Proveedor', 'Total Pedidos', 'Recibidos', 'Problemas', 'Tasa Éxito', 'Tiempo Prom.']
            for col, h in enumerate(headers_prov, 1):
                ws7.cell(row=fila, column=col, value=h)
            aplicar_estilos_header(ws7, fila, len(headers_prov))
            
            for prov in proveedores:
                fila += 1
                ws7.cell(row=fila, column=1, value=prov['proveedor']).border = border_thin
                ws7.cell(row=fila, column=2, value=prov['total_pedidos']).border = border_thin
                ws7.cell(row=fila, column=3, value=prov['recibidos']).border = border_thin
                ws7.cell(row=fila, column=4, value=prov['problemas']).border = border_thin
                cell_tasa = ws7.cell(row=fila, column=5, value=f"{prov['tasa_exito']}%")
                cell_tasa.border = border_thin
                if prov['tasa_exito'] >= 90:
                    cell_tasa.fill = green_fill
                elif prov['tasa_exito'] < 70:
                    cell_tasa.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                tiempo_str = f"{prov['tiempo_promedio']} días" if prov['tiempo_promedio'] else 'Sin datos'
                ws7.cell(row=fila, column=6, value=tiempo_str).border = border_thin
        
        # Problemas
        problemas = analisis_seguimiento.get('problemas_piezas', {})
        if problemas.get('total_problemas', 0) > 0:
            fila += 2
            ws7.cell(row=fila, column=1, value='PROBLEMAS EN PIEZAS').font = Font(bold=True, size=12)
            ws7.cell(row=fila, column=1).fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
            fila += 1
            ws7.cell(row=fila, column=1, value='Piezas incorrectas (WPB):').font = kpi_label_font
            ws7.cell(row=fila, column=2, value=problemas.get('piezas_incorrectas', 0))
            fila += 1
            ws7.cell(row=fila, column=1, value='Piezas dañadas (DOA):').font = kpi_label_font
            ws7.cell(row=fila, column=2, value=problemas.get('piezas_danadas', 0))
            fila += 1
            ws7.cell(row=fila, column=1, value='Tasa de problemas:').font = kpi_label_font
            ws7.cell(row=fila, column=2, value=f"{problemas.get('tasa_problemas', 0)}%")
    else:
        ws7.cell(row=fila, column=1, value='No hay seguimientos de piezas en las cotizaciones aceptadas')
        ws7.cell(row=fila, column=1).font = Font(italic=True, color='999999', size=12)
    
    auto_ajustar_columnas(ws7)
    
    # ========================================================================
    # HOJA 8: RENDIMIENTO POR TÉCNICO
    # ========================================================================
    ws8 = wb.create_sheet('Por Técnico')
    
    ws8.merge_cells('A1:K1')
    ws8['A1'] = 'RENDIMIENTO DE ACEPTACIONES POR TÉCNICO'
    ws8['A1'].font = title_font
    ws8['A1'].fill = title_fill
    ws8['A1'].alignment = title_align
    
    # Calcular métricas por técnico
    tec_metrics = df_aceptadas.groupby('tecnico').agg(
        total_aceptadas=('aceptada', 'count'),
        valor_aceptado=('costo_total_final', 'sum'),
        ticket_promedio=('costo_total_final', 'mean'),
        piezas_promedio=('piezas_aceptadas', 'mean'),
        con_vm=('tiene_venta_mostrador', 'sum'),
        valor_vm=('vm_total_venta', 'sum'),
        valor_combinado=('valor_total_combinado', 'sum'),
        con_descuento=('descontar_mano_obra', 'sum'),
        tiempo_resp=('dias_sin_respuesta', 'mean'),
    ).reset_index()
    tec_metrics['tasa_upsell'] = (tec_metrics['con_vm'] / tec_metrics['total_aceptadas'] * 100).round(1)
    tec_metrics = tec_metrics.sort_values('valor_combinado', ascending=False)
    
    headers_tec = [
        'Técnico', 'Aceptadas', 'Valor Aceptado', 'Ticket Promedio',
        'Piezas Prom.', 'Con VM', 'Tasa Upsell (%)', 'Valor VM',
        'Valor Combinado', 'Con Descuento MO', 'Tiempo Resp. Prom.',
    ]
    
    fila = 3
    for col, h in enumerate(headers_tec, 1):
        ws8.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws8, fila, len(headers_tec))
    
    for _, row in tec_metrics.iterrows():
        fila += 1
        datos = [
            row['tecnico'],
            int(row['total_aceptadas']),
            round(row['valor_aceptado'], 2),
            round(row['ticket_promedio'], 2),
            round(row['piezas_promedio'], 1),
            int(row['con_vm']),
            row['tasa_upsell'],
            round(row['valor_vm'], 2),
            round(row['valor_combinado'], 2),
            int(row['con_descuento']),
            round(row['tiempo_resp'], 1) if not pd.isna(row['tiempo_resp']) else 0,
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws8.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (3, 4, 8, 9):
                cell.number_format = number_fmt
    
    auto_ajustar_columnas(ws8)
    
    # ========================================================================
    # HOJA 9: RENDIMIENTO POR SUCURSAL
    # ========================================================================
    ws9 = wb.create_sheet('Por Sucursal')
    
    ws9.merge_cells('A1:K1')
    ws9['A1'] = 'RENDIMIENTO DE ACEPTACIONES POR SUCURSAL'
    ws9['A1'].font = title_font
    ws9['A1'].fill = title_fill
    ws9['A1'].alignment = title_align
    
    # Calcular métricas por sucursal
    suc_metrics = df_aceptadas.groupby('sucursal').agg(
        total_aceptadas=('aceptada', 'count'),
        valor_aceptado=('costo_total_final', 'sum'),
        ticket_promedio=('costo_total_final', 'mean'),
        piezas_promedio=('piezas_aceptadas', 'mean'),
        con_vm=('tiene_venta_mostrador', 'sum'),
        valor_vm=('vm_total_venta', 'sum'),
        valor_combinado=('valor_total_combinado', 'sum'),
        con_descuento=('descontar_mano_obra', 'sum'),
        tiempo_resp=('dias_sin_respuesta', 'mean'),
    ).reset_index()
    suc_metrics['tasa_upsell'] = (suc_metrics['con_vm'] / suc_metrics['total_aceptadas'] * 100).round(1)
    suc_metrics = suc_metrics.sort_values('valor_combinado', ascending=False)
    
    # Agregar tasa de aceptación global por sucursal
    tasa_por_suc = df_cotizaciones.groupby('sucursal').agg(
        total_cotizaciones=('aceptada', 'count'),
        total_aceptadas_global=('aceptada', lambda x: (x == True).sum()),
    ).reset_index()
    tasa_por_suc['tasa_aceptacion'] = (tasa_por_suc['total_aceptadas_global'] / tasa_por_suc['total_cotizaciones'] * 100).round(1)
    suc_metrics = suc_metrics.merge(tasa_por_suc[['sucursal', 'tasa_aceptacion', 'total_cotizaciones']], on='sucursal', how='left')
    
    headers_suc = [
        'Sucursal', 'Total Cotizaciones', 'Aceptadas', 'Tasa Aceptación (%)',
        'Valor Aceptado', 'Ticket Promedio',
        'Con VM', 'Tasa Upsell (%)', 'Valor VM',
        'Valor Combinado', 'Tiempo Resp. Prom.',
    ]
    
    fila = 3
    for col, h in enumerate(headers_suc, 1):
        ws9.cell(row=fila, column=col, value=h)
    aplicar_estilos_header(ws9, fila, len(headers_suc))
    
    for _, row in suc_metrics.iterrows():
        fila += 1
        datos = [
            row['sucursal'],
            int(row.get('total_cotizaciones', 0)),
            int(row['total_aceptadas']),
            row.get('tasa_aceptacion', 0),
            round(row['valor_aceptado'], 2),
            round(row['ticket_promedio'], 2),
            int(row['con_vm']),
            row['tasa_upsell'],
            round(row['valor_vm'], 2),
            round(row['valor_combinado'], 2),
            round(row['tiempo_resp'], 1) if not pd.isna(row['tiempo_resp']) else 0,
        ]
        for col, valor in enumerate(datos, 1):
            cell = ws9.cell(row=fila, column=col, value=valor)
            cell.border = border_thin
            if col in (5, 6, 9, 10):
                cell.number_format = number_fmt
            # Colorear tasa de aceptación
            if col == 4:
                if valor >= 60:
                    cell.fill = green_fill
                elif valor < 40:
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                else:
                    cell.fill = yellow_fill
    
    auto_ajustar_columnas(ws9)
    
    # ========================================
    # GUARDAR Y RETORNAR RESPUESTA
    # ========================================
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'Analisis_Aceptaciones_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    
    return response

