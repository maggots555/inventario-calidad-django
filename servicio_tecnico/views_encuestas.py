"""
Dashboard y APIs de encuestas de satisfacción + sentimiento IA (Fase 4).

urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
"""

import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from inventario.models import Empleado, Sucursal

from .decorators import permission_required_with_message

logger = logging.getLogger(__name__)


# ============================================================================
# DASHBOARD DE ENCUESTAS DE SATISFACCIÓN (Marzo 2026)
# Panel analítico para visualizar encuestas enviadas, respondidas, pendientes
# y expiradas. Incluye KPIs, gráficos Chart.js y análisis por responsable.
# ============================================================================


def _filtrar_encuestas_satisfaccion(request):
    """
    Helper: construye queryset base de FeedbackCliente tipo 'satisfaccion'
    aplicando los filtros GET comunes (fecha, responsable, sucursal, tipo_orden).
    Retorna el queryset con annotate de fecha_expiracion.
    """
    from .models import FeedbackCliente
    from django.db.models import F, ExpressionWrapper, DateTimeField
    from datetime import timedelta

    qs = FeedbackCliente.objects.filter(tipo='satisfaccion').select_related(
        'orden__responsable_seguimiento',
        'orden__sucursal',
        'orden__detalle_equipo',
    )

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    responsable_id = request.GET.get('responsable_id')
    sucursal_id = request.GET.get('sucursal_id')
    tipo_orden = request.GET.get('tipo_orden')

    if fecha_desde:
        qs = qs.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_creacion__date__lte=fecha_hasta)
    if responsable_id:
        qs = qs.filter(orden__responsable_seguimiento_id=responsable_id)
    if sucursal_id:
        qs = qs.filter(orden__sucursal_id=sucursal_id)
    if tipo_orden and tipo_orden in ('diagnostico', 'venta_mostrador'):
        qs = qs.filter(orden__tipo_servicio=tipo_orden)

    qs = qs.annotate(
        fecha_expiracion=ExpressionWrapper(
            F('fecha_creacion') + timedelta(days=7),
            output_field=DateTimeField()
        )
    )
    return qs




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def dashboard_encuestas(request):
    """
    Vista principal del panel de encuestas de satisfacción.
    Renderiza el template con filtros; la data se carga vía AJAX.
    """
    from django.conf import settings as django_settings

    empleados  = Empleado.objects.filter(activo=True).order_by('nombre_completo')
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')

    return render(request, 'servicio_tecnico/dashboard_encuestas.html', {
        'empleados':   empleados,
        'sucursales':  sucursales,
        'ai_enabled':  getattr(django_settings, 'AI_ENABLED', False),
        'ai_models':   getattr(django_settings, 'AI_MODELS', []),
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_kpis(request):
    """
    API JSON: KPIs globales del dashboard de encuestas.
    """
    from django.db.models import Avg

    now = timezone.now()
    qs = _filtrar_encuestas_satisfaccion(request)

    total_enviadas = qs.filter(correo_enviado=True).count()
    total_respondidas = qs.filter(utilizado=True).count()
    total_pendientes = qs.filter(
        utilizado=False, correo_enviado=True, fecha_expiracion__gte=now
    ).count()
    total_expiradas = qs.filter(
        utilizado=False, fecha_expiracion__lt=now
    ).count()

    tasa_respuesta = round(
        (total_respondidas / total_enviadas * 100) if total_enviadas > 0 else 0, 1
    )

    respondidas_qs = qs.filter(utilizado=True)
    avgs = respondidas_qs.aggregate(
        nps_promedio=Avg('nps'),
        calificacion_promedio=Avg('calificacion_general'),
        calificacion_atencion_promedio=Avg('calificacion_atencion'),
        calificacion_tiempo_promedio=Avg('calificacion_tiempo'),
    )

    total_con_recomendacion = respondidas_qs.filter(recomienda__isnull=False).count()
    total_recomiendan = respondidas_qs.filter(recomienda=True).count()
    tasa_recomendacion = round(
        (total_recomiendan / total_con_recomendacion * 100) if total_con_recomendacion > 0 else 0, 1
    )

    # NPS Score = % promotores (9-10) - % detractores (0-6)
    respondidas_con_nps = respondidas_qs.filter(nps__isnull=False).count()
    promotores = respondidas_qs.filter(nps__gte=9).count()
    detractores = respondidas_qs.filter(nps__lte=6).count()
    nps_score = round(
        ((promotores - detractores) / respondidas_con_nps * 100) if respondidas_con_nps > 0 else 0, 1
    )

    return JsonResponse({
        'total_enviadas': total_enviadas,
        'total_respondidas': total_respondidas,
        'total_pendientes': total_pendientes,
        'total_expiradas': total_expiradas,
        'tasa_respuesta': tasa_respuesta,
        'nps_promedio': round(avgs['nps_promedio'] or 0, 1),
        'calificacion_promedio': round(avgs['calificacion_promedio'] or 0, 1),
        'calificacion_atencion_promedio': round(avgs['calificacion_atencion_promedio'] or 0, 1),
        'calificacion_tiempo_promedio': round(avgs['calificacion_tiempo_promedio'] or 0, 1),
        'tasa_recomendacion': tasa_recomendacion,
        'nps_score': nps_score,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_tendencia(request):
    """
    API JSON: tendencia temporal de métricas (agrupado por semana).
    """
    from django.db.models import Avg
    from django.db.models.functions import TruncWeek

    qs = _filtrar_encuestas_satisfaccion(request).filter(correo_enviado=True)

    datos_por_semana = (
        qs.annotate(semana=TruncWeek('fecha_creacion'))
        .values('semana')
        .annotate(
            total_enviadas=Count('id'),
            total_respondidas=Count('id', filter=Q(utilizado=True)),
            calificacion_promedio=Avg('calificacion_general', filter=Q(utilizado=True)),
            nps_promedio=Avg('nps', filter=Q(utilizado=True)),
        )
        .order_by('semana')
    )

    labels = []
    datasets = {
        'calificacion_promedio': [],
        'nps_promedio': [],
        'tasa_respuesta': [],
        'total_enviadas': [],
        'total_respondidas': [],
    }

    for row in datos_por_semana:
        semana = row['semana']
        labels.append(semana.strftime('%d/%m/%Y'))
        datasets['total_enviadas'].append(row['total_enviadas'])
        datasets['total_respondidas'].append(row['total_respondidas'])
        datasets['calificacion_promedio'].append(
            round(row['calificacion_promedio'] or 0, 1)
        )
        datasets['nps_promedio'].append(
            round(row['nps_promedio'] or 0, 1)
        )
        tasa = round(
            (row['total_respondidas'] / row['total_enviadas'] * 100)
            if row['total_enviadas'] > 0 else 0, 1
        )
        datasets['tasa_respuesta'].append(tasa)

    return JsonResponse({'labels': labels, 'datasets': datasets})




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_por_responsable(request):
    """
    API JSON: métricas agrupadas por responsable de seguimiento.
    """
    from django.db.models import Avg

    qs = _filtrar_encuestas_satisfaccion(request).filter(correo_enviado=True)

    datos = (
        qs.values(
            'orden__responsable_seguimiento__id',
            'orden__responsable_seguimiento__nombre_completo',
        )
        .annotate(
            total_enviadas=Count('id'),
            total_respondidas=Count('id', filter=Q(utilizado=True)),
            calificacion_promedio=Avg('calificacion_general', filter=Q(utilizado=True)),
            nps_promedio=Avg('nps', filter=Q(utilizado=True)),
            total_recomiendan=Count('id', filter=Q(utilizado=True, recomienda=True)),
            total_con_recomendacion=Count('id', filter=Q(utilizado=True, recomienda__isnull=False)),
            promotores=Count('id', filter=Q(utilizado=True, nps__gte=9)),
            detractores=Count('id', filter=Q(utilizado=True, nps__lte=6)),
            respondidas_con_nps=Count('id', filter=Q(utilizado=True, nps__isnull=False)),
        )
        .order_by('-calificacion_promedio')
    )

    responsables = []
    for row in datos:
        nombre = row['orden__responsable_seguimiento__nombre_completo'] or ''
        tasa_rec = round(
            (row['total_recomiendan'] / row['total_con_recomendacion'] * 100)
            if row['total_con_recomendacion'] > 0 else 0, 1
        )
        nps_s = round(
            ((row['promotores'] - row['detractores']) / row['respondidas_con_nps'] * 100)
            if row['respondidas_con_nps'] > 0 else 0, 1
        )
        responsables.append({
            'id': row['orden__responsable_seguimiento__id'],
            'nombre': nombre,
            'total_enviadas': row['total_enviadas'],
            'total_respondidas': row['total_respondidas'],
            'calificacion_promedio': round(row['calificacion_promedio'] or 0, 1),
            'nps_promedio': round(row['nps_promedio'] or 0, 1),
            'tasa_recomendacion': tasa_rec,
            'nps_score': nps_s,
        })

    return JsonResponse({'responsables': responsables})




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_distribucion_nps(request):
    """
    API JSON: distribución NPS (Promotores 9-10 / Pasivos 7-8 / Detractores 0-6).
    """
    qs = _filtrar_encuestas_satisfaccion(request).filter(
        utilizado=True, nps__isnull=False
    )

    datos = qs.aggregate(
        promotores=Count('id', filter=Q(nps__gte=9)),
        pasivos=Count('id', filter=Q(nps__gte=7, nps__lte=8)),
        detractores=Count('id', filter=Q(nps__lte=6)),
        total=Count('id'),
    )

    total = datos['total'] or 0
    nps_score = round(
        ((datos['promotores'] - datos['detractores']) / total * 100)
        if total > 0 else 0, 1
    )

    return JsonResponse({
        'promotores': datos['promotores'],
        'pasivos': datos['pasivos'],
        'detractores': datos['detractores'],
        'total': total,
        'nps_score': nps_score,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_lista(request):
    """
    API JSON: lista paginada de encuestas con búsqueda y filtro por estado.
    """
    from config.paises_config import fecha_local_pais, get_pais_actual
    pais = get_pais_actual()
    now = timezone.now()
    qs = _filtrar_encuestas_satisfaccion(request)

    # Filtro por estado (tab)
    estado = request.GET.get('estado', 'todas')
    if estado == 'respondidas':
        qs = qs.filter(utilizado=True)
    elif estado == 'pendientes':
        qs = qs.filter(utilizado=False, correo_enviado=True, fecha_expiracion__gte=now)
    elif estado == 'expiradas':
        qs = qs.filter(utilizado=False, fecha_expiracion__lt=now)

    # Búsqueda
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        qs = qs.filter(
            Q(orden__numero_orden_interno__icontains=busqueda) |
            Q(orden__detalle_equipo__orden_cliente__icontains=busqueda) |
            Q(orden__detalle_equipo__email_cliente__icontains=busqueda) |
            Q(orden__detalle_equipo__marca__icontains=busqueda) |
            Q(orden__detalle_equipo__modelo__icontains=busqueda)
        )

    qs = qs.order_by('-fecha_creacion')

    # Paginación
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 15))
    paginator = Paginator(qs, page_size)

    try:
        pagina = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        pagina = paginator.page(1)

    encuestas = []
    for fb in pagina.object_list:
        orden = fb.orden
        detalle = getattr(orden, 'detalle_equipo', None)

        if fb.utilizado:
            estado_encuesta = 'respondida'
        elif not fb.correo_enviado:
            estado_encuesta = 'no_enviada'
        elif fb.fecha_expiracion < now:
            estado_encuesta = 'expirada'
        else:
            estado_encuesta = 'pendiente'

        encuestas.append({
            'id': fb.id,
            'orden_numero': detalle.orden_cliente if detalle and detalle.orden_cliente else orden.numero_orden_interno,
            'orden_id': orden.id,
            'equipo': f"{detalle.marca} {detalle.get_tipo_equipo_display()} {detalle.modelo}" if detalle else '',
            'email_cliente': detalle.email_cliente if detalle else '',
            'responsable': str(orden.responsable_seguimiento) if orden.responsable_seguimiento else '',
            'sucursal': str(orden.sucursal) if orden.sucursal else '',
            'tipo_orden': orden.get_tipo_servicio_display(),
            'fecha_envio': fecha_local_pais(fb.fecha_creacion, pais).strftime('%d/%m/%Y %H:%M'),
            'fecha_respuesta': fecha_local_pais(fb.fecha_respuesta, pais).strftime('%d/%m/%Y %H:%M') if fb.fecha_respuesta else None,
            'dias_restantes': fb.dias_restantes,
            'estado': estado_encuesta,
            'calificacion_general': fb.calificacion_general,
            'nps': fb.nps,
            'recomienda': fb.recomienda,
            'calificacion_atencion': fb.calificacion_atencion,
            'calificacion_tiempo': fb.calificacion_tiempo,
            'comentario_cliente': fb.comentario_cliente,
        })

    return JsonResponse({
        'encuestas': encuestas,
        'total': paginator.count,
        'paginas': paginator.num_pages,
        'pagina_actual': pagina.number,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_encuestas_comentarios(request):
    """
    API JSON: últimos comentarios de clientes con calificación.
    """
    from config.paises_config import fecha_local_pais, get_pais_actual
    pais = get_pais_actual()
    qs = _filtrar_encuestas_satisfaccion(request).filter(
        utilizado=True,
    ).exclude(
        comentario_cliente=''
    ).order_by('-fecha_respuesta')[:10]

    comentarios = []
    for fb in qs:
        comentarios.append({
            'orden_numero': (
                fb.orden.detalle_equipo.orden_cliente
                if hasattr(fb.orden, 'detalle_equipo') and fb.orden.detalle_equipo.orden_cliente
                else fb.orden.numero_orden_interno
            ),
            'orden_id': fb.orden.id,
            'responsable': str(fb.orden.responsable_seguimiento) if fb.orden.responsable_seguimiento else '',
            'calificacion': fb.calificacion_general,
            'nps': fb.nps,
            'recomienda': fb.recomienda,
            'comentario': fb.comentario_cliente,
            'fecha': fecha_local_pais(fb.fecha_respuesta, pais).strftime('%d/%m/%Y') if fb.fecha_respuesta else '',
        })

    return JsonResponse({'comentarios': comentarios})




@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def exportar_encuestas_excel(request):
    """
    Exporta las encuestas de satisfacción filtradas a un archivo Excel.
    Genera 3 hojas: Resumen KPIs, Encuestas detalladas, Por Responsable.
    """
    from openpyxl import Workbook
    from django.db.models import Avg

    now = timezone.now()
    qs = _filtrar_encuestas_satisfaccion(request).filter(correo_enviado=True)

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()

    # ── HOJA 1: Resumen KPIs ──────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen KPIs'

    total_enviadas = qs.count()
    total_respondidas = qs.filter(utilizado=True).count()
    total_pendientes = qs.filter(utilizado=False, fecha_expiracion__gte=now).count()
    total_expiradas = qs.filter(utilizado=False, fecha_expiracion__lt=now).count()
    tasa_respuesta = round((total_respondidas / total_enviadas * 100) if total_enviadas > 0 else 0, 1)

    respondidas_qs = qs.filter(utilizado=True)
    avgs = respondidas_qs.aggregate(
        cal_prom=Avg('calificacion_general'),
        nps_prom=Avg('nps'),
        cal_atencion=Avg('calificacion_atencion'),
        cal_tiempo=Avg('calificacion_tiempo'),
    )

    kpis = [
        ('Métrica', 'Valor'),
        ('Total Encuestas Enviadas', total_enviadas),
        ('Total Respondidas', total_respondidas),
        ('Total Pendientes', total_pendientes),
        ('Total Expiradas', total_expiradas),
        ('Tasa de Respuesta (%)', f'{tasa_respuesta}%'),
        ('Calificación General Promedio', round(avgs['cal_prom'] or 0, 2)),
        ('NPS Promedio', round(avgs['nps_prom'] or 0, 2)),
        ('Calificación Atención Promedio', round(avgs['cal_atencion'] or 0, 2)),
        ('Calificación Tiempo Promedio', round(avgs['cal_tiempo'] or 0, 2)),
    ]

    for row_idx, (metrica, valor) in enumerate(kpis, 1):
        ws_resumen.cell(row=row_idx, column=1, value=metrica)
        ws_resumen.cell(row=row_idx, column=2, value=valor)
        if row_idx == 1:
            for col in (1, 2):
                cell = ws_resumen.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        else:
            for col in (1, 2):
                ws_resumen.cell(row=row_idx, column=col).border = thin_border

    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 20

    # ── HOJA 2: Encuestas Detalladas ──────────────────────────────────
    ws_encuestas = wb.create_sheet('Encuestas')
    headers = [
        'Orden', 'Equipo', 'Email Cliente', 'Responsable', 'Sucursal',
        'Tipo Orden', 'Fecha Envío', 'Fecha Respuesta', 'Estado',
        'Calificación General', 'NPS', 'Recomienda',
        'Calificación Atención', 'Calificación Tiempo', 'Comentario'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_encuestas.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, fb in enumerate(qs.order_by('-fecha_creacion'), 2):
        orden = fb.orden
        detalle = getattr(orden, 'detalle_equipo', None)

        if fb.utilizado:
            estado_str = 'Respondida'
        elif fb.fecha_expiracion < now:
            estado_str = 'Expirada'
        else:
            estado_str = 'Pendiente'

        valores = [
            detalle.orden_cliente if detalle and detalle.orden_cliente else orden.numero_orden_interno,
            f"{detalle.marca} {detalle.get_tipo_equipo_display()} {detalle.modelo}" if detalle else '',
            detalle.email_cliente if detalle else '',
            str(orden.responsable_seguimiento) if orden.responsable_seguimiento else '',
            str(orden.sucursal) if orden.sucursal else '',
            orden.get_tipo_servicio_display(),
            fb.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            fb.fecha_respuesta.strftime('%d/%m/%Y %H:%M') if fb.fecha_respuesta else '',
            estado_str,
            fb.calificacion_general or '',
            fb.nps or '',
            'Sí' if fb.recomienda is True else ('No' if fb.recomienda is False else ''),
            fb.calificacion_atencion or '',
            fb.calificacion_tiempo or '',
            fb.comentario_cliente,
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws_encuestas.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws_encuestas.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── HOJA 3: Por Responsable ───────────────────────────────────────
    ws_resp = wb.create_sheet('Por Responsable')
    headers_resp = [
        'Responsable', 'Enviadas', 'Respondidas', 'Tasa Respuesta (%)',
        'Calificación Promedio', 'NPS Promedio', 'NPS Score', 'Tasa Recomendación (%)'
    ]
    for col_idx, header in enumerate(headers_resp, 1):
        cell = ws_resp.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    datos_resp = (
        qs.values(
            'orden__responsable_seguimiento__nombre_completo',
        )
        .annotate(
            total_enviadas=Count('id'),
            total_respondidas=Count('id', filter=Q(utilizado=True)),
            calificacion_promedio=Avg('calificacion_general', filter=Q(utilizado=True)),
            nps_promedio=Avg('nps', filter=Q(utilizado=True)),
            promotores=Count('id', filter=Q(utilizado=True, nps__gte=9)),
            detractores=Count('id', filter=Q(utilizado=True, nps__lte=6)),
            resp_con_nps=Count('id', filter=Q(utilizado=True, nps__isnull=False)),
            total_recomiendan=Count('id', filter=Q(utilizado=True, recomienda=True)),
            total_con_rec=Count('id', filter=Q(utilizado=True, recomienda__isnull=False)),
        )
        .order_by('-calificacion_promedio')
    )

    for row_idx, row in enumerate(datos_resp, 2):
        nombre = row['orden__responsable_seguimiento__nombre_completo'] or ''
        tasa_r = round((row['total_respondidas'] / row['total_enviadas'] * 100) if row['total_enviadas'] > 0 else 0, 1)
        nps_s = round(((row['promotores'] - row['detractores']) / row['resp_con_nps'] * 100) if row['resp_con_nps'] > 0 else 0, 1)
        tasa_rec = round((row['total_recomiendan'] / row['total_con_rec'] * 100) if row['total_con_rec'] > 0 else 0, 1)

        valores = [
            nombre,
            row['total_enviadas'],
            row['total_respondidas'],
            f'{tasa_r}%',
            round(row['calificacion_promedio'] or 0, 2),
            round(row['nps_promedio'] or 0, 2),
            nps_s,
            f'{tasa_rec}%',
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws_resp.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border

    for col_idx in range(1, len(headers_resp) + 1):
        ws_resp.column_dimensions[get_column_letter(col_idx)].width = 22

    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_str = now.strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename=Encuestas_Satisfaccion_{fecha_str}.xlsx'
    wb.save(response)
    return response




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def exportar_encuestas_pdf(request):
    """
    Genera y descarga el Reporte Ejecutivo PDF del Panel de Encuestas.

    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista calcula todos los datos necesarios para el reporte ejecutivo
    (KPIs, tendencia, distribución NPS, ranking por responsable y comentarios),
    los empaqueta en un diccionario y llama al módulo pdf_encuestas.py para
    generar el PDF con ReportLab + matplotlib.

    Comportamiento de comentarios:
      - Si hay filtros activos (fecha, responsable, sucursal o tipo_orden):
        se incluyen TODOS los comentarios del período filtrado.
      - Si no hay filtros aplicados:
        se incluyen solo los últimos 10 comentarios.

    Los mismos parámetros GET que usa el dashboard (fecha_desde, fecha_hasta,
    responsable_id, sucursal_id, tipo_orden) se aplican aquí para mantener
    coherencia entre lo que ve el usuario y lo que descarga en PDF.
    """
    from django.db.models import Avg, Count, Q
    from django.db.models.functions import TruncWeek
    from .pdf_encuestas import generar_pdf_reporte_encuestas
    from config.paises_config import fecha_local_pais, get_pais_actual

    pais = get_pais_actual()
    now  = timezone.now()

    # ---- 1. Queryset base con filtros ----
    qs = _filtrar_encuestas_satisfaccion(request)

    # Detectar si hay filtros activos para decidir el límite de comentarios
    fecha_desde    = request.GET.get('fecha_desde', '').strip()
    fecha_hasta    = request.GET.get('fecha_hasta', '').strip()
    responsable_id = request.GET.get('responsable_id', '').strip()
    sucursal_id    = request.GET.get('sucursal_id', '').strip()
    tipo_orden     = request.GET.get('tipo_orden', '').strip()

    hay_filtros = any([fecha_desde, fecha_hasta, responsable_id, sucursal_id, tipo_orden])

    # ---- 2. KPIs globales ----
    total_enviadas    = qs.filter(correo_enviado=True).count()
    total_respondidas = qs.filter(utilizado=True).count()
    total_pendientes  = qs.filter(utilizado=False, correo_enviado=True,
                                  fecha_expiracion__gte=now).count()
    total_expiradas   = qs.filter(utilizado=False,
                                  fecha_expiracion__lt=now).count()

    tasa_respuesta = round(
        (total_respondidas / total_enviadas * 100) if total_enviadas > 0 else 0, 1
    )

    respondidas_qs = qs.filter(utilizado=True)
    avgs = respondidas_qs.aggregate(
        nps_promedio=Avg('nps'),
        calificacion_promedio=Avg('calificacion_general'),
        calificacion_atencion_promedio=Avg('calificacion_atencion'),
        calificacion_tiempo_promedio=Avg('calificacion_tiempo'),
    )

    total_con_recomendacion = respondidas_qs.filter(recomienda__isnull=False).count()
    total_recomiendan       = respondidas_qs.filter(recomienda=True).count()
    tasa_recomendacion = round(
        (total_recomiendan / total_con_recomendacion * 100)
        if total_con_recomendacion > 0 else 0, 1
    )

    # NPS Score = % Promotores (9-10) − % Detractores (0-6)
    respondidas_con_nps = respondidas_qs.filter(nps__isnull=False).count()
    promotores_kpi  = respondidas_qs.filter(nps__gte=9).count()
    detractores_kpi = respondidas_qs.filter(nps__lte=6).count()
    nps_score_kpi = round(
        ((promotores_kpi - detractores_kpi) / respondidas_con_nps * 100)
        if respondidas_con_nps > 0 else 0, 1
    )

    kpis = {
        'total_enviadas': total_enviadas,
        'total_respondidas': total_respondidas,
        'total_pendientes': total_pendientes,
        'total_expiradas': total_expiradas,
        'tasa_respuesta': tasa_respuesta,
        'nps_promedio': round(avgs['nps_promedio'] or 0, 1),
        'calificacion_promedio': round(avgs['calificacion_promedio'] or 0, 1),
        'calificacion_atencion_promedio': round(avgs['calificacion_atencion_promedio'] or 0, 1),
        'calificacion_tiempo_promedio': round(avgs['calificacion_tiempo_promedio'] or 0, 1),
        'tasa_recomendacion': tasa_recomendacion,
        'nps_score': nps_score_kpi,
    }

    # ---- 3. Tendencia semanal ----
    datos_tendencia = (
        qs.filter(correo_enviado=True)
        .annotate(semana=TruncWeek('fecha_creacion'))
        .values('semana')
        .annotate(
            total_enviadas_s=Count('id'),
            total_respondidas_s=Count('id', filter=Q(utilizado=True)),
            calificacion_promedio_s=Avg('calificacion_general', filter=Q(utilizado=True)),
        )
        .order_by('semana')
    )

    labels_tend = []
    enviadas_tend = []
    respondidas_tend = []
    calificacion_tend = []

    for row in datos_tendencia:
        labels_tend.append(row['semana'].strftime('%d/%m/%Y'))
        enviadas_tend.append(row['total_enviadas_s'])
        respondidas_tend.append(row['total_respondidas_s'])
        calificacion_tend.append(round(row['calificacion_promedio_s'] or 0, 1))

    tendencia = {
        'labels': labels_tend,
        'datasets': {
            'total_enviadas': enviadas_tend,
            'total_respondidas': respondidas_tend,
            'calificacion_promedio': calificacion_tend,
        },
    }

    # ---- 4. Distribución NPS ----
    nps_agg = qs.filter(utilizado=True, nps__isnull=False).aggregate(
        promotores=Count('id', filter=Q(nps__gte=9)),
        pasivos=Count('id', filter=Q(nps__gte=7, nps__lte=8)),
        detractores=Count('id', filter=Q(nps__lte=6)),
        total=Count('id'),
    )
    nps_total = nps_agg['total'] or 0
    nps_score_dist = round(
        ((nps_agg['promotores'] - nps_agg['detractores']) / nps_total * 100)
        if nps_total > 0 else 0, 1
    )
    nps_dist = {
        'promotores': nps_agg['promotores'],
        'pasivos': nps_agg['pasivos'],
        'detractores': nps_agg['detractores'],
        'total': nps_total,
        'nps_score': nps_score_dist,
    }

    # ---- 5. Ranking por responsable ----
    datos_resp = (
        qs.filter(correo_enviado=True)
        .values(
            'orden__responsable_seguimiento__id',
            'orden__responsable_seguimiento__nombre_completo',
        )
        .annotate(
            total_enviadas_r=Count('id'),
            total_respondidas_r=Count('id', filter=Q(utilizado=True)),
            calificacion_promedio=Avg('calificacion_general', filter=Q(utilizado=True)),
            nps_promedio=Avg('nps', filter=Q(utilizado=True)),
            total_recomiendan=Count('id', filter=Q(utilizado=True, recomienda=True)),
            total_con_recomendacion=Count('id', filter=Q(utilizado=True, recomienda__isnull=False)),
            promotores=Count('id', filter=Q(utilizado=True, nps__gte=9)),
            detractores=Count('id', filter=Q(utilizado=True, nps__lte=6)),
            respondidas_con_nps=Count('id', filter=Q(utilizado=True, nps__isnull=False)),
        )
        .order_by('-calificacion_promedio')
    )

    responsables = []
    for row in datos_resp:
        nombre = row['orden__responsable_seguimiento__nombre_completo'] or '(Sin responsable)'
        t_env  = row['total_enviadas_r']
        t_resp = row['total_respondidas_r']
        nps_s  = round(
            ((row['promotores'] - row['detractores']) / row['respondidas_con_nps'] * 100)
            if row['respondidas_con_nps'] > 0 else 0, 1
        )
        tasa_rec = round(
            (row['total_recomiendan'] / row['total_con_recomendacion'] * 100)
            if row['total_con_recomendacion'] > 0 else 0, 1
        )
        responsables.append({
            'id': row['orden__responsable_seguimiento__id'],
            'nombre': nombre,
            'total_enviadas': t_env,
            'total_respondidas': t_resp,
            'calificacion_promedio': round(row['calificacion_promedio'] or 0, 1),
            'nps_promedio': round(row['nps_promedio'] or 0, 1),
            'tasa_recomendacion': tasa_rec,
            'nps_score': nps_s,
        })

    # ---- 6. Comentarios ----
    # Con filtros activos → todos; sin filtros → últimos 10
    comentarios_qs = (
        qs.filter(utilizado=True)
        .exclude(comentario_cliente='')
        .order_by('-fecha_respuesta')
    )
    if not hay_filtros:
        comentarios_qs = comentarios_qs[:10]

    comentarios = []
    for fb in comentarios_qs:
        comentarios.append({
            'orden_numero': (
                fb.orden.detalle_equipo.orden_cliente
                if hasattr(fb.orden, 'detalle_equipo') and fb.orden.detalle_equipo.orden_cliente
                else fb.orden.numero_orden_interno
            ),
            'orden_id': fb.orden.id,
            'responsable': str(fb.orden.responsable_seguimiento) if fb.orden.responsable_seguimiento else '',
            'calificacion': fb.calificacion_general,
            'nps': fb.nps,
            'recomienda': fb.recomienda,
            'comentario': fb.comentario_cliente,
            'fecha': fecha_local_pais(fb.fecha_respuesta, pais).strftime('%d/%m/%Y')
                     if fb.fecha_respuesta else '',
        })

    # ---- 7. Descripción del período ----
    partes_periodo = []
    if fecha_desde:
        partes_periodo.append(f'Desde: {fecha_desde}')
    if fecha_hasta:
        partes_periodo.append(f'Hasta: {fecha_hasta}')
    if responsable_id:
        partes_periodo.append('Responsable filtrado')
    if sucursal_id:
        partes_periodo.append('Sucursal filtrada')
    if tipo_orden:
        partes_periodo.append(f'Tipo: {tipo_orden}')
    periodo = ' | '.join(partes_periodo) if partes_periodo else 'Todos los registros'

    # ---- 8. Buscar análisis IA cacheado (si existe) ----
    # Usamos el mismo cálculo de hash SHA-256 que api_analisis_sentimiento_ia
    # para encontrar el análisis guardado que corresponde exactamente a este
    # conjunto de encuestas filtradas.
    analisis_ia = None
    try:
        import hashlib
        import json as _json
        from .models import AnalisisSentimientoEncuesta

        encuestas_para_hash = list(
            respondidas_qs
            .order_by('fecha_respuesta')
            .values(
                'calificacion_general',
                'calificacion_atencion',
                'calificacion_tiempo',
                'nps',
                'recomienda',
                'comentario_cliente',
            )
        )
        if encuestas_para_hash:
            hash_input = _json.dumps(encuestas_para_hash, sort_keys=True, ensure_ascii=False)
            hash_encuestas = hashlib.sha256(hash_input.encode('utf-8')).hexdigest()
            analisis_ia = (
                AnalisisSentimientoEncuesta.objects
                .filter(hash_encuestas=hash_encuestas)
                .order_by('-fecha_analisis')
                .first()
            )
    except Exception as _e:
        logger.warning(f'No se pudo recuperar análisis IA para el PDF: {_e}')

    # ---- 9. Empaquetar y generar PDF ----
    datos_pdf = {
        'kpis': kpis,
        'tendencia': tendencia,
        'nps_dist': nps_dist,
        'responsables': responsables,
        'comentarios': comentarios,
        'periodo': periodo,
        'filtros_activos': hay_filtros,
        'analisis_ia': analisis_ia,   # None si no hay análisis guardado
    }

    try:
        pdf_buffer = generar_pdf_reporte_encuestas(datos_pdf)
    except Exception as exc:
        logger.error(f'Error generando PDF de encuestas: {exc}', exc_info=True)
        messages.error(request, f'Error al generar el PDF: {exc}')
        return redirect('servicio_tecnico:dashboard_encuestas')

    fecha_str = now.strftime('%Y%m%d_%H%M')
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="Reporte_Encuestas_Satisfaccion_{fecha_str}.pdf"'
    )
    return response


# ============================================================================
# ANÁLISIS DE SENTIMIENTO IA — Encuestas de Satisfacción (Abril 2026)
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['POST'])
def api_analisis_sentimiento_ia(request):
    """
    Endpoint AJAX que genera (o devuelve desde caché) el análisis de
    sentimiento IA sobre el conjunto de encuestas de satisfacción.

    Flujo:
    1. Aplica los mismos filtros del dashboard (fecha, responsable, sucursal…)
    2. Obtiene solo las encuestas respondidas (utilizado=True, tipo='satisfaccion')
    3. Calcula un SHA-256 del conjunto → busca en AnalisisSentimientoEncuesta
    4. Si existe ese hash y no se pidió forzar → devuelve el análisis cacheado
    5. Si no existe o forzar=true → llama a Ollama → guarda → devuelve

    Body JSON esperado (todos opcionales):
        fecha_desde    (str YYYY-MM-DD)
        fecha_hasta    (str YYYY-MM-DD)
        responsable_id (int)
        sucursal_id    (int)
        tipo_orden     (str: 'diagnostico' | 'venta_mostrador')
        forzar         (bool: true para regenerar aunque exista caché)

    EXPLICACIÓN PARA PRINCIPIANTES:
    Esta vista es como un "botón de análisis inteligente". Si ya analizamos
    estos datos antes y no cambiaron, devuelve el resultado guardado al instante.
    Solo llama a la IA cuando es realmente necesario.
    """
    import hashlib
    import json as json_stdlib
    from django.conf import settings as django_settings
    from .models import AnalisisSentimientoEncuesta
    from .ollama_client import analizar_sentimiento_dispatch

    # ── 0. Verificar que la IA está habilitada ──────────────────────────────
    if not getattr(django_settings, 'AI_ENABLED', False):
        return JsonResponse({
            'success': False,
            'error': 'La función de IA no está habilitada en este entorno.',
        }, status=503)

    # ── 1. Parsear el body JSON del POST ────────────────────────────────────
    try:
        body = json_stdlib.loads(request.body or '{}')
    except (json_stdlib.JSONDecodeError, ValueError):
        body = {}

    forzar = bool(body.get('forzar', False))
    # modelo_override: llega con prefijo visual "[Gemini] ..." o "[Ollama] ..."
    # Si está vacío el dispatcher usa el modelo Ollama por defecto.
    modelo_override = str(body.get('modelo', '')).strip()

    # Inyectar los filtros del body como GET para poder reutilizar
    # _filtrar_encuestas_satisfaccion que lee de request.GET
    from django.http import QueryDict
    get_params = QueryDict(mutable=True)
    for campo in ('fecha_desde', 'fecha_hasta', 'responsable_id',
                  'sucursal_id', 'tipo_orden'):
        valor = body.get(campo)
        if valor:
            get_params[campo] = str(valor)

    # Crear un request temporal con los GET params del body
    request_filtrado = request
    request_filtrado.GET = get_params  # noqa: temporal override

    # ── 2. Obtener encuestas respondidas ────────────────────────────────────
    qs = _filtrar_encuestas_satisfaccion(request_filtrado).filter(
        utilizado=True,  # Solo encuestas donde el cliente ya respondió
    ).order_by('fecha_respuesta')

    encuestas_qs = list(qs.values(
        'calificacion_general',
        'calificacion_atencion',
        'calificacion_tiempo',
        'nps',
        'recomienda',
        'comentario_cliente',
    ))

    if not encuestas_qs:
        return JsonResponse({
            'success': False,
            'error': 'No hay encuestas respondidas para analizar con los filtros actuales.',
        }, status=404)

    # ── 3. Calcular hash SHA-256 del conjunto ───────────────────────────────
    # Usamos una representación canónica (sorted keys) para que el hash sea
    # consistente independientemente del orden de las claves en el dict.
    hash_input = json_stdlib.dumps(encuestas_qs, sort_keys=True, ensure_ascii=False)
    hash_encuestas = hashlib.sha256(hash_input.encode('utf-8')).hexdigest()

    # ── 4. Buscar análisis cacheado ─────────────────────────────────────────
    if not forzar:
        analisis_existente = (
            AnalisisSentimientoEncuesta.objects
            .filter(hash_encuestas=hash_encuestas)
            .order_by('-fecha_analisis')
            .first()
        )
        if analisis_existente:
            return JsonResponse({
                'success': True,
                'desde_cache': True,
                'sentimiento_general': analisis_existente.sentimiento_general,
                'resumen_ejecutivo':   analisis_existente.resumen_ejecutivo,
                'temas_positivos':     analisis_existente.temas_positivos,
                'temas_negativos':     analisis_existente.temas_negativos,
                'recomendacion_ia':    analisis_existente.recomendacion_ia,
                'total_encuestas':     analisis_existente.total_encuestas,
                'modelo_usado':        analisis_existente.modelo_usado,
                'fecha_analisis':      analisis_existente.fecha_analisis.strftime(
                    '%d/%m/%Y a las %H:%M'
                ),
                'badge_color':         analisis_existente.badge_color,
                'icono':               analisis_existente.icono,
            })

    # ── 5. Preparar datos para el cliente de IA ─────────────────────────────
    # Normalizar los nombres de campo: el modelo usa 'comentario_cliente'
    # pero la función de análisis espera 'comentario'
    encuestas_para_ia = [
        {
            'calificacion_general':  enc.get('calificacion_general'),
            'calificacion_atencion': enc.get('calificacion_atencion'),
            'calificacion_tiempo':   enc.get('calificacion_tiempo'),
            'nps':                   enc.get('nps'),
            'recomienda':            enc.get('recomienda'),
            'comentario':            enc.get('comentario_cliente', '') or '',
        }
        for enc in encuestas_qs
    ]

    modelo_ia = modelo_override  # Puede ser "[Gemini] gemini-2.0-flash", "[Ollama] gemma4:e4b", etc.

    # ── 6. Llamar a la IA vía dispatcher (Ollama o Gemini según prefijo) ─────
    logger.info(
        f'[api_analisis_sentimiento_ia] Llamando dispatcher con {len(encuestas_para_ia)} '
        f'encuestas. Hash: {hash_encuestas[:12]}… forzar={forzar} modelo="{modelo_override or "(default)"}"'
    )

    resultado_ia = analizar_sentimiento_dispatch(
        encuestas=encuestas_para_ia,
        modelo_override=modelo_override,
    )

    if not resultado_ia.get('success'):
        return JsonResponse({
            'success': False,
            'error': resultado_ia.get('error', 'Error desconocido en el análisis de IA.'),
        }, status=503)

    analisis = resultado_ia['analisis']

    # ── 7. Guardar en base de datos ─────────────────────────────────────────
    filtros_aplicados = {
        k: body.get(k)
        for k in ('fecha_desde', 'fecha_hasta', 'responsable_id', 'sucursal_id', 'tipo_orden')
        if body.get(k)
    }

    registro = AnalisisSentimientoEncuesta.objects.create(
        sentimiento_general = analisis.get('sentimiento_general', 'neutral'),
        resumen_ejecutivo   = analisis.get('resumen_ejecutivo', ''),
        temas_positivos     = analisis.get('temas_positivos', []),
        temas_negativos     = analisis.get('temas_negativos', []),
        recomendacion_ia    = analisis.get('recomendacion_ia', ''),
        total_encuestas     = len(encuestas_qs),
        hash_encuestas      = hash_encuestas,
        filtros_aplicados   = filtros_aplicados,
        modelo_usado        = resultado_ia.get('modelo_usado', modelo_ia),
    )

    # ── 8. Devolver respuesta ───────────────────────────────────────────────
    return JsonResponse({
        'success':            True,
        'desde_cache':        False,
        'sentimiento_general': registro.sentimiento_general,
        'resumen_ejecutivo':  registro.resumen_ejecutivo,
        'temas_positivos':    registro.temas_positivos,
        'temas_negativos':    registro.temas_negativos,
        'recomendacion_ia':   registro.recomendacion_ia,
        'total_encuestas':    registro.total_encuestas,
        'modelo_usado':       registro.modelo_usado,
        'fecha_analisis':     registro.fecha_analisis.strftime('%d/%m/%Y a las %H:%M'),
        'badge_color':        registro.badge_color,
        'icono':              registro.icono,
    })
