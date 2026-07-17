"""
Dashboard y APIs de feedback de rechazo de cotización (Fase 4).

urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
"""

import logging

from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from inventario.models import Empleado, Sucursal

from .decorators import permission_required_with_message

logger = logging.getLogger(__name__)


# ============================================================================
# DASHBOARD DE FEEDBACK DE RECHAZO DE COTIZACIÓN (Marzo 2026)
# ============================================================================

def _filtrar_feedback_rechazo(request):
    """
    Helper: construye queryset base de FeedbackCliente tipo 'rechazo'
    aplicando los filtros GET comunes (fecha, responsable, sucursal, motivo).
    Retorna el queryset con annotate de fecha_expiracion.
    """
    from .models import FeedbackCliente
    from django.db.models import F, ExpressionWrapper, DateTimeField
    from datetime import timedelta

    qs = FeedbackCliente.objects.filter(tipo='rechazo').select_related(
        'orden__responsable_seguimiento',
        'orden__sucursal',
        'orden__detalle_equipo',
        'cotizacion',
    )

    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    responsable_id = request.GET.get('responsable_id')
    sucursal_id = request.GET.get('sucursal_id')
    motivo_rechazo = request.GET.get('motivo_rechazo')

    if fecha_desde:
        qs = qs.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_creacion__date__lte=fecha_hasta)
    if responsable_id:
        qs = qs.filter(orden__responsable_seguimiento_id=responsable_id)
    if sucursal_id:
        qs = qs.filter(orden__sucursal_id=sucursal_id)
    if motivo_rechazo:
        qs = qs.filter(motivo_rechazo_snapshot=motivo_rechazo)

    qs = qs.annotate(
        fecha_expiracion=ExpressionWrapper(
            F('fecha_creacion') + timedelta(days=7),
            output_field=DateTimeField()
        )
    )
    return qs




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
def dashboard_feedback_rechazo(request):
    """
    Vista principal del panel de feedback de rechazo.
    Renderiza el template; la data se carga vía AJAX.
    """
    from config.constants import MOTIVO_RECHAZO_COTIZACION

    empleados = Empleado.objects.filter(activo=True).order_by('nombre_completo')
    sucursales = Sucursal.objects.filter(activa=True).order_by('nombre')

    return render(request, 'servicio_tecnico/dashboard_feedback_rechazo.html', {
        'empleados': empleados,
        'sucursales': sucursales,
        'motivos_rechazo': MOTIVO_RECHAZO_COTIZACION,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_feedback_rechazo_kpis(request):
    """
    API JSON: KPIs globales del dashboard de feedback de rechazo.
    """
    now = timezone.now()
    qs = _filtrar_feedback_rechazo(request)

    total_enviados = qs.filter(correo_enviado=True).count()
    total_respondidos = qs.filter(utilizado=True).count()
    total_pendientes = qs.filter(
        utilizado=False, correo_enviado=True, fecha_expiracion__gte=now
    ).count()
    total_expirados = qs.filter(
        utilizado=False, fecha_expiracion__lt=now
    ).count()

    tasa_respuesta = round(
        (total_respondidos / total_enviados * 100) if total_enviados > 0 else 0, 1
    )

    # Motivo más frecuente
    motivo_top = (
        qs.filter(correo_enviado=True)
        .values('motivo_rechazo_snapshot')
        .annotate(total=Count('id'))
        .order_by('-total')
        .first()
    )

    from config.constants import MOTIVO_RECHAZO_COTIZACION
    motivos_dict = dict(MOTIVO_RECHAZO_COTIZACION)

    motivo_label = ''
    motivo_porcentaje = 0
    if motivo_top and total_enviados > 0:
        motivo_label = motivos_dict.get(motivo_top['motivo_rechazo_snapshot'], motivo_top['motivo_rechazo_snapshot'])
        motivo_porcentaje = round(motivo_top['total'] / total_enviados * 100, 1)

    return JsonResponse({
        'total_enviados': total_enviados,
        'total_respondidos': total_respondidos,
        'total_pendientes': total_pendientes,
        'total_expirados': total_expirados,
        'tasa_respuesta': tasa_respuesta,
        'motivo_mas_frecuente': motivo_label,
        'motivo_mas_frecuente_porcentaje': motivo_porcentaje,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_feedback_rechazo_por_motivo(request):
    """
    API JSON: distribución por motivo de rechazo.
    """
    qs = _filtrar_feedback_rechazo(request).filter(correo_enviado=True)

    from config.constants import MOTIVO_RECHAZO_COTIZACION
    motivos_dict = dict(MOTIVO_RECHAZO_COTIZACION)

    datos = (
        qs.values('motivo_rechazo_snapshot')
        .annotate(
            total=Count('id'),
            respondidos=Count('id', filter=Q(utilizado=True)),
        )
        .order_by('-total')
    )

    motivos = []
    for row in datos:
        clave = row['motivo_rechazo_snapshot']
        motivos.append({
            'motivo': clave,
            'label': motivos_dict.get(clave, clave or 'Sin motivo'),
            'total': row['total'],
            'respondidos': row['respondidos'],
        })

    return JsonResponse({'motivos': motivos})




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_feedback_rechazo_tendencia(request):
    """
    API JSON: tendencia temporal semanal de feedbacks de rechazo.
    """
    from django.db.models.functions import TruncWeek
    from config.paises_config import fecha_local_pais, get_pais_actual
    pais = get_pais_actual()

    qs = _filtrar_feedback_rechazo(request).filter(correo_enviado=True)

    datos_por_semana = (
        qs.annotate(semana=TruncWeek('fecha_creacion'))
        .values('semana')
        .annotate(
            total_enviados=Count('id'),
            total_respondidos=Count('id', filter=Q(utilizado=True)),
        )
        .order_by('semana')
    )

    labels = []
    datasets = {
        'total_enviados': [],
        'total_respondidos': [],
        'tasa_respuesta': [],
    }

    for row in datos_por_semana:
        labels.append(fecha_local_pais(row['semana'], pais).strftime('%d/%m/%Y'))
        datasets['total_enviados'].append(row['total_enviados'])
        datasets['total_respondidos'].append(row['total_respondidos'])
        tasa = round(
            (row['total_respondidos'] / row['total_enviados'] * 100)
            if row['total_enviados'] > 0 else 0, 1
        )
        datasets['tasa_respuesta'].append(tasa)

    return JsonResponse({'labels': labels, 'datasets': datasets})




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_feedback_rechazo_lista(request):
    """
    API JSON: lista paginada de feedbacks de rechazo.
    """
    from config.constants import MOTIVO_RECHAZO_COTIZACION
    from config.paises_config import fecha_local_pais, get_pais_actual
    motivos_dict = dict(MOTIVO_RECHAZO_COTIZACION)
    pais = get_pais_actual()

    now = timezone.now()
    qs = _filtrar_feedback_rechazo(request)

    # Filtro por estado (tab)
    estado = request.GET.get('estado', 'todos')
    if estado == 'respondidos':
        qs = qs.filter(utilizado=True)
    elif estado == 'pendientes':
        qs = qs.filter(utilizado=False, correo_enviado=True, fecha_expiracion__gte=now)
    elif estado == 'expirados':
        qs = qs.filter(utilizado=False, fecha_expiracion__lt=now)

    # Búsqueda
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        qs = qs.filter(
            Q(orden__numero_orden_interno__icontains=busqueda) |
            Q(orden__detalle_equipo__orden_cliente__icontains=busqueda) |
            Q(orden__detalle_equipo__email_cliente__icontains=busqueda) |
            Q(orden__detalle_equipo__marca__icontains=busqueda) |
            Q(orden__detalle_equipo__modelo__icontains=busqueda) |
            Q(comentario_cliente__icontains=busqueda)
        )

    qs = qs.order_by('-fecha_creacion')

    # Paginación
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 15))
    paginator = Paginator(qs, page_size)

    try:
        pagina = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        pagina = paginator.page(1)

    feedbacks = []
    for fb in pagina.object_list:
        orden = fb.orden
        detalle = getattr(orden, 'detalle_equipo', None)

        if fb.utilizado:
            estado_fb = 'respondido'
        elif not fb.correo_enviado:
            estado_fb = 'no_enviado'
        elif fb.fecha_expiracion < now:
            estado_fb = 'expirado'
        else:
            estado_fb = 'pendiente'

        feedbacks.append({
            'id': fb.id,
            'orden_numero': detalle.orden_cliente if detalle and detalle.orden_cliente else orden.numero_orden_interno,
            'orden_id': orden.id,
            'equipo': f"{detalle.marca} {detalle.get_tipo_equipo_display()} {detalle.modelo}" if detalle else '',
            'email_cliente': detalle.email_cliente if detalle else '',
            'responsable': str(orden.responsable_seguimiento) if orden.responsable_seguimiento else '',
            'sucursal': str(orden.sucursal) if orden.sucursal else '',
            'motivo_rechazo': motivos_dict.get(fb.motivo_rechazo_snapshot, fb.motivo_rechazo_snapshot or 'Sin motivo'),
            'fecha_envio': fecha_local_pais(fb.fecha_creacion, pais).strftime('%d/%m/%Y %H:%M'),
            'fecha_respuesta': fecha_local_pais(fb.fecha_respuesta, pais).strftime('%d/%m/%Y %H:%M') if fb.fecha_respuesta else None,
            'dias_restantes': fb.dias_restantes,
            'estado': estado_fb,
            'comentario_cliente': fb.comentario_cliente,
        })

    return JsonResponse({
        'feedbacks': feedbacks,
        'total': paginator.count,
        'paginas': paginator.num_pages,
        'pagina_actual': pagina.number,
    })




@login_required
@permission_required_with_message('servicio_tecnico.view_dashboard_gerencial')
@require_http_methods(['GET'])
def api_feedback_rechazo_comentarios(request):
    """
    API JSON: últimos comentarios de clientes en feedbacks de rechazo.
    """
    from config.constants import MOTIVO_RECHAZO_COTIZACION
    from config.paises_config import fecha_local_pais, get_pais_actual
    motivos_dict = dict(MOTIVO_RECHAZO_COTIZACION)
    pais = get_pais_actual()

    qs = _filtrar_feedback_rechazo(request).filter(
        utilizado=True,
    ).exclude(
        comentario_cliente=''
    ).order_by('-fecha_respuesta')[:10]

    comentarios = []
    for fb in qs:
        detalle = getattr(fb.orden, 'detalle_equipo', None)
        comentarios.append({
            'orden_numero': detalle.orden_cliente if detalle and detalle.orden_cliente else fb.orden.numero_orden_interno,
            'orden_id': fb.orden.id,
            'responsable': str(fb.orden.responsable_seguimiento) if fb.orden.responsable_seguimiento else '',
            'motivo_rechazo': motivos_dict.get(fb.motivo_rechazo_snapshot, fb.motivo_rechazo_snapshot or 'Sin motivo'),
            'comentario': fb.comentario_cliente,
            'fecha': fecha_local_pais(fb.fecha_respuesta, pais).strftime('%d/%m/%Y') if fb.fecha_respuesta else '',
        })

    return JsonResponse({'comentarios': comentarios})




@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def exportar_feedback_rechazo_excel(request):
    """
    Exporta los feedbacks de rechazo filtrados a un archivo Excel.
    Genera 3 hojas: Resumen KPIs, Feedbacks detallados, Por Motivo.
    """
    from openpyxl import Workbook
    from config.constants import MOTIVO_RECHAZO_COTIZACION

    motivos_dict = dict(MOTIVO_RECHAZO_COTIZACION)
    now = timezone.now()
    qs = _filtrar_feedback_rechazo(request).filter(correo_enviado=True)

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()

    # ── HOJA 1: Resumen KPIs ──────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen KPIs'

    total_enviados = qs.count()
    total_respondidos = qs.filter(utilizado=True).count()
    total_pendientes = qs.filter(utilizado=False, fecha_expiracion__gte=now).count()
    total_expirados = qs.filter(utilizado=False, fecha_expiracion__lt=now).count()
    tasa_respuesta = round((total_respondidos / total_enviados * 100) if total_enviados > 0 else 0, 1)

    kpis = [
        ('Métrica', 'Valor'),
        ('Total Feedbacks Enviados', total_enviados),
        ('Total Respondidos', total_respondidos),
        ('Total Pendientes', total_pendientes),
        ('Total Expirados', total_expirados),
        ('Tasa de Respuesta (%)', f'{tasa_respuesta}%'),
    ]

    for row_idx, (metrica, valor) in enumerate(kpis, 1):
        ws_resumen.cell(row=row_idx, column=1, value=metrica)
        ws_resumen.cell(row=row_idx, column=2, value=valor)
        if row_idx == 1:
            for col in (1, 2):
                cell = ws_resumen.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        else:
            for col in (1, 2):
                ws_resumen.cell(row=row_idx, column=col).border = thin_border

    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 20

    # ── HOJA 2: Feedbacks Detallados ──────────────────────────────────
    ws_feedbacks = wb.create_sheet('Feedbacks')
    headers = [
        'Orden', 'Equipo', 'Email Cliente', 'Responsable', 'Sucursal',
        'Motivo de Rechazo', 'Fecha Envío', 'Fecha Respuesta', 'Estado', 'Comentario'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_feedbacks.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, fb in enumerate(qs.order_by('-fecha_creacion'), 2):
        orden = fb.orden
        detalle = getattr(orden, 'detalle_equipo', None)

        if fb.utilizado:
            estado_str = 'Respondido'
        elif fb.fecha_expiracion < now:
            estado_str = 'Expirado'
        else:
            estado_str = 'Pendiente'

        valores = [
            detalle.orden_cliente if detalle and detalle.orden_cliente else orden.numero_orden_interno,
            f"{detalle.marca} {detalle.get_tipo_equipo_display()} {detalle.modelo}" if detalle else '',
            detalle.email_cliente if detalle else '',
            str(orden.responsable_seguimiento) if orden.responsable_seguimiento else '',
            str(orden.sucursal) if orden.sucursal else '',
            motivos_dict.get(fb.motivo_rechazo_snapshot, fb.motivo_rechazo_snapshot or 'Sin motivo'),
            fb.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            fb.fecha_respuesta.strftime('%d/%m/%Y %H:%M') if fb.fecha_respuesta else '',
            estado_str,
            fb.comentario_cliente,
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws_feedbacks.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws_feedbacks.column_dimensions[get_column_letter(col_idx)].width = 20

    # ── HOJA 3: Por Motivo de Rechazo ─────────────────────────────────
    ws_motivos = wb.create_sheet('Por Motivo')
    headers_mot = ['Motivo de Rechazo', 'Total', 'Respondidos', 'Tasa Respuesta (%)']
    for col_idx, header in enumerate(headers_mot, 1):
        cell = ws_motivos.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    datos_motivos = (
        qs.values('motivo_rechazo_snapshot')
        .annotate(
            total=Count('id'),
            respondidos=Count('id', filter=Q(utilizado=True)),
        )
        .order_by('-total')
    )

    for row_idx, row in enumerate(datos_motivos, 2):
        clave = row['motivo_rechazo_snapshot']
        tasa = round((row['respondidos'] / row['total'] * 100) if row['total'] > 0 else 0, 1)
        valores = [
            motivos_dict.get(clave, clave or 'Sin motivo'),
            row['total'],
            row['respondidos'],
            f'{tasa}%',
        ]
        for col_idx, valor in enumerate(valores, 1):
            cell = ws_motivos.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border

    for col_idx in range(1, len(headers_mot) + 1):
        ws_motivos.column_dimensions[get_column_letter(col_idx)].width = 30

    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_str = now.strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename=Feedback_Rechazo_{fecha_str}.xlsx'
    wb.save(response)
    return response
