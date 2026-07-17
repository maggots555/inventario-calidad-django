"""
Vistas de perfil de empleado y directorio (Fase 2 modularización).

EXPLICACIÓN PARA PRINCIPIANTES:
- mi_perfil / exportar_excel_mi_perfil: el técnico ve sus propias métricas.
- directorio_empleados / perfil_empleado: vista gerencial de otros empleados.
- _calcular_metricas_empleado / _calcular_rating_rapido: helpers privados
  compartidos por esas vistas (no son endpoints HTTP).

urls.py sigue usando views.mi_perfil etc. porque views.py reexporta.
"""

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone

from inventario.models import Empleado, Sucursal

from .models import OrdenServicio


def _calcular_metricas_empleado(empleado):
    """
    Calcula métricas de desempeño y rating para un empleado dado.
    Reutilizable por mi_perfil y perfil_empleado (vista gerencial).

    Retorna: (metricas: dict, rating: int, rol: str, rol_display: str)
    """
    from django.db.models import Avg, Sum, Count, Q
    from decimal import Decimal
    from datetime import timedelta
    from .models import Cotizacion, VentaMostrador, PiezaVentaMostrador, FeedbackCliente

    rol = empleado.rol
    ahora = timezone.now()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    hace_90_dias = ahora - timedelta(days=90)

    # Métricas comunes a todos los roles
    ordenes_activas_count = OrdenServicio.objects.filter(
        Q(tecnico_asignado_actual=empleado) | Q(responsable_seguimiento=empleado)
    ).exclude(estado__in=['entregado', 'cancelado']).distinct().count()

    ordenes_entregadas_mes = OrdenServicio.objects.filter(
        Q(tecnico_asignado_actual=empleado) | Q(responsable_seguimiento=empleado),
        estado='entregado',
        fecha_entrega__gte=inicio_mes
    ).distinct().count()

    dias_en_sistema = (ahora - empleado.fecha_ingreso).days if empleado.fecha_ingreso else 0

    metricas = {
        'ordenes_activas': ordenes_activas_count,
        'ordenes_entregadas_mes': ordenes_entregadas_mes,
        'dias_en_sistema': dias_en_sistema,
    }

    # Métricas para recepcionista
    if rol == 'recepcionista':
        ordenes_con_cotizacion = OrdenServicio.objects.filter(
            responsable_seguimiento=empleado,
            cotizacion__isnull=False,
        ).select_related('cotizacion')

        cotizaciones_ordenes = ordenes_con_cotizacion.filter(
            cotizacion__fecha_envio__gte=hace_90_dias
        )

        total_cotizaciones = cotizaciones_ordenes.count()
        cotizaciones_aceptadas = cotizaciones_ordenes.filter(cotizacion__usuario_acepto=True).count()
        cotizaciones_rechazadas = cotizaciones_ordenes.filter(cotizacion__usuario_acepto=False).count()
        cotizaciones_pendientes = cotizaciones_ordenes.filter(cotizacion__usuario_acepto__isnull=True).count()

        tasa_aceptacion = round(
            (cotizaciones_aceptadas / total_cotizaciones * 100) if total_cotizaciones > 0 else 0, 1
        )

        valor_cotizado = Decimal('0.00')
        valor_aceptado = Decimal('0.00')
        for orden in cotizaciones_ordenes.select_related('cotizacion'):
            cot = orden.cotizacion
            valor_cotizado += cot.costo_total
            if cot.usuario_acepto:
                valor_aceptado += cot.costo_total_final

        ventas_mostrador_qs = OrdenServicio.objects.filter(
            responsable_seguimiento=empleado,
            venta_mostrador__isnull=False,
            venta_mostrador__fecha_venta__gte=hace_90_dias,
        ).select_related('venta_mostrador')

        total_ventas_mostrador = ventas_mostrador_qs.count()
        monto_ventas_mostrador = Decimal('0.00')
        for orden in ventas_mostrador_qs:
            monto_ventas_mostrador += orden.venta_mostrador.total_venta

        encuestas_qs = FeedbackCliente.objects.filter(
            tipo='satisfaccion',
            orden__responsable_seguimiento=empleado,
            utilizado=True,
        )
        total_encuestas_respondidas = encuestas_qs.count()
        encuestas_avgs = encuestas_qs.aggregate(
            nps_promedio=Avg('nps'),
            calificacion_promedio=Avg('calificacion_general'),
        )
        nps_promedio = round(encuestas_avgs['nps_promedio'] or 0, 1)
        calificacion_promedio = round(encuestas_avgs['calificacion_promedio'] or 0, 1)

        con_recomendacion = encuestas_qs.filter(recomienda__isnull=False).count()
        recomiendan = encuestas_qs.filter(recomienda=True).count()
        tasa_recomendacion = round(
            (recomiendan / con_recomendacion * 100) if con_recomendacion > 0 else 0, 1
        )

        # Últimos 15 comentarios escritos por clientes (solo si tienen texto)
        comentarios_clientes = list(
            FeedbackCliente.objects.filter(
                tipo='satisfaccion',
                orden__responsable_seguimiento=empleado,
                utilizado=True,
                comentario_cliente__gt='',
            ).select_related('orden__detalle_equipo').order_by('-fecha_respuesta')[:15]
        )

        metricas.update({
            'total_cotizaciones': total_cotizaciones,
            'cotizaciones_aceptadas': cotizaciones_aceptadas,
            'cotizaciones_rechazadas': cotizaciones_rechazadas,
            'cotizaciones_pendientes': cotizaciones_pendientes,
            'tasa_aceptacion': tasa_aceptacion,
            'valor_cotizado': valor_cotizado,
            'valor_aceptado': valor_aceptado,
            'total_ventas_mostrador': total_ventas_mostrador,
            'monto_ventas_mostrador': monto_ventas_mostrador,
            'total_encuestas_respondidas': total_encuestas_respondidas,
            'nps_promedio': nps_promedio,
            'calificacion_promedio': calificacion_promedio,
            'tasa_recomendacion': tasa_recomendacion,
            'comentarios_clientes': comentarios_clientes,
        })

    # Métricas para técnico
    elif rol == 'tecnico':
        stats_activas = empleado.obtener_estadisticas_ordenes_activas()

        ordenes_tecnico_cot = OrdenServicio.objects.filter(
            tecnico_asignado_actual=empleado,
            cotizacion__isnull=False,
            cotizacion__fecha_envio__gte=hace_90_dias,
        ).select_related('cotizacion')

        total_cotizaciones = ordenes_tecnico_cot.count()
        cotizaciones_aceptadas = ordenes_tecnico_cot.filter(cotizacion__usuario_acepto=True).count()
        cotizaciones_rechazadas = ordenes_tecnico_cot.filter(cotizacion__usuario_acepto=False).count()

        tasa_aceptacion = round(
            (cotizaciones_aceptadas / total_cotizaciones * 100) if total_cotizaciones > 0 else 0, 1
        )

        valor_cotizado = Decimal('0.00')
        valor_aceptado = Decimal('0.00')
        for orden in ordenes_tecnico_cot:
            cot = orden.cotizacion
            valor_cotizado += cot.costo_total
            if cot.usuario_acepto:
                valor_aceptado += cot.costo_total_final

        ordenes_completadas_mes = OrdenServicio.objects.filter(
            tecnico_asignado_actual=empleado,
            estado='entregado',
            fecha_entrega__gte=inicio_mes,
        ).count()

        ordenes_rhitso = OrdenServicio.objects.filter(
            tecnico_asignado_actual=empleado,
            es_candidato_rhitso=True,
        ).exclude(estado__in=['cancelado']).count()

        encuestas_qs = FeedbackCliente.objects.filter(
            tipo='satisfaccion',
            orden__tecnico_asignado_actual=empleado,
            utilizado=True,
        )
        total_encuestas_respondidas = encuestas_qs.count()
        encuestas_avgs = encuestas_qs.aggregate(
            nps_promedio=Avg('nps'),
            calificacion_promedio=Avg('calificacion_general'),
        )
        nps_promedio = round(encuestas_avgs['nps_promedio'] or 0, 1)
        calificacion_promedio = round(encuestas_avgs['calificacion_promedio'] or 0, 1)

        # ── Ventas Mostrador del técnico — últimos 90 días ──────────────────
        vm_qs = VentaMostrador.objects.filter(
            orden__tecnico_asignado_actual=empleado,
            fecha_venta__gte=hace_90_dias,
        ).select_related('orden')

        total_vm = vm_qs.count()

        # Monto total: suma la propiedad total_venta de cada instancia
        monto_vm = Decimal('0.00')
        if total_vm > 0:
            for vm_obj in vm_qs:
                monto_vm += vm_obj.total_venta

        # Desglose por tipo de servicio booleano
        vm_limpiezas       = vm_qs.filter(incluye_limpieza=True).count()
        vm_cambios_pieza   = vm_qs.filter(incluye_cambio_pieza=True).count()
        vm_kits_limpieza   = vm_qs.filter(incluye_kit_limpieza=True).count()
        vm_reinstalaciones = vm_qs.filter(incluye_reinstalacion_so=True).count()
        vm_respaldos       = vm_qs.filter(incluye_respaldo=True).count()

        # Distribución de paquetes (solo los que tienen al menos 1 vendido)
        vm_paquetes_raw = {
            'premium': vm_qs.filter(paquete='premium').count(),
            'oro':     vm_qs.filter(paquete='oro').count(),
            'plata':   vm_qs.filter(paquete='plata').count(),
        }
        # Filtramos únicamente los paquetes con conteo > 0
        vm_paquetes = {k: v for k, v in vm_paquetes_raw.items() if v > 0}

        # Top 3 piezas más vendidas en ventas mostrador del técnico
        vm_top_piezas = list(
            PiezaVentaMostrador.objects.filter(
                venta_mostrador__orden__tecnico_asignado_actual=empleado,
                venta_mostrador__fecha_venta__gte=hace_90_dias,
            )
            .values('descripcion_pieza')
            .annotate(total=Count('id'))
            .order_by('-total')[:3]
        )
        # ───────────────────────────────────────────────────────────────────

        metricas.update({
            'ordenes_activas_tecnico': stats_activas['ordenes_activas'],
            'equipos_no_encienden': stats_activas['equipos_no_encienden'],
            'tiene_sobrecarga': stats_activas['tiene_sobrecarga'],
            'total_cotizaciones': total_cotizaciones,
            'cotizaciones_aceptadas': cotizaciones_aceptadas,
            'cotizaciones_rechazadas': cotizaciones_rechazadas,
            'tasa_aceptacion': tasa_aceptacion,
            'valor_cotizado': valor_cotizado,
            'valor_aceptado': valor_aceptado,
            'ordenes_completadas_mes': ordenes_completadas_mes,
            'ordenes_rhitso': ordenes_rhitso,
            'total_encuestas_respondidas': total_encuestas_respondidas,
            'nps_promedio': nps_promedio,
            'calificacion_promedio': calificacion_promedio,
            # Ventas Mostrador
            'total_ventas_mostrador': total_vm,
            'monto_ventas_mostrador': monto_vm,
            'vm_limpiezas': vm_limpiezas,
            'vm_cambios_pieza': vm_cambios_pieza,
            'vm_kits_limpieza': vm_kits_limpieza,
            'vm_reinstalaciones': vm_reinstalaciones,
            'vm_respaldos': vm_respaldos,
            'vm_paquetes': vm_paquetes,
            'vm_top_piezas': vm_top_piezas,
        })

    # Rating de desempeño (1-99)
    rating = 50
    if metricas.get('tasa_aceptacion', 0) > 0:
        rating += min(30, int(metricas['tasa_aceptacion'] * 0.3))
    if metricas.get('calificacion_promedio', 0) > 0:
        rating += min(15, int(metricas['calificacion_promedio'] * 3))
    if metricas.get('total_ventas_mostrador', 0) > 0:
        rating += min(4, int(metricas['total_ventas_mostrador'] * 0.5))
    rating = max(1, min(99, rating))

    rol_display = dict(empleado.ROL_CHOICES).get(rol, rol)
    return metricas, rating, rol, rol_display


def _calcular_rating_rapido(empleado):
    """
    Cálculo rápido del rating para las mini-cards del directorio.
    Evita consultas pesadas: solo usa tasa de aceptación y órdenes activas.
    Retorna un int entre 1 y 99.
    """
    from django.db.models import Q
    from datetime import timedelta
    from .models import Cotizacion

    ahora = timezone.now()
    hace_90_dias = ahora - timedelta(days=90)
    rol = empleado.rol
    rating = 50

    if rol == 'recepcionista':
        ordenes_cot = OrdenServicio.objects.filter(
            responsable_seguimiento=empleado,
            cotizacion__isnull=False,
            cotizacion__fecha_envio__gte=hace_90_dias,
        )
        total = ordenes_cot.count()
        aceptadas = ordenes_cot.filter(cotizacion__usuario_acepto=True).count()
        if total > 0:
            rating += min(30, int((aceptadas / total * 100) * 0.3))
        # Ventas mostrador
        ventas = OrdenServicio.objects.filter(
            responsable_seguimiento=empleado,
            venta_mostrador__isnull=False,
            venta_mostrador__fecha_venta__gte=hace_90_dias,
        ).count()
        rating += min(4, int(ventas * 0.5))

    elif rol == 'tecnico':
        ordenes_cot = OrdenServicio.objects.filter(
            tecnico_asignado_actual=empleado,
            cotizacion__isnull=False,
            cotizacion__fecha_envio__gte=hace_90_dias,
        )
        total = ordenes_cot.count()
        aceptadas = ordenes_cot.filter(cotizacion__usuario_acepto=True).count()
        if total > 0:
            rating += min(30, int((aceptadas / total * 100) * 0.3))

    return max(1, min(99, rating))


# ============================================================================
# MI PERFIL — MÉTRICAS PERSONALES DEL EMPLEADO (Marzo 2026)
# ============================================================================

@login_required
def mi_perfil(request):
    """
    Página "Mi Perfil" con tarjeta de métricas personales.
    Muestra métricas personalizadas según el rol del empleado:
      - Recepcionista: cotizaciones gestionadas, ventas mostrador, encuestas
      - Técnico: órdenes asignadas, cotizaciones donde participó, RHITSO
    """
    empleado = getattr(request.user, 'empleado', None)
    if not empleado:
        messages.warning(request, 'Tu cuenta no tiene un perfil de empleado asociado.')
        return redirect('servicio_tecnico:inicio')

    metricas, rating, rol, rol_display = _calcular_metricas_empleado(empleado)

    context = {
        'empleado': empleado,
        'rol': rol,
        'rol_display': rol_display,
        'metricas': metricas,
        'rating': rating,
    }

    return render(request, 'servicio_tecnico/mi_perfil.html', context)


# ============================================================================
# EXPORTACIÓN EXCEL — REPORTE INDIVIDUAL POR RESPONSABLE (Mi Perfil)
# ============================================================================

@login_required
def exportar_excel_mi_perfil(request):
    """
    Genera y descarga un archivo Excel con el reporte individual OOW-/FL-
    del empleado autenticado como responsable de seguimiento.

    EXPLICACIÓN PARA PRINCIPIANTES:
    ================================
    Esta vista toma el usuario que está logueado, obtiene su registro de Empleado,
    filtra SOLO las órdenes OOW-/FL- donde él es responsable_seguimiento,
    y construye un Excel de 3 hojas:
      1. "Resumen Personal"   → KPIs y métricas propias del responsable
      2. "Órdenes Activas"    → Tabla de órdenes no cerradas
      3. "Órdenes Cerradas"   → Tabla de órdenes entregadas o canceladas

    Seguridad: solo requiere estar autenticado. Cada usuario solo puede
    descargar sus propios datos — nunca los de otro agente.

    Returns:
        HttpResponse: Archivo .xlsx listo para descarga
    """
    from django.http import HttpResponse
    from django.db.models import Q
    from datetime import datetime

    # ------------------------------------------------------------------
    # PASO 0: Verificar que el usuario tiene perfil de empleado
    # ------------------------------------------------------------------
    empleado = getattr(request.user, 'empleado', None)
    if not empleado:
        messages.warning(request, 'Tu cuenta no tiene un perfil de empleado asociado.')
        return redirect('servicio_tecnico:mi_perfil')

    # ------------------------------------------------------------------
    # PASO 1: Importar openpyxl y helpers del módulo excel_exporters
    # ------------------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from .excel_exporters import (
            get_header_style, get_title_style, get_kpi_title_style, get_kpi_value_style,
            get_estado_color, apply_cell_style, auto_adjust_column_width,
        )
    except ImportError as e:
        messages.error(request, f'Error al generar el reporte: {str(e)}')
        return redirect('servicio_tecnico:mi_perfil')

    # ------------------------------------------------------------------
    # PASO 2: Construir QuerySet — solo OOW-/FL- del empleado autenticado
    # ------------------------------------------------------------------
    ordenes = OrdenServicio.objects.filter(
        Q(detalle_equipo__orden_cliente__istartswith='OOW-') |
        Q(detalle_equipo__orden_cliente__istartswith='FL-'),
        responsable_seguimiento=empleado
    ).select_related(
        'detalle_equipo',
        'sucursal',
        'responsable_seguimiento',
        'tecnico_asignado_actual',
        'venta_mostrador',
        'cotizacion'
    ).prefetch_related(
        'historial'
    ).order_by('-fecha_ingreso')

    # ------------------------------------------------------------------
    # PASO 3: Calcular métricas del responsable (sin importar calcular_estadisticas_por_responsable
    # porque aquí solo hay un responsable — calculamos directamente)
    # ------------------------------------------------------------------
    total_ordenes = ordenes.count()
    ordenes_activas_qs = ordenes.exclude(estado__in=['entregado', 'cancelado'])
    ordenes_cerradas_qs = ordenes.filter(estado__in=['entregado', 'cancelado'])

    total_activas    = ordenes_activas_qs.count()
    total_entregadas = ordenes.filter(estado='entregado').count()
    total_canceladas = ordenes.filter(estado='cancelado').count()

    # Acumular montos
    from decimal import Decimal
    monto_vm         = Decimal('0.00')
    monto_cotiz      = Decimal('0.00')
    cotiz_aceptadas  = 0
    cotiz_pendientes = 0
    cotiz_rechazadas = 0
    vm_count         = 0
    dias_acumulados  = 0
    ordenes_en_tiempo = 0

    for orden in ordenes:
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            vm_count  += 1
            monto_vm  += orden.venta_mostrador.total_venta

        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                cotiz_aceptadas += 1
                monto_cotiz     += orden.cotizacion.costo_total_final
            elif orden.cotizacion.usuario_acepto is False:
                cotiz_rechazadas += 1
            else:
                cotiz_pendientes += 1

        dias = orden.dias_habiles_en_servicio
        dias_acumulados += dias
        if dias <= 15:
            ordenes_en_tiempo += 1

    monto_total    = monto_vm + monto_cotiz
    tiempo_promedio = round(dias_acumulados / total_ordenes, 1) if total_ordenes > 0 else 0
    tasa_entrega   = round((total_entregadas / total_ordenes) * 100, 1) if total_ordenes > 0 else 0
    pct_en_tiempo  = round((ordenes_en_tiempo / total_ordenes) * 100, 1) if total_ordenes > 0 else 0

    # ------------------------------------------------------------------
    # PASO 4: Crear Workbook
    # ------------------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)   # Eliminar la hoja vacía predeterminada

    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')

    # ==================================================================
    # HOJA 1 — RESUMEN PERSONAL
    # ==================================================================
    ws_resumen = wb.create_sheet("Resumen Personal")

    # Fila 1: Título principal
    ws_resumen.merge_cells('A1:F1')
    titulo_cell = ws_resumen['A1']
    titulo_cell.value = f"REPORTE INDIVIDUAL OOW-/FL- — {empleado.nombre_completo.upper()}"
    apply_cell_style(titulo_cell, get_title_style())
    ws_resumen.row_dimensions[1].height = 28

    # Fila 2: subtítulo / fecha de generación
    ws_resumen.merge_cells('A2:F2')
    sub_cell = ws_resumen['A2']
    sub_cell.value = f"Generado: {fecha_generacion}  |  Solo órdenes con prefijo OOW- o FL-"
    sub_cell.font      = Font(italic=True, size=9, color="666666")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_resumen.row_dimensions[2].height = 16

    # Fila 4: encabezado de sección
    ws_resumen['A4'] = "ESTADÍSTICAS PERSONALES"
    apply_cell_style(ws_resumen['A4'], get_kpi_title_style())

    # KPIs en dos columnas (A = nombre, B = valor)
    kpis = [
        ('Total de Órdenes OOW-/FL-:',         total_ordenes),
        ('Órdenes Activas:',                     total_activas),
        ('Órdenes Entregadas:',                  total_entregadas),
        ('Órdenes Canceladas:',                  total_canceladas),
        ('',                                     ''),
        ('Tiempo Promedio (días hábiles):',       tiempo_promedio),
        ('% Órdenes en Tiempo (≤ 15 días):',     f'{pct_en_tiempo}%'),
        ('Tasa de Entrega:',                      f'{tasa_entrega}%'),
        ('',                                     ''),
        ('Ventas Mostrador (cantidad):',          vm_count),
        ('Monto Ventas Mostrador:',              f'${float(monto_vm):,.2f}'),
        ('',                                     ''),
        ('Cotizaciones Aceptadas:',               cotiz_aceptadas),
        ('Cotizaciones Pendientes:',              cotiz_pendientes),
        ('Cotizaciones Rechazadas:',              cotiz_rechazadas),
        ('Monto Cotizaciones Aceptadas:',        f'${float(monto_cotiz):,.2f}'),
        ('',                                     ''),
        ('MONTO TOTAL GENERADO:',               f'${float(monto_total):,.2f}'),
    ]

    row = 6
    for nombre_kpi, valor_kpi in kpis:
        if nombre_kpi == '':
            row += 1
            continue

        ws_resumen[f'A{row}'] = nombre_kpi
        ws_resumen[f'B{row}'] = valor_kpi
        apply_cell_style(ws_resumen[f'A{row}'], get_kpi_title_style())

        if 'MONTO TOTAL' in nombre_kpi:
            ws_resumen[f'B{row}'].fill = PatternFill(
                start_color="28a745", end_color="28a745", fill_type="solid"
            )
            ws_resumen[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        else:
            apply_cell_style(ws_resumen[f'B{row}'], get_kpi_value_style())

        row += 1

    # Ajustar anchos de columna
    ws_resumen.column_dimensions['A'].width = 38
    ws_resumen.column_dimensions['B'].width = 22

    # ==================================================================
    # HOJA 2 — ÓRDENES ACTIVAS
    # ==================================================================
    ws_activas = wb.create_sheet("Órdenes Activas")

    # Título de hoja
    ws_activas.merge_cells('A1:O1')
    tit_activas = ws_activas['A1']
    tit_activas.value = (
        f"ÓRDENES ACTIVAS — {empleado.nombre_completo.upper()} "
        f"({total_activas} orden{'es' if total_activas != 1 else ''})"
    )
    tit_activas.fill      = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    tit_activas.font      = Font(bold=True, size=12, color="000000")
    tit_activas.alignment = Alignment(horizontal="left", vertical="center")
    ws_activas.row_dimensions[1].height = 22

    # Encabezados de la tabla
    headers = [
        'N° Orden Cliente', 'N° de Serie', 'Tipo Equipo', 'Marca',
        'Modelo', 'Estado', 'Días Hábiles', 'Días Sin Actualizar',
        'Tipo de Orden', 'Monto', 'Sucursal', 'Fecha Ingreso',
        'Última Actualización', 'Cotización', 'Observaciones'
    ]
    row = 2
    for col_num, header in enumerate(headers, 1):
        cell = ws_activas.cell(row=row, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    ws_activas.row_dimensions[row].height = 20
    row += 1

    # Filas de datos — órdenes activas
    for orden in ordenes_activas_qs.order_by('-fecha_ingreso'):
        ws_activas.cell(row=row, column=1).value  = orden.detalle_equipo.orden_cliente
        ws_activas.cell(row=row, column=2).value  = orden.detalle_equipo.numero_serie or 'N/A'
        ws_activas.cell(row=row, column=3).value  = orden.detalle_equipo.get_tipo_equipo_display()
        ws_activas.cell(row=row, column=4).value  = orden.detalle_equipo.marca
        ws_activas.cell(row=row, column=5).value  = (orden.detalle_equipo.modelo or 'N/A')[:30]
        ws_activas.cell(row=row, column=6).value  = orden.get_estado_display()
        ws_activas.cell(row=row, column=7).value  = orden.dias_habiles_en_servicio
        ws_activas.cell(row=row, column=8).value  = orden.dias_sin_actualizacion_estado

        tipo_orden = 'Servicio Normal'
        monto_orden = 0
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            tipo_orden  = 'Venta Mostrador'
            monto_orden = float(orden.venta_mostrador.total_venta)
        elif hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                tipo_orden  = 'Cotización Aceptada'
                monto_orden = float(orden.cotizacion.costo_total_final)
            elif orden.cotizacion.usuario_acepto is False:
                tipo_orden = 'Cotización Rechazada'
            else:
                tipo_orden = 'Cotización Pendiente'

        ws_activas.cell(row=row, column=9).value  = tipo_orden
        ws_activas.cell(row=row, column=10).value = f'${monto_orden:,.2f}' if monto_orden > 0 else 'N/A'
        ws_activas.cell(row=row, column=11).value = orden.sucursal.nombre
        ws_activas.cell(row=row, column=12).value = orden.fecha_ingreso.strftime('%d/%m/%Y')

        ultima_act = orden.historial.order_by('-fecha_evento').first()
        ws_activas.cell(row=row, column=13).value = (
            ultima_act.fecha_evento.strftime('%d/%m/%Y') if ultima_act else 'N/A'
        )

        cotiz_estado = 'N/A'
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                cotiz_estado = '✅ Aceptada'
            elif orden.cotizacion.usuario_acepto is False:
                cotiz_estado = '❌ Rechazada'
            else:
                cotiz_estado = '⏳ Pendiente'
        ws_activas.cell(row=row, column=14).value = cotiz_estado

        alertas = []
        if orden.dias_habiles_en_servicio > 15:
            alertas.append('⚠️ RETRASADA')
        if orden.dias_sin_actualizacion_estado > 5:
            alertas.append(f'🔴 Sin actualizar {orden.dias_sin_actualizacion_estado}d')
        ws_activas.cell(row=row, column=15).value = ' | '.join(alertas) if alertas else 'OK'

        # Colorear celda de estado
        color_est = get_estado_color(orden.estado)
        ws_activas.cell(row=row, column=6).fill = PatternFill(
            start_color=color_est, end_color=color_est, fill_type="solid"
        )
        ws_activas.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF")

        # Rojo en días si está retrasada
        if orden.dias_habiles_en_servicio > 15:
            ws_activas.cell(row=row, column=7).fill = PatternFill(
                start_color="dc3545", end_color="dc3545", fill_type="solid"
            )
            ws_activas.cell(row=row, column=7).font = Font(bold=True, color="FFFFFF")

        # Resaltar candidatos RHITSO (morado claro)
        if orden.es_candidato_rhitso:
            rhitso_color = "ede9fe"
            for col in range(1, 16):
                cell = ws_activas.cell(row=row, column=col)
                if col not in [6, 7] or (col == 7 and orden.dias_habiles_en_servicio <= 15):
                    cell.fill = PatternFill(
                        start_color=rhitso_color, end_color=rhitso_color, fill_type="solid"
                    )

        row += 1

    auto_adjust_column_width(ws_activas)

    # ==================================================================
    # HOJA 3 — ÓRDENES CERRADAS (ENTREGADAS / CANCELADAS)
    # ==================================================================
    ws_cerradas = wb.create_sheet("Órdenes Cerradas")

    # Título de hoja
    ws_cerradas.merge_cells('A1:O1')
    tit_cerradas = ws_cerradas['A1']
    tit_cerradas.value = (
        f"ÓRDENES CERRADAS — {empleado.nombre_completo.upper()} "
        f"({ordenes_cerradas_qs.count()} orden{'es' if ordenes_cerradas_qs.count() != 1 else ''})"
    )
    tit_cerradas.fill      = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    tit_cerradas.font      = Font(bold=True, size=12, color="FFFFFF")
    tit_cerradas.alignment = Alignment(horizontal="left", vertical="center")
    ws_cerradas.row_dimensions[1].height = 22

    # Encabezados (mismos que órdenes activas)
    row = 2
    for col_num, header in enumerate(headers, 1):
        cell = ws_cerradas.cell(row=row, column=col_num)
        cell.value = header
        apply_cell_style(cell, get_header_style())
    ws_cerradas.row_dimensions[row].height = 20
    row += 1

    # Filas de datos — órdenes cerradas
    for orden in ordenes_cerradas_qs.order_by('-fecha_ingreso'):
        ws_cerradas.cell(row=row, column=1).value  = orden.detalle_equipo.orden_cliente
        ws_cerradas.cell(row=row, column=2).value  = orden.detalle_equipo.numero_serie or 'N/A'
        ws_cerradas.cell(row=row, column=3).value  = orden.detalle_equipo.get_tipo_equipo_display()
        ws_cerradas.cell(row=row, column=4).value  = orden.detalle_equipo.marca
        ws_cerradas.cell(row=row, column=5).value  = (orden.detalle_equipo.modelo or 'N/A')[:30]
        ws_cerradas.cell(row=row, column=6).value  = orden.get_estado_display()
        ws_cerradas.cell(row=row, column=7).value  = orden.dias_habiles_en_servicio
        ws_cerradas.cell(row=row, column=8).value  = orden.dias_sin_actualizacion_estado

        tipo_orden  = 'Servicio Normal'
        monto_orden = 0
        if hasattr(orden, 'venta_mostrador') and orden.venta_mostrador:
            tipo_orden  = 'Venta Mostrador'
            monto_orden = float(orden.venta_mostrador.total_venta)
        elif hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                tipo_orden  = 'Cotización Aceptada'
                monto_orden = float(orden.cotizacion.costo_total_final)
            elif orden.cotizacion.usuario_acepto is False:
                tipo_orden = 'Cotización Rechazada'
            else:
                tipo_orden = 'Cotización Pendiente'

        ws_cerradas.cell(row=row, column=9).value  = tipo_orden
        ws_cerradas.cell(row=row, column=10).value = f'${monto_orden:,.2f}' if monto_orden > 0 else 'N/A'
        ws_cerradas.cell(row=row, column=11).value = orden.sucursal.nombre
        ws_cerradas.cell(row=row, column=12).value = orden.fecha_ingreso.strftime('%d/%m/%Y')

        ultima_act = orden.historial.order_by('-fecha_evento').first()
        ws_cerradas.cell(row=row, column=13).value = (
            ultima_act.fecha_evento.strftime('%d/%m/%Y') if ultima_act else 'N/A'
        )

        cotiz_estado = 'N/A'
        if hasattr(orden, 'cotizacion') and orden.cotizacion:
            if orden.cotizacion.usuario_acepto is True:
                cotiz_estado = '✅ Aceptada'
            elif orden.cotizacion.usuario_acepto is False:
                cotiz_estado = '❌ Rechazada'
            else:
                cotiz_estado = '⏳ Pendiente'
        ws_cerradas.cell(row=row, column=14).value = cotiz_estado

        # Para cerradas: solo indicar si fue cancelada
        ws_cerradas.cell(row=row, column=15).value = (
            '❌ CANCELADA' if orden.estado == 'cancelado' else 'Completada'
        )

        # Colorear celda de estado
        color_est = get_estado_color(orden.estado)
        ws_cerradas.cell(row=row, column=6).fill = PatternFill(
            start_color=color_est, end_color=color_est, fill_type="solid"
        )
        ws_cerradas.cell(row=row, column=6).font = Font(bold=True, color="FFFFFF")

        # Resaltar candidatos RHITSO (morado claro)
        if orden.es_candidato_rhitso:
            rhitso_color = "ede9fe"
            for col in range(1, 16):
                cell = ws_cerradas.cell(row=row, column=col)
                if col != 6:
                    cell.fill = PatternFill(
                        start_color=rhitso_color, end_color=rhitso_color, fill_type="solid"
                    )

        row += 1

    auto_adjust_column_width(ws_cerradas)

    # ------------------------------------------------------------------
    # PASO 5: Construir respuesta HTTP con el archivo Excel
    # ------------------------------------------------------------------
    fecha_str      = datetime.now().strftime('%Y-%m-%d')
    nombre_limpio  = empleado.nombre_completo.replace(' ', '_')[:25]
    nombre_archivo = f'Reporte_OOW_FL_{nombre_limpio}_{fecha_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response


# ============================================================================
# DIRECTORIO DE EMPLEADOS — VISTA GERENCIAL (Marzo 2026)
# ============================================================================

ROLES_GERENCIALES = ('gerente_general', 'gerente_operacional', 'supervisor')


@login_required
def directorio_empleados(request):
    """
    Vista gerencial: muestra una cuadrícula de mini-cards con todos los
    empleados activos. Cada card muestra avatar, nombre, rol y rating rápido.

    Acceso permitido para:
    - Roles gerenciales (gerente_general, gerente_operacional, supervisor)
    - Usuarios con is_superuser=True (aunque no tengan rol gerencial)

    Args:
        request: HttpRequest del usuario autenticado.

    Efectos secundarios:
        Ninguno sobre la BD; solo lectura de empleados activos.
    """
    # EXPLICACIÓN PARA PRINCIPIANTES:
    # Primero dejamos pasar a los superusuarios (casilla "Es superusuario"
    # en el admin de Django). Si no lo son, exigimos perfil Empleado con
    # un rol gerencial. Así un admin técnico puede ver el directorio
    # aunque su rol de negocio sea otro (o no tenga empleado ligado).
    empleado_actual = getattr(request.user, 'empleado', None)
    es_superusuario = request.user.is_superuser
    es_rol_gerencial = bool(
        empleado_actual and empleado_actual.rol in ROLES_GERENCIALES
    )
    if not es_superusuario and not es_rol_gerencial:
        messages.error(request, 'No tienes permiso para acceder al directorio de empleados.')
        return redirect('servicio_tecnico:inicio')

    # Filtros opcionales (GET params)
    filtro_rol = request.GET.get('rol', '')
    filtro_sucursal = request.GET.get('sucursal', '')
    busqueda = request.GET.get('q', '').strip()

    empleados_qs = Empleado.objects.filter(
        activo=True
    ).select_related('user', 'sucursal').order_by('nombre_completo')

    if filtro_rol:
        empleados_qs = empleados_qs.filter(rol=filtro_rol)
    if filtro_sucursal:
        empleados_qs = empleados_qs.filter(sucursal_id=filtro_sucursal)
    if busqueda:
        empleados_qs = empleados_qs.filter(
            Q(nombre_completo__icontains=busqueda) |
            Q(cargo__icontains=busqueda)
        )

    # Calcular rating rápido para cada empleado
    empleados_data = []
    from .models import FeedbackCliente
    for emp in empleados_qs:
        rating = _calcular_rating_rapido(emp)
        # Comentarios de clientes — solo para recepcionistas
        comentarios_clientes = []
        if emp.rol == 'recepcionista':
            comentarios_clientes = list(
                FeedbackCliente.objects.filter(
                    tipo='satisfaccion',
                    orden__responsable_seguimiento=emp,
                    utilizado=True,
                    comentario_cliente__gt='',
                ).select_related('orden__detalle_equipo').order_by('-fecha_respuesta')[:15]
            )
        empleados_data.append({
            'empleado': emp,
            'rating': rating,
            'rol_display': dict(Empleado.ROL_CHOICES).get(emp.rol, emp.rol),
            'comentarios_clientes': comentarios_clientes,
        })

    # Datos para filtros
    roles_disponibles = Empleado.ROL_CHOICES
    sucursales_disponibles = Sucursal.objects.filter(activa=True).order_by('nombre')

    context = {
        'empleados_data': empleados_data,
        'total_empleados': len(empleados_data),
        'roles_disponibles': roles_disponibles,
        'sucursales_disponibles': sucursales_disponibles,
        'filtro_rol': filtro_rol,
        'filtro_sucursal': filtro_sucursal,
        'busqueda': busqueda,
    }

    return render(request, 'servicio_tecnico/directorio_empleados.html', context)


@login_required
def perfil_empleado(request, empleado_id):
    """
    Vista gerencial: muestra la tarjeta completa de un empleado específico.
    Reutiliza la misma template de mi_perfil con contexto adicional.

    Acceso permitido para roles gerenciales o usuarios superusuario.

    Args:
        request: HttpRequest del usuario autenticado.
        empleado_id: PK del Empleado cuyo perfil se quiere ver.

    Efectos secundarios:
        Ninguno sobre la BD; solo lectura.
    """
    # Misma regla de acceso que directorio_empleados (superusuario O rol gerencial)
    empleado_actual = getattr(request.user, 'empleado', None)
    es_superusuario = request.user.is_superuser
    es_rol_gerencial = bool(
        empleado_actual and empleado_actual.rol in ROLES_GERENCIALES
    )
    if not es_superusuario and not es_rol_gerencial:
        messages.error(request, 'No tienes permiso para ver perfiles de otros empleados.')
        return redirect('servicio_tecnico:inicio')

    empleado = get_object_or_404(Empleado, pk=empleado_id, activo=True)
    metricas, rating, rol, rol_display = _calcular_metricas_empleado(empleado)

    context = {
        'empleado': empleado,
        'rol': rol,
        'rol_display': rol_display,
        'metricas': metricas,
        'rating': rating,
        'es_vista_directorio': True,
    }

    return render(request, 'servicio_tecnico/mi_perfil.html', context)
