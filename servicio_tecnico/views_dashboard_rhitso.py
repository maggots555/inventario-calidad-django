"""
Dashboard RHITSO consolidado + exports Excel (Fase 8).

urls.py sigue usando views.<nombre> porque views.py reexporta estos símbolos.
NO incluye gestión RHITSO por orden (eso está en views_rhitso.py, Fase 7).
"""

import openpyxl
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .decorators import permission_required_with_message
from .models import (
    EstadoRHITSO,
    IncidenciaRHITSO,
    OrdenServicio,
    SeguimientoRHITSO,
)
from .utils_rhitso import (
    calcular_dias_en_estatus,
    calcular_dias_habiles,
    obtener_color_por_dias_rhitso,
    obtener_estado_proceso_rhitso,
)


# ============================================================================
# DASHBOARD RHITSO - VISTA CONSOLIDADA DE CANDIDATOS
# ============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def dashboard_rhitso(request):
    """
    Dashboard consolidado de todos los candidatos RHITSO.
 
    Args:
        request: HttpRequest object
    
    Returns:
        HttpResponse con el dashboard renderizado
    """
    from .utils_rhitso_analytics import obtener_embudo_rhitso

    # Filtros compartidos con el reporte de análisis (GET params)
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    sucursal_id = request.GET.get('sucursal', '') or None

    embudo = obtener_embudo_rhitso(
        fecha_inicio=fecha_inicio or None,
        fecha_fin=fecha_fin or None,
        sucursal_id=sucursal_id,
    )

    # =======================================================================
    # PASO 1: CONSULTA OPTIMIZADA DE CANDIDATOS RHITSO
    # =======================================================================

    # Esto evita el "N+1 problem" (hacer una consulta por cada relación)
    candidatos_rhitso = OrdenServicio.objects.filter(
        es_candidato_rhitso=True,
        id__in=embudo['candidatos_qs'].values_list('id', flat=True),
    ).select_related(
        'detalle_equipo',              # Información del equipo
        'sucursal',                    # Sucursal de la orden
        'tecnico_asignado_actual',     # Técnico asignado
        'responsable_seguimiento'      # Responsable del seguimiento
    ).prefetch_related(

        # para relaciones many-to-many o reverse foreign keys
        Prefetch(
            'seguimientos_rhitso',
            queryset=SeguimientoRHITSO.objects.select_related('estado', 'usuario_actualizacion').order_by('-fecha_actualizacion'),
            to_attr='seguimientos_ordenados'
        ),
        Prefetch(
            'incidencias_rhitso',
            queryset=IncidenciaRHITSO.objects.filter(estado__in=['ABIERTA', 'EN_REVISION']),
            to_attr='incidencias_abiertas'
        ),
        Prefetch(
            'incidencias_rhitso',
            queryset=IncidenciaRHITSO.objects.filter(estado='RESUELTA'),
            to_attr='incidencias_resueltas_lista'
        ),
    ).order_by('-fecha_ingreso')
    
    # =======================================================================
    # PASO 2: CALCULAR ESTADÍSTICAS GENERALES
    # =======================================================================
    
    # EXPLICACIÓN: Count() es una función agregada que cuenta registros
    # Usamos Q() para condiciones complejas (OR, AND, NOT)
    total_candidatos = candidatos_rhitso.count()
    total_enviados = candidatos_rhitso.filter(fecha_envio_rhitso__isnull=False).count()
    total_con_diagnostico = candidatos_rhitso.exclude(detalle_equipo__diagnostico_sic='').count()
    
    # Contar incidencias abiertas (en todas las órdenes)
    total_incidencias_abiertas = IncidenciaRHITSO.objects.filter(
        orden__in=candidatos_rhitso,
        estado__in=['ABIERTA', 'EN_REVISION']
    ).count()
    
    # =======================================================================
    # PASO 3: CALCULAR ESTADÍSTICAS POR SUCURSAL
    # =======================================================================
    
    # EXPLICACIÓN: Contamos cuántas órdenes hay en cada sucursal
    # Estas son las 3 sucursales principales según el sistema PHP
    stats_sucursal = {
        'satelite': candidatos_rhitso.filter(sucursal__nombre__icontains='Satelite').count(),
        'drop': candidatos_rhitso.filter(sucursal__nombre__icontains='Drop').count(),
        'mis': candidatos_rhitso.filter(sucursal__nombre__icontains='MIS').count(),
    }
    
    # =======================================================================
    # PASO 4: PREPARAR DATOS DETALLADOS DE CADA ORDEN
    # =======================================================================
    
    # EXPLICACIÓN: Estados que indican órdenes "excluidas" del proceso activo
    estados_excluidos = ['CERRADO', 'USUARIO NO ACEPTA ENVIO A RHITSO']
    estados_pendientes = ['PENDIENTE DE CONFIRMAR ENVIO A RHITSO']
    
    # Listas para separar órdenes por categoría
    activos = []
    pendientes = []
    excluidos = []
    
    # Iterar sobre cada orden para preparar sus datos
    for orden in candidatos_rhitso:
        # ===================================================================
        # 4.1: INFORMACIÓN BÁSICA
        # ===================================================================
        detalle = orden.detalle_equipo
        
        # Estado RHITSO actual (campo de texto simple)
        estado_rhitso_nombre = orden.estado_rhitso if orden.estado_rhitso else 'Pendiente'
        estado_rhitso_display = estado_rhitso_nombre
        
        # Buscar owner del estado (si existe en catálogo)
        try:
            estado_obj = EstadoRHITSO.objects.get(estado=estado_rhitso_nombre)
            owner_actual = estado_obj.owner
        except EstadoRHITSO.DoesNotExist:
            owner_actual = ''
        
        # ===================================================================
        # 4.2: CALCULAR DÍAS HÁBILES
        # ===================================================================
        
        # Días hábiles en SIC (tiempo total del proceso desde ingreso hasta completar)
        # EXPLICACIÓN: Este cálculo representa el tiempo TOTAL de la orden, incluyendo
        # todo el ciclo: diagnóstico SIC + envío a RHITSO + recepción de RHITSO
        if orden.fecha_recepcion_rhitso:
            # Si ya regresó de RHITSO, contar desde ingreso hasta recepción (proceso completado)
            dias_habiles_sic = calcular_dias_habiles(
                orden.fecha_ingreso,
                orden.fecha_recepcion_rhitso
            )
        else:
            # Si no ha regresado (o nunca se envió), contar hasta hoy
            dias_habiles_sic = calcular_dias_habiles(orden.fecha_ingreso)
        
        # Días hábiles en RHITSO (si aplica)
        dias_habiles_rhitso = 0
        if orden.fecha_envio_rhitso:
            if orden.fecha_recepcion_rhitso:
                # Ya regresó de RHITSO
                dias_habiles_rhitso = calcular_dias_habiles(
                    orden.fecha_envio_rhitso,
                    orden.fecha_recepcion_rhitso
                )
            else:
                # Todavía en RHITSO
                dias_habiles_rhitso = calcular_dias_habiles(
                    orden.fecha_envio_rhitso
                )
        
        # ===================================================================
        # 4.3: CALCULAR DÍAS SIN ACTUALIZACIÓN
        # ===================================================================
        
        # EXPLICACIÓN: Buscamos el último comentario de usuario (no del sistema)
        ultimo_seguimiento_usuario = None
        if hasattr(orden, 'seguimientos_ordenados') and orden.seguimientos_ordenados:
            # Buscar el último seguimiento que NO sea automático
            for seg in orden.seguimientos_ordenados:
                if not seg.es_cambio_automatico:
                    ultimo_seguimiento_usuario = seg
                    break
        
        if ultimo_seguimiento_usuario:
            dias_sin_actualizar = calcular_dias_en_estatus(
                ultimo_seguimiento_usuario.fecha_actualizacion
            )
            fecha_ultimo_comentario = ultimo_seguimiento_usuario.fecha_actualizacion
            ultimo_comentario = ultimo_seguimiento_usuario.observaciones
        else:
            # Si no hay seguimientos, contar desde fecha de ingreso
            dias_sin_actualizar = calcular_dias_en_estatus(orden.fecha_ingreso)
            fecha_ultimo_comentario = None
            ultimo_comentario = ''
        
        # ===================================================================
        # 4.4: CONTAR INCIDENCIAS
        # ===================================================================
        
        # EXPLICACIÓN: Usamos los prefetch que definimos arriba
        incidencias_abiertas_count = len(orden.incidencias_abiertas) if hasattr(orden, 'incidencias_abiertas') else 0
        incidencias_resueltas_count = len(orden.incidencias_resueltas_lista) if hasattr(orden, 'incidencias_resueltas_lista') else 0
        total_incidencias = incidencias_abiertas_count + incidencias_resueltas_count
        
        # ===================================================================
        # 4.5: DETERMINAR ESTADO DEL PROCESO
        # ===================================================================
        
        estado_proceso = obtener_estado_proceso_rhitso(orden)
        
        # ===================================================================
        # 4.6: DETERMINAR COLOR SEGÚN DÍAS EN RHITSO
        # ===================================================================
        
        color_badge_dias = obtener_color_por_dias_rhitso(dias_habiles_rhitso)
        
        # ===================================================================
        # 4.7: CONSTRUIR DICCIONARIO CON TODA LA INFORMACIÓN
        # ===================================================================
        
        orden_data = {
            # Información básica
            'id': orden.id,
            'numero_orden_interno': orden.numero_orden_interno,
            'fecha_ingreso': orden.fecha_ingreso,
            'estado_orden': orden.get_estado_display(),
            
            # Información del equipo
            'servicio': detalle.falla_principal if detalle else 'Sin observaciones',
            'numero_serie': detalle.numero_serie if detalle else 'N/A',
            'marca': detalle.marca if detalle else 'N/A',
            'modelo': detalle.modelo if detalle else 'N/A',
            'orden_cliente': detalle.orden_cliente if detalle else 'N/A',
            
            # Sucursal
            'sucursal': orden.sucursal.nombre if orden.sucursal else 'N/A',
            
            # Estado RHITSO
            'estado_rhitso_nombre': estado_rhitso_nombre,
            'estado_rhitso_display': estado_rhitso_display,
            'owner_actual': owner_actual,
            
            # Incidencias
            'incidencias_abiertas': incidencias_abiertas_count,
            'incidencias_resueltas': incidencias_resueltas_count,
            'total_incidencias': total_incidencias,
            
            # Fechas y tiempos
            'fecha_envio_rhitso': orden.fecha_envio_rhitso,
            'dias_habiles_sic': dias_habiles_sic,
            'dias_habiles_rhitso': dias_habiles_rhitso,
            'dias_sin_actualizar': dias_sin_actualizar,
            'fecha_ultimo_comentario': fecha_ultimo_comentario,
            'ultimo_comentario': ultimo_comentario,
            
            # Estado del proceso
            'estado_proceso': estado_proceso,
            'color_badge_dias': color_badge_dias,
            
            # Diagnóstico
            'tiene_diagnostico': bool(detalle and detalle.diagnostico_sic),
        }
        
        # ===================================================================
        # 4.8: CLASIFICAR ORDEN EN CATEGORÍA CORRESPONDIENTE
        # ===================================================================
        
        # EXPLICACIÓN: Separamos las órdenes según su estado RHITSO
        if estado_rhitso_nombre in estados_excluidos:
            excluidos.append(orden_data)
        elif estado_rhitso_nombre in estados_pendientes:
            pendientes.append(orden_data)
        else:
            activos.append(orden_data)
    
    # =======================================================================
    # PASO 5: OBTENER LISTA DE ESTADOS RHITSO PARA FILTROS
    # =======================================================================
    
    # EXPLICACIÓN: Creamos listas de estados únicos para los dropdowns de filtro
    estados_activos = list(set(orden['estado_rhitso_display'] for orden in activos))
    estados_pendientes_lista = list(set(orden['estado_rhitso_display'] for orden in pendientes))
    estados_excluidos_lista = list(set(orden['estado_rhitso_display'] for orden in excluidos))
    
    # Ordenar alfabéticamente
    estados_activos.sort()
    estados_pendientes_lista.sort()
    estados_excluidos_lista.sort()
    
    # =======================================================================
    # PASO 6: PREPARAR CONTEXTO PARA EL TEMPLATE
    # =======================================================================
    
    context = {
        # Estadísticas generales
        'total_candidatos': total_candidatos,
        'total_enviados': total_enviados,
        'total_con_diagnostico': total_con_diagnostico,
        'total_incidencias_abiertas': total_incidencias_abiertas,
        
        # Estadísticas por sucursal
        'stats_sucursal': stats_sucursal,
        
        # Órdenes por categoría
        'activos': activos,
        'pendientes': pendientes,
        'excluidos': excluidos,
        
        # Contadores para pestañas
        'count_activos': len(activos),
        'count_pendientes': len(pendientes),
        'count_excluidos': len(excluidos),
        
        # Listas de estados para filtros
        'estados_activos': estados_activos,
        'estados_pendientes': estados_pendientes_lista,
        'estados_excluidos': estados_excluidos_lista,
        
        # Información adicional
        'fecha_actualizacion': timezone.now(),

        # Embudo de conversión RHITSO
        'embudo': embudo,
        'filtro_fecha_inicio': fecha_inicio,
        'filtro_fecha_fin': fecha_fin,
        'filtro_sucursal': sucursal_id or '',
    }
    
    return render(request, 'servicio_tecnico/rhitso/dashboard_rhitso.html', context)


# =============================================================================
# EXPORTACIÓN EXCEL RHITSO CON OPENPYXL
# =============================================================================

@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def exportar_excel_rhitso(request):
    """
    Genera y descarga un reporte Excel profesional de candidatos RHITSO.

    Args:
        request: HttpRequest object
    
    Returns:
        HttpResponse con archivo Excel para descarga
    """
    # =========================================================================
    # PASO 1: PREPARAR DATOS (REUTILIZAR LÓGICA DEL DASHBOARD)
    # =========================================================================
    
    # EXPLICACIÓN: Reutilizamos la misma consulta optimizada del dashboard
    from .utils_rhitso_analytics import obtener_queryset_candidatos

    fecha_inicio = request.GET.get('fecha_inicio', '') or None
    fecha_fin = request.GET.get('fecha_fin', '') or None
    sucursal_id = request.GET.get('sucursal', '') or None

    candidatos_rhitso = obtener_queryset_candidatos(
        fecha_inicio, fecha_fin, sucursal_id
    ).select_related(
        'detalle_equipo',
        'sucursal',
        'tecnico_asignado_actual',
        'responsable_seguimiento'
    ).prefetch_related(
        Prefetch(
            'seguimientos_rhitso',
            queryset=SeguimientoRHITSO.objects.select_related('estado', 'usuario_actualizacion').order_by('-fecha_actualizacion'),
            to_attr='seguimientos_ordenados'
        ),
        Prefetch(
            'incidencias_rhitso',
            queryset=IncidenciaRHITSO.objects.filter(estado__in=['ABIERTA', 'EN_REVISION']),
            to_attr='incidencias_abiertas'
        ),
        Prefetch(
            'incidencias_rhitso',
            queryset=IncidenciaRHITSO.objects.filter(estado='RESUELTA'),
            to_attr='incidencias_resueltas_lista'
        ),
    ).order_by('-fecha_ingreso')
    
    # Estados para clasificación
    estados_excluidos = ['CERRADO', 'USUARIO NO ACEPTA ENVIO A RHITSO']
    estados_pendientes = ['PENDIENTE DE CONFIRMAR ENVIO A RHITSO']
    
    # Listas para separar órdenes por categoría
    activos = []
    pendientes = []
    excluidos = []
    
    # Procesar cada orden
    for orden in candidatos_rhitso:
        detalle = orden.detalle_equipo
        
        # Estado RHITSO actual
        estado_rhitso_nombre = orden.estado_rhitso if orden.estado_rhitso else 'Pendiente'
        estado_rhitso_display = estado_rhitso_nombre
        
        # Buscar owner del estado
        try:
            estado_obj = EstadoRHITSO.objects.get(estado=estado_rhitso_nombre)
            owner_actual = estado_obj.owner
        except EstadoRHITSO.DoesNotExist:
            owner_actual = ''
        
        # Calcular días hábiles en SIC (tiempo total del proceso)
        if orden.fecha_recepcion_rhitso:
            dias_habiles_sic = calcular_dias_habiles(orden.fecha_ingreso, orden.fecha_recepcion_rhitso)
        else:
            dias_habiles_sic = calcular_dias_habiles(orden.fecha_ingreso)
        
        # Calcular días hábiles en RHITSO
        dias_habiles_rhitso = 0
        if orden.fecha_envio_rhitso:
            if orden.fecha_recepcion_rhitso:
                dias_habiles_rhitso = calcular_dias_habiles(orden.fecha_envio_rhitso, orden.fecha_recepcion_rhitso)
            else:
                dias_habiles_rhitso = calcular_dias_habiles(orden.fecha_envio_rhitso)
        
        # Calcular días sin actualización
        ultimo_seguimiento_usuario = None
        if hasattr(orden, 'seguimientos_ordenados') and orden.seguimientos_ordenados:
            for seg in orden.seguimientos_ordenados:
                if not seg.es_cambio_automatico:
                    ultimo_seguimiento_usuario = seg
                    break
        
        if ultimo_seguimiento_usuario:
            dias_sin_actualizar = calcular_dias_en_estatus(ultimo_seguimiento_usuario.fecha_actualizacion)
            fecha_ultimo_comentario = ultimo_seguimiento_usuario.fecha_actualizacion
            ultimo_comentario = ultimo_seguimiento_usuario.observaciones
        else:
            dias_sin_actualizar = calcular_dias_en_estatus(orden.fecha_ingreso)
            fecha_ultimo_comentario = None
            ultimo_comentario = ''
        
        # Contar incidencias
        incidencias_abiertas_count = len(orden.incidencias_abiertas) if hasattr(orden, 'incidencias_abiertas') else 0
        incidencias_resueltas_count = len(orden.incidencias_resueltas_lista) if hasattr(orden, 'incidencias_resueltas_lista') else 0
        total_incidencias = incidencias_abiertas_count + incidencias_resueltas_count
        
        # Determinar estado del proceso
        estado_proceso = obtener_estado_proceso_rhitso(orden)
        
        # Construir diccionario con toda la información
        orden_data = {
            'orden_cliente': detalle.orden_cliente if detalle else 'Sin orden',
            'numero_serie': detalle.numero_serie if detalle else 'N/A',
            'marca': detalle.marca if detalle else 'N/A',
            'modelo': detalle.modelo if detalle else 'N/A',
            'fecha_ingreso': orden.fecha_ingreso,
            'sucursal': orden.sucursal.nombre if orden.sucursal else 'N/A',
            'estado_orden': orden.get_estado_display(),
            'estado_rhitso_display': estado_rhitso_display,
            'owner_actual': owner_actual,
            'total_incidencias': f"{incidencias_abiertas_count}/{total_incidencias}",
            'fecha_envio_rhitso': orden.fecha_envio_rhitso,
            'dias_habiles_sic': dias_habiles_sic,
            'dias_habiles_rhitso': dias_habiles_rhitso,
            'dias_sin_actualizar': dias_sin_actualizar,
            'estado_proceso': estado_proceso,
            'fecha_ultimo_comentario': fecha_ultimo_comentario,
            'ultimo_comentario': ultimo_comentario if ultimo_comentario else 'Sin comentario',
        }
        
        # Clasificar orden en categoría correspondiente
        if estado_rhitso_nombre in estados_excluidos:
            excluidos.append(orden_data)
        elif estado_rhitso_nombre in estados_pendientes:
            pendientes.append(orden_data)
        else:
            activos.append(orden_data)
    
    # =========================================================================
    # PASO 2: CREAR WORKBOOK DE EXCEL
    # =========================================================================
    
    # EXPLICACIÓN: Workbook es el archivo Excel completo
    wb = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # =========================================================================
    # PASO 3: DEFINIR ESTILOS PROFESIONALES (COMO INVENTARIO)
    # =========================================================================
    
    # EXPLICACIÓN: Definimos estilos reutilizables para mantener consistencia
    
    # Fuente para encabezados: Blanco, negrita, tamaño 11
    header_font = Font(
        name='Calibri',
        bold=True,
        color="FFFFFF",
        size=11
    )
    
    # Relleno azul para encabezados (#366092 es el azul corporativo)
    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid"
    )
    
    # Alineación centrada para encabezados
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Bordes para todas las celdas
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Fuente normal para datos
    normal_font = Font(name='Calibri', size=10)
    
    # Alineación para comentarios (con wrap text)
    wrap_alignment = Alignment(
        horizontal='left',
        vertical='top',
        wrap_text=True
    )
    
    # =========================================================================
    # PASO 4: DEFINIR ENCABEZADOS (17 COLUMNAS)
    # =========================================================================
    
    # EXPLICACIÓN: Estos son los títulos de las columnas que aparecerán en Excel
    headers = [
        'Servicio Cliente',
        'N° Serie',
        'Marca',
        'Modelo',
        'Fecha Ingreso a SIC',
        'Sucursal',
        'Estado General',
        'Estado RHITSO',
        'Owner',
        'Incidencias',
        'Fecha Envío RHITSO',
        'Días Hábiles SIC',
        'Días Hábiles RHITSO',
        'Días en estatus',
        'Estado Proceso',
        'Fecha Último Comentario',
        'Comentario'
    ]
    
    # Anchos óptimos para cada columna (en caracteres)
    column_widths = [20, 15, 15, 15, 18, 15, 18, 25, 15, 12, 18, 18, 18, 15, 20, 20, 50]
    
    # =========================================================================
    # PASO 5: FUNCIÓN AUXILIAR PARA CREAR HOJAS FORMATEADAS
    # =========================================================================
    
    def crear_hoja_excel(nombre_hoja, datos_lista, color_categoria):
        """
        Crea una hoja de Excel con formato profesional.

        """
        # Crear la hoja
        ws = wb.create_sheet(nombre_hoja)
        
        # PASO 5.1: Agregar encabezados con estilo
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # PASO 5.2: Configurar anchos de columna
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width
        
        # PASO 5.3: Congelar primera fila (encabezados)
        # EXPLICACIÓN: freeze_panes permite que los encabezados permanezcan
        # visibles cuando el usuario hace scroll hacia abajo
        ws.freeze_panes = 'A2'
        
        # PASO 5.4: Agregar datos fila por fila
        for row_num, orden in enumerate(datos_lista, start=2):
            # Preparar los valores de cada columna
            valores = [
                orden['orden_cliente'],
                orden['numero_serie'],
                orden['marca'],
                orden['modelo'],
                orden['fecha_ingreso'].strftime('%d/%m/%Y') if orden['fecha_ingreso'] else '',
                orden['sucursal'],
                orden['estado_orden'],
                orden['estado_rhitso_display'],
                orden['owner_actual'],
                orden['total_incidencias'],
                orden['fecha_envio_rhitso'].strftime('%d/%m/%Y') if orden['fecha_envio_rhitso'] else 'No enviado',
                orden['dias_habiles_sic'],
                orden['dias_habiles_rhitso'],
                orden['dias_sin_actualizar'],
                orden['estado_proceso'],
                orden['fecha_ultimo_comentario'].strftime('%d/%m/%Y %H:%M') if orden['fecha_ultimo_comentario'] else 'Sin comentario',
                orden['ultimo_comentario']
            ]
            
            # Escribir valores en las celdas
            for col_num, valor in enumerate(valores, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                cell.font = normal_font
                cell.border = thin_border
                
                # Aplicar wrap text en la columna de comentarios (última columna)
                if col_num == len(headers):  # Columna de comentario
                    cell.alignment = wrap_alignment
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # PASO 5.5: Colorear fila según estado y urgencia
            # EXPLICACIÓN: Aplicamos colores para identificar visualmente el estado
            
            estado_proc = orden['estado_proceso']
            dias_sin_act = orden['dias_sin_actualizar']
            
            # Determinar color de fila
            row_fill = None
            
            if estado_proc == 'Completado':
                # Verde claro para órdenes completadas
                row_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif estado_proc == 'En RHITSO':
                # Amarillo claro para órdenes en RHITSO
                row_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            elif estado_proc == 'Solo en SIC':
                # Gris claro para órdenes solo en SIC
                row_fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
            
            # Sobrescribir con rojo si tiene más de 5 días sin actualizar (URGENTE)
            if dias_sin_act > 5:
                row_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # Aplicar el color a toda la fila
            if row_fill:
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_num).fill = row_fill
        
        # PASO 5.6: Agregar auto-filtro a los encabezados
        # EXPLICACIÓN: Permite al usuario filtrar datos directamente en Excel
        ws.auto_filter.ref = ws.dimensions
    
    # =========================================================================
    # PASO 6: CREAR LAS 3 HOJAS
    # =========================================================================
    
    # Hoja 1: Activos
    crear_hoja_excel(f"Activos ({len(activos)})", activos, "FFF3CD")
    
    # Hoja 2: Pendientes
    crear_hoja_excel(f"Pendientes ({len(pendientes)})", pendientes, "E2E3E5")
    
    # Hoja 3: Excluidos
    crear_hoja_excel(f"Excluidos ({len(excluidos)})", excluidos, "F8D7DA")
    
    # =========================================================================
    # PASO 7: PREPARAR RESPUESTA HTTP PARA DESCARGA
    # =========================================================================
    
    # EXPLICACIÓN: Creamos una respuesta HTTP con el tipo de contenido adecuado
    # para que el navegador sepa que es un archivo Excel y lo descargue
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Definir el nombre del archivo con fecha y hora actual
    nombre_archivo = f'Reporte_RHITSO_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response


@login_required
@permission_required_with_message('servicio_tecnico.view_ordenservicio')
def exportar_analisis_rhitso(request):
    """
    Genera y descarga el reporte Excel del embudo de conversión RHITSO.

    Incluye resumen de KPIs, detalle de candidatos, rechazos de cotización
    con observaciones y órdenes sin decisión de envío.

    Args:
        request: HttpRequest con filtros opcionales fecha_inicio, fecha_fin, sucursal.

    Returns:
        HttpResponse con archivo Excel (.xlsx) para descarga.
    """
    from .utils_rhitso_analytics import (
        obtener_embudo_rhitso,
        obtener_filas_hoja_rechazos_y_no_aptos,
        obtener_detalle_todas_candidatas,
    )

    fecha_inicio = request.GET.get('fecha_inicio', '') or None
    fecha_fin = request.GET.get('fecha_fin', '') or None
    sucursal_id = request.GET.get('sucursal', '') or None

    embudo = obtener_embudo_rhitso(fecha_inicio, fecha_fin, sucursal_id)
    detalle_candidatos = obtener_detalle_todas_candidatas(fecha_inicio, fecha_fin, sucursal_id)
    comentarios_rechazo = obtener_filas_hoja_rechazos_y_no_aptos(embudo['candidatos_qs'])

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    normal_font = Font(name='Calibri', size=10)
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    title_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    kpi_label_font = Font(name='Calibri', bold=True, size=11)
    kpi_value_font = Font(name='Calibri', bold=True, size=12, color='366092')

    filtros_texto = f"Período ingreso: {fecha_inicio or 'Inicio'} — {fecha_fin or 'Hoy'}"
    if sucursal_id:
        filtros_texto += f" | Sucursal ID: {sucursal_id}"

    def aplicar_encabezados(ws, headers, fila=1):
        """Escribe encabezados con estilo corporativo en la hoja indicada."""
        for col_idx, titulo in enumerate(headers, start=1):
            celda = ws.cell(row=fila, column=col_idx, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
            celda.border = thin_border

    def ajustar_anchos(ws, anchos):
        """Configura ancho de columnas según lista de caracteres."""
        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho

    # -------------------------------------------------------------------------
    # HOJA 1: RESUMEN DEL EMBUDO
    # -------------------------------------------------------------------------
    ws_resumen = wb.create_sheet('Resumen')
    ws_resumen.merge_cells('A1:D1')
    titulo = ws_resumen['A1']
    titulo.value = 'Análisis RHITSO — Embudo de Conversión'
    titulo.font = title_font
    titulo.fill = title_fill
    titulo.alignment = Alignment(horizontal='center', vertical='center')

    ws_resumen.merge_cells('A2:D2')
    subtitulo = ws_resumen['A2']
    subtitulo.value = filtros_texto
    subtitulo.font = Font(name='Calibri', italic=True, size=10, color='666666')
    subtitulo.alignment = Alignment(horizontal='center')

    fila_kpi = 4
    resumen_filas = [
        ('NIVEL 1 — Candidatos RHITSO', embudo['total_candidatos'], '100%', ''),
        ('', '', '', ''),
        ('NIVEL 2 — Decisión de envío a RHITSO', '', '', ''),
        ('  Aceptaron envío', embudo['acepto_envio_count'], f"{embudo['acepto_envio_pct']}%", 'sobre candidatos'),
        ('  Rechazaron envío', embudo['rechazo_envio_count'], f"{embudo['rechazo_envio_pct']}%", 'sobre candidatos'),
        ('  Sin decisión de envío', embudo['sin_decision_envio_count'], f"{embudo['sin_decision_envio_pct']}%", 'sobre candidatos'),
        ('', '', '', ''),
        ('NIVEL 3 — Cohorte: aceptaron envío', embudo['total_cohorte_acepto'], '', 'base cohorte'),
        ('  Cliente acepta cotización', embudo['acepto_cotiz_count'], f"{embudo['acepto_cotiz_pct']}%", 'sobre cohorte'),
        ('  Cliente no acepta cotización', embudo['rechazo_cotiz_count'], f"{embudo['rechazo_cotiz_pct']}%", 'sobre cohorte'),
        ('  No apto para reparación', embudo['no_apto_count'], f"{embudo['no_apto_pct']}%", 'sobre cohorte'),
        ('  En proceso (sin esos estados)', embudo['en_proceso_cohorte_count'], f"{embudo['en_proceso_cohorte_pct']}%", 'sobre cohorte'),
    ]

    ws_resumen.cell(row=3, column=1, value='Métrica').font = kpi_label_font
    ws_resumen.cell(row=3, column=2, value='Cantidad').font = kpi_label_font
    ws_resumen.cell(row=3, column=3, value='%').font = kpi_label_font
    ws_resumen.cell(row=3, column=4, value='Referencia').font = kpi_label_font

    for etiqueta, cantidad, pct, ref in resumen_filas:
        ws_resumen.cell(row=fila_kpi, column=1, value=etiqueta).font = normal_font
        if cantidad != '':
            celda_cnt = ws_resumen.cell(row=fila_kpi, column=2, value=cantidad)
            celda_cnt.font = kpi_value_font
        ws_resumen.cell(row=fila_kpi, column=3, value=pct).font = normal_font
        ws_resumen.cell(row=fila_kpi, column=4, value=ref).font = Font(name='Calibri', size=9, color='888888')
        fila_kpi += 1

    nota_fila = fila_kpi + 1
    ws_resumen.merge_cells(f'A{nota_fila}:D{nota_fila + 2}')
    nota = ws_resumen[f'A{nota_fila}']
    nota.value = (
        'Nota: Las métricas se basan en estados RHITSO registrados en SeguimientoRHITSO. '
        'Si el técnico no actualiza el estado en el panel RHITSO, el embudo puede mostrar '
        '"sin decisión" aunque exista respuesta en el flujo SIC de cotización.'
    )
    nota.font = Font(name='Calibri', italic=True, size=9, color='666666')
    nota.alignment = wrap_alignment

    ajustar_anchos(ws_resumen, [42, 14, 10, 22])

    # -------------------------------------------------------------------------
    # HOJA 2: DETALLE CANDIDATOS
    # -------------------------------------------------------------------------
    ws_detalle = wb.create_sheet('Detalle Candidatos')
    headers_detalle = [
        'ID Orden', 'Orden Cliente', 'N° Serie', 'Marca', 'Modelo', 'Sucursal',
        'Fecha Ingreso', 'Técnico Asignado',
        'Estado RHITSO Actual', 'Estado Orden SIC',
        'Aceptó Envío', 'Rechazó Envío', 'Aceptó Cotiz.', 'Rechazó Cotiz.', 'No Apto',
    ]
    aplicar_encabezados(ws_detalle, headers_detalle)

    for row_idx, fila in enumerate(detalle_candidatos, start=2):
        valores = [
            fila['id'],
            fila['orden_cliente'],
            fila['numero_serie'],
            fila['marca'],
            fila['modelo'],
            fila['sucursal'],
            fila['fecha_ingreso'].strftime('%d/%m/%Y %H:%M') if fila['fecha_ingreso'] else '',
            fila['tecnico_asignado'],
            fila['estado_rhitso_actual'],
            fila['estado_orden'],
            'Sí' if fila['acepto_envio'] else 'No',
            'Sí' if fila['rechazo_envio'] else 'No',
            'Sí' if fila['acepto_cotiz'] else 'No',
            'Sí' if fila['rechazo_cotiz'] else 'No',
            'Sí' if fila['no_apto'] else 'No',
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws_detalle.cell(row=row_idx, column=col_idx, value=valor)
            celda.font = normal_font
            celda.border = thin_border
            celda.alignment = wrap_alignment

    ajustar_anchos(ws_detalle, [10, 22, 16, 14, 18, 14, 18, 26, 30, 18, 12, 12, 12, 12, 10])
    ws_detalle.freeze_panes = 'A2'

    # -------------------------------------------------------------------------
    # HOJA 3: RECHAZOS COTIZACIÓN Y NO APTOS (con observaciones)
    # -------------------------------------------------------------------------
    ws_rechazos = wb.create_sheet('Rechazos Cotización')
    headers_rechazos = [
        'Orden Cliente', 'N° Serie', 'Marca', 'Modelo', 'Sucursal',
        'Estado RHITSO', 'Fecha Cambio Estado', 'Usuario', 'Observaciones / Motivo',
    ]
    aplicar_encabezados(ws_rechazos, headers_rechazos)

    for row_idx, seg in enumerate(comentarios_rechazo, start=2):
        detalle = seg.orden.detalle_equipo
        usuario_nombre = ''
        if seg.usuario_actualizacion:
            usuario_nombre = str(seg.usuario_actualizacion)
        estado_rhitso = seg.estado.estado if seg.estado else 'N/A'
        observaciones = (seg.observaciones or '').strip()
        if not observaciones:
            observaciones = 'Sin comentario disponible'
        valores = [
            detalle.orden_cliente if detalle else 'N/A',
            detalle.numero_serie if detalle else 'N/A',
            detalle.marca if detalle else 'N/A',
            detalle.modelo if detalle else 'N/A',
            seg.orden.sucursal.nombre if seg.orden.sucursal else 'N/A',
            estado_rhitso,
            seg.fecha_actualizacion.strftime('%d/%m/%Y %H:%M'),
            usuario_nombre,
            observaciones,
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws_rechazos.cell(row=row_idx, column=col_idx, value=valor)
            celda.font = normal_font
            celda.border = thin_border
            celda.alignment = wrap_alignment

    ajustar_anchos(ws_rechazos, [22, 16, 14, 18, 14, 28, 18, 22, 55])
    ws_rechazos.freeze_panes = 'A2'

    # -------------------------------------------------------------------------
    # HOJA 4: SIN DECISIÓN DE ENVÍO
    # -------------------------------------------------------------------------
    ws_sin_decision = wb.create_sheet('Sin Decisión Envío')
    headers_sin = [
        'ID Orden', 'Orden Cliente', 'N° Serie', 'Marca', 'Modelo',
        'Sucursal', 'Fecha Ingreso', 'Estado RHITSO Actual',
    ]
    aplicar_encabezados(ws_sin_decision, headers_sin)

    sin_decision_filas = [
        construir_fila
        for construir_fila in detalle_candidatos
        if not construir_fila['acepto_envio'] and not construir_fila['rechazo_envio']
    ]

    for row_idx, fila in enumerate(sin_decision_filas, start=2):
        valores = [
            fila['id'],
            fila['orden_cliente'],
            fila['numero_serie'],
            fila['marca'],
            fila['modelo'],
            fila['sucursal'],
            fila['fecha_ingreso'].strftime('%d/%m/%Y %H:%M') if fila['fecha_ingreso'] else '',
            fila['estado_rhitso_actual'],
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws_sin_decision.cell(row=row_idx, column=col_idx, value=valor)
            celda.font = normal_font
            celda.border = thin_border
            celda.alignment = wrap_alignment

    ajustar_anchos(ws_sin_decision, [10, 22, 16, 14, 18, 14, 18, 30])
    ws_sin_decision.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f'Analisis_RHITSO_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


